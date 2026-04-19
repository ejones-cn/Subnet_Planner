#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IP地址管理（IPAM）模块 - SQLite版本
负责IP地址的分配、跟踪、状态管理和历史记录
"""

import sqlite3
import json
import os
import sys
import logging
from datetime import datetime, timedelta
from typing import Any

# 导入国际化模块
from i18n import translate as _
from validators import IPAMValidator

import ipaddress
from ip_subnet_calculator import ip_to_int


class IPAMSQLite:
    """IP地址管理类 - SQLite版本"""
    
    def __init__(self, db_file: str | None = None):
        """初始化IPAM
        
        Args:
            db_file: 数据库文件路径，为None时自动使用程序所在目录
        """
        # 获取应用程序所在目录（确保在单文件模式下也能正确获取）
        self.app_dir: str = self._get_app_directory()
        
        if db_file is None:
            # 使用应用程序目录作为数据库位置
            db_file = os.path.join(self.app_dir, "SubnetPlanner_data.db")
        
        self.db_file: str = db_file
        # 备份目录始终在应用程序目录下，确保数据持久化
        self.backup_dir: str = os.path.join(self.app_dir, "ipam_backups")
        # 创建备份目录
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
        self.init_db()
    
    def _get_app_directory(self) -> str:
        """获取应用程序所在目录
        
        Returns:
            str: 应用程序所在目录路径
        """
        app_dir = None
        
        # 优先使用 sys.argv[0]（这是最可靠的方法）
        if sys.argv and sys.argv[0]:
            exe_path = sys.argv[0]
            if not os.path.isabs(exe_path):
                exe_path = os.path.abspath(exe_path)
            if os.path.exists(exe_path):
                app_dir = os.path.dirname(exe_path)
        
        # 如果 sys.argv[0] 不可用，尝试使用 ctypes（仅限Windows平台）
        if app_dir is None and sys.platform == 'win32':
            try:
                import ctypes
                from ctypes import wintypes
                
                GetModuleFileNameW = ctypes.windll.kernel32.GetModuleFileNameW
                GetModuleFileNameW.argtypes = [wintypes.HMODULE, wintypes.LPWSTR, wintypes.DWORD]
                GetModuleFileNameW.restype = wintypes.DWORD
                
                buffer = ctypes.create_unicode_buffer(260)
                if GetModuleFileNameW(None, buffer, 260) > 0:
                    exe_path = buffer.value
                    if os.path.exists(exe_path):
                        app_dir = os.path.dirname(exe_path)
            except Exception:
                pass
        
        # 如果以上都失败，使用 __file__（非单文件模式）
        if app_dir is None:
            app_dir = os.path.dirname(os.path.abspath(__file__))
        
        return app_dir
    
    def init_db(self) -> None:
        """初始化数据库"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        
        # 创建networks表
        _ = cursor.execute('''
        CREATE TABLE IF NOT EXISTS networks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            network_address TEXT UNIQUE NOT NULL,
            description TEXT,
            vlan TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
        ''')
        
        # 为现有networks表添加vlan列（如果不存在）
        try:
            _ = cursor.execute('ALTER TABLE networks ADD COLUMN vlan TEXT')
            conn.commit()
        except sqlite3.OperationalError:
            # 列已存在，忽略错误
            pass
        
        # 检查是否需要迁移旧的ip_addresses表（移除network_id字段）
        cursor.execute("PRAGMA table_info(ip_addresses)")
        ip_addresses_columns = cursor.fetchall()
        has_network_id = any(col[1] == 'network_id' for col in ip_addresses_columns)
        
        if has_network_id:
            # 需要迁移：重命名旧表，创建新表，迁移数据
            try:
                _ = cursor.execute('ALTER TABLE ip_addresses RENAME TO ip_addresses_old')
                conn.commit()
            except sqlite3.OperationalError:
                pass
            
            # 创建新的ip_addresses表（不含network_id字段）
            _ = cursor.execute('''
            CREATE TABLE ip_addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip_address TEXT UNIQUE NOT NULL,
                status TEXT NOT NULL,
                hostname TEXT,
                description TEXT,
                allocated_at TEXT,
                allocated_by TEXT,
                expiry_date TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                mac_address TEXT
            )
            ''')
            
            # 迁移数据（忽略重复的IP地址）
            _ = cursor.execute('''
            INSERT OR IGNORE INTO ip_addresses (id, ip_address, status, hostname, description, 
            allocated_at, allocated_by, expiry_date, created_at, updated_at, mac_address)
            SELECT id, ip_address, status, hostname, description, 
            allocated_at, allocated_by, expiry_date, created_at, updated_at, 
            COALESCE(mac_address, '') FROM ip_addresses_old
            ''')
            conn.commit()
            
            # 删除旧表
            _ = cursor.execute('DROP TABLE ip_addresses_old')
            conn.commit()
        else:
            # 创建ip_addresses表（如果不存在）
            _ = cursor.execute('''
            CREATE TABLE IF NOT EXISTS ip_addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip_address TEXT UNIQUE NOT NULL,
                status TEXT NOT NULL,
                hostname TEXT,
                description TEXT,
                allocated_at TEXT,
                allocated_by TEXT,
                expiry_date TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                mac_address TEXT
            )
            ''')
        
        # 迁移：将旧的 'available' 状态更新为 'released'
        conn.execute('BEGIN TRANSACTION')
        try:
            cursor.execute('UPDATE ip_addresses SET status = ? WHERE status = ?', ('released', 'available'))
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        
        # 检查是否需要迁移旧的allocation_history表（移除network_id字段）
        cursor.execute("PRAGMA table_info(allocation_history)")
        history_columns = cursor.fetchall()
        has_history_network_id = any(col[1] == 'network_id' for col in history_columns)
        
        if has_history_network_id:
            # 需要迁移：重命名旧表，创建新表，迁移数据
            try:
                _ = cursor.execute('ALTER TABLE allocation_history RENAME TO allocation_history_old')
                conn.commit()
            except sqlite3.OperationalError:
                pass
            
            # 创建新的allocation_history表（不含network_id字段）
            _ = cursor.execute('''
            CREATE TABLE allocation_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip_address TEXT NOT NULL,
                action TEXT NOT NULL,
                hostname TEXT,
                description TEXT,
                performed_by TEXT,
                performed_at TEXT NOT NULL
            )
            ''')
            
            # 迁移数据
            _ = cursor.execute('''
            INSERT INTO allocation_history (id, ip_address, action, hostname, description, 
            performed_by, performed_at)
            SELECT id, ip_address, action, hostname, description, 
            performed_by, performed_at FROM allocation_history_old
            ''')
            conn.commit()
            
            # 删除旧表
            _ = cursor.execute('DROP TABLE allocation_history_old')
            conn.commit()
        else:
            # 创建allocation_history表（如果不存在）
            _ = cursor.execute('''
            CREATE TABLE IF NOT EXISTS allocation_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip_address TEXT NOT NULL,
                action TEXT NOT NULL,
                hostname TEXT,
                description TEXT,
                performed_by TEXT,
                performed_at TEXT NOT NULL
            )
            ''')
        
        # 创建backups表
        _ = cursor.execute('''
        CREATE TABLE IF NOT EXISTS backups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            backup_name TEXT NOT NULL,
            backup_path TEXT NOT NULL,
            backup_type TEXT NOT NULL,
            backup_time TEXT NOT NULL,
            network_count INTEGER NOT NULL,
            ip_count INTEGER NOT NULL
        )
        ''')
        
        # 创建索引（如果不存在）
        try:
            _ = cursor.execute('CREATE INDEX IF NOT EXISTS idx_ip_addresses_status ON ip_addresses(status)')
        except sqlite3.OperationalError:
            pass
        try:
            _ = cursor.execute('CREATE INDEX IF NOT EXISTS idx_ip_addresses_ip_address ON ip_addresses(ip_address)')
        except sqlite3.OperationalError:
            pass
        try:
            _ = cursor.execute('CREATE INDEX IF NOT EXISTS idx_ip_addresses_expiry_date ON ip_addresses(expiry_date)')
        except sqlite3.OperationalError:
            pass
        try:
            _ = cursor.execute('CREATE INDEX IF NOT EXISTS idx_allocation_history_action ON allocation_history(action)')
        except sqlite3.OperationalError:
            pass
        try:
            _ = cursor.execute('CREATE INDEX IF NOT EXISTS idx_allocation_history_performed_at ON allocation_history(performed_at)')
        except sqlite3.OperationalError:
            pass
        
        conn.commit()
        conn.close()
    
    def migrate_from_json(self, json_file: str = "ipam_data.json") -> tuple[bool, str]:
        """从JSON文件迁移数据
        
        Args:
            json_file: JSON数据文件路径
        
        Returns:
            tuple[bool, str]: (是否迁移成功, 错误信息)
        """
        if not os.path.exists(json_file):
            return False, "JSON文件不存在"
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data: dict[str, Any] = json.load(f)
        except Exception as e:
            return False, f"读取JSON文件失败: {str(e)}"
        
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        
        try:
            # 开始事务
            _ = conn.execute('BEGIN TRANSACTION')
            
            # 迁移networks
            networks_map: dict[str, int] = {}  # 用于映射网络地址到ID
            networks_data: dict[str, Any] = data.get('networks', {})
            for net_str, net_data in networks_data.items():
                if net_str and isinstance(net_data, dict):
                    network_created_at = net_data.get('created_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    network_created_at_str = str(network_created_at)
                    description = str(net_data.get('description', ''))
                    _ = cursor.execute('''
                    INSERT OR IGNORE INTO networks (network_address, description, created_at, updated_at)
                    VALUES (?, ?, ?, ?)
                    ''', (net_str, description, network_created_at_str, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            
            # 获取所有网络的ID
            _ = cursor.execute('SELECT id, network_address FROM networks')
            rows: list[tuple[int, str]] = cursor.fetchall()
            for row in rows:
                networks_map[str(row[1])] = int(row[0])
            
            # 迁移ip_addresses
            if networks_data:
                for net_str, net_data in networks_data.items():
                    if net_str and net_data:
                        network_id = networks_map.get(net_str)
                        if not network_id:
                            continue
                        
                        ip_addresses_data: dict[str, dict[str, Any]] = net_data.get('ip_addresses', {}) if isinstance(net_data.get('ip_addresses'), dict) else {}
                        if ip_addresses_data:
                            for ip_str, ip_data in ip_addresses_data.items():
                                if ip_str and ip_data:
                                    ip_created_at = ip_data.get('allocated_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                    if ip_created_at:
                                        _ = cursor.execute('''
                                        INSERT OR IGNORE INTO ip_addresses (ip_address, status, hostname, description, 
                                        allocated_at, allocated_by, expiry_date, created_at, updated_at)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                        ''', (ip_str, ip_data.get('status', 'released'), 
                                              ip_data.get('hostname', ''), ip_data.get('description', ''),
                                              ip_data.get('allocated_at'), ip_data.get('allocated_by'),
                                              ip_data.get('expiry_date'), ip_created_at, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            
            # 迁移allocation_history
            allocation_history_data_raw = data.get('allocation_history')
            allocation_history_data: list[object] = allocation_history_data_raw if isinstance(allocation_history_data_raw, list) else []
            if allocation_history_data:
                for history_item in allocation_history_data:
                    if isinstance(history_item, dict):
                        _ = cursor.execute('''
                        INSERT INTO allocation_history (ip_address, action, hostname, description, 
                        performed_by, performed_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ''', (history_item.get('ip_address'), history_item.get('action'),
                              history_item.get('hostname'), history_item.get('description'),
                              history_item.get('performed_by'), history_item.get('timestamp')))
            
            # 提交事务
            conn.commit()
            conn.close()
            return True, "数据迁移成功"
        except Exception as e:
            conn.rollback()
            conn.close()
            return False, f"数据迁移失败: {str(e)}"
    
    def get_most_specific_network(self, ip_address: str) -> dict[str, str | int] | None:
        """获取IP地址最具体的归属网络
        
        Args:
            ip_address: IP地址
            
        Returns:
            dict: 网络信息，或None
        """
        try:
            target_ip = ipaddress.ip_address(ip_address)
        except ValueError as e:
            print(f"IP地址格式错误: {str(e)}")
            return None
        
        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                _ = cursor.execute('SELECT id, network_address, description, created_at, updated_at FROM networks')
                network_rows: list[tuple[int, str, str | None, str | None, str | None]] = cursor.fetchall()
                most_specific_network: dict[str, str | int] | None = None
                max_prefix_len = 0
                
                for network_row in network_rows:
                    try:
                        if len(network_row) >= 5:
                            network_id: int = int(network_row[0])
                            network_address: str = str(network_row[1])
                            description: str = str(network_row[2]) if network_row[2] else ''
                            created_at: str = str(network_row[3]) if network_row[3] else ''
                            updated_at: str = str(network_row[4]) if network_row[4] else ''
                            network_obj = ipaddress.ip_network(network_address)
                            if target_ip in network_obj:
                                if network_obj.prefixlen > max_prefix_len:
                                    max_prefix_len = network_obj.prefixlen
                                    most_specific_network = {
                                        'id': network_id,
                                        'network_address': network_address,
                                        'description': description,
                                        'created_at': created_at,
                                        'updated_at': updated_at
                                    }
                    except (ValueError, TypeError) as e:
                        print(f"处理网络行时出错：{str(e)}")
                        continue
                    except Exception as e:
                        print(f"处理网络行时发生未知错误：{str(e)}")
                        continue
                
                return most_specific_network
        except sqlite3.Error as e:
            print(f"数据库错误: {str(e)}")
            return None
        except Exception as e:
            print(f"获取最具体网络失败: {str(e)}")
            return None
    
    def add_network(self, network_str: str, description: str = "", vlan: str = "") -> tuple[bool, str]:
        """添加网络
        
        Args:
            network_str: 网络地址（CIDR格式）
            description: 网络描述
            vlan: VLAN ID
        
        Returns:
            tuple[bool, str]: (是否添加成功, 错误信息)
        """
        try:
            # 验证网络格式
            if not network_str:
                return False, "网络地址不能为空"
            
            # 验证VLAN字段
            if vlan:
                if not vlan.isdigit():
                    return False, _('vlan_invalid_format')
                vlan_num: int = int(vlan)
                if vlan_num < 1 or vlan_num > 4094:
                    return False, _('vlan_out_of_range')
            
            try:
                ip_network = ipaddress.ip_network(network_str, strict=False)
                network_str = str(ip_network)
            except ValueError as e:
                return False, f"网络格式错误: {str(e)}"
            
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 检查网络是否已存在
                cursor.execute('SELECT id FROM networks WHERE network_address = ?', (network_str,))
                if cursor.fetchone():
                    return False, "网络已存在"
                
                # 添加网络
                created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute('''
                INSERT INTO networks (network_address, description, vlan, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ''', (network_str, description, vlan, created_at, created_at))
                
                _network_id = cursor.lastrowid
                
                conn.commit()
                return True, "网络添加成功"
        except Exception as e:
            print(f"添加网络失败: {str(e)}")
            return False, f"添加网络失败: {str(e)}"
    
    def allocate_ip(self, network_str: str, ip_address: str, hostname: str, description: str = "", expiry_date: str | None = None, record_id: int | None = None) -> tuple[bool, str]:
        """分配IP地址
        
        Args:
            network_str: 网络地址（CIDR格式）
            ip_address: 要分配的IP地址
            hostname: 主机名
            description: 描述
            expiry_date: 过期日期（ISO格式）
            record_id: 记录ID，用于指定要分配的特定记录
        
        Returns:
            tuple[bool, str]: (是否分配成功, 错误信息)
        """
        try:
            # 验证参数
            if not network_str:
                return False, "网络地址不能为空"
            if not ip_address:
                return False, "IP地址不能为空"
            is_valid, error_msg = IPAMValidator.validate_allocation_params(hostname, description)
            if not is_valid:
                return False, error_msg or "验证失败"
            
            # 验证IP地址格式
            try:
                _ = ipaddress.ip_address(ip_address)
            except ValueError as e:
                return False, f"IP地址格式错误: {str(e)}"
            
            # 验证并标准化过期日期格式
            if expiry_date:
                validated_date = self._validate_expiry_date(expiry_date)
                if validated_date is None:
                    return False, "过期日期格式错误，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS 格式"
                expiry_date = validated_date
            
            # 找到最具体的网络
            most_specific_network = self.get_most_specific_network(ip_address)
            if not most_specific_network:
                # 如果没有找到合适的网络，使用指定的网络
                with sqlite3.connect(self.db_file) as conn:
                    cursor = conn.cursor()
                    _ = cursor.execute('SELECT id FROM networks WHERE network_address = ?', (network_str,))
                    network_row = cursor.fetchone()
                    if not network_row or not isinstance(network_row, tuple) or len(network_row) < 1:
                        return False, "网络不存在"
                    _network_id = int(network_row[0])
            else:
                _network_id = int(most_specific_network['id'])
            
            with sqlite3.connect(self.db_file) as conn:
                conn.execute('BEGIN EXCLUSIVE')
                cursor = conn.cursor()
                
                try:
                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # 查找所有该IP地址的记录
                    _ = cursor.execute('SELECT id, status FROM ip_addresses WHERE ip_address = ?', (ip_address,))
                    ip_rows = cursor.fetchall()
                    
                    success_result: tuple[bool, str] | None = None
                    if record_id:
                        # 使用指定的记录ID
                        _ = cursor.execute('SELECT id, status FROM ip_addresses WHERE id = ? AND ip_address = ?', (record_id, ip_address))
                        specific_row = cursor.fetchone()
                        if specific_row and isinstance(specific_row, tuple) and len(specific_row) >= 2 and str(specific_row[1]) in ['released', 'reserved']:
                            ip_id = int(specific_row[0])
                            _ = cursor.execute('''
                            UPDATE ip_addresses SET status = ?, hostname = ?, description = ?, 
                            allocated_at = ?, allocated_by = ?, expiry_date = ?, updated_at = ?
                            WHERE id = ?
                            ''', ('allocated', hostname, description, now, 'admin', expiry_date, now, ip_id))
                        else:
                            # 检查是否有已分配的记录
                            allocated_rows = [row for row in ip_rows if isinstance(row, tuple) and len(row) >= 2 and str(row[1]) == 'allocated']
                            if allocated_rows:
                                # IP地址已被分配
                                success_result = (False, "IP地址已被分配")
                            else:
                                # 如果指定的记录不存在或状态不符合要求，创建新记录
                                _ = cursor.execute('''
                                INSERT INTO ip_addresses (ip_address, status, hostname, description, 
                                allocated_at, allocated_by, expiry_date, created_at, updated_at)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (ip_address, 'allocated', hostname, description, 
                                      now, 'admin', expiry_date, now, now))
                    else:
                        # 过滤出可用或保留状态的记录
                        available_rows = [row for row in ip_rows if isinstance(row, tuple) and len(row) >= 2 and str(row[1]) in ['released', 'reserved']]
                        
                        # 检查是否已有其他记录处于已分配状态
                        allocated_rows = [row for row in ip_rows if isinstance(row, tuple) and len(row) >= 2 and str(row[1]) == 'allocated']
                        if allocated_rows:
                            # IP地址已被分配
                            success_result = (False, "IP地址已被分配")
                        elif available_rows:
                            # 使用第一条可用记录
                            ip_id = int(available_rows[0][0])
                            _ = cursor.execute('''
                            UPDATE ip_addresses SET status = ?, hostname = ?, description = ?, 
                            allocated_at = ?, allocated_by = ?, expiry_date = ?, updated_at = ?
                            WHERE id = ?
                            ''', ('allocated', hostname, description, now, 'admin', expiry_date, now, ip_id))
                        else:
                            # 新的IP地址，插入记录
                            _ = cursor.execute('''
                            INSERT INTO ip_addresses (ip_address, status, hostname, description, 
                            allocated_at, allocated_by, expiry_date, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (ip_address, 'allocated', hostname, description, 
                                  now, 'admin', expiry_date, now, now))
                    
                    if success_result is None:
                        # 记录分配历史
                        _ = cursor.execute('''
                        INSERT INTO allocation_history (ip_address, action, hostname, description, 
                        performed_by, performed_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ''', (ip_address, 'allocate', hostname, description, 'admin', now))
                        conn.commit()
                        return (True, "IP地址分配成功")
                    else:
                        conn.rollback()
                        return success_result
                except Exception as e:
                    conn.rollback()
                    print(f"分配IP地址时出错: {str(e)}")
                    raise
        except Exception as e:
            print(f"分配IP地址失败: {str(e)}")
            return False, f"分配IP地址失败: {str(e)}"
    
    def get_network_ips(self, network_str: str) -> list[dict[str, str | int]]:
        """获取网络及其所有子网络的IP地址
        
        Args:
            network_str: 网络地址（CIDR格式）
        
        Returns:
            list[dict[str, str | int | None]]: IP地址列表
        """
        try:
            ip_network = ipaddress.ip_network(network_str, strict=False)
            
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 获取所有IP地址，按创建时间降序排序
                _ = cursor.execute('SELECT id, ip_address, status, hostname, mac_address, description, allocated_at, allocated_by, expiry_date, created_at, updated_at FROM ip_addresses ORDER BY created_at DESC')
                ips: list[tuple[int, str, str, str | None, str | None, str | None, str | None, str | None, str | None, str | None, str | None]] = cursor.fetchall()
                
                # 过滤出属于目标网络的IP地址（通过IP地址计算归属关系）
                relevant_ips: list[dict[str, str | int]] = []
                for ip_item in ips:
                    try:
                        if len(ip_item) >= 11:
                            ip_id: int = int(ip_item[0])
                            ip_address: str = str(ip_item[1])
                            status: str = str(ip_item[2])
                            hostname: str = str(ip_item[3]).strip() if ip_item[3] else ''
                            mac_address: str = str(ip_item[4]).strip() if ip_item[4] else ''
                            description: str = str(ip_item[5]).strip() if ip_item[5] else ''
                            allocated_at: str = str(ip_item[6]).strip() if ip_item[6] else ''
                            allocated_by: str = str(ip_item[7]).strip() if ip_item[7] else ''
                            expiry_date: str = str(ip_item[8]).strip() if ip_item[8] else ''
                            created_at: str = str(ip_item[9]).strip() if ip_item[9] else ''
                            updated_at: str = str(ip_item[10]).strip() if ip_item[10] else ''
                            ip_obj = ipaddress.ip_address(ip_address)
                            if ip_obj in ip_network:
                                relevant_ips.append({
                                    'id': ip_id,
                                    'ip_address': ip_address,
                                    'status': status,
                                    'hostname': hostname,
                                    'mac_address': mac_address,
                                    'description': description,
                                    'allocated_at': allocated_at,
                                    'allocated_by': allocated_by,
                                    'expiry_date': expiry_date,
                                    'created_at': created_at,
                                    'updated_at': updated_at
                                })
                    except Exception as e:
                        print(f"处理IP项时出错: {str(e)}")
                        continue
                
                return relevant_ips
        except Exception as e:
            print(f"获取网络IP地址失败: {str(e)}")
            return []
    
    def get_all_networks(self) -> list[dict[str, str | int]]:
        """获取所有网络
        
        Returns:
            list[dict[str, str | int]]: 网络列表
        """
        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                cursor.execute('SELECT id, network_address, description, vlan, created_at, updated_at FROM networks')
                network_rows = cursor.fetchall()
                
                # 一次性获取所有IP地址，避免多次数据库查询
                cursor.execute('SELECT ip_address FROM ip_addresses')
                all_ips = cursor.fetchall()
                
                # 预处理IP地址，避免重复创建ipaddress对象
                ip_objects: list[ipaddress.IPv4Address | ipaddress.IPv6Address] = []
                for ip_item in all_ips:
                    try:
                        if isinstance(ip_item, tuple) and len(ip_item) >= 1:
                            ip_address: str = str(ip_item[0])
                            ip_obj = ipaddress.ip_address(ip_address)
                            ip_objects.append(ip_obj)
                    except Exception as e:
                        print(f"处理IP地址时出错: {str(e)}")
                        continue
                
                network_list: list[dict[str, str | int]] = []
                for network_item in network_rows:
                    # 计算网络及其子网络的IP数量
                    try:
                        if isinstance(network_item, tuple) and len(network_item) >= 6:
                            network_id: int = int(network_item[0])
                            network_address: str = str(network_item[1])
                            description: str = str(network_item[2]).strip() if network_item[2] else ''
                            vlan: str = str(network_item[3]) if network_item[3] else ''
                            created_at: str = str(network_item[4]) if network_item[4] else ''
                            updated_at: str = str(network_item[5]) if network_item[5] else ''
                            ip_count: int = self.get_network_ip_count(network_address, ip_objects)
                            
                            network_list.append({
                                'id': network_id,
                                'network': network_address,
                                'description': description,
                                'vlan': vlan,
                                'created_at': created_at,
                                'updated_at': updated_at,
                                'ip_count': ip_count
                            })
                    except Exception as e:
                        print(f"处理网络项时出错: {str(e)}")
                        continue
                
                return network_list
        except Exception as e:
            print(f"获取所有网络失败: {str(e)}")
            return []
    
    def get_network_ip_count(self, network_address: str, _ip_objects: list[ipaddress.IPv4Address | ipaddress.IPv6Address] | None = None) -> int:
        """计算网络及其所有子网络的IP数量（只计算已分配和已保留的）
        
        Args:
            network_address: 网络地址（CIDR格式）
            ip_objects: 预处理的IP地址对象列表，避免重复计算
        
        Returns:
            int: IP地址数量
        """
        try:
            ip_network = ipaddress.ip_network(network_address, strict=False)
            
            # 计算属于目标网络的已分配和已保留的IP地址数量
            count: int = 0
            
            # 直接从数据库获取IP地址及其状态
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 获取所有IP地址及其状态
            _ = cursor.execute('SELECT ip_address, status FROM ip_addresses')
            ips = cursor.fetchall()
            
            conn.close()
            
            for ip_item in ips:
                try:
                    if isinstance(ip_item, tuple) and len(ip_item) >= 2:
                        ip_address_str: str = str(ip_item[0])
                        status: str = str(ip_item[1])
                        # 只计算已分配和已保留的IP地址
                        if status in ('allocated', 'reserved'):
                            ip_obj = ipaddress.ip_address(ip_address_str)
                            if ip_obj in ip_network:
                                count += 1
                except Exception:
                    pass
            
            return count
        except Exception:
            return 0
    
    def release_ip(self, ip_address: str, release_strategy: str = "all", record_id: int | None = None) -> tuple[bool, str]:
        """释放IP地址
        
        Args:
            ip_address: 要释放的IP地址
            release_strategy: 释放策略，可选值：
                - "all": 释放所有未释放的记录
                - "latest": 释放最新的记录
                - "oldest": 释放最早的记录
                - "allocated": 只释放已分配的记录
                - "reserved": 只释放已保留的记录
                - "specific": 释放指定ID的记录
            record_id: 当 release_strategy 为 "specific" 时，指定要释放的记录ID
        
        Returns:
            tuple[bool, str]: (是否释放成功, 错误信息)
        """
        try:
            # 验证参数
            if not ip_address:
                return False, "IP地址不能为空"
            
            # 验证释放策略
            valid_strategies = ["all", "latest", "oldest", "allocated", "reserved", "specific"]
            if release_strategy not in valid_strategies:
                return False, f"无效的释放策略，有效值：{', '.join(valid_strategies)}"
            
            # 验证特定记录ID
            if release_strategy == "specific" and record_id is None:
                return False, "使用 specific 策略时必须指定 record_id"
            
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 处理 specific 策略
                if release_strategy == "specific":
                    # 查找指定ID的记录
                    _ = cursor.execute('SELECT id, status, created_at FROM ip_addresses WHERE id = ? AND ip_address = ?', (record_id, ip_address))
                    specific_row = cursor.fetchone()
                    if not specific_row or not isinstance(specific_row, tuple) or len(specific_row) < 3:
                        return False, "指定的记录不存在或不属于该IP地址"
                    
                    status = str(specific_row[1])
                    if status == 'released':
                        return False, "IP地址未被分配或已被释放"
                    
                    rows_to_release = [specific_row]
                else:
                    # 查找所有该IP地址的记录
                    _ = cursor.execute('SELECT id, status, created_at FROM ip_addresses WHERE ip_address = ?', (ip_address,))
                    ip_rows = cursor.fetchall()
                    
                    if not ip_rows:
                        return False, "IP地址不存在"
                    
                    # 过滤出可释放的IP地址
                    allocatable_rows: list[tuple[int, str, str]] = []
                    for row in ip_rows:
                        try:
                            if isinstance(row, tuple) and len(row) >= 3:
                                status = str(row[1])
                                if status != 'released':
                                    # 根据策略过滤
                                    if release_strategy == "all":
                                        allocatable_rows.append(row)
                                    elif release_strategy == "allocated" and status == "allocated":
                                        allocatable_rows.append(row)
                                    elif release_strategy == "reserved" and status == "reserved":
                                        allocatable_rows.append(row)
                        except Exception as e:
                            print(f"处理IP行时出错: {str(e)}")
                            continue
                    
                    if not allocatable_rows:
                        return False, "IP地址未被分配或已被释放"
                    
                    # 根据策略选择要释放的记录
                    rows_to_release = []
                    if release_strategy == "latest":
                        # 按创建时间倒序排序，取第一条
                        allocatable_rows.sort(key=lambda x: str(x[2]) if len(x) >= 3 else "", reverse=True)
                        rows_to_release = [allocatable_rows[0]]
                    elif release_strategy == "oldest":
                        # 按创建时间正序排序，取第一条
                        allocatable_rows.sort(key=lambda x: str(x[2]) if len(x) >= 3 else "")
                        rows_to_release = [allocatable_rows[0]]
                    else:
                        # all, allocated, reserved 策略
                        rows_to_release = allocatable_rows
                
                # 释放选定的IP地址记录
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                released_count: int = 0
                
                for ip_row_item in rows_to_release:
                    try:
                        if len(ip_row_item) >= 2:
                            ip_id: int = int(ip_row_item[0])
                            
                            # 释放IP地址
                            _ = cursor.execute('UPDATE ip_addresses SET status = ?, updated_at = ? WHERE id = ?', ('released', now, ip_id))
                            
                            # 记录释放历史
                            _ = cursor.execute('''
                            INSERT INTO allocation_history (ip_address, action, performed_by, performed_at)
                            VALUES (?, ?, ?, ?)
                            ''', (ip_address, 'release', 'admin', now))
                            
                            released_count += 1
                    except Exception as e:
                        print(f"释放IP地址记录时出错: {str(e)}")
                        continue
                
                conn.commit()
                return True, f"成功释放 {released_count} 个IP地址记录"
        except Exception as e:
            print(f"释放IP地址失败: {str(e)}")
            return False, f"释放IP地址失败: {str(e)}"
    
    def delete_ip_by_id(self, ip_id: int) -> tuple[bool, str]:
        """根据ID删除IP地址记录
        
        Args:
            ip_id: IP地址记录ID
            
        Returns:
            Tuple[bool, str]: (是否删除成功, 错误信息)
        """
        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 获取IP地址信息用于历史记录
                _ = cursor.execute('SELECT ip_address FROM ip_addresses WHERE id = ?', (ip_id,))
                ip_row = cursor.fetchone()
                if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 1:
                    return False, "IP地址记录不存在"
                
                ip_address: str = str(ip_row[0])
                
                # 删除IP地址记录
                _ = cursor.execute('DELETE FROM ip_addresses WHERE id = ?', (ip_id,))
                
                # 记录删除历史
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                _ = cursor.execute('''
                INSERT INTO allocation_history (ip_address, action, performed_by, performed_at)
                VALUES (?, ?, ?, ?)
                ''', (ip_address, 'delete_conflict', 'admin', now))
                
                conn.commit()
                return True, "IP地址记录删除成功"
        except Exception as e:
            print(f"删除IP地址记录失败: {str(e)}")
            return False, f"删除IP地址记录失败: {str(e)}"
    
    def get_ip_info(self, ip_address: str) -> dict[str, str | int] | None:
        """获取IP地址信息
        
        Args:
            ip_address: IP地址
        
        Returns:
            dict[str, str | int] or None: IP地址信息，包含hostname, description等字段，失败返回None
        """
        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 查询IP地址信息
                _ = cursor.execute('SELECT hostname, description FROM ip_addresses WHERE ip_address = ?', (ip_address,))
                ip_row = cursor.fetchone()
                
                if ip_row and isinstance(ip_row, tuple) and len(ip_row) >= 2:
                    hostname: str = str(ip_row[0]) if ip_row[0] else ''
                    description: str = str(ip_row[1]) if ip_row[1] else ''
                    return {
                        'hostname': hostname,
                        'description': description
                    }
                return None
        except Exception as e:
            print(f"获取IP地址信息失败: {str(e)}")
            return None
    
    def get_all_ip_records(self, ip_address: str) -> list[dict[str, str | int | None]]:
        """获取指定IP地址的所有记录
        
        Args:
            ip_address: IP地址
        
        Returns:
            list[dict]: IP地址记录列表
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 查询该IP地址的所有记录
            _ = cursor.execute('''
            SELECT id, status, hostname, description, allocated_at, allocated_by, expiry_date, created_at, updated_at 
            FROM ip_addresses 
            WHERE ip_address = ?
            ORDER BY created_at DESC
            ''', (ip_address,))
            
            records = []
            for row in cursor.fetchall():
                if isinstance(row, tuple) and len(row) >= 9:
                    records.append({
                        'id': int(row[0]),
                        'ip_address': ip_address,
                        'status': str(row[1]),
                        'hostname': str(row[2]).strip() if row[2] else '',
                        'description': str(row[3]).strip() if row[3] else '',
                        'allocated_at': str(row[4]).strip() if row[4] else '',
                        'allocated_by': str(row[5]).strip() if row[5] else '',
                        'expiry_date': str(row[6]).strip() if row[6] else '',
                        'created_at': str(row[7]).strip() if row[7] else '',
                        'updated_at': str(row[8]).strip() if row[8] else ''
                    })
            
            conn.close()
            return records
        except Exception:
            return []
    
    def get_ip_record_by_id(self, record_id: int) -> dict[str, str | int] | None:
        """根据记录ID获取IP地址记录
        
        Args:
            record_id: 记录ID
        
        Returns:
            dict[str, str | int | None] or None: IP地址记录，失败返回None
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 查询指定ID的记录
            _ = cursor.execute('''
            SELECT ip_address, status, hostname, description, allocated_at, allocated_by, expiry_date, created_at, updated_at 
            FROM ip_addresses 
            WHERE id = ?
            ''', (record_id,))
            
            row = cursor.fetchone()
            if row and isinstance(row, tuple) and len(row) >= 9:
                record = {
                    'id': record_id,
                    'ip_address': str(row[0]),
                    'status': str(row[1]),
                    'hostname': str(row[2]).strip() if row[2] else '',
                    'description': str(row[3]).strip() if row[3] else '',
                    'allocated_at': str(row[4]).strip() if row[4] else '',
                    'allocated_by': str(row[5]).strip() if row[5] else '',
                    'expiry_date': str(row[6]).strip() if row[6] else '',
                    'created_at': str(row[7]).strip() if row[7] else '',
                    'updated_at': str(row[8]).strip() if row[8] else ''
                }
                conn.close()
                return record
            conn.close()
            return None
        except Exception:
            return None
    
    def update_ip_record(self, record_id: int, hostname: str, mac_address: str, description: str, expiry_date: str | None = None) -> tuple[bool, str]:
        """更新IP地址记录

        Args:
            record_id: 记录ID
            hostname: 主机名
            mac_address: MAC地址
            description: 描述
            expiry_date: 过期日期

        Returns:
            tuple[bool, str]: (是否更新成功, 错误信息)
        """
        is_valid, error_msg = IPAMValidator.validate_allocation_params(hostname, description)
        if not is_valid:
            return False, error_msg or "验证失败"
        
        if expiry_date == "None" or expiry_date == "":
            expiry_date = None
        
        # 验证并标准化过期日期格式
        if expiry_date:
            validated_date = self._validate_expiry_date(expiry_date)
            if validated_date is None:
                return False, "过期日期格式错误，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS 格式"
            expiry_date = validated_date
        
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 检查记录是否存在
            _ = cursor.execute('SELECT ip_address, status FROM ip_addresses WHERE id = ?', (record_id,))
            record = cursor.fetchone()
            if not record or not isinstance(record, tuple) or len(record) < 2:
                conn.close()
                return False, "记录不存在"
            
            ip_address = str(record[0])
            _ = str(record[1])
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 更新记录
            _ = cursor.execute('''
            UPDATE ip_addresses 
            SET hostname = ?, mac_address = ?, description = ?, expiry_date = ?, updated_at = ?
            WHERE id = ?
            ''', (hostname, mac_address, description, expiry_date, now, record_id))
            
            # 检查是否有记录被更新
            if cursor.rowcount == 0:
                conn.close()
                return False, "记录不存在或未更新"
            
            # 记录更新历史
            _ = cursor.execute('''
            INSERT INTO allocation_history (ip_address, action, hostname, description, 
            performed_by, performed_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (ip_address, 'update_record', hostname, description, 'admin', now))
            
            conn.commit()
            conn.close()
            return True, "记录更新成功"
        except Exception as e:
            return False, f"更新记录失败: {str(e)}"
    
    def update_ip_info(self, ip_address: str, hostname: str, description: str) -> tuple[bool, str]:
        """更新IP地址信息
        
        Args:
            ip_address: IP地址
            hostname: 新的主机名
            description: 新的描述
        
        Returns:
            tuple[bool, str]: (是否更新成功, 错误信息)
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 检查IP地址是否存在
            _ = cursor.execute('SELECT id FROM ip_addresses WHERE ip_address = ?', (ip_address,))
            ip_row = cursor.fetchone()
            if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 1:
                conn.close()
                return False, "IP地址不存在"
            
            ip_id: int = int(ip_row[0])
            
            # 更新IP地址信息
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            _ = cursor.execute('''
            UPDATE ip_addresses 
            SET hostname = ?, description = ?, updated_at = ?
            WHERE id = ?
            ''', (hostname, description, now, ip_id))
            
            # 记录更新历史
            _ = cursor.execute('''
            INSERT INTO allocation_history (ip_address, action, hostname, description, 
            performed_by, performed_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (ip_address, 'update', hostname, description, 'admin', now))
            
            conn.commit()
            conn.close()
            return True, "IP地址信息更新成功"
        except Exception as e:
            return False, f"更新IP地址信息失败: {str(e)}"

    def update_ip_expiry(self, ip_address: str, expiry_date: str | None, record_id: int | None = None) -> tuple[bool, str]:
        """更新IP地址过期日期
        
        Args:
            ip_address: IP地址
            expiry_date: 过期日期（格式：YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS，传入None则清除过期日期）
            record_id: 记录ID，如果为None则更新第一条找到的记录
            
        Returns:
            tuple[bool, str]: (是否更新成功, 错误信息)
        """
        # 验证日期格式
        if expiry_date is not None:
            validated_date = self._validate_expiry_date(expiry_date)
            if validated_date is None:
                return False, "过期日期格式错误，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS 格式"
            expiry_date = validated_date
        
        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                if record_id:
                    # 根据记录ID更新
                    _ = cursor.execute('SELECT id, hostname, description FROM ip_addresses WHERE id = ? AND ip_address = ?', (record_id, ip_address))
                    ip_row = cursor.fetchone()
                else:
                    # 更新第一条找到的记录
                    _ = cursor.execute('SELECT id, hostname, description FROM ip_addresses WHERE ip_address = ?', (ip_address,))
                    ip_row = cursor.fetchone()
                
                if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 3:
                    return False, "IP地址不存在"
                
                ip_id: int = int(ip_row[0])
                hostname: str = str(ip_row[1]) if ip_row[1] else ''
                description: str = str(ip_row[2]) if ip_row[2] else ''
                
                # 更新IP地址过期日期
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                _ = cursor.execute('''
                UPDATE ip_addresses 
                SET expiry_date = ?, updated_at = ?
                WHERE id = ?
                ''', (expiry_date, now, ip_id))
                
                # 记录更新历史
                _ = cursor.execute('''
                INSERT INTO allocation_history (ip_address, action, hostname, description, 
                performed_by, performed_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (ip_address, 'update', hostname, description, 'admin', now))
                
                conn.commit()
                return True, "IP地址过期日期更新成功"
        except Exception as e:
            return False, f"更新IP地址过期日期失败: {str(e)}"
    
    def batch_update_ip_expiry(self, ip_addresses: list[str], expiry_date: str | None, record_ids: list[int] | None = None) -> tuple[bool, str, int]:
        """批量更新IP地址过期日期
        
        Args:
            ip_addresses: IP地址列表
            expiry_date: 过期日期（格式：YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS，传入None则清除过期日期）
            record_ids: 记录ID列表，用于更新特定记录
            
        Returns:
            Tuple: (bool, str, int) - (是否更新成功, 错误信息, 更新的IP数量)
        """
        # 验证日期格式
        if expiry_date is not None:
            validated_date = self._validate_expiry_date(expiry_date)
            if validated_date is None:
                return False, "过期日期格式错误，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS 格式", 0
            expiry_date = validated_date
        
        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                updated_count = 0
                
                if record_ids and len(record_ids) == len(ip_addresses):
                    # 使用记录ID更新特定记录
                    for i, record_id in enumerate(record_ids):
                        ip_address = ip_addresses[i]
                        # 检查记录是否存在
                        _ = cursor.execute('SELECT hostname, description FROM ip_addresses WHERE id = ?', (record_id,))
                        ip_row = cursor.fetchone()
                        if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 2:
                            continue
                        
                        hostname_val = str(ip_row[0]) if ip_row[0] else ''
                        description_val = str(ip_row[1]) if ip_row[1] else ''
                        
                        # 更新IP地址过期日期
                        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        _ = cursor.execute('''
                        UPDATE ip_addresses 
                        SET expiry_date = ?, updated_at = ?
                        WHERE id = ?
                        ''', (expiry_date, now, record_id))
                        
                        # 记录更新历史
                        _ = cursor.execute('''
                        INSERT INTO allocation_history (ip_address, action, hostname, description, 
                        performed_by, performed_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ''', (ip_address, 'batch_update', hostname_val, description_val, 'admin', now))
                        
                        updated_count += 1
                else:
                    # 未提供记录ID，更新每条IP地址的第一条记录
                    for ip_address in ip_addresses:
                        # 查找该IP地址的第一条记录
                        _ = cursor.execute('SELECT id, hostname, description FROM ip_addresses WHERE ip_address = ? LIMIT 1', (ip_address,))
                        ip_row = cursor.fetchone()
                        if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 3:
                            continue
                        
                        ip_id: int = int(ip_row[0])
                        hostname_val: str = str(ip_row[1]) if ip_row[1] else ''
                        description_val: str = str(ip_row[2]) if ip_row[2] else ''
                        
                        # 更新IP地址过期日期
                        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        _ = cursor.execute('''
                        UPDATE ip_addresses 
                        SET expiry_date = ?, updated_at = ?
                        WHERE id = ?
                        ''', (expiry_date, now, ip_id))
                        
                        # 记录更新历史
                        _ = cursor.execute('''
                        INSERT INTO allocation_history (ip_address, action, hostname, description, 
                        performed_by, performed_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ''', (ip_address, 'batch_update', hostname_val, description_val, 'admin', now))
                        
                        updated_count += 1
                
                conn.commit()
                return True, f"成功更新 {updated_count} 个IP地址的过期日期", updated_count
        except Exception as e:
            return False, f"批量更新IP地址过期日期失败: {str(e)}", 0
    
    def export_data_to_xlsx(self, file_path: str, networks: list[str] | None = None,
                            include_ips: bool = False, ip_networks: list[str] | None = None) -> bool:
        """导出数据到xlsx文件（双sheet：网段表 + IP地址表）

        Args:
            file_path: 导出文件路径
            networks: 要导出的网段列表，None表示导出所有网段
            include_ips: 是否同时导出IP地址数据
            ip_networks: IP地址导出范围，None表示导出所有IP

        Returns:
            bool: 是否导出成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()

            all_networks = self.get_all_networks()
            if networks:
                filtered_networks = [net for net in all_networks if net['network'] in networks]
            else:
                filtered_networks = all_networks

            ws_net = wb.active
            assert ws_net is not None
            ws_net.title = _('network_sheet_name')
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')

            net_headers = [_('col_network'), _('col_description'), 'VLAN', _('col_created_at')]
            for col_idx, header in enumerate(net_headers, 1):
                cell = ws_net.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align

            for row_idx, net in enumerate(filtered_networks, 2):
                ws_net.cell(row=row_idx, column=1, value=net['network']).border = thin_border
                ws_net.cell(row=row_idx, column=2, value=net['description'] or '').border = thin_border
                ws_net.cell(row=row_idx, column=3, value=net.get('vlan', '') or '').border = thin_border
                ws_net.cell(row=row_idx, column=4, value=net['created_at'] or '').border = thin_border

            for col_idx in range(1, len(net_headers) + 1):
                ws_net.column_dimensions[chr(64 + col_idx)].width = 22

            if include_ips:
                ws_ip = wb.create_sheet(title=_('ip_sheet_name'))

                ip_headers = [_('col_ip_address'), _('col_status'), _('col_hostname'),
                              _('col_mac_address'), _('col_description'), _('col_expiry_date')]
                for col_idx, header in enumerate(ip_headers, 1):
                    cell = ws_ip.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align

                ip_data = self._get_ip_data_for_export(ip_networks)
                for row_idx, ip_item in enumerate(ip_data, 2):
                    ws_ip.cell(row=row_idx, column=1, value=ip_item['ip_address']).border = thin_border
                    ws_ip.cell(row=row_idx, column=2, value=ip_item['status'] or '').border = thin_border
                    ws_ip.cell(row=row_idx, column=3, value=ip_item['hostname'] or '').border = thin_border
                    ws_ip.cell(row=row_idx, column=4, value=ip_item['mac_address'] or '').border = thin_border
                    ws_ip.cell(row=row_idx, column=5, value=ip_item['description'] or '').border = thin_border
                    ws_ip.cell(row=row_idx, column=6, value=ip_item['expiry_date'] or '').border = thin_border

                for col_idx in range(1, len(ip_headers) + 1):
                    ws_ip.column_dimensions[chr(64 + col_idx)].width = 20

            wb.save(file_path)
            return True
        except Exception as e:
            print(f"导出数据失败: {str(e)}")
            return False

    def _get_ip_data_for_export(self, networks: list[str] | None = None) -> list[dict]:
        """获取用于导出的IP地址数据

        Args:
            networks: 网段列表，None表示获取所有IP地址

        Returns:
            list[dict]: IP地址数据列表
        """
        try:
            conn = sqlite3.connect(self.db_file)
            conn.row_factory = sqlite3.Row
            
            conn.create_function('ip_to_int', 1, ip_to_int)  # pyright: ignore[reportArgumentType]
            
            cursor = conn.cursor()

            if networks:
                ip_data = []
                ip_ranges = []
                
                for network_str in networks:
                    try:
                        network = ipaddress.ip_network(network_str, strict=False)
                        network_int = int(network.network_address)
                        broadcast_int = int(network.broadcast_address)
                        ip_ranges.append((network_int, broadcast_int))
                    except ValueError:
                        pass
                
                if ip_ranges:
                    placeholders = ' OR '.join(['(ip_to_int(ip_address) BETWEEN ? AND ?)'] * len(ip_ranges))
                    query = f'''
                    SELECT ip_address, status, hostname, mac_address, description, expiry_date
                    FROM ip_addresses
                    WHERE {placeholders}
                    '''
                    params = []
                    for start, end in ip_ranges:
                        params.extend([start, end])
                    
                    _ = cursor.execute(query, params)
                    for row in cursor.fetchall():
                        ip_data.append({
                            'ip_address': row['ip_address'],
                            'status': row['status'],
                            'hostname': row['hostname'],
                            'mac_address': row['mac_address'],
                            'description': row['description'],
                            'expiry_date': row['expiry_date']
                        })
            else:
                _ = cursor.execute('''
                SELECT ip_address, status, hostname, mac_address, description, expiry_date
                FROM ip_addresses
                ''')
                ip_data = []
                for row in cursor.fetchall():
                    ip_data.append({
                        'ip_address': row['ip_address'],
                        'status': row['status'],
                        'hostname': row['hostname'],
                        'mac_address': row['mac_address'],
                        'description': row['description'],
                        'expiry_date': row['expiry_date']
                    })

            conn.close()
            return ip_data
        except Exception as e:
            print(f"获取IP数据失败: {str(e)}")
            return []

    def parse_xlsx_for_import(self, file_path: str) -> dict:
        """解析xlsx文件用于导入比对

        Args:
            file_path: xlsx文件路径

        Returns:
            dict: {'networks': [...], 'ips': [...]} 包含网段和IP地址数据
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True)
            result = {'networks': [], 'ips': []}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(min_row=1, values_only=True))
                if not rows:
                    continue

                headers = [str(h).strip() if h else '' for h in rows[0]]

                if any(h in headers for h in [_('col_network'), 'Network', 'network']):
                    net_col = self._find_col(headers, [_('col_network'), 'Network', 'network'])
                    desc_col = self._find_col(headers, [_('col_description'), 'Description', 'description'])
                    vlan_col = self._find_col(headers, ['VLAN', 'vlan'])

                    for row in rows[1:]:
                        if not row or not any(row):
                            continue
                        network = str(row[net_col]).strip() if net_col is not None and net_col < len(row) and row[net_col] else ''
                        description = str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else ''
                        vlan = str(row[vlan_col]).strip() if vlan_col is not None and vlan_col < len(row) and row[vlan_col] else ''
                        if network:
                            result['networks'].append({
                                'network': network,
                                'description': description,
                                'vlan': vlan
                            })

                elif any(h in headers for h in [_('col_ip_address'), 'IP Address', 'ip_address']):
                    ip_col = self._find_col(headers, [_('col_ip_address'), 'IP Address', 'ip_address'])
                    status_col = self._find_col(headers, [_('col_status'), 'Status', 'status'])
                    hostname_col = self._find_col(headers, [_('col_hostname'), 'Hostname', 'hostname'])
                    mac_col = self._find_col(headers, [_('col_mac_address'), 'MAC Address', 'mac_address'])
                    desc_col = self._find_col(headers, [_('col_description'), 'Description', 'description'])
                    expiry_col = self._find_col(headers, [_('col_expiry_date'), 'Expiry Date', 'expiry_date'])

                    status_mapping = {
                        _('status_allocated'): 'allocated',
                        _('status_reserved'): 'reserved',
                        _('status_released'): 'released',
                        _('status_conflict'): 'conflict',
                        'allocated': 'allocated',
                        'reserved': 'reserved',
                        'released': 'released',
                        'conflict': 'conflict'
                    }

                    for row in rows[1:]:
                        if not row or not any(row):
                            continue
                        ip_address = str(row[ip_col]).strip() if ip_col is not None and ip_col < len(row) and row[ip_col] else ''
                        if ip_address:
                            status_str = str(row[status_col]).strip() if status_col is not None and status_col < len(row) and row[status_col] else ''
                            status = status_mapping.get(status_str, 'released')
                            result['ips'].append({
                                'ip_address': ip_address,
                                'status': status,
                                'hostname': str(row[hostname_col]).strip() if hostname_col is not None and hostname_col < len(row) and row[hostname_col] else '',
                                'mac_address': str(row[mac_col]).strip() if mac_col is not None and mac_col < len(row) and row[mac_col] else '',
                                'description': str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else '',
                                'expiry_date': str(row[expiry_col]).strip() if expiry_col is not None and expiry_col < len(row) and row[expiry_col] else ''
                            })

            wb.close()
            return result
        except Exception as e:
            print(f"解析xlsx文件失败: {str(e)}")
            return {'networks': [], 'ips': []}

    def _find_col(self, headers: list[str], names: list[str]) -> int | None:
        """在表头中查找列索引

        Args:
            headers: 表头列表
            names: 可能的列名列表

        Returns:
            int | None: 列索引，未找到返回None
        """
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    def validate_import_data(self, data: dict) -> dict:
        """验证导入数据，与现有数据比对

        Args:
            data: parse_xlsx_for_import返回的数据

        Returns:
            dict: {'networks': {'new': [...], 'existing': [...], 'invalid': [...]},
                   'ips': {'new': [...], 'existing': [...], 'invalid': [...]}}
        """
        result = {
            'networks': {'new': [], 'existing': [], 'invalid': []},
            'ips': {'new': [], 'existing': [], 'invalid': []}
        }

        existing_networks = set()
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT network_address FROM networks')
            for row in cursor.fetchall():
                existing_networks.add(row[0])
            conn.close()
        except Exception:
            pass

        for net in data.get('networks', []):
            network = net['network']
            try:
                ipaddress.ip_network(network, strict=False)
                if network in existing_networks:
                    result['networks']['existing'].append(net)
                else:
                    result['networks']['new'].append(net)
            except ValueError:
                result['networks']['invalid'].append(net)

        existing_ips = set()
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT ip_address FROM ip_addresses')
            for row in cursor.fetchall():
                existing_ips.add(row[0])
            conn.close()
        except Exception:
            pass

        valid_statuses = {'released', 'allocated', 'reserved'}
        for ip in data.get('ips', []):
            ip_address = ip['ip_address']
            try:
                ipaddress.ip_address(ip_address)
                status = (ip.get('status') or 'released').strip().lower()
                if status not in valid_statuses:
                    ip['error'] = _('invalid_status_value')
                    result['ips']['invalid'].append(ip)
                elif ip_address in existing_ips:
                    result['ips']['existing'].append(ip)
                else:
                    result['ips']['new'].append(ip)
            except ValueError:
                ip['error'] = _('invalid_ip_format')
                result['ips']['invalid'].append(ip)

        return result

    def import_validated_data(self, data: dict, import_networks: bool = True, import_ips: bool = True,
                              overwrite_existing: bool = False) -> dict:
        """导入验证后的数据

        Args:
            data: validate_import_data返回的比对结果
            import_networks: 是否导入网段数据
            import_ips: 是否导入IP地址数据
            overwrite_existing: 是否覆盖已存在的数据

        Returns:
            dict: 导入统计 {'networks_added': int, 'networks_updated': int,
                           'ips_added': int, 'ips_updated': int}
        """
        stats = {'networks_added': 0, 'networks_updated': 0, 'ips_added': 0, 'ips_updated': 0}
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()

                if import_networks:
                    for net in data['networks']['new']:
                        try:
                            ipaddress.ip_network(net['network'], strict=False)
                            _ = cursor.execute('''
                            INSERT INTO networks (network_address, description, vlan, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?)
                            ''', (net['network'], net.get('description', ''), net.get('vlan', ''), now, now))
                            stats['networks_added'] += 1
                        except (sqlite3.IntegrityError, ValueError):
                            pass

                    if overwrite_existing:
                        for net in data['networks']['existing']:
                            _ = cursor.execute('''
                            UPDATE networks SET description = ?, vlan = ?, updated_at = ?
                            WHERE network_address = ?
                            ''', (net.get('description', ''), net.get('vlan', ''), now, net['network']))
                            stats['networks_updated'] += 1

                if import_ips:
                    for ip in data['ips']['new']:
                        try:
                            ipaddress.ip_address(ip['ip_address'])
                            status = (ip.get('status') or 'released').strip().lower()
                            _ = cursor.execute('''
                            INSERT INTO ip_addresses (ip_address, status, hostname, mac_address, description,
                                   allocated_at, allocated_by, expiry_date, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                ip['ip_address'],
                                status,
                                ip.get('hostname', ''),
                                ip.get('mac_address', ''),
                                ip.get('description', ''),
                                ip.get('allocated_at', '') or now if status == 'allocated' else '',
                                ip.get('allocated_by', '') or 'admin',
                                self._validate_expiry_date(ip.get('expiry_date', '')),
                                now, now
                            ))
                            stats['ips_added'] += 1
                        except (sqlite3.IntegrityError, ValueError):
                            pass

                    if overwrite_existing:
                        for ip in data['ips']['existing']:
                            status = (ip.get('status') or 'released').strip().lower()
                            _ = cursor.execute('''
                            UPDATE ip_addresses SET status = ?, hostname = ?, mac_address = ?, description = ?,
                                   allocated_at = ?, allocated_by = ?, expiry_date = ?, updated_at = ?
                            WHERE ip_address = ?
                            ''', (
                                status,
                                ip.get('hostname', ''),
                                ip.get('mac_address', ''),
                                ip.get('description', ''),
                                ip.get('allocated_at', '') or now if status == 'allocated' else '',
                                ip.get('allocated_by', '') or 'admin',
                                self._validate_expiry_date(ip.get('expiry_date', '')),
                                now,
                                ip['ip_address']
                            ))
                            stats['ips_updated'] += 1

                conn.commit()
        except Exception as e:
            print(f"导入数据失败: {str(e)}")

        return stats

    def generate_template_xlsx(self, file_path: str, include_networks: bool = True, include_ips: bool = True) -> bool:
        """生成导入模板xlsx文件

        Args:
            file_path: 文件保存路径
            include_networks: 是否包含网段模板
            include_ips: 是否包含IP地址模板

        Returns:
            bool: 是否生成成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

            if not include_networks and not include_ips:
                return False

            wb = Workbook()
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')

            if include_networks:
                ws_net = wb.active
                assert ws_net is not None
                ws_net.title = _('network_sheet_name')
                net_headers = [_('col_network'), _('col_description'), 'VLAN']
                for col_idx, header in enumerate(net_headers, 1):
                    cell = ws_net.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align

                examples = [
                    ['192.168.1.0/24', _('template_network_desc1'), '10'],
                    ['10.0.0.0/8', _('template_network_desc2'), ''],
                ]
                for row_idx, example in enumerate(examples, 2):
                    for col_idx, val in enumerate(example, 1):
                        ws_net.cell(row=row_idx, column=col_idx, value=val).border = thin_border

                for col_idx in range(1, len(net_headers) + 1):
                    ws_net.column_dimensions[chr(64 + col_idx)].width = 22

            if include_ips:
                if include_networks:
                    ws_ip = wb.create_sheet(title=_('ip_sheet_name'))
                else:
                    ws_ip = wb.active
                    assert ws_ip is not None
                    ws_ip.title = _('ip_sheet_name')

                ip_headers = [_('col_ip_address'), _('col_status'), _('col_hostname'),
                              _('col_mac_address'), _('col_description'), _('col_expiry_date')]
                for col_idx, header in enumerate(ip_headers, 1):
                    cell = ws_ip.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align

                ip_examples = [
                    ['192.168.1.10', _('status_allocated'), 'server-01', '00:11:22:33:44:55',
                     _('template_ip_desc1'), '2026-12-31 23:59:59'],
                    ['192.168.1.11', _('status_reserved'), '', '', _('template_ip_desc2'), ''],
                    ['192.168.1.12', _('status_released'), '', '', '', ''],
                ]
                for row_idx, example in enumerate(ip_examples, 2):
                    for col_idx, val in enumerate(example, 1):
                        ws_ip.cell(row=row_idx, column=col_idx, value=val).border = thin_border

                for col_idx in range(1, len(ip_headers) + 1):
                    ws_ip.column_dimensions[chr(64 + col_idx)].width = 20

            if include_networks and include_ips:
                wb.active = wb.sheetnames.index(_('network_sheet_name'))

            wb.save(file_path)
            return True
        except Exception as e:
            print(f"生成模板失败: {str(e)}")
            return False
    
    def _validate_expiry_date(self, expiry_date: str | None) -> str | None:
        """验证并标准化过期日期格式
        
        Args:
            expiry_date: 过期日期字符串
            
        Returns:
            str | None: 标准化后的日期字符串（YYYY-MM-DD HH:MM:SS），验证失败返回None
        """
        if not expiry_date:
            return None
        
        expiry_date = expiry_date.strip()
        
        # 支持的日期格式列表
        date_formats = [
            "%Y-%m-%d %H:%M:%S",    # 2023-12-31 23:59:59
            "%Y-%m-%dT%H:%M:%S",    # 2023-12-31T23:59:59
            "%Y-%m-%d",             # 2023-12-31
            "%Y/%m/%d %H:%M:%S",    # 2023/12/31 23:59:59
            "%Y/%m/%d",             # 2023/12/31
        ]
        
        for fmt in date_formats:
            try:
                dt = datetime.strptime(expiry_date, fmt)
                # 如果只有日期部分（格式为%Y-%m-%d或%Y/%m/%d），添加时间部分为23:59:59
                if fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
                    return dt.strftime("%Y-%m-%d") + " 23:59:59"
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        
        return None

    def get_available_ips(self) -> list[dict[str, Any]]:
        """获取所有可用状态的IP地址
        
        Returns:
            list[dict]: 可用状态的IP地址列表
        """
        try:
            conn = sqlite3.connect(self.db_file)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # 查找所有可用状态的IP地址
            _ = cursor.execute('''
            SELECT id, ip_address, status, hostname, mac_address, description, expiry_date 
            FROM ip_addresses 
            WHERE status = ?
            ''', ('released',))
            available_ips = cursor.fetchall()
            
            conn.close()
            
            # 转换为字典列表
            result = []
            for ip in available_ips:
                result.append({
                    'id': ip['id'],
                    'ip_address': ip['ip_address'],
                    'status': ip['status'],
                    'hostname': ip['hostname'],
                    'mac_address': ip['mac_address'],
                    'description': ip['description'],
                    'expiry_date': ip['expiry_date']
                })
            
            return result
        except Exception:
            return []

    def cleanup_available_ips(self) -> tuple[bool, str]:
        """清理所有可用状态的IP地址
        
        Returns:
            tuple[bool, str]: (是否清理成功, 错误信息)
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 查找所有可用状态的IP地址
            _ = cursor.execute('SELECT id, ip_address FROM ip_addresses WHERE status = ?', ('released',))
            available_ips = cursor.fetchall()
            
            if not available_ips:
                conn.close()
                return True, "没有可用状态的IP地址需要清理"
            
            # 记录清理历史
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for ip_item in available_ips:
                if isinstance(ip_item, tuple) and len(ip_item) >= 2:
                    ip_address: str = str(ip_item[1])
                    # 记录清理历史
                    _ = cursor.execute('''
                    INSERT INTO allocation_history (ip_address, action, performed_by, performed_at)
                    VALUES (?, ?, ?, ?)
                    ''', (ip_address, 'cleanup', 'admin', now))
            
            # 删除可用状态的IP地址
            _ = cursor.execute('DELETE FROM ip_addresses WHERE status = ?', ('released',))
            
            conn.commit()
            conn.close()
            return True, f"成功清理 {len(available_ips)} 个可用状态的IP地址"
        except Exception as e:
            return False, f"清理可用IP地址失败: {str(e)}"
    
    def get_ip_status(self, ip_address: str) -> str:
        """获取IP地址状态
        
        Args:
            ip_address: IP地址
        
        Returns:
            str: IP地址状态（allocated, reserved, available）
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            _ = cursor.execute('SELECT status FROM ip_addresses WHERE ip_address = ?', (ip_address,))
            result = cursor.fetchone()
            
            conn.close()
            
            if result and isinstance(result, tuple) and len(result) >= 1:
                status: str = str(result[0])
                return status
            return 'released'
        except Exception:
            return 'released'
    
    def get_expired_ips(self) -> list[dict[str, str | int | None]]:
        """获取所有过期的IP地址
        
        Returns:
            list[dict[str, str | int | None]]: 过期IP地址列表
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 获取所有过期的IP地址，包括allocated和reserved状态
            _ = cursor.execute('''
            SELECT id, ip_address, status, hostname, description, allocated_at, allocated_by, expiry_date, mac_address 
            FROM ip_addresses 
            WHERE expiry_date < ? AND status IN ('allocated', 'reserved')
            ''', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
            
            expired_ips: list[dict[str, str | int | None]] = []
            for row in cursor.fetchall():
                if isinstance(row, tuple) and len(row) >= 9:
                    ip_id: int = int(row[0])
                    ip_address: str = str(row[1])
                    status: str = str(row[2])
                    hostname: str = str(row[3]).strip() if row[3] else ''
                    description: str = str(row[4]).strip() if row[4] else ''
                    allocated_at: str = str(row[5]).strip() if row[5] else ''
                    _allocated_by: str = str(row[6]).strip() if row[6] else ''
                    expiry_date: str = str(row[7]).strip() if row[7] else ''
                    mac_address: str = str(row[8]).strip() if row[8] else ''
                    expired_ips.append({
                        'id': ip_id,
                        'ip_address': ip_address,
                        'status': status,
                        'hostname': hostname,
                        'description': description,
                        'allocated_at': allocated_at,
                        'expiry_date': expiry_date,
                        'mac_address': mac_address
                    })
            
            conn.close()
            return expired_ips
        except Exception:
            return []
    
    def auto_release_expired_ips(self) -> tuple[bool, str, int]:
        """自动释放过期的IP地址
        
        Returns:
            tuple[bool, str, int]: (是否成功, 错误信息, 释放的IP数量)
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 获取过期的IP地址
            expired_ips = self.get_expired_ips()
            released_count = 0
            
            for ip in expired_ips:
                # 释放IP地址
                _ = cursor.execute('UPDATE ip_addresses SET status = ?, updated_at = ? WHERE id = ?', 
                             ('released', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ip['id']))
                
                # 记录释放历史
                _ = cursor.execute('''
                INSERT INTO allocation_history (ip_address, action, hostname, description, 
                performed_by, performed_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (ip['ip_address'], 'auto_release', ip['hostname'], 
                      ip['description'], 'system', datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                
                released_count += 1
            
            conn.commit()
            conn.close()
            
            if released_count > 0:
                return True, f"成功释放 {released_count} 个过期IP地址", released_count
            else:
                return True, "没有过期IP地址需要释放", 0
        except Exception as e:
            return False, f"自动释放过期IP地址失败: {str(e)}", 0
    
    def reserve_ip(self, network_str: str, ip_address: str, hostname: str, description: str = "", expiry_date: str | None = None, record_id: int | None = None) -> tuple[bool, str]:
        """保留IP地址
        
        Args:
            network_str: 网络地址（CIDR格式）
            ip_address: 要保留的IP地址
            hostname: 主机名
            description: 描述
            expiry_date: 过期日期
            record_id: 记录ID，用于指定要保留的特定记录
        
        Returns:
            tuple[bool, str]: (是否保留成功, 错误信息)
        """
        try:
            # 验证参数
            if not network_str:
                return False, "网络地址不能为空"
            if not ip_address:
                return False, "IP地址不能为空"
            is_valid, error_msg = IPAMValidator.validate_allocation_params(hostname, description)
            if not is_valid:
                return False, error_msg or "验证失败"
            
            # 验证IP地址格式
            try:
                _ = ipaddress.ip_address(ip_address)
            except ValueError as e:
                return False, f"IP地址格式错误: {str(e)}"
            
            # 验证并标准化过期日期格式
            if expiry_date:
                validated_date = self._validate_expiry_date(expiry_date)
                if validated_date is None:
                    return False, "过期日期格式错误，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS 格式"
                expiry_date = validated_date
            
            # 找到最具体的网络
            most_specific_network = self.get_most_specific_network(ip_address)
            if not most_specific_network:
                # 如果没有找到合适的网络，使用指定的网络
                conn = sqlite3.connect(self.db_file)
                cursor = conn.cursor()
                _ = cursor.execute('SELECT id FROM networks WHERE network_address = ?', (network_str,))
                network_row = cursor.fetchone()
                if not network_row or not isinstance(network_row, tuple) or len(network_row) < 1:
                    conn.close()
                    return False, "网络不存在"
                _network_id = int(network_row[0])
                conn.close()
            else:
                _network_id = int(most_specific_network['id'])
            
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if record_id:
                # 使用指定的记录ID
                _ = cursor.execute('SELECT id, status FROM ip_addresses WHERE id = ? AND ip_address = ?', (record_id, ip_address))
                specific_row = cursor.fetchone()
                if specific_row and isinstance(specific_row, tuple) and len(specific_row) >= 2 and str(specific_row[1]) == 'released':
                    ip_id = int(specific_row[0])
                    _ = cursor.execute('UPDATE ip_addresses SET status = ?, hostname = ?, description = ?, updated_at = ? WHERE id = ?', 
                                 ('reserved', hostname, description, now, ip_id))
                else:
                    # 允许在已分配状态下创建保留记录
                    # 创建新的IP地址记录（允许多个保留记录存在）
                    _ = cursor.execute('''
                    INSERT INTO ip_addresses (ip_address, status, hostname, description, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ''', (ip_address, 'reserved', hostname, description, now, now))
            else:
                # 查找所有该IP地址的记录
                _ = cursor.execute('SELECT id, status FROM ip_addresses WHERE ip_address = ?', (ip_address,))
                ip_rows = cursor.fetchall()
                
                # 过滤出可用状态的记录（允许在已分配状态下创建新的保留记录）
                available_rows = [row for row in ip_rows if isinstance(row, tuple) and len(row) >= 2 and str(row[1]) == 'released']
                
                if available_rows:
                    # 使用第一条可用记录
                    ip_id = int(available_rows[0][0])
                    _ = cursor.execute('UPDATE ip_addresses SET status = ?, hostname = ?, description = ?, updated_at = ? WHERE id = ?', 
                                 ('reserved', hostname, description, now, ip_id))
                else:
                    # 创建新的IP地址记录
                    _ = cursor.execute('''
                    INSERT INTO ip_addresses (ip_address, status, hostname, description, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ''', (ip_address, 'reserved', hostname, description, now, now))
            
            # 记录保留历史
            _ = cursor.execute('''
            INSERT INTO allocation_history (ip_address, action, hostname, description, performed_by, performed_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (ip_address, 'reserve', hostname, description, 'admin', now))
            
            conn.commit()
            conn.close()
            return True, "IP地址保留成功"
        except Exception as e:
            return False, f"保留IP地址失败: {str(e)}"

    def update_network_description(self, network_str: str, description: str) -> tuple[bool, str]:
        """更新网络描述
        
        Args:
            network_str: 网络地址（CIDR格式）
            description: 新的描述信息
        
        Returns:
            tuple[bool, str]: (是否更新成功, 错误信息)
        """
        try:
            if not network_str:
                return False, "网络地址不能为空"
            
            try:
                ip_network = ipaddress.ip_network(network_str, strict=False)
                network_str = str(ip_network)
            except ValueError as e:
                return False, f"网络格式错误: {str(e)}"
            
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 检查网络是否存在
                _ = cursor.execute('SELECT id FROM networks WHERE network_address = ?', (network_str,))
                network_row = cursor.fetchone()
                if not network_row or not isinstance(network_row, tuple) or len(network_row) < 1:
                    return False, "网络不存在"
                
                # 更新网络描述
                _ = cursor.execute('UPDATE networks SET description = ?, updated_at = datetime("now") WHERE id = ?', 
                             (description, int(network_row[0])))
                
                conn.commit()
            return True, "网络描述更新成功"
        except Exception as e:
            return False, f"更新网络描述失败: {str(e)}"
    
    def update_network_vlan(self, network_str: str, vlan: str) -> tuple[bool, str]:
        """更新网络VLAN
        
        Args:
            network_str: 网络地址（CIDR格式）
            vlan: 新的VLAN ID
        
        Returns:
            tuple[bool, str]: (是否更新成功, 错误信息)
        """
        try:
            if not network_str:
                return False, "网络地址不能为空"
            
            # 验证VLAN字段
            if vlan:
                if not vlan.isdigit():
                    return False, _('vlan_invalid_format')
                vlan_num: int = int(vlan)
                if vlan_num < 1 or vlan_num > 4094:
                    return False, _('vlan_out_of_range')
            
            try:
                ip_network = ipaddress.ip_network(network_str, strict=False)
                network_str = str(ip_network)
            except ValueError as e:
                return False, f"网络格式错误: {str(e)}"
            
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 检查网络是否存在
                cursor.execute('SELECT id FROM networks WHERE network_address = ?', (network_str,))
                network_row = cursor.fetchone()
                if not network_row or not isinstance(network_row, tuple) or len(network_row) < 1:
                    return False, "网络不存在"
                
                # 更新网络VLAN
                cursor.execute('UPDATE networks SET vlan = ?, updated_at = datetime("now") WHERE id = ?', 
                             (vlan, int(network_row[0])))
                
                conn.commit()
            return True, "网络VLAN更新成功"
        except Exception as e:
            return False, f"更新网络VLAN失败: {str(e)}"
    
    def update_network(self, old_network: str, new_network: str) -> tuple[bool, str]:
        """更新网络地址
        
        Args:
            old_network: 旧的网络地址（CIDR格式）
            new_network: 新的网络地址（CIDR格式）
        
        Returns:
            tuple[bool, str]: (是否更新成功, 错误信息)
        """
        try:
            if not old_network or not new_network:
                return False, "网络地址不能为空"
            
            # 验证旧网络格式
            try:
                old_ip_network = ipaddress.ip_network(old_network, strict=False)
                old_network_str = str(old_ip_network)
            except ValueError as e:
                return False, f"旧网络格式错误: {str(e)}"
            
            # 验证新网络格式
            try:
                new_ip_network = ipaddress.ip_network(new_network, strict=False)
                new_network_str = str(new_ip_network)
            except ValueError as e:
                return False, f"新网络格式错误: {str(e)}"
            
            with sqlite3.connect(self.db_file) as conn:
                cursor = conn.cursor()
                
                # 检查旧网络是否存在
                _ = cursor.execute('SELECT id FROM networks WHERE network_address = ?', (old_network_str,))
                network_row = cursor.fetchone()
                if not network_row or not isinstance(network_row, tuple) or len(network_row) < 1:
                    return False, "旧网络不存在"
                
                # 检查新网络是否已存在
                _ = cursor.execute('SELECT id FROM networks WHERE network_address = ? AND id != ?', (new_network_str, int(network_row[0])))
                existing_row = cursor.fetchone()
                if existing_row:
                    return False, "新网络地址已存在"
                
                # 更新网络地址
                _ = cursor.execute('UPDATE networks SET network_address = ?, updated_at = datetime("now") WHERE id = ?', 
                             (new_network_str, int(network_row[0])))
                
                conn.commit()
            return True, "网络地址更新成功"
        except Exception as e:
            return False, f"更新网络地址失败: {str(e)}"

    def check_ip_conflicts(self, ip_address: str) -> list[dict[str, str | int | None]]:
        """检测IP地址冲突
        
        Args:
            ip_address: 要检测的IP地址
        
        Returns:
            list[dict]: 冲突的IP地址记录列表
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 查找该IP地址的所有记录
            _ = cursor.execute('''
            SELECT id, status, hostname, mac_address, description, allocated_at, expiry_date, created_at, updated_at 
            FROM ip_addresses 
            WHERE ip_address = ?
            ''', (ip_address,))
            
            conflict_records = []
            for row in cursor.fetchall():
                if isinstance(row, tuple) and len(row) >= 9:
                    conflict_records.append({
                        'id': int(row[0]),
                        'ip_address': ip_address,
                        'status': str(row[1]),
                        'hostname': str(row[2]).strip() if row[2] else '',
                        'mac_address': str(row[3]).strip() if row[3] else '',
                        'description': str(row[4]).strip() if row[4] else '',
                        'allocated_at': str(row[5]).strip() if row[5] else '',
                        'expiry_date': str(row[6]).strip() if row[6] else '',
                        'created_at': str(row[7]).strip() if row[7] else '',
                        'updated_at': str(row[8]).strip() if row[8] else ''
                    })
            
            conn.close()
            
            # 只有当记录数大于1时才认为有冲突
            if len(conflict_records) > 1:
                return conflict_records
            else:
                return []
        except Exception:
            return []
    
    def resolve_ip_conflicts(self, ip_address: str, keep_record_id: int | None = None) -> tuple[bool, str]:
        """解决IP地址冲突
        
        Args:
            ip_address: 要解决冲突的IP地址
            keep_record_id: 要保留的记录ID，如果为None则保留最新的记录
        
        Returns:
            tuple[bool, str]: (是否解决成功, 错误信息)
        """
        try:
            # 检查是否存在冲突
            conflicts = self.check_ip_conflicts(ip_address)
            if not conflicts:
                return True, "没有冲突需要解决"
            
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            try:
                # 开始事务
                _ = conn.execute('BEGIN EXCLUSIVE')
                
                # 确定要保留的记录
                if keep_record_id:
                    # 保留指定的记录
                    keep_record = next((record for record in conflicts if record['id'] == keep_record_id), None)
                    if not keep_record:
                        conn.rollback()
                        conn.close()
                        return False, "指定的记录不存在"
                else:
                    # 保留最新的记录（按创建时间排序）
                    conflicts.sort(key=lambda x: x['created_at'] or '', reverse=True)
                    keep_record = conflicts[0]
                
                # 删除其他记录
                deleted_count = 0
                for record in conflicts:
                    if record['id'] != keep_record['id']:
                        # 记录删除历史
                        _ = cursor.execute('''
                        INSERT INTO allocation_history (ip_address, action, performed_by, performed_at)
                        VALUES (?, ?, ?, ?)
                        ''', (ip_address, 'delete_conflict', 'admin', datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                        
                        # 删除记录
                        _ = cursor.execute('DELETE FROM ip_addresses WHERE id = ?', (record['id'],))
                        deleted_count += 1
                
                conn.commit()
                conn.close()
                return True, f"成功解决冲突，保留了1个记录，删除了{deleted_count}个记录"
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
        except Exception as e:
            return False, f"解决冲突失败: {str(e)}"
    
    def remove_network(self, network_str: str) -> tuple[bool, str]:
        """移除网络
        
        Args:
            network_str: 网络地址（CIDR格式）
        
        Returns:
            tuple[bool, str]: (是否移除成功, 错误信息)
        """
        try:
            if not network_str:
                return False, "网络地址不能为空"
            
            try:
                ip_network = ipaddress.ip_network(network_str, strict=False)
                network_str = str(ip_network)
            except ValueError as e:
                return False, f"网络格式错误: {str(e)}"
            
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 检查网络是否存在
            _ = cursor.execute('SELECT id FROM networks WHERE network_address = ?', (network_str,))
            network_row = cursor.fetchone()
            if not network_row or not isinstance(network_row, tuple) or len(network_row) < 1:
                conn.close()
                return False, "网络不存在"
            
            network_id = int(network_row[0])
            
            # 检查网络是否有IP地址（通过IP地址计算归属关系）
            ip_network = ipaddress.ip_network(network_str, strict=False)
            _ = cursor.execute('SELECT ip_address FROM ip_addresses')
            ip_count = 0
            for row in cursor.fetchall():
                if isinstance(row, tuple) and len(row) >= 1:
                    try:
                        ip_obj = ipaddress.ip_address(str(row[0]))
                        if ip_obj in ip_network:
                            ip_count += 1
                    except ValueError:
                        continue
            if ip_count > 0:
                conn.close()
                return False, f"网络中存在 {ip_count} 个IP地址，请先删除这些IP地址"
            
            # 移除网络
            _ = cursor.execute('DELETE FROM networks WHERE id = ?', (network_id,))
            
            conn.commit()
            conn.close()
            return True, "网络移除成功"
        except Exception as e:
            return False, f"移除网络失败: {str(e)}"
    
    def get_overall_stats(self) -> dict[str, int]:
        """获取整体统计信息
        
        Returns:
            dict[str, int]: 整体统计信息
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 获取网络总数
            _ = cursor.execute('SELECT COUNT(*) FROM networks')
            total_networks_result = cursor.fetchone()
            total_networks = int(total_networks_result[0]) if total_networks_result and isinstance(total_networks_result, tuple) and len(total_networks_result) >= 1 else 0
            
            # 获取IP总数
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses')
            total_ips_result = cursor.fetchone()
            total_ips = int(total_ips_result[0]) if total_ips_result and isinstance(total_ips_result, tuple) and len(total_ips_result) >= 1 else 0
            
            # 获取已分配IP数
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses WHERE status = ?', ('allocated',))
            allocated_ips_result = cursor.fetchone()
            allocated_ips = int(allocated_ips_result[0]) if allocated_ips_result and isinstance(allocated_ips_result, tuple) and len(allocated_ips_result) >= 1 else 0
            
            # 获取已保留IP数
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses WHERE status = ?', ('reserved',))
            reserved_ips_result = cursor.fetchone()
            reserved_ips = int(reserved_ips_result[0]) if reserved_ips_result and isinstance(reserved_ips_result, tuple) and len(reserved_ips_result) >= 1 else 0
            
            # 获取可用IP数
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses WHERE status = ?', ('released',))
            available_ips_result = cursor.fetchone()
            available_ips = int(available_ips_result[0]) if available_ips_result and isinstance(available_ips_result, tuple) and len(available_ips_result) >= 1 else 0
            
            # 获取过期IP数，包括allocated和reserved状态
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses WHERE expiry_date < ? AND status IN ("allocated", "reserved")', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
            expired_ips_result = cursor.fetchone()
            expired_ips = int(expired_ips_result[0]) if expired_ips_result and isinstance(expired_ips_result, tuple) and len(expired_ips_result) >= 1 else 0
            
            # 获取即将过期IP数（未来7天内过期）
            seven_days_later = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses WHERE expiry_date >= ? AND expiry_date <= ? AND status IN ("allocated", "reserved")', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), seven_days_later))
            expiring_ips_result = cursor.fetchone()
            expiring_ips = int(expiring_ips_result[0]) if expiring_ips_result and isinstance(expiring_ips_result, tuple) and len(expiring_ips_result) >= 1 else 0
            
            # 获取已命名IP数（hostname不为空）
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses WHERE hostname IS NOT NULL AND hostname != ""')
            named_ips_result = cursor.fetchone()
            named_ips = int(named_ips_result[0]) if named_ips_result and isinstance(named_ips_result, tuple) and len(named_ips_result) >= 1 else 0
            
            # 获取IPv4和IPv6网络数
            _ = cursor.execute('SELECT network_address FROM networks')
            network_rows = cursor.fetchall()
            ipv4_networks = 0
            ipv6_networks = 0
            
            for network in network_rows:
                try:
                    if isinstance(network, tuple) and len(network) >= 1:
                        ip_network = ipaddress.ip_network(str(network[0]))
                        if ip_network.version == 4:
                            ipv4_networks += 1
                        elif ip_network.version == 6:
                            ipv6_networks += 1
                except Exception:
                    pass
            
            # 获取IPv4和IPv6地址数
            _ = cursor.execute('SELECT ip_address FROM ip_addresses')
            ip_rows = cursor.fetchall()
            ipv4_ips = 0
            ipv6_ips = 0
            
            for ip_row in ip_rows:
                try:
                    if isinstance(ip_row, tuple) and len(ip_row) >= 1:
                        ip_addr = ipaddress.ip_address(str(ip_row[0]))
                        if ip_addr.version == 4:
                            ipv4_ips += 1
                        elif ip_addr.version == 6:
                            ipv6_ips += 1
                except Exception:
                    pass
            
            # 获取VLAN数量（不重复的VLAN）
            _ = cursor.execute('SELECT DISTINCT vlan FROM networks WHERE vlan IS NOT NULL AND vlan != ""')
            vlan_rows = cursor.fetchall()
            vlan_count = len(vlan_rows)
            
            conn.close()
            
            stats = {
                'total_networks': total_networks,
                'total_ips': total_ips,
                'allocated_ips': allocated_ips,
                'reserved_ips': reserved_ips,
                'available_ips': available_ips,
                'expired_ips': expired_ips,
                'expiring_ips': expiring_ips,
                'named_ips': named_ips,
                'ipv4_networks': ipv4_networks,
                'ipv6_networks': ipv6_networks,
                'ipv4_ips': ipv4_ips,
                'ipv6_ips': ipv6_ips,
                'vlan_count': vlan_count
            }
            
            return stats
        except Exception as e:
            print(f"获取统计信息失败: {str(e)}")
            return {
                'total_networks': 0,
                'total_ips': 0,
                'allocated_ips': 0,
                'reserved_ips': 0,
                'available_ips': 0,
                'expired_ips': 0,
                'expiring_ips': 0,
                'named_ips': 0,
                'ipv4_networks': 0,
                'ipv6_networks': 0,
                'ipv4_ips': 0,
                'ipv6_ips': 0,
                'vlan_count': 0
            }
    
    def list_backups(self) -> list[dict[str, Any]]:
        """获取所有备份记录
        
        Returns:
            list[dict]: 备份记录列表
        """
        try:
            # 从backups表中获取所有备份记录
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 从backups表中获取所有备份记录
            _ = cursor.execute('SELECT id, backup_name, backup_path, backup_time, network_count, ip_count FROM backups ORDER BY backup_time DESC')
            backup_rows = cursor.fetchall()
            
            conn.close()
            
            # 转换为列表字典格式，符合windows_app.py中的预期
            backup_list: list[dict[str, Any]] = []
            backup_names = set()
            
            # 添加数据库中的备份记录
            for backup_row in backup_rows:
                if isinstance(backup_row, tuple) and len(backup_row) >= 6:
                    _ = int(backup_row[0])  # backup_id 未使用
                    backup_name: str = str(backup_row[1])
                    backup_path: str = str(backup_row[2])
                    backup_time: str = str(backup_row[3])
                    network_count: int = int(backup_row[4])
                    ip_count: int = int(backup_row[5])
                    backup_list.append({
                        'filename': backup_name,
                        'file_path': backup_path,
                        'info': {
                            'backup_time': backup_time,
                            'network_count': network_count,
                            'ip_count': ip_count
                        }
                    })
                    backup_names.add(backup_name)
            
            # 检查备份目录中的实际文件，确保所有备份文件都能在列表中显示
            if os.path.exists(self.backup_dir):
                for filename in os.listdir(self.backup_dir):
                    if filename.endswith('.db'):
                        backup_path = os.path.join(self.backup_dir, filename)
                        if filename not in backup_names:
                            # 对于没有数据库记录的备份文件，尝试从文件名提取时间戳
                            # 文件名格式：backup_20260408_183506.db
                            import re
                            match = re.match(r'backup_([0-9]{8})_([0-9]{6})\.db', filename)
                            if match:
                                date_part = match.group(1)
                                time_part = match.group(2)
                                backup_time = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]} {time_part[:2]}:{time_part[2:4]}:{time_part[4:6]}"
                            else:
                                # 如果文件名格式不符合预期，使用文件修改时间
                                mtime = os.path.getmtime(backup_path)
                                backup_time = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
                            
                            # 从备份文件中读取网络数和IP数
                            network_count = 0
                            ip_count = 0
                            try:
                                backup_conn = sqlite3.connect(backup_path)
                                backup_cursor = backup_conn.cursor()
                                
                                backup_cursor.execute('SELECT COUNT(*) FROM networks')
                                network_result = backup_cursor.fetchone()
                                if network_result and isinstance(network_result, tuple) and len(network_result) >= 1:
                                    network_count = int(network_result[0])
                                
                                backup_cursor.execute('SELECT COUNT(*) FROM ip_addresses')
                                ip_result = backup_cursor.fetchone()
                                if ip_result and isinstance(ip_result, tuple) and len(ip_result) >= 1:
                                    ip_count = int(ip_result[0])
                                
                                backup_conn.close()
                            except Exception as e:
                                print(f"读取备份文件信息失败: {str(e)}")
                            
                            # 使用相对路径
                            relative_backup_path = os.path.join('ipam_backups', filename)
                            
                            # 添加到备份列表
                            backup_list.append({
                                'filename': filename,
                                'file_path': relative_backup_path,
                                'info': {
                                    'backup_time': backup_time,
                                    'network_count': network_count,
                                    'ip_count': ip_count
                                }
                            })
            
            # 按备份时间排序
            backup_list.sort(key=lambda x: x['info']['backup_time'], reverse=True)
            
            return backup_list
        except Exception as e:
            print(f"获取备份列表失败: {str(e)}")
            return []
    
    def backup_data(self, backup_name: str = "", backup_type: str = 'manual', frequency: str = 'manual') -> str:
        """备份数据
        
        Args:
            backup_name: 备份文件名（可选）
            backup_type: 备份类型（manual, auto）
            frequency: 备份频率（manual, daily, weekly, monthly）
        
        Returns:
            str: 备份文件路径
        """
        _ = frequency
        try:
            # 确保备份目录存在
            if not os.path.exists(self.backup_dir):
                os.makedirs(self.backup_dir)
            
            # 生成备份文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if not backup_name:
                backup_name = f"backup_{timestamp}.db"
            else:
                # 如果提供了自定义名称，确保包含时间戳和扩展名
                if not backup_name.endswith('.db'):
                    backup_name = f"{backup_name}_{timestamp}.db"
                else:
                    # 如果已经有扩展名，在扩展名前添加时间戳
                    name_part, ext_part = os.path.splitext(backup_name)
                    backup_name = f"{name_part}_{timestamp}{ext_part}"
            backup_path = os.path.join(self.backup_dir, backup_name)
            
            # 复制数据库文件
            import shutil
            _ = shutil.copy2(self.db_file, backup_path)
            
            # 存储相对路径
            relative_backup_path = os.path.join('ipam_backups', backup_name)
            
            # 获取网络和IP数量
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 获取网络数量
            _ = cursor.execute('SELECT COUNT(*) FROM networks')
            network_count_result = cursor.fetchone()
            network_count = int(network_count_result[0]) if network_count_result and isinstance(network_count_result, tuple) and len(network_count_result) >= 1 else 0
            
            # 获取IP数量
            _ = cursor.execute('SELECT COUNT(*) FROM ip_addresses')
            ip_count_result = cursor.fetchone()
            ip_count = int(ip_count_result[0]) if ip_count_result and isinstance(ip_count_result, tuple) and len(ip_count_result) >= 1 else 0
            
            # 记录备份信息到backups表
            backup_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            _ = cursor.execute('''
            INSERT INTO backups (backup_name, backup_path, backup_type, backup_time, network_count, ip_count)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (backup_name, relative_backup_path, backup_type, backup_time, network_count, ip_count))
            
            conn.commit()
            conn.close()
            
            # 更新备份配置文件（只记录必要的备份时间，避免冗余）
            try:
                backup_config_file = os.path.join(self.backup_dir, 'backup_config.json')
                # 只保留必要字段，移除冗余的频率配置
                config = {'last_backup_time': datetime.now().isoformat()}
                
                # 保存配置文件
                with open(backup_config_file, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"更新备份配置失败: {str(e)}")
            
            return backup_path
        except Exception as e:
            print(f"备份失败: {str(e)}")
            return ""
    
    def _get_absolute_backup_path(self, backup_path: str) -> str:
        """将备份路径转换为绝对路径

        Args:
            backup_path: 备份文件路径（相对或绝对）

        Returns:
            str: 绝对路径
        """
        if os.path.isabs(backup_path):
            return backup_path
        return os.path.join(self.app_dir, backup_path)

    def restore_data(self, backup_path: str) -> bool:
        """从备份恢复数据

        Args:
            backup_path: 备份文件路径

        Returns:
            bool: 是否恢复成功
        """
        try:
            # 获取绝对路径
            absolute_backup_path = self._get_absolute_backup_path(backup_path)
            
            if not os.path.exists(absolute_backup_path):
                print(f"备份文件不存在: {absolute_backup_path}")
                return False
            
            # 复制备份文件到当前数据库
            import shutil
            _ = shutil.copy2(absolute_backup_path, self.db_file)
            
            return True
        except Exception as e:
            print(f"恢复数据失败: {str(e)}")
            return False

    def delete_backup(self, backup_path: str) -> bool:
        """删除备份记录

        Args:
            backup_path: 备份文件路径

        Returns:
            bool: 是否删除成功
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 从backups表中删除指定路径的备份记录
            _ = cursor.execute('DELETE FROM backups WHERE backup_path = ?', (backup_path,))
            conn.commit()
            conn.close()
            
            return True
        except Exception as e:
            print(f"删除备份记录失败: {str(e)}")
            return False
    
    def get_last_backup_time(self) -> datetime | None:
        """获取最后一次备份的时间
        
        从 backup_config.json 读取，避免恢复数据库后备份时间回退的问题
        
        Returns:
            datetime | None: 最后一次备份的时间，或None
        """
        try:
            backup_config_file = os.path.join(self.backup_dir, 'backup_config.json')
            if not os.path.exists(backup_config_file):
                return None
            
            with open(backup_config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            if 'last_backup_time' in config:
                last_backup_time_str = config['last_backup_time']
                try:
                    return datetime.fromisoformat(last_backup_time_str)
                except (ValueError, TypeError):
                    return None
            return None
        except Exception as e:
            print(f"获取最后备份时间失败: {str(e)}")
            return None
    
    def get_network_by_id(self, net_id: int) -> dict[str, str | int] | None:
        """根据ID获取网络信息
        
        Args:
            network_id: 网络ID
        
        Returns:
            dict[str, str | int] | None: 网络信息，或None
        """
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # 查询网络信息
            _ = cursor.execute('SELECT id, network_address, description, created_at, updated_at FROM networks WHERE id = ?', (net_id,))
            network_row = cursor.fetchone()
            
            conn.close()
            
            if network_row and isinstance(network_row, tuple) and len(network_row) >= 5:
                network_id: int = int(network_row[0])
                network_address: str = str(network_row[1])
                description: str = str(network_row[2]) if network_row[2] else ''
                created_at: str = str(network_row[3]) if network_row[3] else ''
                updated_at: str = str(network_row[4]) if network_row[4] else ''
                
                return {
                    'id': network_id,
                    'network_address': network_address,
                    'description': description,
                    'created_at': created_at,
                    'updated_at': updated_at
                }
            return None
        except Exception as e:
            print(f"获取网络信息失败: {str(e)}")
            return None

    def migrate_ip(self, ip_address: str, new_network_str: str, record_id: int | None = None) -> tuple[bool, str]:
        """迁移单个IP地址到新网络
        
        由于已移除network_id字段，迁移操作现在仅验证IP地址是否在目标网络范围内。
        
        Args:
            ip_address: 要迁移的IP地址
            new_network_str: 目标网络地址（CIDR格式）
            record_id: 记录ID，如果为None则迁移第一条找到的记录
        
        Returns:
            tuple[bool, str]: (是否迁移成功, 错误信息)
        """
        try:
            if not ip_address:
                return False, "IP地址不能为空"
            if not new_network_str:
                return False, "目标网络地址不能为空"
            
            # 验证IP地址格式
            try:
                ip_obj = ipaddress.ip_address(ip_address)
            except ValueError as e:
                return False, f"IP地址格式错误: {str(e)}"
            
            # 验证目标网络格式
            try:
                new_network = ipaddress.ip_network(new_network_str, strict=False)
                new_network_str = str(new_network)
            except ValueError as e:
                return False, f"目标网络格式错误: {str(e)}"
            
            # 检查IP是否在目标网络内
            if ip_obj not in new_network:
                return False, "IP地址不在目标网络范围内"
            
            # 检查目标网络是否存在
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            try:
                _ = cursor.execute('SELECT id FROM networks WHERE network_address = ?', (new_network_str,))
                new_network_row = cursor.fetchone()
                if not new_network_row or not isinstance(new_network_row, tuple) or len(new_network_row) < 1:
                    conn.close()
                    return False, "目标网络不存在"
                
                # 获取要迁移的记录
                if record_id:
                    _ = cursor.execute('SELECT id, status, hostname, description FROM ip_addresses WHERE id = ? AND ip_address = ?', 
                                 (record_id, ip_address))
                    ip_row = cursor.fetchone()
                else:
                    _ = cursor.execute('SELECT id, status, hostname, description FROM ip_addresses WHERE ip_address = ?', 
                                 (ip_address,))
                    ip_row = cursor.fetchone()
                
                if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 4:
                    conn.close()
                    return False, "IP地址记录不存在"
                
                _ip_id = int(ip_row[0])
                _status = str(ip_row[1])
                hostname = str(ip_row[2]) if ip_row[2] else ''
                description = str(ip_row[3]) if ip_row[3] else ''
                
                # 记录迁移历史（由于已移除network_id，仅记录IP地址变化）
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                _ = cursor.execute('''
                INSERT INTO allocation_history (ip_address, action, hostname, description, performed_by, performed_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (ip_address, 'migrate', hostname, description, 'admin', now))
                
                conn.commit()
                conn.close()
                return True, "IP地址迁移成功"
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
        except Exception as e:
            return False, f"迁移IP地址失败: {str(e)}"

    def batch_migrate_ips(self, ip_records: list[dict[str, str | int]], new_network_str: str) -> tuple[bool, str, int]:
        """批量迁移IP地址到新网络
        
        迁移时会修改IP地址的值，使其成为目标网络中的有效IP地址。
        
        Args:
            ip_records: 要迁移的IP记录列表，每个记录包含 'id' 和 'ip_address'
            new_network_str: 目标网络地址（CIDR格式）
        
        Returns:
            tuple[bool, str, int]: (是否迁移成功, 错误信息, 成功迁移的数量)
        """
        try:
            if not ip_records:
                return False, "没有要迁移的IP地址", 0
            if not new_network_str:
                return False, "目标网络地址不能为空", 0
            
            # 验证目标网络格式
            try:
                new_network = ipaddress.ip_network(new_network_str, strict=False)
                new_network_str = str(new_network)
            except ValueError as e:
                return False, f"目标网络格式错误: {str(e)}", 0
            
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            try:
                # 获取目标网络中已存在的所有IP地址（通过IP地址计算归属关系）
                _ = cursor.execute('SELECT ip_address FROM ip_addresses')
                existing_ips = set()
                for row in cursor.fetchall():
                    if isinstance(row, tuple) and len(row) >= 1:
                        ip_str = str(row[0])
                        try:
                            ip_obj = ipaddress.ip_address(ip_str)
                            if ip_obj in new_network:
                                existing_ips.add(ip_str)
                        except ValueError:
                            continue
                
                # 生成目标网络中可用的IP地址列表（优先按主机位排序）
                available_new_ips = []
                for ip in new_network.hosts():
                    ip_str = str(ip)
                    if ip_str not in existing_ips:
                        available_new_ips.append(ip_str)
                
                if not available_new_ips:
                    conn.close()
                    return False, "目标网络中没有可用的IP地址", 0
                
                migrated_count = 0
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ip_mapping = []  # 记录IP地址的映射关系
                
                skipped_ips = []
                _new_ip_index = 0
                for record in ip_records:
                    try:
                        ip_id = int(record['id'])
                        old_ip_address = str(record['ip_address'])
                        
                        # 获取当前记录信息
                        _ = cursor.execute('SELECT status, hostname, description, mac_address, allocated_at, allocated_by, expiry_date FROM ip_addresses WHERE id = ?', (ip_id,))
                        ip_row = cursor.fetchone()
                        if not ip_row or not isinstance(ip_row, tuple) or len(ip_row) < 7:
                            skipped_ips.append(f"{old_ip_address} (记录不存在)")
                            continue
                        
                        _status = str(ip_row[0])
                        hostname = str(ip_row[1]) if ip_row[1] else ''
                        desc = str(ip_row[2]) if ip_row[2] else ''
                        _mac_address = str(ip_row[3]) if ip_row[3] else ''
                        _allocated_at = str(ip_row[4]) if ip_row[4] else ''
                        _allocated_by = str(ip_row[5]) if ip_row[5] else ''
                        _expiry_date = str(ip_row[6]).strip() if ip_row[6] else ''
                        
                        # 检查原IP是否已经在目标网络中（通过IP地址计算）
                        try:
                            old_ip_obj = ipaddress.ip_address(old_ip_address)
                            if old_ip_obj in new_network:
                                skipped_ips.append(f"{old_ip_address} (已在目标网络)")
                                continue
                        except ValueError:
                            skipped_ips.append(f"{old_ip_address} (IP格式错误)")
                            continue
                        
                        # 获取新的IP地址（优先保持主机位不变）
                        # 提取原IP的主机位（最后一位）
                        old_host_part = int(old_ip_address.split('.')[-1])
                        
                        # 构建期望的新IP地址
                        new_network_parts = str(new_network.network_address).split('.')[:-1]
                        desired_new_ip = '.'.join(new_network_parts) + '.' + str(old_host_part)
                        
                        # 检查期望的IP是否在目标网络中且可用
                        new_ip_address = None
                        if desired_new_ip in available_new_ips:
                            new_ip_address = desired_new_ip
                        else:
                            # 如果期望的IP不可用，查找第一个可用的IP
                            for ip in available_new_ips:
                                if int(ip.split('.')[-1]) >= old_host_part:
                                    new_ip_address = ip
                                    break
                            # 如果没有找到，从头开始找
                            if new_ip_address is None and available_new_ips:
                                new_ip_address = available_new_ips[0]
                        
                        if new_ip_address is None:
                            skipped_ips.append(f"{old_ip_address} (目标网络可用IP不足)")
                            continue
                        
                        # 从可用列表中移除已分配的IP
                        available_new_ips.remove(new_ip_address)
                        
                        # 更新IP地址记录（修改IP地址）
                        _ = cursor.execute('''
                        UPDATE ip_addresses 
                        SET ip_address = ?, updated_at = ? 
                        WHERE id = ?
                        ''', (new_ip_address, now, ip_id))
                        
                        # 删除其他具有相同原IP地址的记录，确保原地址完全消失
                        _ = cursor.execute('DELETE FROM ip_addresses WHERE ip_address = ? AND status = ? AND id != ?', (old_ip_address, 'released', ip_id))
                        
                        # 记录迁移历史
                        _ = cursor.execute('''
                        INSERT INTO allocation_history (ip_address, action, hostname, description, performed_by, performed_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ''', (new_ip_address, 'batch_migrate', hostname, f"{desc} (迁移自: {old_ip_address})", 'admin', now))
                        
                        ip_mapping.append(f"{old_ip_address} -> {new_ip_address}")
                        migrated_count += 1
                    except Exception as e:
                        logging.error(f"迁移IP {record.get('ip_address')} 失败: {str(e)}")
                        skipped_ips.append(f"{record.get('ip_address', '未知')} (迁移异常: {str(e)})")
                        continue
                
                conn.commit()
                conn.close()
                
                if migrated_count > 0:
                    message = f"成功迁移 {migrated_count} 个IP地址"
                    if ip_mapping:
                        message += f": {', '.join(ip_mapping[:5])}"
                        if len(ip_mapping) > 5:
                            message += f"...(共{len(ip_mapping)}个)"
                    if skipped_ips:
                        message += f"，跳过 {len(skipped_ips)} 个IP"
                    return True, message, migrated_count
                else:
                    reason = "没有IP地址被迁移"
                    if skipped_ips:
                        reason = f"所有IP均被跳过: {', '.join(skipped_ips[:5])}"
                        if len(skipped_ips) > 5:
                            reason += f"...(共{len(skipped_ips)}个)"
                    return False, reason, 0
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
        except Exception as e:
            return False, f"批量迁移IP地址失败: {str(e)}", 0
