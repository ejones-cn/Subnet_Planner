#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
子网规划师应用程序 - 主窗口
"""



# 标准库
import base64
import csv
import datetime
from datetime import timedelta
import math
import os
import json
import re
import sqlite3
import sys
import traceback
from collections import deque
from io import BytesIO
from typing import Any, Optional

# 自定义模块
from splash_screen import SplashScreen

# 第三方库
import ipaddress
import tkinter as tk
import tkinter.font as tkfont
import tkinter.messagebox as messagebox
from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.styles import Font, Alignment  # type: ignore
from PIL import Image, ImageTk
from tkinter import ttk, filedialog

# 本地模块
from version import get_version
from i18n import _, set_language, get_language  # _ 是翻译函数，用于国际化
from config_manager import get_config
from ip_subnet_calculator import format_large_number
from ip_subnet_calculator import (
    split_subnet,
    ip_to_int,
    int_to_ip,
    get_subnet_info,
    suggest_subnet_planning,
    merge_subnets,
    get_ip_info,
    range_to_cidr,
    check_subnet_overlap,
    handle_ip_subnet_error,
)
from export_utils import ExportUtils
from chart_utils import draw_text_with_stroke, draw_distribution_chart
from ipam_sqlite import IPAMSQLite
from services.table_column_manager import TableColumnManager
from services.history_repository import HistoryRepository
from services.history_sqlite import HistorySQLite
from services.ipam_repository import IPAMRepository
from services.subnet_split_service import SubnetSplitService
from services.subnet_planning_service import SubnetPlanningService
from services.ip_query_service import IPQueryService
from services.validation_service import ValidationService
from services.crypto_service import get_crypto_service, CryptoService
from validators import IPAMValidator
from visualization import NetworkTopologyVisualizer
from style_manager import (
    init_style_manager,
    update_styles,
    get_current_font_settings,
    get_style_manager,
)
from font_config import (
    get_pin_button_font_size,
    get_function_button_font_size,
    get_info_bar_font_size,
    get_move_button_font,
)

# 尝试导入DateEntry，如果失败则设置为None
DateEntry = None
try:
    from tkcalendar import DateEntry
except ImportError:
    print("Warning: tkcalendar module not found, date picker will be disabled")

__version__ = get_version()


def fix_date_entry_for_modal(date_entry, dialog_toplevel):
    """修复DateEntry在模态对话框中的grab冲突和焦点管理问题
    
    根本原因：
    1. 对话框的grab_set()阻止日历弹窗接收鼠标事件
    2. DateEntry自带的_on_focus_out_cal在焦点移到日历内部按钮时会错误关闭日历
    
    解决方案：
    1. 替换drop_down方法，在日历显示前释放grab，关闭时也正确处理
    2. 替换_on_focus_out_cal方法，正确判断焦点是否仍在日历弹窗内
    3. 日历关闭后恢复grab
    
    Args:
        date_entry: DateEntry控件实例
        dialog_toplevel: 模态对话框的Toplevel窗口（None表示非模态上下文）
    """
    def _is_focus_in_calendar():
        """检查焦点是否在日历弹窗内部（包括年/月导航按钮）"""
        try:
            focus_widget = date_entry.focus_get()
            if focus_widget is None:
                return False
            top_cal = date_entry._top_cal
            widget = focus_widget
            while widget is not None:
                if widget == top_cal:
                    return True
                try:
                    widget = widget.master
                except Exception:
                    break
            if focus_widget == date_entry:
                return True
            return False
        except Exception:
            return False
    
    def _custom_on_focus_out_cal(event):
        """替换DateEntry自带的_on_focus_out_cal，正确处理日历内部焦点转移"""
        if _is_focus_in_calendar():
            return
        try:
            date_entry._top_cal.withdraw()
            date_entry.state(['!pressed'])
        except Exception:
            pass
    
    date_entry._calendar.unbind('<FocusOut>')
    date_entry._calendar.bind('<FocusOut>', _custom_on_focus_out_cal)
    
    if dialog_toplevel is not None:
        _original_drop_down = date_entry.drop_down
        
        def _custom_drop_down():
            if date_entry._calendar.winfo_ismapped():
                # 日历已打开，直接关闭
                try:
                    dialog_toplevel.grab_release()
                except Exception:
                    pass
                date_entry._top_cal.withdraw()
                date_entry.state(['!pressed'])
            else:
                # 日历已关闭，先释放grab再打开
                try:
                    dialog_toplevel.grab_release()
                except Exception:
                    pass
                _original_drop_down()
        
        date_entry.drop_down = _custom_drop_down
        
        def _restore_grab(event=None):
            def _do_restore():
                try:
                    if not date_entry._top_cal.winfo_ismapped():
                        dialog_toplevel.grab_set()
                except Exception:
                    try:
                        dialog_toplevel.grab_set()
                    except Exception:
                        pass
            dialog_toplevel.after(50, _do_restore)
        
        date_entry._top_cal.bind('<Unmap>', _restore_grab)


# 初始化IPAM
def init_ipam():
    """初始化IPAM模块"""
    return IPAMSQLite()


def get_ipam():
    """获取IPAM实例"""
    return init_ipam()


# 全局变量定义
SCALE_FACTOR = 1.0  # DPI缩放因子，默认1.0（96 DPI）
DPI_MODE = None  # DPI模式标记

# 对话框尺寸常量定义
# 标准对话框尺寸
DIALOG_INFO_SMALL = (350, 180)      # 信息对话框
DIALOG_INFO_MEDIUM = (400, 200)     # 警告对话框
DIALOG_INFO_LARGE = (450, 220)      # 错误对话框
DIALOG_CONFIRM = (500, 180)         # 确认对话框
DIALOG_INPUT = (400, 250)           # 输入对话框
DIALOG_SELECT = (500, 300)          # 选择对话框

# 复杂对话框尺寸
DIALOG_SMALL = (400, 200)           # 小型对话框（简单表单）
DIALOG_MEDIUM = (500, 300)          # 中型对话框（中等复杂度）
DIALOG_LARGE = (700, 500)           # 大型对话框（复杂表单）
DIALOG_TABLE = (800, 600)           # 数据表格对话框

# IPAM网络管理按钮数量
NETWORK_MANAGEMENT_BUTTON_COUNT = 6


def create_bordered_entry(parent, border_color="#a9a9a9", **kwargs):
    """创建带边框的Entry组件
    
    Args:
        parent: 父容器
        border_color: 边框颜色，默认为灰色(#a9a9a9)
        **kwargs: Entry的额外参数
        
    Returns:
        tuple: (border_frame, entry) 边框容器和Entry组件
    """
    # 创建带边框的容器
    border_frame = tk.Frame(parent, highlightbackground=border_color, 
                            highlightcolor=border_color, highlightthickness=1, bd=0)
    
    # 创建Entry组件，使用普通的tk.Entry而不是ttk.Entry，避免双重边框
    entry = tk.Entry(border_frame, bd=0, relief="flat", highlightthickness=0, **kwargs)
    # 修改：移除右侧padding，避免边框叠加形成竖线
    entry.pack(fill="both", expand=True, padx=(1, 0), pady=1)
    
    return border_frame, entry


class DialogBase:
    """对话框基类，提供统一的布局和交互模式"""
    
    def __init__(self, parent, title, width, height, resizable=False, modal=True):
        """初始化对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            width: 对话框宽度
            height: 对话框高度
            resizable: 是否允许调整大小
            modal: 是否为模态对话框
        """
        self.parent = parent
        self.title = title
        self.width = width
        self.height = height
        self.resizable = resizable
        self.modal = modal
        
        # 创建对话框
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.resizable(resizable, resizable)
        
        # 设置最小尺寸限制（仅当可调整大小时）
        if resizable:
            self.dialog.minsize(width, height)
        
        self.dialog.transient(parent)
        
        # 为对话框设置与主窗口相同的置顶属性
        if hasattr(parent, 'is_pinned'):
            self.dialog.attributes('-topmost', parent.is_pinned)
        
        if modal:
            self.dialog.grab_set()
        
        # 先隐藏对话框
        self.dialog.withdraw()
        
        # 绑定键盘快捷键
        self._bind_shortcuts()
        
        # 初始化布局
        self._init_layout()
        
        # 设置尺寸并居中
        self.dialog.geometry(f"{self.width}x{self.height}")
        self._center_dialog()
        
        # 对话框在show()方法中显示
    
    def _center_dialog(self):
        """居中对话框"""
        self.dialog.update_idletasks()
        dialog_width = self.dialog.winfo_width()
        dialog_height = self.dialog.winfo_height()
        
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        
        # 计算对话框的居中位置（不包含父窗口标题栏）
        title_bar_height = 30
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + title_bar_height + (parent_height - title_bar_height - dialog_height) // 2
        
        self.dialog.geometry(f"+{x}+{y}")
    
    def _init_layout(self):
        """初始化布局"""
        # 主框架
        self.main_frame = ttk.Frame(self.dialog, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 内容框架
        self.content_frame = ttk.Frame(self.main_frame)
        self.content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        # 配置内容框架的网格布局，使控件垂直居中
        self.content_frame.grid_rowconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(100, weight=1)
        
        # 按钮框架
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.pack(fill=tk.X, pady=(0, 0))
        
        # 配置按钮框架的内部布局，使用pack而不是grid
        self.button_frame.pack_propagate(True)
        
        # 配置按钮框架的网格布局，使按钮靠右对齐
        self.button_frame.grid_columnconfigure(0, weight=1)
        self.button_frame.grid_columnconfigure(1, weight=0)
        self.button_frame.grid_columnconfigure(2, weight=0)
    
    def _bind_shortcuts(self):
        """绑定键盘快捷键"""
        # 绑定Escape键关闭对话框
        self.dialog.bind('<Escape>', lambda e: self.dialog.destroy())
    
    def add_button(self, text, command, column=1, padx=5):
        """添加按钮
        
        Args:
            text: 按钮文本
            command: 按钮命令
            column: 按钮列位置
            padx: 按钮水平间距
        """
        button = ttk.Button(self.button_frame, text=text, command=command, width=10)
        button.pack(side=tk.RIGHT, padx=padx, pady=5)
        return button
    
    def show(self):
        """显示对话框并等待关闭

        Returns:
            对话框的返回值
        """
        # 显示对话框（此时所有控件已添加完毕）
        self.dialog.deiconify()
        self.dialog.focus_force()
        self.parent.wait_window(self.dialog)
        return getattr(self, 'result', None)
    
    def destroy(self):
        """销毁对话框"""
        self.dialog.destroy()
    
    def protocol(self, name, func):
        """设置窗口协议处理程序
        
        Args:
            name: 协议名称
            func: 处理函数
        """
        self.dialog.protocol(name, func)


class InfoDialog(DialogBase):
    """信息对话框，用于显示信息、警告或错误"""
    
    def __init__(self, parent, title, message, dialog_type="info"):
        """初始化信息对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            message: 对话框消息
            dialog_type: 对话框类型，可选值：info, warning, error
        """
        # 根据对话框类型设置不同的默认尺寸
        if dialog_type == "info":
            width, height = 350, 180
        elif dialog_type == "warning":
            width, height = 400, 200
        elif dialog_type == "error":
            width, height = 450, 220
        else:
            width, height = 350, 180
        
        super().__init__(parent, title, width, height, resizable=False, modal=True)
        
        self.message = message
        self.dialog_type = dialog_type
        
        # 初始化界面
        self._init_ui()
    
    def _init_ui(self):
        """初始化界面"""
        font_family, font_size = get_current_font_settings()
        
        # 添加消息文本
        msg_label = ttk.Label(
            self.content_frame, 
            text=self.message, 
            wraplength=self.width - 40, 
            font=(font_family, font_size)
        )
        msg_label.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 添加确定按钮
        def on_ok():
            self.result = True
            self.dialog.destroy()
        
        ok_btn = self.add_button(_('ok'), on_ok, column=1)
        
        # 绑定回车键
        self.dialog.bind('<Return>', lambda e: on_ok())
        
        # 设置焦点到确定按钮
        ok_btn.focus_force()


class ConfirmDialog(DialogBase):
    """确认对话框，用于需要用户确认操作的场景"""
    
    def __init__(self, parent, title, message, ok_text="是", cancel_text="否"):
        """初始化确认对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            message: 对话框消息
            ok_text: 确认按钮文本
            cancel_text: 取消按钮文本
        """
        # 设置默认尺寸
        width, height = 400, 180
        
        super().__init__(parent, title, width, height, resizable=False, modal=True)
        
        self.message = message
        self.ok_text = ok_text
        self.cancel_text = cancel_text
        
        # 初始化界面
        self._init_ui()
    
    def _init_ui(self):
        """初始化界面"""
        font_family, font_size = get_current_font_settings()
        
        # 添加消息文本
        msg_label = ttk.Label(
            self.content_frame, 
            text=self.message, 
            wraplength=self.width - 80, 
            font=(font_family, font_size)
        )
        msg_label.pack(fill=tk.BOTH, expand=True, pady=10, padx=20)
        
        # 添加取消按钮
        def on_cancel():
            self.result = False
            self.dialog.destroy()
        
        cancel_btn = self.add_button(self.cancel_text, on_cancel, column=2)
        
        # 添加确认按钮
        def on_ok():
            self.result = True
            self.dialog.destroy()
        
        ok_btn = self.add_button(self.ok_text, on_ok, column=1)
        
        # 绑定回车键（默认确认）和Escape键（默认取消）
        self.dialog.bind('<Return>', lambda e: on_ok())
        self.dialog.bind('<Escape>', lambda e: on_cancel())
        
        # 设置焦点到确认按钮
        ok_btn.focus_force()


class InputDialog(DialogBase):
    """输入对话框，用于收集用户输入的场景"""
    
    def __init__(self, parent, title, fields, width=400, height=250):
        """初始化输入对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            fields: 输入字段列表，每个字段包含：
                {
                    'label': 字段标签,
                    'default': 默认值,
                    'validator': 验证函数 (可选),
                    'validate_message': 验证失败消息 (可选),
                    'width': 输入框宽度 (可选)
                }
            width: 对话框宽度
            height: 对话框高度
        """
        super().__init__(parent, title, width, height, resizable=False, modal=True)
        
        self.fields = fields
        self.entries = {}
        self.validation_errors = {}
        
        # 初始化界面
        self._init_ui()
    
    def _init_ui(self):
        """初始化界面"""
        font_family, font_size = get_current_font_settings()
        
        # 配置内容框架的网格布局
        self.content_frame.grid_columnconfigure(0, weight=0)
        self.content_frame.grid_columnconfigure(1, weight=1)
        
        # 添加输入字段
        for i, field in enumerate(self.fields):
            # 添加标签
            label = ttk.Label(
                self.content_frame, 
                text=field['label'], 
                font=(font_family, font_size)
            )
            label.grid(row=i, column=0, sticky="e", pady=10, padx=(0, 15))
            
            # 添加输入框
            width = field.get('width', 12)
            border_frame, entry = create_bordered_entry(self.content_frame, width=width)
            border_frame.grid(row=i, column=1, sticky="ew", pady=10, padx=(0, 10))
            
            # 设置默认值
            if 'default' in field:
                entry.insert(0, field['default'])
            
            # 保存输入框引用
            self.entries[field['label']] = entry
            
            # 添加验证错误标签
            error_var = tk.StringVar()
            error_label = ttk.Label(
                self.content_frame, 
                textvariable=error_var, 
                foreground="red", 
                font=(font_family, font_size - 2)
            )
            error_label.grid(row=i + 1, column=1, sticky="w", pady=(0, 5))
            self.validation_errors[field['label']] = error_var
        
        # 添加取消按钮
        def on_cancel():
            self.result = None
            self.dialog.destroy()
        
        cancel_btn = self.add_button(_('cancel'), on_cancel, column=2)
        
        # 添加确认按钮
        def on_ok():
            # 验证输入
            if self._validate_input():
                # 收集输入值
                self.result = {}
                for field in self.fields:
                    label = field['label']
                    self.result[label] = self.entries[label].get().strip()
                self.dialog.destroy()
        
        ok_btn = self.add_button(_('ok'), on_ok, column=1)
        
        # 绑定回车键（默认确认）和Escape键（默认取消）
        self.dialog.bind('<Return>', lambda e: on_ok())
        self.dialog.bind('<Escape>', lambda e: on_cancel())
        
        # 设置焦点到第一个输入框
        if self.entries:
            first_entry = list(self.entries.values())[0]
            first_entry.focus_force()
            first_entry.select_range(0, tk.END)
    
    def _validate_input(self):
        """验证输入
        
        Returns:
            bool: 验证是否通过
        """
        valid = True
        
        for field in self.fields:
            label = field['label']
            value = self.entries[label].get().strip()
            
            # 清除之前的错误信息
            self.validation_errors[label].set("")
            
            # 检查是否为空
            if not value and 'required' in field and field['required']:
                self.validation_errors[label].set("此字段不能为空")
                valid = False
                continue
            
            # 检查是否有验证函数
            if 'validator' in field:
                validator = field['validator']
                validate_message = field.get('validate_message', "输入无效")
                if not validator(value):
                    self.validation_errors[label].set(validate_message)
                    valid = False
        
        return valid


class SelectDialog(DialogBase):
    """选择对话框，用于让用户从选项中选择的场景"""
    
    def __init__(self, parent, title, options, multi_select=False, width=500, height=300):
        """初始化选择对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            options: 选项列表，每个选项可以是字符串或字典 {
                'value': 选项值,
                'label': 选项标签
            }
            multi_select: 是否支持多选
            width: 对话框宽度
            height: 对话框高度
        """
        super().__init__(parent, title, width, height, resizable=True, modal=True)
        
        self.options = options
        self.multi_select = multi_select
        self.selected_values = []
        
        # 初始化界面
        self._init_ui()
    
    def _init_ui(self):
        """初始化界面"""
        font_family, font_size = get_current_font_settings()
        
        # 创建滚动框架
        scroll_frame = ttk.Frame(self.content_frame)
        scroll_frame.pack(fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(scroll_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建列表框
        if self.multi_select:
            self.listbox = tk.Listbox(
                scroll_frame, 
                selectmode=tk.MULTIPLE, 
                yscrollcommand=scrollbar.set, 
                font=(font_family, font_size)
            )
        else:
            self.listbox = tk.Listbox(
                scroll_frame, 
                selectmode=tk.SINGLE, 
                yscrollcommand=scrollbar.set, 
                font=(font_family, font_size)
            )
        
        self.listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        # 填充选项
        self.option_values = []
        for i, option in enumerate(self.options):
            if isinstance(option, dict):
                label = option.get('label', str(option.get('value', '')))
                value = option.get('value', label)
            else:
                label = str(option)
                value = option
            
            self.listbox.insert(tk.END, label)
            self.option_values.append(value)
        
        # 添加取消按钮
        def on_cancel():
            self.result = None
            self.dialog.destroy()
        
        cancel_btn = self.add_button(_('cancel'), on_cancel, column=2)
        
        # 添加确认按钮
        def on_ok():
            # 收集选中的值
            selected_indices = self.listbox.curselection()
            self.selected_values = [self.option_values[i] for i in selected_indices]
            
            if not self.multi_select and self.selected_values:
                self.result = self.selected_values[0]
            else:
                self.result = self.selected_values
            
            self.dialog.destroy()
        
        ok_btn = self.add_button(_('ok'), on_ok, column=1)
        
        # 绑定回车键（默认确认）和Escape键（默认取消）
        self.dialog.bind('<Return>', lambda e: on_ok())
        self.dialog.bind('<Escape>', lambda e: on_cancel())
        
        # 设置焦点到列表框
        self.listbox.focus_force()
        
        # 默认选择第一个选项
        if self.options:
            self.listbox.select_set(0)


class ComplexDialog(DialogBase):
    """复杂对话框，用于包含多个控件的复杂界面"""
    
    def __init__(self, parent, title, width, height, resizable=False, modal=True):
        """初始化复杂对话框
        
        Args:
            parent: 父窗口
            title: 对话框标题
            width: 对话框宽度
            height: 对话框高度
            resizable: 是否允许调整大小
            modal: 是否为模态对话框
        """
        super().__init__(parent, title, width, height, resizable, modal)
    
    def add_field(self, label, row, column, **kwargs):
        """添加输入字段
        
        Args:
            label: 字段标签
            row: 行位置
            column: 列位置
            **kwargs: 额外参数
            
        Returns:
            tuple: (label_widget, entry_widget) 标签和输入框组件
        """
        font_family, font_size = get_current_font_settings()
        
        # 添加标签（自动添加冒号）
        label_widget = ttk.Label(
            self.content_frame, 
            text=label + ':', 
            font=(font_family, font_size)
        )
        label_widget.grid(row=row, column=column, sticky="e", pady=5, padx=(0, 15))
        
        # 添加输入框
        width = kwargs.get('width', 12)
        border_frame, entry_widget = create_bordered_entry(self.content_frame, width=width)
        border_frame.grid(row=row, column=column + 1, sticky="ew", pady=5, padx=(0, 10))
        
        # 设置默认值
        if 'default' in kwargs:
            entry_widget.insert(0, kwargs['default'])
        
        return label_widget, entry_widget
    
    def add_label(self, text, row, column, **kwargs):
        """添加标签
        
        Args:
            text: 标签文本
            row: 行位置
            column: 列位置
            **kwargs: 额外参数
            
        Returns:
            ttk.Label: 标签组件
        """
        font_family, font_size = get_current_font_settings()
        
        label = ttk.Label(
            self.content_frame, 
            text=text, 
            font=(font_family, font_size),
            **kwargs
        )
        label.grid(row=row, column=column, **kwargs.get('grid', {}))
        return label
    
    def add_button(self, text, command, column=1, padx=5):
        """添加按钮
        
        Args:
            text: 按钮文本
            command: 按钮命令
            column: 按钮列位置
            padx: 按钮水平间距
        """
        button = ttk.Button(self.button_frame, text=text, command=command, width=10)
        button.pack(side=tk.RIGHT, padx=padx, pady=5)
        return button



def center_window(window, parent=None):
    """居中显示窗口
    
    Args:
        window: 要居中的窗口对象
        parent: 父窗口对象，如果为None则居中屏幕
    """
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    
    if parent:
        # 居中父窗体
        parent.update_idletasks()
        parent_x = parent.winfo_x()
        parent_y = parent.winfo_y()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        x = parent_x + (parent_width // 2) - (width // 2)
        y = parent_y + (parent_height // 2) - (height // 2)
    else:
        # 居中屏幕
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
    
    window.geometry(f"{width}x{height}+{x}+{y}")


if sys.platform == 'win32':
    try:
        import ctypes

        PROCESS_DPI_UNAWARE = 0
        PROCESS_SYSTEM_DPI_AWARE = 1
        PROCESS_PER_MONITOR_DPI_AWARE = 2
        PROCESS_PER_MONITOR_DPI_AWARE_V2 = 3

        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(PROCESS_PER_MONITOR_DPI_AWARE_V2)
            DPI_MODE = "PROCESS_PER_MONITOR_DPI_AWARE_V2"  # type: ignore
        except AttributeError:
            ctypes.windll.shcore.SetProcessDpiAwareness(PROCESS_PER_MONITOR_DPI_AWARE)
            DPI_MODE = "PROCESS_PER_MONITOR_DPI_AWARE"  # type: ignore
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
            DPI_MODE = "SetProcessDPIAware"  # type: ignore

        # 获取当前DPI和缩放因子
        hdc = ctypes.windll.user32.GetDC(None)
        LOGPIXELSX = 88  # 水平DPI
        LOGPIXELSY = 90  # 垂直DPI
        dpi_x = ctypes.windll.gdi32.GetDeviceCaps(hdc, LOGPIXELSX)
        dpi_y = ctypes.windll.gdi32.GetDeviceCaps(hdc, LOGPIXELSY)
        ctypes.windll.user32.ReleaseDC(None, hdc)

        # 计算缩放因子
        SCALE_FACTOR = dpi_x / 96.0  # type: ignore
        # 只在直接运行应用程序时打印DPI信息，不在模块导入时打印
        if __name__ == "__main__":
            print(f"[OK] Windows DPI设置: {dpi_x}x{dpi_y} DPI, 缩放因子: {SCALE_FACTOR:.2f}, 模式: {DPI_MODE}")

    except Exception as e:
        # 只在直接运行应用程序时打印错误信息，不在模块导入时打印
        if __name__ == "__main__":
            print(f"[WARN] 设置DPI感知失败: {e}")
        # 使用默认缩放因子
        SCALE_FACTOR = 1.0  # type: ignore
        DPI_MODE = "Default"  # type: ignore


# 自定义的ColoredNotebook类，支持每个标签不同颜色
class ColoredNotebook(ttk.Frame):
    """自定义的ColoredNotebook组件，用于显示带有颜色的标签页

    这个组件扩展了ttk.Frame，实现了类似Notebook的标签页功能，
    支持为不同标签页设置不同的背景颜色，提供更好的视觉区分。

    参数:
        master: 父容器
        style: 样式对象
        tab_change_callback: 标签页切换回调函数
        is_top_level: 是否为顶级标签页
        **kwargs: 传递给ttk.Frame的其他参数
    """
    def __init__(self, master, style=None, tab_change_callback=None, is_top_level=False, **kwargs):
        super().__init__(master, **kwargs)
        # 标识是否为顶级标签页
        self.is_top_level = is_top_level

        # 确保框架能够填充父容器的空间
        self.pack_propagate(True)
        self.grid_propagate(True)

        # 绑定大小变化事件，确保内容区域能正确调整大小
        self.bind('<Configure>', self.on_configure)

        # 保存样式对象
        self.style = style
        # 保存标签页切换回调函数
        self.tab_change_callback = tab_change_callback

        # 为每个实例生成唯一ID，用于创建唯一的样式名称
        self.unique_id = id(self)

        # 为每个Notebook实例创建唯一的样式名称
        unique_id = self.unique_id
        self.light_blue_style = f"LightBlue{unique_id}.TFrame"
        self.light_green_style = f"LightGreen{unique_id}.TFrame"
        self.light_orange_style = f"LightOrange{unique_id}.TFrame"
        self.light_purple_style = f"LightPurple{unique_id}.TFrame"
        self.light_pink_style = f"LightPink{unique_id}.TFrame"

        # 颜色映射字典，用于优化重复的条件判断
        self.color_styles = {
            "#e3f2fd": self.light_blue_style,  # 蓝色标签
            "#e8f5e9": self.light_green_style,  # 绿色标签
            "#fff3e0": self.light_orange_style,  # 橙色标签
            "#f3e5f5": self.light_purple_style,  # 紫色标签
            "#fce4ec": self.light_pink_style,  # 粉色标签
            "#e0f2f1": self.light_blue_style,  # 青色标签
        }

        # 鼠标按下时的激活颜色映射
        self.mouse_down_colors = {
            "#e3f2fd": "#bbdefb",  # 蓝色标签
            "#e8f5e9": "#c8e6c9",  # 绿色标签
            "#fff3e0": "#ffe0b2",  # 橙色标签
            "#f3e5f5": "#e1bee7",  # 紫色标签
            "#fce4ec": "#f8bbd0",  # 粉色标签
            "#e0f2f1": "#b2dfdb",  # 青色标签
        }

        # 鼠标释放时的激活颜色映射
        self.mouse_up_colors = {
            "#e3f2fd": "#90caf9",  # 蓝色标签
            "#e8f5e9": "#a5d6a7",  # 绿色标签
            "#fff3e0": "#ffcc80",  # 橙色标签
            "#f3e5f5": "#ce93d8",  # 紫色标签
            "#fce4ec": "#f48fb1",  # 粉色标签
            "#e0f2f1": "#80deea",  # 青色标签
        }

        # 创建标签栏容器，使用ttk.Frame并继承默认样式
        self.tab_bar_container = ttk.Frame(self)
        self.tab_bar_container.pack(side="top", fill="x")

        # 创建标签栏 - 使用ttk.Frame并继承默认样式
        self.tab_bar = ttk.Frame(self.tab_bar_container)
        self.tab_bar.pack(side="left", fill="y")

        # 创建一个占位Frame，使用ttk.Frame并继承默认样式
        self.tab_bar_spacer = ttk.Frame(self.tab_bar_container)
        self.tab_bar_spacer.pack(side="left", fill="both", expand=True)

        # 创建右侧按钮容器 - 使用ttk.Frame并继承默认样式
        self.tab_bar_right_buttons = ttk.Frame(self.tab_bar_container)
        self.tab_bar_right_buttons.pack(side="right", fill="y")

        # 创建内容区域 - 移除箭头指向的灰色框线
        self.content_area = ttk.Frame(self, borderwidth=0, relief="flat")
        self.content_area.pack(side="top", fill="both", expand=True, padx=0, pady=0)

        # 确保content_area能完全填充笔记本控件的空间
        self.content_area.pack_propagate(True)
        self.content_area.grid_propagate(True)

        # 标签配置
        self.tabs = []
        self.active_tab = None

    def _get_font_settings(self):
        """获取当前字体设置

        Returns:
            tuple: (字体名称, 字体大小)
        """
        return get_current_font_settings()

    def on_configure(self, event):
        """当笔记本控件大小变化时调用，确保内容区域能正确调整大小"""
        if hasattr(self, 'content_area'):
            # 更新content_area的布局，确保它能完全填充笔记本控件的空间
            self.content_area.pack_configure(fill='both', expand=True)

            # 如果有选中的标签，确保其内容框架也能正确调整大小
            if hasattr(self, 'active_tab') and self.active_tab is not None and 0 <= self.active_tab < len(self.tabs):
                selected_tab = self.tabs[self.active_tab]
                selected_tab["content"].pack_configure(fill='both', expand=True)

    def _on_tab_mouse_down(self, button, color):
        """当鼠标按下标签页时，更新内容区域背景色为按下状态颜色"""
        if hasattr(self, "active_tab") and button.tab_index == self.active_tab:
            self.mouse_down_colors.get(color, "#e1bee7")

    def _on_tab_mouse_up(self, button, color):
        """当鼠标释放标签页时，恢复内容区域背景色为激活状态颜色"""
        if hasattr(self, "active_tab") and button.tab_index == self.active_tab:
            self.mouse_up_colors.get(color, "#ce93d8")

    def _update_background_to_result_frame_color(self):
        """更新标签栏背景色以匹配result_frame"""
        try:
            # 直接获取父容器的背景色，而不是通过style.lookup
            # 对于ttk组件，我们需要先获取其内部的label或content组件
            bg_color = None

            # 尝试多种方式获取父容器的背景色
            if hasattr(self.master, 'winfo_children'):
                # 获取父容器的子组件
                children = self.master.winfo_children()
                for child in children:
                    # 检查子组件是否有cget方法
                    if not hasattr(child, 'cget'):
                        continue

                    # 尝试从子组件获取背景色
                    try:
                        child_bg = child.cget("background")
                        if child_bg and not child_bg.startswith("system."):
                            bg_color = child_bg
                            break
                    except (AttributeError, tk.TclError):
                        continue

            # 如果无法从子组件获取背景色，尝试直接从父容器获取
            if not bg_color or bg_color.startswith("system."):
                try:
                    bg_color = self.master.cget("background")
                except (AttributeError, tk.TclError):
                    pass

            # 如果还是无法获取背景色，使用默认的背景色
            if not bg_color or bg_color.startswith("system."):
                bg_color = self.style.lookup("TFrame", "background") or "#f0f0f0"

            # 使用style来设置ttk.Frame的背景色，为每个实例创建唯一的样式名称
            temp_style_name = f"TempColoredNotebookStyle{self.unique_id}.TFrame"
            self.style.configure(temp_style_name, background=bg_color)

            # 应用样式到所有ttk.Frame组件
            self.tab_bar_container.configure(style=temp_style_name)
            self.tab_bar.configure(style=temp_style_name)
            self.tab_bar_spacer.configure(style=temp_style_name)
            self.content_area.configure(style=temp_style_name)

        except (tk.TclError, AttributeError, ValueError):
            # 发生错误时，不设置自定义背景色，使用默认样式
            pass

    def add_tab(self, label, content_frame, color="#e0e0e0", tab_id=None):
        """添加一个新标签
        
        Args:
            label: 标签显示文本
            content_frame: 标签内容框架
            color: 标签颜色
            tab_id: 标签唯一标识符（用于跨语言保存次序），如果为None则使用label
        """
        tab = {"label": label, "content": content_frame, "color": color, "button": None, "tab_id": tab_id if tab_id else label}

        font_family, font_size = self._get_font_settings()
        style_manager = get_style_manager()
        tab_width = style_manager.get_tab_width() if style_manager else 10
        tab_pady = style_manager.get_tab_vertical_padding() if style_manager else 5

        button_params = {
            "text": label,
            "bg": color,
            "relief": "flat",
            "borderwidth": 0,
            "padx": 5,
            "pady": tab_pady,
            "font": (font_family, font_size, "normal"),
            "width": tab_width,
            "foreground": "#333333",
            "activeforeground": "#333333",
        }

        if self.is_top_level:
            button_params["activebackground"] = "#ffb74d"
        else:
            button_params["activebackground"] = self.mouse_down_colors.get(color, "#e1bee7")

        button = tk.Button(self.tab_bar, **button_params)  # type: ignore

        # 保存按钮对应的标签索引，以便在事件处理中使用
        button.tab_index = len(self.tabs)  # type: ignore

        # 绑定标签页切换事件 - 使用按钮的tab_index属性获取当前索引
        button.bind("<Button-1>", lambda e: self.select_tab(e.widget.tab_index))

        # 只有内部标签页需要为鼠标按下/释放添加内容区域背景色变化效果
        if not self.is_top_level:
            # 绑定鼠标按下事件 - 更新内容区域背景色为按下状态颜色
            button.bind("<Button-1>", lambda e, c=color: self._on_tab_mouse_down(e.widget, c), add="+")
            # 绑定鼠标释放事件 - 恢复内容区域背景色为激活状态颜色
            button.bind("<ButtonRelease-1>", lambda e, c=color: self._on_tab_mouse_up(e.widget, c), add="+")
        button.pack(side="left", padx=0, pady=0)

        tab["button"] = button
        self.tabs.append(tab)

        # 如果是第一个标签，自动选中
        if len(self.tabs) == 1:
            self.select_tab(0)
        else:
            # 确保所有标签页都应用正确的样式
            current_active_tab = getattr(self, "active_tab", 0)
            self.select_tab(current_active_tab)

    def select_tab(self, tab_index):
        """选中一个标签"""
        if tab_index < 0 or tab_index >= len(self.tabs):
            return

        font_family, font_size = self._get_font_settings()

        # 隐藏所有内容
        for tab in self.tabs:
            tab["content"].pack_forget()

            # 根据是否为顶级标签页应用不同的非激活样式
            if self.is_top_level:
                # 顶级标签页非激活状态：灰底深灰色文字不加粗
                tab["button"].config(
                    relief="flat",
                    bg="#808080",  # 灰色背景
                    font=(font_family, font_size, "normal"),
                    foreground="#fcfcfc",  # 白色文字
                )
            else:
                # 内部标签页非激活状态：保持原有背景色，深灰色文字
                tab["button"].config(
                    relief="flat",
                    bg=tab["color"],
                    font=(font_family, font_size, "normal"),
                    foreground="#808080",  # 默认深灰色文字
                )

        # 显示选中的标签内容
        selected_tab = self.tabs[tab_index]
        selected_tab["content"].pack(fill="both", expand=True, padx=0, pady=0)

        # 更新当前激活的标签页索引
        self.active_tab = tab_index

        # 根据是否为顶级标签页应用不同的激活样式
        selected_color = "#ff9800"  # 默认橙色背景
        if self.is_top_level:
            # 顶级标签页激活状态：橙底黑字并加粗
            selected_tab["button"].config(
                relief="flat",
                bg="#ff9800",  # 橙色背景
                font=(font_family, font_size, "normal"),  # 加粗字体
                foreground="#000000",  # 黑色文字
            )
        else:
            # 内部标签页激活状态：使用更亮的颜色（之前的鼠标按下颜色）
            selected_color = self.mouse_up_colors.get(selected_tab["color"], "#ce93d8")

            selected_tab["button"].config(
                relief="flat", bg=selected_color, font=(font_family, font_size, "normal"), foreground="#000000"
            )

        # 更新对应内容框架样式的背景色，使其与选中标签的颜色保持一致
        # 只有内部标签页需要更新样式，顶级标签页不需要
        if not self.is_top_level:
            style_name = self.color_styles.get(selected_tab["color"], self.light_blue_style)
            self.style.configure(style_name, background=selected_color)

        # 更新背景色以匹配result_frame
        self._update_background_to_result_frame_color()

        # 调用标签页切换回调函数
        if self.tab_change_callback:
            self.tab_change_callback(tab_index)

    def move_tab_up(self, tab_index):
        """将指定标签上移一位
        
        Args:
            tab_index: 要移动的标签索引
        """
        if tab_index <= 0 or tab_index >= len(self.tabs):
            return False
        
        # 交换标签位置
        self.tabs[tab_index], self.tabs[tab_index - 1] = self.tabs[tab_index - 1], self.tabs[tab_index]
        
        # 更新按钮的tab_index属性
        for i, tab in enumerate(self.tabs):
            tab["button"].tab_index = i
        
        # 重新排列按钮显示顺序
        for tab in self.tabs:
            tab["button"].pack_forget()
        
        for tab in self.tabs:
            tab["button"].pack(side="left", padx=0, pady=0)
        
        # 如果移动的是当前激活标签，更新active_tab
        if self.active_tab == tab_index:
            self.active_tab = tab_index - 1
        elif self.active_tab == tab_index - 1:
            self.active_tab = tab_index
        
        # 刷新显示
        self.select_tab(self.active_tab)
        return True

    def move_tab_down(self, tab_index):
        """将指定标签下移一位
        
        Args:
            tab_index: 要移动的标签索引
        """
        if tab_index < 0 or tab_index >= len(self.tabs) - 1:
            return False
        
        # 交换标签位置
        self.tabs[tab_index], self.tabs[tab_index + 1] = self.tabs[tab_index + 1], self.tabs[tab_index]
        
        # 更新按钮的tab_index属性
        for i, tab in enumerate(self.tabs):
            tab["button"].tab_index = i
        
        # 重新排列按钮显示顺序
        for tab in self.tabs:
            tab["button"].pack_forget()
        
        for tab in self.tabs:
            tab["button"].pack(side="left", padx=0, pady=0)
        
        # 如果移动的是当前激活标签，更新active_tab
        if self.active_tab == tab_index:
            self.active_tab = tab_index + 1
        elif self.active_tab == tab_index + 1:
            self.active_tab = tab_index
        
        # 刷新显示
        self.select_tab(self.active_tab)
        return True

    def move_tab_to_top(self, tab_index):
        """将指定标签移动到最顶部
        
        Args:
            tab_index: 要移动的标签索引
        """
        if tab_index <= 0 or tab_index >= len(self.tabs):
            return False
        
        # 将标签移到顶部
        tab = self.tabs.pop(tab_index)
        self.tabs.insert(0, tab)
        
        # 更新按钮的tab_index属性
        for i, tab_item in enumerate(self.tabs):
            tab_item["button"].tab_index = i
        
        # 重新排列按钮显示顺序
        for tab_item in self.tabs:
            tab_item["button"].pack_forget()
        
        for tab_item in self.tabs:
            tab_item["button"].pack(side="left", padx=0, pady=0)
        
        # 如果移动的是当前激活标签，更新active_tab
        if self.active_tab == tab_index:
            self.active_tab = 0
        elif self.active_tab < tab_index:
            self.active_tab += 1
        
        # 刷新显示
        self.select_tab(self.active_tab)
        return True

    def move_tab_to_bottom(self, tab_index):
        """将指定标签移动到最底部
        
        Args:
            tab_index: 要移动的标签索引
        """
        if tab_index < 0 or tab_index >= len(self.tabs) - 1:
            return False
        
        # 将标签移到底部
        tab = self.tabs.pop(tab_index)
        self.tabs.append(tab)
        
        # 更新按钮的tab_index属性
        for i, tab_item in enumerate(self.tabs):
            tab_item["button"].tab_index = i
        
        # 重新排列按钮显示顺序
        for tab_item in self.tabs:
            tab_item["button"].pack_forget()
        
        for tab_item in self.tabs:
            tab_item["button"].pack(side="left", padx=0, pady=0)
        
        # 如果移动的是当前激活标签，更新active_tab
        if self.active_tab == tab_index:
            self.active_tab = len(self.tabs) - 1
        elif self.active_tab > tab_index:
            self.active_tab -= 1
        
        # 刷新显示
        self.select_tab(self.active_tab)
        return True

    def get_tab_labels(self):
        """获取所有标签的显示文本列表
        
        Returns:
            list: 标签显示文本列表
        """
        return [tab['label'] for tab in self.tabs]
    
    def get_tab_ids(self):
        """获取所有标签的唯一标识符列表
        
        Returns:
            list: 标签唯一标识符列表
        """
        return [tab['tab_id'] for tab in self.tabs]

    def save_tab_order(self):
        """保存当前标签次序到配置文件（使用唯一标识符，支持跨语言）"""
        from config_manager import get_config
        
        try:
            config = get_config()
            tab_ids = self.get_tab_ids()
            config.set_ui_tab_order(tab_ids)
            return True
        except Exception as e:
            print(f"保存标签次序失败: {e}")
            return False

    def apply_tab_order(self, ordered_ids):
        """根据给定的标签次序重新排列标签（使用唯一标识符，支持跨语言）
        
        Args:
            ordered_ids: 标签唯一标识符列表，按期望顺序排列
            
        Returns:
            bool: 是否成功
        """
        if not ordered_ids or not isinstance(ordered_ids, list):
            return False
            
        # 创建原始标签字典（使用tab_id作为键）
        original_tabs = {tab['tab_id']: tab for tab in self.tabs}
        
        # 验证所有指定的标签是否存在
        for tab_id in ordered_ids:
            if tab_id not in original_tabs:
                return False
        
        # 按指定顺序重新排列标签
        new_tabs = []
        for tab_id in ordered_ids:
            if tab_id in original_tabs:
                new_tabs.append(original_tabs[tab_id])
        
        # 添加未在配置中指定的标签（保持原有顺序）
        for tab in self.tabs:
            if tab['tab_id'] not in ordered_ids:
                new_tabs.append(tab)
        
        # 更新标签列表
        self.tabs = new_tabs
        
        # 更新按钮的tab_index属性
        for i, tab in enumerate(self.tabs):
            tab["button"].tab_index = i
        
        # 重新排列按钮显示顺序
        for tab in self.tabs:
            tab["button"].pack_forget()
        
        for tab in self.tabs:
            tab["button"].pack(side="left", padx=0, pady=0)
        
        # 刷新显示（保持当前选中的标签）
        if self.active_tab is not None and 0 <= self.active_tab < len(self.tabs):
            self.select_tab(self.active_tab)
        elif self.tabs:
            self.select_tab(0)
        
        return True


class SubnetPlannerApp:
    """子网规划师主应用程序类

    这个类实现了一个子网规划的GUI应用程序，
    支持子网分割、子网规划、IP信息查询等功能。
    """
    
    # 状态相关的类常量
    VALID_STATUSES = {'released', 'allocated', 'reserved'}
    
    @classmethod
    def _get_status_map(cls):
        """获取状态映射表（动态本地化版本）
        
        Returns:
            dict: 本地化状态到英文状态的映射
        """
        return {
            _('released'): 'released',
            _('allocated'): 'allocated',
            _('reserved'): 'reserved'
        }
    
    @classmethod
    def _get_all_valid_statuses(cls):
        """获取所有有效状态值（英文和本地化版本）
        
        Returns:
            set: 包含所有有效状态值的集合
        """
        return cls.VALID_STATUSES | set(cls._get_status_map().keys())
    
    def autocomplete_ipv6(self, event):
        """IPv6地址自动补全功能
        
        当用户输入IPv6地址时，提供智能自动补全功能
        - 自动补全双冒号
        - 智能补全常见IPv6地址片段
        - 支持零压缩
        - 验证并格式化IPv6地址
        
        Args:
            event: 键盘事件对象
        """
        try:
            entry = event.widget
            current_text = entry.get().strip()
            cursor_pos = entry.index(tk.INSERT)
            
            # 处理退格键和删除键，不进行补全
            if event.keysym in ['BackSpace', 'Delete']:
                return
            
            # 智能补全常见IPv6地址片段
            # 补全链路本地地址前缀: fe80 -> fe80::
            if current_text == 'fe80':
                entry.delete(0, tk.END)
                entry.insert(0, 'fe80::')
                entry.icursor(5)
                return
            
            # 补全链路本地地址前缀: fe80: -> fe80::
            if current_text == 'fe80:':
                entry.delete(0, tk.END)
                entry.insert(0, 'fe80::')
                entry.icursor(5)
                return
            
            # 补全唯一本地地址前缀: fd -> fd00::
            if current_text == 'fd':
                entry.delete(0, tk.END)
                entry.insert(0, 'fd00::')
                entry.icursor(5)
                return
            
            # 补全唯一本地地址前缀: fd: -> fd00::
            if current_text == 'fd:':
                entry.delete(0, tk.END)
                entry.insert(0, 'fd00::')
                entry.icursor(5)
                return
            
            # 补全文档地址前缀: 2001 -> 2001:db8::
            if current_text == '2001':
                entry.delete(0, tk.END)
                entry.insert(0, '2001:db8::')
                entry.icursor(9)
                return
            
            # 补全文档地址前缀: 2001: -> 2001:db8::
            if current_text == '2001:':
                entry.delete(0, tk.END)
                entry.insert(0, '2001:db8::')
                entry.icursor(9)
                return
            
            # 补全文档地址前缀: 2001:db8 -> 2001:db8::
            if current_text == '2001:db8':
                entry.delete(0, tk.END)
                entry.insert(0, '2001:db8::')
                entry.icursor(9)
                return
            
            # 补全文档地址前缀: 2001:db8: -> 2001:db8::
            if current_text == '2001:db8:':
                entry.delete(0, tk.END)
                entry.insert(0, '2001:db8::')
                entry.icursor(9)
                return
            
            # 双冒号补全: 当用户输入单个冒号且前面不是冒号时
            if event.char == ':' and cursor_pos > 0:
                # 获取当前光标前的字符
                before_cursor = current_text[:cursor_pos]
                after_cursor = current_text[cursor_pos:]
                
                # 检查光标前是否已有冒号
                if before_cursor and before_cursor[-1] != ':' and '::' not in current_text:
                    # 在光标位置插入另一个冒号，形成双冒号
                    entry.delete(0, tk.END)
                    entry.insert(0, before_cursor + '::' + after_cursor)
                    entry.icursor(cursor_pos + 1)
                    return
                    
        except Exception as e:
            # 自动补全失败时不影响用户输入
            pass
    
    def validate_cidr(self, text, entry=None, style_based=False, ip_version=None, require_prefix=None):
        """通用CIDR验证函数（UI层包装）

        Args:
            text: 要验证的CIDR字符串
            entry: 可选的输入框对象，用于显示验证结果
            style_based: 是否使用样式来显示验证结果，否则使用前景色
            ip_version: 可选的IP版本字符串，用于指定要验证的IP版本，如"IPv4"或"IPv6"
            require_prefix: 前缀要求模式:
                - None: 带不带前缀都可以（默认）
                - True: 必须带前缀（如 10.0.0.0/8）
                - False: 必须不带前缀（纯IP地址，如 10.0.0.1）

        Returns:
            验证结果，True表示有效，False表示无效，"1"表示用于validatecommand的有效
        """
        # 调用服务层的验证方法
        result = self.validation_service.validate_cidr(text, ip_version, require_prefix)
        is_valid = result['valid']
        
        # 处理视觉反馈
        if entry:
            if style_based:
                entry.config(style='Valid.TEntry' if is_valid else 'Invalid.TEntry')
            else:
                entry.config(foreground='black' if is_valid else 'red')

        # 对于validatecommand，始终返回"1"，允许所有输入，只做视觉提示
        # 对于直接调用，返回布尔值表示验证结果
        return "1" if entry else is_valid

    def _get_font(self, font_size=None):
        """获取当前字体对象

        Args:
            font_size: 可选的字体大小，如果不提供则使用默认字体大小

        Returns:
            tkfont.Font: 字体对象
        """
        try:
            font_family, default_font_size = get_current_font_settings()
            # 如果提供了字体大小，则使用提供的值，否则使用默认值
            current_font_size = font_size if font_size is not None else default_font_size
            return tkfont.Font(family=font_family, size=current_font_size)
        except tk.TclError:
            return tkfont.Font(family="Arial", size=10 if font_size is None else font_size)

    def _calculate_pixel_width(self, text, font=None):
        """计算文本的像素宽度

        Args:
            text: 要计算的文本
            font: 可选的字体对象，如果不提供则使用默认字体

        Returns:
            int: 文本的像素宽度
        """
        if font is None:
            font = self._get_font()
        return font.measure(text)

    def _truncate_text_by_pixel(self, text, icon, max_pixel_width, font=None):
        """基于像素宽度截断文本

        Args:
            text: 要截断的文本
            icon: 图标文本
            max_pixel_width: 最大像素宽度
            font: 可选的字体对象

        Returns:
            str: 截断后的文本
        """
        if font is None:
            font = self._get_font()

        # 计算图标的宽度
        icon_width = self._calculate_pixel_width(icon, font)

        # 可用宽度：总宽度减去图标宽度
        available_width = max_pixel_width - icon_width

        # 先尝试显示完整文本（加上图标）
        full_text_with_icon = icon + text
        full_width = self._calculate_pixel_width(full_text_with_icon, font)

        # 如果完整文本可以显示，直接返回
        if full_width <= max_pixel_width:
            return text

        # 计算省略号的宽度
        ellipsis_width = self._calculate_pixel_width("...", font)

        # 二分查找合适的截断位置，考虑省略号宽度
        low = 0
        high = len(text)
        best_length = 0

        while low <= high:
            mid = (low + high) // 2
            current_text = text[:mid]
            current_width = self._calculate_pixel_width(current_text, font)

            if current_width <= available_width - ellipsis_width:
                best_length = mid
                low = mid + 1
            else:
                high = mid - 1

        # 确保截断后的文本不会过长
        truncated = text[:best_length]

        # 调整截断位置，确保加上省略号和图标后不会超过最大宽度
        while best_length > 0:
            truncated = text[:best_length]
            truncated_width = self._calculate_pixel_width(truncated, font) + ellipsis_width + icon_width
            if truncated_width <= max_pixel_width:
                return truncated + "..."
            best_length -= 1

        return "..."

    def _format_datetime(self, datetime_str, output_format="%Y-%m-%d %H:%M:%S"):
        """格式化日期时间字符串
        
        Args:
            datetime_str: 日期时间字符串，支持多种格式
            output_format: 输出格式，默认为"%Y-%m-%d %H:%M:%S"
        
        Returns:
            str: 格式化后的日期时间字符串，如果解析失败则返回原始字符串
        """
        # 处理字符串 "None" 的情况
        if not datetime_str or datetime_str == "None":
            return ""
        
        try:
            if 'T' in datetime_str:
                # ISO格式: 2023-12-31T23:59:59.123456
                # 替换T为空格，然后使用strptime解析
                dt_str = datetime_str.replace('T', ' ')
                if '.' in dt_str:
                    dt = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S.%f")
                else:
                    dt = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            elif ' ' in datetime_str:
                # 普通格式: 2023-12-31 23:59:59.123456 或 2023-12-31 23:59:59
                if '.' in datetime_str:
                    dt = datetime.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S.%f")
                else:
                    dt = datetime.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
            else:
                # 只有日期部分: 2023-12-31
                dt = datetime.datetime.strptime(datetime_str, "%Y-%m-%d")
            
            return dt.strftime(output_format)
        except (ValueError, TypeError):
            # 只捕获预期的异常类型，避免捕获KeyboardInterrupt等
            return datetime_str

    def __init__(self, main_window):
        # 应用程序信息
        self.app_name = _("app_name")
        self.app_version = get_version()

        # 初始化历史记录仓库
        self.history_repo = HistoryRepository()
        self.deleted_history = self.history_repo.deleted_history

        # 高级工具历史记录列表 - 委托给HistoryRepository（从数据库加载）
        self.ipv4_history = self.history_repo.ipv4_history
        self.ipv6_history = self.history_repo.ipv6_history

        # 图表相关属性（预声明，避免Attribute-defined-outside-init警告）
        self.planning_chart_frame = None
        self.planning_chart_canvas = None
        self.planning_chart_v_scrollbar = None
        self.planning_chart_data = None

        # 窗口背景色（预声明，动态更新）
        self.bg_color = None
        self.range_start_history = self.history_repo.range_start_history
        self.range_end_history = self.history_repo.range_end_history

        # 切分子网相关属性 - 从数据库加载历史记录
        self.split_parent_networks = self.history_repo.split_parent_networks
        self.split_networks = self.history_repo.split_networks
        self.parent_entry = None
        self.split_entry = None
        self.execute_btn = None
        self.reexecute_btn = None
        self.history_tree = None
        self.history_listbox = None
        self.export_btn = None
        self.notebook = None
        self.split_info_frame = None
        self.split_tree = None
        self.remaining_frame = None
        self.remaining_tree = None
        self.chart_frame = None
        self.chart_scrollbar = None
        self.chart_canvas = None
        self.remaining_scroll_v = None
        self.chart_data = None

        # 网段规划相关属性 - 从数据库加载历史记录
        self.planning_parent_networks = self.history_repo.planning_parent_networks
        self.planning_parent_entry = None
        self.pool_tree = None
        self.pool_scrollbar = None
        self.requirements_tree = None
        self.requirements_scrollbar = None

        # 功能调试面板相关属性
        self.test_dialog = None  # 用于存储调试面板的引用，确保只能打开一个
        self.undo_delete_btn = None
        self.swap_btn = None
        self.planning_notebook = None
        self.execute_planning_btn = None
        self.allocated_frame = None
        self.allocated_tree = None
        self.planning_remaining_frame = None
        self.planning_remaining_tree = None

        # 窗口锁定相关属性
        self.width_locked = True  # 默认锁定窗口横向尺寸
        self.width_lock_var = None  # 初始化width_lock_var属性

        # 编辑相关属性
        self.edit_entry = None
        self.current_edit_item = None
        self.current_edit_column = None
        self.current_edit_column_index = None
        self.current_edit_tree = None

        # 高级工具相关属性
        self.advanced_notebook = None
        self.ipv4_info_frame = None
        self.ipv6_info_frame = None
        self.merge_frame = None
        self.overlap_frame = None
        self.ipv6_info_entry = None
        self.ipv6_cidr_var = None
        self.ipv6_cidr_combobox = None
        self.ipv6_info_btn = None
        self.ipv6_info_tree = None
        self.subnet_merge_text = None
        self.merge_btn = None
        self.range_start_entry = None
        self.range_end_entry = None
        self.range_to_cidr_btn = None
        self.merge_result_frame = None
        self.merge_result_tree = None
        self.merge_result_scrollbar = None
        self.ip_info_entry = None
        self.subnet_mask_cidr_map = None
        self.cidr_subnet_mask_map = None
        self.ip_mask_var = None
        self.ip_mask_combobox = None
        self.ip_cidr_var = None
        self.ip_cidr_combobox = None
        self.ip_info_btn = None
        self.ip_info_tree = None
        self.overlap_text = None
        self.overlap_btn = None
        self.overlap_result_tree = None

        # 其他属性
        self.theme_var = None
        self.is_pinned = None

        # 初始化导出工具
        self.export_utils = ExportUtils()

        # 初始化业务服务
        self.validation_service = ValidationService(self)
        self.ip_query_service = IPQueryService(self)
        
        # 初始化网络扫描器（重用实例以避免内存泄露）
        from services.network_scanner import NetworkScanner
        self.network_scanner = NetworkScanner()

        # 获取系统双击间隔设置
        try:
            self.double_click_interval = self.root.tk.call('tk', 'getDoubleClickTime')
        except Exception:
            self.double_click_interval = 500  # 默认值

        self.root = main_window
        self.root.title(f"{_("app_name")} v{self.app_version}")
        # 设置应用图标
        try:
            # 使用PIL加载高分辨率图标
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Subnet_Planner.ico")
            if os.path.exists(icon_path):
                # 打开图标文件
                icon_image = Image.open(icon_path)
                # 转换为PhotoImage对象
                photo = ImageTk.PhotoImage(icon_image)
                # 设置应用图标
                self.root.iconphoto(True, photo)
                # 保存引用，防止被GC回收
                self._icon_photo = photo
        except Exception as e:
            print(f"设置图标失败: {e}")
            # 降级方案：使用传统iconbitmap方法
            try:
                icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Subnet_Planner.ico")
                if os.path.exists(icon_path):
                    self.root.iconbitmap(icon_path)
            except Exception as fallback_e:
                print(f"降级方案也失败: {fallback_e}")
        # 所有窗口大小、位置和限制设置都由主程序入口统一管理
        # 这里只设置窗口标题

        # 设置样式
        self.style = ttk.Style()
        
        # 从配置文件加载主题设置，如果没有配置则使用默认主题
        from config_manager import get_config
        config = get_config()
        saved_theme = config.get('ui.theme', 'vista')
        
        # 获取系统可用的主题列表
        available_themes = self.style.theme_names()
        
        # 如果保存的主题不在可用列表中，使用默认主题
        if saved_theme not in available_themes:
            saved_theme = 'vista'
        
        self.style.theme_use(saved_theme)

        # 初始化样式管理器
        self.style_manager = init_style_manager(self.root)
        update_styles()
        
        # 初始化IPAM仓库
        self.ipam_repo = IPAMRepository()
        self.ipam = self.ipam_repo.ipam

        # 初始化子网业务服务
        self.subnet_split_service = SubnetSplitService(self)
        self.subnet_planning_service = SubnetPlanningService(self)

        # 初始化历史记录相关属性 - 委托给HistoryRepository
        self.history_states = self.history_repo.history_states
        self.current_history_index = self.history_repo.current_history_index
        self.planning_history_records = self.history_repo.planning_history_records

        # 功能调试面板快捷键绑定（彩蛋功能）
        # Ctrl+Shift+I: 打开功能调试面板，同时绑定大小写版本确保 Caps Lock 状态下正常工作
        # - I: 主快捷键 (Info/Inspect)
        self.root.bind_all('<Control-Shift-I>', self.toggle_test_info_bar)
        self.root.bind_all('<Control-Shift-i>', self.toggle_test_info_bar)
        self.test_info_bar_enabled = False

        # 隐藏信息管理快捷键绑定
        # Ctrl+Shift+H: 打开选中IP地址的隐藏信息管理对话框
        # - H: 主快捷键 (Hidden)
        self.root.bind_all('<Control-Shift-H>', self.open_hidden_info_dialog)
        self.root.bind_all('<Control-Shift-h>', self.open_hidden_info_dialog)

        # 创建主框架 - 调整内边距使其更加紧凑
        self.main_frame = ttk.Frame(self.root, padding="15")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建信息栏 spacer - 固定高度的占位空间
        self.info_spacer = ttk.Frame(self.main_frame, style="Placeholder.TFrame")
        self.info_spacer.pack(side="bottom", fill="x")
        self.info_spacer.pack_forget()
        self.info_spacer.configure(height=30)

        # 创建信息栏框架 - 放在 spacer 内，使用 place 布局
        self.info_bar_frame = ttk.Frame(self.info_spacer, style="InfoBar.TFrame")
        # 初始时不显示，等待需要时才显示
        self.info_bar_frame.place_forget()

        # 创建顶级标签页控件，用于切换子网切分和子网规划两大功能模块
        self.create_top_level_notebook()

        # 在右上角添加关于链接按钮、主题切换按钮和钉住按钮（延迟执行，确保窗口完全渲染）
        self.root.after(100, self.create_about_link)

        # 绑定窗口大小变化事件，动态调整右上角按钮位置
        self.root.bind('<Configure>', self.on_window_configure, add='+')

        # 确保信息栏框架的grid布局配置正确
        self.info_bar_frame.grid_rowconfigure(0, weight=1)
        self.info_bar_frame.grid_columnconfigure(0, weight=1)
        self.info_bar_frame.grid_columnconfigure(1, weight=0)

        # 获取当前语言的字体设置
        font_family, font_size = get_current_font_settings()
        # 获取信息栏的独立字体大小配置
        info_bar_font_size = get_info_bar_font_size()

        self.info_bar_frame.grid_rowconfigure(0, weight=1)
        self.info_bar_frame.grid_columnconfigure(0, weight=1)
        self.info_bar_frame.grid_columnconfigure(1, weight=0)

        # 获取当前主题的背景色
        
        # 启动时执行自动备份检查
        try:
            from datetime import datetime
            # 从配置管理器获取自动备份频率
            config = get_config()
            frequency = config.get_auto_backup_frequency()
            if frequency:
                # 获取最后一次备份时间
                last_backup_time = self.ipam.get_last_backup_time()
                current_time = datetime.now()
                
                # 根据频率检查是否需要备份
                should_backup = False
                if not last_backup_time:
                    # 没有备份记录，执行备份
                    should_backup = True
                else:
                    try:
                        # 验证 last_backup_time 类型
                        if isinstance(last_backup_time, str):
                            # 尝试从 ISO 格式字符串转换
                            last_backup_time = datetime.fromisoformat(last_backup_time)
                        elif not isinstance(last_backup_time, datetime):
                            # 处理无效类型，执行备份
                            should_backup = True
                        else:
                            # 计算时间差
                            time_diff = current_time - last_backup_time
                            
                            # 验证频率值
                            valid_frequencies = {'hourly': 3600, 'daily': 86400, 'weekly': 604800, 'monthly': 2592000}
                            if frequency in valid_frequencies:
                                # 根据频率判断是否需要备份
                                if time_diff.total_seconds() >= valid_frequencies[frequency]:
                                    should_backup = True
                            else:
                                # 无效频率值，执行备份
                                should_backup = True
                    except Exception as e:
                        # 发生异常，执行备份
                        print(f"自动备份时间检查异常: {e}")
                        should_backup = True
                
                if should_backup:
                    self.ipam.backup_data(backup_type='auto', frequency=frequency)
        except Exception as e:
            print(f"自动备份检查失败: {e}")
        
        # 启动时不自动检查过期IP地址
        # try:
        #     self.check_expired_ips()
        # except Exception as e:
        #     print(f"检查过期IP失败: {e}")
        
        bg_color = self.style.lookup("TFrame", "background")

        # 使用Label组件替代Text，以简化实现
        self.info_label = tk.Label(
            self.info_bar_frame,
            padx=0, pady=0,  # 增加内边距以避免被边框遮挡，使用单一值而非元组
            font=(font_family, info_bar_font_size),  # 使用信息栏独立的字体大小配置
            takefocus=False,  # 不接受焦点
            cursor="arrow",  # 显示普通箭头光标
            background=bg_color,  # 设置背景色跟随主题
            anchor="nw",  # 文本左上对齐，支持多行文本左对齐
            justify="left",  # 多行文本时左对齐
        )
        self.info_label.grid(row=0, column=0, sticky="ew", padx=(5, 0), pady=0)

        self.info_close_btn = tk.Button(
            self.info_bar_frame,
            text="✕",
            command=self.hide_info_bar,
            cursor="hand2",
            takefocus=False,
            bg=bg_color,
            fg="#9E9E9E",
            font=("Arial", 8),  # 此处字体硬编码是程序需要，禁止修改
            borderwidth=0,
            highlightthickness=0,
            relief="flat",
            padx=0,
            pady=0,
        )
        self.info_close_btn.grid(row=0, column=1, padx=(0, 0), pady=(0, 4), sticky="se")

        self.info_auto_hide_id = None
        self.info_auto_hide_scheduled_time = None  # 记录定时器设置的时间
        self.info_auto_hide_paused = False  # 信息栏自动隐藏暂停标志
        self.info_bar_animating = False

        # 初始化时获取并保存参考宽度
        self.root.update_idletasks()
        self.info_bar_ref_width = max(self.main_frame.winfo_width() - 10, 100)

        # 创建信息栏字体对象并保存，避免重复创建导致渲染不一致
        self._info_font = tkfont.Font(family=font_family, size=font_size)

        self.info_label.lift(self.info_close_btn)

        # 初始化图表数据
        self.chart_data = None

        # 初始化历史记录 - 委托给HistoryRepository
        self.history_records = self.history_repo.history_records

        # 创建临时标签用于测量文本宽度，避免重复创建和销毁
        if not hasattr(self, '_temp_label'):
            self._temp_label = tk.Label(self.root)
            self._temp_label.pack_forget()

        # 信息栏相关常量
        self.info_bar_left_offset = 235
        self.info_bar_right_offset = 136
        self.info_bar_padding = 3
        self.min_info_bar_width = 300
        self.close_btn_width = 30
        self.min_pixel_width = 50
        self.info_bar_place_left = 238
        self.info_bar_place_right = 136
        self.info_bar_place_y = 21.5
        self.info_bar_place_height = 30
        self.min_info_bar_place_width = 300

        """验证CIDR格式是否有效

        Args:
            cidr: 要验证的CIDR字符串

        Returns:
            bool: 如果CIDR格式有效则返回True，否则返回False
        """

    def update_history_listbox(self):
        """更新历史记录列表"""
        try:
            # 清空现有历史记录
            self.history_listbox.delete(0, tk.END)

            # 重新插入所有历史记录
            for index, history_record in enumerate(self.history_records, 1):
                # 格式化为: 1.  10.0.0.8/5 | 10.21.50.0/23
                formatted_record = f"{index}. {history_record['parent']}  |  {history_record['split']}"
                self.history_listbox.insert(tk.END, formatted_record)

            # 应用斑马条纹效果
            for index in range(self.history_listbox.size()):
                bg_color = "#d8d8d8" if (index + 1) % 2 == 0 else "#ffffff"
                self.history_listbox.itemconfigure(index, bg=bg_color)
        except (tk.TclError, AttributeError) as e:
            # 错误处理，确保GUI更新失败不会导致程序崩溃
            print(f"更新历史记录列表失败: {str(e)}")

    def reexecute_split(self):
        """从历史记录重新执行切分操作"""
        # 获取选中的历史记录索引
        selected_indices = self.history_listbox.curselection()
        if not selected_indices:
            self.show_info(_("hint"), _("please_select_a_history_record"))
            return

        # 获取选中项的索引
        selected_index = selected_indices[0]

        # 获取对应索引的历史记录
        if selected_index >= len(self.history_records):
            return

        history_record = self.history_records[selected_index]
        parent = history_record['parent']
        split = history_record['split']

        # 调用统一的切分方法
        self.perform_split(parent, split, from_history=True, auto_switch_version=True, fill_inputs=True)

    def save_current_state(self, action_type):
        """保存当前状态到操作记录中

        Args:
            action_type: 操作类型描述
        """
        # 获取当前子网需求
        subnet_requirements = []
        for item in self.requirements_tree.get_children():
            values = self.requirements_tree.item(item, "values")
            subnet_requirements.append((values[1], int(values[2])))

        # 获取当前需求池
        pool_requirements = []
        for item in self.pool_tree.get_children():
            values = self.pool_tree.item(item, "values")
            pool_requirements.append((values[1], int(values[2])))

        # 获取当前父网段
        parent = self.planning_parent_entry.get().strip()

        # 格式化需求信息
        req_str = ", ".join([f"{name}({hosts})" for name, hosts in subnet_requirements])
        pool_str = ", ".join([f"{name}({hosts})" for name, hosts in pool_requirements])

        # 创建操作记录
        history_record = {
            'action_type': action_type,
            'parent': parent,
            'requirements': subnet_requirements,
            'pool': pool_requirements,
            'timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'req_str': req_str,
            'pool_str': pool_str,
        }

        # 检查当前状态是否与上一个状态相同，如果相同则不保存
        if self.history_states:
            last_state = self.history_states[-1]
            # 比较子网需求、需求池和父网段
            if (last_state['requirements'] == subnet_requirements
                    and last_state['pool'] == pool_requirements
                    and last_state['parent'] == parent):
                return

        # 如果当前不是最新状态，截断历史记录
        if self.current_history_index < len(self.history_states) - 1:
            history_list = list(self.history_states)
            history_list = history_list[: self.current_history_index + 1]
            self.history_states = deque(history_list, maxlen=20)
            
            records_list = list(self.planning_history_records)
            records_list = records_list[: self.current_history_index + 1]
            self.planning_history_records = deque(records_list, maxlen=20)

        # 添加新状态(deque会自动管理大小,无需手动pop)
        self.history_states.append(history_record)
        self.planning_history_records.append(history_record)
        self.current_history_index += 1

        # 持久化子网需求和需求池数据到数据库
        self._persist_requirements_data()

    def _move_records_between_trees(self, source_tree, target_tree, selected_items, move_from, move_to):
        """通用方法：在两个树之间移动记录（支持多条记录，移动后保持选中）

        Args:
            source_tree: 源树控件
            target_tree: 目标树控件
            selected_items: 选中的项目ID列表
            move_from: 移动来源的描述（用于错误提示）
            move_to: 移动目标的描述（用于错误提示）

        Returns:
            list: 新插入的项目ID列表，用于后续选中操作
        """
        if not selected_items:
            self.show_info(_("hint"), _("please_select_records_to_move").format(record_type=move_from))
            return []

        # 先检查所有选中记录是否都可以移动
        # 同时收集要移动的记录数据
        items_to_move = []
        for selected_item in selected_items:
            values = source_tree.item(selected_item, "values")
            name = values[1]
            hosts = values[2]
            items_to_move.append({"name": name, "hosts": hosts})

            # 检查目标表中是否已存在相同名称的记录
            for item in target_tree.get_children():
                target_values = target_tree.item(item, "values")
                if target_values[1] == name:
                    self.show_error(_("error"), f"{move_to}{_("record_already_exists", name=name)}")
                    return []

        # 执行移动操作，并保存新插入记录的ID
        new_target_items = []

        # 先从源树删除所有选中记录
        for selected_item in selected_items:
            source_tree.delete(selected_item)

        # 然后插入到目标树，并保存新记录的ID
        for data in items_to_move:
            new_item_id = target_tree.insert("", tk.END, values=("", data["name"], data["hosts"]))
            new_target_items.append(new_item_id)

        # 更新序号和斑马条纹
        self.update_planning_tables_zebra_stripes()

        return new_target_items

    def move_left(self):
        """向左移：从子网需求表向需求池移动记录（支持多条记录，移动后保持选中）"""
        selected_items = self.requirements_tree.selection()
        new_items = self._move_records_between_trees(
            source_tree=self.requirements_tree,
            target_tree=self.pool_tree,
            selected_items=selected_items,
            move_from="子网需求表",
            move_to="需求池"
        )
        if new_items:
            self.save_current_state("移动记录：从子网需求表到需求池")

    def move_right(self):
        """向右移：从需求池向子网需求表移动记录（支持多条记录，移动后保持选中）"""
        selected_items = self.pool_tree.selection()
        new_items = self._move_records_between_trees(
            source_tree=self.pool_tree,
            target_tree=self.requirements_tree,
            selected_items=selected_items,
            move_from="需求池",
            move_to="子网需求表"
        )
        if new_items:
            self.save_current_state("移动记录：从需求池到子网需求表")

    def move_records(self):
        """根据选中情况自动判断移动方向：
        - 仅选中子网需求表数据：移动到需求池
        - 仅选中需求池数据：移动到子网需求表
        - 同时选中两个表数据：交换数据
        """
        # 获取两个表格中的选中记录
        selected_requirements = self.requirements_tree.selection()
        selected_pool_items = self.pool_tree.selection()

        # 情况1：仅选中子网需求表数据，移动到需求池
        if selected_requirements and not selected_pool_items:
            new_items = self._move_records_between_trees(
                source_tree=self.requirements_tree,
                target_tree=self.pool_tree,
                selected_items=selected_requirements,
                move_from="子网需求表",
                move_to="需求池"
            )
            if new_items:
                self.save_current_state("移动记录：从子网需求表到需求池")

        # 情况2：仅选中需求池数据，移动到子网需求表
        elif not selected_requirements and selected_pool_items:
            new_items = self._move_records_between_trees(
                source_tree=self.pool_tree,
                target_tree=self.requirements_tree,
                selected_items=selected_pool_items,
                move_from="需求池",
                move_to="子网需求表"
            )
            if new_items:
                self.save_current_state("移动记录：从需求池到子网需求表")

        # 情况3：同时选中两个表数据，交换数据
        elif selected_requirements and selected_pool_items:
            # 准备交换的记录数据
            req_items_to_move = []
            for item in selected_requirements:
                values = self.requirements_tree.item(item, "values")
                req_items_to_move.append({"name": values[1], "hosts": values[2]})

            pool_items_to_move = []
            for item in selected_pool_items:
                values = self.pool_tree.item(item, "values")
                pool_items_to_move.append({"name": values[1], "hosts": values[2]})

            # 检查交换后是否会导致重复名称
            req_names_to_swap = [data["name"] for data in req_items_to_move]
            pool_names_to_swap = [data["name"] for data in pool_items_to_move]

            # 检查需求池表
            all_pool_names = []
            for item in self.pool_tree.get_children():
                if item not in selected_pool_items:
                    values = self.pool_tree.item(item, "values")
                    all_pool_names.append(values[1])

            for name in req_names_to_swap:
                if name in all_pool_names:
                    self.show_error(_("error"), f"{_("requirements_pool")}{_("record_already_exists", name=name)}")
                    return

            # 检查子网需求表
            all_req_names = []
            for item in self.requirements_tree.get_children():
                if item not in selected_requirements:
                    values = self.requirements_tree.item(item, "values")
                    all_req_names.append(values[1])

            for name in pool_names_to_swap:
                if name in all_req_names:
                    self.show_error(_("error"), f"{_("subnet_requirements")}{_("record_already_exists", name=name)}")
                    return

            # 执行交换操作
            new_req_items = []
            new_pool_items = []

            # 删除所有选中的记录
            for item in selected_requirements:
                self.requirements_tree.delete(item)

            for item in selected_pool_items:
                self.pool_tree.delete(item)

            # 将需求池的记录添加到子网需求表
            for data in pool_items_to_move:
                new_item_id = self.requirements_tree.insert("", tk.END, values=("", data["name"], data["hosts"]))
                new_req_items.append(new_item_id)

            # 将子网需求表的记录添加到需求池
            for data in req_items_to_move:
                new_item_id = self.pool_tree.insert("", tk.END, values=("", data["name"], data["hosts"]))
                new_pool_items.append(new_item_id)

            # 更新两个表格的序号和斑马条纹
            self.update_planning_tables_zebra_stripes()

            # 保存交换操作到历史记录
            self.save_current_state("交换记录：子网需求表和需求池")

        # 情况4：未选中任何记录
        else:
            self.show_info(_("hint"), _("please_select_record_to_move_or_swap"))
            return

    def swap_records(self):
        """交换两个表格中选中的记录（支持多条记录，完全交换所有选中记录）"""
        selected_requirements = self.requirements_tree.selection()
        selected_pool_items = self.pool_tree.selection()

        # 检查是否同时选中了两个表格中的记录
        if not selected_requirements or not selected_pool_items:
            self.show_info(_("hint"), _("please_select_records_to_swap"))
            return

        # 1. 从子网需求表收集所有选中记录的数据
        req_items_to_move = []
        for item in selected_requirements:
            values = self.requirements_tree.item(item, "values")
            req_items_to_move.append({"name": values[1], "hosts": values[2]})

        # 2. 从需求池表收集所有选中记录的数据
        pool_items_to_move = []
        for item in selected_pool_items:
            values = self.pool_tree.item(item, "values")
            pool_items_to_move.append({"name": values[1], "hosts": values[2]})

        # 第一阶段：检查所有交换后是否会导致重复名称
        # 收集所有要交换的名称
        req_names_to_swap = [data["name"] for data in req_items_to_move]
        pool_names_to_swap = [data["name"] for data in pool_items_to_move]

        # 检查需求池表：要交换到需求池的req_names是否与需求池中已有的名称冲突（排除当前选中的pool_items）
        all_pool_names = []
        for item in self.pool_tree.get_children():
            if item not in selected_pool_items:
                values = self.pool_tree.item(item, "values")
                all_pool_names.append(values[1])

        for name in req_names_to_swap:
            if name in all_pool_names:
                self.show_error(_("error"), f"{_("requirements_pool")}{_("record_already_exists", name=name)}")
                return

        # 检查子网需求表：要交换到子网需求表的pool_names是否与子网需求表中已有的名称冲突（排除当前选中的req_items）
        all_req_names = []
        for item in self.requirements_tree.get_children():
            if item not in selected_requirements:
                values = self.requirements_tree.item(item, "values")
                all_req_names.append(values[1])

        for name in pool_names_to_swap:
            if name in all_req_names:
                self.show_error(_("error"), f"{_("subnet_requirements")}{_("record_already_exists", name=name)}")
                return

        # 第二阶段：执行完全交换操作
        swapped_records = []

        # 保存新插入记录的ID，用于交换后重新选中
        new_req_items = []
        new_pool_items = []

        # 1. 先删除所有选中的记录
        # 从子网需求表删除选中记录
        for item in selected_requirements:
            self.requirements_tree.delete(item)

        # 从需求池表删除选中记录
        for item in selected_pool_items:
            self.pool_tree.delete(item)

        # 2. 然后将对方的记录添加到自己的表格中
        # 将需求池的记录添加到子网需求表，并保存新插入记录的ID
        for data in pool_items_to_move:
            new_item_id = self.requirements_tree.insert("", tk.END, values=("", data["name"], data["hosts"]))
            new_req_items.append(new_item_id)
            swapped_records.append(f"{data['name']} ↔ ...")

        # 将子网需求表的记录添加到需求池，并保存新插入记录的ID
        for data in req_items_to_move:
            new_item_id = self.pool_tree.insert("", tk.END, values=("", data["name"], data["hosts"]))
            new_pool_items.append(new_item_id)

        self.update_planning_tables_zebra_stripes()

        # 保存交换操作到历史记录
        action_type = (
            f"交换记录: 子网需求表 {len(selected_requirements)} 条记录 ↔ 需求池 {len(selected_pool_items)} 条记录"
        )
        self.save_current_state(action_type)

    def create_split_input_section(self):
        """创建子网切分功能的输入区域"""
        # 设置 split_frame 的 grid 布局，实现两列等宽
        self.split_frame.grid_columnconfigure(0, weight=1, uniform="equal")
        self.split_frame.grid_columnconfigure(1, weight=1, uniform="equal")
        self.split_frame.grid_rowconfigure(0, weight=0)
        self.split_frame.grid_rowconfigure(1, weight=1)

        # 左侧：输入参数面板
        input_frame = ttk.LabelFrame(
            self.split_frame, text=_("input_parameters"), padding=(10, 10, 10, 10)
        )
        input_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=(0, 10))

        # 右侧：历史记录面板
        history_frame = ttk.LabelFrame(self.split_frame, text=_("history"), padding=(10, 10, 10, 10))
        history_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=(0, 10))

        # 添加IP版本切换框架
        ip_version_frame = ttk.Frame(input_frame)
        ip_version_frame.grid(row=0, column=0, columnspan=3, sticky="ew", pady=0)
        
        # 初始化IP版本变量
        self.split_ip_version_var = tk.StringVar(value="IPv4")
        
        # 添加IPv4/IPv6切换按钮
        ipv4_btn = ttk.Radiobutton(ip_version_frame, text="IPv4", variable=self.split_ip_version_var, value="IPv4", 
                                  command=self.on_split_ip_version_change, style="IpVersion.TRadiobutton")
        ipv4_btn.pack(side=tk.LEFT, padx=(10, 10))
        
        ipv6_btn = ttk.Radiobutton(ip_version_frame, text="IPv6", variable=self.split_ip_version_var, value="IPv6", 
                                  command=self.on_split_ip_version_change, style="IpVersion.TRadiobutton")
        ipv6_btn.pack(side=tk.LEFT)

        # 配置 input_frame 的 grid 行列
        input_frame.grid_columnconfigure(0, minsize=30, weight=0)
        input_frame.grid_columnconfigure(1, minsize=0, weight=1)
        input_frame.grid_columnconfigure(2, weight=0)
        input_frame.grid_rowconfigure(0, weight=0, minsize=0)
        input_frame.grid_rowconfigure(1, weight=0)
        input_frame.grid_rowconfigure(2, weight=0)
        input_frame.grid_rowconfigure(3, weight=0, minsize=0)
        input_frame.grid_rowconfigure(4, weight=0, minsize=0)

        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()

        # 父网段 - 统一pady、sticky和字体，确保与文本框垂直对齐
        ttk.Label(input_frame, text=_("parent_network"), anchor="w", font=(font_family, font_size)).grid(
            row=1, column=0, sticky=tk.W + tk.N + tk.S, pady=4, padx=(10, 0)
        )
        # 初始化IP版本相关数据
        # 从数据库加载每个IP版本的历史记录列表
        self.split_parent_networks_v4 = self.history_repo.split_parent_networks_v4
        self.split_parent_networks_v6 = self.history_repo.split_parent_networks_v6
        self.split_networks_v4 = self.history_repo.split_networks_v4
        self.split_networks_v6 = self.history_repo.split_networks_v6
        
        # 根据IP版本选择对应的历史记录列表
        ip_version = self.split_ip_version_var.get()
        if ip_version == "IPv4":
            # 使用IPv4历史记录
            self.split_parent_networks = self.split_parent_networks_v4
            self.split_networks = self.split_networks_v4
            default_parent = self.split_parent_networks_v4[0] if self.split_parent_networks_v4 else "10.0.0.0/8"
            default_split = self.split_networks_v4[0] if self.split_networks_v4 else "10.21.50.0/23"
        else:
            # 使用IPv6历史记录
            self.split_parent_networks = self.split_parent_networks_v6
            self.split_networks = self.split_networks_v6
            default_parent = self.split_parent_networks_v6[0] if self.split_parent_networks_v6 else "2001:0db8::/32"
            default_split = self.split_networks_v6[0] if self.split_networks_v6 else "2001:0db8::/64"

        # 父网段 - 使用Combobox，支持下拉选择和即时验证
        def validate_split_parent(p):
            return self.validate_cidr(p, self.parent_entry, ip_version=self.split_ip_version_var.get(), require_prefix=True)
        vcmd = (self.root.register(validate_split_parent), '%P')
        self.parent_entry = ttk.Combobox(
            input_frame,
            values=list(self.split_parent_networks),  # 使用过滤后的记录列表
            font=(font_family, font_size),
            validate='all',
            validatecommand=vcmd,
        )
        self.parent_entry.grid(row=1, column=1, padx=10, pady=4, sticky=tk.EW + tk.N + tk.S)
        self.parent_entry.insert(0, default_parent)  # 默认值
        self.parent_entry.config(state="normal")  # 允许手动输入
        # 添加IPv6自动补全功能
        self.parent_entry.bind('<KeyRelease>', self.autocomplete_ipv6)

        # 切分段 - 统一pady、sticky和字体，确保与文本框垂直对齐
        ttk.Label(input_frame, text=_("split_segments"), anchor="w", font=(font_family, font_size)).grid(
            row=2, column=0, sticky=tk.W + tk.N + tk.S, pady=4, padx=(10, 0)
        )

        def validate_split_segment(p):
            return self.validate_cidr(p, self.split_entry, ip_version=self.split_ip_version_var.get(), require_prefix=True)
        vcmd = (self.root.register(validate_split_segment), '%P')
        self.split_entry = ttk.Combobox(
            input_frame,
            values=list(self.split_networks),  # 使用过滤后的记录列表
            font=(font_family, font_size),
            validate='all',
            validatecommand=vcmd,
        )
        self.split_entry.grid(row=2, column=1, padx=10, pady=4, sticky=tk.EW + tk.N + tk.S)
        self.split_entry.insert(0, default_split)  # 默认值
        self.split_entry.config(state="normal")  # 允许手动输入
        # 添加IPv6自动补全功能
        self.split_entry.bind('<KeyRelease>', self.autocomplete_ipv6)

        # 按钮区域 - 执行按钮，跨四行的方形样式
        self.execute_btn = ttk.Button(input_frame, text=_("execute_split"), command=self.execute_split, width=10)
        self.execute_btn.grid(row=0, column=2, rowspan=4, padx=(0, 0), pady=0, sticky=tk.NSEW)

        # 配置 history_frame 的 grid 布局
        history_frame.grid_rowconfigure(0, weight=1)
        history_frame.grid_rowconfigure(1, weight=1)
        history_frame.grid_columnconfigure(0, weight=1)  # 表格列
        history_frame.grid_columnconfigure(1, weight=0)  # 滚动条列
        history_frame.grid_columnconfigure(2, weight=0)  # 按钮列

        # 创建历史记录列表 - 撑满整个框架
        self.history_listbox = tk.Listbox(
            history_frame, height=3, font=(font_family, font_size),
            highlightthickness=1, highlightbackground="#999999", highlightcolor="#999999",
            bd=0, selectbackground="#0078D7", selectforeground="white", takefocus=False
        )
        self.history_listbox.configure(activestyle="none")
        self.history_listbox.grid(row=0, column=0, sticky="nsew", rowspan=2)

        # 添加垂直滚动条
        history_scroll = ttk.Scrollbar(history_frame, orient=tk.VERTICAL)

        # 使用通用滚动条配置
        self._setup_scrollbar(history_scroll, self.history_listbox, initial_hidden=True)

        # 绑定右键菜单
        self.bind_listbox_right_click(self.history_listbox)

        # 初始化历史记录 - 确保在更新列表框之前赋值
        self.history_records = self.history_repo.history_records

        # 初始化历史记录列表（从数据库加载已有记录）
        self.update_history_listbox()

        # 创建重新切分按钮 - 与执行切分按钮样式一致
        self.reexecute_btn = ttk.Button(
            history_frame, text=_("reexecute_split"), command=self.reexecute_split, width=10
        )
        self.reexecute_btn.grid(row=0, column=2, rowspan=2, sticky="nsew", padx=(5, 0))

    def adjust_remaining_tree_width(self):
        """调整剩余网段表表格的宽度，使其自适应窗口大小"""
        self.remaining_tree.update_idletasks()
        frame_width = self.remaining_frame.winfo_width()
        self.remaining_tree.column("index", width=35)

        items = self.remaining_tree.get_children()
        columns = ["cidr", "network", "netmask", "wildcard", "broadcast", "usable"]

        if not items and frame_width > 0:
            # 表格为空时均分宽度
            total_columns = 6
            available_width = frame_width - 70
            column_width = max(100, available_width // total_columns)
            for col in columns:
                # 跳过已经隐藏的列
                if self.is_column_hidden(self.remaining_tree, col):
                    continue
                self.remaining_tree.column(col, width=column_width)
        elif items:
            # 表格有数据时自适应内容
            for col in columns:
                # 跳过已经隐藏的列
                if self.is_column_hidden(self.remaining_tree, col):
                    continue
                self.remaining_tree.column(col, width="0")
                self.remaining_tree.update_idletasks()
                auto_width = self.remaining_tree.column(col, "width")
                self.remaining_tree.column(col, width=max(100, auto_width))

    def is_column_hidden(self, tree, col):
        """检查表格列是否被隐藏
        
        Args:
            tree: Treeview对象
            col: 列名
            
        Returns:
            bool: 如果列被隐藏返回True，否则返回False
        """
        # 检查列是否被隐藏（宽度为0且stretch为False）
        col_width = tree.column(col, "width")
        col_stretch = tree.column(col, "stretch")
        return col_width == 0 and not col_stretch
    

    
    def adjust_allocated_tree_width(self):
        """调整已分配子网表表格的宽度，使其根据内容自动调整"""
        self.allocated_tree.update_idletasks()
        
        items = self.allocated_tree.get_children()
        columns = ["name", "cidr", "required", "available", "network", "netmask", "wildcard", "broadcast"]
        
        if items:
            # 表格有数据时，根据内容自适应列宽
            for col in columns:
                # 跳过已经隐藏的列
                if self.is_column_hidden(self.allocated_tree, col):
                    continue
                    
                # 将列宽设为0，触发自动计算
                self.allocated_tree.column(col, width="0")
                self.allocated_tree.update_idletasks()
                # 获取自动计算的宽度
                auto_width = self.allocated_tree.column(col, "width")
                # 设置列宽为基于内容的宽度
                self.allocated_tree.column(col, width=auto_width)

    def on_tab_change(self, tab_index):
        """标签页切换时的处理函数"""
        if tab_index == 2 and hasattr(self, 'chart_canvas'):
            self.draw_distribution_chart()
            
    def on_planning_tab_change(self, tab_index):
        """规划结果标签页切换时的处理函数"""
        # 确保UI更新完成
        self.root.update_idletasks()
        
        # 检查 _temp_label 是否存在，如果不存在则创建
        if not hasattr(self, '_temp_label'):
            self._temp_label = tk.Label(self.root)
            self._temp_label.pack_forget()
        
        if tab_index == 0:  # 已分配子网标签页
            # 调整已分配子网表的列宽
            self.auto_resize_columns(self.allocated_tree)
        elif tab_index == 1:  # 剩余网段标签页
            # 调整剩余网段表的列宽
            self.auto_resize_columns(self.planning_remaining_tree)
        elif tab_index == 2:  # 网段分布图标签页
            # 重新绘制图表
            if hasattr(self, 'chart_canvas'):
                self.draw_distribution_chart()

    def on_top_level_tab_change(self, tab_index):
        """顶级标签页切换时的处理函数"""
        # 检查是否有正在编辑的状态，有的话保存并清理
        if hasattr(self, 'current_edit_item') and self.current_edit_item is not None:
            # 直接销毁编辑框并清理状态，不保存
            if hasattr(self, 'edit_entry') and self.edit_entry:
                self.edit_entry.destroy()
            self._cleanup_edit_state()
        
    def _create_requirements_tree(self, parent, height=5, columns=("index", "name", "hosts"),
                                double_click_handler=None):
        """创建需求表格（子网需求表或需求池表）的通用方法

        Args:
            parent: 父容器
            height: 表格高度（行数）
            columns: 列名列表
            double_click_handler: 双击事件处理器

        Returns:
            ttk.Treeview: 创建的表格对象
        """
        tree = ttk.Treeview(parent, columns=columns, show="headings", height=height)
        self.bind_treeview_right_click(tree)

        # 设置表头
        tree.heading("index", text=_("index"))
        tree.heading("name", text=_("subnet_name"))
        tree.heading("hosts", text=_("host_count"))

        # 设置列宽
        tree.column("index", width=35, minwidth=35, stretch=False, anchor="e")
        tree.column("name", width=80, minwidth=80, stretch=True)
        tree.column("hosts", width=80, minwidth=40, stretch=False)

        self.configure_treeview_styles(tree)

        # 绑定事件
        if double_click_handler:
            tree.bind("<Double-1>", double_click_handler)
        tree.bind("<Button-1>", self.on_treeview_click)

        return tree

    def create_top_level_notebook(self):
        """创建顶级标签页控件，用于切换子网切分和子网规划两大功能模块"""
        self.top_level_notebook = ColoredNotebook(self.main_frame, style=self.style,
                                                 is_top_level=True, tab_change_callback=self.on_top_level_tab_change)
        self.top_level_notebook.pack(fill=tk.BOTH, expand=True)

        # 子网切分模块
        self.split_frame = ttk.Frame(self.top_level_notebook.content_area, padding="10")
        self.create_split_input_section()
        self.create_split_result_section()

        # 子网规划模块
        self.planning_frame = ttk.Frame(self.top_level_notebook.content_area, padding="10")
        self.setup_planning_page()

        # 高级工具模块
        self.advanced_frame = ttk.Frame(self.top_level_notebook.content_area, padding="10")
        self.setup_advanced_tools_page()
        
        # IP地址管理（IPAM）模块
        self.ipam_frame = ttk.Frame(self.top_level_notebook.content_area, padding="10")
        self.setup_ipam_page()

        # 添加顶级标签页（传入tab_id以支持跨语言保存标签次序）
        self.top_level_notebook.add_tab(_("ip_address_management"), self.ipam_frame, "#e3f2fd", tab_id="ip_address_management")
        self.top_level_notebook.add_tab(_("subnet_planning"), self.planning_frame, "#fce4ec", tab_id="subnet_planning")
        self.top_level_notebook.add_tab(_("subnet_split"), self.split_frame, "#fff3e0", tab_id="subnet_split")
        self.top_level_notebook.add_tab(_("advanced_tools"), self.advanced_frame, "#e8f5e9", tab_id="advanced_tools")

        # 加载保存的标签次序
        self.load_saved_tab_order()
    
    def load_saved_tab_order(self):
        """从配置文件加载保存的标签次序"""
        from config_manager import get_config
        
        try:
            config = get_config()
            saved_order = config.get_ui_tab_order()
            if saved_order and isinstance(saved_order, list) and len(saved_order) > 0:
                self.top_level_notebook.apply_tab_order(saved_order)
        except Exception as e:
            print(f"加载标签次序失败: {e}")

    def create_split_result_section(self):
        """创建子网切分功能的结果显示区域"""
        result_frame = ttk.LabelFrame(self.split_frame, text=_("split_result"), padding="10")
        result_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=(0, 0), pady=(0, 0))

        # 配置 result_frame 的 grid 布局
        result_frame.grid_rowconfigure(0, weight=1)
        result_frame.grid_columnconfigure(0, weight=1)

        # 导出结果按钮
        style_manager = get_style_manager()
        btn_width, __ = style_manager.get_button_size("export_result") if style_manager else (10, 25)
        self.export_btn = ttk.Button(result_frame, text=_("export_result"), command=self.export_result, width=btn_width)
        self.export_btn.place(relx=1.0, rely=0.0, anchor=tk.NE, x=0, y=-3)

        # 创建自定义笔记本控件
        self.notebook = ColoredNotebook(result_frame, style=self.style, tab_change_callback=self.on_tab_change)
        self.notebook.grid(row=0, column=0, sticky="nsew")
        self.export_btn.lift()

        # 切分段信息页面
        self.split_info_frame = ttk.Frame(self.notebook.content_area, padding="5", style=self.notebook.light_blue_style)
        
        # 创建切分段信息表格
        self.split_tree = ttk.Treeview(self.split_info_frame, columns=("item", "value"), show="headings", height=5)
        # 添加右键复制功能
        self.bind_treeview_right_click(self.split_tree)
        self.split_tree.heading("item", text=_("item"))
        self.split_tree.heading("value", text=_("value"))
        # 设置合适的列宽
        self.split_tree.column("item", width=120, minwidth=120, stretch=False)
        self.split_tree.column("value", width=250)
        self.split_tree.pack(fill=tk.BOTH, expand=True, pady=0)

        # 配置斑马条纹样式和信息标签样式
        self.configure_treeview_styles(self.split_tree, include_special_tags=True)

        # 剩余网段表页面
        self.remaining_frame = ttk.Frame(self.notebook.content_area, padding="5", style=self.notebook.light_green_style)

        # 创建剩余网段信息表格
        self.remaining_tree = ttk.Treeview(
            self.remaining_frame,
            columns=("index", "cidr", "network", "end_address", "netmask", "wildcard", "broadcast", "usable"),
            show="headings",
            height=5,
        )
        self.bind_treeview_right_click(self.remaining_tree)
        self.remaining_tree.heading("index", text=_("index"))
        self.remaining_tree.heading("cidr", text=_("cidr"))
        self.remaining_tree.heading("network", text=_("network_address"))
        self.remaining_tree.heading("end_address", text=_("network_end_address"))
        self.remaining_tree.heading("netmask", text=_("subnet_mask"))
        self.remaining_tree.heading("wildcard", text=_("wildcard_mask"))
        self.remaining_tree.heading("broadcast", text=_("broadcast_address"))
        self.remaining_tree.heading("usable", text=_("usable_address_count"))

        # 设置列宽，使用minwidth替代width，让列可以自适应
        self.remaining_tree.column("index", minwidth=35, width=35, stretch=False, anchor="e")
        self.remaining_tree.column("cidr", minwidth=100, width=120, stretch=True)
        self.remaining_tree.column("network", minwidth=100, width=120, stretch=True)
        self.remaining_tree.column("end_address", minwidth=100, width=0, stretch=False)  # 初始隐藏网段结束地址列
        self.remaining_tree.column("netmask", minwidth=100, width=120, stretch=True)
        self.remaining_tree.column("wildcard", minwidth=100, width=120, stretch=True)
        self.remaining_tree.column("broadcast", minwidth=100, width=120, stretch=True)
        self.remaining_tree.column("usable", minwidth=70, width=80, stretch=True)  # 初始就窄化可用地址数列，因为使用科学计数法

        # 配置斑马条纹样式
        self.configure_treeview_styles(self.remaining_tree)

        # 网段分布图页面
        self.chart_frame = ttk.Frame(self.notebook.content_area, padding="5", style=self.notebook.light_purple_style)

        # 添加标签页，每个标签页设置不同的颜色
        self.notebook.add_tab(_("split_segment_info"), self.split_info_frame, "#e3f2fd")  # 浅蓝色
        self.notebook.add_tab(_("remaining_subnets"), self.remaining_frame, "#e8f5e9")  # 浅绿色
        self.notebook.add_tab(_("distribution_chart"), self.chart_frame, "#f3e5f5")  # 浅紫色

        # 配置chart_frame的grid布局
        self.chart_frame.grid_rowconfigure(0, weight=1)
        self.chart_frame.grid_columnconfigure(0, weight=1)
        self.chart_frame.grid_columnconfigure(1, weight=0)

        # 添加滚动条
        self.chart_scrollbar = ttk.Scrollbar(self.chart_frame, orient=tk.VERTICAL)

        # 创建Canvas用于绘制柱状图，设置背景色为深灰色以匹配图表背景
        # 禁止水平滚动，只允许垂直滚动
        self.chart_canvas = tk.Canvas(self.chart_frame, bg="#333333", highlightthickness=0)
        self.chart_canvas.grid(row=0, column=0, sticky=tk.NSEW, pady=0)

        # 使用通用滚动条配置
        self._setup_scrollbar(self.chart_scrollbar, self.chart_canvas, initial_hidden=False, widget_command=self.chart_canvas.yview)
        self.chart_canvas.config(xscrollcommand=None)
        self.chart_scrollbar.grid(row=0, column=1, sticky=tk.NS)

        # 绑定窗口大小变化事件，实现图表自适应
        self.chart_canvas.bind("<Configure>", self.on_chart_resize)
        # 绑定鼠标滚轮事件
        self.chart_canvas.bind("<MouseWheel>", self.on_chart_mousewheel)

        # 配置remaining_frame的grid布局
        self.remaining_frame.grid_rowconfigure(0, weight=1)
        self.remaining_frame.grid_columnconfigure(0, weight=1)
        self.remaining_frame.grid_columnconfigure(1, weight=0)

        # 垂直滚动条
        self.remaining_scroll_v = ttk.Scrollbar(
            self.remaining_frame, orient=tk.VERTICAL
        )
        # 水平滚动条
        self.remaining_scroll_h = ttk.Scrollbar(
            self.remaining_frame, orient=tk.HORIZONTAL
        )

        # 使用通用方法创建带自动隐藏垂直滚动条的Treeview，滚动条隐藏时不添加右边距
        self.create_scrollable_treeview(self.remaining_frame, self.remaining_tree, self.remaining_scroll_v, no_scrollbar_padx=(0, 0))
        
        # 使用通用方法创建带自动隐藏功能的水平滚动条
        self.create_horizontal_scrollbar(self.remaining_frame, self.remaining_tree, self.remaining_scroll_h)
        
        # 配置remaining_frame的列布局
        self.remaining_frame.grid_columnconfigure(0, weight=1)
        self.remaining_frame.grid_columnconfigure(1, weight=0)

        # 绑定窗口大小变化事件，实现表格自适应
        self.root.bind("<Configure>", self.on_window_resize, add='+')

        # 初始提示
        self.clear_result()

        # Treeview表格线样式已在初始化时设置

        # 在窗口完全渲染后再调用动态计算方法，确保获取准确的高度
        self.root.after(100, self.setup_table_styles)

    def setup_table_styles(self):
        """设置表格样式"""
        # 图表滚动条已在初始化时配置，此处仅需要确保滚动条可见
        self.chart_scrollbar.grid(row=0, column=1, sticky=tk.NS)

        self.chart_canvas.bind("<Configure>", self.on_chart_resize)
        self.chart_canvas.bind("<MouseWheel>", self.on_chart_mousewheel)

    def _add_result_tabs(self):
        """添加标签页到笔记本"""
        self.notebook.add_tab(_("split_segment_info"), self.split_info_frame, "#e3f2fd")  # 浅蓝色
        self.notebook.add_tab(_("remaining_subnets"), self.remaining_frame, "#e8f5e9")  # 浅绿色
        self.notebook.add_tab(_("distribution_chart"), self.chart_frame, "#f3e5f5")  # 浅紫色

    def _setup_scrollbars(self):
        """配置滚动条"""
        self.remaining_frame.grid_rowconfigure(0, weight=1)
        self.remaining_frame.grid_columnconfigure(0, weight=1)
        self.remaining_frame.grid_columnconfigure(1, weight=0)

        self.remaining_scroll_v = ttk.Scrollbar(
            self.remaining_frame, orient=tk.VERTICAL, command=self.remaining_tree.yview
        )

        def remaining_scrollbar_callback(*args):
            self.remaining_scroll_v.set(*args)
            if float(args[0]) <= 0.0 and float(args[1]) >= 1.0:
                self.remaining_scroll_v.grid_remove()
            else:
                self.remaining_scroll_v.grid(row=0, column=1, sticky=tk.NS)

        self.remaining_tree.configure(yscrollcommand=remaining_scrollbar_callback)

        self.remaining_tree.grid(row=0, column=0, sticky=tk.NSEW)
        self.remaining_scroll_v.grid(row=0, column=1, sticky=tk.NS)
        remaining_scrollbar_callback(0.0, 1.0)

    def _setup_initial_state(self):
        """设置初始状态"""
        self.root.bind("<Configure>", self.on_window_resize, add='+')
        self.clear_result()
        self.root.after(100, self.setup_table_styles)

    def setup_planning_page(self):
        """设置子网规划功能的界面"""
        # 直接使用self.planning_frame，移除中间层main_planning_frame

        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 设置grid布局
        self.planning_frame.grid_columnconfigure(0, weight=1)  # 左侧列可伸缩
        self.planning_frame.grid_columnconfigure(1, weight=1)  # 右侧列可伸缩
        self.planning_frame.grid_rowconfigure(0, weight=0)  # 父网段设置行，固定高度
        self.planning_frame.grid_rowconfigure(1, weight=0)  # 需求池和子网需求行，固定高度
        self.planning_frame.grid_rowconfigure(2, weight=1)  # 规划结果行，可伸缩

        # 父网段设置区域
        parent_frame = ttk.LabelFrame(self.planning_frame, text=_('parent_network_settings'), padding=(5, 10, 10, 10))
        parent_frame.grid(row=0, column=0, sticky="ew", padx=(0, 5), pady=(0, 0))  # 左上角
        # 设置父网段设置面板的固定宽度
        parent_frame.configure(width=250)

        # 初始化IP版本变量
        self.ip_version_var = tk.StringVar(value="IPv4")
        
        # 初始化父网段列表 - 从数据库加载每个IP版本的历史记录
        self.planning_parent_networks_v4 = self.history_repo.planning_parent_networks_v4
        self.planning_parent_networks_v6 = self.history_repo.planning_parent_networks_v6
        self.planning_parent_networks = self.planning_parent_networks_v4

        # 创建父网段输入区域框架，用于水平排列IP选项和输入框
        parent_input_frame = ttk.Frame(parent_frame)
        parent_input_frame.pack(fill=tk.X, pady=(0, 5))
        
        # 添加IPv4/IPv6切换按钮 - 水平排列
        ipv4_btn = ttk.Radiobutton(parent_input_frame, text="IPv4", variable=self.ip_version_var, value="IPv4", 
                                  command=self.on_ip_version_change, style="IpVersion.TRadiobutton")
        ipv4_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        ipv6_btn = ttk.Radiobutton(parent_input_frame, text="IPv6", variable=self.ip_version_var, value="IPv6", 
                                  command=self.on_ip_version_change, style="IpVersion.TRadiobutton")
        ipv6_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 父网段下拉文本框 - 与IP选项水平排列
        def validate_planning_parent(p):
            return self.validate_cidr(p, self.planning_parent_entry, ip_version=self.ip_version_var.get(), require_prefix=True)
        vcmd = (self.root.register(validate_planning_parent), '%P')
        self.planning_parent_entry = ttk.Combobox(
            parent_input_frame,
            values=list(self.planning_parent_networks),  # 使用包含两条记录的列表
            width=8,
            font=(font_family, font_size),
            validate='all',
            validatecommand=vcmd,
        )
        self.planning_parent_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        # 使用历史记录的第一个元素作为默认值，如果没有则使用默认值
        default_parent = self.planning_parent_networks[0] if self.planning_parent_networks else "10.21.48.0/20"
        self.planning_parent_entry.insert(0, default_parent)  # 默认值
        self.planning_parent_entry.config(state="normal")  # 允许手动输入
        # 添加IPv6自动补全功能
        self.planning_parent_entry.bind('<KeyRelease>', self.autocomplete_ipv6)

        # 需求池区域
        history_frame = ttk.LabelFrame(self.planning_frame, text=_('requirements_pool'), padding=(10, 10, 0, 10))
        history_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 5), pady=(0, 10))  # 左下角
        # 设置需求池面板的固定宽度
        history_frame.configure(width=250)

        # 子网需求区域
        requirements_frame = ttk.LabelFrame(self.planning_frame, text=_('subnet_requirements'), padding=(10, 10, 0, 10))
        requirements_frame.grid(
            row=0, column=1, rowspan=2, sticky="nsew", padx=(5, 0), pady=(0, 10)
        )  # 右侧跨两行
        # 设置子网需求面板的固定宽度
        requirements_frame.configure(width=250)

        # 内部容器框架，用于组织表格和按钮
        inner_frame = ttk.Frame(requirements_frame)
        inner_frame.pack(fill=tk.BOTH, expand=True)

        history_frame.grid_rowconfigure(0, weight=1)
        history_frame.grid_rowconfigure(1, weight=0)
        history_frame.grid_columnconfigure(0, weight=1)
        history_frame.grid_columnconfigure(1, weight=0)

        # 创建需求池表格
        self.pool_tree = self._create_requirements_tree(
            history_frame, height=6, double_click_handler=None
        )


        # 添加滚动条，确保只作用于表格，位于表格右侧
        self.pool_scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL)

        # 直接创建Treeview和滚动条
        self.pool_tree.grid(row=0, column=0, sticky=tk.NSEW)
        self.pool_scrollbar.grid(row=0, column=1, sticky=tk.NS)
        self.pool_scrollbar.config(command=self.pool_tree.yview)
        self.pool_tree.config(yscrollcommand=self.pool_scrollbar.set)

        inner_frame.grid_rowconfigure(0, weight=1)
        inner_frame.grid_columnconfigure(0, weight=0)  # 按钮列，固定宽度
        inner_frame.grid_columnconfigure(1, weight=1)  # 表格列，可伸缩
        inner_frame.grid_columnconfigure(2, weight=0)  # 滚动条列，固定宽度

        # 子网需求操作按钮框架
        button_frame = ttk.Frame(inner_frame)
        button_frame.grid(row=0, column=0, sticky="nsew")
        # 设置按钮框架的最小宽度，确保两个按钮大小一致
        button_frame.configure(width=70)

        # 子网需求表格
        self.requirements_tree = self._create_requirements_tree(
            inner_frame, height=5, double_click_handler=None
        )

        # 放置表格
        self.requirements_tree.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        self.requirements_scrollbar = ttk.Scrollbar(inner_frame, orient=tk.VERTICAL)
        
        # 注册子网需求表和需求池表的内联编辑配置
        # 两表结构相同，可编辑name和hosts列
        for tree_name, tree in [('requirements', self.requirements_tree), ('pool', self.pool_tree)]:
            # 注册内联编辑配置
            self.register_inline_edit_config(tree_name, {
                'editable_columns': [1, 2],  # 允许编辑name和hosts列
                'column_types': {
                    1: 'entry',  # name列使用文本框
                    2: 'entry'   # hosts列使用文本框
                }
            })
            
            # 注册内联编辑处理器
            self.register_inline_edit_handler(tree_name, {
                'get_row_data': lambda item, tree_name=tree_name: self._get_requirements_row_data(item, tree_name),
                'validate': self._validate_requirements_edit,
                'save': lambda new_value, column_name, row_data, item, tree_name=tree_name: self._save_requirements_edit(new_value, column_name, row_data, item, tree_name)
            })
            
            # 绑定双击事件
            tree.bind('<Double-1>', lambda event, tree=tree, tree_name=tree_name: self.on_generic_tree_double_click(tree, tree_name, event))

        # self.create_scrollable_treeview_with_grid(
        #     inner_frame, self.requirements_tree, self.requirements_scrollbar,
        #     tree_row=0, tree_column=1, scrollbar_row=0, scrollbar_column=2,
        #     tree_padx=(10, 0), scrollbar_padx=(0, 0)
        # )

        self.requirements_tree.grid(row=0, column=1, sticky=tk.NSEW)
        self.requirements_scrollbar.grid(row=0, column=2, sticky=tk.NS)
        self.requirements_scrollbar.config(command=self.requirements_tree.yview)
        self.requirements_tree.config(yscrollcommand=self.requirements_scrollbar.set)

        # 允许同时选择两张表中的记录，移除选择事件绑定

        # 按钮框架内部布局 - 按照用户要求设置行权重
        button_frame.grid_rowconfigure(0, weight=0)  # 添加按钮
        button_frame.grid_rowconfigure(1, weight=0)  # 删除按钮
        button_frame.grid_rowconfigure(2, weight=0)  # 撤销按钮
        button_frame.grid_rowconfigure(3, weight=0)  # 导入按钮
        button_frame.grid_rowconfigure(4, weight=1)  # 空白区域，将底部按钮推到底部
        button_frame.grid_rowconfigure(5, weight=0)  # 空白行，保持原有结构
        button_frame.grid_rowconfigure(6, weight=0)  # 交换记录按钮
        button_frame.grid_columnconfigure(0, weight=1)

        # 添加按钮
        add_btn = ttk.Button(button_frame, text=_('add'), command=self.add_subnet_requirement, width=7, style="Function.TButton")
        add_btn.grid(row=0, column=0, sticky="ew", pady=(0, 5))

        # 删除按钮
        delete_btn = ttk.Button(button_frame, text=_('delete'), command=self.delete_subnet_requirement, width=7, style="Function.TButton")
        delete_btn.grid(row=1, column=0, sticky="ew", pady=(0, 5))

        # 撤销按钮
        self.undo_delete_btn = ttk.Button(button_frame, text=_('undo'), command=self.undo, width=7, style="Function.TButton")
        self.undo_delete_btn.grid(row=2, column=0, sticky="ew", pady=(0, 5))

        # 移动/交换按钮（根据选中情况自动判断操作）
        # 交换记录按钮 - 使用交换图标
        self.swap_btn = ttk.Button(button_frame, text=_('move_records'), command=self.move_records, width=7, style="Move.TButton")
        self.swap_btn.grid(row=3, column=0, sticky="ew", pady=(0, 5))

        # 导入按钮
        import_btn = ttk.Button(button_frame, text=_('import'), command=self.import_requirements, width=7, style="Function.TButton")
        import_btn.grid(row=6, column=0, sticky="ew", pady=(0, 0))


        # 从数据库加载子网需求和需求池数据
        requirements_data = self.history_repo.load_requirements_data(HistorySQLite.TABLE_REQUIREMENTS)
        if requirements_data:
            for index, (name, hosts) in enumerate(requirements_data, 1):
                tag = "even" if index % 2 == 0 else "odd"
                self.requirements_tree.insert("", tk.END, values=("", name, str(hosts)), tags=(tag,))
        else:
            sample_requirements = [
                ("office", "20"), ("hr_department", "10"), ("finance_department", "10"),
                ("planning_department", "30"), ("legal", "10"), ("procurement", "10"),
                ("security", "10"), ("party", "20"), ("discipline", "10"), ("it_department", "20"),
            ]
            for index, (name_key, hosts) in enumerate(sample_requirements, 1):
                tag = "even" if index % 2 == 0 else "odd"
                self.requirements_tree.insert("", tk.END, values=("", _(name_key), hosts), tags=(tag,))

        pool_data = self.history_repo.load_requirements_data(HistorySQLite.TABLE_POOL)
        if pool_data:
            for index, (name, hosts) in enumerate(pool_data, 1):
                tag = "even" if index % 2 == 0 else "odd"
                self.pool_tree.insert("", tk.END, values=("", name, str(hosts)), tags=(tag,))
        else:
            sample_pool = [
                ("engineering", "20"), ("sales", "20"), ("rd", "15"),
                ("production", "100"), ("transportation", "20"),
            ]
            for index, (name_key, hosts) in enumerate(sample_pool, 1):
                tag = "even" if index % 2 == 0 else "odd"
                self.pool_tree.insert("", tk.END, values=("", _(name_key), hosts), tags=(tag,))

        # 调用方法更新序号和斑马条纹
        self.update_planning_tables_zebra_stripes()

        self.configure_treeview_styles(self.requirements_tree)
        self.configure_treeview_styles(self.pool_tree)  # 配置需求池表格样式

        # 设置表格选择模式为多选，允许一次选择多条记录
        self.requirements_tree.configure(selectmode=tk.EXTENDED)
        self.pool_tree.configure(selectmode=tk.EXTENDED)

        # 删除原来的执行规划按钮容器
        # 按钮已移动到删除按钮下方

        # 规划结果区域 - 使用grid布局，跨两列显示
        result_frame = ttk.LabelFrame(self.planning_frame, text=_('planning_result'), padding="10")
        result_frame.grid(row=2, column=0, columnspan=2, sticky="nwse", pady=(0, 0))

        # 创建笔记本控件显示规划结果
        self.planning_notebook = ColoredNotebook(result_frame, style=self.style, tab_change_callback=self.on_planning_tab_change)
        self.planning_notebook.pack(fill=tk.BOTH, expand=True)

        # 保存初始状态到历史记录
        self.save_current_state("初始状态")

        # 设置统一的按钮宽度，使用合适的宽度确保文字完全显示
        style_manager = get_style_manager()
        button_width, __ = style_manager.get_button_size("export_planning") if style_manager else (10, 25)

        # 使用通用方法配置按钮样式
        self._setup_accent_button_style("Accent.TButton", "#1565c0", "#0d47a1", "#0d47a1")
        self._setup_accent_button_style("RedAccent.TButton", "#2e7d32", "#1b5e20", "#1b5e20")

        # 按钮间距
        button_gap = 10

        # 同步到地址管理按钮 - 位于最右边
        sync_btn_width, __ = style_manager.get_button_size("sync_to_ipam") if style_manager else (12, 25)
        self.sync_to_ipam_btn = ttk.Button(
            result_frame, text=_("sync_to_ipam"), command=self.sync_allocated_to_ipam, width=sync_btn_width
        )
        # 先更新窗口，确保能获取到按钮的实际宽度
        self.root.update_idletasks()
        self.sync_to_ipam_btn.place(relx=1.0, rely=0.0, anchor=tk.NE, x=0, y=-3)
        sync_btn_width_actual = self.sync_to_ipam_btn.winfo_reqwidth()

        # 导出规划按钮 - 位于同步按钮左边
        export_btn_width, __ = style_manager.get_button_size("export_planning") if style_manager else (10, 25)
        export_planning_btn = ttk.Button(
            result_frame, text=_('export_planning'), command=self.export_planning_result, width=export_btn_width
        )
        export_btn_x = -sync_btn_width_actual - button_gap
        export_planning_btn.place(relx=1.0, rely=0.0, anchor=tk.NE, x=export_btn_x, y=-3)
        export_btn_width_actual = export_planning_btn.winfo_reqwidth()

        # 执行规划按钮 - 位于导出规划按钮左边（最左边）
        execute_btn_width, __ = style_manager.get_button_size("execute_planning") if style_manager else (10, 25)
        self.execute_planning_btn = ttk.Button(
            result_frame, text=_("execute_planning"), command=self.execute_subnet_planning, width=execute_btn_width
        )
        execute_btn_x = -sync_btn_width_actual - button_gap - export_btn_width_actual - button_gap
        self.execute_planning_btn.place(relx=1.0, rely=0.0, anchor=tk.NE, x=execute_btn_x, y=-3)

        # 已分配子网页面
        self.allocated_frame = ttk.Frame(
            self.planning_notebook.content_area, padding="5", style=self.planning_notebook.light_blue_style
        )
        self.allocated_tree = ttk.Treeview(
            self.allocated_frame,
            columns=("index", "name", "cidr", "required", "available", "network", "end_address", "netmask", "wildcard", "broadcast"),
            show="headings",
            height=5,  # 设置为5行高度
        )

        self.bind_treeview_right_click(self.allocated_tree)

        # 设置列标题
        self.allocated_tree.heading("index", text=_("index"))
        self.allocated_tree.heading("name", text=_("subnet_name"))
        self.allocated_tree.heading("cidr", text=_("cidr"))
        self.allocated_tree.heading("required", text=_("required_count"))
        self.allocated_tree.heading("available", text=_("available_count"))
        self.allocated_tree.heading("network", text=_("network_address"))
        self.allocated_tree.heading("end_address", text=_("network_end_address"))
        self.allocated_tree.heading("netmask", text=_("subnet_mask"))
        self.allocated_tree.heading("wildcard", text=_("wildcard_mask"))
        self.allocated_tree.heading("broadcast", text=_("broadcast_address"))

        # 设置列宽为自动，根据内容自动调整宽度
        self.allocated_tree.column("index", width=35, minwidth=35, stretch=False, anchor="e")  # 序号列固定宽度40
        self.allocated_tree.column("name", width=0, minwidth=90, stretch=True)  # 子网名称列自动宽度
        self.allocated_tree.column("cidr", width=0, minwidth=90, stretch=True)  # CIDR列自动宽度
        self.allocated_tree.column("required", width=0, minwidth=35, stretch=True)  # 需求数列自动宽度
        self.allocated_tree.column("available", width=0, minwidth=35, stretch=True)  # 可用数列自动宽度
        self.allocated_tree.column("network", width=0, minwidth=80, stretch=True)  # 网络地址列自动宽度
        self.allocated_tree.column("end_address", width=0, minwidth=80, stretch=False)  # 网段结束地址列初始隐藏
        self.allocated_tree.column("netmask", width=0, minwidth=80, stretch=True)  # 子网掩码列自动宽度
        self.allocated_tree.column("wildcard", width=0, minwidth=80, stretch=True)  # 通配符掩码列自动宽度
        self.allocated_tree.column("broadcast", width=0, minwidth=80, stretch=True)  # 广播地址列自动宽度

        # 垂直滚动条
        allocated_v_scrollbar = ttk.Scrollbar(
            self.allocated_frame, orient=tk.VERTICAL
        )
        # 水平滚动条
        allocated_h_scrollbar = ttk.Scrollbar(
            self.allocated_frame, orient=tk.HORIZONTAL
        )

        # 使用通用方法创建带自动隐藏垂直滚动条的Treeview，滚动条隐藏时不添加右边距
        self.create_scrollable_treeview(self.allocated_frame, self.allocated_tree, allocated_v_scrollbar, no_scrollbar_padx=(0, 0))
        
        # 使用通用方法创建带自动隐藏功能的水平滚动条
        self.create_horizontal_scrollbar(self.allocated_frame, self.allocated_tree, allocated_h_scrollbar)
        
        # 配置allocated_frame的grid布局
        self.allocated_frame.grid_rowconfigure(0, weight=1)
        self.allocated_frame.grid_columnconfigure(0, weight=1)

        self.configure_treeview_styles(self.allocated_tree)
        
        # 初始化时根据表头自动调整列宽
        self.auto_resize_columns(self.allocated_tree)

        # 剩余网段页面
        self.planning_remaining_frame = ttk.Frame(
            self.planning_notebook.content_area, padding="5", style=self.planning_notebook.light_green_style
        )
        self.planning_remaining_tree = ttk.Treeview(
            self.planning_remaining_frame,
            columns=("index", "cidr", "network", "end_address", "netmask", "wildcard", "broadcast", "usable"),
            show="headings",
            height=5,  # 设置为5行高度
        )

        self.bind_treeview_right_click(self.planning_remaining_tree)

        self.planning_remaining_tree.heading("index", text=_("index"))
        self.planning_remaining_tree.heading("cidr", text=_("cidr"))
        self.planning_remaining_tree.heading("network", text=_("network_address"))
        self.planning_remaining_tree.heading("end_address", text=_("network_end_address"))
        self.planning_remaining_tree.heading("netmask", text=_("subnet_mask"))
        self.planning_remaining_tree.heading("wildcard", text=_("wildcard_mask"))
        self.planning_remaining_tree.heading("broadcast", text=_("broadcast_address"))
        self.planning_remaining_tree.heading("usable", text=_("usable_address_count"))

        # 设置列宽，所有列都启用拉伸以实现自适应
        self.planning_remaining_tree.column("index", width=35, minwidth=35, stretch=False, anchor="e")
        self.planning_remaining_tree.column("cidr", width=120, minwidth=100, stretch=True)
        self.planning_remaining_tree.column(
            "network", width=120, minwidth=100, stretch=True
        )  # 加宽网络地址列以完整显示IPv6地址
        self.planning_remaining_tree.column("end_address", width=0, minwidth=100, stretch=False)  # 网段结束地址列，初始隐藏
        self.planning_remaining_tree.column("netmask", width=120, minwidth=100, stretch=True)
        self.planning_remaining_tree.column("wildcard", width=120, minwidth=100, stretch=True)
        self.planning_remaining_tree.column("broadcast", width=120, minwidth=100, stretch=True)
        self.planning_remaining_tree.column("usable", width=70, minwidth=60, stretch=True)  # 窄化可用地址数列，因为使用科学计数法

        # 垂直滚动条
        remaining_v_scrollbar = ttk.Scrollbar(
            self.planning_remaining_frame,
            orient=tk.VERTICAL,
        )
        # 水平滚动条
        remaining_h_scrollbar = ttk.Scrollbar(
            self.planning_remaining_frame,
            orient=tk.HORIZONTAL,
        )

        # 使用通用方法创建带自动隐藏垂直滚动条的Treeview，滚动条隐藏时不添加右边距
        self.create_scrollable_treeview(self.planning_remaining_frame, self.planning_remaining_tree, remaining_v_scrollbar, no_scrollbar_padx=(0, 0))
        
        # 使用通用方法创建带自动隐藏功能的水平滚动条
        self.create_horizontal_scrollbar(self.planning_remaining_frame, self.planning_remaining_tree, remaining_h_scrollbar)
        
        # 配置planning_remaining_frame的grid布局
        self.planning_remaining_frame.grid_rowconfigure(0, weight=1)
        self.planning_remaining_frame.grid_columnconfigure(0, weight=1)

        self.configure_treeview_styles(self.planning_remaining_tree)
        
        # 初始化时根据表头自动调整列宽
        self.auto_resize_columns(self.planning_remaining_tree)

        # 添加标签页 - 使用与切分结果一致的颜色
        self.planning_notebook.add_tab(_("allocated_subnets"), self.allocated_frame, "#e3f2fd")  # 浅蓝色
        self.planning_notebook.add_tab(_("remaining_subnets"), self.planning_remaining_frame, "#e8f5e9")  # 浅绿色

        self.planning_chart_frame = ttk.Frame(
            self.planning_notebook.content_area, padding="5", style=self.planning_notebook.light_purple_style
        )
        
        # 创建Canvas用于绘制图表
        self.planning_chart_canvas = tk.Canvas(
            self.planning_chart_frame,
            bg="#333333",
            highlightthickness=0
        )
        
        self.planning_chart_v_scrollbar = ttk.Scrollbar(
            self.planning_chart_frame,
            orient=tk.VERTICAL,
        )

        # 使用通用滚动条配置
        self._setup_scrollbar(self.planning_chart_v_scrollbar, self.planning_chart_canvas, initial_hidden=False, widget_command=self.planning_chart_canvas.yview)

        # 使用grid布局
        self.planning_chart_frame.grid_rowconfigure(0, weight=1)
        self.planning_chart_frame.grid_columnconfigure(0, weight=1)

        self.planning_chart_canvas.grid(row=0, column=0, sticky="nsew")
        self.planning_chart_v_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # 添加鼠标滚轮事件支持
        self.planning_chart_canvas.bind("<MouseWheel>", self.on_planning_chart_mousewheel)
        
        # 添加resize事件支持，确保图表能自适应窗口大小
        self.planning_chart_canvas.bind("<Configure>", self.on_planning_chart_resize)
        
        # 为规划模块添加图表标签页
        self.planning_notebook.add_tab(_("distribution_chart"), self.planning_chart_frame, "#f3e5f5")

        # 添加窗口大小变化事件处理，确保表格能自适应宽度
        self.planning_notebook.content_area.bind('<Configure>', lambda e: self.resize_tables())

        # 为规划模块表格添加空行或示例数据，显示斑马条纹效果
        for item in self.allocated_tree.get_children():
            self.allocated_tree.delete(item)
        for item in self.planning_remaining_tree.get_children():
            self.planning_remaining_tree.delete(item)
        
        # 在主循环启动后，再调整列宽，确保表格完全渲染
        # 使用after(100)确保表格完全渲染后再调整列宽
        self.root.after(100, self.auto_resize_columns, self.allocated_tree)
        self.root.after(100, self.auto_resize_columns, self.planning_remaining_tree)

    def setup_table_zebra_styles(self):
        """在窗口完全渲染后初始化表格斑马纹样式"""
        try:
            # 更新表格的斑马条纹样式
            if hasattr(self, 'split_tree'):
                self.update_table_zebra_stripes(self.split_tree)
            if hasattr(self, 'remaining_tree'):
                self.update_table_zebra_stripes(self.remaining_tree)
            if hasattr(self, 'allocated_tree'):
                self.update_table_zebra_stripes(self.allocated_tree)
            if hasattr(self, 'planning_remaining_tree'):
                self.update_table_zebra_stripes(self.planning_remaining_tree)
        except (tk.TclError, AttributeError):
            pass

    def configure_treeview_styles(self, tree, include_special_tags=False):
        """配置Treeview控件的基本样式（斑马条纹、错误和信息标签）

        Args:
            tree: 要配置的Treeview对象
            include_special_tags: 是否包含错误和信息标签配置
        """
        try:
            tree.tag_configure("even", background="#ffffff")
            tree.tag_configure("odd", background="#d8d8d8")

            tree.tag_configure("section", background="#d8d8d8", foreground="#000000")

            # 配置当前选中行样式，避免字体类型错误
            try:
                tree.tag_configure("current", font=("TkDefaultFont", 11), foreground="#000000")
            except tk.TclError:
                # 如果自定义字体失败，只设置前景色
                tree.tag_configure("current", foreground="#000000")
            
            # 配置错误行样式
            tree.tag_configure("error_row", foreground="red")
            
            # 配置错误和信息标签
            tree.tag_configure("error", foreground="red")
            tree.tag_configure("info", foreground="blue")
        except (tk.TclError, AttributeError) as e:
            # 只捕获与Tkinter配置相关的错误，避免隐藏其他重要错误
            print(f"配置Treeview样式时出错: {str(e)}")

    def update_table_zebra_stripes(self, tree, update_index=False):
        """更新表格的斑马条纹标签

        Args:
            tree: 要处理的Treeview对象
            update_index: 是否更新序号列（适用于包含序号的表格）
        """
        try:
            # 递归处理所有可见的子节点
            def process_children(parent, start_index=1):
                index = start_index
                children = tree.get_children(parent)
                for item in children:
                    tag = "even" if index % 2 == 0 else "odd"

                    # 获取当前行的标签，保留原有标签
                    current_tags = list(tree.item(item, "tags"))
                    
                    # 如果没有当前标签，创建空列表
                    if not current_tags:
                        current_tags = []
                    
                    # 如果当前标签是字符串而不是列表，转换为列表
                    if isinstance(current_tags, str):
                        current_tags = [current_tags]
                    
                    # 保存特殊标签
                    special_tags = [t for t in current_tags if t not in ["even", "odd"]]
                    
                    # 重新构建标签列表，特殊标签在前，斑马条纹标签在后
                    new_tags = special_tags.copy()
                    if tag not in new_tags:
                        new_tags.append(tag)

                    if update_index:
                        # 更新序号列
                        values = list(tree.item(item, "values"))
                        if values and values[0] != index:  # 只有当序号不一致时才更新
                            values[0] = index
                            tree.item(item, values=values, tags=tuple(new_tags))
                        else:
                            # 只更新斑马条纹标签，减少不必要的UI更新
                            tree.item(item, tags=tuple(new_tags))
                    else:
                        # 只更新斑马条纹标签
                        tree.item(item, tags=tuple(new_tags))
                    
                    # 只有当节点被展开时，才递归处理其子节点
                    if tree.item(item, 'open'):
                        index = process_children(item, index + 1)
                    else:
                        index += 1
                return index
            
            # 从根节点开始处理
            process_children('')
        except AttributeError:
            # 忽略属性不存在的错误
            pass
        except (tk.TclError, TypeError):
            # 忽略Tcl和类型错误，不影响程序运行
            pass

    def update_all_zebra_stripes(self):
        """批量更新所有表格的斑马条纹

        此方法统一更新所有需要斑马条纹效果的表格,
        避免重复调用 update_table_zebra_stripes 方法。
        """
        trees_to_update = []

        # 收集所有需要更新的表格
        if hasattr(self, 'split_tree'):
            trees_to_update.append(self.split_tree)
        if hasattr(self, 'remaining_tree'):
            trees_to_update.append(self.remaining_tree)
        if hasattr(self, 'allocated_tree'):
            trees_to_update.append(self.allocated_tree)
        if hasattr(self, 'planning_remaining_tree'):
            trees_to_update.append(self.planning_remaining_tree)
        if hasattr(self, 'ipam_network_tree'):
            trees_to_update.append(self.ipam_network_tree)
        if hasattr(self, 'ipam_ip_tree'):
            trees_to_update.append(self.ipam_ip_tree)

        # 批量更新所有表格
        for tree in trees_to_update:
            self.update_table_zebra_stripes(tree)

    def update_requirements_tree_zebra_stripes(self):
        """更新子网需求表的斑马条纹"""
        if hasattr(self, 'requirements_tree'):
            self.update_table_zebra_stripes(self.requirements_tree, update_index=True)

    def update_pool_tree_zebra_stripes(self):
        """更新需求池的斑马条纹"""
        if hasattr(self, 'pool_tree'):
            self.update_table_zebra_stripes(self.pool_tree, update_index=True)

    def _persist_requirements_data(self):
        """将当前子网需求表和需求池的数据持久化到数据库"""
        try:
            requirements = []
            for item in self.requirements_tree.get_children():
                values = self.requirements_tree.item(item, "values")
                requirements.append((values[1], int(values[2])))
            self.history_repo.save_requirements_data(HistorySQLite.TABLE_REQUIREMENTS, requirements)

            pool = []
            for item in self.pool_tree.get_children():
                values = self.pool_tree.item(item, "values")
                pool.append((values[1], int(values[2])))
            self.history_repo.save_requirements_data(HistorySQLite.TABLE_POOL, pool)
        except Exception:
            pass

    def _persist_text_data(self, category, text_widget):
        """将文本控件的内容持久化到数据库

        Args:
            category: 数据类别标识
            text_widget: tk.Text 控件
        """
        try:
            content = text_widget.get(1.0, tk.END).strip()
            self.history_repo.save_text_data(category, content)
        except Exception:
            pass

    def update_planning_tables_zebra_stripes(self):
        """批量更新子网需求表和需求池的斑马条纹

        此方法统一更新规划模块的两个表格,减少重复代码。
        """
        self.update_requirements_tree_zebra_stripes()
        self.update_pool_tree_zebra_stripes()

    def auto_resize_columns(self, tree):
        """自动调整表格列宽以适应内容

        Args:
            tree: 要调整列宽的Treeview对象
        """
        
        # 检查 _temp_label 是否存在，如果不存在则创建
        if not hasattr(self, '_temp_label'):
            self._temp_label = tk.Label(self.root)
            self._temp_label.pack_forget()

        # 为每列设置一个合理的默认最小宽度（基于列类型）
        default_min_widths = {
            'index': 40,
            'name': 80,  # 减小默认宽度，让name列能自适应
            'cidr': 80,  # 增加CIDR列的默认宽度，确保能显示完整的CIDR和中文标题
            'required': 70,
            'released': 70,
            'network': 100,  # 增加网络地址列的默认宽度，确保能显示完整的中文标题
            'netmask': 100,  # 增加子网掩码列的默认宽度，确保能显示完整的中文标题
            'broadcast': 120,  # 适中的广播地址列宽度，确保能显示完整的中文标题
            'wildcard': 80,  # 减小通配符掩码列的默认宽度
            'usable': 60,  # 增加可用地址数列的默认宽度，确保能显示完整的中文标题
            'size': 60,  # 增加主机数列的默认宽度，确保能显示完整的中文标题
        }

        # 调整列宽以适应表头
        for col in tree['columns']:
            # 跳过已经隐藏的列
            if self.is_column_hidden(tree, col):
                continue
                
            # 获取表头文本
            header = tree.heading(col, 'text') or ''  # 确保header不是None

            # 跳过序号列，保持固定宽度
            if col == 'index':
                continue
            
            # 获取当前列宽
            current_width = tree.column(col, 'width')
            
            # 设置临时标签文本并测量宽度
            self._temp_label.config(text=header)
            header_width = self._temp_label.winfo_reqwidth() + 20  # 增加一些边距

            # 获取列中内容的最大宽度
            max_width = header_width
            for item in tree.get_children():
                value = tree.item(item, 'values')
                if value and len(value) > list(tree['columns']).index(col):
                    cell_value = str(value[list(tree['columns']).index(col)])
                    self._temp_label.config(text=cell_value)
                    cell_width = self._temp_label.winfo_reqwidth() + 20  # 增加一些边距
                    # 确保cell_width和max_width都是有效的数值
                    max_width = max(max_width, cell_width)

            # 应用默认最小宽度，确保列宽合理
            if col in default_min_widths:
                # 使用更合理的最小宽度，确保能显示完整内容
                min_width = default_min_widths[col]
                # 总是应用最小宽度，确保表格列宽一致
                if max_width < min_width:
                    max_width = min_width
            
            # 当表格没有数据时，确保使用默认最小宽度
            if not tree.get_children() and col in default_min_widths:
                max_width = default_min_widths[col]
            
            # 无论当前宽度如何，都根据内容调整列宽，允许缩小
            tree.column(col, width=max_width, stretch=True)

    def resize_tables(self):
        """调整表格列宽以适应容器大小并更新空行数"""
        try:
            # 动态更新所有表格的空行数
            tree_names = ['split_tree', 'remaining_tree', 'allocated_tree', 'planning_remaining_tree']
            for tree_name in tree_names:
                if hasattr(self, tree_name):
                    self.update_table_zebra_stripes(getattr(self, tree_name))

            # 仅调整规划结果区域的表格列宽，不影响子网需求区域
            if hasattr(self, 'planning_notebook') and hasattr(self.planning_notebook, 'content_area'):
                # 调整已分配子网表格，根据内容自动调整列宽
                if hasattr(self, 'allocated_tree'):
                    self.auto_resize_columns(self.allocated_tree)

                # 调整剩余网段表格，根据内容自动调整列宽
                if hasattr(self, 'planning_remaining_tree'):
                    self.auto_resize_columns(self.planning_remaining_tree)
        except AttributeError:
            pass
        except ValueError:
            # 忽略值错误
            pass
        except (tk.TclError, TypeError):
            # 忽略Tcl和类型错误
            pass

    def add_subnet_requirement(self):
        """添加子网需求"""
        # 创建临时窗口
        # 创建添加子网需求对话框
        dialog = ComplexDialog(self.root, _('add_subnet_requirement'), 400, 200)

        # 配置内容框架的网格布局
        dialog.content_frame.columnconfigure(0, weight=1)
        dialog.content_frame.columnconfigure(1, weight=0)
        dialog.content_frame.columnconfigure(2, weight=0)
        dialog.content_frame.columnconfigure(3, weight=1)

        # 子网名称 - 标签在中间列左侧，输入框在中间列右侧
        name_var = tk.StringVar()
        label, name_entry = dialog.add_field(_("subnet_name"), 1, 1, width=20)
        name_entry.config(textvariable=name_var)
        # 自动获得焦点，方便直接输入
        name_entry.focus_set()

        # 为子网名称添加验证
        def validate_name(text):
            is_valid = bool(text.strip())
            if name_entry.winfo_exists():
                name_entry.config(foreground='black' if is_valid else 'red')
            return "1"  # 始终允许输入，只做视觉提示
        name_entry.config(validate="all", validatecommand=(dialog.dialog.register(validate_name), "%P"))

        # 主机数量 - 标签在中间列左侧，输入框在中间列右侧
        hosts_var = tk.StringVar()
        label, hosts_entry = dialog.add_field(_("host_count"), 2, 1, width=20)
        hosts_entry.config(textvariable=hosts_var)

        # 为主机数量添加验证
        def validate_hosts(text):
            # 允许空输入，只验证非空时是否为正整数
            if not text:
                if hosts_entry.winfo_exists():
                    hosts_entry.config(foreground='black')
                return "1"
            is_valid = text.isdigit() and int(text) > 0
            if hosts_entry.winfo_exists():
                hosts_entry.config(foreground='black' if is_valid else 'red')
            return "1"  # 始终允许输入，只做视觉提示
        hosts_entry.config(validate="all", validatecommand=(dialog.dialog.register(validate_hosts), "%P"))

        # 定义回车键事件处理函数
        def on_return_key(event):
            save_requirement()

        # 只在窗口创建时绑定一次回车键事件
        dialog.dialog.bind("<Return>", on_return_key)

        def save_requirement(target_table="requirements"):
            """保存子网需求

            Args:
                target_table: 目标表，"requirements"表示子网需求表，"pool"表示需求池表
            """
            name = name_var.get().strip()
            hosts = hosts_var.get().strip()

            if not name:
                self.show_error(_("error"), _("please_enter_subnet_name"))
                # 重新将焦点设置到子网名称输入框
                name_entry.focus_set()
                return

            if not hosts.isdigit() or int(hosts) <= 0:
                self.show_error(_("error"), _("please_enter_valid_host_count"))
                # 重新将焦点设置到主机数量输入框
                hosts_entry.focus_set()
                return

            # 检查是否存在相同名称的子网，同时检查子网需求表和需求池表
            for item in self.requirements_tree.get_children():
                values = self.requirements_tree.item(item, "values")
                existing_name = values[1]  # 子网名称在第二列
                if existing_name == name:
                    self.show_error(_("error"), _("subnet_already_exists", name=name))
                    name_entry.focus_set()
                    return

            for item in self.pool_tree.get_children():
                values = self.pool_tree.item(item, "values")
                existing_name = values[1]  # 子网名称在第二列
                if existing_name == name:
                    self.show_error(_("error"), f"{name}{_("msg_already_exists", name=name)}")
                    return

            if target_table == "requirements":
                # 添加到子网需求表 - 带斑马条纹标签
                # 获取当前表格中的行数，计算新行的索引（从1开始）
                current_rows = len(self.requirements_tree.get_children())
                new_index = current_rows + 1
                tag = "even" if new_index % 2 == 0 else "odd"
                self.requirements_tree.insert("", tk.END, values=(new_index, name, hosts), tags=(tag,))

                # 重新应用所有行的斑马条纹，确保一致性
                self.update_requirements_tree_zebra_stripes()
            else:
                # 添加到需求池表 - 带斑马条纹标签
                current_rows = len(self.pool_tree.get_children())
                new_index = current_rows + 1
                tag = "even" if new_index % 2 == 0 else "odd"
                self.pool_tree.insert("", tk.END, values=(new_index, name, hosts), tags=(tag,))

                self.update_pool_tree_zebra_stripes()

            # 保存当前状态到操作记录，包含添加的子网信息
            self.save_current_state(f"添加子网: {name}({hosts})")

            dialog.destroy()

        # 创建按钮
        dialog.add_button(_('save_requirement'), lambda: save_requirement("requirements"), column=1)
        dialog.add_button(_('save_to_pool'), lambda: save_requirement("pool"), column=2)
        
        # 显示对话框
        dialog.show()

    def center_window(self, window, width, height, parent=None):
        """将窗口居中显示在指定父窗口中

        Args:
            window: 要居中的窗口对象
            width: 窗口宽度
            height: 窗口高度
            parent: 父窗口对象，如果为None则使用主窗口
        """
        # 确定父窗口
        parent_window = parent if parent else self.root
        
        # 获取父窗口的位置和尺寸
        parent_x = parent_window.winfo_x()
        parent_y = parent_window.winfo_y()
        parent_width = parent_window.winfo_width()
        parent_height = parent_window.winfo_height()

        # 计算对话框的居中位置（不包含父窗口标题栏）
        title_bar_height = 30  # 通常标题栏高度约为30像素
        dialog_x = parent_x + (parent_width - width) // 2
        dialog_y = parent_y + title_bar_height + (parent_height - title_bar_height - height) // 2

        # 一次性设置对话框的尺寸和位置
        window.geometry(f"{width}x{height}+{dialog_x}+{dialog_y}")

    def create_dialog(self, title, width, height, resizable=False, modal=True, parent=None):
        """创建统一居中的对话框

        Args:
            title: 对话框标题
            width: 对话框宽度
            height: 对话框高度
            resizable: 是否允许调整大小
            modal: 是否为模态对话框
            parent: 父窗口对象，如果为None则使用主窗口

        Returns:
            tk.Toplevel: 居中的对话框对象
        """
        # 确定父窗口
        parent_window = parent if parent else self.root
        
        # 创建对话框
        dialog = tk.Toplevel(parent_window)
        dialog.title(title)
        dialog.resizable(resizable, resizable)
        dialog.transient(parent_window)
        
        # 设置最小尺寸限制（仅当可调整大小时）
        if resizable:
            dialog.minsize(width, height)
        
        # 为对话框设置与主窗口相同的置顶属性，确保在主窗口置顶时对话框也能保持在前面
        dialog.attributes('-topmost', self.is_pinned)
        
        if modal:
            dialog.grab_set()
        
        # 先隐藏对话框，避免定位过程中的闪现
        dialog.withdraw()
        
        # 居中对话框
        self.center_window(dialog, width, height, parent_window)
        
        # 显示对话框并设置焦点
        dialog.deiconify()
        dialog.focus_force()
        
        return dialog

    def delete_subnet_requirement(self):
        """删除选中的子网需求或需求池记录，并重新应用斑马条纹"""
        # 检查两个表格中是否有选中的记录
        selected_requirements = self.requirements_tree.selection()
        selected_pool_items = self.pool_tree.selection()

        if not selected_requirements and not selected_pool_items:
            self.show_warning(_("hint"), _("please_select_record_to_delete"))
            return

        # 显示自定义的居中确认对话框
        confirm = self.show_custom_confirm(_("confirm_delete"), _("delete_confirmation_message"))
        if not confirm:
            return

        # 收集要删除的子网信息和详细记录
        deleted_subnets = []
        deleted_records = []

        # 删除子网需求表中的选中记录
        for item in selected_requirements:
            values = self.requirements_tree.item(item, "values")
            deleted_subnets.append(f"{values[1]}({values[2]})")
            # 保存详细记录，包括表格类型和记录数据
            deleted_records.append({"tree": "requirements", "values": tuple(values), "item": item})
            self.requirements_tree.delete(item)

        # 删除需求池表中的选中记录
        for item in selected_pool_items:
            values = self.pool_tree.item(item, "values")
            deleted_subnets.append(f"{values[1]}({values[2]})")
            deleted_records.append({"tree": "pool", "values": tuple(values), "item": item})
            self.pool_tree.delete(item)

        # 保存删除记录到历史列表，支持多次撤销
        if deleted_records:
            self.deleted_history.append(deleted_records)

        # 删除后重新应用斑马条纹
        self.update_requirements_tree_zebra_stripes()
        self.update_pool_tree_zebra_stripes()

        # 保存当前状态到操作记录，包含删除的子网信息
        if deleted_subnets:
            action_type = f"删除子网: {', '.join(deleted_subnets)}"
        else:
            action_type = "删除子网"
        self.save_current_state(action_type)

    def import_requirements(self):
        """导入子网需求数据"""
        self._import_data()

    def _import_data(self):
        """导入数据的主方法"""
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 显示导入选项对话框
        # 创建导入数据对话框
        dialog = ComplexDialog(self.root, _("import_data"), 400, 200)

        # 按钮框架 - 纵向排列，居中放置
        button_frame = ttk.Frame(dialog.content_frame)
        button_frame.pack(fill=tk.X, pady=10)

        # 导入文件按钮 - 宽度等于下面两个按钮的总和
        import_file_btn = ttk.Button(button_frame, text=_("import_from_file"),
                                    command=lambda: self._import_from_file(dialog))
        import_file_btn.pack(pady=10, fill=tk.X, padx=20)

        # 将焦点聚焦到第一个按钮上
        import_file_btn.focus_force()

        # 下载模板按钮框架 - 横向排列
        template_frame = ttk.Frame(button_frame)
        template_frame.pack(pady=10, fill=tk.X, padx=20)

        # 下载Excel模板按钮 - 左对齐并撑满，添加右边距10
        download_excel_btn = ttk.Button(
            template_frame,
            text=_("download_excel_template"),
            command=lambda: self._generate_template("excel")
        )
        download_excel_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        # 下载CSV模板按钮 - 右对齐并撑满，添加左边距10
        download_csv_btn = ttk.Button(
            template_frame,
            text=_("download_csv_template"),
            command=lambda: self._generate_template("csv")
        )
        download_csv_btn.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))

        # 取消按钮
        dialog.add_button(_("cancel"), dialog.destroy, column=1)
        
        # 显示对话框
        dialog.show()

    def _import_from_file(self, parent_dialog):
        """从文件导入数据

        Args:
            parent_dialog: 父对话框
        """
        parent_dialog.destroy()

        # 选择文件
        file_path = filedialog.askopenfilename(
            title=_("select_file_to_import"),
            filetypes=[
                ("Excel文件", "*.xlsx"),
                ("CSV文件", "*.csv"),
            ],
            initialdir=""
        )

        if not file_path:
            return

        # 解析文件
        try:
            data_list = self._parse_import_file(file_path)
        except (ValueError, IOError, UnicodeDecodeError, TypeError) as e:
            self.show_error(_("error"), f"{_("msg_file_parse_failed")}: {str(e)}")
            return

        if not data_list:
            self.show_info(_("hint"), _("no_valid_data_found"))
            return

        # 验证数据（包含重复性检查）
        errors = self._validate_import_data(data_list)

        # 显示验证结果对话框
        self._show_import_result(errors, data_list)

    def _parse_import_file(self, file_path):
        """解析导入文件

        Args:
            file_path: 文件路径

        Returns:
            list: 解析后的数据列表，每个元素是包含"name"和"hosts"的字典
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        data_list = []

        if file_ext == ".xlsx":
            # Excel文件解析
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            # 跳过表头，从第二行开始读取
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and len(row) >= 2:
                    name = str(row[0]).strip() if row[0] else ""
                    hosts = str(row[1]).strip() if row[1] else ""
                    if name and hosts:
                        data_list.append({"name": name, "hosts": hosts, "row": row_idx})

            wb.close()

        elif file_ext == ".csv":
            # CSV文件解析，尝试多种编码
            encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']
            csv_data = None

            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        csv_data = list(csv.reader(f))
                    break
                except UnicodeDecodeError:
                    continue

            if csv_data is None:
                raise ValueError("无法识别文件编码，请确保文件使用UTF-8或GBK编码")

            for row_idx, row in enumerate(csv_data[1:], start=2):
                if row and len(row) >= 2:
                    name = str(row[0]).strip() if row[0] else ""
                    hosts = str(row[1]).strip() if row[1] else ""
                    if name and hosts:
                        data_list.append({"name": name, "hosts": hosts, "row": row_idx})

        else:
            raise ValueError("不支持的文件格式，请使用Excel (.xlsx) 或CSV (.csv) 文件")

        return data_list

    def _validate_import_data(self, data_list):
        """验证导入数据

        Args:
            data_list: 数据列表

        Returns:
            list: 错误列表，每个元素是包含"row"、"name"、"hosts"、"error"的字典
        """
        errors = []

        # 获取现有数据名称（同时检查子网需求表和需求池表）
        existing_names = set()

        for item in self.requirements_tree.get_children():
            values = self.requirements_tree.item(item, "values")
            existing_names.add(values[1])

        for item in self.pool_tree.get_children():
            values = self.pool_tree.item(item, "values")
            existing_names.add(values[1])

        # 验证每条数据
        for idx, data in enumerate(data_list):
            row = data.get("row", idx + 1)
            name = data.get("name", "")
            hosts = data.get("hosts", "")

            # 检查必填字段
            if not name:
                errors.append({"row": row, "name": name, "hosts": hosts,
                              "error": "子网名称不能为空"})
                continue

            if not hosts:
                errors.append({"row": row, "name": name, "hosts": hosts,
                              "error": "主机数量不能为空"})
                continue

            # 检查主机数量是否为正整数
            if not hosts.isdigit():
                errors.append({"row": row, "name": name, "hosts": hosts,
                              "error": "主机数量必须是正整数"})
                continue

            if int(hosts) <= 0:
                errors.append({"row": row, "name": name, "hosts": hosts,
                              "error": "主机数量必须大于0"})
                continue

            # 检查同批次重复
            if any(d["name"] == name for d in data_list[:idx]):
                errors.append({"row": row, "name": name, "hosts": hosts,
                              "error": "文件中存在重复的子网名称"})
                continue

            # 检查重复性（同时检查两张表）
            if name in existing_names:
                errors.append({"row": row, "name": name, "hosts": hosts,
                              "error": _("subnet_already_exists", name=name)})
                continue

            # 添加到已存在名称集合中，避免同批次重复
            existing_names.add(name)

        return errors

    def _show_import_result(self, errors, data_list):
        """显示导入验证结果对话框（显示所有数据，包括有效和无效）

        Args:
            errors: 错误列表
            data_list: 数据列表
        """
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 创建导入结果对话框
        dialog = ComplexDialog(self.root, _("import_data"), 800, 600, resizable=True)

        dialog.dialog.focus_force()

        main_frame = ttk.Frame(dialog.content_frame, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 统计信息
        error_count = len(errors)
        total_count = len(data_list)
        valid_count = total_count - error_count

        # 创建统计信息框架，左对齐显示
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        summary_text = _("data_import_summary").format(total_count=total_count, valid_count=valid_count, error_count=error_count)
        ttk.Label(stats_frame, text=summary_text, font=(font_family, font_size)).pack(anchor=tk.W)

        # 创建表格显示所有数据
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        result_tree = ttk.Treeview(tree_frame, columns=("row", "name", "hosts", "status"),
                                  show="headings", height=12)
        result_tree.bind("<Button-3>", lambda event, t=result_tree: self.copy_cell_data(event, t))
        result_tree.heading("row", text=_("row_number"))
        result_tree.heading("name", text=_("subnet_name"))
        result_tree.heading("hosts", text=_("host_count"))
        result_tree.heading("status", text=_("status"))

        result_tree.column("row", width=40, minwidth=20, anchor="e")
        result_tree.column("name", width=100, minwidth=80)
        result_tree.column("hosts", width=60, minwidth=20)
        result_tree.column("status", width=400, minwidth=80)

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)

        # 创建滚动条回调函数，实现自动隐藏和外边距调整
        def scrollbar_callback(*args):
            # 设置滚动条位置
            scrollbar.set(*args)

            yview = result_tree.yview()
            need_scrollbar = not (float(yview[0]) <= 0.0 and float(yview[1]) >= 1.0)

            # 根据是否需要滚动条调整Treeview的右边距
            if need_scrollbar:
                # 显示滚动条
                scrollbar.grid(row=0, column=1, sticky=tk.NS, padx=0)
                # 调整Treeview的grid配置，移除内边距，与主框架内边距保持一致
                result_tree.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=0)

            else:
                # 隐藏滚动条
                scrollbar.grid_remove()
                # 调整Treeview的grid配置，移除内边距，与主框架内边距保持一致
                result_tree.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=0)

        # 绑定滚动条和Treeview
        scrollbar.config(command=result_tree.yview)
        result_tree.config(yscrollcommand=scrollbar_callback)

        # 使用grid布局，确保Treeview和滚动条正确对齐
        result_tree.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar.grid(row=0, column=1, sticky=tk.NS, padx=0)

        # 配置grid权重，使Treeview可以扩展
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 初始调用一次回调函数，设置初始状态
        scrollbar_callback(0.0, 1.0)

        # 将错误列表转换为字典，提高查找效率
        error_dict = {e["row"]: e for e in errors}

        # 填充所有数据
        for data in data_list:
            row = data.get("row", 0)
            name = data.get("name", "")
            hosts = data.get("hosts", "")

            # 从字典中查找对应的错误信息（O(1)复杂度）
            error = error_dict.get(row)
            if error:
                status = error["error"]
                # 无效数据用红色标签
                tag = "invalid"
            else:
                status = _("valid")
                # 有效数据用绿色标签
                tag = "valid"

            result_tree.insert("", tk.END, values=(row, name, hosts, status), tags=(tag,))

        # 配置标签样式
        result_tree.tag_configure("valid", foreground="green")
        result_tree.tag_configure("invalid", foreground="red")

        # 使用主配置方法配置斑马纹样式
        self.configure_treeview_styles(result_tree)

        for index, item in enumerate(result_tree.get_children()):
            # 获取当前标签
            current_tags = result_tree.item(item, "tags")
            # 添加斑马纹标签
            stripe_tag = "even" if index % 2 == 0 else "odd"
            # 合并标签，保持原有状态标签和新的斑马纹标签
            result_tree.item(item, tags=(*current_tags, stripe_tag))

        # 预先计算有效数据列表，避免在按钮点击时重复计算
        valid_data = [d for d in data_list if d.get("row", 0) not in error_dict]

        # 使用dialog.button_frame放置按钮
        # 左侧按钮
        ttk.Button(dialog.button_frame, text=_('import_requirements_pool'),
                  command=lambda: self._import_valid_data(valid_data, "pool", dialog),
                  width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(dialog.button_frame, text=_('import_subnet_requirements'),
                  command=lambda: self._import_valid_data(valid_data, "requirements", dialog),
                  width=15).pack(side=tk.LEFT, padx=5)
        # 右侧取消按钮
        ttk.Button(dialog.button_frame, text=_('cancel'),
                  command=dialog.destroy,
                  width=10).pack(side=tk.RIGHT, padx=5)
        
        # 显示对话框
        dialog.show()

    def _import_valid_data(self, valid_data, target_table, dialog=None):
        """导入有效数据

        Args:
            valid_data: 有效数据列表
            target_table: 目标表
            dialog: 对话框（可选）
        """
        if not valid_data:
            self.show_info(_("hint"), _("no_data_to_import"))
            if dialog:
                dialog.destroy()
            return

        # 导入数据
        for data in valid_data:
            name = data["name"]
            hosts = data["hosts"]

            if target_table == "requirements":
                # 添加到子网需求表
                current_rows = len(self.requirements_tree.get_children())
                new_index = current_rows + 1
                tag = "even" if new_index % 2 == 0 else "odd"
                self.requirements_tree.insert("", tk.END, values=(new_index, name, hosts), tags=(tag,))
            else:
                # 添加到需求池表
                current_rows = len(self.pool_tree.get_children())
                new_index = current_rows + 1
                tag = "even" if new_index % 2 == 0 else "odd"
                self.pool_tree.insert("", tk.END, values=(new_index, name, hosts), tags=(tag,))

        # 更新斑马条纹
        if target_table == "requirements":
            self.update_requirements_tree_zebra_stripes()
        else:
            self.update_pool_tree_zebra_stripes()

        # 保存状态
        self.save_current_state(f"导入数据: {len(valid_data)} 条记录")

        # 显示成功消息
        target_name = "子网需求表" if target_table == "requirements" else "需求池表"
        self.show_info(_("success"), f"{_("successfully_imported")} {len(valid_data)} {_("records_to")}{target_name}")

        # 关闭对话框
        if dialog:
            dialog.destroy()

    def _generate_template(self, template_type):
        """生成模板文件

        Args:
            template_type: 模板类型，"excel"或"csv"
        """
        # 选择保存位置
        if template_type == "excel":
            default_ext = ".xlsx"
            filetypes = [("Excel文件", "*.xlsx")]
        else:
            default_ext = ".csv"
            filetypes = [("CSV文件", "*.csv")]

        file_path = filedialog.asksaveasfilename(
            title=_("save_template"),
            defaultextension=default_ext,
            filetypes=filetypes,
            initialfile=f"{_("subnet_requirement_import_template")}{default_ext}",
            initialdir=""
        )

        if not file_path:
            return

        try:
            if template_type == "excel":
                # 生成Excel模板
                wb = Workbook()
                ws = wb.active
                ws.title = _("subnet_requirements")

                # 设置表头
                headers = [_("subnet_name"), _("host_count")]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # 添加示例数据（包含IPv4和IPv6示例）
                example_data = [
                    [_("office"), "20"],
                    [_("hr_department"), "10"],
                    [_("finance_department"), "10"],
                    [_("planning_department"), "30"],
                    [_("it_department"), "20"],
                    [_("ipv6_lab"), "50"],
                    [_("ipv6_infrastructure"), "100"],
                ]
                for row_idx, row_data in enumerate(example_data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                wb.save(file_path)

            else:
                # 生成CSV模板
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    # 写入表头
                    writer.writerow([_("subnet_name"), _("host_count")])
                    # 写入示例数据
                    writer.writerow([_("office"), "20"])
                    writer.writerow([_("hr_department"), "10"])
                    writer.writerow([_("finance_department"), "10"])
                    writer.writerow([_("planning_department"), "30"])
                    writer.writerow([_("it_department"), "20"])

            self.show_info(_("success"), _("template_saved_to").format(file_path=file_path))

        except (IOError, ValueError, TypeError) as e:
            self.show_error(_("error"), f"{_("msg_template_generation_failed")}: {str(e)}")

    def copy_cell_data(self, event, tree):
        """复制表格中单元格数据的通用功能"""
        # 获取点击位置的行和列
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        # 获取点击的行ID和列
        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)

        if not item:
            return

        # 获取列索引
        column_index = int(column.replace("#", "")) - 1

        # 获取行数据
        values = tree.item(item, "values")
        if not values or column_index >= len(values):
            return

        # 获取单元格数据
        cell_data = str(values[column_index])

        # 特殊处理：如果是隐藏信息表且复制的是密码列（索引2），获取真实密码
        # 检查 _hidden_info_tree 是否存在，防止在未初始化时访问
        if hasattr(self, '_hidden_info_tree') and tree == self._hidden_info_tree and column_index == 2:
            record_id = int(item)
            for record in getattr(self, '_hidden_info_raw_data', []):
                if record.get('id') == record_id:
                    cell_data = record.get('password', '')
                    break

        # 将数据复制到剪贴板
        self.root.clipboard_clear()
        self.root.clipboard_append(cell_data)

        # 可选：显示复制成功的提示
        self.show_result(_("copied_to_clipboard"), keep_data=True)

    def bind_treeview_right_click(self, tree):
        """为Treeview绑定右键复制功能"""
        # 绑定右键菜单事件
        tree.bind("<Button-3>", lambda event, t=tree: self.copy_cell_data(event, t))

    def bind_listbox_right_click(self, listbox):
        """为Listbox绑定右键复制功能"""
        listbox.bind("<Button-3>", lambda event, lb=listbox: self.copy_listbox_data(event, lb))

    def copy_listbox_data(self, event, listbox):
        """复制Listbox选中项到剪贴板"""
        # 禁用事件，防止列表框捕获事件
        listbox.selection_clear(0, tk.END)
        # 获取鼠标点击位置的索引
        index = listbox.nearest(event.y)
        if index >= 0 and index < listbox.size():
            # 选中该项
            listbox.selection_set(index)
            # 获取数据
            cell_data = listbox.get(index)
            # 复制到剪贴板
            self.root.clipboard_clear()
            self.root.clipboard_append(cell_data)
            self.show_result(_("copied_to_clipboard"), keep_data=True)

    def _center_dialog(self, dialog):
        """将对话框居中显示在主窗口中"""
        dialog.update_idletasks()
        dialog_width = dialog.winfo_width()
        dialog_height = dialog.winfo_height()

        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()

        dialog_x = root_x + (root_width - dialog_width) // 2
        dialog_y = root_y + (root_height - dialog_height) // 2

        dialog.geometry(f"+{dialog_x}+{dialog_y}")
        dialog.deiconify()

    def _set_dialog_focus(self, dialog, ok_btn):
        """设置对话框焦点"""
        def set_focus():
            dialog.lift()
            dialog.focus_force()
            if ok_btn:
                ok_btn.focus_force()
        dialog.after_idle(set_focus)

    def show_custom_dialog(self, title, message, dialog_type="info"):
        """显示自定义的居中对话框，支持info、error、warning类型"""
        if dialog_type in ["info", "error", "warning"]:
            # 使用InfoDialog类
            dialog = InfoDialog(self.root, title, message, dialog_type)
            return dialog.show()
        elif dialog_type == "confirm":
            # 使用ConfirmDialog类
            dialog = ConfirmDialog(self.root, title, message, _("yes"), _("no"))
            return dialog.show()
        return None

    def show_info(self, title, message):
        """显示信息（使用信息栏）"""
        return self.show_result(message, error=False)

    def show_error(self, title, message):
        """显示错误（使用信息栏）"""
        return self.show_result(message, error=True)

    def show_warning(self, title, message):
        """显示警告（使用信息栏）"""
        # 警告也使用错误样式的信息栏
        return self.show_result(message, error=True)

    def show_info_dialog(self, title, message):
        """显示信息对话框"""
        dialog = InfoDialog(self.root, title, message, "info")
        return dialog.show()

    def show_warning_dialog(self, title, message):
        """显示警告对话框"""
        dialog = InfoDialog(self.root, title, message, "warning")
        return dialog.show()

    def show_error_dialog(self, title, message):
        """显示错误对话框"""
        dialog = InfoDialog(self.root, title, message, "error")
        return dialog.show()
    
    def show_yes_no_dialog(self, title, message):
        """显示是/否确认对话框"""
        return self.show_custom_dialog(title, message, "confirm")

    def _create_modal_dialog(self, title, width=500, height=150, parent_window=None, resizable=False):
        """创建模态对话框并居中显示

        Args:
            title: 对话框标题
            width: 对话框宽度（可选）
            height: 对话框高度（可选）
            parent_window: 父窗口（可选，默认为当前焦点窗口或主窗口）
            resizable: 是否允许调整大小（可选，默认为False）

        Returns:
            tk.Toplevel: 创建的对话框对象
        """
        # 对于主窗口的对话框，使用统一的创建方法
        if not parent_window or parent_window == self.root:
            return self.create_dialog(title, width, height, resizable)
        
        # 对于其他父窗口的对话框，保持原有逻辑
        self.root.update_idletasks()

        dialog = tk.Toplevel(parent_window)
        dialog.title(title)
        dialog.resizable(resizable, resizable)
        dialog.transient(parent_window)
        dialog.grab_set()

        # 先隐藏对话框，避免定位过程中的闪现
        dialog.withdraw()

        # 设置对话框最小尺寸
        dialog.minsize(width=width, height=height)

        # 计算居中位置
        parent_width = parent_window.winfo_width()
        parent_height = parent_window.winfo_height()
        parent_x = parent_window.winfo_rootx()
        parent_y = parent_window.winfo_rooty()

        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2

        dialog.geometry(f"{width}x{height}+{x}+{y}")
        dialog.deiconify()

        return dialog

    def show_custom_confirm(self, title, message):
        """显示自定义的居中确认对话框"""
        # 使用ConfirmDialog类
        dialog = ConfirmDialog(self.root, title, message, _("ok"), _("cancel"))
        return dialog.show()

    def undo(self):
        """撤销最近的操作，支持多次撤销，包括删除、移动、导入等操作"""
        # 检查是否有可撤销的操作
        if self.current_history_index <= 0:
            self.show_info(_("hint"), _("no_undoable_operation"))
            return

        # 回到上一个状态
        self.current_history_index -= 1
        # 获取上一个状态
        previous_state = self.history_states[self.current_history_index]

        # 清空当前的子网需求表和需求池表
        for item in self.requirements_tree.get_children():
            self.requirements_tree.delete(item)
        for item in self.pool_tree.get_children():
            self.pool_tree.delete(item)

        # 恢复子网需求表
        for req in previous_state['requirements']:
            name, hosts = req
            self.requirements_tree.insert("", tk.END, values=("", name, hosts))

        # 恢复需求池表
        for pool_item in previous_state['pool']:
            name, hosts = pool_item
            self.pool_tree.insert("", tk.END, values=("", name, hosts))

        # 更新序号和斑马条纹
        self.update_requirements_tree_zebra_stripes()
        self.update_pool_tree_zebra_stripes()

        # 显示成功提示 - 使用信息栏而不是对话框
        # 获取当前状态的 action_type，即被撤销的操作类型
        current_state = self.history_states[self.current_history_index + 1] if self.current_history_index + 1 < len(self.history_states) else {"action_type": "未知操作"}
        success_message = f"{_("successfully_undone")} {current_state['action_type']}"
        self.show_result(success_message, error=False)

    def undo_delete(self):
        """撤销最近的删除操作，支持多次撤销"""
        # 检查是否有删除记录历史
        if not self.deleted_history:
            self.show_info(_("hint"), _("no_undoable_delete_operation"))
            return

        # 从历史记录中取出最近一次删除的记录批次
        deleted_records = self.deleted_history.pop()

        # 恢复被删除的记录
        restored_subnets = []

        for record in deleted_records:
            tree_type = record["tree"]
            values = record["values"]

            # 根据记录类型选择对应的表格
            if tree_type == "requirements":
                # 恢复到子网需求表
                self.requirements_tree.insert("", tk.END, values=values)
            elif tree_type == "pool":
                # 恢复到需求池表
                self.pool_tree.insert("", tk.END, values=values)

            # 收集恢复的子网信息
            restored_subnets.append(f"{values[1]}({values[2]})")

        # 恢复后重新应用斑马条纹
        self.update_requirements_tree_zebra_stripes()
        self.update_pool_tree_zebra_stripes()

        # 保存当前状态到操作记录，包含恢复的子网信息
        if restored_subnets:
            action_type = f"撤销删除: 恢复了 {', '.join(restored_subnets)}"
        else:
            action_type = "撤销删除"
        self.save_current_state(action_type)

        # 显示成功提示 - 使用信息栏而不是对话框
        success_message = f"{_("successfully_restored")} {len(deleted_records)} {_("records")}"
        self.show_result(success_message, error=False)

    def _on_treeview_double_click(self, tree, tree_name, event):
        """通用的双击Treeview单元格编辑处理函数

        Args:
            tree: Treeview组件
            tree_name: 表格名称标识
            event: 事件对象
        """
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)

        if not item or not column:
            return

        column_index = int(column[1:]) - 1
        if column_index == 0:  # 不允许编辑序号列
            return
        column_name = tree["columns"][column_index]

        current_value = tree.item(item, "values")[column_index]

        try:
            cell_x, cell_y, width, height = tree.bbox(item, column)
        except tk.TclError:
            return

        self.edit_entry = ttk.Entry(tree, width=width // 10)
        self.edit_entry.insert(0, current_value)
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.focus()

        def validate_edit(text):
            if column_index == 1:  # 子网名称列
                is_valid = bool(text.strip())
            elif column_index == 2:  # 主机数量列
                is_valid = text.isdigit() and int(text) > 0 if text else True
            else:
                is_valid = True
            if hasattr(self, 'edit_entry') and self.edit_entry is not None:
                self.edit_entry.config(foreground='black' if is_valid else 'red')
            return "1"
        self.edit_entry.config(validate="all", validatecommand=(self.root.register(validate_edit), "%P"))

        self.edit_entry.place(x=cell_x, y=cell_y, width=width, height=height)

        self.current_edit_item = item
        self.current_edit_column = column_name
        self.current_edit_column_index = column_index
        self.current_edit_tree = tree_name

        self.edit_entry.bind("<FocusOut>", self.on_edit_focus_out)
        self.edit_entry.bind("<Return>", self.on_edit_enter)
        self.edit_entry.bind("<Escape>", self.on_edit_escape)

    def on_requirements_tree_double_click(self, event):
        """双击Treeview单元格时触发编辑功能（子网需求表）"""
        self._on_treeview_double_click(self.requirements_tree, "requirements", event)

    def on_pool_tree_double_click(self, event):
        """双击Treeview单元格时触发编辑功能（需求池表）"""
        self._on_treeview_double_click(self.pool_tree, "pool", event)

    def on_edit_focus_out(self, event):
        """编辑框失去焦点时保存数据"""
        if hasattr(self, 'edit_entry') and self.edit_entry:
            # 获取当前编辑框的值
            new_value = self.edit_entry.get().strip()
            
            # 保存原始值，用于验证失败时恢复
            if self.current_edit_tree == "requirements":
                tree = self.requirements_tree
            else:
                tree = self.pool_tree
            original_value = tree.item(self.current_edit_item, "values")[self.current_edit_column_index]
            
            # 简单验证：如果值为空，使用原始值
            if not new_value:
                # 直接销毁编辑框并清理状态，不显示错误
                self.edit_entry.destroy()
                self._cleanup_edit_state()
                return
            
            # 验证新值
            validation_error = self._validate_edit_value(new_value)
            if validation_error:
                # 验证失败，直接销毁编辑框并清理状态，不显示错误
                self.edit_entry.destroy()
                self._cleanup_edit_state()
                return
            
            # 验证通过，调用save_edit保存，并标记是从焦点丢失事件调用
            self.save_edit(from_focus_out=True)

    def on_edit_enter(self, event):
        """按下Enter键时保存数据"""
        if hasattr(self, 'edit_entry') and self.edit_entry:
            self.save_edit()

    def on_edit_escape(self, event):
        """按下Escape键时取消编辑"""
        if hasattr(self, 'edit_entry') and self.edit_entry:
            self.edit_entry.destroy()
        self._cleanup_edit_state()

    def on_treeview_click(self, event):
        """处理Treeview左键单击事件，实现取消选择功能"""
        # 检查是否正在编辑状态
        if hasattr(self, 'inline_edit_data') and self.inline_edit_data:
            # 正在编辑状态，点击任何位置都会退出编辑并选中点击的行
            return
        
        # 检查是否是双击事件的一部分
        # 当发生双击时，单击事件会被触发两次，然后双击事件才会被触发
        # 我们使用定时器来区分单击和双击
        if hasattr(self, '_click_timer'):
            self.root.after_cancel(self._click_timer)
        
        # 检查是否有正在编辑的状态（旧机制）
        if hasattr(self, 'current_edit_item') and self.current_edit_item is not None:
            # 获取当前编辑的表格
            if self.current_edit_tree == "requirements":
                current_tree = self.requirements_tree
            else:
                current_tree = self.pool_tree
            
            # 保存当前编辑
            self.save_edit(from_focus_out=True)
        
        # 检查是否有正在编辑的状态（新机制）
        if hasattr(self, 'inline_edit_widget') and hasattr(self, 'inline_edit_data'):
            # 检查点击是否在DateEntry的日历弹出窗口内
            should_save = True
            try:
                if hasattr(self.inline_edit_widget, '_top_cal'):
                    top_cal = self.inline_edit_widget._top_cal
                    if top_cal:
                        # 不再依赖ismapped，直接检查坐标
                        x = event.x_root
                        y = event.y_root
                        try:
                            tx = top_cal.winfo_rootx()
                            ty = top_cal.winfo_rooty()
                            tw = top_cal.winfo_width()
                            th = top_cal.winfo_height()
                            if tw > 0 and th > 0 and tx <= x <= tx + tw and ty <= y <= ty + th:
                                should_save = False
                        except Exception:
                            pass
            except Exception:
                pass
            
            if should_save:
                self.on_generic_inline_edit_save(None)
        
        # 获取点击位置的信息
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region not in ("cell", "row"):
            return "break"

        # 获取点击的行
        item = tree.identify_row(event.y)
        if not item:
            return "break"

        # 获取当前选中的所有项
        selected_items = list(tree.selection())

        # 获取按键状态
        is_ctrl = event.state & 0x4  # Ctrl键
        is_shift = event.state & 0x1  # Shift键

        # Ctrl+点击：切换选择状态
        if is_ctrl:
            if item in selected_items:
                # 已选中，取消选择
                tree.selection_remove(item)
            else:
                # 未选中，添加到选择
                tree.selection_add(item)
        # Shift+点击：选择范围
        elif is_shift:
            if selected_items:
                all_items = tree.get_children()
                last_selected = selected_items[-1]
                start_idx = all_items.index(last_selected)
                end_idx = all_items.index(item)

                # 先取消所有选中项
                tree.selection_remove(selected_items)

                # 选择范围
                for idx in range(min(start_idx, end_idx), max(start_idx, end_idx) + 1):
                    tree.selection_add(all_items[idx])
            else:
                # 没有选中项，直接选择当前项
                tree.selection_set(item)
        # 普通点击
        else:
            if item in selected_items:
                # 点击已选中项
                if len(selected_items) == 1:
                    # 唯一选中项，使用定时器延迟处理取消选择
                    def handle_single_click_cancel():
                        # 重新获取选择状态，因为可能在延迟期间发生变化
                        current_selected = list(tree.selection())
                        current_item = tree.identify_row(event.y)
                        
                        # 如果点击的是已选中的项，取消选择
                        if current_item and current_item in current_selected and len(current_selected) == 1:
                            tree.selection_remove(current_item)
                    
                    # 设置定时器，延迟30毫秒执行取消选择
                    # 这样在双击时，定时器会被取消，避免与编辑功能冲突
                    self._click_timer = self.root.after(30, handle_single_click_cancel)
                    # 允许默认的选择行为
                    return
                else:
                    # 多个选中项，只选择当前项
                    tree.selection_set(item)
            else:
                # 点击未选中项，使用定时器延迟处理
                def handle_single_click_select():
                    # 重新获取选择状态，因为可能在延迟期间发生变化
                    current_selected = list(tree.selection())
                    current_item = tree.identify_row(event.y)
                    
                    # 如果点击的是未选中的项，只选择当前项
                    if current_item and current_item not in current_selected:
                        tree.selection_set(current_item)
                
                # 设置定时器，延迟30毫秒执行选择
                # 这样在双击时，定时器会被取消，避免与编辑功能冲突
                self._click_timer = self.root.after(30, handle_single_click_select)
                # 允许默认的选择行为
                return

        # 阻止事件继续传递，避免默认行为冲突
        return "break"
        
        # 定义单击处理函数
        def handle_single_click():
            # 这里处理其他单击事件，比如点击空白区域等
            pass
        
        # 设置定时器，延迟30毫秒执行单击处理
        self._click_timer = self.root.after(30, handle_single_click)

    def save_edit(self, from_focus_out=False):
        """保存编辑的数据
        
        Args:
            from_focus_out: 是否从焦点丢失事件调用，True表示不要重新设置焦点
        """
        if not hasattr(self, 'current_edit_item'):
            return

        new_value = self.edit_entry.get().strip()

        # 获取原始值
        if self.current_edit_tree == "requirements":
            tree = self.requirements_tree
            original_value = tree.item(self.current_edit_item, "values")[self.current_edit_column_index]
        else:
            tree = self.pool_tree
            original_value = tree.item(self.current_edit_item, "values")[self.current_edit_column_index]

        # 如果值没有变化，直接更新并清理
        if new_value == original_value:
            self._update_tree_and_cleanup(tree)
            return

        # 验证数据
        if not new_value:
            if not from_focus_out:
                self.show_error(_("error"), _("input_cannot_be_empty"))
                self.edit_entry.focus_set()
            else:
                # 从焦点丢失事件调用，直接销毁编辑框并清理
                self.edit_entry.destroy()
                self._cleanup_edit_state()
            return

        # 验证新值
        validation_error = self._validate_edit_value(new_value)
        if validation_error:
            if not from_focus_out:
                self.show_error(_("error"), validation_error)
                self.edit_entry.focus_set()
            else:
                # 从焦点丢失事件调用，直接销毁编辑框并清理
                self.edit_entry.destroy()
                self._cleanup_edit_state()
            return

        # 更新数据
        values = list(tree.item(self.current_edit_item, "values"))
        values[self.current_edit_column_index] = new_value
        tree.item(self.current_edit_item, values=values)

        # 更新斑马条纹
        if self.current_edit_tree == "requirements":
            self.update_requirements_tree_zebra_stripes()
        else:
            self.update_pool_tree_zebra_stripes()

        # 销毁编辑框并清理
        self.edit_entry.destroy()
        self._cleanup_edit_state()

    def _update_tree_and_cleanup(self, tree):
        """更新Treeview数据并清理编辑状态"""
        values = list(tree.item(self.current_edit_item, "values"))
        tree.item(self.current_edit_item, values=values)
        self.update_table_zebra_stripes(tree)
        self.edit_entry.destroy()
        self._cleanup_edit_state()

    def _validate_edit_value(self, new_value):
        """验证编辑的值

        Args:
            new_value: 新值

        Returns:
            str or None: 错误消息，如果验证通过则返回None
        """
        if self.current_edit_column == "name":
            # 检查是否存在相同名称的子网
            for tree, tree_name in [(self.requirements_tree, "requirements"), (self.pool_tree, "pool")]:
                for item in tree.get_children():
                    # 排除当前编辑的行
                    if tree_name == self.current_edit_tree and item == self.current_edit_item:
                        continue
                    values = tree.item(item, "values")
                    if values[1] == new_value:  # 子网名称在第二列
                        return _("subnet_already_exists", name=new_value)

        elif self.current_edit_column == "hosts":
            try:
                hosts = int(new_value)
                if hosts <= 0:
                    return _("host_count_must_be_greater_than_0")
            except ValueError:
                return _("host_count_must_be_integer")

        return None

    def _cleanup_edit_state(self):
        """清理编辑状态"""
        if hasattr(self, 'current_edit_item'):
            del self.current_edit_item
        if hasattr(self, 'current_edit_column'):
            del self.current_edit_column
        if hasattr(self, 'current_edit_column_index'):
            del self.current_edit_column_index
        if hasattr(self, 'current_edit_tree'):
            del self.current_edit_tree
        if hasattr(self, 'edit_entry'):
            del self.edit_entry
        self.edit_entry = None

    def perform_planning(self, parent=None, subnet_requirements=None, from_history=False, update_history=True, save_state=True, generate_chart=True):
        """执行子网规划操作

        Args:
            parent: 父网段，如果为None则从输入框获取
            subnet_requirements: 子网需求列表，如果为None则从表格获取
            from_history: 是否从历史记录重新执行
            update_history: 是否更新下拉表历史
            save_state: 是否保存当前状态到操作记录
            generate_chart: 是否生成网段分布图

        Returns:
            dict: 规划结果，如果失败则返回None
        """
        # 获取父网段
        if parent is None:
            parent = self.planning_parent_entry.get().strip()

        # 验证输入
        validation_result = self._validate_planning_input(parent)
        if not validation_result['valid']:
            self.show_error(_("error"), validation_result['error'])
            return None
        
        # 重新获取修正后的父网段
        parent = self.planning_parent_entry.get().strip()

        # 获取子网需求
        if subnet_requirements is None:
            subnet_requirements = []
            for item in self.requirements_tree.get_children():
                values = self.requirements_tree.item(item, "values")
                subnet_requirements.append((values[1], int(values[2])))

        if not subnet_requirements:
            self.show_error(_("error"), _("please_add_at_least_one_requirement"))
            return None

        try:
            # 只验证IP地址格式，不自动修正格式，保留用户输入的原始格式
            parent_network = ipaddress.ip_network(parent, strict=False)
            
            # 执行子网规划
            # 转换子网需求格式以匹配函数参数要求
            formatted_requirements = [{'name': name, 'hosts': hosts} for name, hosts in subnet_requirements]

            # 调用子网规划函数
            plan_result = suggest_subnet_planning(parent, formatted_requirements)

            # 检查是否有错误
            if 'error' in plan_result:
                self.show_error(_("error"), f"{_("subnet_planning_failed")}: {plan_result['error']}")
                return None

            # 直接使用第一个规划方案
            selected_plan = plan_result['plans'][0]

            # 清空结果表格
            self.clear_tree_items(self.allocated_tree)
            self.clear_tree_items(self.planning_remaining_tree)

            # 检测IP版本
            is_ipv6 = TableColumnManager.is_ipv6_network(parent)
            
            # 根据IP版本显示或隐藏相应的列
            TableColumnManager.configure_planning_allocated_tree(self.allocated_tree, is_ipv6)
            
            # 显示已分配子网
            for i, subnet in enumerate(selected_plan['allocated_subnets'], 1):
                # 设置斑马条纹标签
                tags = ("even",) if i % 2 == 0 else ("odd",)
                self.allocated_tree.insert(
                    "",
                    tk.END,
                    values=(
                        i,
                        subnet["name"],
                        subnet["cidr"],
                        format_large_number(subnet["required_hosts"]),
                        format_large_number(subnet["available_hosts"]),
                        subnet["info"]["network"],
                        subnet["info"]["broadcast"],  # 网段结束地址
                        subnet["info"]["netmask"],
                        subnet["info"]["wildcard"],
                        subnet["info"]["broadcast"] if not is_ipv6 else "-",
                    ),
                    tags=tags,
                )
            # 斑马条纹样式已在初始化时配置

            # 数据添加完成后，自动调整列宽以适应内容
            self.auto_resize_columns(self.allocated_tree)

            # 根据IP版本显示或隐藏相应的列
            TableColumnManager.configure_planning_remaining_tree(self.planning_remaining_tree, is_ipv6)
            
            # 显示剩余网段
            for i, subnet in enumerate(selected_plan['remaining_subnets_info'], 1):
                tags = ("even",) if i % 2 == 0 else ("odd",)
                hidden_vals = TableColumnManager.get_hidden_values_for_ipv6(subnet, is_ipv6)
                
                self.planning_remaining_tree.insert(
                    "",
                    tk.END,
                    values=(
                        i,
                        selected_plan['remaining_subnets'][i - 1],
                        subnet["network"],
                        subnet["host_range_end"],
                        hidden_vals["netmask"],
                        hidden_vals["wildcard"],
                        hidden_vals["broadcast"],
                        format_large_number(subnet["usable_addresses"]),
                    ),
                    tags=tags,
                )

            self.auto_resize_columns(self.planning_remaining_tree)

            # 子网规划完成，不显示对话框提示

            # 如果需要更新历史记录
            if update_history and not from_history:
                # 使用通用方法更新父网段历史记录
                current_parent = self.planning_parent_entry.get().strip()
                self._update_history_entry(current_parent, self.planning_parent_networks, self.planning_parent_entry)

            # 如果需要保存状态
            if save_state and not from_history:
                # 保存当前状态到操作记录
                self.save_current_state("执行规划")

            # 如果需要生成图表
            if generate_chart:
                # 生成网段分布图数据并绘制
                # 构建兼容的plan_result结构
                compatible_plan_result = {
                    "parent_cidr": plan_result["parent_cidr"],
                    "allocated_subnets": selected_plan["allocated_subnets"],
                    "remaining_subnets": selected_plan["remaining_subnets"],
                    "remaining_subnets_info": selected_plan["remaining_subnets_info"],
                    "ip_version": plan_result["ip_version"]
                }
                self.generate_planning_chart_data(compatible_plan_result)

            return plan_result

        except ValueError as e:
            # 导入handle_ip_subnet_error函数
            from ip_subnet_calculator import handle_ip_subnet_error
            # 使用通用错误处理函数处理IP相关错误
            error_dict = handle_ip_subnet_error(e)
            self.show_error(_("error"), f"{_("subnet_planning_failed")}: {error_dict.get('error', str(e))}")
            return None
        except (tk.TclError, AttributeError, TypeError) as e:
            self.show_error(_("error"), f"{_("subnet_planning_failed")}: {_("unknown_error_occurred")} - {str(e)}")
            return None

    def execute_subnet_planning(self, from_history=False):
        """执行子网规划

        Args:
            from_history: 是否从历史记录重新执行，True表示不将操作记入历史
        """
        self.perform_planning(from_history=from_history)

    def sync_allocated_to_ipam(self):
        """将已分配子网表中的数据同步到地址管理的网段表"""
        # 获取已分配子网表中的所有数据
        subnet_items = self.allocated_tree.get_children()
        
        if not subnet_items:
            self.show_info(_("hint"), _("no_subnet_to_sync"))
            return
        
        # 先获取并添加父网段（如果不存在）
        parent_cidr = self.planning_parent_entry.get().strip()
        if parent_cidr:
            try:
                ipaddress.ip_network(parent_cidr, strict=False)
            except ValueError:
                self.show_error(_("error"), _("invalid_cidr_format"))
                return
            # 添加父网段，名称设置为"同步规划网段"加上CIDR
            result = self.ipam.add_network(parent_cidr, f"同步规划网段 - {parent_cidr}")
            # 如果已存在会返回False，但不需要处理这个错误
        
        success_count = 0
        failed_count = 0
        
        for item in subnet_items:
            values = self.allocated_tree.item(item, "values")
            if len(values) >= 3:
                cidr = values[2]  # CIDR列
                name = values[1]  # 子网名称列
                
                # 调用IPAM的add_network方法添加网段
                result = self.ipam.add_network(cidr, name)
                if result[0]:
                    success_count += 1
                else:
                    failed_count += 1
        
        # 刷新IPAM网络列表
        self.refresh_ipam_networks()
        
        # 显示同步结果
        if failed_count == 0:
            self.show_info(_("hint"), _("sync_success").format(count=success_count))
        else:
            self.show_info(_("hint"), f"{_("sync_success").format(count=success_count)} ({failed_count} 个失败)")


    def generate_planning_chart_data(self, plan_result):
        """生成规划图表数据并绘制"""
        # 准备图表数据
        parent_cidr = plan_result["parent_cidr"]
        parent_info = get_subnet_info(parent_cidr)
        
        chart_data = {
            "parent": {
                "name": parent_cidr,
                "range": parent_info["num_addresses"]
            },
            "networks": [],
            "type": "plan"  # 添加图表类型字段，确保导出PDF时能正确识别
        }
        
        # 添加已分配子网（作为split类型）
        for subnet in plan_result["allocated_subnets"]:
            chart_data["networks"].append({
                "name": subnet["name"],
                "cidr": subnet["cidr"],
                "range": subnet["info"]["num_addresses"],
                "type": "split"
            })
        
        # 添加剩余子网（作为remaining类型）
        for i, subnet in enumerate(plan_result["remaining_subnets_info"]):
            chart_data["networks"].append({
                "name": plan_result["remaining_subnets"][i],
                "range": subnet["num_addresses"],
                "type": "remaining"
            })
        
        # 保存图表数据为实例属性，方便resize事件使用
        self.planning_chart_data = chart_data
        
        # 调用通用图表绘制函数
        draw_distribution_chart(self.planning_chart_canvas, chart_data, self.planning_chart_frame, chart_type="plan")

    def _create_scrollbar_callback(self, scrollbar, treeview=None, padx_adjust=0):
        """创建滚动条回调函数，实现自动隐藏功能

        Args:
            scrollbar: 滚动条组件
            treeview: Treeview组件（可选），用于调整padx
            padx_adjust: 滚动条隐藏时Treeview的右边距调整量

        Returns:
            function: 滚动条回调函数
        """
        import weakref
        # 使用弱引用避免内存泄露
        scrollbar_ref = weakref.ref(scrollbar)
        treeview_ref = weakref.ref(treeview) if treeview else None
        
        def scrollbar_callback(*args):
            # 获取滚动条和树视图的弱引用
            scrollbar = scrollbar_ref()
            treeview = treeview_ref() if treeview_ref else None
            
            # 如果滚动条已被销毁，直接返回
            if not scrollbar:
                return
            
            try:
                # 获取滚动条方向
                scrollbar_orient = scrollbar.cget('orient')
                scrollbar.set(*args)
                if float(args[0]) <= 0.0 and float(args[1]) >= 1.0:
                    scrollbar.grid_remove()
                    if treeview and padx_adjust:
                        try:
                            treeview.grid_configure(padx=padx_adjust)
                        except (tk.TclError, AttributeError):
                            pass
                else:
                    # 根据滚动条方向设置正确的网格位置
                    if scrollbar_orient == 'horizontal':
                        scrollbar.grid(row=1, column=0, sticky=tk.EW)
                    else:  # vertical
                        scrollbar.grid(row=0, column=1, sticky=tk.NS)
                    if treeview and padx_adjust:
                        try:
                            treeview.grid_configure(padx=0)
                        except (tk.TclError, AttributeError):
                            pass
            except (tk.TclError, AttributeError):
                # 如果组件已被销毁，直接返回
                return
        return scrollbar_callback

    def _validate_split_input(self, parent, split):
        """验证子网切分输入

        Args:
            parent: 父网段CIDR
            split: 切分段CIDR

        Returns:
            dict: 包含验证结果的字典 {'valid': bool, 'error': str or None}
        """
        # 验证输入不为空
        if not parent or not split:
            return {
                'valid': False,
                'error': _("parent_and_split_cidr_cannot_be_empty"),
                'error_code': 'empty_input'
            }

        # 验证并自动修正父网段CIDR格式
        try:
            parent_net = ipaddress.ip_network(parent, strict=False)
            parent_address = ipaddress.ip_address(parent.split('/')[0])
            if parent_address != parent_net.network_address:
                # 输入地址包含主机位，自动修正为正确的网络地址
                correct_parent = f"{parent_net.network_address}/{parent_net.prefixlen}"
                # 更新输入框内容
                self.parent_entry.delete(0, tk.END)
                self.parent_entry.insert(0, correct_parent)
        except ValueError:
            return {
                'valid': False,
                'error': _("invalid_parent_network_cidr"),
                'error_code': 'invalid_parent'
            }

        # 验证并自动修正切分段CIDR格式
        try:
            split_net = ipaddress.ip_network(split, strict=False)
            split_address = ipaddress.ip_address(split.split('/')[0])
            if split_address != split_net.network_address:
                # 输入地址包含主机位，自动修正为正确的网络地址
                correct_split = f"{split_net.network_address}/{split_net.prefixlen}"
                # 更新输入框内容
                self.split_entry.delete(0, tk.END)
                self.split_entry.insert(0, correct_split)
        except ValueError:
            return {
                'valid': False,
                'error': _("invalid_split_segment_cidr"),
                'error_code': 'invalid_split'
            }

        return {'valid': True, 'error': None, 'error_code': None}

    def on_ip_version_change(self):
        """处理IP版本切换事件
        
        当用户切换IPv4/IPv6时，更新父网段输入框的默认值和历史记录
        """
        try:
            ip_version = self.ip_version_var.get()
            
            # 清空当前输入框
            self.planning_parent_entry.delete(0, tk.END)
            
            # 根据IP版本选择对应的历史记录列表
            if ip_version == "IPv4":
                # 使用IPv4历史记录
                self.planning_parent_networks = self.planning_parent_networks_v4
                default_parent = self.planning_parent_networks_v4[0] if self.planning_parent_networks_v4 else "10.21.48.0/20"
            else:
                # 使用IPv6历史记录
                self.planning_parent_networks = self.planning_parent_networks_v6
                default_parent = self.planning_parent_networks_v6[0] if self.planning_parent_networks_v6 else "2001:0db8::/32"
            
            # 更新输入框默认值和下拉列表
            self.planning_parent_entry.insert(0, default_parent)
            self.planning_parent_entry.config(values=list(self.planning_parent_networks))
            
        except Exception as e:
            self.show_error(_("error"), f"{_("ip_version_change_failed")}: {str(e)}")
    
    def on_split_ip_version_change(self):
        """处理子网切分功能的IP版本切换事件
        
        当用户切换IPv4/IPv6时，更新子网切分输入框的默认值和历史记录
        """
        try:
            ip_version = self.split_ip_version_var.get()
            
            # 清空当前输入框
            self.parent_entry.delete(0, tk.END)
            self.split_entry.delete(0, tk.END)
            
            # 根据IP版本选择对应的历史记录列表
            if ip_version == "IPv4":
                # 使用IPv4历史记录
                self.split_parent_networks = self.split_parent_networks_v4
                self.split_networks = self.split_networks_v4
                default_parent = self.split_parent_networks_v4[0] if self.split_parent_networks_v4 else "10.0.0.0/8"
                default_split = self.split_networks_v4[0] if self.split_networks_v4 else "10.21.50.0/23"
            else:
                # 使用IPv6历史记录
                self.split_parent_networks = self.split_parent_networks_v6
                self.split_networks = self.split_networks_v6
                default_parent = self.split_parent_networks_v6[0] if self.split_parent_networks_v6 else "2001:0db8::/32"
                default_split = self.split_networks_v6[0] if self.split_networks_v6 else "2001:0db8::/64"
            
            # 更新输入框默认值和下拉列表
            self.parent_entry.insert(0, default_parent)
            self.parent_entry.config(values=list(self.split_parent_networks))
            
            self.split_entry.insert(0, default_split)
            self.split_entry.config(values=list(self.split_networks))
            
        except Exception as e:
            self.show_error(_("error"), f"{_("ip_version_change_failed")}: {str(e)}")

    def _validate_planning_input(self, parent):
        """验证子网规划输入

        Args:
            parent: 父网段CIDR

        Returns:
            dict: 包含验证结果的字典 {'valid': bool, 'error': str or None}
        """
        # 验证父网段不为空
        if not parent:
            return {
                'valid': False,
                'error': _("please_enter_parent_network"),
                'error_code': 'empty_parent'
            }

        # 验证父网段CIDR格式
        try:
            # 使用strict=False创建网络对象，允许主机位设置
            ip_network = ipaddress.ip_network(parent, strict=False)
            # 检查IP版本是否与当前选中的版本匹配
            selected_version = self.ip_version_var.get()
            actual_version = f"IPv{ip_network.version}"
            if selected_version != actual_version:
                return {
                    'valid': False,
                    'error': _("ip_versions_not_compatible"),
                    'error_code': 'version_mismatch'
                }
            
            # 检查输入的地址是否与网络地址完全匹配
            input_address = ipaddress.ip_address(parent.split('/')[0])
            if input_address != ip_network.network_address:
                # 输入地址包含主机位，自动修正为正确的网络地址
                correct_cidr = f"{ip_network.network_address}/{ip_network.prefixlen}"
                # 更新输入框内容
                self.planning_parent_entry.delete(0, tk.END)
                self.planning_parent_entry.insert(0, correct_cidr)
        except ValueError:
            return {
                'valid': False,
                'error': _("invalid_parent_network_format"),
                'error_code': 'invalid_parent'
            }

        return {'valid': True, 'error': None, 'error_code': None}

    def _update_history_entry(self, value, history_container, entry_widget):
        self.history_repo.update_history_entry(value, history_container, entry_widget)

    def _setup_scrollbar(self, scrollbar, widget, initial_hidden=True, widget_command=None, padx_adjust=0):
        """通用的滚动条配置方法

        Args:
            scrollbar: 滚动条组件
            widget: 关联的组件（Treeview/Canvas/Listbox等）
            initial_hidden: 初始时是否隐藏滚动条
            widget_command: 滚动条回调命令（如widget.yview）
            padx_adjust: 隐藏滚动条时widget的padx调整量
        """
        # 设置滚动条的命令
        if widget_command:
            scrollbar.config(command=widget_command)
        
        # 获取滚动条方向
        scrollbar_orient = scrollbar.cget('orient')
        
        # 根据滚动条方向设置相应的滚动命令
        if scrollbar_orient == 'horizontal':
            # 水平滚动条
            widget.configure(xscrollcommand=scrollbar.set)
            
            # 创建滚动条回调
            scrollbar_callback = self._create_scrollbar_callback(scrollbar, widget, padx_adjust)
            widget.configure(xscrollcommand=scrollbar_callback)
        else:
            # 垂直滚动条
            widget.configure(yscrollcommand=scrollbar.set)
            
            # 创建滚动条回调
            scrollbar_callback = self._create_scrollbar_callback(scrollbar, widget, padx_adjust)
            widget.configure(yscrollcommand=scrollbar_callback)

        # 初始隐藏或显示滚动条
        if initial_hidden:
            scrollbar.grid_forget()
        scrollbar_callback(0.0, 1.0)

    def _setup_accent_button_style(self, style_name, background_color, active_color, pressed_color):
        """配置强调按钮样式

        Args:
            style_name: 样式名称（如 "Accent.TButton"）
            background_color: 背景颜色
            active_color: 悬停颜色
            pressed_color: 按下颜色
        """
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        self.style.configure(
            style_name,
            background=background_color,
            foreground="white",
            font=(font_family, font_size, "bold"),
            padding=6,
        )
        self.style.map(
            style_name,
            background=[
                ("active", active_color),
                ("!active", background_color),
                ("pressed", pressed_color),
            ],
            foreground=[("active", "white"), ("!active", "white"), ("pressed", "white")],
        )

    def _perform_split_operation(self, parent, split):
        """执行核心切分操作

        Args:
            parent: 父网段
            split: 切分段

        Returns:
            bool: 切分是否成功
        """
        # 验证输入
        validation_result = self._validate_split_input(parent, split)
        if not validation_result['valid']:
            # 清空表格并显示错误信息
            self.clear_result()
            self.clear_tree_items(self.split_tree)
            self.show_error(_("input_error"), validation_result['error'])
            return False
        
        # 重新获取修正后的父网段和切分段
        parent = self.parent_entry.get().strip()
        split = self.split_entry.get().strip()

        try:
            # 只验证IP地址格式，不自动修正格式，保留用户输入的原始格式
            parent_network = ipaddress.ip_network(parent, strict=False)
            split_network = ipaddress.ip_network(split, strict=False)
            
            # 调用切分函数
            result = split_subnet(parent, split)

            # 清空现有结果
            self.clear_tree_items(self.split_tree)
            self.clear_tree_items(self.remaining_tree)

            if "error" in result:
                # 显示错误信息
                self.show_error(_("error"), result["error"])
                return False

            # 添加切分段信息，同时设置斑马条纹标签
            row_index = 0
            self.split_tree.insert("", tk.END, values=(_("parent_network"), result["parent_info"]["cidr"]), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            
            # 切分网段
            self.split_tree.insert("", tk.END, values=(_("split_segment"), result["split_info"]["cidr"]), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            self.split_tree.insert("", tk.END, values=(("-" * 10), ("-" * 20)), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1

            # 添加切分后的网段信息
            split_info = result["split_info"]
            self.split_tree.insert("", tk.END, values=(_("network_address"), split_info["network"]), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            
            # 根据IP版本显示不同的字段
            is_ipv6 = TableColumnManager.is_ipv6_network(parent)
            if not is_ipv6:
                # IPv4显示子网掩码、通配符掩码和广播地址
                self.split_tree.insert("", tk.END, values=(_("subnet_mask"), split_info["netmask"]), tags=("odd" if row_index % 2 == 0 else "even",))
                row_index += 1
                self.split_tree.insert("", tk.END, values=(_("wildcard_mask"), split_info["wildcard"]), tags=("odd" if row_index % 2 == 0 else "even",))
                row_index += 1
                self.split_tree.insert("", tk.END, values=(_("broadcast_address"), split_info["broadcast"]), tags=("odd" if row_index % 2 == 0 else "even",))
                row_index += 1
            
            self.split_tree.insert("", tk.END, values=(_("start_address"), split_info["host_range_start"]), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            self.split_tree.insert("", tk.END, values=(_("end_address"), split_info["host_range_end"]), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            
            # 调整顺序：可用地址数在前，总地址数在后
            self.split_tree.insert("", tk.END, values=(_("usable_addresses"), format_large_number(split_info["usable_addresses"], use_scientific=True)), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            self.split_tree.insert("", tk.END, values=(_("total_addresses"), format_large_number(split_info["num_addresses"], use_scientific=True)), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            
            self.split_tree.insert("", tk.END, values=(_("prefix_length"), split_info["prefixlen"]), tags=("odd" if row_index % 2 == 0 else "even",))
            row_index += 1
            self.split_tree.insert("", tk.END, values=(_("cidr"), split_info["cidr"]), tags=("odd" if row_index % 2 == 0 else "even",))

            # 检测IP版本
            is_ipv6 = TableColumnManager.is_ipv6_network(parent)
            
            # 根据IP版本显示或隐藏剩余网段表的列
            TableColumnManager.configure_split_remaining_tree(self.remaining_tree, is_ipv6)
            
            # 显示剩余网段表表格
            if result["remaining_subnets_info"]:
                for i, network in enumerate(result["remaining_subnets_info"], 1):
                    tags = ("even",) if i % 2 == 0 else ("odd",)
                    hidden_vals = TableColumnManager.get_hidden_values_for_ipv6(network, is_ipv6)
                    
                    self.remaining_tree.insert(
                        "",
                        tk.END,
                        values=(
                            i,
                            network["cidr"],
                            network["network"],
                            network["host_range_end"],
                            hidden_vals["netmask"],
                            hidden_vals["wildcard"],
                            hidden_vals["broadcast"],
                            format_large_number(network["usable_addresses"], use_scientific=True),
                        ),
                        tags=tags,
                    )

            else:
                self.remaining_tree.insert("", tk.END, values=(1, _("none"), _("none"), _("none"), _("none"), _("none"), _("none"), _("none")))

            # 不再手动调整表格宽度，依靠Tkinter的stretch=True自动处理

            # 优化滚动条状态更新，减少不必要的计算
            if hasattr(self, 'remaining_scroll_v'):
                # 获取当前滚动位置
                yview = self.remaining_tree.yview()
                need_scrollbar = not (float(yview[0]) <= 0.0 and float(yview[1]) >= 1.0)

                # 只在状态变化时才更新UI，减少不必要的刷新
                current_state = self.remaining_scroll_v.winfo_ismapped()
                if need_scrollbar != current_state:
                    if need_scrollbar:
                        self.remaining_scroll_v.grid(row=0, column=1, sticky=tk.NS)
                        self.remaining_scroll_v.set(yview[0], yview[1])
                    else:
                        self.remaining_scroll_v.grid_remove()

            self.prepare_chart_data(result, split_info, result["remaining_subnets_info"])

            # 绘制图表
            self.draw_distribution_chart()

            return True

        except ValueError as e:
            error_result = handle_ip_subnet_error(e)
            message = error_result.get('error', str(e))
            self.clear_result()
            self.split_tree.insert("", tk.END, values=(_("error"), message), tags=("error",))
            return False
        except (tk.TclError, AttributeError, TypeError) as e:
            self.clear_result()
            self.split_tree.insert("", tk.END, values=(_("error"), f"{_("unknown_error_occurred")}: {str(e)}"), tags=("error",))
            return False

    def perform_split(self, parent, split, from_history=False, auto_switch_version=False, fill_inputs=False, update_history=True):
        """执行切分操作

        Args:
            parent: 父网段
            split: 切分段
            from_history: 是否从历史记录重新执行，True表示不将操作记入历史
            auto_switch_version: 是否自动切换IP版本
            fill_inputs: 是否填充到输入框
            update_history: 是否更新历史记录

        Returns:
            bool: 切分是否成功
        """
        # 检测IP版本并自动切换
        detected_version = self.validation_service.detect_ip_version(parent)
        if not detected_version:
            # 如果父网段检测失败，尝试检测切分段
            detected_version = self.validation_service.detect_ip_version(split)
        
        if auto_switch_version and detected_version and detected_version != self.split_ip_version_var.get():
            self.split_ip_version_var.set(detected_version)
            self.on_split_ip_version_change()

        # 填充到输入框
        if fill_inputs:
            self.parent_entry.delete(0, tk.END)
            self.parent_entry.insert(0, parent)
            self.split_entry.delete(0, tk.END)
            self.split_entry.insert(0, split)

        # 执行核心切分操作
        success = self._perform_split_operation(parent, split)

        # 如果需要更新历史记录，且切分成功
        if success and update_history:
            # 重新获取修正后的父网段和切分段
            parent = self.parent_entry.get().strip()
            split = self.split_entry.get().strip()

            # 根据IP版本选择对应的历史容器
            if detected_version == 'IPv4':
                parent_history = self.split_parent_networks
                split_history = self.split_networks
            else:
                parent_history = self.split_parent_networks_v6
                split_history = self.split_networks_v6

            # 更新父网段历史（排到最前面）
            self._update_history_entry(parent, parent_history, self.parent_entry)
            # 更新切分段历史（排到最前面）
            self._update_history_entry(split, split_history, self.split_entry)

            # 添加到历史记录（通过history_repo自动处理重复记录并更新时间戳）
            self.history_repo.add_split_record(parent, split)

            # 更新历史记录列表
            self.update_history_listbox()

        return success

    def execute_split(self, from_history=False):
        """执行切分操作

        Args:
            from_history: 是否从历史记录重新执行，True表示不将操作记入历史
        """
        parent = self.parent_entry.get().strip()
        split = self.split_entry.get().strip()
        self.perform_split(parent, split, from_history=from_history, update_history=not from_history)

    def clear_tree_items(self, tree):
        """清空表格中的所有项

        Args:
            tree: 要清空的Treeview对象
        """
        # 批量删除所有子项，减少UI更新次数
        children = tree.get_children()
        if children:
            tree.delete(*children)

    def animate_info_bar(self, animation_type):
        """通用信息栏动画函数

        Args:
            animation_type: 动画类型，'show' 或 'hide'
        """
        # 如果已经在动画中，则跳过
        if self.info_bar_animating:
            return

        # 取消之前的自动隐藏定时器
        if self.info_auto_hide_id:
            self.root.after_cancel(self.info_auto_hide_id)
            self.info_auto_hide_id = None

        # 如果是隐藏动画且信息栏未映射，则跳过
        if animation_type == 'hide' and not self.info_bar_frame.winfo_ismapped():
            return

        self.info_bar_animating = True

        # 获取主窗口宽度
        main_width = self.main_frame.winfo_width()
        bar_x = 10
        bar_width = int(main_width * 1) - 50

        # 确保宽度在合理范围内
        bar_width = max(bar_width, 100)
        bar_width = min(bar_width, main_width - 20)

        # 动画参数配置
        animation_config = {
            'show': {
                'current_y': 30,
                'target_y': 0,
                'step': 1,
                'delay': 15,
                'condition': lambda y, t: y <= t,
                'update_y': lambda y, s: y - s,
                'on_complete': lambda: self._on_show_animation_complete(bar_x, bar_width)
            },
            'hide': {
                'current_y': 0,
                'target_y': 30,
                'step': 2,
                'delay': 10,
                'condition': lambda y, t: y >= t,
                'update_y': lambda y, s: y + s,
                'on_complete': lambda: self._on_hide_animation_complete(bar_x, bar_width)
            }
        }

        config = animation_config[animation_type]
        current_y = config['current_y']
        target_y = config['target_y']
        step = config['step']
        delay = config['delay']
        condition = config['condition']
        update_y = config['update_y']
        on_complete = config['on_complete']

        def animate():
            nonlocal current_y
            current_y = update_y(current_y, step)
            if condition(current_y, target_y):
                current_y = target_y
                on_complete()
            else:
                # 确保 spacer 在显示动画时可见
                if animation_type == 'show' and not self.info_spacer.winfo_ismapped():
                    self.info_spacer.pack(side="bottom", fill="x")
                self.info_bar_frame.place(x=bar_x, y=current_y, width=bar_width, height=30)
                self.root.after(delay, animate)

        animate()

    def _on_show_animation_complete(self, bar_x, bar_width):
        """显示动画完成后的回调"""
        import time
        self.info_bar_frame.place(x=bar_x, y=0, width=bar_width, height=30)
        self.info_bar_animating = False
        # 设置5秒后自动隐藏（仅当未暂停时）
        if not self.info_auto_hide_paused:
            self.info_auto_hide_id = self.root.after(5000, lambda: self.hide_info_bar(from_timer=True))
            self.info_auto_hide_scheduled_time = time.time()

    def _on_hide_animation_complete(self, _bar_x, _bar_width):
        """隐藏动画完成后的回调"""
        self.info_bar_frame.place_forget()
        self.info_spacer.pack_forget()
        self.info_bar_animating = False
        # 重置暂停状态
        self.info_auto_hide_paused = False

    def _on_info_bar_click(self, event=None):
        """点击信息栏时暂停自动消失"""
        import time
        # 如果已经暂停，恢复自动消失
        if self.info_auto_hide_paused:
            self.info_auto_hide_paused = False
            # 设置新的5秒定时器
            self.info_auto_hide_id = self.root.after(5000, lambda: self.hide_info_bar(from_timer=True))
            self.info_auto_hide_scheduled_time = time.time()
        else:
            # 暂停自动消失
            if self.info_auto_hide_id:
                self.root.after_cancel(self.info_auto_hide_id)
                self.info_auto_hide_id = None
            self.info_auto_hide_paused = True

    def hide_info_bar(self, from_timer=False):
        """隐藏信息栏"""
        import time
        
        # 如果是从定时器调用的，检查时间戳
        if from_timer and self.info_auto_hide_scheduled_time:
            current_time = time.time()
            if (current_time - self.info_auto_hide_scheduled_time) < 4.5:
                return
        
        self.animate_info_bar('hide')
        
    def toggle_info_bar_expand(self, event=None):
        """切换信息栏文本显示状态（完整/截断）"""
        if not hasattr(self, '_info_truncated') or not self._info_truncated:
            return
            
        # 切换显示状态
        self._info_currently_expanded = not self._info_currently_expanded
        
        if self._info_currently_expanded:
            font_family, default_font_size = get_current_font_settings()
            # 使用信息栏的独立字体大小配置
            info_bar_font_size = get_info_bar_font_size()
            font = tkfont.Font(family=font_family, size=info_bar_font_size)
            
            # 获取当前信息栏宽度（保持不变）
            current_width = self.info_bar_frame.winfo_width()
            
            # 智能换行算法（符合中日英韩文字习惯）
            def smart_wrap_text(text, max_width):
                """智能文本换行算法，支持中日英韩文字的混合排版"""
                if not text:
                    return ""
                

                
                lines = []
                current_line = ""
                current_word = ""
                
                # 定义字符类型
                def get_char_type(char):
                    code = ord(char)
                    if ((0x4E00 <= code <= 0x9FFF
                            or 0x3400 <= code <= 0x4DBF
                            or 0x20000 <= code <= 0x2A6DF
                            or 0x2A700 <= code <= 0x2B73F
                            or 0x2B740 <= code <= 0x2B81F
                            or 0x2B820 <= code <= 0x2CEAF
                            or 0x2CEB0 <= code <= 0x2EBEF
                            or 0x30000 <= code <= 0x3134F
                            or 0x31350 <= code <= 0x323AF)):
                        return 'cjk'
                    elif ((0x3040 <= code <= 0x309F  # 平假名
                            or 0x30A0 <= code <= 0x30FF)):  # 片假名
                        return 'cjk'
                    elif 0xAC00 <= code <= 0xD7AF:
                        return 'cjk'
                    elif ((0x3000 <= code <= 0x303F  # CJK标点
                            or 0xFF00 <= code <= 0xFFEF)):  # 全角字符
                        return 'cjk_punct'
                    # 英文字母
                    elif char.isalpha():
                        return 'alpha'
                    # 数字
                    elif char.isdigit():
                        return 'digit'
                    # 空格
                    elif char.isspace():
                        return 'space'
                    # 其他标点符号
                    else:
                        return 'punct'
                
                # 判断是否可以在字符后换行
                def can_break_after(char_type, char):
                    if char_type == 'cjk':
                        return True
                    elif char_type == 'cjk_punct':
                        return True
                    elif char_type == 'space':
                        return True
                    elif char_type == 'punct':
                        return True
                    return False
                
                # 判断是否可以在字符前换行
                def can_break_before(char_type, char):
                    if char_type == 'cjk':
                        return True
                    elif char_type == 'cjk_punct':
                        return True
                    elif char_type == 'punct':
                        return True
                    return False
                
                i = 0
                while i < len(text):
                    char = text[i]
                    char_type = get_char_type(char)
                    
                    if char_type in ['alpha', 'digit']:
                        # 英文单词或数字，累积整个单词
                        current_word += char
                        i += 1
                        # 检查下一个字符是否是单词的一部分
                        while i < len(text):
                            next_char = text[i]
                            next_type = get_char_type(next_char)
                            if next_type in ['alpha', 'digit']:
                                current_word += next_char
                                i += 1
                            else:
                                break
                        
                        # 尝试添加单词到当前行
                        test_line = current_line + current_word
                        test_width = font.measure(test_line)
                        
                        if test_width <= max_width:
                            current_line = test_line
                        else:
                            # 如果当前行不为空，先换行
                            if current_line:
                                lines.append(current_line)
                                current_line = current_word
                            else:
                                # 单词本身超过最大宽度，需要拆分
                                word_chars = list(current_word)
                                temp_line = ""
                                for wc in word_chars:
                                    test_w = temp_line + wc
                                    if font.measure(test_w) <= max_width:
                                        temp_line = test_w
                                    else:
                                        if temp_line:
                                            lines.append(temp_line)
                                        temp_line = wc
                                current_line = temp_line
                        current_word = ""
                    
                    elif char_type == 'cjk':
                        # CJK字符，可以单独换行
                        test_line = current_line + char
                        test_width = font.measure(test_line)
                        
                        if test_width <= max_width:
                            current_line = test_line
                        else:
                            if current_line:
                                lines.append(current_line)
                            current_line = char
                        i += 1
                    
                    elif char_type == 'cjk_punct':
                        # CJK标点符号
                        test_line = current_line + char
                        test_width = font.measure(test_line)
                        
                        if test_width <= max_width:
                            current_line = test_line
                        else:
                            # 标点符号不应该单独出现在行首
                            # 如果当前行不为空且回退一个字符后仍有内容，则回退换行
                            if current_line and len(current_line) > 0:
                                # 回退到前一个字符位置换行
                                lines.append(current_line[:-1])
                                current_line = current_line[-1] + char
                            elif current_line:
                                # 当前行只有一个字符，直接换行
                                lines.append(current_line)
                                current_line = char
                            else:
                                # 如果当前行为空，标点符号单独成行
                                lines.append(char)
                                current_line = ""
                        i += 1
                    
                    elif char_type == 'space':
                        # 空格，尝试添加到当前行
                        test_line = current_line + char
                        test_width = font.measure(test_line)
                        
                        if test_width <= max_width:
                            current_line = test_line
                        else:
                            # 空格会导致超出宽度，先换行
                            if current_line:
                                lines.append(current_line)
                            current_line = ""
                        i += 1
                    
                    elif char_type == 'punct':
                        # 英文标点符号
                        test_line = current_line + char
                        test_width = font.measure(test_line)
                        
                        if test_width <= max_width:
                            current_line = test_line
                        else:
                            # 标点符号不应该单独出现在行首
                            # 如果当前行不为空且回退一个字符后仍有内容，则回退换行
                            if current_line and len(current_line) > 0:
                                # 回退到前一个字符位置换行
                                lines.append(current_line[:-1])
                                current_line = current_line[-1] + char
                            elif current_line:
                                # 当前行只有一个字符，直接换行
                                lines.append(current_line)
                                current_line = char
                            else:
                                # 如果当前行为空，标点符号单独成行
                                lines.append(char)
                                current_line = ""
                        i += 1
                
                # 添加最后一行
                if current_line:
                    lines.append(current_line)
                
                # 实现两端对齐
                def justify_line(line, target_width, font):
                    """对单行进行两端对齐"""
                    if not line:
                        return line
                    
                    line_width = font.measure(line)
                    
                    # 如果行宽已经达到或超过目标宽度，不需要对齐
                    if line_width >= target_width:
                        return line
                    
                    # 检查是否包含空格
                    has_space = ' ' in line
                    
                    # 计算需要填充的宽度
                    padding_width = target_width - line_width
                    
                    if has_space:
                        # 有空格的文本（如英文）：在空格位置均匀分配额外空格
                        # 找到所有空格的位置
                        space_positions = []
                        for i, char in enumerate(line):
                            if char == ' ':
                                space_positions.append(i)
                        
                        if not space_positions:
                            return line
                        
                        # 计算每个空格需要增加的宽度
                        num_spaces = len(space_positions)
                        extra_width_per_space = padding_width / num_spaces
                        
                        # 计算空格的宽度
                        space_width = font.measure(' ')
                        
                        # 计算需要增加的空格数量（向上取整以确保达到目标宽度）
                        extra_spaces_per_gap = int(extra_width_per_space / space_width) + 1
                        
                        # 在每个空格位置增加空格
                        result = []
                        space_count = 0
                        for i, char in enumerate(line):
                            result.append(char)
                            if char == ' ' and space_count < num_spaces - 1:
                                # 增加额外的空格
                                for i in range(extra_spaces_per_gap):
                                    result.append(' ')
                                space_count += 1
                        
                        justified_line = ''.join(result)
                        
                        # 检查是否超出目标宽度，如果超出则减少一个空格
                        justified_width = font.measure(justified_line)
                        if justified_width > target_width and extra_spaces_per_gap > 0:
                            # 减少一个空格
                            result = []
                            space_count = 0
                            for i, char in enumerate(line):
                                result.append(char)
                                if char == ' ' and space_count < num_spaces - 1:
                                    for space in range(extra_spaces_per_gap - 1):
                                        result.append(' ')
                                    space_count += 1
                            justified_line = ''.join(result)
                        
                        return justified_line
                    else:
                        # 没有空格的文本（如中文）：在汉字之间均匀添加空格
                        # 找到所有非标点符号的汉字位置
                        char_positions = []
                        for i, char in enumerate(line):
                            # 跳过标点符号
                            if char not in '，。、；：「」『』（）【】！？.,;:!?':
                                char_positions.append(i)
                        
                        if len(char_positions) <= 1:
                            return line
                        
                        # 计算每个间隙需要增加的宽度
                        num_gaps = len(char_positions) - 1
                        extra_width_per_gap = padding_width / num_gaps
                        
                        # 计算空格的宽度
                        space_width = font.measure(' ')
                        
                        # 计算需要增加的空格数量（向上取整以确保达到目标宽度）
                        extra_spaces_per_gap = int(extra_width_per_gap / space_width) + 1
                        
                        if extra_spaces_per_gap <= 0:
                            return line
                        
                        # 在每个汉字之间添加空格
                        result = []
                        gap_count = 0
                        for i, char in enumerate(line):
                            result.append(char)
                            # 在汉字之间添加空格
                            if i in char_positions and gap_count < num_gaps:
                                for space in range(extra_spaces_per_gap):
                                    result.append(' ')
                                gap_count += 1
                        
                        justified_line = ''.join(result)
                        
                        # 检查是否超出目标宽度，如果超出则减少一个空格
                        justified_width = font.measure(justified_line)
                        if justified_width > target_width and extra_spaces_per_gap > 0:
                            # 减少一个空格
                            result = []
                            gap_count = 0
                            for i, char in enumerate(line):
                                result.append(char)
                                if i in char_positions and gap_count < num_gaps:
                                    for space in range(extra_spaces_per_gap - 1):
                                        result.append(' ')
                                    gap_count += 1
                            justified_line = ''.join(result)
                        
                        return justified_line
                
                # 对除最后一行外的所有行进行两端对齐
                justified_lines = []
                for i, line in enumerate(lines):
                    if i == len(lines) - 1:
                        # 最后一行不进行两端对齐
                        justified_lines.append(line)
                    else:
                        # 对其他行进行两端对齐
                        justified_line = justify_line(line, max_width, font)
                        justified_lines.append(justified_line)
                
                return '\n'.join(justified_lines)
            # 显示完整文本，首行加上图标
            # 先将图标添加到文本开头
            text_with_icon = self._info_icon + self._full_info_text
            
            # 计算最大行宽（不包括额外边距）
            max_line_width = current_width - 20  # 减去左右内边距
            
            # 对带图标的完整文本进行智能换行处理
            final_text = smart_wrap_text(text_with_icon, max_line_width)
            
            # 使用Text组件的方法设置文本
            # 强制将焦点从Text组件移开，避免渲染问题
            self.root.focus_set()
            
            self.info_label.config(text=final_text)
            
            # 根据消息类型设置文本颜色
            if "Error" in self._info_label_style:
                self.info_label.configure(fg="#c62828")  # 错误信息显示红色
            else:
                self.info_label.configure(fg="#424242")  # 正确信息显示灰色
            
            # 强制将焦点从Text组件移开，避免渲染问题
            # 使用after延迟确保焦点转移在禁用状态之后生效
            self.root.after(1, lambda: self.root.focus_set())
            
            # 强制更新布局，让label计算出正确的高度
            self.root.update_idletasks()
            
            # 获取label的实际高度
            label_height = self.info_label.winfo_reqheight()
            
            # 计算新的信息栏高度，添加额外的上下边距以确保文本完整显示
            # 给最后一行文字留出足够空间，添加4px额外高度
            new_height = label_height + 0  # 额外添加0px高度，确保最后一行完整显示
            new_height = max(new_height, 30)  # 最小高度30px
            
            # 更新信息栏框架高度
            self.info_bar_frame.place_configure(height=new_height)
            
            # 更新spacer高度，确保有足够空间显示
            self.info_spacer.configure(height=new_height)
            
            # 展开时停止自动消失计时，并显式设置暂停状态
            self.info_auto_hide_paused = True
            if hasattr(self, 'info_auto_hide_id') and self.info_auto_hide_id:
                self.root.after_cancel(self.info_auto_hide_id)
                self.info_auto_hide_id = None
            
            # 强制将焦点从Text组件移开，避免渲染问题
            self.root.focus_set()
        else:
            # 显示截断文本，使用信息栏的字体大小
            info_bar_font_size = get_info_bar_font_size()
            font = self._get_font(info_bar_font_size)
            truncated_text = self._truncate_text_by_pixel(self._full_info_text, self._info_icon, self._info_max_pixel_width, font)

            # 使用Text组件的方法设置文本
            # 强制将焦点从Text组件移开，避免渲染问题
            self.root.focus_set()
            
            self.info_label.config(text=self._info_icon + truncated_text)
            
            # 根据消息类型设置文本颜色
            if "Error" in self._info_label_style:
                self.info_label.configure(fg="#c62828")  # 错误信息显示红色
            else:
                self.info_label.configure(fg="#424242")  # 正确信息显示灰色
            
            # 恢复单行显示
            # 对于Label组件，高度由内容决定，不需要特别设置
            
            # 恢复原始高度，宽度保持不变
            original_height = 30
            self.info_bar_frame.place_configure(height=original_height)
            self.info_spacer.configure(height=original_height)
            
            # 强制将焦点从Text组件移开，避免渲染问题
            # 使用after延迟确保焦点转移在禁用状态之后生效
            self.root.after(1, lambda: self.root.focus_set())
            
            # 收起时显式重置暂停状态，并重新开始自动消失计时
            self.info_auto_hide_paused = False
            if hasattr(self, 'root'):
                if self.info_auto_hide_id:
                    self.root.after_cancel(self.info_auto_hide_id)
                self.info_auto_hide_id = self.root.after(5000, lambda: self.hide_info_bar(from_timer=True))
            
            # 强制将焦点从Text组件移开，避免渲染问题
            self.root.focus_set()

    def setup_advanced_tools_page(self):
        """设置高级工具功能的界面"""
        # 创建一个笔记本控件来显示不同的高级工具功能
        self.advanced_notebook = ColoredNotebook(self.advanced_frame, style=self.style)
        self.advanced_notebook.pack(fill=tk.BOTH, expand=True)

        # 1. IPv4地址信息查询功能 - 浅蓝色
        self.ipv4_info_frame = ttk.Frame(
            self.advanced_notebook.content_area, padding="5", style=self.advanced_notebook.light_blue_style
        )
        self.create_ipv4_info_section()

        # 2. IPv6地址信息查询功能 - 浅绿色
        self.ipv6_info_frame = ttk.Frame(
            self.advanced_notebook.content_area, padding="5", style=self.advanced_notebook.light_green_style
        )
        self.create_ipv6_info_section()

        # 3. 子网合并与范围转CIDR功能 - 浅紫色
        self.merge_frame = ttk.Frame(
            self.advanced_notebook.content_area, padding="5", style=self.advanced_notebook.light_purple_style
        )
        self.create_merged_subnets_and_cidr_section()

        # 5. 子网重叠检测功能 - 淡粉色
        self.overlap_frame = ttk.Frame(
            self.advanced_notebook.content_area, padding="5", style=self.advanced_notebook.light_pink_style
        )
        self.create_subnet_overlap_section()

        # 添加高级工具标签页
        self.advanced_notebook.add_tab(_("ipv4_query"), self.ipv4_info_frame, "#e3f2fd")  # 浅蓝色
        self.advanced_notebook.add_tab(_("ipv6_query"), self.ipv6_info_frame, "#e8f5e9")  # 浅绿色
        self.advanced_notebook.add_tab(_("subnet_merge"), self.merge_frame, "#f3e5f5")  # 浅紫色
        self.advanced_notebook.add_tab(_("overlap_detection"), self.overlap_frame, "#fce4ec")  # 淡粉色

    def create_ipv6_info_section(self):
        """创建IPv6地址信息查询功能界面"""
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 在ipv6_info_frame中增加中间容器，内边距10
        content_container = ttk.Frame(self.ipv6_info_frame, padding="10")
        content_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # 创建输入区域
        input_frame = ttk.LabelFrame(content_container, text=_("ipv6_address_info"), padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 10))

        # IPv6地址输入 - 使用Combobox，支持下拉选择和记忆功能
        ttk.Label(input_frame, text=_('ipv6_address')).pack(side=tk.LEFT, padx=(0, 5))
        self.ipv6_info_entry = ttk.Combobox(input_frame, values=self.ipv6_history, width=48, font=(font_family, font_size))
        self.ipv6_info_entry.pack(side=tk.LEFT, padx=(0, 10))
        # 使用历史记录的第一个元素作为默认值，如果没有则使用默认值
        default_ipv6 = self.ipv6_history[0] if self.ipv6_history else "2001:0db8:85a3:0000:0000:8a2e:0370:7334"
        self.ipv6_info_entry.insert(0, default_ipv6)
        self.ipv6_info_entry.config(state="normal")  # 允许手动输入

        # IPv6地址验证函数 - 使用统一的UI层验证方法
        def validate_ipv6(text):
            """验证IPv6地址格式"""
            return self.validate_cidr(text, self.ipv6_info_entry, ip_version="IPv6", require_prefix=False)

        # 配置验证
        self.ipv6_info_entry.config(validate="all", validatecommand=(self.root.register(validate_ipv6), "%P"))

        # 初始验证一次
        validate_ipv6(self.ipv6_info_entry.get())

        # CIDR下拉列表（IPv6支持1-128）
        ttk.Label(input_frame, text=_("cidr")).pack(side=tk.LEFT, padx=(0, 5))
        self.ipv6_cidr_var = tk.StringVar()
        self.ipv6_cidr_combobox = ttk.Combobox(
            input_frame, textvariable=self.ipv6_cidr_var, width=3, state="readonly", font=(font_family, font_size)
        )
        self.ipv6_cidr_combobox['values'] = list(range(1, 129))
        self.ipv6_cidr_combobox.current(63)  # 默认选择64
        self.ipv6_cidr_combobox.pack(side=tk.LEFT, padx=(0, 10))

        self.ipv6_info_btn = ttk.Button(input_frame, text=_("query_info"), command=self.execute_ipv6_info)
        self.ipv6_info_btn.pack(side=tk.RIGHT)

        # 创建结果区域
        result_frame = ttk.LabelFrame(content_container, text=_('query_result'), padding=(10, 10, 0, 10))
        result_frame.pack(fill=tk.BOTH, expand=True)

        # 创建Treeview和垂直滚动条
        self.ipv6_info_tree = ttk.Treeview(result_frame, columns=("item", "value"), show="headings")
        self.bind_treeview_right_click(self.ipv6_info_tree)
        self.ipv6_info_tree.heading("item", text=_("project"))
        self.ipv6_info_tree.heading("value", text=_("value"))

        self.ipv6_info_tree.column("item", width=185, minwidth=185, stretch=False)
        self.ipv6_info_tree.column("value", width=200)

        ipv6_info_scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL)

        self.create_scrollable_treeview(result_frame, self.ipv6_info_tree, ipv6_info_scrollbar)

        self.configure_treeview_styles(self.ipv6_info_tree, include_special_tags=True)

    def create_merged_subnets_and_cidr_section(self):
        """创建子网合并和范围转CIDR功能界面"""
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 创建输入部分的容器，包含所有组件
        input_container = ttk.Frame(self.merge_frame, padding="10")
        input_container.pack(fill=tk.BOTH, expand=True)

        # 创建两列框架，放置在输入容器中
        left_frame = ttk.Frame(input_container)
        right_frame = ttk.Frame(input_container)

        # 使用grid布局，固定左侧宽度，右侧自适应
        input_container.grid_columnconfigure(0, minsize=210, weight=0)  # 固定左侧宽度，与重叠检测面板左侧列宽度一致
        input_container.grid_columnconfigure(1, weight=1)  # 右侧自适应
        input_container.grid_rowconfigure(0, weight=1)  # 确保行能够撑满高度

        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        # 确保左侧面板内部的行能撑满高度
        left_frame.grid_rowconfigure(0, weight=1)  # 子网列表面板行
        left_frame.grid_rowconfigure(1, weight=0)  # IP地址范围面板行（固定高度）

        # 左侧上方：子网合并列表 - 使用grid布局
        subnet_frame = ttk.LabelFrame(left_frame, text=_('merge_subnets'), padding=(10, 10, 0, 10))
        subnet_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 5))

        # 配置左侧面板的grid布局
        left_frame.grid_rowconfigure(0, weight=1)  # 子网列表面板随窗体变化
        left_frame.grid_rowconfigure(1, weight=0)  # IP地址范围面板固定高度
        left_frame.grid_columnconfigure(0, weight=1)  # 第一列占满宽度

        # 子网合并列表输入文本框
        self.subnet_merge_text = tk.Text(subnet_frame, height=8, width=17, font=(font_family, font_size - 1), 
                                        bd=0, relief="flat", highlightthickness=0)

        subnet_merge_scrollbar = ttk.Scrollbar(subnet_frame, orient=tk.VERTICAL)
        subnet_merge_content = self.history_repo.load_text_data(HistorySQLite.CATEGORY_SUBNET_MERGE)
        if subnet_merge_content:
            self.subnet_merge_text.insert(tk.END, subnet_merge_content)
        else:
            self.subnet_merge_text.insert(tk.END, "192.168.0.0/24\n192.168.1.0/24\n192.168.2.0/24\n10.21.16.0/24\n10.21.17.0/24\n10.21.18.0/24\n10.21.19.128/26\n10.21.19.192/26\n2001:0db8::/127\n2001:0db8::2/127\n2001:0db8::4/127\n2001:0db8::6/127\n2001:0db8:1::/64\n2001:0db8:2::/64\n2001:0db8:3::/64")
        
        # 为文本框添加边框，与其他文本框风格一致
        self.subnet_merge_text.config(highlightbackground="#a9a9a9", 
                                      highlightcolor="#a9a9a9", 
                                      highlightthickness=1)
        
        # 实时验证子网格式
        def validate_subnet_merge_text(event=None):
            text = self.subnet_merge_text.get(1.0, tk.END).strip()
            self.subnet_merge_text.tag_remove("invalid", "1.0", tk.END)
            self.subnet_merge_text.tag_configure("invalid", foreground="red")
            
            lines = text.splitlines()
            for i, line in enumerate(lines, 1):
                subnet = line.strip()
                if subnet:
                    result = self.validation_service.validate_cidr(subnet, require_prefix=True)
                    if not result['valid']:
                        start = f"{i}.0"
                        end = f"{i}.end"
                        self.subnet_merge_text.tag_add("invalid", start, end)
        
        self.validate_subnet_merge_text = validate_subnet_merge_text
        self.subnet_merge_text.bind('<KeyRelease>', validate_subnet_merge_text)
        self.subnet_merge_text.bind('<FocusOut>', validate_subnet_merge_text)

        # 配置子网合并列表面板的grid布局
        subnet_frame.grid_columnconfigure(0, weight=1)  # 文本框列
        subnet_frame.grid_columnconfigure(1, weight=0)  # 滚动条列
        subnet_frame.grid_rowconfigure(0, weight=1)  # 文本框行
        subnet_frame.grid_rowconfigure(1, weight=0)  # 按钮行

        # 使用通用方法创建带自动隐藏滚动条的Text组件
        self.create_scrollable_text(subnet_frame, self.subnet_merge_text, subnet_merge_scrollbar)

        # 子网合并按钮 - 固定在右下角
        self.merge_btn = ttk.Button(subnet_frame, text=_("merge_subnet"), command=self.execute_merge_subnets)
        self.merge_btn.grid(row=1, column=0, columnspan=1, sticky="w", pady=(5, 0), padx=(0, 10))

        # 左侧下方：IP地址范围 - 使用grid布局
        range_frame = ttk.LabelFrame(left_frame, text=_("ip_address_range"), padding="10")
        range_frame.grid(row=1, column=0, sticky="ew", pady=(5, 0))  # 仅水平填充，固定高度

        # 起始IP - 使用Combobox，支持下拉选择和记忆功能
        start_frame = ttk.Frame(range_frame)
        start_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(start_frame, text=_('start')).pack(side=tk.LEFT, padx=(0, 5), pady=(0, 5))
        self.range_start_entry = ttk.Combobox(
            start_frame, values=self.range_start_history, width=13, font=(font_family, font_size)
        )
        self.range_start_entry.pack(side=tk.RIGHT, pady=(0, 5))
        # 使用历史记录的第一个元素作为默认值，如果没有则使用默认值
        default_start = self.range_start_history[0] if self.range_start_history else "192.168.0.1"
        self.range_start_entry.insert(0, default_start)
        self.range_start_entry.config(state="normal")  # 允许手动输入

        # IP范围地址验证函数 - 使用统一的UI层验证方法
        def validate_range_ip(text, entry):
            """验证IP范围地址格式，支持IPv4和IPv6"""
            return self.validate_cidr(text, entry, require_prefix=False)

        # 为起始IP添加验证
        def validate_start_ip(text):
            return validate_range_ip(text, self.range_start_entry)
        self.range_start_entry.config(validate="all", validatecommand=(self.range_start_entry.register(validate_start_ip), "%P"))

        validate_start_ip(self.range_start_entry.get())

        # 结束IP - 使用Combobox，支持下拉选择和记忆功能
        end_frame = ttk.Frame(range_frame)
        end_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(end_frame, text=_('end')).pack(side=tk.LEFT, padx=(0, 5), pady=(0, 5))
        self.range_end_entry = ttk.Combobox(end_frame, values=self.range_end_history, width=13, font=(font_family, font_size))
        self.range_end_entry.pack(side=tk.RIGHT, pady=(0, 5))
        # 使用历史记录的第一个元素作为默认值，如果没有则使用默认值
        default_end = self.range_end_history[0] if self.range_end_history else "192.168.30.254"
        self.range_end_entry.insert(0, default_end)
        self.range_end_entry.config(state="normal")  # 允许手动输入

        # 为结束IP添加验证
        def validate_end_ip(text):
            return validate_range_ip(text, self.range_end_entry)
        self.range_end_entry.config(validate="all", validatecommand=(self.range_end_entry.register(validate_end_ip), "%P"))

        validate_end_ip(self.range_end_entry.get())

        # 范围转CIDR按钮 - 靠左放置
        self.range_to_cidr_btn = ttk.Button(range_frame, text=_("convert_to_cidr"), command=self.execute_range_to_cidr)
        self.range_to_cidr_btn.pack(side=tk.LEFT, pady=(5, 0))

        # 右侧：直接放置IPv4和IPv6结果，去掉总框架
        # 创建IPv4结果框，右侧内边距为0，和其他框架保持一致
        self.ipv4_frame = ttk.LabelFrame(right_frame, text=_('IPv4结果'), padding=(10, 10, 0, 10))
        self.ipv4_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 创建IPv6结果框，右侧内边距为0，和其他框架保持一致
        self.ipv6_frame = ttk.LabelFrame(right_frame, text=_('IPv6结果'), padding=(10, 10, 0, 10))
        self.ipv6_frame.pack(fill=tk.BOTH, expand=True)

        # 创建正常的列结构 - IPv4和IPv6使用不同的列配置
        ipv4_columns = [_("cidr"), _("network_address"), _("subnet_mask"), _("broadcast_address"), _("host_count")]
        ipv6_columns = [_("cidr"), _("network_address"), _("host_count")]
        
        # 创建IPv4结果树
        self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=ipv4_columns, show="headings")
        self.bind_treeview_right_click(self.ipv4_result_tree)
        
        # 创建IPv6结果树
        self.ipv6_result_tree = ttk.Treeview(self.ipv6_frame, columns=ipv6_columns, show="headings")
        self.bind_treeview_right_click(self.ipv6_result_tree)
        
        # 使用公共方法配置IPv4结果树列宽
        self.configure_treeview_columns(self.ipv4_result_tree, ipv4_columns, config_type="ipv4_result")
        
        # 使用公共方法配置IPv6结果树列宽
        self.configure_treeview_columns(self.ipv6_result_tree, ipv6_columns, config_type="ipv6_result")
            
        # 添加垂直滚动条 - IPv4结果树
        ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
        self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
        self.configure_treeview_styles(self.ipv4_result_tree)
        
        # 添加垂直滚动条 - IPv6结果树
        ipv6_scrollbar = ttk.Scrollbar(self.ipv6_frame, orient=tk.VERTICAL)
        self.create_scrollable_treeview(self.ipv6_frame, self.ipv6_result_tree, ipv6_scrollbar)
        self.configure_treeview_styles(self.ipv6_result_tree)

    def create_scrollable_treeview(self, parent_frame, treeview, scrollbar, no_scrollbar_padx=(0, 10)):
        """
        创建带自动隐藏滚动条的Treeview，并实现滚动条隐藏时自动调整外边距

        参数:
            parent_frame: Treeview和滚动条的父容器
            treeview: 要添加滚动条的Treeview组件
            scrollbar: 滚动条组件
            no_scrollbar_padx: 滚动条隐藏时Treeview的右边距，默认(0, 10)
        """

        # 创建滚动条回调函数，实现自动隐藏和外边距调整
        def scrollbar_callback(*args):
            # 设置滚动条位置
            scrollbar.set(*args)

            yview = treeview.yview()
            need_scrollbar = not (float(yview[0]) <= 0.0 and float(yview[1]) >= 1.0)

            # 根据是否需要滚动条调整Treeview的右边距
            if need_scrollbar:
                # 显示滚动条
                scrollbar.grid(row=0, column=1, sticky=tk.NS, padx=0)
                # 调整Treeview的grid配置，移除右边距
                treeview.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=0)

            else:
                # 隐藏滚动条
                scrollbar.grid_remove()
                # 调整Treeview的grid配置，添加右边距
                treeview.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=no_scrollbar_padx)

        # 绑定滚动条和Treeview
        scrollbar.config(command=treeview.yview)
        treeview.config(yscrollcommand=scrollbar_callback)

        # 使用grid布局，确保Treeview和滚动条正确对齐
        treeview.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar.grid(row=0, column=1, sticky=tk.NS, padx=0)

        # 配置grid权重，使Treeview可以扩展
        parent_frame.grid_rowconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)

        # 初始调用一次回调函数，设置初始状态
        scrollbar_callback(0.0, 1.0)

    def create_treeview_component(self, parent_frame, columns, show_headings=True):
        """创建并配置Treeview组件
        
        参数:
            parent_frame: Treeview的父容器
            columns: 列配置列表
            show_headings: 是否显示表头，默认为True
        
        返回:
            tuple: (treeview, scrollbar) - 创建的Treeview组件和滚动条组件
        """
        # 创建Treeview组件
        tree = ttk.Treeview(parent_frame, columns=columns, show="headings" if show_headings else "")
        self.bind_treeview_right_click(tree)
        
        # 使用公共方法配置列宽
        self.configure_treeview_columns(tree, columns, config_type="default")
        
        # 创建并配置滚动条
        scrollbar = ttk.Scrollbar(parent_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # 使用grid布局显示Treeview和滚动条
        parent_frame.grid_rowconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(1, weight=0)
        
        tree.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar.grid(row=0, column=1, sticky=tk.NS)
        
        # 配置Treeview样式
        self.configure_treeview_styles(tree)
        
        return tree, scrollbar
        
    def configure_treeview_columns(self, tree, columns, config_type="default"):
        """配置Treeview组件的列宽
        
        参数:
            tree: Treeview组件
            columns: 列配置列表
            config_type: 配置类型，可选值："default", "ipv4_result", "ipv6_result", "transposed"
        """
        # 设置列标题
        for col in columns:
            tree.heading(col, text=col)
        
        # 根据配置类型设置列宽
        if config_type == "transposed":
            # 转置Treeview的列配置
            tree.column(columns[0], width=180, minwidth=180, stretch=False)
            for col in columns[1:]:
                tree.column(col, width=100, stretch=True)
        elif config_type == "ipv4_result":
            # IPv4结果树的列配置
            for i, col in enumerate(columns):
                if i == 0:  # CIDR列，需要更大宽度以适应长IP地址
                    tree.column(col, width=130, minwidth=130, stretch=True)
                elif i == 1:  # 网络地址列
                    tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 2:  # 子网掩码列
                    tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 3:  # 广播地址列
                    tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 4:  # 主机数列，可适当缩小宽度
                    tree.column(col, width=80, minwidth=80, stretch=False)
        elif config_type == "ipv6_result":
            # IPv6结果树的列配置
            for i, col in enumerate(columns):
                if i == 0:  # CIDR列
                    tree.column(col, width=120, minwidth=120, stretch=True)
                elif i == 1:  # 网络地址列
                    tree.column(col, width=120, minwidth=120, stretch=True)
                elif i == 2:  # 主机数列
                    tree.column(col, width=100, minwidth=100, stretch=False)
        else:  # default
            # 默认列配置，适用于create_treeview_component方法
            for i, col in enumerate(columns):
                if i == 0:  # CIDR列
                    tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 1:  # 网络地址列
                    tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 2:  # 子网掩码列或主机数列
                    if len(columns) <= 3:  # IPv6配置，只有3列
                        tree.column(col, width=100, minwidth=100, stretch=False)
                    else:  # IPv4配置，有5列
                        tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 3:  # 广播地址列
                    tree.column(col, width=100, minwidth=100, stretch=True)
                elif i == 4:  # 主机数列
                    tree.column(col, width=80, minwidth=80, stretch=False)

    def create_transposed_treeview(self, parent_frame, cidrs, is_ipv6=False):
        """创建并配置转置后的Treeview组件
        
        参数:
            parent_frame: Treeview的父容器
            cidrs: CIDR列表
            is_ipv6: 是否为IPv6，默认为False
        
        返回:
            tuple: (treeview, scrollbar) - 创建的转置Treeview组件和滚动条组件
        """
        # 转换为转置后的列结构
        transposed_columns = [_("attribute")] + cidrs
        
        # 创建Treeview组件
        tree = ttk.Treeview(parent_frame, columns=transposed_columns, show="headings")
        self.bind_treeview_right_click(tree)
        
        # 使用公共方法配置列宽
        self.configure_treeview_columns(tree, transposed_columns, config_type="transposed")
        
        # 创建并配置滚动条
        scrollbar = ttk.Scrollbar(parent_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # 使用grid布局显示Treeview和滚动条
        parent_frame.grid_rowconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(1, weight=0)
        
        tree.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar.grid(row=0, column=1, sticky=tk.NS)
        
        # 配置Treeview样式
        self.configure_treeview_styles(tree)
        
        return tree, scrollbar

    def create_horizontal_scrollbar(self, parent_frame, treeview, scrollbar):
        """
        创建带自动隐藏功能的水平滚动条

        参数:
            parent_frame: Treeview和滚动条的父容器
            treeview: 要添加滚动条的Treeview组件
            scrollbar: 水平滚动条组件
        """
        # 初始状态：完全隐藏滚动条，不占用任何空间
        scrollbar.grid_remove()
        
        # 创建滚动条回调函数，实现自动隐藏
        def scrollbar_callback(*args):
            # 设置滚动条位置
            scrollbar.set(*args)
            
            # 只有当Treeview实际可见时才检查是否需要滚动条
            if treeview.winfo_ismapped():
                # 检查是否需要滚动条
                xview = treeview.xview()
                need_scrollbar = not (float(xview[0]) <= 0.0 and float(xview[1]) >= 1.0)
                
                if need_scrollbar:
                    # 显示水平滚动条，设置columnspan确保横跨整个容器
                    scrollbar.grid(row=1, column=0, columnspan=2, sticky=tk.EW)
                else:
                    # 隐藏水平滚动条
                    scrollbar.grid_remove()
        
        # 绑定滚动条和Treeview
        scrollbar.config(command=treeview.xview)
        treeview.config(xscrollcommand=scrollbar_callback)
        
        # 配置父容器的grid布局
        parent_frame.grid_rowconfigure(1, weight=0)
        
        # 不立即调用回调函数，等待Treeview完全渲染后由系统自动触发
        # 这样可以避免初始化时的闪烁问题

    def create_scrollable_treeview_with_grid(self, parent_frame, treeview, scrollbar,
                                           tree_row=0, tree_column=0, scrollbar_row=0, scrollbar_column=1,
                                           tree_padx=(0, 0), scrollbar_padx=(0, 0), no_scrollbar_padx=(0, 10)):
        """
        创建带自动隐藏滚动条的Treeview，并实现滚动条隐藏时自动调整外边距
        支持自定义grid位置

        参数:
            parent_frame: Treeview和滚动条的父容器
            treeview: 要添加滚动条的Treeview组件
            scrollbar: 滚动条组件
            tree_row: Treeview的grid行位置，默认0
            tree_column: Treeview的grid列位置，默认0
            scrollbar_row: 滚动条的grid行位置，默认0
            scrollbar_column: 滚动条的grid列位置，默认1
            tree_padx: Treeview的grid padx参数，默认(0, 0)
            scrollbar_padx: 滚动条的grid padx参数，默认(0, 0)
            no_scrollbar_padx: 滚动条隐藏时Treeview的右边距，默认(0, 10)
        """

        def scrollbar_callback(*args):
            scrollbar.set(*args)

            yview = treeview.yview()
            need_scrollbar = not (float(yview[0]) <= 0.0 and float(yview[1]) >= 1.0)

            if need_scrollbar:
                scrollbar.grid(row=scrollbar_row, column=scrollbar_column, sticky=tk.NS, padx=scrollbar_padx)
                treeview.grid_configure(row=tree_row, column=tree_column, sticky=tk.NSEW, padx=tree_padx)

                # 如果是需求池表或子网需求表，减小name列宽度为滚动条留出空间
                if treeview in [getattr(self, 'pool_tree', None), getattr(self, 'requirements_tree', None)]:
                    try:
                        treeview.column("name", width=110)  # 减小name列宽度
                    except tk.TclError:
                        pass  # 如果列不存在则忽略
            else:
                scrollbar.grid_remove()
                adjusted_padx = (tree_padx[0], tree_padx[1] + no_scrollbar_padx[1]) if tree_padx else no_scrollbar_padx
                treeview.grid_configure(row=tree_row, column=tree_column, sticky=tk.NSEW, padx=adjusted_padx)

        scrollbar.config(command=treeview.yview)
        treeview.config(yscrollcommand=scrollbar_callback)

        treeview.grid(row=tree_row, column=tree_column, sticky=tk.NSEW, padx=tree_padx)
        scrollbar.grid(row=scrollbar_row, column=scrollbar_column, sticky=tk.NS, padx=scrollbar_padx)

        parent_frame.grid_rowconfigure(tree_row, weight=1)
        parent_frame.grid_columnconfigure(tree_column, weight=1)

        scrollbar_callback(0.0, 1.0)

    def create_scrollable_text(self, parent_frame, text_widget, scrollbar, no_scrollbar_padx=(0, 10)):
        """
        创建带自动隐藏滚动条的Text组件，并实现滚动条隐藏时自动调整外边距

        参数:
            parent_frame: Text组件和滚动条的父容器
            text_widget: 要添加滚动条的Text组件
            scrollbar: 滚动条组件
            no_scrollbar_padx: 滚动条隐藏时Text组件的右边距，默认(0, 10)
        """

        def scrollbar_callback(*args):
            scrollbar.set(*args)

            yview = text_widget.yview()
            need_scrollbar = not (float(yview[0]) <= 0.0 and float(yview[1]) >= 1.0)

            # 根据是否需要滚动条调整Text组件的右边距
            if need_scrollbar:
                scrollbar.grid(row=0, column=1, sticky=tk.NS)
                # 调整Text组件的grid配置，移除右边距
                text_widget.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=0)
            else:
                scrollbar.grid_remove()
                # 调整Text组件的grid配置，添加右边距
                text_widget.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=no_scrollbar_padx)

        # 绑定滚动条和Text组件
        scrollbar.config(command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar_callback)

        # 使用grid布局，确保Text组件和滚动条正确对齐
        text_widget.grid(row=0, column=0, sticky=tk.NSEW)
        scrollbar.grid(row=0, column=1, sticky=tk.NS)

        # 配置grid权重，使Text组件可以扩展
        parent_frame.grid_rowconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)

        scrollbar_callback(0.0, 1.0)

    def create_ipv4_info_section(self):
        """创建IPv4地址信息查询功能界面"""
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 在ipv4_info_frame中增加中间容器
        content_container = ttk.Frame(self.ipv4_info_frame, padding="10")
        content_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        input_frame = ttk.LabelFrame(content_container, text=_("ipv4_address_info"), padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 10))

        # IP地址输入 - 使用Combobox，支持下拉选择和记忆功能
        ttk.Label(input_frame, text=_('ipv4_address')).pack(side=tk.LEFT, padx=(0, 5))
        self.ip_info_entry = ttk.Combobox(input_frame, values=self.ipv4_history, width=21, font=(font_family, font_size))
        self.ip_info_entry.pack(side=tk.LEFT, padx=(0, 10))
        # 使用历史记录的第一个元素作为默认值，如果没有则使用默认值
        default_ipv4 = self.ipv4_history[0] if self.ipv4_history else "192.168.1.1"
        self.ip_info_entry.insert(0, default_ipv4)
        self.ip_info_entry.config(state="normal")  # 允许手动输入

        # IPv4地址验证函数 - 使用统一的UI层验证方法
        def validate_ipv4(text):
            """验证IPv4地址格式"""
            return self.validate_cidr(text, self.ip_info_entry, ip_version="IPv4", require_prefix=False)

        self.ip_info_entry.config(validate="all", validatecommand=(self.root.register(validate_ipv4), "%P"))

        validate_ipv4(self.ip_info_entry.get())

        # 常用子网掩码与CIDR的映射关系，包含所有CIDR 1~32
        self.subnet_mask_cidr_map = {
            "128.0.0.0": "1",
            "192.0.0.0": "2",
            "224.0.0.0": "3",
            "240.0.0.0": "4",
            "248.0.0.0": "5",
            "252.0.0.0": "6",
            "254.0.0.0": "7",
            "255.0.0.0": "8",
            "255.128.0.0": "9",
            "255.192.0.0": "10",
            "255.224.0.0": "11",
            "255.240.0.0": "12",
            "255.248.0.0": "13",
            "255.252.0.0": "14",
            "255.254.0.0": "15",
            "255.255.0.0": "16",
            "255.255.128.0": "17",
            "255.255.192.0": "18",
            "255.255.224.0": "19",
            "255.255.240.0": "20",
            "255.255.248.0": "21",
            "255.255.252.0": "22",
            "255.255.254.0": "23",
            "255.255.255.0": "24",
            "255.255.255.128": "25",
            "255.255.255.192": "26",
            "255.255.255.224": "27",
            "255.255.255.240": "28",
            "255.255.255.248": "29",
            "255.255.255.252": "30",
            "255.255.255.254": "31",
            "255.255.255.255": "32",
        }

        # 创建反向映射，用于从CIDR获取子网掩码
        self.cidr_subnet_mask_map = {v: k for k, v in self.subnet_mask_cidr_map.items()}

        # 子网掩码下拉列表
        ttk.Label(input_frame, text=_("subnet_mask")).pack(side=tk.LEFT, padx=(0, 5))
        self.ip_mask_var = tk.StringVar()
        self.ip_mask_combobox = ttk.Combobox(
            input_frame, textvariable=self.ip_mask_var, width=15, state="readonly", font=(font_family, font_size)
        )
        self.ip_mask_combobox['values'] = list(self.subnet_mask_cidr_map.keys())
        self.ip_mask_combobox.current(list(self.subnet_mask_cidr_map.keys()).index("255.255.255.0"))
        self.ip_mask_combobox.pack(side=tk.LEFT, padx=(0, 10))
        # 绑定子网掩码选择事件
        self.ip_mask_combobox.bind("<<ComboboxSelected>>", self.on_subnet_mask_change)

        # CIDR下拉列表
        ttk.Label(input_frame, text="CIDR").pack(side=tk.LEFT, padx=(0, 5))
        self.ip_cidr_var = tk.StringVar()
        self.ip_cidr_combobox = ttk.Combobox(
            input_frame, textvariable=self.ip_cidr_var, width=3, state="readonly", font=(font_family, font_size)
        )
        self.ip_cidr_combobox['values'] = list(range(1, 33))
        self.ip_cidr_combobox.current(23)  # 默认选择24
        self.ip_cidr_combobox.pack(side=tk.LEFT, padx=(0, 10))
        # 绑定CIDR选择事件
        self.ip_cidr_combobox.bind("<<ComboboxSelected>>", self.on_cidr_change)

        self.ip_info_btn = ttk.Button(input_frame, text=_("query_info"), command=self.execute_ipv4_info)
        self.ip_info_btn.pack(side=tk.RIGHT)

        result_frame = ttk.LabelFrame(content_container, text=_("query_result"), padding=(10, 10, 0, 10))
        result_frame.pack(fill=tk.BOTH, expand=True)

        self.ip_info_tree = ttk.Treeview(result_frame, columns=("item", "value"), show="headings")
        self.bind_treeview_right_click(self.ip_info_tree)
        self.ip_info_tree.heading("item", text=_("project"))
        self.ip_info_tree.heading("value", text=_("value"))

        self.ip_info_tree.column("item", width=185, minwidth=185, stretch=False)
        self.ip_info_tree.column("value", width=200)

        ip_info_scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL)
        ip_info_scrollbar.config(command=self.ip_info_tree.yview)

        def scrollbar_callback(*args):
            ip_info_scrollbar.set(*args)

            yview = self.ip_info_tree.yview()
            need_scrollbar = not (float(yview[0]) <= 0.0 and float(yview[1]) >= 1.0)

            if need_scrollbar:
                ip_info_scrollbar.grid(row=0, column=1, sticky=tk.NS)
                self.ip_info_tree.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=0)
            else:
                ip_info_scrollbar.grid_remove()
                self.ip_info_tree.grid_configure(row=0, column=0, sticky=tk.NSEW, padx=(0, 10))

        self.ip_info_tree.config(yscrollcommand=scrollbar_callback)

        self.ip_info_tree.grid(row=0, column=0, sticky=tk.NSEW)
        ip_info_scrollbar.grid(row=0, column=1, sticky=tk.NS)

        result_frame.grid_rowconfigure(0, weight=1)
        result_frame.grid_columnconfigure(0, weight=1)

        scrollbar_callback(0.0, 1.0)

        self.configure_treeview_styles(self.ip_info_tree, include_special_tags=True)

    def create_subnet_overlap_section(self):
        """创建子网重叠检测功能界面"""
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 在overlap_frame中增加中间容器，内边距10
        content_container = ttk.Frame(self.overlap_frame, padding="10")
        content_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # 创建左右两列框架
        left_frame = ttk.Frame(content_container)
        right_frame = ttk.Frame(content_container)

        content_container.grid_columnconfigure(0, minsize=210, weight=0)  # 固定左侧宽度，与子网合并页面保持一致
        content_container.grid_columnconfigure(1, weight=1)  # 右侧自适应
        content_container.grid_rowconfigure(0, weight=1)  # 确保行能够撑满高度

        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        # 左侧：子网列表
        left_frame.grid_columnconfigure(0, weight=1)  # 确保左侧框架第一列占满宽度
        left_frame.grid_rowconfigure(0, weight=1)  # 确保左侧框架第一行撑满高度
        
        input_frame = ttk.LabelFrame(left_frame, text=_('overlap_subnets'), padding=(10, 10, 0, 10))
        input_frame.pack(fill=tk.BOTH, expand=True)

        # 子网输入文本框和滚动条
        text_frame = ttk.Frame(input_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)

        self.overlap_text = tk.Text(text_frame, height=10, width=17, font=(font_family, font_size - 1), 
                                    bd=0, relief="flat", highlightthickness=1, 
                                    highlightbackground="#a9a9a9", highlightcolor="#a9a9a9")
        overlap_content = self.history_repo.load_text_data(HistorySQLite.CATEGORY_OVERLAP_DETECTION)
        if overlap_content:
            self.overlap_text.insert(tk.END, overlap_content)
        else:
            self.overlap_text.insert(tk.END, "192.168.0.0/24\n192.168.0.128/25\n10.0.0.0/16\n10.0.0.128/25\n10.0.10.0/20\n10.10.0.0/23\n2001:0db8::/64\n2001:0db8::1000/120\n2001:0db8:1::/64\n2001:0db8:2::/64\n2001:0db8:1:0::/66\n2001:0db8:1:1000::/66")
        
        # 实时验证子网格式
        def validate_overlap_text(event=None):
            text = self.overlap_text.get(1.0, tk.END).strip()
            self.overlap_text.tag_remove("invalid", "1.0", tk.END)
            self.overlap_text.tag_configure("invalid", foreground="red")
            
            lines = text.splitlines()
            for i, line in enumerate(lines, 1):
                subnet = line.strip()
                if subnet:
                    result = self.validation_service.validate_cidr(subnet, require_prefix=True)
                    if not result['valid']:
                        start = f"{i}.0"
                        end = f"{i}.end"
                        self.overlap_text.tag_add("invalid", start, end)
        
        self.validate_overlap_text = validate_overlap_text
        self.overlap_text.bind('<KeyRelease>', validate_overlap_text)
        self.overlap_text.bind('<FocusOut>', validate_overlap_text)

        # 添加垂直滚动条，并使用通用方法创建带自动隐藏滚动条的Text组件
        overlap_text_scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL)

        self.create_scrollable_text(text_frame, self.overlap_text, overlap_text_scrollbar)

        # 检测重叠按钮 - 靠左放置
        self.overlap_btn = ttk.Button(input_frame, text=_("check_overlap"), command=self.execute_check_overlap)
        self.overlap_btn.pack(side=tk.LEFT, pady=(5, 0), padx=(0, 10))

        # 右侧：检测结果
        result_frame = ttk.LabelFrame(right_frame, text=_('detection_result'), padding=(10, 10, 0, 10))
        result_frame.pack(fill=tk.BOTH, expand=True)

        self.overlap_result_tree = ttk.Treeview(result_frame, columns=("status", "message"), show="headings", height=5)
        self.bind_treeview_right_click(self.overlap_result_tree)
        self.overlap_result_tree.heading("status", text=_("status"))
        self.overlap_result_tree.heading("message", text=_("description"))

        self.overlap_result_tree.column("status", width=60, minwidth=60, stretch=True)
        self.overlap_result_tree.column("message", width=400, minwidth=400, stretch=True)

        overlap_result_scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL)

        self.create_scrollable_treeview(result_frame, self.overlap_result_tree, overlap_result_scrollbar)

        self.configure_treeview_styles(self.overlap_result_tree, include_special_tags=True)

    def execute_merge_subnets(self):
        """执行子网合并操作"""
        try:
            # 定义默认列配置（IPv4）
            ipv4_columns = [_("cidr"), _("network_address"), _("subnet_mask"), _("broadcast_address"), _("host_count")]
            # 定义IPv6列配置（不包含子网掩码和广播地址）
            ipv6_columns = [_("cidr"), _("network_address"), _("host_count")]
            
            # 检查并恢复Treeview组件的列结构
            # 1. 处理IPv4结果树
            # 检查当前列配置是否与需要的配置一致
            current_ipv4_cols = self.ipv4_result_tree["columns"]
            if list(current_ipv4_cols) != ipv4_columns:
                # 如果列配置不一致，重新创建Treeview组件
                for child in self.ipv4_result_tree.winfo_children():
                    child.destroy()
                self.ipv4_result_tree.destroy()
                
                # 重新创建IPv4结果树
                self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=ipv4_columns, show="headings")
                self.bind_treeview_right_click(self.ipv4_result_tree)
                
                # 创建IPv4结果树滚动条并使用按需显示功能
                ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
            
            # 清空现有内容
            self.clear_tree_items(self.ipv4_result_tree)
            
            # 使用公共方法配置IPv4结果树列宽
            self.configure_treeview_columns(self.ipv4_result_tree, ipv4_columns, config_type="ipv4_result")
            self.configure_treeview_styles(self.ipv4_result_tree)
            
            # 2. 处理IPv6结果树
            # 检查当前列配置是否与需要的配置一致
            current_ipv6_cols = self.ipv6_result_tree["columns"]
            if list(current_ipv6_cols) != ipv6_columns:
                # 如果列配置不一致，重新创建Treeview组件
                for child in self.ipv6_result_tree.winfo_children():
                    child.destroy()
                self.ipv6_result_tree.destroy()
                
                # 重新创建IPv6结果树
                self.ipv6_result_tree = ttk.Treeview(self.ipv6_frame, columns=ipv6_columns, show="headings")
                self.bind_treeview_right_click(self.ipv6_result_tree)
                
                # 创建IPv6结果树滚动条并使用按需显示功能
                ipv6_scrollbar = ttk.Scrollbar(self.ipv6_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv6_frame, self.ipv6_result_tree, ipv6_scrollbar)
            
            # 清空现有内容
            self.clear_tree_items(self.ipv6_result_tree)
            
            # 使用公共方法配置IPv6结果树列宽
            self.configure_treeview_columns(self.ipv6_result_tree, ipv6_columns, config_type="ipv6_result")
            self.configure_treeview_styles(self.ipv6_result_tree)
            
            # 获取输入的子网合并列表
            subnets_text = self.subnet_merge_text.get(1.0, tk.END).strip()
            if not subnets_text:
                self.show_info(_("hint"), _("enter_subnet_merge_list"))
                return

            # 解析子网合并列表
            subnets = [line.strip() for line in subnets_text.splitlines() if line.strip()]

            # 执行合并
            result = merge_subnets(subnets)

            # 清空结果表格
            self.clear_tree_items(self.ipv4_result_tree)
            self.clear_tree_items(self.ipv6_result_tree)

            # 重新运行实时验证，保持标记与当前输入一致
            self.validate_subnet_merge_text()

            # 获取无效子网列表（如果有）
            invalid_subnets = []
            if isinstance(result, dict) and "error" in result:
                invalid_subnets = result.get("invalid_subnets", [])
            
            # 如果有无效子网，在结果表格中显示错误信息（但不中断执行）
            row_index = 0
            if invalid_subnets:
                for invalid in invalid_subnets:
                    tag = "error_row"
                    simplified_error = _("invalid_format")
                    subnet = invalid["subnet"]
                    
                    # 使用评分机制智能判断IP版本类型
                    ipv4_score = 0
                    ipv6_score = 0
                    
                    # 检查IPv4特征
                    if '.' in subnet:
                        ipv4_score += 1
                    if subnet.endswith('.'):
                        ipv4_score += 1
                        ipv6_score -= 1
                    
                    # 检查IPv6特征
                    if ':' in subnet:
                        ipv6_score += 1
                    if subnet.endswith(':'):
                        ipv6_score += 1
                        ipv4_score -= 1
                    
                    # 检查十六进制字符（IPv6特有）
                    has_hex = any(c in 'abcdefABCDEF' for c in subnet)
                    if has_hex:
                        ipv6_score += 1
                        ipv4_score -= 1
                    
                    # 纯数字且长度较短（更可能是IPv4的一部分）
                    if subnet.isdigit() and len(subnet) <= 3:
                        ipv4_score += 1
                        ipv6_score -= 1
                    
                    # 根据评分判断插入位置
                    if ipv4_score > ipv6_score:
                        # 更像IPv4
                        self.ipv4_result_tree.insert("", tk.END, values=(
                            subnet,
                            simplified_error,
                            "",
                            "",
                            ""
                        ), tags=(tag,))
                    elif ipv6_score > ipv4_score:
                        # 更像IPv6
                        self.ipv6_result_tree.insert("", tk.END, values=(
                            subnet,
                            simplified_error,
                            ""
                        ), tags=(tag,))
                    else:
                        # 无法判断，两张表都插
                        self.ipv4_result_tree.insert("", tk.END, values=(
                            subnet,
                            simplified_error,
                            "",
                            "",
                            ""
                        ), tags=(tag,))
                        self.ipv6_result_tree.insert("", tk.END, values=(
                            subnet,
                            simplified_error,
                            ""
                        ), tags=(tag,))
                    row_index += 1

            # 分离IPv4和IPv6结果（无论是否有错误，都尝试获取合并结果）
            merged_subnets = []
            if isinstance(result, dict):
                merged_subnets = result.get("merged_subnets", [])
            ipv4_results = []
            ipv6_results = []
            
            for subnet in merged_subnets:
                info = get_subnet_info(subnet)
                if info["version"] == 4:  # 使用整数比较，因为get_subnet_info返回的是整数
                    ipv4_results.append((subnet, info))
                else:
                    ipv6_results.append((subnet, info))

            # 填充IPv4结果表格
            row_index = 0
            for subnet, info in ipv4_results:
                row_values = [
                    subnet,  # CIDR
                    info["network"],  # 网络地址
                    info["netmask"],  # 子网掩码
                    info["broadcast"],  # 广播地址
                    format_large_number(info["usable_addresses"], use_scientific=True),  # 可用主机数，使用科学计数法
                ]
                tag = "odd" if row_index % 2 == 0 else "even"
                self.ipv4_result_tree.insert("", tk.END, values=row_values, tags=(tag,))
                row_index += 1
            
            # 填充IPv6结果表格
            row_index = 0
            for subnet, info in ipv6_results:
                row_values = [
                    subnet,  # CIDR
                    info["network"],  # 网络地址
                    format_large_number(info["usable_addresses"], use_scientific=True),  # 可用主机数，使用科学计数法
                ]
                tag = "odd" if row_index % 2 == 0 else "even"
                self.ipv6_result_tree.insert("", tk.END, values=row_values, tags=(tag,))
                row_index += 1

            # 更新斑马条纹
            self.update_table_zebra_stripes(self.ipv4_result_tree)
            self.update_table_zebra_stripes(self.ipv6_result_tree)

            # 操作成功完成，添加到历史记录
            self.update_range_start_history()
            self.update_range_end_history()

            # 持久化子网合并列表文本到数据库
            self._persist_text_data(HistorySQLite.CATEGORY_SUBNET_MERGE, self.subnet_merge_text)

        except ValueError as e:
            # 在结果表格中显示错误信息
            self.clear_tree_items(self.ipv4_result_tree)
            self.clear_tree_items(self.ipv6_result_tree)
            
            # 尝试获取无效子网列表
            error_result = handle_ip_subnet_error(e)
            error_msg = f"{_("merge_subnet")}{_("failed")}: {error_result.get('error', str(e))}"
            tag = "error_row"
            self.ipv4_result_tree.insert("", tk.END, values=(
                _("error"),
                error_msg,
                "",
                "",
                ""
            ), tags=(tag,))
            
            # 清除无效子网高亮
            self.subnet_merge_text.tag_remove("invalid", "1.0", tk.END)
        except (tk.TclError, AttributeError, TypeError) as e:
            # 在结果表格中显示错误信息
            self.clear_tree_items(self.ipv4_result_tree)
            self.clear_tree_items(self.ipv6_result_tree)
            
            # 尝试获取无效子网列表
            error_msg = f"{_("operation_failed")}: {str(e)}"
            tag = "error_row"
            self.ipv4_result_tree.insert("", tk.END, values=(
                _("error"),
                error_msg,
                "",
                "",
                ""
            ), tags=(tag,))
            
            # 清除无效子网高亮
            self.subnet_merge_text.tag_remove("invalid", "1.0", tk.END)

    def execute_ipv6_info(self):
        """执行IPv6地址信息查询"""
        try:
            for item in self.ipv6_info_tree.get_children():
                self.ipv6_info_tree.delete(item)

            ipv6_full = self.ipv6_info_entry.get().strip()
            if not ipv6_full:
                self.show_info(_("hint"), _("enter_ipv6_address"))
                return

            # 移除CIDR前缀，获取纯IPv6地址
            ipv6 = ipv6_full.split('/')[0]
            cidr = self.ipv6_cidr_var.get()

            # 验证IPv6地址格式
            try:
                ipaddress.IPv6Address(ipv6)
            except ValueError as e:
                try:
                    # 使用handle_ip_subnet_error函数获取友好的错误信息
                    self.show_error(_("error"), handle_ip_subnet_error(e)["error"])
                    return
                except ValueError:
                    # 如果handle_ip_subnet_error失败，使用通用错误信息
                    self.show_error(_("error"), f"{_("ipv6_address")}{_("invalid_ip_format")}")
                    return

            # 构建网络字符串
            network_str = f"{ipv6}/{cidr}"
            ipv6_info = get_ip_info(network_str)
            original_ip_info = get_ip_info(ipv6)

            # 插入基本信息
            self.ipv6_info_tree.insert("", tk.END, values=(_("ip_address"), ipv6_info.get("ip_address", "")))
            self.ipv6_info_tree.insert("", tk.END, values=(_("version"), ipv6_info.get("version", "")))
            ip_address = ipv6_info.get("ip_address", "")
            
            # 直接从ipv6_info获取地址类型
            address_type = ipv6_info.get("address_type", "unknown")
            self.ipv6_info_tree.insert("", tk.END, values=(_("address_type"), _(address_type)))
            self.ipv6_info_tree.insert("", tk.END, values=(_("cidr_prefix"), ipv6_info.get("cidr", "")))
            self.ipv6_info_tree.insert("", tk.END, values=(_("network_address"), ipv6_info.get("network_address", "")))
            # 添加多播地址说明
            self.ipv6_info_tree.insert("", tk.END, values=(_("multicast_address_note"), _("ipv6_multicast_note")))
            # IPv6没有广播地址，显示多播地址说明
            first_host = ipv6_info.get("first_host", "")
            if first_host:
                # IPv6使用多播地址代替广播，第一个可用主机
                self.ipv6_info_tree.insert("", tk.END, values=(_("first_usable_host"), first_host))
            last_host = ipv6_info.get("last_host", "")
            if last_host:
                # IPv6使用多播地址代替广播，最后一个可用主机不是多播地址
                self.ipv6_info_tree.insert("", tk.END, values=(_("last_usable_host"), last_host))
            self.ipv6_info_tree.insert("", tk.END, values=(_("usable_hosts"), ipv6_info.get("usable_hosts", "")))
            self.ipv6_info_tree.insert("", tk.END, values=(_("total_hosts"), ipv6_info.get("total_hosts", "")))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("address_format"), ""), tags=("section",))
            self.ipv6_info_tree.insert("", tk.END, values=(_("compressed_format"), ipv6_info.get("compressed", "")))
            self.ipv6_info_tree.insert("", tk.END, values=(_("expanded_format"), ipv6_info.get("exploded", "")))
            self.ipv6_info_tree.insert("", tk.END, values=(_("reverse_dns_format"), ipv6_info.get("reverse_dns", "")))

            if "::ffff:" in ip_address:
                mapped_ipv4 = ip_address.split("::ffff:")[-1]
                self.ipv6_info_tree.insert("", tk.END, values=(_("mapped_ipv4_address"), mapped_ipv4))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("address_properties"), ""), tags=("section",))
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_global_routable"), _("yes") if ipv6_info.get("is_global") else _("no"))
            )
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_private_address"), _("yes") if ipv6_info.get("is_private") else _("no"))
            )
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_link_local"), _("yes") if ipv6_info.get("is_link_local") else _("no"))
            )
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_loopback"), _("yes") if ipv6_info.get("is_loopback") else _("no"))
            )
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_multicast"), _("yes") if ipv6_info.get("is_multicast") else _("no"))
            )
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_unspecified"), _("yes") if ipv6_info.get("is_unspecified") else _("no"))
            )
            self.ipv6_info_tree.insert(
                "", tk.END, values=(_("is_reserved"), _("yes") if ipv6_info.get("is_reserved") else _("no"))
            )
            self.ipv6_info_tree.insert("", tk.END, values=(_("is_ipv4_mapped"), _("yes") if "::ffff:" in ip_address else _("no")))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("address_structure_analysis"), ""), tags=("section",))

            # 直接从ipv6_info获取前缀分析
            prefix_analysis = ipv6_info.get("prefix_analysis", "unknown_prefix")
            user_cidr = ipv6_info.get("prefix_length", ipv6_info.get("cidr", 128))

            full_prefix_analysis = f"{_(prefix_analysis)} {_('network_prefix')}：/{user_cidr}"
            self.ipv6_info_tree.insert("", tk.END, values=(_("prefix_analysis"), full_prefix_analysis))

            # 使用原始IP地址的展开格式计算段数（总是8段）
            segments_count = original_ip_info.get("exploded", "").split(":")
            if len(segments_count) > 1:
                self.ipv6_info_tree.insert("", tk.END, values=(_("address_segment_count"), f"{len(segments_count)}"))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("binary_representation"), ""), tags=("section",))
            self.ipv6_info_tree.insert("", tk.END, values=(_("ip_address"), ipv6_info.get("binary", "")))

            if ipv6_info.get("network_address"):
                network_addr_value = ipv6_info["network_address"]
                network_bin_value = network_addr_value.replace(':', '').zfill(32)
                network_bin_grouped = ' '.join([network_bin_value[i:i + 4] for i in range(0, 32, 4)])
                self.ipv6_info_tree.insert("", tk.END, values=(_("network_address"), network_bin_grouped))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("hexadecimal_representation"), ""), tags=("section",))
            self.ipv6_info_tree.insert("", tk.END, values=(_("ip_address"), ipv6_info.get("hexadecimal", "")))
            self.ipv6_info_tree.insert("", tk.END, values=(_("network_address"), ipv6_info.get("network_address", "")))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("decimal_value_representation"), ""), tags=("section",))
            if "integer" in ipv6_info:
                self.ipv6_info_tree.insert("", tk.END, values=(_("ip_address"), ipv6_info["integer"]))

            if ipv6_info.get("network_address"):
                network_addr = ipv6_info["network_address"]
                network_int = int(ipaddress.IPv6Address(network_addr))
                self.ipv6_info_tree.insert("", tk.END, values=(_("network_address"), network_int))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("address_segment_details"), ""), tags=("section",))
            
            # 直接从original_ip_info获取段详情
            segment_details = original_ip_info.get("segment_details", [])
            
            for detail in segment_details:
                segment = detail["segment"]
                if segment:
                    dec_value = detail["decimal"]
                    bin_value = detail["binary"]
                    self.ipv6_info_tree.insert(
                        "",
                        tk.END,
                        values=(_("segment_index").format(detail["index"]), _("segment_value").format(segment, dec_value, bin_value)),
                    )
                else:
                    self.ipv6_info_tree.insert(
                        "", tk.END, values=(_("segment_index").format(detail["index"]), _("segment_value_zero"))
                    )

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("network_scale_and_usage"), ""), tags=("section",))

            prefix_length = ipv6_info.get("prefix_length", ipv6_info.get("cidr", 128))
            size_desc = ""
            if prefix_length == 128:
                size_desc = _("single_host_address")
            elif prefix_length == 64:
                size_desc = _("small_network")
            elif prefix_length == 48:
                size_desc = _("medium_network")
            elif 40 <= prefix_length <= 47:
                size_desc = _("regional_network").format(prefix_length)
            elif 32 < prefix_length <= 39:
                size_desc = _("large_network").format(prefix_length)
            elif prefix_length <= 32:
                size_desc = _("extra_large_network")
            else:
                size_desc = _("special_network").format(prefix_length)
            self.ipv6_info_tree.insert("", tk.END, values=(_("subnet_size"), size_desc))

            usage_desc = ""
            if ipv6_info.get("is_loopback"):
                usage_desc = _("purpose_loopback_ipv6")
            elif ipv6_info.get("is_link_local"):
                usage_desc = _("purpose_link_local")
            elif ipv6_info.get("is_multicast"):
                usage_desc = _("purpose_multicast_ipv6")
            elif "::ffff:" in ip_address:
                usage_desc = _("purpose_ipv4_mapped")
            elif ip_address.startswith("64:ff9b::"):
                usage_desc = _("purpose_ipv4_ipv6_translation")
            elif ip_address.startswith("fc00:") or ip_address.startswith("fd00:"):
                usage_desc = _("purpose_ula")
            elif ip_address.startswith("2000:") or ip_address.startswith("2001:") or ip_address.startswith("2002:"):
                usage_desc = _("purpose_global_unicast")
            elif ip_address.startswith("2001:db8::"):
                usage_desc = _("purpose_documentation")
            elif ip_address.startswith("100::"):
                usage_desc = _("purpose_blackhole")
            elif ip_address.startswith("2001:10::"):
                usage_desc = _("purpose_orchid")
            elif ip_address == "::":
                usage_desc = _("purpose_unspecified")
            elif ip_address.startswith("fec0:"):
                usage_desc = _("purpose_deprecated_site_local")
            elif ipv6_info.get("is_global"):
                usage_desc = _("purpose_global_unicast")
            elif ipv6_info.get("is_private"):
                usage_desc = _("purpose_ula")
            else:
                usage_desc = _("purpose_specific")
            self.ipv6_info_tree.insert("", tk.END, values=(_("main_usage"), usage_desc))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("configuration_advice"), ""), tags=("section",))

            advice = ""
            if ipv6_info.get("is_global"):
                advice = _("advice_global_routable")
            elif ipv6_info.get("is_private"):
                advice = _("advice_private_network")
            if ipv6_info.get("prefix_length", 0) < 64:
                advice += _("advice_prefix_length")
            self.ipv6_info_tree.insert("", tk.END, values=(_("network_configuration"), advice))

            self.ipv6_info_tree.insert("", tk.END, values=())
            self.ipv6_info_tree.insert("", tk.END, values=(_("rfc_standards_reference"), ""), tags=("section",))
            # 直接从ipv6_info字典中获取RFC参考信息
            rfc_ref = ipv6_info.get("rfc_reference", "")
            self.ipv6_info_tree.insert("", tk.END, values=(_("related_rfc"), rfc_ref))

            if ip_address.startswith("::ffff:"):
                ipv4_mapped = ip_address.replace("::ffff:", "")
                self.ipv6_info_tree.insert("", tk.END, values=())
                self.ipv6_info_tree.insert("", tk.END, values=(_("extended_information"), ""), tags=("section",))
                self.ipv6_info_tree.insert("", tk.END, values=(_("ipv4_mapped_address"), ipv4_mapped))

            elif ip_address.startswith("2001:0db8:"):
                self.ipv6_info_tree.insert("", tk.END, values=())
                self.ipv6_info_tree.insert("", tk.END, values=(_("extended_information"), ""), tags=("section",))
                self.ipv6_info_tree.insert("", tk.END, values=(_("address_usage"), "文档/测试地址 (RFC 3849)"))
                self.ipv6_info_tree.insert("", tk.END, values=(_("rfc_specification"), "RFC 3849 - IPv6文档地址分配"))
            elif ip_address.startswith("fc00:") or ip_address.startswith("fd00:"):
                self.ipv6_info_tree.insert("", tk.END, values=())
                self.ipv6_info_tree.insert("", tk.END, values=(_("extended_information"), ""), tags=("section",))
                self.ipv6_info_tree.insert("", tk.END, values=(_("address_usage"), "唯一本地地址 (ULA)"))
                self.ipv6_info_tree.insert("", tk.END, values=(_("rfc_specification"), "RFC 4193 - IPv6唯一本地地址"))
            elif ip_address.startswith("fe80:"):
                self.ipv6_info_tree.insert("", tk.END, values=())
                self.ipv6_info_tree.insert("", tk.END, values=(_("extended_information"), ""), tags=("section",))
                self.ipv6_info_tree.insert("", tk.END, values=(_("address_usage"), "链路本地地址"))
                self.ipv6_info_tree.insert("", tk.END, values=(_("rfc_specification"), "RFC 4291 - IPv6寻址架构"))
            elif ipv6_info.get("is_multicast"):
                self.ipv6_info_tree.insert("", tk.END, values=())
                self.ipv6_info_tree.insert("", tk.END, values=(_("extended_information"), ""), tags=("section",))
                self.ipv6_info_tree.insert("", tk.END, values=(_("address_usage"), "组播地址"))
                self.ipv6_info_tree.insert("", tk.END, values=(_("rfc_specification"), "RFC 4291 - IPv6寻址架构"))

            self.update_ipv6_history()

        except ValueError as e:
            error_result = handle_ip_subnet_error(e)
            self.show_error(_("error"), f"{_("query_failed")}: {error_result.get('error', str(e))}")
        except (tk.TclError, AttributeError, TypeError) as e:
            self.show_error(_("error"), f"{_("operation_failed")}: {str(e)}")

    def execute_ipv4_info(self):
        """执行IPv4地址信息查询"""
        try:
            for item in self.ip_info_tree.get_children():
                self.ip_info_tree.delete(item)

            ip = self.ip_info_entry.get().strip()
            if not ip:
                self.show_info(_("hint"), _("enter_ip_address"))
                return

            # 验证IPv4地址格式
            try:
                ipaddress.IPv4Address(ip)
            except ValueError as e:
                try:
                    error_info = handle_ip_subnet_error(e)
                    self.show_error(_("error"), error_info["error"])
                    return
                except ValueError:
                    self.show_error(_("error"), f"{_("ipv4_address")}{_("invalid_ip_format")}")
                    return

            subnet_mask = self.ip_mask_var.get()
            cidr = self.ip_cidr_var.get()

            network_str = None
            # 优先使用子网掩码计算网络地址
            if subnet_mask:
                try:
                    # 使用IP地址和子网掩码计算网络地址
                    ip_int = ip_to_int(ip)
                    mask_int = ip_to_int(subnet_mask)
                    network_int = ip_int & mask_int
                    network_address = int_to_ip(network_int)
                    prefix_len = bin(mask_int).count('1')
                    network_str = f"{network_address}/{prefix_len}"
                except (ValueError, TypeError):
                    pass

            # 如果子网掩码计算失败，再尝试使用CIDR
            if not network_str and cidr:
                try:
                    # 使用CIDR直接构建网络地址
                    temp_network = ipaddress.ip_network(f"{ip}/{cidr}", strict=False)
                    network_str = str(temp_network)
                except (ValueError, TypeError):
                    pass

            basic_info = True
            subnet_info = None

            if network_str:
                try:
                    # 自动修正CIDR格式（将IP地址转换为正确的网络地址）
                    corrected_network = ipaddress.ip_network(network_str, strict=False)
                    corrected_network_str = str(corrected_network)
                    subnet_info = get_subnet_info(corrected_network_str)
                    basic_info = False
                except (ValueError, TypeError):
                    pass

            info = get_ip_info(ip)

            if not basic_info and subnet_info:
                self.ip_info_tree.insert("", tk.END, values=(_("ip_address"), ip))
                self.ip_info_tree.insert("", tk.END, values=(_("subnet_mask"), subnet_info["netmask"]))
                wildcard_mask = '.'.join(str(255 - int(octet)) for octet in subnet_info["netmask"].split('.'))
                self.ip_info_tree.insert("", tk.END, values=(_("wildcard_mask"), wildcard_mask))
                self.ip_info_tree.insert("", tk.END, values=(_("cidr"), subnet_info["cidr"]))
                # 网络类别
                network_class = info.get("class", "")
                class_text = _(f"class_{network_class.lower()}") if network_class else ""
                self.ip_info_tree.insert("", tk.END, values=(_("network_class"), class_text))
                self.ip_info_tree.insert("", tk.END, values=(_("default_netmask"), info.get("default_netmask", "")))

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_address_range"), ""), tags=("section",))
                self.ip_info_tree.insert("", tk.END, values=(_("network_address"), subnet_info["network"]))
                self.ip_info_tree.insert("", tk.END, values=(_("broadcast_address"), subnet_info["broadcast"]))
                self.ip_info_tree.insert("", tk.END, values=(_("first_usable_address"), subnet_info["host_range_start"]))
                self.ip_info_tree.insert("", tk.END, values=(_("last_usable_address"), subnet_info["host_range_end"]))
                self.ip_info_tree.insert("", tk.END, values=(_("usable_hosts"), format_large_number(subnet_info["usable_addresses"], use_scientific=True)))
                self.ip_info_tree.insert("", tk.END, values=(_("total_hosts"), format_large_number(subnet_info["num_addresses"], use_scientific=True)))

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_binary_representation"), ""), tags=("section",))
                self.ip_info_tree.insert("", tk.END, values=(_("ip_address"), info["binary"]))
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(_("subnet_mask"), '.'.join(f'{int(octet):08b}' for octet in subnet_info["netmask"].split('.'))),
                )
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(
                        _("wildcard_mask"),
                        '.'.join(f'{255 - int(octet):08b}' for octet in subnet_info["netmask"].split('.')),
                    ),
                )
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(_("network_address"), '.'.join(f'{int(octet):08b}' for octet in subnet_info["network"].split('.'))),
                )
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(_("broadcast_address"), '.'.join(f'{int(octet):08b}' for octet in subnet_info["broadcast"].split('.'))),
                )

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_hexadecimal_representation"), ""), tags=("section",))
                self.ip_info_tree.insert("", tk.END, values=(_("ip_address"), info["hexadecimal"]))
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(_("subnet_mask"), '.'.join(f'{int(octet):02x}' for octet in subnet_info["netmask"].split('.'))),
                )
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(
                        _("wildcard_mask"),
                        '.'.join(f'{255 - int(octet):02x}' for octet in subnet_info["netmask"].split('.')),
                    ),
                )
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(_("network_address"), '.'.join(f'{int(octet):02x}' for octet in subnet_info["network"].split('.'))),
                )
                self.ip_info_tree.insert(
                    "",
                    tk.END,
                    values=(_("broadcast_address"), '.'.join(f'{int(octet):02x}' for octet in subnet_info["broadcast"].split('.'))),
                )

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_decimal_representation"), ""), tags=("section",))
                self.ip_info_tree.insert("", tk.END, values=(_("ip_address"), info["integer"]))
                self.ip_info_tree.insert("", tk.END, values=(_("subnet_mask"), str(ip_to_int(subnet_info["netmask"]))))
                wildcard_int = ip_to_int('.'.join(str(255 - int(octet)) for octet in subnet_info["netmask"].split('.')))
                self.ip_info_tree.insert("", tk.END, values=(_("wildcard_mask"), str(wildcard_int)))
                self.ip_info_tree.insert("", tk.END, values=(_("network_address"), str(ip_to_int(subnet_info["network"]))))
                self.ip_info_tree.insert("", tk.END, values=(_("broadcast_address"), str(ip_to_int(subnet_info["broadcast"]))))

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_ip_properties"), ""), tags=("section",))
                self.ip_info_tree.insert("", tk.END, values=(_("version"), info["version"]))
                self.ip_info_tree.insert("", tk.END, values=(_("is_private"), _("yes") if info["is_private"] else _("no")))
                self.ip_info_tree.insert("", tk.END, values=(_("is_reserved"), _("yes") if info["is_reserved"] else _("no")))
                self.ip_info_tree.insert("", tk.END, values=(_("is_loopback"), _("yes") if info["is_loopback"] else _("no")))
                self.ip_info_tree.insert("", tk.END, values=(_("is_multicast"), _("yes") if info["is_multicast"] else _("no")))
                self.ip_info_tree.insert("", tk.END, values=(_("is_global"), _("yes") if info["is_global"] else _("no")))
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_link_local"), _("yes") if info["is_link_local"] else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_unspecified"), _("yes") if info["is_unspecified"] else _("no"))
                )

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_extended_info"), ""), tags=("section",))

                # IP地址用途
                if info["is_loopback"]:
                    ip_purpose = _("purpose_loopback")
                elif info["is_private"]:
                    ip_purpose = _("purpose_private")
                elif info["is_multicast"]:
                    ip_purpose = _("purpose_multicast")
                elif info["is_reserved"]:
                    ip_purpose = _("purpose_reserved")
                elif info["is_global"]:
                    ip_purpose = _("purpose_global")
                else:
                    ip_purpose = _("purpose_unknown")
                self.ip_info_tree.insert("", tk.END, values=(_("ip_purpose"), ip_purpose))

                # 子网规模
                subnet_size = subnet_info["usable_addresses"]
                if subnet_size <= 254:
                    size_desc = _("size_small")
                elif subnet_size <= 65534:
                    size_desc = _("size_medium")
                else:
                    size_desc = _("size_large")
                self.ip_info_tree.insert("", tk.END, values=(_("subnet_size"), size_desc))

                # 配置建议
                if subnet_size > 65534:
                    config_advice = _("advice_large_subnet")
                elif info["is_private"]:
                    config_advice = _("advice_private_network")
                else:
                    config_advice = _("advice_public_network")
                self.ip_info_tree.insert("", tk.END, values=(_("configuration_advice"), config_advice))
            else:
                # 基本信息模式
                self.ip_info_tree.insert("", tk.END, values=(_("ip_address"), info.get("ip_address", ip)))
                self.ip_info_tree.insert("", tk.END, values=(_("version"), info.get("version", "")))
                network_class = info.get("class", "")
                class_text = _(f"class_{network_class.lower()}") if network_class else ""
                self.ip_info_tree.insert("", tk.END, values=(_("network_class"), class_text))
                self.ip_info_tree.insert("", tk.END, values=(_("default_netmask"), info.get("default_netmask", "")))

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("value_representation"), ""), tags=("section",))
                self.ip_info_tree.insert("", tk.END, values=(_("binary_representation"), info.get("binary", "")))
                self.ip_info_tree.insert("", tk.END, values=(_("hexadecimal_representation"), info.get("hexadecimal", "")))
                self.ip_info_tree.insert("", tk.END, values=(_("integer_representation"), info.get("integer", "")))

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=(_("section_ip_properties"), ""), tags=("section",))
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_private"), _("yes") if info.get("is_private", False) else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_reserved"), _("yes") if info.get("is_reserved", False) else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_loopback"), _("yes") if info.get("is_loopback", False) else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_multicast"), _("yes") if info.get("is_multicast", False) else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_global_routable"), _("yes") if info.get("is_global", False) else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_link_local"), _("yes") if info.get("is_link_local", False) else _("no"))
                )
                self.ip_info_tree.insert(
                    "", tk.END, values=(_("is_unspecified"), _("yes") if info.get("is_unspecified", False) else _("no"))
                )

                self.ip_info_tree.insert("", tk.END, values=())
                self.ip_info_tree.insert("", tk.END, values=("扩展信息", ""), tags=("section",))

                ip_purpose = ""
                if info.get("is_loopback", False):
                    ip_purpose = "本地回环地址，用于测试本地网络"
                elif info.get("is_private", False):
                    ip_purpose = "私有地址，用于内部网络"
                elif info.get("is_multicast", False):
                    ip_purpose = "组播地址，用于一对多通信"
                elif info.get("is_reserved", False):
                    ip_purpose = "保留地址，用于特殊用途"
                elif info.get("is_global", False):
                    ip_purpose = "全球可路由地址，用于公网通信"
                else:
                    ip_purpose = "未知用途"
                self.ip_info_tree.insert("", tk.END, values=("IP地址用途", ip_purpose))

                config_advice = ""
                if info.get("is_private", False):
                    config_advice = "建议使用DHCP服务器自动分配IP地址"
                else:
                    config_advice = "建议配置静态路由和防火墙规则"
                self.ip_info_tree.insert("", tk.END, values=("配置建议", config_advice))

            self.update_ipv4_history()

        except ValueError as e:
            error_result = handle_ip_subnet_error(e)
            self.show_error(_("error"), f"{_("query_failed")}: {error_result.get('error', str(e))}")
        except (tk.TclError, AttributeError, TypeError) as e:
            self.show_error(_("error"), f"{_("operation_failed")}: {str(e)}")

    def execute_range_to_cidr(self):
        """执行IP地址范围转CIDR操作"""
        try:
            # 定义默认列配置
            default_columns = [_("cidr"), _("network_address"), _("subnet_mask"), _("broadcast_address"), _("host_count")]
            ipv6_columns = [_("cidr"), _("network_address"), _("host_count")]
            
            # 1. 重置两个Treeview组件到默认状态
            # 1.1 处理IPv4结果树
            # 检查当前列配置是否与默认配置一致
            current_ipv4_cols = self.ipv4_result_tree["columns"]
            if list(current_ipv4_cols) != default_columns:
                # 如果列配置不一致，重新创建Treeview组件
                for child in self.ipv4_result_tree.winfo_children():
                    child.destroy()
                self.ipv4_result_tree.destroy()
                
                # 重新创建IPv4结果树
                self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=default_columns, show="headings")
                self.bind_treeview_right_click(self.ipv4_result_tree)
                
                # 创建滚动条并配置
                ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
                self.configure_treeview_styles(self.ipv4_result_tree)
            
            # 1.2 处理IPv6结果树
            # 检查当前列配置是否与IPv6配置一致
            current_ipv6_cols = self.ipv6_result_tree["columns"]
            if list(current_ipv6_cols) != ipv6_columns:
                # 如果列配置不一致，重新创建Treeview组件
                for child in self.ipv6_result_tree.winfo_children():
                    child.destroy()
                self.ipv6_result_tree.destroy()
                
                # 重新创建IPv6结果树
                self.ipv6_result_tree = ttk.Treeview(self.ipv6_frame, columns=ipv6_columns, show="headings")
                self.bind_treeview_right_click(self.ipv6_result_tree)
                
                # 创建滚动条并配置
                ipv6_scrollbar = ttk.Scrollbar(self.ipv6_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv6_frame, self.ipv6_result_tree, ipv6_scrollbar)
                self.configure_treeview_styles(self.ipv6_result_tree)
            
            # 2. 清空现有内容并设置表头
            # 清空IPv4结果树内容
            self.clear_tree_items(self.ipv4_result_tree)
            # 设置IPv4结果树表头
            for i, col in enumerate(default_columns):
                self.ipv4_result_tree.heading(col, text=col)
            # 设置IPv4结果树列宽
            self.ipv4_result_tree.column(default_columns[0], width=130, minwidth=130, stretch=True)
            self.ipv4_result_tree.column(default_columns[1], width=100, minwidth=100, stretch=True)
            self.ipv4_result_tree.column(default_columns[2], width=100, minwidth=100, stretch=True)
            self.ipv4_result_tree.column(default_columns[3], width=100, minwidth=100, stretch=True)
            self.ipv4_result_tree.column(default_columns[4], width=80, minwidth=80, stretch=False)
            
            # 清空IPv6结果树内容
            self.clear_tree_items(self.ipv6_result_tree)
            # 设置IPv6结果树表头
            for i, col in enumerate(ipv6_columns):
                self.ipv6_result_tree.heading(col, text=col)
            # 设置IPv6结果树列宽
            self.ipv6_result_tree.column(ipv6_columns[0], width=120, minwidth=120, stretch=True)
            self.ipv6_result_tree.column(ipv6_columns[1], width=120, minwidth=120, stretch=True)
            self.ipv6_result_tree.column(ipv6_columns[2], width=100, minwidth=100, stretch=False)

            # 获取输入的IP范围
            start_ip = self.range_start_entry.get().strip()
            end_ip = self.range_end_entry.get().strip()

            # 执行转换
            result = range_to_cidr(start_ip, end_ip)

            if isinstance(result, dict) and "error" in result:
                # 创建两列的错误表格
                error_columns = [_("network_segment"), _("error_message")]
                
                # 获取详细错误信息并清理格式
                error_msg = result["error"]
                if f"{_('ip_subnet')} {_('error')}:" in error_msg:
                    error_msg = error_msg.split(f"{_('ip_subnet')} {_('error')}:")[1].strip()
                
                # 强制重建IPv4结果树
                for child in self.ipv4_result_tree.winfo_children():
                    child.destroy()
                self.ipv4_result_tree.destroy()
                
                # 创建两列的IPv4错误表格
                self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=error_columns, show="headings")
                self.bind_treeview_right_click(self.ipv4_result_tree)
                
                ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
                self.configure_treeview_styles(self.ipv4_result_tree)
                
                # 设置IPv4表头和列宽
                self.ipv4_result_tree.heading(error_columns[0], text=error_columns[0])
                self.ipv4_result_tree.heading(error_columns[1], text=error_columns[1])
                self.ipv4_result_tree.column(error_columns[0], width=140, minwidth=140, stretch=True)
                self.ipv4_result_tree.column(error_columns[1], width=360, minwidth=360, stretch=True)
                
                # 在IPv4表格中显示错误信息
                tag = "error_row"
                self.ipv4_result_tree.insert("", tk.END, values=(
                    f"{start_ip} - {end_ip}",
                    error_msg
                ), tags=(tag,))
                self.update_table_zebra_stripes(self.ipv4_result_tree)
                
                # 同样创建IPv6错误表格
                for child in self.ipv6_result_tree.winfo_children():
                    child.destroy()
                self.ipv6_result_tree.destroy()
                
                # 创建两列的IPv6错误表格
                self.ipv6_result_tree = ttk.Treeview(self.ipv6_frame, columns=error_columns, show="headings")
                self.bind_treeview_right_click(self.ipv6_result_tree)
                
                ipv6_scrollbar = ttk.Scrollbar(self.ipv6_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv6_frame, self.ipv6_result_tree, ipv6_scrollbar)
                self.configure_treeview_styles(self.ipv6_result_tree)
                
                # 设置IPv6表头和列宽
                self.ipv6_result_tree.heading(error_columns[0], text=error_columns[0])
                self.ipv6_result_tree.heading(error_columns[1], text=error_columns[1])
                self.ipv6_result_tree.column(error_columns[0], width=140, minwidth=140, stretch=True)
                self.ipv6_result_tree.column(error_columns[1], width=360, minwidth=360, stretch=True)
                
                # 在IPv6表格中显示错误信息
                self.ipv6_result_tree.insert("", tk.END, values=(
                    f"{start_ip} - {end_ip}",
                    error_msg
                ), tags=(tag,))
                self.update_table_zebra_stripes(self.ipv6_result_tree)
                
                return

            cidr_list = result.get("cidr_list", [])

            # 检查CIDR列表是否为空
            if not cidr_list:
                # 创建两列的错误表格
                error_columns = [_("network_segment"), _("error_message")]
                
                # 强制重建IPv4结果树
                for child in self.ipv4_result_tree.winfo_children():
                    child.destroy()
                self.ipv4_result_tree.destroy()
                
                # 创建两列的错误表格
                self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=error_columns, show="headings")
                self.bind_treeview_right_click(self.ipv4_result_tree)
                
                ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
                self.configure_treeview_styles(self.ipv4_result_tree)
                
                # 设置表头和列宽
                self.ipv4_result_tree.heading(error_columns[0], text=error_columns[0])
                self.ipv4_result_tree.heading(error_columns[1], text=error_columns[1])
                self.ipv4_result_tree.column(error_columns[0], width=140, minwidth=140, stretch=True)
                self.ipv4_result_tree.column(error_columns[1], width=360, minwidth=360, stretch=True)
                
                # 在表格中显示错误信息
                tag = "error_row"
                self.ipv4_result_tree.insert("", tk.END, values=(
                    f"{start_ip} - {end_ip}",
                    _("cannot_convert_ip_range_to_cidr")
                ), tags=(tag,))
                self.update_table_zebra_stripes(self.ipv4_result_tree)
                return

            # 分离IPv4和IPv6结果
            ipv4_cidrs = []
            ipv6_cidrs = []
            
            for cidr in cidr_list:
                info = get_subnet_info(cidr)
                if info["version"] == 4:  # 使用整数比较，因为get_subnet_info返回的是整数
                    ipv4_cidrs.append(cidr)
                elif info["version"] == 6:  # 使用整数比较
                    ipv6_cidrs.append(cidr)

            # 处理IPv4结果 - 如果有结果则转置显示
            if ipv4_cidrs:
                # 转换为转置后的列结构
                transposed_columns = [_("attribute")] + ipv4_cidrs
                
                # 重新创建IPv4结果树用于转置显示
                for child in self.ipv4_result_tree.winfo_children():
                    child.destroy()
                self.ipv4_result_tree.destroy()
                
                # 创建转置后的IPv4结果树
                self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=transposed_columns, show="headings")
                self.bind_treeview_right_click(self.ipv4_result_tree)
                
                # 设置列标题
                for col in transposed_columns:
                    self.ipv4_result_tree.heading(col, text=col)
                
                # 设置列宽
                self.ipv4_result_tree.column(transposed_columns[0], width=180, minwidth=180, stretch=False)
                for col in transposed_columns[1:]:
                    self.ipv4_result_tree.column(col, width=100, stretch=True)
                
                # 创建滚动条并配置 - 使用按需显示功能
                ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
                self.configure_treeview_styles(self.ipv4_result_tree)
                
                # 定义要显示的属性列表（IPv4）
                properties = [
                    ("cidr", lambda info, cidr: cidr),
                    ("network_address", lambda info, cidr: info["network"]),
                    ("subnet_mask", lambda info, cidr: info["netmask"]),
                    ("prefix_length", lambda info, cidr: info["prefixlen"]),
                    ("broadcast_address", lambda info, cidr: info["broadcast"]),
                    ("first_host", lambda info, cidr: info["host_range_start"]),
                    ("last_host", lambda info, cidr: info["host_range_end"]),
                    ("usable_addresses", lambda info, cidr: format_large_number(info["usable_addresses"], use_scientific=True)),
                    ("total_addresses", lambda info, cidr: format_large_number(info["num_addresses"], use_scientific=True)),
                ]
                
                # 填充转置后的数据
                row_index = 0
                for prop_key, prop_func in properties:
                    row_values = [_(prop_key)]
                    for cidr in ipv4_cidrs:
                        info = get_subnet_info(cidr)
                        row_values.append(prop_func(info, cidr))
                    tag = "odd" if row_index % 2 == 0 else "even"
                    self.ipv4_result_tree.insert("", tk.END, values=row_values, tags=(tag,))
                    row_index += 1
            else:
                # 没有IPv4结果时，确保IPv4结果树保持默认配置，不被转置
                # 检查当前IPv4结果树是否是转置状态
                current_columns = self.ipv4_result_tree['columns']
                if current_columns and len(current_columns) > 5:  # 转置状态通常有更多列
                    # 重新创建默认状态的IPv4结果树
                    for child in self.ipv4_result_tree.winfo_children():
                        child.destroy()
                    self.ipv4_result_tree.destroy()
                    
                    # 重新创建默认状态的IPv4结果树
                    self.ipv4_result_tree = ttk.Treeview(self.ipv4_frame, columns=default_columns, show="headings")
                    self.bind_treeview_right_click(self.ipv4_result_tree)
                    
                    # 设置默认状态的列标题和宽度
                    for i, col in enumerate(default_columns):
                        self.ipv4_result_tree.heading(col, text=col)
                        if i == 0:  # CIDR列，需要更大宽度以适应长IP地址
                            self.ipv4_result_tree.column(col, width=130, minwidth=130, stretch=True)
                        elif i == 1:  # 网络地址列
                            self.ipv4_result_tree.column(col, width=100, minwidth=100, stretch=True)
                        elif i == 2:  # 子网掩码列
                            self.ipv4_result_tree.column(col, width=100, minwidth=100, stretch=True)
                        elif i == 3:  # 广播地址列
                            self.ipv4_result_tree.column(col, width=100, minwidth=100, stretch=True)
                        elif i == 4:  # 主机数列，可适当缩小宽度
                            self.ipv4_result_tree.column(col, width=80, minwidth=80, stretch=False)
                    
                    # 创建滚动条并配置 - 使用按需显示功能
                    ipv4_scrollbar = ttk.Scrollbar(self.ipv4_frame, orient=tk.VERTICAL)
                    self.create_scrollable_treeview(self.ipv4_frame, self.ipv4_result_tree, ipv4_scrollbar)
                    self.configure_treeview_styles(self.ipv4_result_tree)
                else:
                    # 如果已经是默认状态，确保所有列宽度正常，表头可见（与初始化和合并子网保持一致）
                    for i, col in enumerate(default_columns):
                        if i == 0:  # CIDR列，需要更大宽度以适应长IP地址
                            self.ipv4_result_tree.column(col, width=130, minwidth=130, stretch=True)
                        elif i == 1:  # 网络地址列
                            self.ipv4_result_tree.column(col, width=100, minwidth=100, stretch=True)
                        elif i == 2:  # 子网掩码列
                            self.ipv4_result_tree.column(col, width=100, minwidth=100, stretch=True)
                        elif i == 3:  # 广播地址列
                            self.ipv4_result_tree.column(col, width=100, minwidth=100, stretch=True)
                        elif i == 4:  # 主机数列，可适当缩小宽度
                            self.ipv4_result_tree.column(col, width=80, minwidth=80, stretch=False)

            # 处理IPv6结果 - 如果有结果则转置显示
            if ipv6_cidrs:
                # 转换为转置后的列结构
                transposed_columns = [_("attribute")] + ipv6_cidrs
                
                # 重新创建IPv6结果树用于转置显示
                for child in self.ipv6_result_tree.winfo_children():
                    child.destroy()
                self.ipv6_result_tree.destroy()
                
                # 创建转置后的IPv6结果树
                self.ipv6_result_tree = ttk.Treeview(self.ipv6_frame, columns=transposed_columns, show="headings")
                self.bind_treeview_right_click(self.ipv6_result_tree)
                
                # 设置列标题
                for col in transposed_columns:
                    self.ipv6_result_tree.heading(col, text=col)
                
                # 设置列宽
                self.ipv6_result_tree.column(transposed_columns[0], width=180, minwidth=180, stretch=False)
                for col in transposed_columns[1:]:
                    self.ipv6_result_tree.column(col, width=100, stretch=True)
                
                # 创建滚动条并配置 - 使用按需显示功能
                ipv6_scrollbar = ttk.Scrollbar(self.ipv6_frame, orient=tk.VERTICAL)
                self.create_scrollable_treeview(self.ipv6_frame, self.ipv6_result_tree, ipv6_scrollbar)
                self.configure_treeview_styles(self.ipv6_result_tree)
                
                # 定义要显示的属性列表（IPv6）
                properties = [
                    ("cidr", lambda info, cidr: cidr),
                    ("network_address", lambda info, cidr: info["network"]),
                    ("prefix_length", lambda info, cidr: info["prefixlen"]),
                    ("first_host", lambda info, cidr: info["host_range_start"]),
                    ("last_host", lambda info, cidr: info["host_range_end"]),
                    ("usable_addresses", lambda info, cidr: format_large_number(info["usable_addresses"], use_scientific=True)),
                    ("total_addresses", lambda info, cidr: format_large_number(info["num_addresses"], use_scientific=True)),
                ]
                
                # 填充转置后的数据
                row_index = 0
                for prop_key, prop_func in properties:
                    row_values = [_(prop_key)]
                    for cidr in ipv6_cidrs:
                        info = get_subnet_info(cidr)
                        row_values.append(prop_func(info, cidr))
                    tag = "odd" if row_index % 2 == 0 else "even"
                    self.ipv6_result_tree.insert("", tk.END, values=row_values, tags=(tag,))
                    row_index += 1
            else:
                # 没有IPv6结果时，确保IPv6结果树保持默认配置，不被转置
                # 检查当前IPv6结果树是否是转置状态
                current_columns = self.ipv6_result_tree['columns']
                if current_columns and len(current_columns) > 3:  # 转置状态通常有更多列
                    # 重新创建默认状态的IPv6结果树
                    for child in self.ipv6_result_tree.winfo_children():
                        child.destroy()
                    self.ipv6_result_tree.destroy()
                    
                    # 重新创建默认状态的IPv6结果树
                    self.ipv6_result_tree = ttk.Treeview(self.ipv6_frame, columns=ipv6_columns, show="headings")
                    self.bind_treeview_right_click(self.ipv6_result_tree)
                    
                    # 设置默认状态的列标题和宽度
                    for i, col in enumerate(ipv6_columns):
                        self.ipv6_result_tree.heading(col, text=col)
                        if i == 0:  # CIDR列
                            self.ipv6_result_tree.column(col, width=120, minwidth=120, stretch=True)
                        elif i == 1:  # 网络地址列
                            self.ipv6_result_tree.column(col, width=120, minwidth=120, stretch=True)
                        elif i == 2:  # 主机数列
                            self.ipv6_result_tree.column(col, width=180, minwidth=180, stretch=True)
                    
                    # 创建滚动条并配置 - 使用按需显示功能
                    ipv6_scrollbar = ttk.Scrollbar(self.ipv6_frame, orient=tk.VERTICAL)
                    self.create_scrollable_treeview(self.ipv6_frame, self.ipv6_result_tree, ipv6_scrollbar)
                    self.configure_treeview_styles(self.ipv6_result_tree)

            self.update_range_start_history()
            self.update_range_end_history()

        except ValueError as e:
            error_result = handle_ip_subnet_error(e)
            self.show_error(_("error"), f"{_("conversion_failed")}: {error_result.get('error', str(e))}")
        except (tk.TclError, AttributeError, TypeError) as e:
            self.show_error(_("error"), f"{_("operation_failed")}: {str(e)}")

    def _process_overlap_detection(self, subnets, ip_version_label, row_index):
        """
        处理子网重叠检测结果并更新UI
        
        参数:
        subnets: 子网列表
        ip_version_label: IP版本标签
        row_index: 当前行索引
        
        返回:
        更新后的行索引
        """
        result = check_subnet_overlap(subnets)
        if isinstance(result, dict) and "error" in result:
            tag = "error_row"
            self.overlap_result_tree.insert("", tk.END, values=(_("error"), result["error"]), tags=(tag,))
            row_index += 1
        else:
            overlaps = result.get("overlaps", [])
            if not overlaps:
                tag = "odd" if row_index % 2 == 0 else "even"
                self.overlap_result_tree.insert("", tk.END, values=(_("no_overlap"), f"{ip_version_label}: {_('no_subnet_overlap_detected')}"), tags=(tag,))
                row_index += 1
            else:
                for overlap in overlaps:
                    tag = "odd" if row_index % 2 == 0 else "even"
                    status = _("overlap")
                    # 直接使用overlap['type']作为描述，因为它已经是完整的重叠关系描述
                    description = overlap['type']
                    self.overlap_result_tree.insert("", tk.END, values=(status, description), tags=(tag,))
                    row_index += 1
        return row_index

    def execute_check_overlap(self):
        """执行子网重叠检测操作，支持同时检测IPv4和IPv6子网"""
        try:
            for item in self.overlap_result_tree.get_children():
                self.overlap_result_tree.delete(item)

            # 获取输入的子网列表
            subnets_text = self.overlap_text.get(1.0, tk.END).strip()
            if not subnets_text:
                self.show_info(_("hint"), _("enter_subnet_list"))
                return

            # 解析子网列表
            subnets = [line.strip() for line in subnets_text.splitlines() if line.strip()]

            # 按IP版本分类子网，用于分别检测IPv4和IPv6子网重叠
            ipv4_subnets = []
            ipv6_subnets = []
            invalid_subnets = []
            
            # 验证子网格式并分类
            for subnet in subnets:
                try:
                    network = ipaddress.ip_network(subnet, strict=False)
                    if isinstance(network, ipaddress.IPv4Network):
                        ipv4_subnets.append(subnet)
                    else:
                        ipv6_subnets.append(subnet)
                except ValueError:
                    invalid_subnets.append(subnet)

            # 显示无效子网信息
            row_index = 0
            if invalid_subnets:
                for subnet in invalid_subnets:
                    tag = "error_row"
                    self.overlap_result_tree.insert("", tk.END, values=(_("error"), f"{subnet}: {_('invalid_cidr_format')}"), tags=(tag,))
                    row_index += 1

            # 检测IPv4子网重叠
            if ipv4_subnets:
                row_index = self._process_overlap_detection(ipv4_subnets, _("ipv4"), row_index)

            # 检测IPv6子网重叠
            if ipv6_subnets:
                row_index = self._process_overlap_detection(ipv6_subnets, _("ipv6"), row_index)

            # 如果没有任何子网（除了无效的），显示提示信息
            if not ipv4_subnets and not ipv6_subnets and not invalid_subnets:
                tag = "odd" if row_index % 2 == 0 else "even"
                self.overlap_result_tree.insert("", tk.END, values=(_("no_overlap"), _("no_subnet_overlap_detected")), tags=(tag,))
            
            # 重新运行实时验证，保持标记与当前输入一致
            self.validate_overlap_text()
            
            self.auto_resize_columns(self.overlap_result_tree)
            # 更新斑马条纹，确保错误行同时显示红色和斑马条纹效果
            self.update_table_zebra_stripes(self.overlap_result_tree)

            # 持久化重叠检测列表文本到数据库
            self._persist_text_data(HistorySQLite.CATEGORY_OVERLAP_DETECTION, self.overlap_text)

        except (ValueError, tk.TclError, AttributeError, TypeError) as e:
            self.show_error(_('error'), f'{_('execute_subnet_overlap_detection_failed')}: {str(e)}')

    def _update_history(self, entry, history_list, value=None, max_items=10):
        self.history_repo.update_history(entry, history_list, value, max_items)

    def update_ipv4_history(self, event=None):
        """更新IPv4地址查询历史记录"""
        self._update_history(self.ip_info_entry, self.ipv4_history)

    def update_ipv6_history(self, event=None):
        """更新IPv6地址查询历史记录"""
        self._update_history(self.ipv6_info_entry, self.ipv6_history)

    def update_range_start_history(self, event=None):
        """更新IP范围起始地址历史记录"""
        self._update_history(self.range_start_entry, self.range_start_history)

    def update_range_end_history(self, event=None):
        """更新IP范围结束地址历史记录"""
        self._update_history(self.range_end_entry, self.range_end_history)

    def on_subnet_mask_change(self, event):
        """当子网掩码改变时，更新CIDR值"""
        selected_mask = self.ip_mask_var.get()
        if selected_mask in self.subnet_mask_cidr_map:
            cidr = self.subnet_mask_cidr_map[selected_mask]
            self.ip_cidr_var.set(cidr)

    def on_cidr_change(self, event):
        """当CIDR改变时，更新子网掩码值"""
        selected_cidr = self.ip_cidr_var.get()
        if selected_cidr in self.cidr_subnet_mask_map:
            subnet_mask = self.cidr_subnet_mask_map[selected_cidr]
            self.ip_mask_var.set(subnet_mask)

    def toggle_test_info_bar(self, event=None):
        """打开功能调试对话框（彩蛋功能）
        快捷键：Ctrl+Shift+I
        """
        # 检查调试面板是否已经存在
        if hasattr(self, 'test_dialog') and self.test_dialog is not None:
            try:
                # 尝试将对话框显示在最上层并设置焦点
                self.test_dialog.dialog.lift()
                self.test_dialog.dialog.focus_force()
                return
            except tk.TclError:
                # 如果对话框已被销毁，忽略错误并创建新对话框
                self.test_dialog = None

        # 计算对话框尺寸
        dialog_width = 550  # 加大窗体宽度，适应不同语言
        dialog_height = 600  # 增加对话框高度，确保所有控件能完整显示
        
        # 创建功能调试对话框，使用统一的create_dialog方法
        # 创建功能调试对话框
        self.test_dialog = ComplexDialog(self.root, _("function_debug"), dialog_width, dialog_height, resizable=False, modal=False)
        
        # 绑定关闭事件，确保对话框关闭时更新状态
        self.test_dialog.dialog.protocol("WM_DELETE_WINDOW", self.close_test_dialog)

        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()

        # 创建对话框内容框架
        content_frame = ttk.Frame(self.test_dialog.content_frame, padding="15")
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 使用grid布局管理器来精确控制各个组件的位置
        content_frame.grid_rowconfigure(0, weight=0)  # 标题行不扩展
        content_frame.grid_rowconfigure(1, weight=0)  # 说明行不扩展
        content_frame.grid_rowconfigure(2, weight=0)  # 按钮区行不扩展
        content_frame.grid_rowconfigure(3, weight=0)  # 底部区域不扩展
        content_frame.grid_rowconfigure(4, weight=0)  # 底部区域不扩展
        content_frame.grid_rowconfigure(5, weight=1)  # 关闭按钮行扩展，使其靠底部
        content_frame.grid_columnconfigure(0, weight=1)  # 左列（主题切换/窗口锁定）
        content_frame.grid_columnconfigure(1, weight=1)  # 右列（标签次序）

        # 添加标题标签
        title_label = ttk.Label(content_frame, text=_("function_debug_panel"), font=(font_family, 12, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 15))

        # 添加说明标签
        desc_label = ttk.Label(content_frame, text=_("test_info_display_effect"))
        desc_label.grid(row=1, column=0, columnspan=2, pady=(0, 15))

        # 创建按钮框架（使用grid布局实现3x2矩阵）
        button_frame = ttk.Frame(content_frame)
        button_frame.grid(row=2, column=0, columnspan=2, sticky=tk.N)  # 跨两列，顶部对齐

        # 按钮样式
        button_style = "TButton"
        button_width = 25  # 恢复原来的按钮宽度
        original_button_width = 15  # 应用主题和关闭按钮保持原来宽度
        
        # 创建复选框样式
        checkbutton_style = ttk.Style()
        checkbutton_style.configure("Custom.TCheckbutton", font=(font_family, font_size))

        # 第一行按钮
        success_btn = ttk.Button(
            button_frame,
            text=_("test_success_message"),
            width=button_width,
            style=button_style,
            command=lambda: self.show_result(_("test_success_content"), error=False),
        )
        success_btn.grid(row=0, column=0, padx=5, pady=5)

        success_btn.focus_force()

        error_btn = ttk.Button(
            button_frame,
            text=_("test_error_message"),
            width=button_width,
            style=button_style,
            command=lambda: self.show_result(_("test_error_content"), error=True, keep_data=True),
        )
        error_btn.grid(row=0, column=1, padx=5, pady=5)

        # 第二行按钮
        long_text = _("test_long_text_content") * 5
        long_text_btn = ttk.Button(
            button_frame,
            text=_("test_long_text_message"),
            width=button_width,
            style=button_style,
            command=lambda: self.show_result(long_text, error=False),
        )
        long_text_btn.grid(row=1, column=0, padx=5, pady=5)

        # 中英文混排长文本测试按钮
        mixed_text = (
            (_("test_mixed_text_content") + "This is a long text with mixed " + _("db_language") + " and English characters. "
             ) * 5
        )
        mixed_text_btn = ttk.Button(
            button_frame,
            text=_("test_mixed_language"),
            width=button_width,
            style=button_style,
            command=lambda: self.show_result(mixed_text, error=False),
        )
        mixed_text_btn.grid(row=1, column=1, padx=5, pady=5)

        # 添加第三行按钮：隐藏信息栏和清空结果
        hide_info_btn = ttk.Button(
            button_frame, text=_(
            "hide_info_bar"), width=button_width, style=button_style, command=self.hide_info_bar
        )
        hide_info_btn.grid(row=2, column=0, padx=5, pady=5)

        clear_result_btn = ttk.Button(
            button_frame, text=_("clear_subnet_split_and_planning"), width=button_width, style=button_style, command=self.clear_result
        )
        clear_result_btn.grid(row=2, column=1, padx=5, pady=5)

        # 添加一键导出按钮
        one_click_export_btn = ttk.Button(
            button_frame, text=_("one_click_export"), width=button_width, style=button_style, command=self.one_click_export
        )
        one_click_export_btn.grid(row=3, column=0, padx=5, pady=5)
        
        # 添加一键PDF按钮
        one_click_pdf_btn = ttk.Button(
            button_frame, text=_("one_click_pdf"), width=button_width, style=button_style, command=self.one_click_pdf
        )
        one_click_pdf_btn.grid(row=3, column=1, padx=5, pady=5)

        # 主题切换部分（放在左下方）
        theme_frame = ttk.LabelFrame(content_frame, text=_("theme_switch"), padding="10")
        theme_frame.grid(row=3, column=0, sticky=tk.EW, pady=(15, 10), padx=(0, 10))

        # 配置主题切换框架的列
        theme_frame.grid_columnconfigure(0, weight=0)  # 标签列
        theme_frame.grid_columnconfigure(1, weight=1)  # 下拉列表列
        theme_frame.grid_columnconfigure(2, weight=0)  # 按钮列

        # 主题选择标签
        theme_label = ttk.Label(theme_frame, text=_("select_theme") + ":")
        theme_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=5)

        # 获取系统可用的内置主题列表
        theme_list = self.style.theme_names()

        # 创建主题选择下拉列表
        self.theme_var = tk.StringVar(value=self.style.theme_use())  # 设置默认主题为当前使用的主题

        # 创建下拉列表控件
        theme_combobox = ttk.Combobox(theme_frame, textvariable=self.theme_var, values=theme_list, state="readonly")
        theme_combobox.grid(row=0, column=1, sticky=tk.EW, pady=5)

        # 主题切换函数
        def switch_theme():
            new_theme = self.theme_var.get()
            try:
                # 使用系统内置主题切换，彻底移除sv-ttk，解决黑色底色问题
                self.style.theme_use(new_theme)
                
                # theme_use() 会重置所有样式，需要重新应用全局字体设置
                update_styles()
                
                # 重新配置Treeview样式，确保在新主题下表格线仍然可见
                tree_names = [
                    'split_tree',
                    'remaining_tree',
                    'allocated_tree',
                    'planning_remaining_tree',
                    'pool_tree',
                    'requirements_tree',
                    'history_tree',
                ]
                for tree_name in tree_names:
                    if hasattr(self, tree_name):
                        tree = getattr(self, tree_name)
                        # 确保tree不是None
                        if tree is None:
                            continue
                        # 检查split_tree是否需要包含特殊标签
                        include_special = tree_name == 'split_tree'
                        self.configure_treeview_styles(tree, include_special)

                # 重新配置信息栏标签样式，确保错误信息颜色正确
                info_bar_font_size = get_info_bar_font_size()
                base_info_label_style = {"borderwidth": 0, "font": (font_family, info_bar_font_size), "relief": "flat"}
                self.style.configure("Success.TLabel", foreground="#424242", **base_info_label_style)
                self.style.configure("Error.TLabel", foreground="#c62828", **base_info_label_style)
                self.style.configure("Info.TLabel", foreground="#424242", **base_info_label_style)

                # 重新配置信息栏框架样式 - 所有信息栏框架使用相同的基础样式
                info_bar_frame_style = {"borderwidth": 0, "relief": "flat"}
                for frame_style in ["InfoBar.TFrame", "SuccessInfoBar.TFrame", "ErrorInfoBar.TFrame", "InfoInfoBar.TFrame"]:
                    self.style.configure(frame_style, **info_bar_frame_style)
                
                # 更新Text组件的背景色，确保跟随主题
                if hasattr(self, 'info_label'):
                    bg_color = self.style.lookup("TFrame", "background")
                    self.info_label.configure(background=bg_color)
                # 更新信息栏关闭按钮的背景色，确保跟随主题
                if hasattr(self, 'info_close_btn'):
                    bg_color = self.style.lookup("TFrame", "background")
                    self.info_close_btn.configure(bg=bg_color)
                
                # 刷新二级notebook的内容区域背景色
                self._refresh_secondary_notebooks()
                
                # 刷新主窗口界面，确保所有控件颜色正确更新
                self.root.update_idletasks()
                self.root.update()
                
                # 保存主题配置到文件
                from config_manager import get_config
                config = get_config()
                config.set_ui_theme(new_theme)
            except (tk.TclError, AttributeError) as e:
                print(f"{_('theme_switch_error')}: {e}")
                # 出错时恢复到默认主题
                self.style.theme_use("vista")

        # 创建应用主题按钮 - 使用原来宽度
        theme_switch_btn = ttk.Button(
            theme_frame, text=_('apply_theme'), width=original_button_width, style=button_style, command=switch_theme
        )
        theme_switch_btn.grid(row=0, column=2, padx=(10, 0), pady=5)

        # 窗口横向锁定控制部分（放在主题切换下方，左列）
        lock_frame = ttk.LabelFrame(content_frame, text=_("window_lock"), padding="10")
        lock_frame.grid(row=4, column=0, sticky=tk.EW, pady=(10, 10), padx=(0, 10))

        # 配置锁定框架的列
        lock_frame.grid_columnconfigure(0, weight=1)

        # 窗口横向锁定复选框 - 检查窗口的实际可调整大小状态
        width_locked = not self.root.resizable()[0]  # 获取窗口宽度是否可调整
        self.width_lock_var = tk.BooleanVar(value=width_locked)
        width_lock_cb = ttk.Checkbutton(
            lock_frame,
            text=_('lock_window_width'),
            variable=self.width_lock_var,
            command=self.toggle_width_lock,
            style="Custom.TCheckbutton"
        )
        width_lock_cb.grid(row=0, column=0, sticky=tk.W, pady=5)

        # 顶级标签次序调整部分（放在右侧，跨两行对齐左侧两个面板）
        tab_order_frame = ttk.LabelFrame(content_frame, text=_("top_level_tab_order"), padding="10")
        tab_order_frame.grid(row=3, column=1, rowspan=2, sticky=tk.NSEW, pady=(15, 10))

        # 配置标签次序调整框架的行列
        tab_order_frame.grid_columnconfigure(0, weight=1)  # 标签列表列
        tab_order_frame.grid_columnconfigure(1, weight=0)  # 操作按钮列
        tab_order_frame.grid_rowconfigure(0, weight=1)    # 第一行扩展填充空间

        # 创建标签列表（使用Treeview替代Listbox以更好地控制选中样式）
        self.tab_order_tree = ttk.Treeview(
            tab_order_frame,
            columns=('name'),
            show='',
            height=0
        )
        
        # 配置列
        self.tab_order_tree.column('name', anchor='w', width=150)
        
        # 创建自定义样式来移除选中时的下划线
        tree_style = ttk.Style()
        tree_style.configure('TabOrder.Treeview', 
                            rowheight=22,
                            foreground='black',
                            fieldbackground='white')
        tree_style.map('TabOrder.Treeview',
                       background=[('selected', '#0078D7')],
                       foreground=[('selected', '#ffffff')],
                       fieldbackground=[('selected', '#0078D7'), ('!selected', 'white')])
        # 移除选中时的焦点边框
        tree_style.configure('TabOrder.Treeview', highlightthickness=0)
        tree_style.map('TabOrder.Treeview',
                       highlightcolor=[('focus', '#0078D7')],
                       highlightbackground=[('focus', '#0078D7')])
        
        self.tab_order_tree.config(style='TabOrder.Treeview')
        self.tab_order_tree.grid(row=0, column=0, sticky="nsew")

        # 填充标签列表
        self.update_tab_order_listbox()

        # 添加显示启动画面按钮
        show_splash_btn = ttk.Button(
            tab_order_frame, text=_('show_splash'), width=20, style=button_style, command=self.show_splash
        )
        show_splash_btn.grid(row=1, column=0, columnspan=2, sticky=tk.EW, padx=5, pady=10)

        # 创建按钮框架，垂直排列操作按钮
        button_frame = ttk.Frame(tab_order_frame)
        button_frame.grid(row=0, column=1, sticky=tk.N, padx=(5, 0), pady=(0, 2))

        # 上移按钮
        move_up_btn = ttk.Button(
            button_frame,
            text="▲",
            width=3,
            command=self.move_selected_tab_up
        )
        move_up_btn.grid(row=0, column=0, pady=(0, 2))

        # 下移按钮
        move_down_btn = ttk.Button(
            button_frame,
            text="▼",
            width=3,
            command=self.move_selected_tab_down
        )
        move_down_btn.grid(row=1, column=0, pady=2)

        # 关闭按钮框架
        close_frame = ttk.Frame(content_frame)
        close_frame.grid(row=5, column=0, columnspan=2, sticky=tk.SE, pady=(15, 0), padx=(0, 0))

        # 添加关闭按钮到右下角 - 使用原来宽度
        close_btn = ttk.Button(
            close_frame, text=_('close'), width=original_button_width, style=button_style, command=self.close_test_dialog
        )
        close_btn.pack(side=tk.RIGHT)

        # 显示对话框
        self.test_dialog.show()

    def _refresh_secondary_notebooks(self):
        """刷新所有二级notebook的内容区域背景色，确保主题切换后颜色正确显示"""
        # 刷新子网分割页面的二级notebook
        if hasattr(self, 'notebook') and self.notebook is not None:
            self._refresh_single_notebook(self.notebook)
        
        # 刷新规划页面的二级notebook
        if hasattr(self, 'planning_notebook') and self.planning_notebook is not None:
            self._refresh_single_notebook(self.planning_notebook)
        
        # 刷新地址管理页面的二级notebook
        if hasattr(self, 'ipam_notebook') and self.ipam_notebook is not None:
            self._refresh_single_notebook(self.ipam_notebook)
        
        # 刷新高级工具页面的二级notebook
        if hasattr(self, 'advanced_notebook') and self.advanced_notebook is not None:
            self._refresh_single_notebook(self.advanced_notebook)

    def _refresh_single_notebook(self, notebook):
        """刷新单个notebook的样式"""
        if not hasattr(notebook, 'tabs') or not hasattr(notebook, 'active_tab'):
            return
        
        # 如果有活动标签，重新选中以刷新样式
        if notebook.active_tab is not None and 0 <= notebook.active_tab < len(notebook.tabs):
            notebook.select_tab(notebook.active_tab)

    def close_test_dialog(self):
        """关闭功能调试对话框并更新状态"""
        if hasattr(self, 'test_dialog') and self.test_dialog is not None:
            try:
                self.test_dialog.destroy()
            finally:
                # 确保无论如何都将test_dialog设置为None
                self.test_dialog = None

    def show_splash(self):
        """显示启动画面，用于调试
        启动画面会一直显示，直到用户点击画面才退出
        """
        # 直接使用已定义的SplashScreen类，不需要导入
        splash = SplashScreen(self.root)
        
        # 添加点击事件，点击启动画面时关闭
        def close_splash_on_click(event):
            try:
                if splash.splash:
                    splash.splash.destroy()
            except:
                pass
        
        # 绑定点击事件到启动画面窗口
        if splash.splash:
            splash.splash.bind('<Button-1>', close_splash_on_click)
            
            # 确保启动画面在最上层
            splash.splash.lift()
            splash.splash.attributes('-topmost', True)

    def open_hidden_info_dialog(self, event=None):
        """打开隐藏信息管理对话框（快捷键：Ctrl+Shift+H）

        需要先在IP地址列表中选中一个IP地址
        """
        if not hasattr(self, 'ipam_ip_tree'):
            return

        selected_items = self.ipam_ip_tree.selection()
        if not selected_items:
            self.show_info(_("hint"), _("please_select_ip_first"))
            return

        item = selected_items[0]
        values = self.ipam_ip_tree.item(item, 'values')
        if not values:
            return

        ip_address = values[0]
        ip_record_id = self._get_db_record_id(item)
        if not ip_record_id:
            self.show_info(_("hint"), _("cannot_access_hidden_info"))
            return
        self._show_hidden_info_dialog(ip_address, ip_record_id)

    def _show_hidden_info_dialog(self, ip_address, ip_record_id):
        """显示指定IP地址的隐藏信息管理对话框

        Args:
            ip_address: IP地址（仅用于显示）
            ip_record_id: IP记录ID（关联到ip_addresses表的主键）
        """
        dialog = self.create_dialog(
            _('hidden_info_for_ip').format(ip=ip_address),
            720, 480,
            resizable=True,
            modal=True
        )

        self._hidden_info_dialog = dialog
        self._hidden_info_ip = ip_address
        self._hidden_info_ip_record_id = ip_record_id
        self._hidden_info_passwords_visible = False
        self._hidden_info_raw_data = []

        font_family, font_size = get_current_font_settings()

        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        tree_frame = ttk.Frame(main_frame)
        tree_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        columns = ('url', 'username', 'password', 'notes')
        self._hidden_info_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            height=10,
            selectmode='extended'
        )

        self._hidden_info_tree.heading('url', text=_('access_url'))
        self._hidden_info_tree.heading('username', text=_('username'))
        self._hidden_info_tree.heading('password', text=_('password'))
        self._hidden_info_tree.heading('notes', text=_('notes'))

        self._hidden_info_tree.column('url', width=220, minwidth=150, stretch=True)
        self._hidden_info_tree.column('username', width=120, minwidth=80, stretch=True)
        self._hidden_info_tree.column('password', width=120, minwidth=80, stretch=True)
        self._hidden_info_tree.column('notes', width=200, minwidth=100, stretch=True)

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        self.create_scrollable_treeview_with_grid(tree_frame, self._hidden_info_tree, scrollbar)

        self.configure_treeview_styles(self._hidden_info_tree)
        
        # 绑定右键复制功能
        self.bind_treeview_right_click(self._hidden_info_tree)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, sticky="ew")

        left_btn_frame = ttk.Frame(button_frame)
        left_btn_frame.pack(side=tk.LEFT)

        btn_font = (font_family, font_size)
        add_btn = ttk.Button(left_btn_frame, text=_('add_record'), command=lambda: self._add_hidden_record(ip_record_id, dialog), width=12)
        add_btn.pack(side=tk.LEFT, padx=(0, 5))

        edit_btn = ttk.Button(left_btn_frame, text=_('edit_record'), command=lambda: self._edit_hidden_record(ip_record_id, dialog), width=12)
        edit_btn.pack(side=tk.LEFT, padx=(0, 5))

        delete_btn = ttk.Button(left_btn_frame, text=_('delete_record'), command=lambda: self._delete_hidden_record(ip_record_id, dialog), width=12)
        delete_btn.pack(side=tk.LEFT, padx=(0, 5))

        right_btn_frame = ttk.Frame(button_frame)
        right_btn_frame.pack(side=tk.RIGHT)

        self._toggle_pwd_btn = ttk.Button(
            right_btn_frame,
            text=_('show_password'),
            command=lambda: self._toggle_password_visibility(ip_address),
            width=12
        )
        self._toggle_pwd_btn.pack(side=tk.LEFT, padx=(0, 5))

        close_btn = ttk.Button(right_btn_frame, text=_('close'), command=dialog.destroy, width=12)
        close_btn.pack(side=tk.LEFT)

        self._refresh_hidden_info_tree(ip_record_id)

        self._hidden_info_tree.bind('<Double-1>', lambda e: self._edit_hidden_record(ip_record_id, dialog))

        # 绑定ESC键关闭对话框
        dialog.bind('<Escape>', lambda e: dialog.destroy())

        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)

    def _refresh_hidden_info_tree(self, ip_record_id):
        """刷新隐藏信息表格数据

        Args:
            ip_record_id: IP记录ID
        """
        for item in self._hidden_info_tree.get_children():
            self._hidden_info_tree.delete(item)

        records = self.ipam_repo.get_hidden_info(ip_record_id)
        self._hidden_info_raw_data = records

        crypto = get_crypto_service()
        for record in records:
            password = record.get('password', '')
            if self._hidden_info_passwords_visible:
                display_pwd = password
            else:
                display_pwd = crypto.mask_password(password)

            self._hidden_info_tree.insert('', tk.END, iid=str(record['id']), values=(
                record.get('url', ''),
                record.get('username', ''),
                display_pwd,
                record.get('notes', '')
            ))

        self.update_table_zebra_stripes(self._hidden_info_tree)

    def _toggle_password_visibility(self, ip_address):
        """切换密码字段的显示/隐藏状态

        Args:
            ip_address: IP地址
        """
        self._hidden_info_passwords_visible = not self._hidden_info_passwords_visible
        if self._hidden_info_passwords_visible:
            self._toggle_pwd_btn.configure(text=_('hide_password'))
        else:
            self._toggle_pwd_btn.configure(text=_('show_password'))
        self._refresh_hidden_info_tree(self._hidden_info_ip_record_id)

    def _add_hidden_record(self, ip_record_id, parent_dialog):
        """添加隐藏信息记录

        Args:
            ip_record_id: IP记录ID
            parent_dialog: 父对话框
        """
        self._show_hidden_record_edit_dialog(ip_record_id, parent_dialog, None)

    def _edit_hidden_record(self, ip_record_id, parent_dialog):
        """编辑隐藏信息记录

        Args:
            ip_record_id: IP记录ID
            parent_dialog: 父对话框
        """
        selected = self._hidden_info_tree.selection()
        if not selected:
            self.show_info(_("hint"), _("please_select_record"))
            return

        record_id = int(selected[0])
        record = None
        for r in self._hidden_info_raw_data:
            if r['id'] == record_id:
                record = r
                break

        if not record:
            return

        self._show_hidden_record_edit_dialog(ip_record_id, parent_dialog, record)

    def _delete_hidden_record(self, ip_record_id, parent_dialog):
        """删除隐藏信息记录（支持批量删除）

        Args:
            ip_record_id: IP记录ID
            parent_dialog: 父对话框
        """
        selected = self._hidden_info_tree.selection()
        if not selected:
            self.show_info(_("hint"), _("please_select_record"))
            return

        if len(selected) > 1:
            confirm_msg = _("confirm_delete_hidden_records").format(count=len(selected))
        else:
            confirm_msg = _("confirm_delete_hidden_record")

        if not self.show_custom_confirm(_("hint"), confirm_msg):
            return

        success_count = 0
        for item in selected:
            record_id = int(item)
            success, msg = self.ipam_repo.delete_hidden_info(record_id)
            if success:
                success_count += 1

        if success_count > 0:
            self._refresh_hidden_info_tree(ip_record_id)
            if success_count == len(selected):
                self.show_info(_("success"), _("delete_success"))
            else:
                self.show_info(_("success"), _("partial_delete_success").format(success=success_count, total=len(selected)))
        else:
            self.show_error(_("operation_failed"), msg)

    def _show_hidden_record_edit_dialog(self, ip_record_id, parent_dialog, record=None):
        """显示隐藏信息记录的编辑对话框

        Args:
            ip_record_id: IP记录ID
            parent_dialog: 父对话框
            record: 要编辑的记录，为None时表示添加新记录
        """
        is_edit = record is not None
        title = _('edit_record') if is_edit else _('add_record')

        edit_dialog = self.create_dialog(title, 420, 360, resizable=False, modal=True, parent=parent_dialog)

        font_family, font_size = get_current_font_settings()

        frame = ttk.Frame(edit_dialog, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)

        # 创建两个子框架：一个用于表单字段，一个用于按钮
        content_frame = ttk.Frame(frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(side=tk.BOTTOM, anchor=tk.SE, pady=(10, 5))

        # 配置内容框架的列布局
        content_frame.grid_columnconfigure(0, weight=0)
        content_frame.grid_columnconfigure(1, weight=1)

        label_font = (font_family, font_size)

        fields = [
            (_('access_url'), 'url'),
            (_('username'), 'username'),
            (_('password'), 'password'),
            (_('notes'), 'notes'),
        ]

        entries = {}
        for i, (label_text, field_key) in enumerate(fields):
            # 添加标签（带冒号，与应用风格一致）
            label = ttk.Label(content_frame, text=label_text + ':', font=label_font, anchor='e')
            label.grid(row=i, column=0, padx=(0, 15), pady=5, sticky='e')

            if field_key == 'notes':
                # 备注字段使用8行的Text组件，标签顶部对齐
                label.grid(row=i, column=0, padx=(0, 15), pady=5, sticky='ne')
                notes_border_frame = tk.Frame(content_frame, highlightbackground="#a9a9a9",
                                              highlightcolor="#a9a9a9", highlightthickness=1, bd=0)
                notes_entry = tk.Text(notes_border_frame, height=8, font=(font_family, font_size),
                                      bd=0, relief="flat", highlightthickness=0, wrap=tk.WORD)
                notes_entry.pack(fill="both", expand=True, padx=(1, 0), pady=1)
                notes_border_frame.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                entries[field_key] = notes_entry
            else:
                # 使用应用风格的带边框Entry
                border_frame, entry = create_bordered_entry(content_frame, width=30)
                border_frame.grid(row=i, column=1, padx=5, pady=5, sticky='ew')
                entries[field_key] = entry

            if is_edit:
                if field_key == 'password':
                    entries[field_key].insert(0, record.get('password', ''))
                elif field_key == 'notes':
                    entries[field_key].insert(tk.END, record.get('notes', ''))
                else:
                    entries[field_key].insert(0, record.get(field_key, ''))

        def save_record():
            url = entries['url'].get().strip()
            username = entries['username'].get().strip()
            password = entries['password'].get().strip()
            notes = entries['notes'].get('1.0', tk.END).strip()

            if is_edit:
                success, msg = self.ipam_repo.update_hidden_info(
                    record['id'], url, username, password, notes
                )
            else:
                success, msg, _ = self.ipam_repo.add_hidden_info(
                    ip_record_id, url, username, password, notes
                )

            if success:
                self._refresh_hidden_info_tree(ip_record_id)
                edit_dialog.destroy()
            else:
                import tkinter.messagebox as messagebox
                messagebox.showerror("操作失败", msg)

        cancel_btn = ttk.Button(btn_frame, text=_('cancel'), command=edit_dialog.destroy, width=10)
        cancel_btn.pack(side=tk.RIGHT, padx=5)

        save_btn = ttk.Button(btn_frame, text=_('save'), command=save_record, width=10)
        save_btn.pack(side=tk.RIGHT, padx=5)

        # 绑定ESC键关闭对话框
        edit_dialog.bind('<Escape>', lambda e: edit_dialog.destroy())

        entries['url'].focus_set()

    def update_tab_order_listbox(self):
        """更新调试面板中的标签列表"""
        if not hasattr(self, 'tab_order_tree'):
            return
        
        # 清空列表
        for item in self.tab_order_tree.get_children():
            self.tab_order_tree.delete(item)
        
        # 填充标签列表
        if hasattr(self, 'top_level_notebook') and self.top_level_notebook:
            for i, tab in enumerate(self.top_level_notebook.tabs):
                # 添加序号前缀
                self.tab_order_tree.insert('', 'end', values=(f"{i + 1}. {tab['label']}",))

    def move_selected_tab_up(self):
        """将选中的标签上移一位"""
        if not hasattr(self, 'tab_order_tree'):
            return
        
        selected_items = self.tab_order_tree.selection()
        if not selected_items:
            return
        
        # 获取选中项的索引
        all_items = self.tab_order_tree.get_children()
        tab_index = all_items.index(selected_items[0])
        
        if self.top_level_notebook.move_tab_up(tab_index):
            self.update_tab_order_listbox()
            # 重新选中移动后的位置
            new_items = self.tab_order_tree.get_children()
            if tab_index - 1 >= 0 and tab_index - 1 < len(new_items):
                self.tab_order_tree.selection_set(new_items[tab_index - 1])
            # 保存到配置文件
            self.top_level_notebook.save_tab_order()

    def move_selected_tab_down(self):
        """将选中的标签下移一位"""
        if not hasattr(self, 'tab_order_tree'):
            return
        
        selected_items = self.tab_order_tree.selection()
        if not selected_items:
            return
        
        # 获取选中项的索引
        all_items = self.tab_order_tree.get_children()
        tab_index = all_items.index(selected_items[0])
        
        if self.top_level_notebook.move_tab_down(tab_index):
            self.update_tab_order_listbox()
            # 重新选中移动后的位置
            new_items = self.tab_order_tree.get_children()
            if tab_index + 1 < len(new_items):
                self.tab_order_tree.selection_set(new_items[tab_index + 1])
            # 保存到配置文件
            self.top_level_notebook.save_tab_order()

    def toggle_width_lock(self):
        """切换窗口横向锁定状态"""
        self.width_locked = self.width_lock_var.get()

        if self.width_locked:
            self.root.resizable(width=False, height=True)
            self.root.minsize(850, 750)
            self.root.maxsize(1100, 10000)
        else:
            self.root.resizable(width=True, height=True)
            self.root.minsize(850, 750)
            self.root.maxsize(10000, 10000)

    def show_result(self, text, error=False, keep_data=False):
        """显示结果

        Args:
            text: 要显示的文本
            error: 是否为错误信息
            keep_data: 是否保留数据
        """
        # 重置暂停状态，新消息来时恢复自动消失
        self.info_auto_hide_paused = False
        
        # 立即取消所有可能的自动隐藏定时器
        if self.info_auto_hide_id:
            self.root.after_cancel(self.info_auto_hide_id)
            self.info_auto_hide_id = None
        
        # 只有在不保留数据且显示错误信息时才清空表格
        if not keep_data and error:
            self.clear_result()

        # 根据信息类型设置样式和图标，使用带框风格，保持一致
        if error:
            label_style = "Error.TLabel"
            frame_style = "ErrorInfoBar.TFrame"
            icon = "❎ "  # 使用明确的带框叉号 (U+274E)
        else:
            label_style = "Success.TLabel"
            frame_style = "SuccessInfoBar.TFrame"
            icon = "✅ "  # 使用带框对勾 (U+2705)，与带框叉风格一致

        # 更新信息标签和框架样式

        # 获取信息栏的实际宽度
        # 确保使用一致的宽度计算逻辑，无论是第一次还是第二次显示
        # 始终使用主窗口宽度作为参考，确保第一次和第二次显示时截断一致
        main_window_width = self.root.winfo_width()
        # 使用主窗口宽度的88%作为信息栏宽度，放大截断位置
        info_bar_width = int(main_window_width * 1)
        # 确保不小于原始的最小宽度
        info_bar_width = max(info_bar_width, self.min_info_bar_width)

        # 确保info_bar_frame已经添加到父容器中
        # 使用place布局时winfo_manager()返回"place"，所以用not判断
        if not self.info_bar_frame.winfo_manager():
            # 先临时显示在隐藏位置，避免闪现
            self.info_spacer.pack(side="bottom", fill="x")
            self.info_bar_frame.place(x=10, y=30, width=max(self.info_bar_ref_width, 400), height=30)

        # 更新窗口，确保能获取到准确的宽度
        self.root.update_idletasks()

        # 设置最大像素宽度（考虑信息栏的实际宽度、关闭按钮宽度和内边距）
        # 可用宽度 = 信息栏宽度 - 内边距 - 关闭按钮宽度
        # 增加内边距减去值，确保能显示更多字符
        max_pixel_width = info_bar_width - 40 - self.close_btn_width  # 减去更小的内边距和关闭按钮宽度

        # 确保最大像素宽度为正数
        max_pixel_width = max(max_pixel_width, self.min_pixel_width)

        # 移除文本中的换行符，确保在信息框中单行显示
        text = text.replace('\n', ' ')

        # 调用截断函数，传递信息栏的字体大小
        font_family, default_font_size = get_current_font_settings()
        info_bar_font_size = get_info_bar_font_size()
        font = self._get_font(info_bar_font_size)
        truncated_text = self._truncate_text_by_pixel(text, icon, max_pixel_width, font)

        # 保存完整文本和相关信息到实例变量
        self._full_info_text = text
        self._info_icon = icon
        self._info_label_style = label_style
        self._info_frame_style = frame_style
        self._info_max_pixel_width = max_pixel_width
        self._info_truncated = truncated_text != text
        self._info_currently_expanded = False
        
        # 显示截断文本（带有图标）
        # 使用Label组件的方法设置文本
        self.info_label.config(text=icon + truncated_text)
        # Text组件不支持style参数，通过直接设置样式属性来实现
        # self.info_label.configure(bg="#f0f0f0")  # 设置背景色
        
        # 根据消息类型设置文本颜色
        if error:
            self.info_label.configure(fg="#c62828")  # 错误信息显示红色
        else:
            self.info_label.configure(fg="#424242")  # 正确信息显示灰色
        self.info_bar_frame.configure(style=frame_style)
        
        # 确保双击事件能够正常触发展开/折叠功能
        # 先解绑可能存在的冲突绑定
        self.info_label.unbind("<Button-1>")
        self.info_label.unbind("<Double-Button-1>")
        # 单击暂停自动消失，双击触发展开/折叠
        self.info_label.bind("<Button-1>", self._on_info_bar_click)
        self.info_label.bind("<Double-Button-1>", self.toggle_info_bar_expand)
        
        # 绑定信息栏框架的点击事件，点击时暂停自动消失
        self.info_bar_frame.unbind("<Button-1>")
        self.info_bar_frame.unbind("<Double-Button-1>")
        self.info_bar_frame.bind("<Button-1>", self._on_info_bar_click)
        self.info_bar_frame.bind("<Double-Button-1>", self.toggle_info_bar_expand)

        # 显示信息栏 - 使用高度动画实现滑入效果

        if self.info_bar_animating:
            return

        # 显示 spacer，固定高度30px
        self.info_spacer.pack(side="bottom", fill="x", before=self.top_level_notebook)
        self.info_spacer.configure(height=30)
        
        # 强制更新布局
        self.root.update_idletasks()

        # 延迟一下再获取宽度，让布局稳定
        def show_with_width():
            # 计算左边距和宽度
            bar_x = 10
            # 使用主窗口宽度的88%作为信息栏宽度，减去左右边距
            main_width = self.main_frame.winfo_width()
            bar_width = int(main_width * 0.88) - 10

            bar_width = max(bar_width, 100)
            bar_width = min(bar_width, main_width - 10)

            # 先强制更新布局，确保spacer已正确pack
            self.root.update_idletasks()

            # 先将 info_bar_frame 放在隐藏位置（y=30），避免闪现
            self.info_bar_frame.place(x=bar_x, y=30, width=bar_width, height=30)
            self.info_bar_frame.lift()

            # 使用通用动画函数执行显示动画
            self.animate_info_bar('show')

        self.root.after(50, show_with_width)

    def prepare_chart_data(self, result, split_info, remaining_subnets):
        """准备图表数据"""
        try:
            # 获取父网段信息
            parent_cidr = result.get("parent", "")
            if not parent_cidr:
                self.chart_data = None
                return

            parent_info = get_subnet_info(parent_cidr)
            if "error" in parent_info:
                self.chart_data = None
                return

            parent_start = ip_to_int(parent_info.get("network", "0.0.0.0"))
            parent_end = ip_to_int(parent_info.get("broadcast", "0.0.0.0"))
            parent_range = parent_end - parent_start + 1

            # 准备所有网段数据
            self.chart_data = {
                "parent": {
                    "start": parent_start,
                    "end": parent_end,
                    "range": parent_range,
                    "name": parent_info.get("cidr", parent_cidr),
                    "color": "#f3e5f5",  # 浅紫色背景
                },
                "networks": [],
                "type": "split"  # 添加图表类型字段，确保导出PDF时能正确识别
            }

            # 添加切分网段
            if split_info:
                split_start = ip_to_int(split_info.get("network", "0.0.0.0"))
                split_end = ip_to_int(split_info.get("broadcast", "0.0.0.0"))
                self.chart_data["networks"].append(
                    {
                        "start": split_start,
                        "end": split_end,
                        "range": split_end - split_start + 1,
                        "name": split_info.get("cidr", result.get("split", "")),
                        "color": "#2196f3",  # 现代蓝色
                        "type": "split",
                    }
                )

            # 添加剩余网段 - 使用更现代化的颜色方案
            # 优化：使用元组存储，减少内存分配和提高访问速度
            subnet_colors = (
                "#4caf50",
                "#ff9800",
                "#f44336",
                "#9c27b0",
                "#00bcd4",
                "#795548",
                "#ffeb3b",
                "#607d8b",
            )  # 现代化颜色列表
            # 优化：预先计算颜色数量，避免循环中重复计算
            colors_count = len(subnet_colors)

            for index, subnet in enumerate(remaining_subnets):
                subnet_start = ip_to_int(subnet.get("network", "0.0.0.0"))
                subnet_end = ip_to_int(subnet.get("broadcast", "0.0.0.0"))
                self.chart_data["networks"].append(
                    {
                        "start": subnet_start,
                        "end": subnet_end,
                        "range": subnet_end - subnet_start + 1,
                        "name": subnet.get("cidr", ""),
                        "color": subnet_colors[index % colors_count],  # 循环使用颜色
                        "type": "remaining",
                    }
                )

            # 按起始地址排序
            self.chart_data["networks"].sort(key=lambda x: x["start"])
        except (ValueError, TypeError, AttributeError):
            # 如果出现任何错误，就不绘制图表
            self.chart_data = None

    def _on_chart_resize(self, canvas, chart_data, chart_frame, chart_type="split"):
        """通用图表尺寸变化处理

        Args:
            canvas: Canvas对象
            chart_data: 图表数据
            chart_frame: 图表框架
            chart_type: 图表类型（"split"或"plan"）
        """
        if chart_data:
            draw_distribution_chart(canvas, chart_data, chart_frame, chart_type=chart_type)

    def on_chart_resize(self, event):
        """Canvas尺寸变化时重新绘制图表"""
        if hasattr(self, 'chart_data') and self.chart_data:
            self._on_chart_resize(self.chart_canvas, self.chart_data, self.chart_frame, "split")

    def on_planning_chart_resize(self, event):
        """规划图表尺寸变化时重新绘制"""
        if hasattr(self, 'planning_chart_data') and self.planning_chart_data:
            self._on_chart_resize(self.planning_chart_canvas, self.planning_chart_data, self.planning_chart_frame, "plan")

    def _on_chart_mousewheel(self, canvas, event):
        """通用鼠标滚轮事件处理

        Args:
            canvas: Canvas对象
            event: 事件对象
        """
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_chart_mousewheel(self, event):
        """处理鼠标滚轮事件"""
        self._on_chart_mousewheel(self.chart_canvas, event)

    def on_planning_chart_mousewheel(self, event):
        """处理规划图表的鼠标滚轮事件"""
        self._on_chart_mousewheel(self.planning_chart_canvas, event)

    def draw_text_with_stroke(
        self,
        text,
        x,
        y,
        font,
        anchor=tk.W,
        fill="#ffffff",
        stroke_color="#666666",
        stroke_width=2,
    ):
        """绘制带描边的文字（包装通用函数）"""
        draw_text_with_stroke(
            self.chart_canvas, text, x, y, font,
            anchor=anchor, fill=fill,
            stroke_color=stroke_color, stroke_width=stroke_width
        )

    def draw_distribution_chart(self):
        """绘制网段分布柱状图（包装通用函数）"""
        if hasattr(self, 'chart_data') and self.chart_data:
            draw_distribution_chart(self.chart_canvas, self.chart_data, self.chart_frame)

    def on_window_resize(self, event):
        """窗口大小变化时的处理函数，实现表格和图表自适应"""

    def _prepare_export_data(self, data_source):
        """准备导出数据

        Args:
            data_source: 字典，包含导出数据的源信息

        Returns:
            tuple: (main_data, main_headers, remaining_data, remaining_headers)
        """
        main_data = []
        main_tree = data_source["main_tree"]
        main_filter = data_source.get("main_filter", None)
        main_headers = data_source.get("main_headers")

        if main_headers is None:
            main_headers = [main_tree.heading(col, "text") or "" for col in main_tree["columns"]]

        # 用于去重的集合，存储已经添加过的项目
        added_items = set()
        for item in main_tree.get_children():
            values = main_tree.item(item, "values")
            if main_filter:
                if main_filter(values):
                    # 去重：如果是键值对格式，确保每个项目只出现一次
                    if len(values) >= 2 and values[0] != "":
                        item_key = values[0]
                        if item_key not in added_items:
                            added_items.add(item_key)
                            main_data.append(values)
                    else:
                        main_data.append(values)
            elif values:
                if len(values) >= 2 and values[0] != "":
                    item_key = values[0]
                    if item_key not in added_items:
                        added_items.add(item_key)
                        main_data.append(values)
                else:
                    main_data.append(values)

        # 二次去重：确保所有数据行都是唯一的，解决切分段信息重复的问题
        unique_main_data = []
        seen_rows = set()
        for row in main_data:
            # 将行转换为可哈希的元组
            row_tuple = tuple(row)
            if row_tuple not in seen_rows:
                seen_rows.add(row_tuple)
                unique_main_data.append(row)
        main_data = unique_main_data

        # 准备剩余数据
        remaining_tree = data_source["remaining_tree"]
        remaining_headers = [remaining_tree.heading(col, "text") or "" for col in remaining_tree["columns"]]
        remaining_data = []
        for item in remaining_tree.get_children():
            values = remaining_tree.item(item, "values")
            if values:
                remaining_data.append(dict(zip(remaining_headers, values)))

        return main_data, main_headers, remaining_data, remaining_headers

    def _export_data(self, data_source, title, success_msg, failure_msg):
        """通用数据导出函数

        Args:
            data_source: 字典，包含导出数据的源信息
                - main_tree: 主数据表格控件
                - main_headers: 主数据表格标题（可选）
                - main_name: 主数据名称
                - main_filter: 主数据过滤函数（可选）
                - remaining_tree: 剩余数据表格控件
                - remaining_name: 剩余数据名称
                - pdf_title: PDF文档标题
                - main_table_cols: PDF主表格列宽（可选）
                - remaining_table_cols: PDF剩余表格列宽（可选）
            title: 文件对话框标题
            success_msg: 成功消息格式字符串
            failure_msg: 失败消息格式字符串
        """
        try:
            # 先准备默认文件名
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 获取IP版本信息
            ip_version = data_source.get("ip_version", "IPv4")
            
            # 根据数据类型生成默认文件名前缀
            if data_source["main_name"] == _("split_segment_info"):
                # 子网切分结果
                default_file_name = f"{_("subnet_split")}_{ip_version}_{timestamp}"
            else:
                # 子网规划结果
                default_file_name = f"{_("subnet_planning")}_{ip_version}_{timestamp}"
            
            # 先显示文件选择对话框，不准备数据
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[
                    ("PDF文件", "*.pdf"),
                    ("CSV文件", "*.csv"),
                    ("JSON文件", "*.json"),
                    ("文本文件", "*.txt"),
                    ("Excel文件", "*.xlsx"),
                    ("所有文件", "*.*"),
                ],
                title=title,
                initialdir="",
                initialfile=default_file_name
            )

            if not file_path:
                return

            # 用户确认文件路径后，再准备数据
            success, message, data = self.export_utils.export_data(data_source, title, success_msg, failure_msg)

            if not success:
                self.show_error("提示", message)
                return

            main_data, main_headers, remaining_data, remaining_headers = data

            # 使用导出工具导出到文件
            success, message = self.export_utils.export_to_file(
                file_path, data_source, main_data, main_headers, remaining_data, remaining_headers
            )

            if success:
                self.show_result(success_msg.format(file_path=file_path), keep_data=True)
            else:
                self.show_result(message, error=True)

        except (IOError, ValueError, TypeError, PermissionError) as e:
            error_msg = f"{failure_msg.format(error=str(e))}\n堆栈跟踪：{traceback.format_exc()}"
            self.show_result(error_msg, error=True)

    def export_result(self):
        """导出子网切分结果为多种格式（CSV、JSON、TXT、PDF、Excel）"""
        data_source = {
            "main_tree": self.split_tree,
            "main_name": _("split_segment_info"),
            "main_filter": lambda values: values[0] not in [_("hint"), _("error"), "-", _("split_segment_info"), _("remaining_segment_info")],
            "main_headers": [_("item"), _("value")],
            "remaining_tree": self.remaining_tree,
            "remaining_name": _("remaining_segment_info"),
            "pdf_title": f"{_("subnet_planner")} - {_("split_result")}",
            "main_table_cols": None,  # 使用默认列宽
            "remaining_table_cols": [40, 80, 80, 100, 90, 80, 50],  # 剩余网段表格列宽
            "chart_data": getattr(self, 'chart_data', None),  # 添加网段分布图数据
            "ip_version": self.split_ip_version_var.get(),  # 添加IP版本信息
        }

        self._export_data(data_source, _("save_subnet_split_result"), _("result_successfully_exported"), _("export_failed"))

    def export_planning_result(self):
        """导出子网规划结果为多种格式（CSV、JSON、TXT、PDF、Excel）"""
        data_source = {
            "main_tree": self.allocated_tree,
            "main_name": _("allocated_subnets"),
            "main_filter": None,  # 不需要过滤，直接导出所有数据
            "main_headers": None,  # 自动从表格获取
            "remaining_tree": self.planning_remaining_tree,
            "remaining_name": _("remaining_segment_info"),
            "pdf_title": f"{_("subnet_planner")} - {_("planning_result")}",
            "main_table_cols": [10, 100, 90, 30, 40, 80, 110, 80],  # 已分配子网表格列宽
            "remaining_table_cols": [40, 90, 80, 110, 80, 60],  # 剩余网段表格列宽
            "chart_data": getattr(self, 'planning_chart_data', None),  # 添加网段分布图数据
            "ip_version": self.ip_version_var.get(),  # 添加IP版本信息
        }

        self._export_data(data_source, _("save_subnet_planning_result"), _("result_successfully_exported"), _("export_failed"))

    def clear_result(self):
        """清空子网切分和子网规划的结果表格和图表"""
        try:
            # 清空切分段信息表格
            if hasattr(self, 'split_tree') and self.split_tree:
                self.clear_tree_items(self.split_tree)
                # 更新切分段表格的斑马条纹标签
                self.update_table_zebra_stripes(self.split_tree)

            # 清空剩余网段表表格
            if hasattr(self, 'remaining_tree') and self.remaining_tree:
                self.clear_tree_items(self.remaining_tree)
                # 更新剩余网段表的斑马条纹标签
                self.update_table_zebra_stripes(self.remaining_tree)

            # 处理剩余网段表的滚动条，确保清空结果时滚动条隐藏
            if hasattr(self, 'remaining_scroll_v') and self.remaining_scroll_v:
                try:
                    # 重置滚动条位置
                    self.remaining_scroll_v.set(0.0, 1.0)
                    self.remaining_scroll_v.grid_remove()
                except Exception:
                    pass

            # 清空图表
            if hasattr(self, 'chart_canvas') and self.chart_canvas:
                try:
                    self.chart_canvas.winfo_exists()
                    self.chart_canvas.delete("all")
                    self.chart_data = None

                    # 更新Canvas滚动区域，设置为不可滚动状态
                    self.chart_canvas.config(scrollregion=(0, 0, self.chart_canvas.winfo_width(), 100))
                except Exception:
                    pass

            # 调用滚动条回调函数，确保滚动条隐藏
            # 模拟内容不可滚动的情况，让滚动条隐藏
            if hasattr(self, 'chart_scrollbar') and self.chart_scrollbar:
                try:
                    self.chart_scrollbar.set(0.0, 1.0)
                    # 使用grid_remove()直接隐藏滚动条
                    self.chart_scrollbar.grid_remove()
                except Exception:
                    pass
        
            # 清空子网规划的已分配子网表格
            if hasattr(self, 'allocated_tree') and self.allocated_tree:
                self.clear_tree_items(self.allocated_tree)
                # 更新已分配表格的斑马条纹标签
                self.update_table_zebra_stripes(self.allocated_tree)
        
            # 清空子网规划的剩余子网表格
            if hasattr(self, 'planning_remaining_tree') and self.planning_remaining_tree:
                self.clear_tree_items(self.planning_remaining_tree)
                # 更新剩余子网表格的斑马条纹标签
                self.update_table_zebra_stripes(self.planning_remaining_tree)
            # 清空子网规划的图表
            if hasattr(self, 'planning_chart_canvas') and self.planning_chart_canvas:
                try:
                    # 检查canvas是否仍然有效
                    self.planning_chart_canvas.winfo_exists()
                    self.planning_chart_canvas.delete("all")
                    self.planning_chart_data = None
                    
                    # 更新规划图表Canvas滚动区域，设置为不可滚动状态
                    self.planning_chart_canvas.config(scrollregion=(0, 0, self.planning_chart_canvas.winfo_width(), 100))
                    
                    # 处理规划图表的滚动条，确保清空结果时滚动条隐藏
                    if hasattr(self, 'planning_chart_v_scrollbar') and self.planning_chart_v_scrollbar:
                        try:
                            self.planning_chart_v_scrollbar.set(0.0, 1.0)
                            self.planning_chart_v_scrollbar.grid_remove()
                        except Exception:
                            pass
                except Exception:
                    # 如果canvas已经失效，忽略错误
                    pass
        except Exception as e:
            # 捕获所有异常，确保语言切换等操作不会失败
            pass
    
    def one_click_export(self):
        """一键导出功能：自动执行规划和切分，然后导出所有格式的结果"""
        try:
            # 1. 自动执行规划（不更新历史记录）
            plan_result = self.perform_planning(from_history=True, update_history=False, save_state=False)
            if plan_result is None:
                return
            
            # 2. 自动执行切分（不更新历史记录）
            parent = self.parent_entry.get().strip()
            split = self.split_entry.get().strip()
            success = self.perform_split(parent, split, from_history=True, update_history=False)
            if not success:
                return
            
            # 3. 让用户选择导出目录
            export_dir = filedialog.askdirectory(title=_("select_export_directory"))
            if not export_dir:
                return
            
            # 4. 导出所有格式的结果
            self.batch_export_all_formats(export_dir)
            
            # 5. 显示成功消息
            self.show_result(_("one_click_export_success").format(dir=export_dir), keep_data=True)
        except Exception as e:
            self.show_result(_("one_click_export_failed").format(error=str(e)), error=True)
    
    def one_click_pdf(self):
        """一键PDF功能：自动执行规划和切分，然后只导出PDF格式的结果"""
        try:
            # 1. 自动执行规划（不更新历史记录）
            plan_result = self.perform_planning(from_history=True, update_history=False, save_state=False)
            if plan_result is None:
                return
            
            # 2. 自动执行切分（不更新历史记录）
            parent = self.parent_entry.get().strip()
            split = self.split_entry.get().strip()
            success = self.perform_split(parent, split, from_history=True, update_history=False)
            if not success:
                return
            
            # 3. 让用户选择导出目录
            export_dir = filedialog.askdirectory(title=_("select_export_directory"))
            if not export_dir:
                return
            
            # 4. 只导出PDF格式的结果
            self.batch_export_all_formats(export_dir, formats=['.pdf'])
            
            # 5. 显示成功消息
            self.show_result(_("one_click_pdf_success").format(dir=export_dir), keep_data=True)
        except Exception as e:
            self.show_result(_("one_click_pdf_failed").format(error=str(e)), error=True)
    
    def batch_export_all_formats(self, export_dir, formats=None):
        """批量导出所有支持的格式或指定格式
        
        Args:
            export_dir: 导出目录路径
            formats: 要导出的格式列表，默认为所有支持的格式
        """
        import os
        from datetime import datetime
        
        # 获取当前时间作为文件名前缀
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 支持的导出格式
        default_formats = ['.pdf', '.csv', '.json', '.txt', '.xlsx']
        export_formats = formats if formats is not None else default_formats
        
        # 保存当前的IP版本状态
        current_split_version = self.split_ip_version_var.get()
        current_planning_version = self.ip_version_var.get()
        
        # 要导出的IP版本列表
        ip_versions = ["IPv4", "IPv6"]
        
        try:
            # 导出两个IP版本的结果
            for ip_version in ip_versions:
                # 设置子网切分的IP版本并重新执行
                self.split_ip_version_var.set(ip_version)
                self.on_split_ip_version_change()
                self.execute_split(from_history=True)
                
                # 设置子网规划的IP版本并重新执行
                self.ip_version_var.set(ip_version)
                self.on_ip_version_change()
                self.execute_subnet_planning(from_history=True)
                
                # 导出子网切分结果
                for fmt in export_formats:
                    # 准备数据
                    split_data_source = {
                        "main_tree": self.split_tree,
                        "main_name": _("split_segment_info"),
                        "main_filter": lambda values: values[0] not in [_("hint"), _("error"), "-", _("split_segment_info"), _("remaining_segment_info")],
                        "main_headers": [_("item"), _("value")],
                        "remaining_tree": self.remaining_tree,
                        "remaining_name": _("remaining_segment_info"),
                        "pdf_title": f"{_("subnet_planner")} - {_("split_result")}",
                        "main_table_cols": None,
                        "remaining_table_cols": [40, 80, 80, 100, 90, 80, 50],
                        "chart_data": getattr(self, 'chart_data', None),
                        "ip_version": ip_version,
                    }
                    
                    # 准备文件名 - 使用翻译后的名称，并添加IP版本标识
                    split_file_name = _("subnet_split")
                    split_file_path = os.path.join(export_dir, f"{split_file_name}_{ip_version}_{timestamp}{fmt}")
                    
                    # 导出数据
                    self._export_data_to_format(split_file_path, split_data_source)
                
                # 导出子网规划结果
                for fmt in export_formats:
                    # 准备数据
                    planning_data_source = {
                        "main_tree": self.allocated_tree,
                        "main_name": _("allocated_subnets"),
                        "main_filter": None,
                        "main_headers": None,
                        "remaining_tree": self.planning_remaining_tree,
                        "remaining_name": _("remaining_segment_info"),
                        "pdf_title": f"{_("subnet_planner")} - {_("planning_result")}",
                        "main_table_cols": [10, 100, 90, 30, 40, 80, 110, 80],
                        "remaining_table_cols": [40, 90, 80, 110, 80, 60],
                        "chart_data": getattr(self, 'planning_chart_data', None),
                        "ip_version": ip_version,
                    }
                    
                    # 准备文件名 - 使用翻译后的名称，并添加IP版本标识
                    planning_file_name = _("subnet_planning")
                    planning_file_path = os.path.join(export_dir, f"{planning_file_name}_{ip_version}_{timestamp}{fmt}")
                    
                    # 导出数据
                    self._export_data_to_format(planning_file_path, planning_data_source)
        finally:
            # 恢复原来的IP版本状态
            self.split_ip_version_var.set(current_split_version)
            self.on_split_ip_version_change()
            self.ip_version_var.set(current_planning_version)
            self.on_ip_version_change()
    
    def _export_data_to_format(self, file_path, data_source):
        """导出数据到指定格式文件
        
        Args:
            file_path: 导出文件路径
            data_source: 数据源字典
        """
        try:
            # 准备数据
            success, message, data = self.export_utils.export_data(data_source, "", "", "")
            if not success:
                raise Exception(message)
            
            main_data, main_headers, remaining_data, remaining_headers = data
            
            # 使用导出工具导出到文件
            success, message = self.export_utils.export_to_file(
                file_path, data_source, main_data, main_headers, remaining_data, remaining_headers
            )
            
            if not success:
                raise Exception(message)
        except Exception as e:
            # 记录错误但继续导出其他格式
            print(f"导出失败 {file_path}: {e}")

    def create_about_link(self):
        """在主窗体标题栏右侧（红框位置）创建关于链接按钮和钉住按钮"""
        # 强制更新窗口，确保完全渲染
        self.root.update_idletasks()
        
        # 直接在root窗口创建关于链接，不使用框架
        # 使用普通tk.Label直接控制样式，确保悬停效果可靠

        # 获取窗口背景色以确保完全一致
        self.bg_color = self.root.cget("background")
        self.hover_bg_color = "#e0e0e0"  # 更浅的灰色背景，柔和过渡
        self.hover_fg_color = "#333333"  # 深灰色文字，保持可读性
        self.normal_fg_color = "#666666"
        # 使用系统默认控件边框颜色，与ttk.Combobox边框颜色一致
        border_color = "#a9a9a9"  # 系统默认灰色边框，与ttk.Combobox一致

        # 初始化窗口置顶状态
        self.is_pinned = False

        # 创建语言选择下拉菜单
        self.language_var = tk.StringVar()
        # 设置当前语言
        current_lang = get_language()
        if current_lang == "zh":
            self.language_var.set("简体中文")
        elif current_lang == "zh_tw":
            self.language_var.set("繁體中文")
        elif current_lang == "en":
            self.language_var.set("English")
        elif current_lang == "ja":
            self.language_var.set("日本語")
        elif current_lang == "ko":
            self.language_var.set("한국어")
        
        # 使用ttk.Combobox创建语言选择控件
        self.language_combobox = ttk.Combobox(
            self.root,
            textvariable=self.language_var,
            values=["简体中文", "繁體中文", "English", "日本語", "한국어"],
            state="readonly",
            width=10
        )
        
        # 绑定语言切换事件
        self.language_combobox.bind("<<ComboboxSelected>>", self.on_language_change)

        # 使用普通tk.Label创建关于标签，直接设置所有样式属性，高度与信息框一致
        font_family, font_size = get_current_font_settings()
        # 获取钉住按钮的单独字体大小配置
        pin_font_size = get_pin_button_font_size()

        self.about_label = tk.Label(
            self.root,
            text=_('about') + '…',
            font=(font_family, font_size, 'bold'),  # 使用统一的字体设置，加粗
            fg=self.normal_fg_color,  # 文字颜色调淡为浅灰色
            bg=self.bg_color,  # 背景色与窗口完全一致
            padx=12,  # 水平内边距，与子网标签一致
            pady=4.4,  # 调整垂直内边距，高度调小1px
            bd=0,  # 取消默认边框
            relief="flat",  # 平坦样式
            highlightthickness=1,  # 高亮边框宽度，模拟边框
            highlightbackground=border_color,  # 边框颜色
            highlightcolor=border_color,  # 边框颜色（确保一致性）
            cursor="hand2",  # 鼠标指针为手形
        )
        
        self.about_pressed = False
        
        def on_about_press(e):
            self.about_pressed = True
        
        def on_about_release(e):
            if self.about_pressed:
                self.about_pressed = False
                self.show_about_dialog()
        
        def on_about_enter(e):
            self.on_about_link_enter(e)
        
        def on_about_leave(e):
            self.about_pressed = False
            self.on_about_link_leave(e)
        
        self.about_label.bind("<Button-1>", on_about_press)
        self.about_label.bind("<ButtonRelease-1>", on_about_release)
        self.about_label.bind("<Enter>", on_about_enter)
        self.about_label.bind("<Leave>", on_about_leave)

        # 创建钉住按钮，使用图片替代emoji，确保所有语言下大小一致
        try:
            from PIL import Image, ImageDraw, ImageFont
            import io
            
            # 创建28x26的图片（与容器尺寸一致）
            img_size = (28, 26)
            img = Image.new('RGBA', img_size, (0, 0, 0, 0))  # 透明背景
            
            # 使用Segoe UI Emoji字体绘制📌图标（使用适中的字体大小14）
            try:
                font = ImageFont.truetype("seguiemj.ttf", 14)  # Windows emoji字体
            except Exception:
                try:
                    font = ImageFont.truetype("arial.ttf", 14)  # 备用字体
                except Exception:
                    font = ImageFont.load_default()
            
            draw = ImageDraw.Draw(img)
            
            # 绘制📌图标，使用精确的居中算法
            bbox = draw.textbbox((0, 0), "📌", font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
            
            # 精确居中计算（根据实际显示效果微调）
            x = (img_size[0] - text_width) // 2 + 0   # 水平居中
            y = (img_size[1] - text_height) // 2 + 2    # 垂直向下移动2px
            
            draw.text((x, y), "📌", fill="#000000", font=font)  # 使用黑色，更清晰可见
            
            # 转换为PhotoImage
            self.pin_icon = ImageTk.PhotoImage(img)
            
            self.pin_label = tk.Label(
                self.root,
                image=self.pin_icon,
                bg=self.bg_color,
                bd=0,
                relief="flat",
                highlightthickness=1,
                highlightbackground=border_color,
                cursor="hand2",
            )
        except Exception as e:
            print(f"创建钉住按钮图片失败: {e}，使用备用方案")
            # 备用方案：使用文字符号
            self.pin_label = tk.Label(
                self.root,
                text="⚙",
                font=("Segoe UI Symbol", 14, 'bold'),
                fg=self.normal_fg_color,
                bg=self.bg_color,
                padx=4.4,
                pady=4.4,
                bd=0,
                relief="flat",
                highlightthickness=1,
                highlightbackground=border_color,
                highlightcolor=border_color,
                cursor="hand2",
            )
        self.pin_pressed = False
        
        def on_pin_press(e):
            self.pin_pressed = True
        
        def on_pin_release(e):
            if self.pin_pressed:
                self.pin_pressed = False
                self.toggle_pin_window()
        
        def on_pin_enter(e):
            self.on_about_link_enter(e)
        
        def on_pin_leave(e):
            self.pin_pressed = False
            self.on_about_link_leave(e)
        
        self.pin_label.bind("<Button-1>", on_pin_press)
        self.pin_label.bind("<ButtonRelease-1>", on_pin_release)
        self.pin_label.bind("<Enter>", on_pin_enter)
        self.pin_label.bind("<Leave>", on_pin_leave)
        
        # 使用固定像素位置放置所有右上角控件，避免动态计算导致的偏移问题
        # 关于按钮（右侧第一个）
        self.about_label.place(
            relx=1.0, rely=0.0, anchor=tk.NE,
            x=-25, y=22,
            width=88,
            height=26
        )
        
        # 钉住按钮（右侧第二个，在关于按钮左侧）
        self.pin_label.place(
            relx=1.0, rely=0.0, anchor=tk.NE,
            x=-120, y=22,  # 固定x坐标：-25(关于按钮) - 88(关于宽度) - 5(间距) - 2(边框) = -120
            width=28,
            height=26
        )
        
        # 语言选择框（右侧第三个，在钉住按钮左侧）
        self.language_combobox.place(
            relx=1.0, rely=0.0, anchor=tk.NE,
            x=-155, y=22,  # 固定x坐标：-120(钉住) - 28(钉住宽度) - 5(间距) - 2(边框) = -155
            height=26
        )



    def on_about_link_enter(self, event):
        """鼠标进入关于链接或钉住按钮时的处理函数"""
        # 获取事件源，判断是哪个标签触发的事件
        widget = event.widget

        # 修改标签的前景色和背景色
        if widget == self.about_label:
            self.about_label.config(fg=self.hover_fg_color, bg=self.hover_bg_color)
        elif widget == self.pin_label:
            self.pin_label.config(fg="#333333", bg=self.hover_bg_color)

    def on_about_link_leave(self, event):
        """鼠标离开关于链接或钉住按钮时的处理函数"""
        widget = event.widget

        # 恢复标签的默认前景色和背景色
        if widget == self.about_label:
            self.about_label.config(fg=self.normal_fg_color, bg=self.bg_color)
        elif widget == self.pin_label:
            # 根据当前状态设置合适的颜色（扁平化风格，无relief属性）
            if self.is_pinned:
                self.pin_label.config(fg="#333333", bg="#e0e0e0")  # 深灰色文字，浅灰色背景
            else:
                self.pin_label.config(fg=self.normal_fg_color, bg=self.bg_color)  # 浅灰色文字，原始背景


    
    def on_language_change(self, event):
        """语言选择变化时的处理函数"""
        selected_language = self.language_var.get()
        if selected_language == "简体中文":
            lang_code = "zh"
        elif selected_language == "繁體中文":
            lang_code = "zh_tw"
        elif selected_language == "English":
            lang_code = "en"
        elif selected_language == "日本語":
            lang_code = "ja"
        else:  # 한국어
            lang_code = "ko"
        
        # 清除字体缓存，确保新语言使用正确的字体
        ExportUtils.clear_font_cache()
        
        # 设置当前语言
        set_language(lang_code)
        
        # 更新应用程序名称
        self.app_name = _("app_name")
        self.root.title(f"{self.app_name} v{self.app_version}")
        
        # 更新样式，确保新创建的UI元素使用正确的样式
        update_styles()
        
        # 重新创建所有UI元素，实现语言更新
        self.destroy_all_widgets()
        self.recreate_ui()
        
        # 重新绘制拓扑图，应用新的字体设置
        if hasattr(self, 'topology_visualizer'):
            # 尝试获取当前选中的网络
            selected_network = None
            if hasattr(self, 'ipam_network_tree'):
                selected_items = self.ipam_network_tree.selection()
                if selected_items:
                    selected_network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
            # 刷新可视化
            self.refresh_visualization(selected_network)
    
    def destroy_all_widgets(self):
        """销毁所有子组件"""
        # 关闭功能调试面板，如果它是打开的
        if hasattr(self, 'test_dialog') and self.test_dialog is not None:
            self.test_dialog.destroy()
            self.test_dialog = None
        
        # 销毁主框架，如果存在的话
        if hasattr(self, 'main_frame'):
            self.main_frame.destroy()
        
        # 销毁右上角的控件
        if hasattr(self, 'about_label'):
            self.about_label.destroy()
        if hasattr(self, 'pin_label'):
            self.pin_label.destroy()
        if hasattr(self, 'language_combobox'):
            self.language_combobox.destroy()
        
        # 销毁信息栏相关控件
        if hasattr(self, 'info_spacer'):
            self.info_spacer.destroy()
        if hasattr(self, 'info_bar_frame'):
            self.info_bar_frame.destroy()
        if hasattr(self, 'info_close_btn'):
            self.info_close_btn.destroy()
        if hasattr(self, 'info_label'):
            self.info_label.destroy()
    
    def recreate_ui(self):
        """重新创建UI元素"""
        # 重新初始化所有必要的组件属性
        self.initialize_component_properties()
        
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        # 获取信息栏的独立字体大小配置
        info_bar_font_size = get_info_bar_font_size()
        
        # 重新创建主框架
        self.main_frame = ttk.Frame(self.root, padding="15")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 重新创建信息栏 spacer
        self.info_spacer = ttk.Frame(self.main_frame, style="Placeholder.TFrame")
        self.info_spacer.pack(side="bottom", fill="x")
        self.info_spacer.pack_forget()
        self.info_spacer.configure(height=30)
        
        # 重新创建信息栏框架
        self.info_bar_frame = ttk.Frame(self.info_spacer, style="InfoBar.TFrame")
        self.info_bar_frame.place_forget()
        
        # 重新创建顶级标签页控件
        self.create_top_level_notebook()
        
        # 重新创建右上角的关于链接按钮和钉住按钮
        self.create_about_link()
        
        self.info_bar_frame.grid_rowconfigure(0, weight=1)
        self.info_bar_frame.grid_columnconfigure(0, weight=1)
        self.info_bar_frame.grid_columnconfigure(1, weight=0)
        
        # 重新创建信息栏关闭按钮
        bg_color = self.style.lookup("TFrame", "background")
        self.info_close_btn = tk.Button(
            self.info_bar_frame,
            text="✕",
            command=self.hide_info_bar,
            cursor="hand2",
            takefocus=False,
            bg=bg_color,
            fg="#9E9E9E",
            font=("Arial", 8),  # 此处字体硬编码是程序需要，禁止修改
            borderwidth=0,
            highlightthickness=0,
            relief="flat",
            padx=0,
            pady=0,
        )
        self.info_close_btn.grid(row=0, column=1, padx=(0, 0), pady=(0, 4), sticky="se")
        
        # 重新创建信息标签（使用Label组件替代Text，以简化实现）
        # 获取背景色
        bg_color = self.style.lookup("TFrame", "background")
        self.info_label = tk.Label(
            self.info_bar_frame,
            padx=0, pady=0,  # 使用内边距控制间距，使用单一值而非元组
            font=(font_family, info_bar_font_size),  # 使用信息栏独立的字体大小配置
            takefocus=False,  # 不接受焦点
            cursor="arrow",  # 显示普通箭头光标
            background=bg_color,  # 设置背景色跟随主题
            anchor="w",  # 文本左对齐
            justify="left",  # 多行文本时左对齐，与第一次创建保持一致
        )
        self.info_label.grid(row=0, column=0, sticky="ew", padx=(5, 0), pady=0)
        self.info_label.lift(self.info_close_btn)
        
        # 重新初始化图表数据
        self.chart_data = None
        
        # 重新初始化历史记录 - 语言切换不应保留撤销历史
        self.history_repo.clear_split_history()
        self.history_records = self.history_repo.history_records
        self.history_states = deque(maxlen=20)
        self.current_history_index = -1
        self.deleted_history = []
        
        self.root.update_idletasks()
        self.info_bar_ref_width = max(self.main_frame.winfo_width() - 10, 100)
    
    def initialize_component_properties(self):
        """重新初始化所有必要的组件属性"""
        self.planning_parent_networks = self.history_repo.planning_parent_networks
        self.planning_parent_entry = None
        self.pool_tree = None
        self.pool_scrollbar = None
        self.requirements_tree = None
        self.requirements_scrollbar = None
        
        # 重新初始化export_utils，确保使用当前语言的字体
        from export_utils import ExportUtils
        self.export_utils = ExportUtils()
        
        self.test_dialog = None  # 用于存储调试面板的引用，确保只能打开一个
        self.undo_delete_btn = None
        self.swap_btn = None
        self.planning_notebook = None
        self.execute_planning_btn = None
        self.allocated_frame = None
        self.allocated_tree = None
        self.planning_remaining_frame = None
        self.planning_remaining_tree = None
        
        self.edit_entry = None
        self.current_edit_item = None
        self.current_edit_column = None
        self.current_edit_column_index = None
        self.current_edit_tree = None
        
        self.advanced_notebook = None
        self.ipv4_info_frame = None
        self.ipv6_info_frame = None
        self.merge_frame = None
        self.overlap_frame = None
        self.ipv6_info_entry = None
        self.ipv6_cidr_var = None
        self.ipv6_cidr_combobox = None
        self.ipv6_info_btn = None
        self.ipv6_info_tree = None
        self.subnet_merge_text = None
        self.merge_btn = None
        self.range_start_entry = None
        self.range_end_entry = None
        self.range_to_cidr_btn = None
        self.merge_result_frame = None
        self.merge_result_tree = None
        self.merge_result_scrollbar = None
        self.ip_info_entry = None
        self.subnet_mask_cidr_map = None
        self.cidr_subnet_mask_map = None
        self.ip_mask_var = None
        self.ip_mask_combobox = None
        self.ip_cidr_var = None
        self.ip_cidr_combobox = None
        self.ip_info_btn = None
        self.ip_info_tree = None
        self.overlap_text = None
        self.overlap_btn = None
        self.overlap_result_tree = None
        
        self.split_parent_networks = deque(maxlen=100)
        self.split_networks = deque(maxlen=100)
        self.parent_entry = None
        self.split_entry = None
        self.execute_btn = None
        self.reexecute_btn = None
        self.history_tree = None
        self.export_btn = None
        self.notebook = None
        self.split_info_frame = None
        self.split_tree = None
        self.remaining_frame = None
        self.remaining_tree = None
        self.chart_frame = None
        self.chart_scrollbar = None
        self.chart_canvas = None
        self.remaining_scroll_v = None
    
    def on_window_configure(self, event):
        """窗口大小变化时动态调整右上角按钮位置"""
        # 窗口大小变化时重新获取窗口背景色，确保按钮背景色与窗口一致
        self.bg_color = self.root.cget("background")
        
        # 更新按钮背景色，先检查控件是否存在
        if hasattr(self, 'about_label') and self.about_label.winfo_exists():
            self.about_label.config(bg=self.bg_color)
        if hasattr(self, 'pin_label') and self.pin_label.winfo_exists():
            if not self.is_pinned:
                # Check if normal_fg_color is set before accessing it
                if hasattr(self, 'normal_fg_color'):
                    self.pin_label.config(bg=self.bg_color)
                else:
                    # If normal_fg_color isn't set yet, just update the background
                    self.pin_label.config(bg=self.bg_color)
        
        # 这里不需要调整按钮位置，因为按钮使用了相对定位（relx=1.0）
        # 窗口大小变化时，按钮会自动保持在右上角位置
    
    def toggle_pin_window(self):
        """切换窗口置顶状态"""
        # 切换置顶状态
        self.is_pinned = not self.is_pinned

        # 设置窗口置顶属性
        self.root.attributes("-topmost", self.is_pinned)

        # 更新钉住按钮的视觉效果（扁平化风格，使用颜色变化表示状态）
        if self.is_pinned:
            # 置顶状态：深色图标和文字
            self.pin_label.config(fg="#333333", bg="#e0e0e0")  # 深灰色文字，浅灰色背景
        else:
            # 非置顶状态：浅色图标和文字
            self.pin_label.config(fg=self.normal_fg_color, bg=self.bg_color)  # 浅灰色文字，原始背景

    def show_about_dialog(self):
        """显示关于对话框"""
        # 设置对话框宽度，高度将根据内容自动调整
        dialog_width = 400
        dialog_height = 380  # 初始高度设置为较大值以容纳所有内容
        
        # 创建关于对话框
        about_window = ComplexDialog(self.root, f"{_("about")} {self.app_name}", dialog_width, dialog_height, resizable=False, modal=True)
        
        # 设置背景色为白色
        about_window.dialog.configure(bg="#ffffff")

        # 创建内容框架，移除所有边框和焦点指示
        content_frame = ttk.Frame(about_window.content_frame, padding=(20, 20, 20, 15), relief="flat", borderwidth=0)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 创建内部框架放置实际内容，不使用占位符框架
        inner_frame = ttk.Frame(content_frame)
        inner_frame.pack(side="top", fill="both", expand=True)
        inner_frame.configure(style="About.TFrame")

        # 移除可能影响焦点的事件绑定
        about_window.dialog.unbind("<FocusIn>")
        about_window.dialog.unbind("<FocusOut>")

        # 为关于对话框中的标签和按钮添加焦点样式，移除虚线
        # 创建对话框专用的样式，避免影响主窗口
        # 只在样式未配置时才配置，避免重复配置
        try:
            self.style.configure("About.TLabel", focuscolor="none")
            self.style.configure("About.TButton",
                               focuscolor="none",
                               focuswidth=0,
                               padding=(10, 2))
            self.style.map("About.TButton",
                          focuscolor=[("focus", "none")],
                          focuswidth=[("focus", 0)])
        except tk.TclError:
            pass  # 样式已配置或配置失败，忽略错误

        # 获取当前字体设置，确保与应用程序其他部分一致
        font_family, __ = get_current_font_settings()

        # 标题区域
        title_frame = ttk.Frame(inner_frame)
        title_frame.pack(pady=(10, 8))

        # 添加应用名称作为主要标题，使用蓝色主题色
        app_name_label = ttk.Label(title_frame, 
                                 text=self.app_name, 
                                 font=(font_family, 18, "bold"), 
                                 style="About.TLabel",
                                 foreground="#1976d2")
        app_name_label.pack(anchor=tk.CENTER)

        # 添加版本号，使用灰色调
        version_label = ttk.Label(
            title_frame, 
            text=f"{_("version")} {self.app_version}", 
            font=(font_family, 11, "bold"), 
            style="About.TLabel",
            foreground="#666666"
        )
        version_label.pack(anchor=tk.CENTER, pady=(5, 0))

        # 添加分隔线，增强视觉层次
        separator = ttk.Separator(title_frame, orient='horizontal')
        separator.pack(fill=tk.X, pady=(5, 0))

        # 添加应用描述，使用翻译函数确保多语言适配
        desc_label = ttk.Label(
            title_frame, 
            text=_("app_description"), 
            font=(font_family, 11), 
            style="About.TLabel",
            foreground="#555555"
        )
        desc_label.pack(anchor=tk.CENTER, pady=(5, 0))

        # 信息区域
        info_frame = ttk.Frame(inner_frame)
        info_frame.pack(pady=(10, 5))
        
        # 添加作者信息
        # 使用ttk.Label与其他标签保持一致
        author_label = ttk.Label(
            info_frame, 
            text=f"{_("author")}：Ejones", 
            font=(font_family, 10), 
            style="About.TLabel",
            foreground="#444444"
        )
        author_label.pack(anchor=tk.CENTER, pady=(2, 0))

        # 添加联系方式，使用蓝色调，可点击复制
        # 使用ttk.Label与其他标签保持一致
        email_label = ttk.Label(
            info_frame, 
            text=f"{_("email")}：ejones.cn@hotmail.com", 
            font=(font_family, 10), 
            style="About.TLabel",
            foreground="#1976d2",
            cursor="hand2"
        )
        email_label.pack(anchor=tk.CENTER, pady=(0, 0))
        
        # 为邮箱添加点击事件，启动邮件客户端
        import webbrowser
        
        email_pressed = [False]
        original_email_color = "#1976d2"
        hover_email_color = "#42a5f5"
        pressed_email_color = "#d32f2f"
        
        def on_email_enter(e):
            if not email_pressed[0]:
                email_label.config(foreground=hover_email_color)
        
        def on_email_press(e):
            email_pressed[0] = True
            email_label.config(foreground=pressed_email_color)
        
        def on_email_release(e):
            email_label.config(foreground=hover_email_color)
            if email_pressed[0]:
                email_pressed[0] = False
                webbrowser.open("mailto:ejones.cn@hotmail.com")
        
        def on_email_leave(e):
            email_pressed[0] = False
            email_label.config(foreground=original_email_color)
        
        email_label.bind("<Enter>", on_email_enter)
        email_label.bind("<Button-1>", on_email_press)
        email_label.bind("<ButtonRelease-1>", on_email_release)
        email_label.bind("<Leave>", on_email_leave)

        # 按钮区域
        button_frame = ttk.Frame(inner_frame)
        button_frame.pack(anchor=tk.CENTER, pady=(10, 5))

        # 添加扫码捐赠按钮
        def show_donate_qr():
            # 创建二维码捐赠对话框
            # 设置对话框宽度，高度将根据内容自动调整
            qr_window = ComplexDialog(self.root, _("donate"), 420, 420, resizable=False, modal=True)
            qr_window.dialog.configure(bg="#ffffff")
            
            # 创建内容框架
            qr_content = ttk.Frame(qr_window.content_frame, padding=(20, 20, 20, 10))
            qr_content.pack(fill=tk.BOTH, expand=True)
            
            # 添加标题
            qr_title = ttk.Label(qr_content, 
                               text=_("scan_qr_donate"), 
                               font=(font_family, 14, "bold"),
                               style="About.TLabel")
            qr_title.pack(pady=(0, 15))
            
            # 二维码容器
            qr_container = ttk.Frame(qr_content)
            qr_container.pack()
            
            # 添加微信二维码
            wechat_frame = ttk.Frame(qr_container)
            wechat_frame.pack(side=tk.LEFT, padx=(0, 10))
            
            wechat_label = ttk.Label(wechat_frame, 
                                   text=_("wechat"), 
                                   font=(font_family, 12, "bold"),
                                   style="About.TLabel")
            wechat_label.pack(pady=(0, 5))
            
            # 加载并显示微信二维码
            try:
                from PIL import Image, ImageTk
                import os
                
                # 微信二维码图片路径（如果存在）
                wechat_qr_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Picture", "33144271b0126dc527a3697193132c8f.jpg")
                
                if os.path.exists(wechat_qr_path):
                    # 加载图片
                    wechat_img = Image.open(wechat_qr_path)
                    wechat_img = wechat_img.resize((150, 150), Image.Resampling.LANCZOS)
                    wechat_photo = ImageTk.PhotoImage(wechat_img)
                    
                    # 创建标签显示图片
                    wechat_qr_label = ttk.Label(wechat_frame, image=wechat_photo, style="About.TLabel")
                    wechat_qr_label.image = wechat_photo  # 保存引用，防止被GC回收
                    wechat_qr_label.pack(pady=(5, 5))
                else:
                    # 微信二维码占位符
                    wechat_qr_placeholder = ttk.Label(wechat_frame, 
                                                    text=_("wechat_qr"), 
                                                    font=(font_family, 10),
                                                    style="About.TLabel",
                                                    foreground="#999999")
                    wechat_qr_placeholder.pack(pady=(5, 5))
            except Exception as e:
                # 加载图片失败时显示占位符
                print(f"加载微信二维码失败: {e}")
                wechat_qr_placeholder = ttk.Label(wechat_frame, 
                                                text=_("wechat_qr"), 
                                                font=(font_family, 10),
                                                style="About.TLabel",
                                                foreground="#999999")
                wechat_qr_placeholder.pack(pady=(5, 5))
            
            # 添加支付宝二维码
            alipay_frame = ttk.Frame(qr_container)
            alipay_frame.pack(side=tk.LEFT, padx=(5, 0))
            
            alipay_label = ttk.Label(alipay_frame, 
                                   text=_("alipay"), 
                                   font=(font_family, 12, "bold"),
                                   style="About.TLabel")
            alipay_label.pack(pady=(0, 5))
            
            # 加载并显示支付宝二维码
            try:
                from PIL import Image, ImageTk
                import os
                
                # 支付宝二维码图片路径（如果存在）
                alipay_qr_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Picture", "df981e51d905be6bfc2eda8666621d93.jpg")
                
                if os.path.exists(alipay_qr_path):
                    # 加载图片
                    alipay_img = Image.open(alipay_qr_path)
                    alipay_img = alipay_img.resize((150, 150), Image.Resampling.LANCZOS)
                    alipay_photo = ImageTk.PhotoImage(alipay_img)
                    
                    # 创建标签显示图片
                    alipay_qr_label = ttk.Label(alipay_frame, image=alipay_photo, style="About.TLabel")
                    alipay_qr_label.image = alipay_photo  # 保存引用，防止被GC回收
                    alipay_qr_label.pack(pady=(5, 5))
                else:
                    # 支付宝二维码占位符
                    alipay_qr_placeholder = ttk.Label(alipay_frame, 
                                                    text=_("alipay_qr"), 
                                                    font=(font_family, 10),
                                                    style="About.TLabel",
                                                    foreground="#999999")
                    alipay_qr_placeholder.pack(pady=(5, 5))
            except Exception as e:
                # 加载图片失败时显示占位符
                print(f"加载支付宝二维码失败: {e}")
                alipay_qr_placeholder = ttk.Label(alipay_frame, 
                                                text=_("alipay_qr"), 
                                                font=(font_family, 10),
                                                style="About.TLabel",
                                                foreground="#999999")
                alipay_qr_placeholder.pack(pady=(5, 5))
            
            # 添加提示信息
            tip_label = ttk.Label(qr_content, 
                               text=_("donate_tip"), 
                               font=(font_family, 10),
                               style="About.TLabel",
                               foreground="#666666",
                               wraplength=360,  # 减小换行宽度，确保文字完全显示
                               justify=tk.LEFT)  # 设置文本左对齐
            tip_label.pack(pady=(5, 5))
            
            # 添加开源地址链接
            import webbrowser
            
            github_pressed = [False]
            original_github_color = "#1976d2"
            hover_github_color = "#42a5f5"
            pressed_github_color = "#d32f2f"
            
            def on_github_enter(e):
                if not github_pressed[0]:
                    github_label.config(foreground=hover_github_color)
            
            def on_github_press(e):
                github_pressed[0] = True
                github_label.config(foreground=pressed_github_color)
            
            def on_github_release(e):
                github_label.config(foreground=hover_github_color)
                if github_pressed[0]:
                    github_pressed[0] = False
                    webbrowser.open("https://gitcode.com/ejones-cn/Subnet_Planner")
            
            def on_github_leave(e):
                github_pressed[0] = False
                github_label.config(foreground=original_github_color)
            
            github_label = ttk.Label(qr_content, 
                                   text="https://gitcode.com/ejones-cn/Subnet_Planner", 
                                   font=(font_family, 9),
                                   style="About.TLabel",
                                   foreground="#1976d2",
                                   cursor="hand2")
            github_label.pack(pady=(5, 5))
            github_label.bind("<Enter>", on_github_enter)
            github_label.bind("<Button-1>", on_github_press)
            github_label.bind("<ButtonRelease-1>", on_github_release)
            github_label.bind("<Leave>", on_github_leave)
            
            # 关闭按钮
            close_button = ttk.Button(qr_content, 
                                    text=_('close'), 
                                    command=qr_window.destroy, 
                                    width=10,
                                    style="About.TButton")
            close_button.pack()
            
            # 显示对话框并自动调整高度
            qr_window.show()
        
        # 添加扫码捐赠按钮
        qr_button = ttk.Button(button_frame, 
                            text=_('donate'), 
                            command=show_donate_qr, 
                            width=10,
                            style="About.TButton")
        qr_button.pack(side=tk.LEFT, padx=(0, 5))

        # 添加确定按钮，使用更大的宽度和更好的居中效果
        ok_button = ttk.Button(button_frame, 
                             text=_("ok"), 
                             command=about_window.destroy, 
                             width=10,
                             style="About.TButton")
        ok_button.pack(side=tk.LEFT)

        # 绑定回车键事件，确保按回车键能关闭对话框
        about_window.dialog.bind('<Return>', lambda e: ok_button.invoke())
        about_window.dialog.bind('<Escape>', lambda e: ok_button.invoke())

        # 添加开源地址链接
        import webbrowser
        
        github_pressed = [False]
        original_github_color = "#1976d2"
        hover_github_color = "#42a5f5"
        pressed_github_color = "#d32f2f"
        
        def on_github_enter(e):
            if not github_pressed[0]:
                github_label.config(foreground=hover_github_color)
        
        def on_github_press(e):
            github_pressed[0] = True
            github_label.config(foreground=pressed_github_color)
        
        def on_github_release(e):
            github_label.config(foreground=hover_github_color)
            if github_pressed[0]:
                github_pressed[0] = False
                webbrowser.open("https://gitcode.com/ejones-cn/Subnet_Planner")
        
        def on_github_leave(e):
            github_pressed[0] = False
            github_label.config(foreground=original_github_color)
        
        github_label = ttk.Label(
            inner_frame, 
            text="https://gitcode.com/ejones-cn/Subnet_Planner", 
            font=(font_family, 9),
            style="About.TLabel",
            foreground="#1976d2",
            cursor="hand2"
        )
        github_label.pack(anchor=tk.CENTER, pady=(5, 5))
        github_label.bind("<Enter>", on_github_enter)
        github_label.bind("<Button-1>", on_github_press)
        github_label.bind("<ButtonRelease-1>", on_github_release)
        github_label.bind("<Leave>", on_github_leave)
        
        # 添加版权信息，使用动态字体设置，灰色调
        copyright_label = ttk.Label(
            inner_frame, 
            text=_('copyright').format(app_name=self.app_name), 
            font=(font_family, 8),  # 使用动态字体，保持9号大小
            style="About.TLabel",
            foreground="#888888"
        )
        copyright_label.pack(anchor=tk.CENTER, pady=(5, 5))
        
        # 显示对话框并自动调整高度
        about_window.show()

    def setup_ipam_page(self):
        """设置IP地址管理（IPAM）页面"""
        # 配置IPAM页面的网格布局
        self.ipam_frame.grid_rowconfigure(0, weight=0)
        self.ipam_frame.grid_rowconfigure(1, weight=1)
        self.ipam_frame.grid_columnconfigure(0, weight=1)
        self.ipam_frame.grid_columnconfigure(1, weight=1)
        
        # 网络管理区域
        network_frame = ttk.LabelFrame(self.ipam_frame, text=_('network_management'), padding=(10, 0, 0, 10))
        network_frame.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        
        # 配置网络管理区域的网格布局
        network_frame.grid_columnconfigure(0, weight=1)
        network_frame.grid_rowconfigure(0, weight=0)  # 按钮行
        network_frame.grid_rowconfigure(1, weight=1)  # 表格容器行
        

        

        
        # 网络管理按钮 - 使用网格布局，更加紧凑
        button_frame = ttk.Frame(network_frame)
        button_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        # 配置按钮框架的网格布局
        for i in range(NETWORK_MANAGEMENT_BUTTON_COUNT):
            button_frame.grid_columnconfigure(i, weight=1)
        
        # 所有按钮排成一行
        ttk.Button(button_frame, text=_('add_network'), command=self.add_ipam_network).grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        ttk.Button(button_frame, text=_('remove_network'), command=self.remove_ipam_network).grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        ttk.Button(button_frame, text=_('check_conflicts'), command=self.check_ip_conflicts).grid(row=0, column=2, padx=2, pady=2, sticky="ew")
        ttk.Button(button_frame, text=_('auto_scan'), command=self.auto_scan_network).grid(row=0, column=3, padx=2, pady=2, sticky="ew")
        ttk.Button(button_frame, text=_('backup_restore'), command=self.backup_restore_data).grid(row=0, column=4, padx=2, pady=2, sticky="ew")
        ttk.Button(button_frame, text=_('import_export'), command=self.import_export_network_data).grid(row=0, column=5, padx=(2, 10), pady=2, sticky="ew")
        
        # 创建表格容器 - 专门用来容纳表格和滚动条
        table_container = ttk.Frame(network_frame)
        table_container.grid(row=1, column=0, sticky="nsew")
        
        # 配置 table_container 的 grid 布局
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)
        
        # 网络列表 - 在table_container上创建
        self.ipam_network_tree = ttk.Treeview(table_container, columns=('network', 'description', 'vlan', 'created_at', 'ip_count'), show='tree headings', height=8, selectmode='extended')
        
        # 启用层次关系线
        style = ttk.Style()
        style.configure('Treeview', indent=10)
        
        # 禁用焦点指示器和选中状态的边框效果，避免空单元格显示选中标记
        style.map('Treeview')
        
        # 不设置特定主题，使用系统默认主题
        # 设置列标题
        self.ipam_network_tree.heading('#0', text=_('network_segment'))  # 树状结构列的标题
        self.ipam_network_tree.heading('network', text=_('network_address'))
        self.ipam_network_tree.heading('description', text=_('description'))
        self.ipam_network_tree.heading('vlan', text='VLAN')
        self.ipam_network_tree.heading('created_at', text=_('created_at'))
        self.ipam_network_tree.heading('ip_count', text=_('ip_count'))
        
        # 调整列宽，提高空间利用率
        self.ipam_network_tree.column('network', width=0, stretch=False)  # 隐藏重复的网段地址列
        self.ipam_network_tree.column('description', width=200, minwidth=150, stretch=True)
        self.ipam_network_tree.column('vlan', width=80, minwidth=60, stretch=True)
        self.ipam_network_tree.column('created_at', width=150, minwidth=120, stretch=True)
        self.ipam_network_tree.column('ip_count', width=70, minwidth=60, stretch=True)
        
        # 创建垂直滚动条 - 在table_container上创建
        network_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL)
        
        # 使用通用方法创建带自动隐藏功能的垂直滚动条
        self.create_scrollable_treeview(table_container, self.ipam_network_tree, network_scrollbar)
        
        # 绑定网络选择事件
        self.ipam_network_tree.bind('<<TreeviewSelect>>', self.on_ipam_network_select)
        # 绑定鼠标点击事件用于取消选择
        self.ipam_network_tree.bind('<Button-1>', self.on_ipam_network_click)
        # 为网络管理表添加内联编辑功能
        # 注册网络管理表的内联编辑配置
        self.register_inline_edit_config('ipam_network', {
            'editable_columns': [0, 1, 2],  # 网段地址列、描述列和VLAN列
            'column_types': {
                0: 'entry',  # 网段地址列使用文本框
                1: 'entry',  # 描述列使用文本框
                2: 'entry'   # VLAN列使用文本框
            }
        })
        
        # 注册网络管理表的内联编辑处理器
        self.register_inline_edit_handler('ipam_network', {
            'get_row_data': self._get_network_row_data,
            'validate': self._validate_network_edit,
            'save': self._save_network_edit
        })
        
        # 绑定双击事件用于内联编辑
        self.ipam_network_tree.bind('<Double-1>', lambda event: self.on_generic_tree_double_click(self.ipam_network_tree, 'ipam_network', event))
        # 绑定展开/收缩事件，用于刷新斑马纹
        self.ipam_network_tree.bind('<<TreeviewOpen>>', lambda event: self.ipam_network_tree.after(100, lambda: self.update_table_zebra_stripes(self.ipam_network_tree)))
        self.ipam_network_tree.bind('<<TreeviewClose>>', lambda event: self.ipam_network_tree.after(100, lambda: self.update_table_zebra_stripes(self.ipam_network_tree)))
        
        # 配置Treeview样式，包括斑马条纹
        self.configure_treeview_styles(self.ipam_network_tree)
        # 添加斑马纹样式
        self.update_table_zebra_stripes(self.ipam_network_tree)
        
        # 创建自定义笔记本控件
        self.ipam_notebook = ColoredNotebook(self.ipam_frame, style=self.style)
        self.ipam_notebook.grid(row=1, column=0, columnspan=2, sticky="nsew")
        
        # IP地址管理标签页 - 蓝色容器保持padding为5
        ip_management_frame = ttk.Frame(self.ipam_notebook.content_area, padding="5", style=self.ipam_notebook.light_blue_style)
        
        # 配置IP地址管理区域的网格布局
        ip_management_frame.grid_columnconfigure(0, weight=1)
        ip_management_frame.grid_rowconfigure(2, weight=1)
        
        # 创建内容容器，灰色容器 - 加上10的内部边距
        content_container = ttk.Frame(ip_management_frame, padding="10")
        content_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # IP地址管理按钮 - 使用网格布局，更加紧凑
        ip_button_frame = ttk.Frame(content_container)
        ip_button_frame.pack(fill=tk.X, pady=(0, 5))
        
        # 一行按钮，按要求排序
        ttk.Button(ip_button_frame, text=_('allocate_reserve_address'), command=self.allocate_reserve_ip).pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        ttk.Button(ip_button_frame, text=_('release_address'), command=self.release_ip_address).pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        ttk.Button(ip_button_frame, text=_('batch_migrate'), command=self.batch_migrate_ip_addresses).pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        ttk.Button(ip_button_frame, text=_('cleanup_unused'), command=self.cleanup_available_ips).pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        ttk.Button(ip_button_frame, text=_('check_expired_ips'), command=self.check_expired_ips).pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        ttk.Button(ip_button_frame, text=_('batch_set_expiry_date'), command=self.batch_set_expiry_date).pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        
        # IP地址列表 - 使用LabelFrame，和IPv6查询结果表保持一致
        # padding=(10, 0, 0, 10) 表示左边距10，上边距0，右边距0，下边距10
        ip_list_frame = ttk.LabelFrame(content_container, text=_('ip_address_list'), padding=(10, 0, 0, 10))
        ip_list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 配置 ip_list_frame 的 grid 布局
        ip_list_frame.grid_rowconfigure(0, weight=0)  # 搜索控件行
        ip_list_frame.grid_rowconfigure(1, weight=1)  # 表格容器行
        ip_list_frame.grid_columnconfigure(0, weight=1)
        
        # 搜索和过滤功能 - 移到LabelFrame内部，放在表格上部
        search_frame = ttk.Frame(ip_list_frame)
        search_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        # 配置搜索框架的网格布局
        search_frame.grid_columnconfigure(0, weight=0)
        search_frame.grid_columnconfigure(1, weight=3)  # 关键词输入框权重更大
        search_frame.grid_columnconfigure(2, weight=0)
        search_frame.grid_columnconfigure(3, weight=1)  # 范围下拉列表所在列
        search_frame.grid_columnconfigure(4, weight=0)
        search_frame.grid_columnconfigure(5, weight=1)  # 搜索模式下拉列表所在列
        search_frame.grid_columnconfigure(6, weight=0)
        search_frame.grid_columnconfigure(7, weight=1)  # 状态下拉列表所在列
        search_frame.grid_columnconfigure(8, weight=0)
        search_frame.grid_columnconfigure(9, weight=1)  # 过期日期下拉列表所在列
        search_frame.grid_columnconfigure(10, weight=0)
        
        # 第一行：所有输入控件
        ttk.Label(search_frame, text=_('keyword') + ':').grid(row=0, column=0, sticky="e", padx=5, pady=0)
        search_entry_border, self.ipam_search_entry = create_bordered_entry(search_frame)
        search_entry_border.grid(row=0, column=1, sticky="ew", padx=5, pady=0)
        
        ttk.Label(search_frame, text=_('scope') + ':').grid(row=0, column=2, sticky="e", padx=5, pady=0)
        self.search_scope = ttk.Combobox(search_frame, values=[_('all'), _('ip_address'), _('hostname'), _('description'), _('mac_address')], width=6)
        self.search_scope.set(_('all'))
        self.search_scope.grid(row=0, column=3, sticky="ew", padx=5, pady=0)
        
        ttk.Label(search_frame, text=_('search_mode') + ':').grid(row=0, column=4, sticky="e", padx=5, pady=0)
        self.search_mode = ttk.Combobox(search_frame, values=[_('exact_match'), _('contains'), _('regex')], width=6)
        self.search_mode.set(_('contains'))
        self.search_mode.grid(row=0, column=5, sticky="ew", padx=5, pady=0)
        
        ttk.Label(search_frame, text=_('status') + ':').grid(row=0, column=6, sticky="e", padx=5, pady=0)
        self.ipam_status_filter = ttk.Combobox(search_frame, values=[_('all'), _('allocated'), _('reserved'), _('released')], width=6)
        self.ipam_status_filter.set(_('all'))
        self.ipam_status_filter.grid(row=0, column=7, sticky="ew", padx=5, pady=0)
        
        # 过期日期筛选
        ttk.Label(search_frame, text=_('expiry_date') + ':').grid(row=0, column=8, sticky="e", padx=5, pady=0)
        self.ipam_expiry_filter = ttk.Combobox(search_frame, values=[_('all'), _('expired'), _('expiring_soon'), _('not_expired')], width=6)
        self.ipam_expiry_filter.set(_('all'))
        self.ipam_expiry_filter.grid(row=0, column=9, sticky="ew", padx=5, pady=0)
        
        # 重置按钮（小方块图标）
        reset_button = ttk.Button(search_frame, text="↺", command=self.reset_search, width=3)
        reset_button.grid(row=0, column=10, padx=(5, 10), pady=0)
        
        # 添加实时搜索和自动过滤事件监听器
        self.ipam_search_entry.bind('<KeyRelease>', self.on_search_input)
        self.search_scope.bind('<<ComboboxSelected>>', self.on_filter_change)
        self.search_mode.bind('<<ComboboxSelected>>', self.on_filter_change)
        self.ipam_status_filter.bind('<<ComboboxSelected>>', self.on_filter_change)
        self.ipam_expiry_filter.bind('<<ComboboxSelected>>', self.on_filter_change)
        
        # 创建表格容器 - 专门用来容纳表格和滚动条
        table_container = ttk.Frame(ip_list_frame)
        table_container.grid(row=1, column=0, sticky="nsew")
        
        # 配置 table_container 的 grid 布局
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)
        
        # IP 地址列表 - 在table_container上创建
        self.ipam_ip_tree = ttk.Treeview(table_container, columns=('ip_address', 'status', 'hostname', 'mac_address', 'description', 'allocated_at', 'expiry_date'), show='headings', height=8, selectmode='extended')
        
        # 为表头添加点击事件，实现排序功能
        self.ipam_ip_tree.heading('ip_address', text=_('ip_address'), command=lambda: self.sort_ip_table('ip_address'))
        self.ipam_ip_tree.heading('status', text=_('status'), command=lambda: self.sort_ip_table('status'))
        self.ipam_ip_tree.heading('hostname', text=_('hostname'), command=lambda: self.sort_ip_table('hostname'))
        self.ipam_ip_tree.heading('mac_address', text=_('mac_address'), command=lambda: self.sort_ip_table('mac_address'))
        self.ipam_ip_tree.heading('description', text=_('description'), command=lambda: self.sort_ip_table('description'))
        self.ipam_ip_tree.heading('allocated_at', text=_('allocated_time'), command=lambda: self.sort_ip_table('allocated_at'))
        self.ipam_ip_tree.heading('expiry_date', text=_('expiry_date'), command=lambda: self.sort_ip_table('expiry_date'))
        
        # 排序状态变量
        self.sort_column = 'ip_address'
        self.sort_order = 'asc'
        
        # 调整列宽，优化空间利用率，确保能显示所有列
        # 为每列设置合适的初始宽度和最小宽度，stretch=True让列能根据容器调整
        # 总初始宽度约为 100+70+100+120+130+130=650，适应大部分窗口宽度
        self.ipam_ip_tree.column('ip_address', width=100, minwidth=80, stretch=True)
        self.ipam_ip_tree.column('status', width=70, minwidth=50, stretch=True)
        self.ipam_ip_tree.column('hostname', width=100, minwidth=80, stretch=True)
        self.ipam_ip_tree.column('mac_address', width=120, minwidth=100, stretch=False)
        self.ipam_ip_tree.column('description', width=120, minwidth=90, stretch=True)
        self.ipam_ip_tree.column('allocated_at', width=130, minwidth=100, stretch=True)
        self.ipam_ip_tree.column('expiry_date', width=130, minwidth=100, stretch=True)
        
        # 设置合适的初始宽度，确保能显示所有列，同时避免不必要的水平滚动条
        
        # 创建垂直滚动条 - 在table_container上创建
        ip_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL)
        
        # 使用通用方法创建带自动隐藏功能的垂直滚动条
        # 和IPv6查询结果表保持一致
        self.create_scrollable_treeview(table_container, self.ipam_ip_tree, ip_scrollbar)
        
        # 移除水平滚动条，只保留垂直滚动条
        # 配置Treeview不使用水平滚动
        self.ipam_ip_tree.configure(xscrollcommand=None)
        
        # 为IP地址表格添加右键菜单
        self.ipam_ip_tree.bind('<Button-3>', self.on_ip_tree_right_click)
        # 为IP地址表格添加双击编辑功能 - 支持内联编辑
        # 注册IP地址表的内联编辑配置
        self.register_inline_edit_config('ipam_ip', {
            'editable_columns': [0, 1, 2, 3, 4, 6],  # 允许编辑IP地址、状态、主机名、MAC地址、描述和过期日期
            'column_types': {
                1: 'combobox',  # 状态列使用下拉框
                6: 'datepicker'  # 过期日期列使用日期选择器
            },
            'combobox_values': {
                1: [_('released'), _('allocated'), _('reserved')]  # 状态选项
            }
        })
        
        # 注册IP地址表的内联编辑处理器
        self.register_inline_edit_handler('ipam_ip', {
            'get_row_data': self._get_ip_row_data,
            'validate': self._validate_ip_edit,
            'save': self._save_ip_edit
        })
        
        # 绑定双击事件
        self.ipam_ip_tree.bind('<Double-1>', lambda event: self.on_generic_tree_double_click(self.ipam_ip_tree, 'ipam_ip', event))
        # 绑定左键点击事件，实现取消选择功能
        self.ipam_ip_tree.bind('<Button-1>', self.on_ipam_ip_click)
        
        # 配置Treeview样式，包括斑马条纹
        self.configure_treeview_styles(self.ipam_ip_tree)
        # 添加斑马纹样式
        self.update_table_zebra_stripes(self.ipam_ip_tree)
        
        # 统计和图表标签页（使用浅紫色样式，与标签颜色匹配）
        stats_frame = ttk.Frame(self.ipam_notebook.content_area, padding="5", style=self.ipam_notebook.light_purple_style)
        
        # 配置统计区域的网格布局（仅一列，让stats_container自适应整个区域）
        stats_frame.grid_rowconfigure(0, weight=1)
        stats_frame.grid_columnconfigure(0, weight=1)
        
        # 添加标签页（调整顺序：地址管理 → 网络拓扑 → 统计分析，颜色次序遵循规范：浅蓝色→浅绿色→浅紫色）
        self.ipam_notebook.add_tab(_('ip_address_management'), ip_management_frame, "#e3f2fd")  # 浅蓝色（第一个位置）
        
        # 网络拓扑标签页（第二个位置，使用浅绿色）
        topology_frame = ttk.Frame(self.ipam_notebook.content_area, padding="5", style=self.ipam_notebook.light_green_style)
        self.ipam_notebook.add_tab(_('network_topology'), topology_frame, "#e8f5e9")  # 浅绿色（第二个位置）
        
        # 配置网络拓扑区域的网格布局
        topology_frame.grid_rowconfigure(0, weight=1)
        topology_frame.grid_columnconfigure(0, weight=1)
        
        # 创建拓扑可视化器，直接放在拓扑页面上
        self.topology_visualizer = NetworkTopologyVisualizer(topology_frame)
        
        # 统计分析标签页（第三个位置，使用浅紫色）
        self.ipam_notebook.add_tab(_('statistical_analysis'), stats_frame, "#f3e5f5")  # 浅紫色（第三个位置）
        
        # 创建父容器，包裹统计信息和图表（自适应整个stats_frame）
        stats_container = ttk.Frame(stats_frame, padding="10")
        stats_container.grid(row=0, column=0, sticky="nsew")
        stats_container.rowconfigure(0, weight=1)
        stats_container.columnconfigure(0, weight=1)
        stats_container.columnconfigure(1, weight=1)
        
        # 统计信息框架（增加内边距）
        stats_info_frame = ttk.Frame(stats_container, padding="15")
        stats_info_frame.grid(row=0, column=0, padx=(0, 20), pady=5, sticky="nsew")
        
        # 统计数据 - 分两列显示
        self.stats_labels = {}
        stats_items = [
            ('total_networks', _('stats_total_networks')),
            ('total_ips', _('stats_total_ips')),
            ('ipv4_networks', _('stats_ipv4_networks')),
            ('ipv4_ips', _('stats_ipv4_ips')),
            ('ipv6_networks', _('stats_ipv6_networks')),
            ('ipv6_ips', _('stats_ipv6_ips')),
            ('allocated_ips', _('stats_allocated_ips')),
            ('reserved_ips', _('stats_reserved_ips')),
            ('available_ips', _('stats_released_ips')),
            ('expired_ips', _('stats_expired_ips')),
            ('expiring_ips', _('stats_expiring_ips')),
            ('named_ips', _('stats_named_ips')),
            ('vlan_count', _('stats_vlan_count')),
            ('utilization_rate', _('stats_utilization_rate'))
        ]
        
        for i, (key, label) in enumerate(stats_items):
            row = i // 2
            column = i % 2 * 2
            ttk.Label(stats_info_frame, text=label).grid(row=row, column=column, sticky="e", pady=3, padx=(0, 10))
            self.stats_labels[key] = ttk.Label(stats_info_frame, text="0")
            self.stats_labels[key].grid(row=row, column=column + 1, sticky="w", pady=3, padx=(5, 30))
        
        # 添加可视化图表
        chart_frame = ttk.LabelFrame(stats_container, text=_('ip_usage_statistics'), padding="10")
        chart_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        
        # 创建Canvas用于绘制饼图（使用应用标准背景色，增加内边距）
        self.stats_canvas = tk.Canvas(chart_frame, width=400, height=300, bg="#2c3e50", highlightthickness=0)
        self.stats_canvas.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # 添加画布大小变化事件监听器
        def on_canvas_configure(event):
            # 当画布尺寸变化时，重新绘制图表
            self.refresh_ipam_stats()
        
        self.stats_canvas.bind('<Configure>', on_canvas_configure)
        
        # 刷新统计数据
        self.refresh_ipam_stats()
        
        # 初始化IPAM数据
        self.refresh_ipam_networks()
        # 添加样例数据
        self.add_ipam_sample_data()
        
        # 自动选择第一个网段并刷新网络拓扑（延迟执行，避免初始化闪现）
        self.root.after(100, self._auto_select_first_network)
    
    def refresh_ipam_networks(self):
        """刷新IPAM网络列表，实现分层显示"""
        # 保存展开状态
        expanded_items = set()

        def save_expanded_state(item):
            if self.ipam_network_tree.item(item, 'open'):
                expanded_items.add(self.ipam_network_tree.item(item, 'values')[0])
            for child in self.ipam_network_tree.get_children(item):
                save_expanded_state(child)
        
        for item in self.ipam_network_tree.get_children():
            save_expanded_state(item)
        
        # 保存选中状态
        selected_network = None
        selected_items = self.ipam_network_tree.selection()
        if selected_items:
            selected_network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        
        # 清空现有网络列表
        for item in self.ipam_network_tree.get_children():
            self.ipam_network_tree.delete(item)
        
        # 获取所有网络
        networks = self.ipam.get_all_networks()
        
        # 构建网段层次结构
        full_hierarchy, top_level_networks = self._build_network_hierarchy(networks)
        
        # 递归插入网段到Treeview
        self._insert_networks_recursively('', top_level_networks, full_hierarchy)
        
        # 恢复展开状态
        def restore_expanded_state(item):
            network_value = self.ipam_network_tree.item(item, 'values')[0]
            if network_value in expanded_items:
                self.ipam_network_tree.item(item, open=True)
            for child in self.ipam_network_tree.get_children(item):
                restore_expanded_state(child)
        
        for item in self.ipam_network_tree.get_children():
            restore_expanded_state(item)
        
        # 恢复选中状态
        if selected_network:
            def find_and_select(item):
                network_value = self.ipam_network_tree.item(item, 'values')[0]
                if network_value == selected_network:
                    self.ipam_network_tree.selection_set(item)
                    return True
                for child in self.ipam_network_tree.get_children(item):
                    if find_and_select(child):
                        return True
                return False
            
            for item in self.ipam_network_tree.get_children():
                find_and_select(item)
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(self.ipam_network_tree)
    
    def _auto_select_first_network(self):
        """自动选择第一个网段并刷新网络拓扑"""
        # 获取所有顶层网段
        top_level_items = self.ipam_network_tree.get_children()
        if top_level_items:
            # 选择第一个网段
            first_item = top_level_items[0]
            self.ipam_network_tree.selection_set(first_item)
            
            # 获取网段信息并刷新相关视图
            network = self.ipam_network_tree.item(first_item, 'values')[0]
            if network:
                self.refresh_ipam_ips(network)
                self.refresh_visualization(network)
    
    def _build_network_hierarchy(self, networks):
        """构建网段层次结构
        
        Args:
            networks: 网段列表
            
        Returns:
            tuple: (full_hierarchy, top_level_networks)
                full_hierarchy: 完整的网段层次结构
                top_level_networks: 只包含顶层网络的层次结构
        """
        
        # 先将所有网段转换为ipaddress对象并排序
        network_objects = []
        for network in networks:
            try:
                net_obj = ipaddress.ip_network(network['network'], strict=False)
                network_objects.append((net_obj, network))
            except ValueError:
                # 跳过无效的网段
                pass
        
        # 按网络大小排序，大网络在前
        network_objects.sort(key=lambda x: (x[0].version, x[0].prefixlen, x[0].network_address))
        
        # 构建层次结构
        hierarchy = {}
        # 首先将所有网络添加到层次结构中
        for net_obj, network_data in network_objects:
            network_str = str(net_obj)
            hierarchy[network_str] = {'data': network_data, 'children': []}
        
        # 然后构建父子关系
        for net_obj, network_data in network_objects:
            network_str = str(net_obj)
            
            # 查找父网络
            parent = None
            for existing_net, existing_data in hierarchy.items():
                try:
                    existing_net_obj = ipaddress.ip_network(existing_net, strict=False)
                    # 检查子网段是否完全包含在父网段中
                    if net_obj.prefixlen > existing_net_obj.prefixlen:
                        # 检查网络地址是否在父网段范围内
                        if net_obj.network_address in existing_net_obj:
                            # 找到一个父网络，检查是否是最直接的父网络
                            if parent is None:
                                parent = existing_net
                            else:
                                # 检查哪个父网络更具体（prefixlen更大）
                                parent_obj = ipaddress.ip_network(parent, strict=False)
                                if existing_net_obj.prefixlen > parent_obj.prefixlen:
                                    parent = existing_net
                except ValueError:
                    pass
            
            # 添加到层次结构
            if parent:
                hierarchy[parent]['children'].append(network_str)
        
        # 过滤掉作为子节点的网络，只保留顶层网络
        top_level_networks = {}
        for network_str, network_data in hierarchy.items():
            # 检查是否是顶层网络（没有被其他网络作为子节点）
            is_top_level = True
            for existing_net, existing_data in hierarchy.items():
                if network_str in existing_data['children']:
                    is_top_level = False
                    break
            if is_top_level:
                top_level_networks[network_str] = network_data
        
        return hierarchy, top_level_networks
    
    def _insert_networks_recursively(self, parent_item, network_hierarchy, full_hierarchy):
        """递归插入网段到Treeview
        
        Args:
            parent_item: 父Treeview项ID
            network_hierarchy: 当前层次的网段层次结构
            full_hierarchy: 完整的网段层次结构
        """
        # 对网段进行排序，确保显示顺序一致
        sorted_networks = sorted(network_hierarchy.keys())
        
        for network_str in sorted_networks:
            network_data = network_hierarchy[network_str]['data']
            children = network_hierarchy[network_str]['children']
            
            # 格式化时间戳，只显示到秒
            created_at = network_data['created_at']
            formatted_time = self._format_datetime(created_at)
            
            # 插入网段
            item = self.ipam_network_tree.insert(parent_item, tk.END, text=network_data['network'], values=(
                network_data['network'],
                network_data.get('description', ''),
                network_data.get('vlan', ''),
                formatted_time,
                network_data['ip_count']
            ))
            
            # 递归插入子网段
            if children:
                # 构建子网段的层次结构
                child_hierarchy = {}
                for child_net in children:
                    # 从完整的层次结构中查找子网段的完整数据
                    if child_net in full_hierarchy:
                        child_hierarchy[child_net] = full_hierarchy[child_net]
                
                # 递归插入子网络
                self._insert_networks_recursively(item, child_hierarchy, full_hierarchy)

    def refresh_ipam_ips(self, network):
        """刷新IP地址列表"""
        # 清空现有IP地址列表
        for item in self.ipam_ip_tree.get_children():
            self.ipam_ip_tree.delete(item)
        
        # 直接调用get_network_ips获取网络及其所有子网络的IP地址
        # 由于get_network_ips已经会返回网络及其所有子网络的IP地址
        # 所以不需要再单独处理子网络
        all_ips = self.ipam.get_network_ips(network)
        
        # 使用新的排序功能
        sorted_ips = self._sort_ip_list(all_ips)
        
        for ip in sorted_ips:
            status_text = ip['status']
            # 翻译状态文本
            if status_text == 'reserved':
                status_text = _('reserved')
            elif status_text == 'released':
                status_text = _('released')
            elif status_text == 'allocated':
                status_text = _('allocated')
            
            # 格式化分配时间，只显示到秒
            allocated_at = ip.get('allocated_at', '')
            formatted_allocated_at = self._format_datetime(allocated_at)
            
            # 格式化过期日期，只显示日期部分
            expiry_date = ip.get('expiry_date', '')
            formatted_expiry_date = self._format_datetime(expiry_date, "%Y-%m-%d")
            
            # 使用数据库记录ID作为树项的tags，确保能可靠获取数据库ID
            db_record_id = ip.get('id', None)
            tags = (f'dbid_{db_record_id}',) if db_record_id is not None else ()
            # 生成唯一的iid，使用数据库ID或IP地址+描述的组合
            if db_record_id is not None:
                iid = f'rec_{db_record_id}'
            else:
                iid = f'ip_{ip["ip_address"]}_{ip.get("description", "")}_{ip.get("status", "")}'
            self.ipam_ip_tree.insert('', tk.END, iid=iid, tags=tags, values=(
                ip['ip_address'],
                status_text,
                ip.get('hostname', ''),
                ip.get('mac_address', ''),
                ip.get('description', ''),
                formatted_allocated_at,
                formatted_expiry_date
            ))
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(self.ipam_ip_tree)
    
    def batch_release_ip(self):
        """批量释放IP地址"""
        # 检查是否选择了IP地址
        selected_ip_items = self.ipam_ip_tree.selection()
        if not selected_ip_items:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 批量释放IP地址
        success_count = 0
        error_count = 0
        for item in selected_ip_items:
            ip_address = self.ipam_ip_tree.item(item, 'values')[0]
            status = self.ipam_ip_tree.item(item, 'values')[1]
            # 只有状态为"已分配"或"已保留"的IP地址才能被释放
            if status in [_('allocated'), _('reserved')]:
                success, message = self.ipam.release_ip(ip_address)
                if success:
                    success_count += 1
                else:
                    error_count += 1
                    # 记录释放失败的原因
                    print(f"释放IP {ip_address} 失败: {message}")
            else:
                # 跳过已经是已释放状态的IP地址
                print(f"跳过IP {ip_address}，状态为: {status}")
        
        # 显示结果
        if success_count > 0 and error_count > 0:
            # 既有成功又有失败
            self.show_info(_('info'), f"成功释放 {success_count} 个IP地址，释放失败 {error_count} 个IP地址")
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        elif success_count > 0:
            # 全部成功
            self.show_info(_('success'), f"{_('successfully_released_ips', count=success_count)}")
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        elif error_count > 0:
            # 全部失败
            self.show_error(_('error'), f"释放失败 {error_count} 个IP地址")
        else:
            # 检查是否所有选中的IP地址都是可用状态
            all_available = True
            for item in selected_ip_items:
                status = self.ipam_ip_tree.item(item, 'values')[1]
                if status not in [_('released')]:
                    all_available = False
                    break
            
            if all_available:
                self.show_info(_('hint'), _('ip_already_available_no_need_to_release'))
            else:
                self.show_error(_('error'), _('failed_to_release_ip'))
    

    
    def _match_search_pattern(self, text, keyword, search_mode):
        """根据搜索模式匹配文本
        
        Args:
            text: 要匹配的文本
            keyword: 搜索关键词
            search_mode: 搜索模式（精确匹配、包含、正则表达式）
            
        Returns:
            bool: 是否匹配成功
        """
        try:
            text = str(text).lower()
            keyword = keyword.lower()
            
            if search_mode == "精确匹配":
                return text == keyword
            elif search_mode == "包含":
                return keyword in text
            elif search_mode == "正则表达式":
                import re
                return bool(re.search(keyword, text))
            else:
                return keyword in text
        except Exception:
            return False
    
    def _get_localized_status(self, status):
        """获取本地化的状态文本"""
        if status == 'reserved':
            return _('reserved')
        elif status == 'released':
            return _('released')
        elif status == 'allocated':
            return _('allocated')
        return status
    
    
    def apply_filter(self):
        """应用过滤和排序"""
        from datetime import datetime
        selected_items = self.ipam_network_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_network'))
            return
        
        network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        status = self.ipam_status_filter.get()
        
        # 清空现有IP地址列表
        for item in self.ipam_ip_tree.get_children():
            self.ipam_ip_tree.delete(item)
        
        # 获取网络中的所有IP地址
        ips = self.ipam.get_network_ips(network)
        
        # 过滤IP地址
        filtered_ips = []
        
        # 获取过期日期筛选选项
        expiry_filter = self.ipam_expiry_filter.get()
        now = datetime.now()
        
        # 获取搜索关键词和范围
        search_text = self.ipam_search_entry.get().strip()
        search_scope = self.search_scope.get()
        search_mode = self.search_mode.get()
        
        for ip in ips:
            # 按状态过滤
            if status != _('all'):
                # 使用类方法获取状态映射表
                status_map = self._get_status_map()
                if status in status_map:
                    if ip['status'] != status_map[status]:
                        continue
                elif ip['status'] != status:
                    continue
            
            # 按过期日期过滤
            if expiry_filter != _('all'):
                expiry_date = ip.get('expiry_date')
                if not expiry_date:
                    if expiry_filter != _('not_expired'):
                        continue
                else:
                    try:
                        # 解析过期日期
                        if 'T' in expiry_date:
                            # ISO格式
                            exp_date = datetime.fromisoformat(expiry_date)
                        else:
                            # 普通格式
                            exp_date = datetime.strptime(expiry_date, "%Y-%m-%d %H:%M:%S")
                        
                        # 根据筛选选项过滤
                        if expiry_filter == _('expired'):
                            if exp_date >= now:
                                continue
                        elif expiry_filter == _('expiring_soon'):
                            # 7天内过期
                            seven_days_later = now + timedelta(days=7)
                            if exp_date < now or exp_date > seven_days_later:
                                continue
                        elif expiry_filter == _('not_expired'):
                            if exp_date < now:
                                continue
                    except (ValueError, TypeError):
                        # 日期格式错误，跳过
                        if expiry_filter != _('not_expired'):
                            continue
            
            # 按搜索关键词过滤
            if search_text:
                # 支持多关键词搜索，按空格分割
                keywords = search_text.split()
                match = False
                
                # 检查IP地址
                ip_match = self._match_search_pattern(ip['ip_address'], ' '.join(keywords), search_mode)
                # 检查主机名
                hostname_match = self._match_search_pattern(ip.get('hostname', ''), ' '.join(keywords), search_mode)
                # 检查描述
                desc_match = self._match_search_pattern(ip.get('description', ''), ' '.join(keywords), search_mode)
                # 检查MAC地址
                mac_match = self._match_search_pattern(ip.get('mac_address', ''), ' '.join(keywords), search_mode)
                
                if search_scope == _('all'):
                    match = ip_match or hostname_match or desc_match or mac_match
                elif search_scope == _('ip_address'):
                    match = ip_match
                elif search_scope == _('hostname'):
                    match = hostname_match
                elif search_scope == _('description'):
                    match = desc_match
                elif search_scope == _('mac_address'):
                    match = mac_match
                
                if not match:
                    continue
            
            filtered_ips.append(ip)
        
        # 排序
        sorted_ips = self._sort_ip_list(filtered_ips)
        
        # 显示过滤结果
        for ip in sorted_ips:
            status_text = ip['status']
            # 翻译状态文本
            if status_text == 'reserved':
                status_text = _('reserved')
            elif status_text == 'released':
                status_text = _('released')
            elif status_text == 'allocated':
                status_text = _('allocated')
            
            # 格式化分配时间，只显示到秒，与初始加载保持一致
            allocated_at = ip.get('allocated_at', '')
            try:
                # 尝试解析ISO格式的时间戳
                from datetime import datetime
                if 'T' in allocated_at:
                    # ISO格式: 2023-12-31T23:59:59.123456
                    dt = datetime.fromisoformat(allocated_at)
                else:
                    # 普通格式: 2023-12-31 23:59:59.123456
                    dt = datetime.strptime(allocated_at, "%Y-%m-%d %H:%M:%S.%f")
                # 格式化为只显示到秒
                formatted_allocated_at = dt.strftime("%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                # 如果解析失败，使用原始时间戳
                formatted_allocated_at = allocated_at
            
            # 格式化过期日期，只显示日期部分
            expiry_date = ip.get('expiry_date', '')
            formatted_expiry_date = expiry_date
            try:
                from datetime import datetime
                if expiry_date:
                    if 'T' in expiry_date:
                        # ISO格式: 2023-12-31T23:59:59
                        dt = datetime.fromisoformat(expiry_date)
                    elif ' ' in expiry_date:
                        # 普通格式: 2023-12-31 23:59:59
                        dt = datetime.strptime(expiry_date, "%Y-%m-%d %H:%M:%S")
                    else:
                        # 只有日期部分: 2023-12-31
                        dt = datetime.strptime(expiry_date, "%Y-%m-%d")
                    # 格式化为只显示日期部分
                    formatted_expiry_date = dt.strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                # 如果解析失败，使用原始日期
                pass
            
            # 使用数据库记录ID作为树项的tags，确保能可靠获取数据库ID
            db_record_id = ip.get('id', None)
            tags = (f'dbid_{db_record_id}',) if db_record_id is not None else ()
            # 生成唯一的iid，使用数据库ID或IP地址+描述的组合
            if db_record_id is not None:
                iid = f'rec_{db_record_id}'
            else:
                iid = f'ip_{ip["ip_address"]}_{ip.get("description", "")}_{ip.get("status", "")}'
            self.ipam_ip_tree.insert('', tk.END, iid=iid, tags=tags, values=(
                ip['ip_address'],
                status_text,
                ip.get('hostname', ''),
                ip.get('mac_address', ''),
                ip.get('description', ''),
                formatted_allocated_at,
                formatted_expiry_date
            ))
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(self.ipam_ip_tree)
    
    def _sort_ip_list(self, ips):
        """对IP地址列表进行排序"""
        def ip_key(ip_item):
            try:
                addr = ipaddress.ip_address(ip_item['ip_address'])
                return (addr.version, int(addr))
            except ValueError:
                return (0, ip_item['ip_address'])
        return sorted(ips, key=ip_key)
    
    def reset_search(self):
        """重置搜索"""
        # 清空搜索输入框
        self.ipam_search_entry.delete(0, tk.END)
        # 重置搜索范围
        if hasattr(self, 'search_scope'):
            self.search_scope.set(_('all'))
        # 重置状态过滤器
        if hasattr(self, 'ipam_status_filter'):
            self.ipam_status_filter.set(_('all'))
        # 重置搜索模式
        if hasattr(self, 'search_mode'):
            self.search_mode.set(_('contains'))
        # 重新应用过滤
        self.apply_filter()
        if hasattr(self, 'ipam_expiry_filter'):
            self.ipam_expiry_filter.set("全部")
        # 重置分配时间过滤器
        if hasattr(self, 'ipam_time_filter'):
            self.ipam_time_filter.set("全部")
        
        # 刷新IP地址列表
        selected_items = self.ipam_network_tree.selection()
        if selected_items:
            network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
            self.refresh_ipam_ips(network)
    
    def filter_ip_by_status(self):
        """按状态过滤IP地址"""
        selected_items = self.ipam_network_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_network'))
            return
        
        network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        status = self.ipam_status_filter.get()
        
        # 获取网络中的所有IP地址
        ips = self.ipam.get_network_ips(network)
        
        # 清空当前列表
        for item in self.ipam_ip_tree.get_children():
            self.ipam_ip_tree.delete(item)
        
        # 过滤并显示IP地址
        for ip in ips:
            if status == "全部" or ip['status'] == status:
                # 翻译状态文本
                status_text = ip['status']
                if status_text == 'reserved':
                    status_text = _('reserved')
                elif status_text == 'released':
                    status_text = _('released')
                elif status_text == 'allocated':
                    status_text = _('allocated')
                
                # 格式化分配时间和过期日期
                allocated_at = ip.get('allocated_at', '')
                formatted_allocated_at = self._format_datetime(allocated_at)
                
                expiry_date = ip.get('expiry_date', '')
                formatted_expiry_date = self._format_datetime(expiry_date, "%Y-%m-%d")
                
                # 使用记录ID作为树项的ID
                # 使用数据库记录ID作为树项的tags，确保能可靠获取数据库ID
                db_record_id = ip.get('id', None)
                tags = (f'dbid_{db_record_id}',) if db_record_id is not None else ()
                # 生成唯一的iid，使用数据库ID或IP地址+描述的组合
                if db_record_id is not None:
                    iid = f'rec_{db_record_id}'
                else:
                    iid = f'ip_{ip["ip_address"]}_{ip.get("description", "")}_{ip.get("status", "")}'
                self.ipam_ip_tree.insert('', tk.END, iid=iid, tags=tags, values=(
                    ip['ip_address'],
                    status_text,
                    ip.get('hostname', ''),
                    ip.get('mac_address', ''),
                    ip.get('description', ''),
                    formatted_allocated_at,
                    formatted_expiry_date
                ))
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(self.ipam_ip_tree)
    
    def check_ip_conflicts(self):
        """检查IP冲突"""
        selected_items = self.ipam_network_tree.selection()
        
        # 检查IP冲突
        try:
            if not selected_items:
                # 没有选中网段，检查所有网段
                # 收集所有IP地址记录，避免重复检查
                all_ip_records = {}
                
                networks = self.ipam.get_all_networks()
                for network_info in networks:
                    network = network_info['network']
                    # 获取网络中的所有IP地址
                    ips = self.ipam.get_network_ips(network)
                    
                    for ip in ips:
                        if ip['id'] not in all_ip_records:
                            all_ip_records[ip['id']] = {
                                'id': ip['id'],
                                'network': network,
                                'ip_address': ip['ip_address'],
                                'status': ip['status'],
                                'hostname': ip.get('hostname', ''),
                                'mac_address': ip.get('mac_address', ''),
                                'description': ip.get('description', ''),
                                'allocated_at': ip.get('allocated_at', ''),
                                'expiry_date': ip.get('expiry_date', '')
                            }
                
                # 统计每个IP地址的分配情况
                ip_counts = {}
                for ip_record in all_ip_records.values():
                    ip_addr = ip_record['ip_address']
                    if ip_addr in ip_counts:
                        ip_counts[ip_addr].append(ip_record)
                    else:
                        ip_counts[ip_addr] = [ip_record]
                
                # 找出冲突的IP地址（被分配多次的）
                all_conflicts = []
                for ip_addr, ip_list in ip_counts.items():
                    if len(ip_list) > 1:
                        # 为每个冲突的IP地址记录添加到冲突列表
                        for ip in ip_list:
                            all_conflicts.append(ip)
                
                # 按IP地址排序
                conflicts = self._sort_ip_list(all_conflicts)
                selected_networks = []  # 空列表表示检查所有网段
            else:
                # 有选中网段，检查所有选中的网段
                # 收集所有IP地址记录，避免重复检查
                all_ip_records = {}
                selected_networks = []
                
                for item in selected_items:
                    network = self.ipam_network_tree.item(item, 'values')[0]
                    selected_networks.append(network)
                    # 获取网络中的所有IP地址
                    ips = self.ipam.get_network_ips(network)
                    
                    for ip in ips:
                        if ip['id'] not in all_ip_records:
                            all_ip_records[ip['id']] = {
                                'id': ip['id'],
                                'network': network,
                                'ip_address': ip['ip_address'],
                                'status': ip['status'],
                                'hostname': ip.get('hostname', ''),
                                'mac_address': ip.get('mac_address', ''),
                                'description': ip.get('description', ''),
                                'allocated_at': ip.get('allocated_at', ''),
                                'expiry_date': ip.get('expiry_date', '')
                            }
                
                # 统计每个IP地址的分配情况
                ip_counts = {}
                for ip_record in all_ip_records.values():
                    ip_addr = ip_record['ip_address']
                    if ip_addr in ip_counts:
                        ip_counts[ip_addr].append(ip_record)
                    else:
                        ip_counts[ip_addr] = [ip_record]
                
                # 找出冲突的IP地址（被分配多次的）
                conflicts = []
                for ip_addr, ip_list in ip_counts.items():
                    if len(ip_list) > 1:
                        # 为每个冲突的IP地址记录添加到冲突列表
                        for ip in ip_list:
                            conflicts.append(ip)
                
                # 按IP地址排序
                conflicts = self._sort_ip_list(conflicts)
            
            if not conflicts:
                self.show_info(_('hint'), _('no_ip_conflicts_found'))
                return
            
            # 显示冲突信息窗口，传递选中的网段信息
            self.show_conflict_window(conflicts, selected_networks)
        except Exception as e:
            self.show_error(_('error'), f"{_('check_conflict_failed')}: {str(e)}")
            return
    
    def show_conflict_window(self, conflicts, selected_networks=None):
        """显示IP冲突结果窗口"""
        # 确保selected_networks是列表
        if selected_networks is None:
            selected_networks = []
        
        # 创建窗口
        dialog = ComplexDialog(self.root, _('ip_conflicts'), 800, 600, resizable=True, modal=True)
        
        # 设置字体
        font_family, font_size = get_current_font_settings()
        
        # 创建主框架
        main_frame = ttk.Frame(dialog.content_frame, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 添加统计信息
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 显示统计信息
        stats_label = ttk.Label(stats_frame, text=_('found_ip_conflicts_count', count=len(conflicts)), font=(font_family, font_size))
        stats_label.pack(anchor=tk.W)
        
        # 创建结果树 - 不显示ID和网络地址列，使用item的iid存储ID
        columns = ('ip_address', 'status', 'hostname', 'mac_address', 'description', 'allocated_at', 'expiry_date')
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', selectmode='extended')
        
        # 设置列宽
        tree.column('ip_address', width=90, stretch=False)
        tree.column('status', width=60, stretch=False)
        tree.column('hostname', width=100, stretch=False)
        tree.column('mac_address', width=110, stretch=False)
        tree.column('description', width=100, stretch=True)
        tree.column('allocated_at', width=110, stretch=False)
        tree.column('expiry_date', width=90, stretch=False)
        
        # 设置表头
        tree.heading('ip_address', text=_('ip_address'))
        tree.heading('status', text=_('status'))
        tree.heading('hostname', text=_('hostname'))
        tree.heading('mac_address', text=_('mac_address'))
        tree.heading('description', text=_('description'))
        tree.heading('allocated_at', text=_('allocated_at'))
        tree.heading('expiry_date', text=_('expiry_date'))
        
        # 添加冲突数据，使用ID作为item的iid
        for conflict in conflicts:
            # 翻译状态值
            status = conflict['status']
            if status == 'released':
                translated_status = _('released')
            elif status == 'allocated':
                translated_status = _('allocated')
            elif status == 'reserved':
                translated_status = _('reserved')
            else:
                translated_status = status
            
            # 处理分配时间，只显示日期部分
            allocated_at = conflict.get('allocated_at', '')
            if allocated_at:
                if ' ' in allocated_at:
                    allocated_at = allocated_at.split(' ')[0]
                elif 'T' in allocated_at:
                    allocated_at = allocated_at.split('T')[0]
            
            # 处理过期日期，只显示日期部分
            expiry_date = conflict.get('expiry_date', '')
            if expiry_date:
                if ' ' in expiry_date:
                    expiry_date = expiry_date.split(' ')[0]
                elif 'T' in expiry_date:
                    expiry_date = expiry_date.split('T')[0]
            
            tree.insert('', tk.END, iid=str(conflict['id']), values=(
                conflict['ip_address'],
                translated_status,
                conflict['hostname'],
                conflict.get('mac_address', ''),
                conflict['description'],
                allocated_at,
                expiry_date
            ))
        
        # 配置斑马纹样式并应用
        self.configure_treeview_styles(tree)
        self.update_table_zebra_stripes(tree)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        
        # 创建滚动条回调函数，实现自动隐藏功能
        def scrollbar_callback(*args):
            scrollbar.set(*args)
            if float(args[1]) - float(args[0]) >= 1.0 - 1e-9:
                scrollbar.pack_forget()
            else:
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 配置滚动条和Treeview
        tree.configure(yscroll=scrollbar_callback)
        scrollbar.config(command=tree.yview)
        
        # 使用pack布局
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 初始调用回调函数
        scrollbar_callback(0.0, 1.0)
        
        # 处理冲突按钮
        def handle_conflict():
            selected_items = tree.selection()
            if not selected_items:
                self.show_info(_('hint'), _('please_select_conflict'))
                return
            
            # 收集所有选中的记录
            selected_records = []
            for item in selected_items:
                item_values = tree.item(item, 'values')
                if len(item_values) >= 1:
                    selected_records.append({
                        'id': item,  # 使用item的iid作为ID
                        'ip_address': item_values[0]  # IP地址现在是第一个值
                    })
            
            # 打开冲突处理对话框（处理第一个选中的记录）
            if selected_records:
                first_record = selected_records[0]
                self.show_conflict_resolution_dialog(dialog, tree, first_record['id'], first_record['ip_address'], conflicts, selected_records, selected_networks)
        
        # 使用dialog.button_frame放置按钮
        # 左侧按钮
        ttk.Button(dialog.button_frame, text=_('handle_conflict'), command=handle_conflict).pack(side=tk.LEFT, padx=5)
        ttk.Button(dialog.button_frame, text=_('refresh'), command=lambda: self.refresh_conflicts(tree, conflicts, selected_networks)).pack(side=tk.LEFT, padx=5)
        # 右侧关闭按钮
        ttk.Button(dialog.button_frame, text=_('close'), command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
        
        # 显示对话框
        dialog.show()
    
    def show_conflict_resolution_dialog(self, parent_dialog, tree, ip_id, ip_address, conflicts, selected_records=None, selected_networks=None):
        """显示冲突处理对话框
        
        Args:
            parent_dialog: 父对话框
            tree: 冲突列表树形控件
            ip_id: 选中的IP记录ID
            ip_address: IP地址
            conflicts: 冲突列表
            selected_records: 选中的记录列表
            selected_networks: 选中的网段列表
        """
        if selected_records is None:
            selected_records = [{'id': ip_id, 'ip_address': ip_address}]
        # 创建对话框，使用parent_dialog.dialog作为父窗口
        dialog = ComplexDialog(parent_dialog.dialog, _('resolve_conflict'), 400, 260, resizable=False, modal=True)
        
        # 获取当前字体设置
        font_family, font_size = get_current_font_settings()
        
        # 主框架
        main_frame = ttk.Frame(dialog.content_frame, padding=(15, 0))
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 说明文字
        if len(selected_records) > 1:
            ttk.Label(main_frame, text=f"{_('selected_ips')}: {len(selected_records)} 条记录", 
                      font=(font_family, font_size, 'bold')).pack(anchor='w', pady=(0, 10))
        else:
            ttk.Label(main_frame, text=f"{_('selected_ip')}: {ip_address}", 
                      font=(font_family, font_size, 'bold')).pack(anchor='w', pady=(0, 10))
        
        # 操作选项框架
        action_frame = ttk.LabelFrame(main_frame, text=_('select_action'), padding=(25,10))
        action_frame.pack(fill=tk.X, pady=(0, 0))
        
        # 操作选项
        action_var = tk.StringVar(value='keep')
        
        # 创建单选按钮样式
        radiobutton_style = ttk.Style()
        radiobutton_style.configure("Custom.TRadiobutton", font=(font_family, font_size))
        
        ttk.Radiobutton(action_frame, text=_('keep_this_record'), 
                        variable=action_var, value='keep', style="Custom.TRadiobutton").pack(anchor='w', pady=5)
        
        ttk.Radiobutton(action_frame, text=_('delete_this_record'), 
                        variable=action_var, value='delete', style="Custom.TRadiobutton").pack(anchor='w', pady=5)
        
        ttk.Radiobutton(action_frame, text=_('release_this_ip'), 
                        variable=action_var, value='release', style="Custom.TRadiobutton").pack(anchor='w', pady=5)
        
        def execute_action():
            action = action_var.get()

            if action == 'keep':
                # 保留选中记录，删除该IP的其他冲突记录
                self.keep_selected_record_and_delete_others(ip_id, ip_address)
            elif action == 'delete':
                # 删除所有选中的记录
                deleted_count = 0
                for record in selected_records:
                    success, message = self.ipam.delete_ip_by_id(int(record['id']))
                    if success:
                        deleted_count += 1
                if deleted_count > 0:
                    self.show_info(_('success'), f"已删除 {deleted_count} 条记录")
                else:
                    self.show_error(_('error'), "删除记录失败")
            elif action == 'release':
                # 释放IP地址
                success, message = self.ipam.release_ip(ip_address)
                if success:
                    self.show_info(_('success'), message)
                else:
                    self.show_error(_('error'), message)

            dialog.destroy()
            # 刷新冲突列表 - 重新检查冲突
            self.refresh_conflicts(tree, conflicts, selected_networks)
            
            # 刷新地址管理表
            try:
                selected_items = self.ipam_network_tree.selection()
                if selected_items:
                    network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
                    self.refresh_ipam_ips(network)
            except Exception:
                pass

        # 按钮框架（放在dialog底部，固定位置）
        button_frame = ttk.Frame(dialog.dialog)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=(0, 15))
        
        button_container = ttk.Frame(button_frame)
        button_container.pack(side=tk.RIGHT)

        # 创建按钮
        confirm_btn = ttk.Button(button_container, text=_('confirm'), command=execute_action)
        confirm_btn.pack(side=tk.LEFT, padx=10)

        cancel_btn = ttk.Button(button_container, text=_('cancel'), command=dialog.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=0)
        # 显示对话框
        dialog.show()
    
    def keep_selected_record_and_delete_others(self, keep_id, ip_address):
        """保留选中的记录，删除该IP的其他冲突记录
        
        Args:
            keep_id: 要保留的记录ID
            ip_address: IP地址
        """
        try:
            # 获取该IP的所有记录
            conn = sqlite3.connect(self.ipam.db_file)
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM ip_addresses WHERE ip_address = ? AND id != ?', 
                          (ip_address, keep_id))
            other_records = cursor.fetchall()
            conn.close()
            
            # 删除其他记录
            deleted_count = 0
            for record in other_records:
                success, _message = self.ipam.delete_ip_by_id(record[0])
                if success:
                    deleted_count += 1
            
            self.show_info(_('success'), f"{_('conflict_resolved')}: {_('deleted_records').format(count=deleted_count)}")
        except Exception as e:
            self.show_error(_('error'), f"{_('resolve_conflict_failed')}: {str(e)}")
    
    def refresh_conflicts(self, tree, conflicts, selected_networks=None):
        """刷新冲突列表"""
        # 清空树
        for item in tree.get_children():
            tree.delete(item)
        
        # 重新检查冲突
        try:
            # 收集所有IP地址记录，避免重复检查
            all_ip_records = {}
            
            if not selected_networks:
                networks = self.ipam.get_all_networks()
                for network_info in networks:
                    network = network_info['network']
                    ips = self.ipam.get_network_ips(network)
                    
                    for ip in ips:
                        if ip['id'] not in all_ip_records:
                            all_ip_records[ip['id']] = {
                                'id': ip['id'],
                                'network': network,
                                'ip_address': ip['ip_address'],
                                'status': ip['status'],
                                'hostname': ip.get('hostname', ''),
                                'mac_address': ip.get('mac_address', ''),
                                'description': ip.get('description', ''),
                                'allocated_at': ip.get('allocated_at', ''),
                                'expiry_date': ip.get('expiry_date', '')
                            }
            else:
                for network in selected_networks:
                    ips = self.ipam.get_network_ips(network)
                    
                    for ip in ips:
                        if ip['id'] not in all_ip_records:
                            all_ip_records[ip['id']] = {
                                'id': ip['id'],
                                'network': network,
                                'ip_address': ip['ip_address'],
                                'status': ip['status'],
                                'hostname': ip.get('hostname', ''),
                                'mac_address': ip.get('mac_address', ''),
                                'description': ip.get('description', ''),
                                'allocated_at': ip.get('allocated_at', ''),
                                'expiry_date': ip.get('expiry_date', '')
                            }
            
            # 统计每个IP地址的分配情况
            ip_counts = {}
            for ip_record in all_ip_records.values():
                ip_addr = ip_record['ip_address']
                if ip_addr in ip_counts:
                    ip_counts[ip_addr].append(ip_record)
                else:
                    ip_counts[ip_addr] = [ip_record]
            
            # 找出冲突的IP地址（被分配多次的）
            updated_conflicts = []
            for ip_addr, ip_list in ip_counts.items():
                if len(ip_list) > 1:
                    # 为每个冲突的IP地址记录添加到冲突列表
                    for ip in ip_list:
                        updated_conflicts.append(ip)
            
            # 按IP地址排序，与初始打开时保持一致
            updated_conflicts = self._sort_ip_list(updated_conflicts)
            
            # 显示更新后的冲突
            for conflict in updated_conflicts:
                # 翻译状态值
                status = conflict['status']
                if status == 'released':
                    translated_status = _('released')
                elif status == 'allocated':
                    translated_status = _('allocated')
                elif status == 'reserved':
                    translated_status = _('reserved')
                else:
                    translated_status = status
                
                # 处理分配时间，只显示日期部分
                allocated_at = conflict.get('allocated_at', '')
                if allocated_at:
                    if ' ' in allocated_at:
                        allocated_at = allocated_at.split(' ')[0]
                    elif 'T' in allocated_at:
                        allocated_at = allocated_at.split('T')[0]
                
                # 处理过期日期，只显示日期部分
                expiry_date = conflict.get('expiry_date', '')
                if expiry_date:
                    if ' ' in expiry_date:
                        expiry_date = expiry_date.split(' ')[0]
                    elif 'T' in expiry_date:
                        expiry_date = expiry_date.split('T')[0]
                
                tree.insert('', tk.END, iid=str(conflict['id']), values=(
                    conflict['ip_address'],
                    translated_status,
                    conflict['hostname'],
                    conflict.get('mac_address', ''),
                    conflict['description'],
                    allocated_at,
                    expiry_date
                ))
            
            # 更新斑马纹样式
            self.update_table_zebra_stripes(tree)
            
            # 如果没有冲突了，显示提示
            if not updated_conflicts:
                self.show_info(_('hint'), _('no_ip_conflicts_found'))
            
            # 更新统计信息
            for widget in tree.master.winfo_children():
                if isinstance(widget, ttk.Frame) and widget.winfo_children():
                    for label in widget.winfo_children():
                        if isinstance(label, ttk.Label):
                            label.config(text=_('found_ip_conflicts_count', count=len(updated_conflicts)))
                            break
                    break
        except Exception as e:
            self.show_error(_('error'), f"{_('check_conflict_failed')}: {str(e)}")
    
    def import_export_network_data(self):
        """导入/导出网段数据对话框"""
        try:
            dialog = ComplexDialog(self.root, _('network_import_export'), 500, 370, resizable=False, modal=True)

            main_frame = ttk.Frame(dialog.content_frame, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(main_frame, text=_('choose_import_export_method'), font=('', 11, 'bold')).pack(pady=(0, 10))

            export_frame = ttk.LabelFrame(main_frame, text=_('export_options'))
            export_frame.pack(fill=tk.X, pady=(0, 5), padx=15)

            export_btn_frame = ttk.Frame(export_frame)
            export_btn_frame.pack(pady=5, fill=tk.X, padx=10)

            btn_row1 = ttk.Frame(export_btn_frame)
            btn_row1.pack(fill=tk.X, pady=2)

            sync_ip_var = tk.BooleanVar(value=False)

            ttk.Button(btn_row1, text=_('export_selected_networks'),
                      command=lambda: self._do_export_network(dialog, selected=True, sync_ip_var=sync_ip_var)).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
            ttk.Button(btn_row1, text=_('export_all_networks'),
                      command=lambda: self._do_export_network(dialog, selected=False, sync_ip_var=sync_ip_var)).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

            ttk.Checkbutton(export_btn_frame, text=_('sync_export_ips'), variable=sync_ip_var).pack(anchor=tk.W, padx=5, pady=(5, 2))

            import_frame = ttk.LabelFrame(main_frame, text=_('import_options'))
            import_frame.pack(fill=tk.X, pady=(0, 5), padx=15)

            import_btn_frame = ttk.Frame(import_frame)
            import_btn_frame.pack(pady=5, fill=tk.X, padx=10)

            import_btn = ttk.Button(import_btn_frame, text=_('import_from_file'),
                                   command=lambda: self._do_import_from_file(dialog))
            import_btn.pack(fill=tk.X, padx=5, pady=(0, 5))

            template_frame = ttk.LabelFrame(main_frame, text=_('download_template'))
            template_frame.pack(fill=tk.X, pady=(0, 5), padx=15)

            template_inner = ttk.Frame(template_frame)
            template_inner.pack(pady=5, fill=tk.X, padx=10)

            tmpl_net_var = tk.BooleanVar(value=True)
            tmpl_ip_var = tk.BooleanVar(value=True)

            ttk.Checkbutton(template_inner, text=_('include_network_template'), variable=tmpl_net_var,
                           command=lambda: self._validate_template_checks(True, tmpl_net_var, tmpl_ip_var)).pack(side=tk.LEFT, padx=5)
            ttk.Checkbutton(template_inner, text=_('include_ip_template'), variable=tmpl_ip_var,
                           command=lambda: self._validate_template_checks(False, tmpl_net_var, tmpl_ip_var)).pack(side=tk.LEFT, padx=5)
            ttk.Button(template_inner, text=_('download'),
                      command=lambda: self._do_download_template(dialog, tmpl_net_var, tmpl_ip_var)).pack(side=tk.RIGHT, padx=5, pady=(0, 5))

            # 添加关闭按钮（使用对话框标准按钮布局）
            dialog.add_button(_('close'), dialog.destroy)

            # 显示对话框并自动调整高度
            dialog.show()

        except Exception as e:
            self.show_error(_('error'), f"{_('operation_failed')}: {str(e)}")

    def _validate_template_checks(self, is_network, net_var, ip_var):
        """确保模板复选框至少选中一个，回勾另一个未被点击的复选框"""
        if not net_var.get() and not ip_var.get():
            if is_network:
                ip_var.set(True)
            else:
                net_var.set(True)

    def _do_export_network(self, parent_dialog, selected: bool, sync_ip_var=None):
        """执行网段导出操作"""
        try:
            import tkinter.filedialog as filedialog
            from datetime import datetime

            parent_dialog.destroy()

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"network_export_{timestamp}.xlsx"

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            if not file_path:
                return

            include_ips = sync_ip_var.get() if sync_ip_var else False

            if selected:
                selected_items = self.ipam_network_tree.selection()
                if not selected_items:
                    self.show_error(_('error'), _('please_select_network'))
                    return
                selected_networks = [self.ipam_network_tree.item(item, 'values')[0] for item in selected_items]
                ip_networks = selected_networks if include_ips else None
                if self.ipam.export_data_to_xlsx(
                        file_path, networks=selected_networks,
                        include_ips=include_ips, ip_networks=ip_networks):
                    self.show_info(_('success'), _('export_success'))
                else:
                    self.show_error(_('error'), _('export_failed'))
            else:
                if self.ipam.export_data_to_xlsx(file_path, include_ips=include_ips):
                    self.show_info(_('success'), _('export_success'))
                else:
                    self.show_error(_('error'), _('export_failed'))
        except Exception as e:
            self.show_error(_('error'), f"{_('export_failed')}: {str(e)}")

    def _do_import_from_file(self, parent_dialog):
        """从文件导入数据，弹出比对对话框"""
        try:
            import tkinter.filedialog as filedialog

            parent_dialog.destroy()

            file_path = filedialog.askopenfilename(
                title=_('select_file_to_import'),
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not file_path:
                return

            data = self.ipam.parse_xlsx_for_import(file_path)
            if not data['networks'] and not data['ips']:
                self.show_info(_('hint'), _('no_valid_data_found'))
                return

            validated = self.ipam.validate_import_data(data)
            self._show_import_compare_dialog(validated)
        except Exception as e:
            self.show_error(_('error'), f"{_('import_failed')}: {str(e)}")

    def _show_import_compare_dialog(self, validated_data: dict):
        """显示导入数据对话框"""
        try:
            font_family, font_size = get_current_font_settings()

            dialog = ComplexDialog(self.root, _('import_data'), 800, 600, resizable=True, modal=True)
            dialog.dialog.focus_force()

            # 清除content_frame的默认网格配置（取消垂直居中）
            for i in range(101):  # 清除DialogBase中设置的101行
                try:
                    dialog.content_frame.grid_rowconfigure(i, weight=0)
                except Exception:
                    pass

            main_frame = ttk.Frame(dialog.content_frame)
            main_frame.pack(fill=tk.BOTH, expand=True)

            # 构建标签页内容（占据主要空间）
            notebook = ColoredNotebook(main_frame, style=self.style)
            notebook.pack(fill=tk.BOTH, expand=True)

            # 计算数据统计
            net_data = validated_data['networks']
            ip_data = validated_data['ips']
            net_new = len(net_data['new'])
            net_existing = len(net_data['existing'])
            net_invalid = len(net_data['invalid'])
            ip_new = len(ip_data['new'])
            ip_existing = len(ip_data['existing'])
            ip_invalid = len(ip_data['invalid'])

            net_summary = f"{_('network_data_summary')}: {_('new_count')}={net_new}, {_('existing_count')}={net_existing}, {_('invalid_count')}={net_invalid}"
            ip_summary = f"{_('ip_data_summary')}: {_('new_count')}={ip_new}, {_('existing_count')}={ip_existing}, {_('invalid_count')}={ip_invalid}"

            net_tab = ttk.Frame(notebook.content_area, padding="2", style=notebook.light_blue_style)
            net_tab.pack(fill=tk.BOTH, expand=True)
            self._build_compare_table(net_tab, net_data, 'network')
            notebook.add_tab(f"{_('network_sheet_name')} ({net_new + net_existing + net_invalid})", net_tab, "#e3f2fd")

            ip_tab = ttk.Frame(notebook.content_area, padding="2", style=notebook.light_green_style)
            ip_tab.pack(fill=tk.BOTH, expand=True)
            self._build_compare_table(ip_tab, ip_data, 'ip')
            notebook.add_tab(f"{_('ip_sheet_name')} ({ip_new + ip_existing + ip_invalid})", ip_tab, "#e8f5e9")

            for tab in notebook.tabs:
                current_width = tab["button"].cget("width")
                tab["button"].config(width=int(current_width * 1.5))

            import_net_var = tk.BooleanVar(value=net_new > 0 or net_existing > 0)
            import_ip_var = tk.BooleanVar(value=ip_new > 0 or ip_existing > 0)
            overwrite_var = tk.BooleanVar(value=False)

            # 底部控件放入dialog.button_frame，天然在对话框底部
            # 左侧放置复选框和统计信息
            left_frame = ttk.Frame(dialog.button_frame)
            left_frame.pack(side=tk.LEFT, fill=tk.Y)

            # 复选框
            check_frame = ttk.Frame(left_frame)
            check_frame.pack(fill=tk.X, pady=(0, 5))

            ttk.Checkbutton(check_frame, text=_('import_network_data'), variable=import_net_var).pack(side=tk.LEFT, padx=10)
            ttk.Checkbutton(check_frame, text=_('import_ip_data'), variable=import_ip_var).pack(side=tk.LEFT, padx=10)
            ttk.Checkbutton(check_frame, text=_('overwrite_existing'), variable=overwrite_var).pack(side=tk.LEFT, padx=10)

            # 统计信息
            summary_label_frame = ttk.Frame(left_frame)
            summary_label_frame.pack(fill=tk.X)

            ttk.Label(summary_label_frame, text=net_summary, font=(font_family, font_size - 2)).pack(anchor=tk.W)
            ttk.Label(summary_label_frame, text=ip_summary, font=(font_family, font_size - 2)).pack(anchor=tk.W)

            # 右侧放置导入按钮（靠右下）
            btn_container = ttk.Frame(dialog.button_frame)
            btn_container.pack(side=tk.RIGHT, fill=tk.Y)
            ttk.Button(btn_container, text=_('import'),
                      command=lambda: self._execute_import(dialog, validated_data,
                                                           import_net_var, import_ip_var, overwrite_var)).pack(side=tk.BOTTOM)

        except Exception as e:
            self.show_error(_('error'), f"{_('show_compare_dialog_failed')}: {str(e)}")

    def _build_compare_table(self, parent, data: dict, data_type: str):
        """构建比对表格"""
        try:
            tree_frame = ttk.Frame(parent)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
            tree_frame.grid_rowconfigure(0, weight=1)
            tree_frame.grid_columnconfigure(0, weight=1)

            if data_type == 'network':
                columns = ("network", "description", "vlan", "status")
                headers = (_('col_network'), _('col_description'), 'VLAN', _('col_status'))
            else:
                columns = ("ip_address", "status", "hostname", "mac_address", "description", "compare_status")
                headers = (_('col_ip_address'), _('col_status'), _('col_hostname'),
                          _('col_mac_address'), _('col_description'), _('col_compare_status'))

            # 不限制高度，让表格自适应窗口大小
            tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
            for col, header in zip(columns, headers):
                tree.heading(col, text=header)
                tree.column(col, width=100, minwidth=60)

            tree.column(columns[-1], width=80, minwidth=60)

            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)

            def scrollbar_callback(*args):
                scrollbar.set(*args)
                if float(args[0]) <= 0.0 and float(args[1]) >= 1.0:
                    scrollbar.grid_remove()
                else:
                    scrollbar.grid(row=0, column=1, sticky=tk.NS)

            tree.configure(yscrollcommand=scrollbar_callback)
            tree.grid(row=0, column=0, sticky=tk.NSEW)
            scrollbar_callback(0.0, 1.0)

            self.configure_treeview_styles(tree)
            tree.tag_configure('new', foreground='green')
            tree.tag_configure('existing', foreground='#CC7700')
            tree.tag_configure('invalid', foreground='red')

            status_mapping = {
                'allocated': _('status_allocated'),
                'reserved': _('status_reserved'),
                'released': _('status_released'),
                'conflict': _('status_conflict')
            }

            if data_type == 'network':
                for item in data['new']:
                    tree.insert("", tk.END, values=(item['network'], item.get('description', ''), item.get('vlan', ''), _('status_new')), tags=('new',))
                for item in data['existing']:
                    tree.insert("", tk.END, values=(item['network'], item.get('description', ''), item.get('vlan', ''), _('status_existing')), tags=('existing',))
                for item in data['invalid']:
                    tree.insert("", tk.END, values=(item['network'], item.get('description', ''), item.get('vlan', ''), _('status_invalid')), tags=('invalid',))
            else:
                for item in data['new']:
                    status = status_mapping.get(item.get('status', ''), item.get('status', ''))
                    tree.insert("", tk.END, values=(item['ip_address'], status, item.get('hostname', ''), item.get('mac_address', ''), item.get('description', ''), _('status_new')), tags=('new',))
                for item in data['existing']:
                    status = status_mapping.get(item.get('status', ''), item.get('status', ''))
                    tree.insert("", tk.END, values=(item['ip_address'], status, item.get('hostname', ''), item.get('mac_address', ''), item.get('description', ''), _('status_existing')), tags=('existing',))
                for item in data['invalid']:
                    status = status_mapping.get(item.get('status', ''), item.get('status', ''))
                    error_msg = item.get('error', _('status_invalid'))
                    tree.insert("", tk.END, values=(item['ip_address'], status, item.get('hostname', ''), item.get('mac_address', ''), item.get('description', ''), error_msg), tags=('invalid',))

            self.update_table_zebra_stripes(tree)
        except Exception as e:
            print(f"{_('build_compare_table_failed')}: {str(e)}")

    def _execute_import(self, dialog, validated_data, import_net_var, import_ip_var, overwrite_var):
        """执行导入操作"""
        try:
            import_networks = import_net_var.get()
            import_ips = import_ip_var.get()

            if not import_networks and not import_ips:
                self.show_info(_('hint'), _('please_select_import_data'))
                return

            stats = self.ipam.import_validated_data(
                validated_data,
                import_networks=import_networks,
                import_ips=import_ips,
                overwrite_existing=overwrite_var.get()
            )

            dialog.destroy()

            msg_parts = []
            if import_networks:
                if stats['networks_added'] > 0:
                    msg_parts.append(f"{_('networks_added')}: {stats['networks_added']}")
                if stats['networks_updated'] > 0:
                    msg_parts.append(f"{_('networks_updated')}: {stats['networks_updated']}")
            if import_ips:
                if stats['ips_added'] > 0:
                    msg_parts.append(f"{_('ips_added')}: {stats['ips_added']}")
                if stats['ips_updated'] > 0:
                    msg_parts.append(f"{_('ips_updated')}: {stats['ips_updated']}")

            if msg_parts:
                self.show_info(_('success'), f"{_('import_success')}: {', '.join(msg_parts)}")
            else:
                self.show_info(_('hint'), _('no_data_imported'))

            self.refresh_ipam_networks()
            self.refresh_ipam_stats()
        except Exception as e:
            self.show_error(_('error'), f"{_('import_failed')}: {str(e)}")

    def _do_download_template(self, parent_dialog, tmpl_net_var, tmpl_ip_var):
        """下载导入模板"""
        try:
            import tkinter.filedialog as filedialog
            from datetime import datetime

            include_networks = tmpl_net_var.get()
            include_ips = tmpl_ip_var.get()

            if not include_networks and not include_ips:
                self.show_info(_('hint'), _('please_select_template'))
                return

            parent_dialog.destroy()

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"import_template_{timestamp}.xlsx"

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            if not file_path:
                return

            if self.ipam.generate_template_xlsx(file_path, include_networks=include_networks, include_ips=include_ips):
                self.show_info(_('success'), f"{_('download_template_success')}: {file_path}")
            else:
                self.show_error(_('error'), _('download_template_failed'))
        except Exception as e:
            self.show_error(_('error'), f"{_('download_template_failed')}: {str(e)}")
    

    

    def backup_restore_data(self):
        """备份/恢复数据对话框"""
        try:
            # 创建对话框
            dialog = ComplexDialog(self.root, _('backup_restore'), 700, 500, resizable=False, modal=True)
            
            # 配置对话框的行和列，移除默认的垂直居中设置
            # 清除所有现有的行配置
            # 注意：grid_size()[1]可能返回0（空框架），需要确保至少执行一次循环
            rows = dialog.content_frame.grid_size()[1] if dialog.content_frame.grid_size()[1] > 0 else 1
            for i in range(rows):
                try:
                    dialog.content_frame.grid_rowconfigure(i, weight=0)
                except Exception:
                    pass
            # 重新配置行
            dialog.content_frame.grid_rowconfigure(0, weight=0)
            dialog.content_frame.grid_rowconfigure(1, weight=1)
            dialog.content_frame.grid_columnconfigure(0, weight=1)
            
            def on_backup():
                backup_path = self.ipam.backup_data(backup_type='manual', frequency='manual')
                if backup_path:
                    self.show_info(_('success'), f"{_('backup_success')}: {backup_path}")
                    # 刷新备份列表
                    refresh_backup_list()
                else:
                    self.show_error(_('error'), _('backup_failed'))
            
            # 添加统计信息
            stats_frame = ttk.Frame(dialog.content_frame)
            stats_frame.grid(row=0, column=0, sticky="ew", padx=10)
            
            # 获取当前系统中的网络和IP数量
            stats = self.ipam.get_overall_stats()
            total_networks = stats.get('total_networks', 0)
            total_ips = stats.get('total_ips', 0)
            
            # 显示统计信息
            stats_label = ttk.Label(stats_frame, text=_('backup_stats_info', network_count=total_networks, ip_count=total_ips, backup_count=len(self.ipam.list_backups())))
            stats_label.pack(anchor=tk.W)
            
            # 创建一个框架来包含表格和滚动条
            tree_frame = ttk.Frame(dialog.content_frame, borderwidth=0, relief="flat")
            tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(10, 10))
            
            # 创建树状视图显示备份信息
            backup_tree = ttk.Treeview(tree_frame, columns=('filename', 'backup_time', 'network_count', 'ip_count'), show='headings')
            backup_tree.heading('filename', text=_('file_name'))
            backup_tree.heading('backup_time', text=_('backup_time'))
            backup_tree.heading('network_count', text=_('network_count'))
            backup_tree.heading('ip_count', text=_('ip_count'))
            
            backup_tree.column('filename', width=200)
            backup_tree.column('backup_time', width=150)
            backup_tree.column('network_count', width=80, anchor=tk.CENTER)
            backup_tree.column('ip_count', width=80, anchor=tk.CENTER)
            
            # 添加滚动条，放在表格右侧
            scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
            # 使用通用方法设置滚动条，支持自动隐藏
            self._setup_scrollbar(scrollbar, backup_tree, initial_hidden=True)
            # 使用grid布局
            tree_frame.grid_rowconfigure(0, weight=1)
            tree_frame.grid_columnconfigure(0, weight=1)
            backup_tree.grid(row=0, column=0, sticky="nsew")
            scrollbar.grid(row=0, column=1, sticky="ns")
            
            # 配置斑马纹样式
            self.configure_treeview_styles(backup_tree)
            
            # 加载备份列表的函数
            def refresh_backup_list():
                # 清空现有列表
                for item in backup_tree.get_children():
                    backup_tree.delete(item)
                
                # 加载备份列表
                backups = self.ipam.list_backups()
                for backup in backups:
                    filename = backup['filename']
                    info = backup['info']
                    backup_tree.insert('', tk.END, values=(
                        filename,
                        info.get('backup_time', ''),
                        info.get('network_count', 0),
                        info.get('ip_count', 0)
                    ), tags=(backup['file_path'],))
                
                # 刷新滚动条状态
                if backup_tree.yview():
                    # 触发滚动条回调，更新滚动条显示状态
                    backup_tree.yview_moveto(0)
                    backup_tree.yview_moveto(0)
                
                # 更新斑马纹样式
                self.update_table_zebra_stripes(backup_tree)
            
            # 初始加载备份列表
            refresh_backup_list()
            
            # 使用DialogBase提供的button_frame
            button_frame = dialog.button_frame
            
            # 左侧按钮区域
            left_buttons = ttk.Frame(button_frame)
            left_buttons.pack(side=tk.LEFT, padx=0, pady=(0, 10))
            
            # 添加一个占位框架，增加左右区域之间的间距
            spacer = ttk.Frame(button_frame)
            spacer.pack(side=tk.LEFT, expand=True, fill=tk.X)
            
            # 右侧自动备份设置区域
            right_settings = ttk.Frame(button_frame)
            right_settings.pack(side=tk.RIGHT, padx=10, pady=(0, 10))
            
            def on_restore():
                selected_items = backup_tree.selection()
                if not selected_items:
                    self.show_info(_('hint'), _('please_select_backup_to_restore'))
                    return
                
                item = selected_items[0]
                file_path = backup_tree.item(item, 'tags')[0]
                
                # 确认恢复
                confirm = messagebox.askyesno(_('confirm'), f"{_('confirm_restore_backup')}: {backup_tree.item(item, 'values')[0]} {_('question_mark')}")
                if confirm:
                    if self.ipam.restore_data(file_path):
                        self.show_info(_('success'), f"{_('data_restore_success')}: {file_path}")
                        self.refresh_ipam_networks()
                        self.refresh_ipam_stats()
                    else:
                        self.show_error(_('error'), _('data_restore_failed'))
            
            def on_delete():
                selected_items = backup_tree.selection()
                if not selected_items:
                    self.show_info(_('hint'), _('please_select_backup_to_delete'))
                    return
                
                # 获取选中的备份文件信息
                selected_backups = []
                for item in selected_items:
                    file_path = backup_tree.item(item, 'tags')[0]
                    filename = backup_tree.item(item, 'values')[0]
                    selected_backups.append((item, file_path, filename))
                
                # 确认删除
                if len(selected_backups) == 1:
                    confirm = messagebox.askyesno(_('confirm'), f"{_('confirm_delete_backup')}: {selected_backups[0][2]} {_('question_mark')}")
                else:
                    confirm = messagebox.askyesno(_('confirm'), f"{_('confirm_delete_multiple_backups', count=len(selected_backups))}")
                
                if confirm:
                    try:
                        import os
                        deleted_count = 0
                        for item, file_path, filename in selected_backups:
                            try:
                                # 尝试删除物理文件
                                try:
                                    os.remove(file_path)
                                except (FileNotFoundError, PermissionError):
                                    # 文件不存在或权限不足时忽略错误
                                    pass
                                # 从数据库中删除备份记录
                                self.ipam.delete_backup(file_path)
                                backup_tree.delete(item)
                                deleted_count += 1
                            except Exception as e:
                                self.show_error(_('error'), f"{_('delete_failed', name=filename)}: {str(e)}")
                        
                        if deleted_count > 0:
                            self.show_info(_('success'), f"{_('successfully_deleted_backups', count=deleted_count)}")
                    except Exception as e:
                        self.show_error(_('error'), f"{_('delete_failed')}: {str(e)}")
            # 频率映射（英文到本地化）

            def get_frequency_map():
                return {
                    'disabled': _('disabled'),
                    'hourly': _('hourly'),
                    'daily': _('daily'),
                    'weekly': _('weekly'),
                    'monthly': _('monthly')
                }
            
            # 反向频率映射（本地化到英文）
            def get_reverse_frequency_map():
                frequency_map = get_frequency_map()
                return {v: k for k, v in frequency_map.items()}
            
            # 加载配置
            def load_config():
                try:
                    # 使用配置管理器获取自动备份频率
                    config = get_config()
                    english_frequency = config.get_auto_backup_frequency()
                    # 转换为本地化值
                    frequency_map = get_frequency_map()
                    return frequency_map.get(english_frequency, _('daily'))
                except Exception:
                    pass
                return _('daily')
            
            # 保存配置
            def save_config(frequency):
                try:
                    # 将本地化频率值转换为英文
                    reverse_map = get_reverse_frequency_map()
                    english_frequency = reverse_map.get(frequency, 'daily')
                    
                    # 使用配置管理器保存自动备份频率
                    config = get_config()
                    return config.set_auto_backup_frequency(english_frequency)
                except Exception:
                    return False
            
            # 自动备份周期设置
            def on_auto_backup_change(event=None):
                frequency = auto_backup_var.get()
                if save_config(frequency):
                    self.show_info(_('success'), f"{_('auto_backup_frequency_set_to')}: {frequency}")
                else:
                    self.show_error(_('error'), _('save_config_failed'))
            
            # 创建自动备份周期选择控件
            ttk.Label(right_settings, text=_('auto_backup_frequency') + ':').pack(side=tk.LEFT, padx=5)
            auto_backup_var = tk.StringVar(value=load_config())
            auto_backup_options = [_('disabled'), _('hourly'), _('daily'), _('weekly'), _('monthly')]
            auto_backup_menu = ttk.Combobox(right_settings, textvariable=auto_backup_var, values=auto_backup_options, width=10)
            auto_backup_menu.pack(side=tk.LEFT, padx=5)
            # 添加事件处理程序，当选择变化时自动保存
            auto_backup_menu.bind("<<ComboboxSelected>>", on_auto_backup_change)
            
            # 左侧按钮
            ttk.Button(left_buttons, text=_('backup'), command=on_backup).pack(side=tk.LEFT, padx=5)
            ttk.Button(left_buttons, text=_('restore'), command=on_restore).pack(side=tk.LEFT, padx=5)
            ttk.Button(left_buttons, text=_('delete'), command=on_delete).pack(side=tk.LEFT, padx=5)
            
            # 绑定对话框关闭事件
            dialog.protocol("WM_DELETE_WINDOW", dialog.destroy)
            
            # 显示对话框
            dialog.show()
            
        except Exception as e:
            self.show_error(_('error'), f"{_('operation_failed')}: {str(e)}")
    

    def auto_scan_network(self):
        """打开网络扫描配置对话框"""
        port_presets = {
            'fast': [80, 443, 445, 135],
            'standard': [80, 443, 22, 21, 25, 53, 110, 135, 139, 445, 3389, 8080],
            'complete': [80, 443, 22, 21, 23, 25, 53, 110, 135, 139, 445, 465, 587, 993, 995, 3389, 8080, 8443, 3306, 5432]
        }
        
        dialog = self.create_dialog(_('auto_scan_network'), 440, 200, resizable=False, modal=True)
        
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(20, 8))
        
        # 内容区域
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15)
        content_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(content_frame, text=_('network_address_cidr_format') + ':').grid(
            row=0, column=0, padx=5, pady=8, sticky=tk.W)
        network_border, network_entry = create_bordered_entry(content_frame)
        network_border.grid(row=0, column=1, padx=5, pady=8, sticky=tk.EW)
        
        def validate_network_input(event=None):
            self.validate_cidr(network_entry.get(), network_entry, require_prefix=True)
        
        network_entry.bind('<FocusOut>', validate_network_input)
        network_entry.bind('<KeyRelease>', validate_network_input)

        options_frame = ttk.Frame(content_frame)
        options_frame.grid(row=1, column=0, columnspan=2, pady=8, sticky=tk.EW)
        options_frame.grid_columnconfigure(1, weight=1)
        options_frame.grid_columnconfigure(3, weight=1)

        ttk.Label(options_frame, text=_('thread_count') + ':').grid(row=0, column=0, padx=5, sticky=tk.W)
        thread_var = tk.StringVar(value="20")
        thread_border, thread_entry = create_bordered_entry(options_frame, textvariable=thread_var)
        thread_border.grid(row=0, column=1, padx=5, sticky=tk.EW)

        ttk.Label(options_frame, text=_('timeout_ms') + ':').grid(row=0, column=2, padx=10, sticky=tk.W)
        timeout_var = tk.StringVar(value="500")
        timeout_border, timeout_entry = create_bordered_entry(options_frame, textvariable=timeout_var)
        timeout_border.grid(row=0, column=3, padx=5, sticky=tk.EW)

        method_frame = ttk.Frame(content_frame)
        method_frame.grid(row=2, column=0, columnspan=2, pady=4, sticky=tk.W)
        
        ttk.Label(method_frame, text=_('scan_method') + ':').pack(side=tk.LEFT, padx=5)
        method_var = tk.StringVar(value="ping")
        
        ping_radio = ttk.Radiobutton(method_frame, text="ICMP Ping", variable=method_var, 
                                     value="ping", command=lambda: on_method_change(True))
        ping_radio.pack(side=tk.LEFT, padx=2)
        tcp_radio = ttk.Radiobutton(method_frame, text="TCP", variable=method_var, 
                                    value="tcp", command=lambda: on_method_change(True))
        tcp_radio.pack(side=tk.LEFT, padx=2)

        tcp_frame = ttk.Frame(content_frame)
        tcp_frame.grid(row=3, column=0, columnspan=2, pady=5, sticky=tk.W)
        
        port_preset_frame = ttk.Frame(tcp_frame)
        port_preset_frame.pack(side=tk.TOP, fill=tk.X)
        
        ttk.Label(port_preset_frame, text=_('port_preset') + ':').pack(side=tk.LEFT, padx=5)
        preset_var = tk.StringVar(value="standard")
        
        ttk.Radiobutton(port_preset_frame, text=_('fast') + ' (4)', variable=preset_var, value="fast").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(port_preset_frame, text=_('standard') + ' (12)', variable=preset_var, value="standard").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(port_preset_frame, text=_('complete') + ' (20)', variable=preset_var, value="complete").pack(side=tk.LEFT, padx=2)
        
        custom_ports_frame = ttk.Frame(tcp_frame)
        custom_ports_frame.pack(side=tk.TOP, fill=tk.X, pady=(10, 0))
        
        ttk.Label(custom_ports_frame, text=_('custom_ports') + ':').pack(side=tk.LEFT, padx=5, anchor=tk.N)
        
        border_frame = tk.Frame(custom_ports_frame, highlightbackground="#a9a9a9", 
                                highlightcolor="#a9a9a9", highlightthickness=1, bd=0)
        border_frame.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True, anchor=tk.N)
        
        ports_text = tk.Text(border_frame, height=3, wrap=tk.WORD, bd=0, relief="flat", highlightthickness=0)
        ports_text.pack(fill="both", expand=True, padx=1, pady=1)
        
        ports_var = tk.StringVar(value="")
        
        def update_ports_text():
            preset = preset_var.get()
            ports = port_presets[preset]
            ports_text.delete(1.0, tk.END)
            ports_text.insert(tk.END, ', '.join(map(str, ports)))
        
        def on_preset_change():
            update_ports_text()
        
        update_ports_text()
        preset_var.trace('w', lambda *args: on_preset_change())
        
        def on_method_change(show):
            if method_var.get() == 'tcp':
                tcp_frame.grid(row=3, column=0, columnspan=2, pady=4, sticky=tk.W)
                dialog.geometry('440x280')
            else:
                tcp_frame.grid_remove()
                dialog.geometry('440x200')

        if method_var.get() != 'tcp':
            tcp_frame.grid_remove()

        # 按钮区域 - 固定在底部右侧
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=6, side=tk.BOTTOM)
        
        cancel_btn = ttk.Button(button_frame, text=_('cancel'), command=dialog.destroy)
        cancel_btn.pack(side=tk.RIGHT, padx=5)
        
        def on_scan():
            network = network_entry.get().strip()
            thread_count = thread_var.get()
            timeout_ms = timeout_var.get()
            scan_method = method_var.get()
            ports = ports_text.get(1.0, tk.END).strip()
            
            if not network:
                self.show_error(_('error'), _('please_enter_network_address'))
                return
            
            is_valid = self.validate_cidr(network)
            if is_valid:
                dialog.destroy()
                self._start_scan_from_dialog_simple(network, thread_count, timeout_ms, scan_method, ports)
        
        scan_btn = ttk.Button(button_frame, text=_('start_scan'), command=on_scan)
        scan_btn.pack(side=tk.RIGHT, padx=5)
        
        dialog.bind('<Return>', lambda e: on_scan())
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def _start_scan_from_dialog_simple(self, network, thread_count_str, timeout_ms_str, scan_method, ports_str=None):
        """简化的扫描启动方法"""
        try:
            thread_count = int(thread_count_str)
            if thread_count < 1 or thread_count > 200:
                self.show_error(_('error'), _('thread_count_range_error'))
                return
        except ValueError:
            self.show_error(_('error'), _('thread_count_invalid'))
            return

        try:
            timeout_ms = int(timeout_ms_str)
            if timeout_ms < 100 or timeout_ms > 10000:
                self.show_error(_('error'), _('timeout_range_error'))
                return
        except ValueError:
            self.show_error(_('error'), _('timeout_invalid'))
            return

        custom_ports = None
        if scan_method == 'tcp' and ports_str:
            ports_str = ports_str.strip()
            if ports_str:
                try:
                    custom_ports = [int(p.strip()) for p in ports_str.split(',') if p.strip()]
                    if not custom_ports:
                        self.show_error(_('error'), _('port_list_empty'))
                        return
                    for port in custom_ports:
                        if port < 1 or port > 65535:
                            self.show_error(_('error'), _('port_range_error'))
                            return
                except ValueError:
                    self.show_error(_('error'), _('port_invalid'))
                    return

        try:
            ip_network = ipaddress.ip_network(network, strict=False)
            total_hosts = len(list(ip_network.hosts()))
            if total_hosts > 1024:
                confirm = self.show_custom_confirm(
                    _('large_network_warning'),
                    _('large_network_confirm_message').format(total=total_hosts, network=network)
                )
                if not confirm:
                    return
        except ValueError as e:
            error_result = handle_ip_subnet_error(e)
            self.show_error(_('error'), error_result.get('error', str(e)))
            return

        self._show_scan_progress_dialog(network, thread_count, timeout_ms, scan_method, custom_ports)

    def _start_scan_from_dialog(self, config_dialog, network_entry, thread_var, timeout_var, method_var, ports_var=None):
        """从配置对话框启动扫描

        Args:
            config_dialog: 配置对话框
            network_entry: 网络地址输入框
            thread_var: 线程数变量
            timeout_var: 超时时间变量
            method_var: 扫描方式变量
            ports_var: 端口列表变量（可选）
        """
        network = network_entry.get().strip()
        if not network:
            self.show_error(_('error'), _('please_enter_network_address'))
            return

        try:
            ip_network = ipaddress.ip_network(network, strict=False)
        except ValueError as e:
            error_result = handle_ip_subnet_error(e)
            self.show_error(_('error'), error_result.get('error', str(e)))
            return

        try:
            thread_count = int(thread_var.get())
            if thread_count < 1 or thread_count > 200:
                self.show_error(_('error'), _('thread_count_range_error'))
                return
        except ValueError:
            self.show_error(_('error'), _('thread_count_invalid'))
            return

        try:
            timeout_ms = int(timeout_var.get())
            if timeout_ms < 100 or timeout_ms > 10000:
                self.show_error(_('error'), _('timeout_range_error'))
                return
        except ValueError:
            self.show_error(_('error'), _('timeout_invalid'))
            return

        scan_method = method_var.get()
        custom_ports = None
        
        if scan_method == 'tcp' and ports_var:
            ports_str = ports_var.get().strip()
            if ports_str:
                try:
                    custom_ports = [int(p.strip()) for p in ports_str.split(',') if p.strip()]
                    if not custom_ports:
                        self.show_error(_('error'), _('port_list_empty'))
                        return
                    for port in custom_ports:
                        if port < 1 or port > 65535:
                            self.show_error(_('error'), _('port_range_error'))
                            return
                except ValueError:
                    self.show_error(_('error'), _('port_invalid'))
                    return

        total_hosts = len(list(ip_network.hosts()))
        if total_hosts > 1024:
            confirm = self.show_custom_confirm(
                _('large_network_warning'),
                _('large_network_confirm_message').format(total=total_hosts, network=network)
            )
            if not confirm:
                return

        config_dialog.destroy()

        self._show_scan_progress_dialog(network, thread_count, timeout_ms, scan_method, custom_ports)

    def _show_scan_progress_dialog(self, network, thread_count, timeout_ms, scan_method, custom_ports=None):
        """显示扫描进度对话框

        Args:
            network: 网络地址
            thread_count: 线程数
            timeout_ms: 超时时间（毫秒）
            scan_method: 扫描方式
            custom_ports: 自定义端口列表（可选）
        """
        progress_dialog = self.create_dialog(_('network_scan'), 800, 600, resizable=True, modal=True)

        main_frame = ttk.Frame(progress_dialog, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 上半部分：信息+进度（固定高度）
        top_frame = ttk.Frame(main_frame, padding="10 2 10 2")  # 左右10，上下2
        top_frame.pack(fill=tk.X)

        info_frame = ttk.Frame(top_frame)
        info_frame.pack(fill=tk.X, pady=(0, 1))

        scan_info_label = ttk.Label(info_frame, text=f"{_('scanning_network')} {network}")
        scan_info_label.pack(anchor=tk.W)

        # 创建包含状态和百分比的子框架
        status_pct_frame = ttk.Frame(info_frame)
        status_pct_frame.pack(fill=tk.X, pady=(0, 0))

        status_label = ttk.Label(status_pct_frame, text=_('preparing_to_scan'))
        status_label.pack(side=tk.LEFT, anchor=tk.W)

        progress_pct_label = ttk.Label(status_pct_frame, text="0%")
        progress_pct_label.pack(side=tk.RIGHT, anchor=tk.E)

        progress_frame = ttk.Frame(top_frame)
        progress_frame.pack(fill=tk.X, pady=1)

        progress_var = tk.DoubleVar(value=0)
        progress_bar = ttk.Progressbar(progress_frame, variable=progress_var, maximum=100, mode='determinate')
        progress_bar.pack(fill=tk.X, pady=0)

        # 下半部分：表格（自动填充剩余空间）
        tree_frame = ttk.Frame(main_frame, padding="10 2 10 2")  # 左右10，上下2
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=1)

        columns = ('ip_address', 'hostname')
        result_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
        result_tree.heading('ip_address', text=_('ip_address'))
        result_tree.heading('hostname', text=_('hostname'))
        result_tree.column('ip_address', width=160, anchor=tk.W)
        result_tree.column('hostname', width=300, anchor=tk.W)

        # 使用主配置方法配置斑马纹样式
        self.configure_treeview_styles(result_tree)

        # 创建滚动条并应用自动隐藏功能
        tree_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        result_tree.grid(row=0, column=0, sticky=tk.NSEW)
        tree_scroll.grid(row=0, column=1, sticky=tk.NS)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 应用自动隐藏滚动条
        self._setup_scrollbar(tree_scroll, result_tree, initial_hidden=True, widget_command=result_tree.yview)

        # 按钮区域（固定在底部）
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(5, 0))

        # 使用应用程序级别的网络扫描器实例，避免每次创建新实例
        scanner = self.network_scanner
        scanner.reset()  # 重置扫描状态
        scan_state = {'completed': False, 'active_ips': []}  # 移除对对话框的引用

        def cancel_scan():
            """取消扫描或关闭对话框"""
            if not scan_state['completed']:
                scanner.cancel()
            # 清理事件绑定
            try:
                progress_dialog.unbind('<Escape>')
                progress_dialog.unbind('<Return>')
            except Exception:
                pass
            # 销毁对话框
            try:
                progress_dialog.destroy()
            except Exception:
                pass

        def on_closing():
            """处理对话框关闭事件"""
            cancel_scan()

        def import_results():
            """导入扫描结果到数据库"""
            active_ips = scan_state['active_ips'].copy()
            cancel_scan()
            self._process_scan_results(network, active_ips)

        progress_dialog.protocol("WM_DELETE_WINDOW", on_closing)

        # 统计结果标签
        result_label = ttk.Label(button_frame, text=f"{_('found_active_ips')}: 0")
        
        # 取消按钮（始终显示）
        cancel_button = ttk.Button(button_frame, text=_('cancel'), command=cancel_scan, width=10)
        
        # 导入按钮（扫描完成后显示）
        import_button = ttk.Button(button_frame, text=_('import'), command=import_results, width=10)
        
        # 使用grid布局实现垂直居中对齐
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, pad=5)
        button_frame.grid_columnconfigure(2, pad=5)
        
        result_label.grid(row=0, column=0, sticky=tk.NW, pady=0, padx=(10, 0))
        cancel_button.grid(row=0, column=2, sticky=tk.E, pady=(5, 0))
        
        # 绑定Escape键取消扫描
        progress_dialog.bind('<Escape>', lambda e: cancel_scan())
        
        # 绑定回车键导入结果
        def on_return_key(e):
            """处理回车键事件"""
            if scan_state['completed']:
                import_results()
        
        progress_dialog.bind('<Return>', on_return_key)

        def on_progress(scanned, active, total, current_ip):
            """扫描进度回调"""
            if not progress_dialog.winfo_exists():
                return
            try:
                pct = (scanned / total * 100) if total > 0 else 0
                progress_var.set(pct)
                progress_pct_label.config(text=f"{pct:.1f}%")
                status_label.config(text=f"{_('scanning')}: {current_ip} ({scanned}/{total})")
            except Exception:
                pass

        def ip_to_int(ip_str):
            """将IP地址转换为整数以便排序"""
            return int(''.join(f'{int(octet):03d}' for octet in ip_str.split('.')))

        def on_ip_found(ip_info):
            """发现活动IP回调，实时插入到正确位置保持排序"""
            if not progress_dialog.winfo_exists():
                return
            try:
                new_ip_int = ip_to_int(ip_info['ip_address'])
                insert_index = 0
                
                for child in result_tree.get_children():
                    values = result_tree.item(child)['values']
                    if values and ip_to_int(values[0]) > new_ip_int:
                        break
                    insert_index += 1
                
                result_tree.insert('', insert_index, values=(ip_info['ip_address'], ip_info['hostname']))
                scan_state['active_ips'].append(ip_info)
                
                self.update_table_zebra_stripes(result_tree)
                result_label.config(text=f"{_('found_active_ips')}: {len(scan_state['active_ips'])}")
            except Exception:
                pass

        def on_complete(active_ips):
            """扫描完成回调"""
            scan_state['completed'] = True
            
            def ip_to_int(ip_str):
                """将IP地址转换为整数以便排序"""
                return int(''.join(f'{int(octet):03d}' for octet in ip_str.split('.')))
            
            sorted_ips = sorted(active_ips, key=lambda x: ip_to_int(x['ip_address']))
            scan_state['active_ips'] = sorted_ips
            
            if not progress_dialog.winfo_exists():
                return
            try:
                progress_var.set(100)
                progress_pct_label.config(text="100%")
                status_label.config(text=_('scan_complete'))
                
                result_label.config(text=f"{_('found_active_ips')}: {len(sorted_ips)}")

                import_button.grid(row=0, column=1, sticky=tk.E, pady=(5, 0))
            except Exception:
                pass

        def on_error(error_msg):
            """扫描错误回调"""
            scan_state['completed'] = True
            if not progress_dialog.winfo_exists():
                return
            try:
                status_label.config(text=f"{_('scan_failed')}: {error_msg}")
                cancel_button.config(text=_('ok'))
                cancel_button.config(command=progress_dialog.destroy)
            except Exception:
                pass

        import threading
        
        def progress_callback(scanned, active, total, current_ip):
            if progress_dialog.winfo_exists():
                self.root.after(0, lambda: on_progress(scanned, active, total, current_ip))
        
        def ip_found_callback(ip_info):
            if progress_dialog.winfo_exists():
                self.root.after(0, lambda: on_ip_found(ip_info.copy()))
        
        def complete_callback(active_ips):
            if progress_dialog.winfo_exists():
                self.root.after(0, lambda: on_complete(active_ips))
        
        def error_callback(error_msg):
            if progress_dialog.winfo_exists():
                self.root.after(0, lambda: on_error(error_msg))
        
        scan_kwargs = {
            'network': network,
            'thread_count': thread_count,
            'timeout_ms': timeout_ms,
            'scan_method': scan_method,
            'on_progress': progress_callback,
            'on_ip_found': ip_found_callback,
            'on_complete': complete_callback,
            'on_error': error_callback,
        }
        
        if custom_ports:
            scan_kwargs['ports'] = custom_ports
        
        scan_thread = threading.Thread(
            target=scanner.scan_network,
            kwargs=scan_kwargs,
            daemon=True
        )
        scan_thread.start()

    def _on_scan_dialog_close(self, dialog, network, active_ips):
        """扫描结果对话框关闭处理

        Args:
            dialog: 对话框
            network: 网络地址
            active_ips: 活动IP列表
        """
        dialog.destroy()
        
        # 显示确认对话框，让用户确认是否导入扫描结果
        if active_ips:
            message = _('confirm_import_scan_results').format(count=len(active_ips), network=network)
            confirm = self.show_custom_confirm(_('scan_complete'), message)
            if confirm:
                self._process_scan_results(network, active_ips)
            else:
                self.show_info(_('import_canceled'), _('scan_results_not_imported'))
        else:
            self.show_info(_('scan_complete'), _('no_active_ips_found'))

    def _process_scan_results(self, network, active_ips):
        """处理扫描结果，将活动IP添加到IPAM

        Args:
            network: 网络地址
            active_ips: 活动IP信息列表
        """
        if not active_ips:
            self.show_info(_('scan_complete'), _('no_active_ips_found'))
            self.root.after(0, self.refresh_ipam_networks)
            return

        network_str = str(ipaddress.ip_network(network, strict=False))

        networks = self.ipam.get_all_networks()
        network_exists = any(net['network'] == network_str for net in networks)
        if not network_exists:
            self.ipam.add_network(network_str, f"{_('auto_scan_network')} - {network_str}")
            self.root.after(0, self.refresh_ipam_networks)

        for ip_info in active_ips:
            ip_info['description'] = _('auto_scan_discovered')

        success_count, failed_ips = self.ipam.batch_allocate_ips(network_str, active_ips)

        self.root.after(0, self.refresh_ipam_networks)
        self.root.after(100, self.refresh_ipam_stats)

        if failed_ips:
            self.show_info(_('scan_complete'),
                           f"{_('scan_result_with_failures').format(found=len(active_ips), added=success_count, failed=len(failed_ips))}")
        else:
            self.show_info(_('scan_complete'),
                           f"{_('scan_result_success').format(found=len(active_ips), added=success_count)}")
    

    

    

    

    
    def refresh_ipam_stats(self):
        """刷新IPAM统计数据"""
        try:
            # 使用新的get_overall_stats方法获取统计信息
            stats = self.ipam.get_overall_stats()
            
            total_networks = stats.get('total_networks', 0)
            total_ips = stats.get('total_ips', 0)
            ipv4_networks = stats.get('ipv4_networks', 0)
            ipv4_ips = stats.get('ipv4_ips', 0)
            ipv6_networks = stats.get('ipv6_networks', 0)
            ipv6_ips = stats.get('ipv6_ips', 0)
            allocated_ips = stats.get('allocated_ips', 0)
            reserved_ips = stats.get('reserved_ips', 0)
            available_ips = stats.get('available_ips', 0)
            expired_ips = stats.get('expired_ips', 0)
            expiring_ips = stats.get('expiring_ips', 0)
            named_ips = stats.get('named_ips', 0)
            vlan_count = stats.get('vlan_count', 0)
            
            utilization_rate = 0
            if total_ips > 0:
                utilization_rate = (allocated_ips + reserved_ips) / total_ips * 100
            
            if 'total_networks' in self.stats_labels:
                self.stats_labels['total_networks'].config(text=str(total_networks))
            if 'total_ips' in self.stats_labels:
                self.stats_labels['total_ips'].config(text=str(total_ips))
            if 'ipv4_networks' in self.stats_labels:
                self.stats_labels['ipv4_networks'].config(text=str(ipv4_networks))
            if 'ipv4_ips' in self.stats_labels:
                self.stats_labels['ipv4_ips'].config(text=str(ipv4_ips))
            if 'ipv6_networks' in self.stats_labels:
                self.stats_labels['ipv6_networks'].config(text=str(ipv6_networks))
            if 'ipv6_ips' in self.stats_labels:
                self.stats_labels['ipv6_ips'].config(text=str(ipv6_ips))
            if 'allocated_ips' in self.stats_labels:
                self.stats_labels['allocated_ips'].config(text=str(allocated_ips))
            if 'reserved_ips' in self.stats_labels:
                self.stats_labels['reserved_ips'].config(text=str(reserved_ips))
            if 'available_ips' in self.stats_labels:
                self.stats_labels['available_ips'].config(text=str(available_ips))
            if 'expired_ips' in self.stats_labels:
                self.stats_labels['expired_ips'].config(text=str(expired_ips))
            if 'expiring_ips' in self.stats_labels:
                self.stats_labels['expiring_ips'].config(text=str(expiring_ips))
            if 'named_ips' in self.stats_labels:
                self.stats_labels['named_ips'].config(text=str(named_ips))
            if 'vlan_count' in self.stats_labels:
                self.stats_labels['vlan_count'].config(text=str(vlan_count))
            if 'utilization_rate' in self.stats_labels:
                self.stats_labels['utilization_rate'].config(text=f"{utilization_rate:.2f}%")
            
            # 绘制饼图
            self.draw_ip_usage_pie_chart(allocated_ips, reserved_ips, available_ips, utilization_rate)
        except Exception as e:
            print(f"刷新统计数据失败: {e}")
    
    def draw_ip_usage_pie_chart(self, allocated, reserved, available, utilization_rate=0, avg_allocation_rate=0):
        """绘制IP地址使用情况饼图"""
        try:
            self.stats_canvas.delete("all")
            
            width = self.stats_canvas.winfo_width()
            height = self.stats_canvas.winfo_height()
            
            if width == 1 or height == 1:
                return
            
            center_x = width // 2
            center_y = height // 2
            radius = min(center_x, center_y) - 60
            
            total = allocated + reserved + available
            
            if total == 0:
                self._draw_text_with_stroke(center_x, center_y, _('no_data_available'), font=("微软雅黑", 12))
                return
            
            colors = ["#4CAF50", "#2196F3", "#E0E0E0"]
            labels = [_('stats_allocated'), _('stats_reserved'), _('stats_released')]
            values = [allocated, reserved, available]
            
            legend_x = 15
            legend_y = 15
            for i, (color, label, value) in enumerate(zip(colors, labels, values)):
                if value == 0:
                    continue
                self.stats_canvas.create_rectangle(legend_x, legend_y + i * 20, legend_x + 12, legend_y + 12 + i * 20, fill=color, outline="#ffffff")
                percentage = (value / total) * 100 if total > 0 else 0
                legend_text = f"{label}: {value} ({percentage:.1f}%)"
                self._draw_text_with_stroke(legend_x + 20, legend_y + 6 + i * 20, legend_text, font=("微软雅黑", 8), anchor=tk.W)
            
            start_angle = 0
            for i, value in enumerate(values):
                if value == 0:
                    continue
                
                angle = (value / total) * 360
                end_angle = start_angle + angle
                
                self.stats_canvas.create_arc(
                    center_x - radius, center_y - radius,
                    center_x + radius, center_y + radius,
                    start=start_angle, extent=angle,
                    fill=colors[i], outline="white", width=2
                )
                
                mid_angle_rad = (start_angle + end_angle) / 2 * 3.14159 / 180
                label_radius = radius + 20
                label_x = center_x + label_radius * math.cos(mid_angle_rad)
                label_y = center_y - label_radius * math.sin(mid_angle_rad)
                
                mid_angle_deg = (start_angle + end_angle) / 2
                if mid_angle_deg > 30 and mid_angle_deg < 150:
                    anchor = tk.W
                elif mid_angle_deg > 210 and mid_angle_deg < 330:
                    anchor = tk.E
                else:
                    anchor = tk.CENTER
                
                percentage = (value / total) * 100
                label_text = f"{labels[i]}: {value}"
                self._draw_text_with_stroke(label_x, label_y, label_text, font=("微软雅黑", 8), anchor=anchor)
                
                start_angle = end_angle
            
            self._draw_text_with_stroke(
                center_x, center_y, 
                f"{_('stats_total_ip')}: {total}",
                font=("微软雅黑", 12, "bold")
            )
            
            self._draw_text_with_stroke(
                center_x, height - 25, 
                f"{_('stats_utilization_rate')} {utilization_rate:.1f}%",
                font=("微软雅黑", 10, "bold")
            )
        except Exception as e:
            print(f"绘制饼图失败: {e}")
    
    def _draw_text_with_stroke(self, x, y, text, font=("微软雅黑", 10), anchor=tk.CENTER, fill="#ffffff"):
        """绘制带描边效果的文本，提高文字在各种背景色上的可读性"""
        # 先绘制4个黑色偏移文字作为描边
        for dx in (-1, 1):
            for dy in (-1, 1):
                self.stats_canvas.create_text(
                    x + dx, y + dy,
                    text=text,
                    font=font,
                    anchor=anchor,
                    fill="#000000"
                )
        # 再绘制主文字
        self.stats_canvas.create_text(
            x, y,
            text=text,
            font=font,
            anchor=anchor,
            fill=fill
        )
    

    
    def add_ipam_network(self):
        """添加网络"""
        # 创建添加网络对话框
        dialog = ComplexDialog(self.root, _('add_network'), 400, 200)
        
        # 配置内容框架的网格布局
        dialog.content_frame.grid_columnconfigure(0, weight=1)  # 左侧填充列
        dialog.content_frame.grid_columnconfigure(1, weight=0)  # 标签列
        dialog.content_frame.grid_columnconfigure(2, weight=0)  # 输入框列
        dialog.content_frame.grid_columnconfigure(3, weight=1)  # 右侧填充列
        
        # 网络CIDR输入
        _label, network_entry = dialog.add_field(_('network_cidr'), 1, 1, width=25)
        
        # CIDR实时验证（必须带前缀）
        def validate_cidr_input(event=None):
            self.validate_cidr(network_entry.get(), network_entry, require_prefix=True)
        
        network_entry.bind('<FocusOut>', validate_cidr_input)
        network_entry.bind('<KeyRelease>', validate_cidr_input)
        
        # 描述输入
        _label, description_entry = dialog.add_field(_('description'), 2, 1, width=25)
        
        # VLAN 输入
        _label, vlan_entry = dialog.add_field('VLAN', 3, 1, width=25)
        
        # 填充当前选中网段的前缀
        selected_items = self.ipam_network_tree.selection()
        if selected_items:
            selected_network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
            try:
                network_obj = ipaddress.ip_network(selected_network, strict=False)
                network_address = str(network_obj.network_address)
                prefix_len = network_obj.prefixlen
                
                network_prefix = ""
                
                if network_obj.version == 4:
                    # IPv4: 根据前缀长度动态提取网段
                    octets = network_address.split('.')
                    if prefix_len < 8:
                        # 小于/8网段，需要用户完整输入4个8位段
                        network_prefix = ""
                    elif 8 <= prefix_len <= 15:
                        # /8 到 /15 网段，显示前1个8位段
                        network_prefix = f"{octets[0]}."
                    elif 16 <= prefix_len <= 23:
                        # /16 到 /23 网段，显示前2个8位段
                        network_prefix = f"{octets[0]}.{octets[1]}."
                    elif prefix_len >= 24:
                        # 大于等于/24网段，显示前3个8位段
                        network_prefix = f"{octets[0]}.{octets[1]}.{octets[2]}."
                else:
                    # IPv6: 使用 exploded 属性提取前缀
                    if prefix_len == 0:
                        network_prefix = ""
                    elif prefix_len >= 128:
                        network_prefix = str(network_obj.network_address)
                    else:
                        blocks_needed = min(8, (prefix_len + 15) // 16)
                        exploded_parts = network_obj.network_address.exploded.split(':')
                        prefix_parts = [part.lstrip('0') or '0' for part in exploded_parts[:blocks_needed]]
                        network_prefix = ':'.join(prefix_parts) + '::'
                
                if network_prefix:
                    network_entry.insert(0, network_prefix)
                    # 清除选择，设置光标位置到前缀末尾
                    network_entry.selection_clear()
                    network_entry.icursor(len(network_prefix))
            except Exception:
                # 如果解析失败，不填充前缀
                pass
        
        def on_add():
            network = network_entry.get().strip()
            description = description_entry.get().strip()
            vlan = vlan_entry.get().strip()
            
            if not network:
                self.show_error(_('error'), _('please_enter_network'))
                return
            
            # 验证CIDR格式（必须带前缀）
            if not self.validate_cidr(network, require_prefix=True):
                self.show_error(_('error'), _('invalid_cidr_format'))
                return
            
            # 验证VLAN字段
            if vlan:
                if not vlan.isdigit():
                    self.show_info(_('hint'), _('vlan_invalid_format'))
                    return
                vlan_num = int(vlan)
                if vlan_num < 1 or vlan_num > 4094:
                    self.show_info(_('hint'), _('vlan_out_of_range'))
                    return
            
            result = self.ipam.add_network(network, description, vlan)
            # 处理返回值，可能是 (success, message) 或 (success, message, is_overlap)
            if len(result) == 3:
                success, message, is_overlap = result
            else:
                success, message = result
                is_overlap = False
            
            if success:
                self.show_info(_('success'), message)
                self.refresh_ipam_networks()
                dialog.destroy()
            else:
                if is_overlap:
                    # 弹出确认对话框，让用户决定是否继续添加
                    confirm_dialog = ConfirmDialog(self.root, _('confirm'), f"{message}\n\n{_('continue_adding')}？", _('yes'), _('no'))
                    confirm = confirm_dialog.show()
                    if confirm:
                        # 强制添加网络，跳过重叠检查
                        try:
                            from datetime import datetime
                            ip_network = ipaddress.ip_network(network, strict=False)
                            network_str = str(ip_network)
                            self.ipam.networks[network_str] = {
                                'description': description,
                                'ip_addresses': {},
                                'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                            self.ipam.save_data()
                            self.show_info(_('success'), _('network_added_successfully'))
                            self.refresh_ipam_networks()
                            dialog.destroy()
                        except Exception as e:
                            self.show_error(_('error'), f"{_('add_network_failed')}: {str(e)}")
                else:
                    self.show_error(_('error'), message)
        
        # 添加取消按钮
        def on_cancel():
            dialog.destroy()
        
        # 添加按钮
        dialog.add_button(_('cancel'), on_cancel, column=1)
        dialog.add_button(_('ok'), on_add, column=2)
        
        # 绑定回车键（默认确认）和Escape键（默认取消）
        dialog.dialog.bind('<Return>', lambda e: on_add())
        dialog.dialog.bind('<Escape>', lambda e: on_cancel())
        
        # 设置焦点到第一个输入框
        network_entry.focus_force()
        
        # 显示对话框
        dialog.show()
    
    def remove_ipam_network(self):
        """移除网络（支持多选）"""
        selected_items = self.ipam_network_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_network'))
            return
        
        # 获取所有选中的网络
        networks_to_remove = []
        for item in selected_items:
            network = self.ipam_network_tree.item(item, 'values')[0]
            if network:
                networks_to_remove.append(network)
        
        # 根据选择数量显示不同的确认消息
        if len(networks_to_remove) == 1:
            confirm_msg = _('confirm_remove_network')
        else:
            confirm_msg = _('confirm_remove_networks').format(count=len(networks_to_remove))
        
        # 确认删除
        if self.show_yes_no_dialog(_('confirmation'), confirm_msg):
            success_count = 0
            failed_count = 0
            
            for network in networks_to_remove:
                success, message = self.ipam.remove_network(network)
                if success:
                    success_count += 1
                else:
                    failed_count += 1
            
            if success_count > 0:
                self.refresh_ipam_networks()
                # 清空IP地址列表
                for item in self.ipam_ip_tree.get_children():
                    self.ipam_ip_tree.delete(item)
                
                if failed_count == 0:
                    self.show_info(_('success'), _('remove_networks_success').format(count=success_count))
                else:
                    self.show_info(_('success'), _('remove_networks_partial').format(success=success_count, failed=failed_count))
            else:
                self.show_error(_('error'), _('remove_networks_failed'))
    
    def add_ipam_sample_data(self):
        """添加IPAM样例数据"""
        # 检查是否已有数据，如果有则不添加样例数据
        networks = self.ipam.get_all_networks()
        if networks:
            return
        
        from i18n import get_language, translator
        current_lang = get_language()
        
        # 获取当前语言的样例网络数据
        sample_networks = translator.translations.get('sample_networks', {}).get(current_lang, [])
        if not sample_networks:
            # 如果当前语言没有样例数据，使用默认数据
            sample_networks = [
                ['10.0.0.0/8', 'Corporate Network'],
                ['192.168.0.0/16', 'Office Network'],
                ['10.0.0.0/16', 'Headquarters Network'],
                ['192.168.1.0/24', 'Office Room Network'],
                ['10.0.0.0/24', 'Server Network'],
                ['10.0.1.0/24', 'HR Department Network'],
                ['10.0.2.0/24', 'Employee Wireless Network'],
                ['2001:db8::/32', 'Corporate IPv6 Network'],
                ['fd00::/16', 'Internal IPv6 Network'],
                ['2001:db8:1::/48', 'Headquarters IPv6 Network'],
                ['fd00:1::/48', 'Office IPv6 Network'],
                ['2001:db8:1:1::/64', 'Server IPv6 Network'],
                ['2001:db8:1:2::/64', 'HR Department IPv6 Network'],
                ['fd00:1:1::/64', 'Employee Wireless IPv6 Network']
            ]
        
        # 添加样例网络
        for network in sample_networks:
            if len(network) >= 2:
                network_address, description = network[0], network[1]
                self.ipam.add_network(network_address, description)
        
        # 获取当前语言的样例IP数据
        sample_ips = translator.translations.get('sample_ips', {}).get(current_lang, [])
        if not sample_ips:
            # 如果当前语言没有样例数据，使用默认数据
            sample_ips = [
                ['192.168.1.0/24', '192.168.1.10', 'PC-001', 'General Manager Office'],
                ['192.168.1.0/24', '192.168.1.11', 'PC-002', 'Technical Department Manager'],
                ['192.168.1.0/24', '192.168.1.20', 'PC-003', 'Technical Department Employee'],
                ['192.168.1.0/24', '192.168.1.21', 'PC-004', 'Technical Department Employee'],
                ['192.168.1.0/24', '192.168.1.1', 'Router', ''],
                ['192.168.1.0/24', '192.168.1.254', 'Gateway', ''],
                ['10.0.0.0/24', '10.0.0.10', 'Server-001', 'File Server'],
                ['10.0.0.0/24', '10.0.0.11', 'Server-002', 'Database Server'],
                ['10.0.0.0/24', '10.0.0.12', 'Server-003', 'Web Server'],
                ['10.0.0.0/24', '10.0.0.1', 'Router', ''],
                ['10.0.0.0/24', '10.0.0.254', 'Gateway', ''],
                ['10.0.1.0/24', '10.0.1.10', 'HR-001', 'HR Manager'],
                ['10.0.1.0/24', '10.0.1.11', 'HR-002', 'HR Staff'],
                ['2001:db8:1:1::/64', '2001:db8:1:1::10', 'Server-IPv6-001', 'File Server'],
                ['2001:db8:1:1::/64', '2001:db8:1:1::11', 'Server-IPv6-002', 'Database Server'],
                ['2001:db8:1:1::/64', '2001:db8:1:1::12', 'Server-IPv6-003', 'Web Server'],
                ['2001:db8:1:1::/64', '2001:db8:1:1::1', 'Router', ''],
                ['2001:db8:1:1::/64', '2001:db8:1:1::ffff', 'Gateway', ''],
                ['2001:db8:1:2::/64', '2001:db8:1:2::10', 'HR-IPv6-001', 'HR Manager'],
                ['2001:db8:1:2::/64', '2001:db8:1:2::11', 'HR-IPv6-002', 'HR Staff'],
                ['fd00:1:1::/64', 'fd00:1:1::10', 'WiFi-001', 'Wireless Terminal 1'],
                ['fd00:1:1::/64', 'fd00:1:1::11', 'WiFi-002', 'Wireless Terminal 2'],
                ['fd00:1:1::/64', 'fd00:1:1::1', 'Wireless Router', '']
            ]
        
        # 获取当前语言的保留设备名称
        current_language = get_language()
        
        # 读取翻译文件获取设备名称列表（使用全局缓存避免重复读取）
        if not hasattr(self, '_translations_cache'):
            translations_file = os.path.join(os.path.dirname(__file__), 'translations.json')
            try:
                with open(translations_file, 'r', encoding='utf-8') as f:
                    self._translations_cache = json.load(f)
            except (FileNotFoundError, json.JSONDecodeError) as e:
                print(f"Failed to load translations: {e}")
                self._translations_cache = {}
        
        reserved_device_names = self._translations_cache.get('sample_device_names', {}).get(current_language, [])
        
        # 添加样例IP地址
        for ip_data in sample_ips:
            if len(ip_data) >= 4:
                network, ip_address, hostname, description = ip_data[0], ip_data[1], ip_data[2], ip_data[3]
                if hostname in reserved_device_names:
                    # 保留地址
                    self.ipam.reserve_ip(network, ip_address, hostname, description)
                else:
                    # 分配地址
                    self.ipam.allocate_ip(network, ip_address, hostname, description)
        
        # 刷新IPAM数据
        self.refresh_ipam_networks()
    
    def _get_all_visible_network_items(self):
        """获取网段树中所有可见的项ID列表（用于Shift范围选择）"""
        items = []

        def collect_visible(parent=''):
            for child in self.ipam_network_tree.get_children(parent):
                items.append(child)
                if self.ipam_network_tree.item(child, 'open'):
                    collect_visible(child)
        collect_visible()
        return items
    
    def on_ipam_network_click(self, event):
        """网络点击事件处理（用于取消选择）"""
        result = self._handle_inline_edit_validation_before_click()
        if result == 'break':
            return result
        
        # 获取点击的列
        column = self.ipam_network_tree.identify_column(event.x)
        
        # 获取点击的项
        clicked_item = self.ipam_network_tree.identify_row(event.y)
        selected_items = self.ipam_network_tree.selection()
        
        # 检查是否点击了展开/折叠图标
        if column == '#0' and clicked_item:
            # 检查当前项是否有子节点（即是否有展开/折叠图标）
            children = self.ipam_network_tree.get_children(clicked_item)
            if children:
                # 获取点击的x坐标相对于Treeview的位置
                x = event.x
                
                # 获取当前项的深度（缩进级别）
                depth = 0
                parent = self.ipam_network_tree.parent(clicked_item)
                while parent:
                    depth += 1
                    parent = self.ipam_network_tree.parent(parent)
                
                # 计算展开/折叠图标的位置
                # 从截图观察，每个层级的缩进量是10px
                indent_per_level = 10
                
                # 计算图标起始和结束位置
                # 第一层级图标在8-18px位置，每个层级向右移动10px
                # 增加检测范围到10px，让用户更容易点击
                icon_x_start = 2 + (depth * indent_per_level)
                icon_x_end = 20 + (depth * indent_per_level)
                
                # 检查点击的x坐标是否在展开/折叠图标的范围内
                if icon_x_start <= x < icon_x_end:
                    # 点击的是展开/折叠图标，直接处理，不使用定时器
                    # 切换节点的展开/折叠状态
                    if self.ipam_network_tree.item(clicked_item, 'open'):
                        self.ipam_network_tree.item(clicked_item, open=False)
                    else:
                        self.ipam_network_tree.item(clicked_item, open=True)
                    # 刷新斑马纹
                    self.ipam_network_tree.after(100, lambda: self.update_table_zebra_stripes(self.ipam_network_tree))
                    # 阻止默认的选择行为
                    return 'break'
        
        # 检查点击的区域是否是单元格、行或树区域
        # 如果是这些区域，可能是双击编辑的一部分，不执行展开/收缩功能
        region = self.ipam_network_tree.identify_region(event.x, event.y)
        if region in ("cell", "row", "tree"):
            # 获取按键状态
            is_ctrl = event.state & 0x4
            is_shift = event.state & 0x1
            
            # Ctrl+点击：切换选择状态
            if is_ctrl and clicked_item:
                if clicked_item in selected_items:
                    self.ipam_network_tree.selection_remove(clicked_item)
                    # 如果没有选中项了，清空IP地址列表
                    if not self.ipam_network_tree.selection():
                        for item in self.ipam_ip_tree.get_children():
                            self.ipam_ip_tree.delete(item)
                else:
                    self.ipam_network_tree.selection_add(clicked_item)
                    self.on_ipam_network_select(None)
                return 'break'
            # Shift+点击：范围选择
            elif is_shift and clicked_item and selected_items:
                all_items = self._get_all_visible_network_items()
                last_selected = selected_items[-1]
                try:
                    start_idx = all_items.index(last_selected)
                    end_idx = all_items.index(clicked_item)
                    self.ipam_network_tree.selection_clear()
                    for idx in range(min(start_idx, end_idx), max(start_idx, end_idx) + 1):
                        self.ipam_network_tree.selection_add(all_items[idx])
                    self.on_ipam_network_select(None)
                except ValueError:
                    pass
                return 'break'
            
            # 检查是否点击了已选中的项
            if clicked_item and clicked_item in selected_items and len(selected_items) == 1:
                # 使用定时器延迟处理取消选择，这样在双击时，定时器会被取消
                def handle_single_click_cancel():
                    # 重新获取选择状态，因为可能在延迟期间发生变化
                    current_selected = self.ipam_network_tree.selection()
                    current_clicked = self.ipam_network_tree.identify_row(event.y)
                    
                    # 如果点击的是已选中的项，取消选择
                    if current_clicked and current_clicked in current_selected and len(current_selected) == 1:
                        # 取消选择
                        self.ipam_network_tree.selection_remove(current_clicked)
                        # 清空IP地址列表
                        for item in self.ipam_ip_tree.get_children():
                            self.ipam_ip_tree.delete(item)
                
                # 设置定时器，延迟30毫秒执行取消选择
                # 这样在双击时，定时器会被取消，避免与编辑功能冲突
                self._click_timer = self.root.after(30, handle_single_click_cancel)
                # 允许默认的选择行为
                return
            
            # 如果点击的是未选中的项，使用定时器延迟处理
            def handle_single_click_cell():
                # 重新获取选择状态，因为可能在延迟期间发生变化
                current_selected = self.ipam_network_tree.selection()
                current_clicked = self.ipam_network_tree.identify_row(event.y)
                
                # 如果点击的是未选中的项，手动选中该行
                if current_clicked and current_clicked not in current_selected:
                    # 清除之前的选择
                    self.ipam_network_tree.selection_clear()
                    # 选中点击的项
                    self.ipam_network_tree.selection_add(current_clicked)
                    # 触发选择事件
                    self.on_ipam_network_select(None)
            
            # 设置定时器，延迟30毫秒执行单击处理
            # 这样在双击时，定时器会被取消，避免与编辑功能冲突
            self._click_timer = self.root.after(30, handle_single_click_cell)
            # 允许默认的选择行为
            return
        
        # 检查是否是双击事件的一部分
        # 当发生双击时，单击事件会被触发两次，然后双击事件才会被触发
        # 我们使用定时器来区分单击和双击
        if hasattr(self, '_click_timer'):
            self.root.after_cancel(self._click_timer)
        
        # 定义单击处理函数
        def handle_single_click():
            # 这里处理其他单击事件，比如点击空白区域等
            pass
        
        # 设置定时器，延迟150毫秒执行单击处理
        self._click_timer = self.root.after(30, handle_single_click)
    
    def on_ipam_ip_click(self, event):
        """IP地址表点击事件处理（用于取消选择）"""
        result = self._handle_inline_edit_validation_before_click()
        if result == 'break':
            return result
        
        # 获取点击的项
        clicked_item = self.ipam_ip_tree.identify_row(event.y)
        selected_items = self.ipam_ip_tree.selection()
        
        # 检查点击的区域是否是单元格或行
        region = self.ipam_ip_tree.identify_region(event.x, event.y)
        if region in ("cell", "row"):
            # 获取按键状态
            is_ctrl = event.state & 0x4
            is_shift = event.state & 0x1
            
            # Ctrl+点击：切换选择状态
            if is_ctrl and clicked_item:
                if clicked_item in selected_items:
                    self.ipam_ip_tree.selection_remove(clicked_item)
                else:
                    self.ipam_ip_tree.selection_add(clicked_item)
                return 'break'
            # Shift+点击：范围选择
            elif is_shift and clicked_item and selected_items:
                all_items = self.ipam_ip_tree.get_children()
                last_selected = selected_items[-1]
                try:
                    start_idx = list(all_items).index(last_selected)
                    end_idx = list(all_items).index(clicked_item)
                    self.ipam_ip_tree.selection_clear()
                    for idx in range(min(start_idx, end_idx), max(start_idx, end_idx) + 1):
                        self.ipam_ip_tree.selection_add(all_items[idx])
                except ValueError:
                    pass
                return 'break'
            
            # 检查是否点击了已选中的项
            if clicked_item and clicked_item in selected_items and len(selected_items) == 1:
                # 使用定时器延迟处理取消选择，这样在双击时，定时器会被取消
                def handle_single_click_cancel():
                    # 重新获取选择状态，因为可能在延迟期间发生变化
                    current_selected = self.ipam_ip_tree.selection()
                    current_clicked = self.ipam_ip_tree.identify_row(event.y)
                    
                    # 如果点击的是已选中的项，取消选择
                    if current_clicked and current_clicked in current_selected and len(current_selected) == 1:
                        # 取消选择
                        self.ipam_ip_tree.selection_remove(current_clicked)
                
                # 设置定时器，延迟30毫秒执行取消选择
                # 这样在双击时，定时器会被取消，避免与编辑功能冲突
                self._click_timer = self.root.after(30, handle_single_click_cancel)
                # 允许默认的选择行为
                return
            
            # 如果点击的是未选中的项，使用定时器延迟处理
            def handle_single_click_cell():
                # 重新获取选择状态，因为可能在延迟期间发生变化
                current_selected = self.ipam_ip_tree.selection()
                current_clicked = self.ipam_ip_tree.identify_row(event.y)
                
                # 如果点击的是未选中的项，手动选中该行
                if current_clicked and current_clicked not in current_selected:
                    # 清除之前的选择
                    self.ipam_ip_tree.selection_clear()
                    # 选中点击的项
                    self.ipam_ip_tree.selection_add(current_clicked)
            
            # 设置定时器，延迟30毫秒执行单击处理
            # 这样在双击时，定时器会被取消，避免与编辑功能冲突
            self._click_timer = self.root.after(30, handle_single_click_cell)
            # 允许默认的选择行为
            return
        
        # 检查是否是双击事件的一部分
        # 当发生双击时，单击事件会被触发两次，然后双击事件才会被触发
        # 我们使用定时器来区分单击和双击
        if hasattr(self, '_click_timer'):
            self.root.after_cancel(self._click_timer)
        
        # 定义单击处理函数
        def handle_single_click():
            # 这里处理其他单击事件，比如点击空白区域等
            pass
        
        # 设置定时器，延迟150毫秒执行单击处理
        self._click_timer = self.root.after(30, handle_single_click)
    
    def on_ipam_network_double_click(self, event):
        """网络双击事件处理（用于内联编辑）"""
        try:
            # 取消单击定时器，防止触发展开/收缩功能
            if hasattr(self, '_click_timer'):
                self.root.after_cancel(self._click_timer)
            
            # 获取双击的行和列
            item = self.ipam_network_tree.identify_row(event.y)
            column = self.ipam_network_tree.identify_column(event.x)
            
            if not item or not column:
                return
            
            # 检查是否点击了展开/折叠图标
            if column == '#0':
                # 检查当前项是否有子节点（即是否有展开/折叠图标）
                children = self.ipam_network_tree.get_children(item)
                if children:
                    # 获取点击的x坐标相对于Treeview的位置
                    x = event.x
                    
                    # 获取当前项的深度（缩进级别）
                    depth = 0
                    parent = self.ipam_network_tree.parent(item)
                    while parent:
                        depth += 1
                        parent = self.ipam_network_tree.parent(parent)
                    
                    # 计算展开/折叠图标的位置
                    indent_per_level = 10
                    icon_x_start = 2 + (depth * indent_per_level)
                    icon_x_end = 20 + (depth * indent_per_level)
                    
                    # 检查点击的x坐标是否在展开/折叠图标的范围内
                    if icon_x_start <= x < icon_x_end:
                        # 点击的是展开/折叠图标，不触发内联编辑
                        return
            
            # 获取列索引
            if column == '#0':
                # 对于#0列（网段字段），直接设置为0
                column_index = 0
            else:
                # 对于其他列，正常计算
                column_index = int(column[1:]) - 1
            
            # 确保内联编辑数据结构已初始化
            self._init_inline_edit_data()
            
            # 检查是否是可编辑列
            config = self._inline_edit_configs.get('ipam_network', {})
            if column_index not in config.get('editable_columns', []):
                return
            
            # 获取当前单元格的值
            if column_index == 0 and column == '#0':
                # 对于#0列（网段字段），值存储在'text'属性中
                current_value = self.ipam_network_tree.item(item, 'text')
            else:
                # 对于其他列，值存储在'values'属性中
                values = self.ipam_network_tree.item(item, 'values')
                current_value = values[column_index]
            
            # 获取单元格的坐标
            try:
                x, y, width, height = self.ipam_network_tree.bbox(item, column)
            except tk.TclError:
                return
            
            # 安全销毁所有旧的编辑控件，防止内存泄漏
            self._destroy_inline_edit_widgets()
            
            # 获取列名
            columns = self.ipam_network_tree["columns"]
            column_name = columns[column_index] if column_index < len(columns) else f"column_{column_index}"
            
            # 保存相关信息
            self.inline_edit_data = {
                'item': item,
                'column_index': column_index,
                'column_name': column_name,
                'original_value': current_value,
                'tree': self.ipam_network_tree,
                'tree_name': 'ipam_network'
            }
            
            # 绑定事件 - 先定义事件处理函数
            def on_widget_save(event):
                # 用户主动保存，清除验证失败标记，让保存逻辑重新验证
                self._inline_edit_validation_failed = False
                # 直接保存
                self.on_generic_inline_edit_save(None)
            
            def on_widget_focus_out(event):
                # 焦点离开时，主动验证当前输入
                is_valid, error_msg = self._validate_current_inline_edit()
                if not is_valid:
                    self._inline_edit_validation_failed = True
                    self.show_error("", error_msg)
                    self.root.after(1, self._refocus_inline_edit)
                    return
                # 验证通过，直接保存
                self.on_generic_inline_edit_save(None)
            
            def on_widget_cancel(event):
                # 直接取消
                self.on_generic_inline_edit_cancel(event)
            
            def on_widget_input(event):
                # 用户修改输入时，清除验证失败标记，允许后续保存
                self._inline_edit_validation_failed = False
            
            # 根据列类型创建编辑控件
            column_type = config.get('column_types', {}).get(column_index, 'entry')
            
            # 获取字体设置
            font_family, font_size = get_current_font_settings()
            
            if column_type == 'combobox':
                # 为Combobox创建一个特殊的框架，确保下拉列表正常显示
                self.inline_edit_frame = ttk.Frame(self.ipam_network_tree)
                self.inline_edit_frame.place(x=x, y=y, width=width, height=height)
                
                # 获取下拉框值
                combobox_values = config.get('combobox_values', {}).get(column_index, [])
                
                self.inline_edit_widget = ttk.Combobox(self.inline_edit_frame, 
                                                       values=combobox_values,
                                                       width=width // 10 - 2,
                                                       font=(font_family, font_size))
                self.inline_edit_widget.current(combobox_values.index(current_value) if current_value in combobox_values else 0)
            else:
                # 创建Entry控件
                self.inline_edit_widget = ttk.Entry(self.ipam_network_tree, 
                                                   width=width // 10 - 2,
                                                   font=(font_family, font_size))
                self.inline_edit_widget.insert(0, current_value)
            
            # 放置编辑控件
            if hasattr(self, 'inline_edit_frame') and self.inline_edit_frame:
                self.inline_edit_widget.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
            else:
                self.inline_edit_widget.place(x=x, y=y, width=width, height=height)
            
            # 绑定事件
            self.inline_edit_widget.bind('<Return>', on_widget_save)
            self.inline_edit_widget.bind('<FocusOut>', on_widget_focus_out)
            self.inline_edit_widget.bind('<Escape>', on_widget_cancel)
            self.inline_edit_widget.bind('<Key>', on_widget_input)
            
            # 获取焦点并全选文本
            self.inline_edit_widget.focus_set()
            self.inline_edit_widget.select_range(0, tk.END)
        except Exception as e:
            print(f"双击事件处理错误: {str(e)}")
        
    def on_ipam_network_select(self, event):
        """网络选择事件处理"""
        selected_items = self.ipam_network_tree.selection()
        
        if not selected_items:
            return
        
        network = None
        if len(selected_items) == 1:
            # 只选中一个网段，显示该网段的IP地址（带过滤）
            item = selected_items[0]
            network = self.ipam_network_tree.item(item, 'values')[0]
            # 使用apply_filter而不是refresh_ipam_ips，以应用过滤条件
            self.apply_filter()
            # 刷新网络拓扑图
            self.refresh_visualization(network)
        else:
            # 选中多个网段，显示所有选中网段的IP地址（带过滤）
            from datetime import datetime
            # 收集所有IP地址记录，避免重复显示
            all_ip_records = {}
            
            for item in selected_items:
                network_item = self.ipam_network_tree.item(item, 'values')[0]
                network_ips = self.ipam.get_network_ips(network_item)
                
                for ip in network_ips:
                    # 用ID作为唯一标识，避免重复记录
                    if ip['id'] not in all_ip_records:
                        all_ip_records[ip['id']] = ip
            
            # 应用过滤条件
            status = self.ipam_status_filter.get()
            expiry_filter = self.ipam_expiry_filter.get()
            now = datetime.now()
            search_text = self.ipam_search_entry.get().strip()
            search_scope = self.search_scope.get()
            search_mode = self.search_mode.get()
            
            filtered_ips = []
            for ip in all_ip_records.values():
                # 按状态过滤
                if status != _('all'):
                    status_map = self._get_status_map()
                    if status in status_map:
                        if ip['status'] != status_map[status]:
                            continue
                    elif ip['status'] != status:
                        continue
                
                # 按过期日期过滤
                if expiry_filter != _('all'):
                    expiry_date = ip.get('expiry_date')
                    if not expiry_date:
                        if expiry_filter != _('not_expired'):
                            continue
                    else:
                        try:
                            if 'T' in expiry_date:
                                exp_date = datetime.fromisoformat(expiry_date)
                            else:
                                exp_date = datetime.strptime(expiry_date, "%Y-%m-%d %H:%M:%S")
                            
                            if expiry_filter == _('expired'):
                                if exp_date >= now:
                                    continue
                            elif expiry_filter == _('expiring_soon'):
                                seven_days_later = now + timedelta(days=7)
                                if exp_date < now or exp_date > seven_days_later:
                                    continue
                            elif expiry_filter == _('not_expired'):
                                if exp_date < now:
                                    continue
                        except (ValueError, TypeError):
                            if expiry_filter != _('not_expired'):
                                continue
                
                # 按搜索关键词过滤
                if search_text:
                    keywords = search_text.split()
                    match = False
                    
                    ip_match = self._match_search_pattern(ip['ip_address'], ' '.join(keywords), search_mode)
                    hostname_match = self._match_search_pattern(ip.get('hostname', ''), ' '.join(keywords), search_mode)
                    desc_match = self._match_search_pattern(ip.get('description', ''), ' '.join(keywords), search_mode)
                    mac_match = self._match_search_pattern(ip.get('mac_address', ''), ' '.join(keywords), search_mode)
                    
                    if search_scope == _('all'):
                        match = ip_match or hostname_match or desc_match or mac_match
                    elif search_scope == _('ip_address'):
                        match = ip_match
                    elif search_scope == _('hostname'):
                        match = hostname_match
                    elif search_scope == _('description'):
                        match = desc_match
                    elif search_scope == _('mac_address'):
                        match = mac_match
                    
                    if not match:
                        continue
                
                filtered_ips.append(ip)
            
            # 显示过滤后的IP地址
            self.ipam_ip_tree.delete(*self.ipam_ip_tree.get_children())
            # 转换为列表并排序
            sorted_ips = self._sort_ip_list(filtered_ips)
            
            for ip in sorted_ips:
                status_text = ip['status']
                if status_text == 'reserved':
                    status_text = _('reserved')
                elif status_text == 'released':
                    status_text = _('released')
                elif status_text == 'allocated':
                    status_text = _('allocated')
                
                allocated_at = ip.get('allocated_at', '')
                try:
                    from datetime import datetime
                    if 'T' in allocated_at:
                        dt = datetime.fromisoformat(allocated_at)
                    else:
                        dt = datetime.strptime(allocated_at, "%Y-%m-%d %H:%M:%S.%f")
                    formatted_allocated_at = dt.strftime("%Y-%m-%d %H:%M:%S")
                except (ValueError, TypeError):
                    formatted_allocated_at = allocated_at
                
                expiry_date = ip.get('expiry_date', '')
                formatted_expiry_date = expiry_date
                try:
                    from datetime import datetime
                    if expiry_date:
                        if 'T' in expiry_date:
                            dt = datetime.fromisoformat(expiry_date)
                        elif ' ' in expiry_date:
                            dt = datetime.strptime(expiry_date, "%Y-%m-%d %H:%M:%S")
                        else:
                            dt = datetime.strptime(expiry_date, "%Y-%m-%d")
                        formatted_expiry_date = dt.strftime("%Y-%m-%d")
                except (ValueError, TypeError):
                    pass
                
                # 使用记录ID作为树项的ID
                # 使用数据库记录ID作为树项的tags，确保能可靠获取数据库ID
                db_record_id = ip.get('id', None)
                tags = (f'dbid_{db_record_id}',) if db_record_id is not None else ()
                # 生成唯一的iid，使用数据库ID或IP地址+描述的组合
                if db_record_id is not None:
                    iid = f'rec_{db_record_id}'
                else:
                    iid = f'ip_{ip["ip_address"]}_{ip.get("description", "")}_{ip.get("status", "")}'
                self.ipam_ip_tree.insert('', tk.END, iid=iid, tags=tags, values=(
                    ip['ip_address'],
                    status_text,
                    ip.get('hostname', ''),
                    ip.get('mac_address', ''),
                    ip.get('description', ''),
                    formatted_allocated_at,
                    formatted_expiry_date
                ))
            
            # 更新斑马纹样式
            self.update_table_zebra_stripes(self.ipam_ip_tree)
            
            # 选中多个网段时，使用第一个网段刷新网络拓扑
            if selected_items:
                network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
                self.refresh_visualization(network)
        
        # 显示网络前缀到IP输入框
        if network:
            try:
                network_obj = ipaddress.ip_network(network, strict=False)
                network_address = str(network_obj.network_address)
                
                # 根据网络前缀长度确定显示的前缀段数
                if network_obj.prefixlen >= 24:
                    # 对于/24及更窄的网络，显示前三个段
                    prefix_parts = network_address.split('.')[:3]
                    prefix = '.'.join(prefix_parts) + '.'
                elif network_obj.prefixlen >= 16:
                    # 对于/16到/23的网络，显示前两个段
                    prefix_parts = network_address.split('.')[:2]
                    prefix = '.'.join(prefix_parts) + '.'
                else:
                    # 对于更宽的网络，显示第一个段
                    prefix_parts = network_address.split('.')[:1]
                    prefix = '.'.join(prefix_parts) + '.'
                

            except Exception as e:
                print(f"网络选择处理失败: {str(e)}")
                pass
    
    def on_ip_input(self, event):
        """IP输入事件处理 - 已简化，不再需要"""
        pass
    
    def on_ip_tree_right_click(self, event):
        """IP地址表格右键菜单处理"""
        try:
            # 获取当前右键点击的行
            item = self.ipam_ip_tree.identify_row(event.y)
            if not item:
                return
        
            # 处理多选逻辑
            selected_items = self.ipam_ip_tree.selection()
            if item not in selected_items:
                self.ipam_ip_tree.selection_set(item)
                selected_items = [item]
            
            # 收集所有选中行的状态
            statuses = set()
            for item in selected_items:
                row_values = self.ipam_ip_tree.item(item, 'values')
                if row_values and len(row_values) >= 2:
                    statuses.add(row_values[1])
            
            if not statuses:
                return
            
            # 使用原生 tk.Menu 创建右键菜单（保持原生样式和阴影）
            menu = tk.Menu(self.root, tearoff=0)
            
            # 根据选中行的状态动态添加菜单项
            has_commands = False
            if any(status != _('released') for status in statuses):
                menu.add_command(label=_('release_address'), command=lambda: self.on_ip_menu_action('release'))
                has_commands = True
            
            has_quick_actions = False
            
            if _('released') in statuses or _('reserved') in statuses:
                if has_commands:
                    menu.add_separator()
                menu.add_command(label=_('convert_to_allocated'), command=lambda: self.on_ip_menu_action('quick_allocate'))
                has_quick_actions = True
                has_commands = True
            
            if _('released') in statuses or _('allocated') in statuses:
                if has_commands and not has_quick_actions:
                    menu.add_separator()
                menu.add_command(label=_('convert_to_reserved'), command=lambda: self.on_ip_menu_action('quick_reserve'))
                has_quick_actions = True
                has_commands = True
            
            # 添加清理地址菜单项
            if _('released') in statuses:
                if has_commands:
                    menu.add_separator()
                menu.add_command(label=_('cleanup_address'), command=lambda: self.on_ip_menu_action('cleanup'))
                has_commands = True
            
            if len(selected_items) == 1:
                row_values = self.ipam_ip_tree.item(selected_items[0], 'values')
                current_status = row_values[1]
                
                if current_status == _('released'):
                    if has_commands:
                        menu.add_separator()
                    menu.add_command(label=_('reallocate_address'), command=lambda: self.on_ip_menu_action('allocate'))
                    menu.add_command(label=_('reserve_address_again'), command=lambda: self.on_ip_menu_action('reserve'))
            
            # 显示菜单（使用原生样式和阴影）
            try:
                # 显示原生菜单
                menu.post(event.x_root, event.y_root)
                
                # 通过DWM API禁用非客户区渲染（解决叠影问题）
                # 这样只保留tkinter菜单自己的阴影，去掉Windows 11添加的额外圆角阴影
                try:
                    import ctypes
                    
                    # 获取菜单窗口的句柄
                    hwnd = ctypes.windll.user32.GetParent(menu.winfo_id())
                    
                    # 定义常量
                    DWMNCRP_DISABLED = 1  # 禁用非客户区渲染
                    DWMWA_NCRENDERING_POLICY = 2  # 非客户区渲染策略属性ID
                    
                    # 禁用DWM的非客户区渲染（包括额外的圆角阴影）
                    ctypes.windll.dwmapi.DwmSetWindowAttribute(
                        hwnd,
                        DWMWA_NCRENDERING_POLICY,
                        ctypes.byref(ctypes.c_int(DWMNCRP_DISABLED)),
                        ctypes.sizeof(ctypes.c_int())
                    )
                except Exception as e:
                    # 如果API调用失败，不影响菜单正常显示
                    pass
                
            except Exception as e:
                print(f"菜单显示错误: {e}")
        except Exception as e:
            print(f"右键菜单处理错误: {e}")

    def _on_ip_popup_action(self, action, popup):
        """处理弹出菜单的操作"""
        # 先关闭所有阴影
        try:
            for s in self._current_ip_shadows:
                try:
                    s.destroy()
                except tk.TclError:
                    pass
        except tk.TclError:
            pass
        # 再关闭菜单
        try:
            self._current_ip_popup.destroy()
        except tk.TclError:
            pass
        # 移除事件绑定
        if self._ip_popup_click_bind_id:
            try:
                self.root.unbind('<Button-1>', self._ip_popup_click_bind_id)
            except Exception:
                pass
        # 执行菜单操作
        self.on_ip_menu_action(action)
    
    def _cleanup_ip_popup_menu(self):
        """清理弹出菜单"""
        # 清理事件绑定
        if hasattr(self, '_ip_popup_click_bind_id') and self._ip_popup_click_bind_id:
            try:
                self.root.unbind('<Button-1>', self._ip_popup_click_bind_id)
            except Exception:
                pass
            self._ip_popup_click_bind_id = None
        
        # 清理阴影窗口
        if hasattr(self, '_ip_shadow_windows') and self._ip_shadow_windows:
            for shadow in self._ip_shadow_windows:
                try:
                    shadow.destroy()
                except Exception:
                    pass
            self._ip_shadow_windows = []
        
        # 清理当前菜单引用
        if hasattr(self, '_current_ip_shadows'):
            self._current_ip_shadows = []
        if hasattr(self, '_current_ip_popup'):
            self._current_ip_popup = None
        
        # 清理菜单窗口
        if hasattr(self, '_ip_popup_menu') and self._ip_popup_menu:
            try:
                self._ip_popup_menu.destroy()
            except Exception:
                pass
            self._ip_popup_menu = None
    
    def _cleanup_ip_menu(self):
        """清理IP树菜单资源"""
        if hasattr(self, 'ip_tree_menu'):
            try:
                if self.ip_tree_menu:
                    self.ip_tree_menu.unpost()
                    self.ip_tree_menu.destroy()
            except Exception:
                pass
            finally:
                delattr(self, 'ip_tree_menu')
    
    def _execute_menu_action(self, action):
        """执行菜单操作，不使用实例变量保存菜单"""
        print(f"执行菜单操作: {action}")
        # 直接执行操作，不通过实例变量菜单
        self.on_ip_menu_action(action)
    
    def _cleanup_all_menus(self):
        """清理所有可能存在的菜单资源"""
        # 清理IP树菜单
        self._cleanup_ip_menu()
    
    def __del__(self):
        """销毁对象时清理所有资源"""
        self._cleanup_all_menus()
    
    def _destroy_inline_edit_widgets(self):
        """安全销毁内联编辑控件，防止内存泄漏
        
        此方法确保所有内联编辑相关的控件和数据被正确清理
        """
        try:
            if hasattr(self, '_inline_edit_save_after_id') and self._inline_edit_save_after_id:
                self.root.after_cancel(self._inline_edit_save_after_id)
                delattr(self, '_inline_edit_save_after_id')
        except Exception:
            pass
        try:
            if hasattr(self, '_inline_edit_save_timer') and self._inline_edit_save_timer:
                self.root.after_cancel(self._inline_edit_save_timer)
                delattr(self, '_inline_edit_save_timer')
        except Exception:
            pass
        try:
            if hasattr(self, '_calendar_close_check_timer') and self._calendar_close_check_timer:
                self.root.after_cancel(self._calendar_close_check_timer)
                delattr(self, '_calendar_close_check_timer')
        except Exception:
            pass
        try:
            if hasattr(self, 'inline_edit_widget'):
                # 对于DateEntry，先关闭日历窗口再销毁
                if DateEntry is not None and isinstance(self.inline_edit_widget, DateEntry):
                    try:
                        if hasattr(self.inline_edit_widget, '_top_cal'):
                            top_cal = self.inline_edit_widget._top_cal
                            if top_cal:
                                top_cal.withdraw()
                    except Exception:
                        pass
                self.inline_edit_widget.destroy()
                delattr(self, 'inline_edit_widget')
        except Exception:
            pass
        try:
            if hasattr(self, 'inline_edit_frame'):
                self.inline_edit_frame.destroy()
                delattr(self, 'inline_edit_frame')
        except Exception:
            pass
        try:
            if hasattr(self, 'inline_edit_data'):
                delattr(self, 'inline_edit_data')
        except Exception:
            pass
        
        try:
            # 移除根窗口点击事件绑定
            if hasattr(self, '_inline_edit_click_handler_id'):
                self.root.unbind('<Button-1>', self._inline_edit_click_handler_id)
                delattr(self, '_inline_edit_click_handler_id')
        except Exception:
            pass
    
    def _init_inline_edit_data(self):
        """初始化内联编辑相关的数据属性
        
        确保所有需要的属性都已初始化，避免AttributeError
        """
        if not hasattr(self, '_inline_edit_handlers'):
            self._inline_edit_handlers = {}
        if not hasattr(self, '_inline_edit_configs'):
            self._inline_edit_configs = {}
    
    def register_inline_edit_handler(self, tree_name, handlers):
        """注册内联编辑处理器
        
        Args:
            tree_name: 表格名称标识
            handlers: 处理器字典，包含以下键：
                - validate: 验证函数，接收(new_value, column_name, row_data)，返回(是否有效, 错误消息或None)
                - save: 保存函数，接收(new_value, column_name, row_data, item)，返回(是否成功, 消息)
                - get_row_data: 获取行数据函数，接收(item)，返回行数据字典
        """
        self._init_inline_edit_data()
        self._inline_edit_handlers[tree_name] = handlers
    
    def register_inline_edit_config(self, tree_name, config):
        """注册内联编辑配置
        
        Args:
            tree_name: 表格名称标识
            config: 配置字典，包含以下键：
                - editable_columns: 可编辑列的索引列表
                - column_types: 列类型字典，键为列索引，值为'entry'或'combobox'
                - combobox_values: 下拉框值字典，键为列索引，值为可选值列表
        """
        self._init_inline_edit_data()
        self._inline_edit_configs[tree_name] = config
    
    def on_generic_tree_double_click(self, tree, tree_name, event):
        """通用的双击Treeview单元格编辑处理函数
        
        Args:
            tree: Treeview组件
            tree_name: 表格名称标识
            event: 事件对象
        """
        result = self._handle_inline_edit_validation_before_click()
        if result == 'break':
            return result
        
        # 检查是否有注册的配置和处理器
        if tree_name not in self._inline_edit_configs:
            return
        config = self._inline_edit_configs[tree_name]
        if tree_name not in self._inline_edit_handlers:
            return
        
        # 检查是否处于双击冷却期（防止三击被误判为两次双击）
        if hasattr(self, '_in_double_click_cooldown') and self._in_double_click_cooldown:
            return 'break'
        
        # 获取双击的行和列
        region = tree.identify_region(event.x, event.y)
        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        
        # 无论是否是真正的双击，都阻止默认的展开/收缩行为
        # 这样网段表的双击就不会触发展开/收缩层级
        
        # 允许在cell或tree区域触发编辑，特别是#0列（网段字段）
        if region not in ("cell", "tree"):
            return
        
        if not item or not column:
            return
        
        # 确保双击的行被选中
        tree.selection_set(item)
        
        # 检查是否点击了展开/折叠图标
        if column == '#0':
            # 检查当前项是否有子节点（即是否有展开/折叠图标）
            children = tree.get_children(item)
            if children:
                # 获取点击的x坐标相对于Treeview的位置
                x = event.x
                
                # 获取当前项的深度（缩进级别）
                depth = 0
                parent = tree.parent(item)
                while parent:
                    depth += 1
                    parent = tree.parent(parent)
                
                # 计算展开/折叠图标的位置
                indent_per_level = 10
                icon_x_start = 2 + (depth * indent_per_level)
                icon_x_end = 20 + (depth * indent_per_level)
                
                # 检查点击的x坐标是否在展开/折叠图标的范围内
                if icon_x_start <= x < icon_x_end:
                    # 点击的是展开/折叠图标，不触发内联编辑
                    return
        
        # 获取列索引
        if column == '#0':
            # 对于#0列（网段字段），直接设置为0
            column_index = 0
        else:
            # 对于其他列，正常计算
            column_index = int(column[1:]) - 1
        
        # 检查是否是可编辑列
        if column_index not in config['editable_columns']:
            return
        
        # 获取当前单元格的值
        if column_index == 0 and column == '#0':
            # 对于#0列（网段字段），值存储在'text'属性中
            current_value = tree.item(item, 'text')
        else:
            # 对于其他列，值存储在'values'属性中
            values = tree.item(item, 'values')
            current_value = values[column_index]
        
        # 获取单元格的坐标
        try:
            x, y, width, height = tree.bbox(item, column)
        except tk.TclError:
            return
        
        # 安全销毁所有旧的编辑控件，防止内存泄漏
        self._destroy_inline_edit_widgets()
        
        # 获取列名
        columns = tree["columns"]
        column_name = columns[column_index] if column_index < len(columns) else f"column_{column_index}"
        
        # 保存相关信息
        self.inline_edit_data = {
            'item': item,
            'column_index': column_index,
            'column_name': column_name,
            'original_value': current_value,
            'tree': tree,
            'tree_name': tree_name
        }
        
        # 绑定事件 - 先定义事件处理函数
        def on_widget_save(event):
            # 用户主动保存，清除验证失败标记，让保存逻辑重新验证
            self._inline_edit_validation_failed = False
            # 直接保存
            self.on_generic_inline_edit_save(None)
        
        def on_widget_focus_out(event):
            # 焦点离开时，主动验证当前输入
            is_valid, error_msg = self._validate_current_inline_edit()
            if not is_valid:
                self._inline_edit_validation_failed = True
                self.show_error("", error_msg)
                self.root.after(1, self._refocus_inline_edit)
                return
            # 验证通过，执行保存
            self.on_generic_inline_edit_save(None)
        
        def on_widget_cancel(event):
            # 直接取消
            self.on_generic_inline_edit_cancel(event)
        
        def on_widget_input(event):
            # 用户修改输入时，清除验证失败标记，允许后续保存
            self._inline_edit_validation_failed = False
        
        # 根据列类型创建编辑控件
        column_type = config.get('column_types', {}).get(column_index, 'entry')
        
        if column_type == 'combobox':
            # 为Combobox创建一个特殊的框架，确保下拉列表正常显示
            self.inline_edit_frame = ttk.Frame(tree)
            self.inline_edit_frame.place(x=x, y=y, width=width, height=height)
            
            # 获取下拉框值
            combobox_values = config.get('combobox_values', {}).get(column_index, [])
            
            self.inline_edit_widget = ttk.Combobox(self.inline_edit_frame, 
                                                   values=combobox_values,
                                                   width=width // 10 - 2,
                                                   state='readonly')
            self.inline_edit_widget.pack(fill=tk.BOTH, expand=True)
            self.inline_edit_widget.set(current_value)
            
            def on_combobox_save(event=None):
                # 用户选择了新值，清除验证失败标记
                self._inline_edit_validation_failed = False
                # 直接保存
                self.on_generic_inline_edit_save(event)
            
            def on_combobox_cancel(event=None):
                # 直接取消
                self.on_generic_inline_edit_cancel(event)
            
            # 绑定Combobox特定事件
            self.inline_edit_widget.bind('<<ComboboxSelected>>', on_combobox_save)
            
            # 绑定基本事件
            self.inline_edit_widget.bind('<Return>', on_combobox_save)
            self.inline_edit_widget.bind('<Escape>', on_combobox_cancel)
            
            # 确保Combobox始终保持焦点
            self.inline_edit_widget.focus_force()
            # 初始化验证失败标记
            self._inline_edit_validation_failed = False
            
            # 绑定根窗口点击事件，用于关闭编辑状态
            def on_root_click(event):
                # 主动验证当前输入，如果无效则阻止点击其他位置
                is_valid, error_msg = self._validate_current_inline_edit()
                if not is_valid:
                    self._inline_edit_validation_failed = True
                    self.show_error("", error_msg)
                    self.root.after(1, self._refocus_inline_edit)
                    return
                
                # 检查点击的是否是编辑控件或其内部组件
                target = event.widget
                is_edit_widget = False
                is_treeview_click = False
                clicked_item = None
                
                # 遍历控件树，检查是否是编辑控件或其内部组件
                while target:
                    if (hasattr(self, 'inline_edit_widget') and target == self.inline_edit_widget) or (hasattr(self, 'inline_edit_frame') and target == self.inline_edit_frame):
                        is_edit_widget = True
                        break
                    # 检查是否是日期选择器的日历窗口
                    if str(target).endswith('.toplevel') and 'dateentry' in str(target):
                        is_edit_widget = True
                        break
                    # 检查是否是Treeview控件
                    if str(target).endswith('.treeview'):
                        is_treeview_click = True
                        # 获取点击的行
                        try:
                            clicked_item = target.identify_row(event.y)
                        except Exception:
                            pass
                        break
                    target = target.master
                
                # 如果点击的不是编辑控件，关闭编辑状态
                if not is_edit_widget:
                    # 保存当前编辑
                    self.on_generic_inline_edit_save(None)
                    
                    # 如果点击的是Treeview控件且点击了有效的行
                    if is_treeview_click and clicked_item:
                        # 选中点击的行
                        target.selection_set(clicked_item)
            
            self.root.bind('<Button-1>', on_root_click, add='+')
        elif column_type == 'datepicker':
            # 检查DateEntry是否可用
            if DateEntry is not None:
                try:
                    # 日期选择器类型，使用tkcalendar的DateEntry控件
                    self.inline_edit_widget = DateEntry(
                        tree, 
                        width=width // 10 - 2, 
                        background='white', 
                        foreground='black', 
                        borderwidth=1,
                        date_pattern='yyyy-MM-dd',
                        showweeknumbers=False,
                        showothermonthdays=False
                    )
                    
                    # 设置初始日期
                    if current_value and current_value != 'None':
                        try:
                            # 支持多种日期格式解析
                            date_obj = None
                            date_str = current_value
                            
                            # 如果有时间部分，只取日期部分
                            if ' ' in date_str:
                                date_str = date_str.split(' ')[0]
                            
                            # 支持多种日期格式
                            date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d', '%m-%d-%Y', '%m/%d/%Y']
                            for fmt in date_formats:
                                try:
                                    date_obj = datetime.datetime.strptime(date_str, fmt)
                                    break
                                except ValueError:
                                    continue
                            
                            if date_obj:
                                # 设置DateEntry的日期
                                self.inline_edit_widget.set_date(date_obj)
                        except Exception:
                            # 如果解析失败，使用默认日期
                            default_date = (datetime.datetime.now() + datetime.timedelta(days=365)).date()
                            self.inline_edit_widget.set_date(default_date)
                    else:
                        # 如果没有当前值，使用当前日期 + 365天（与过期设置对话框保持一致）
                        default_date = (datetime.datetime.now() + datetime.timedelta(days=365)).date()
                        self.inline_edit_widget.set_date(default_date)
                    
                    # 放置控件
                    self.inline_edit_widget.place(x=x, y=y, width=width, height=height)
                    
                    # 绑定必要的事件
                    def on_date_save(event=None):
                        # 用户选择了新日期，清除验证失败标记
                        self._inline_edit_validation_failed = False
                        # 延迟执行保存，让tkcalendar内部的_select方法先完成
                        self.root.after(10, lambda: self.on_generic_inline_edit_save(None))
                    
                    def on_date_focus_out(event):
                        # 检查日历是否打开
                        calendar_open = False
                        try:
                            if hasattr(self.inline_edit_widget, '_top_cal'):
                                top_cal = self.inline_edit_widget._top_cal
                                if top_cal and top_cal.winfo_exists() and top_cal.winfo_ismapped():
                                    calendar_open = True
                        except Exception:
                            pass
                        
                        # 日历关闭且焦点离开时，保存编辑
                        if not calendar_open:
                            # 主动验证当前输入
                            is_valid, error_msg = self._validate_current_inline_edit()
                            if not is_valid:
                                self._inline_edit_validation_failed = True
                                self.show_error("", error_msg)
                                self.root.after(1, self._refocus_inline_edit)
                                return
                            self.on_generic_inline_edit_save(None)
                    
                    # 绑定事件
                    self.inline_edit_widget.bind('<Return>', on_date_save)
                    self.inline_edit_widget.bind('<Escape>', on_widget_cancel)
                    self.inline_edit_widget.bind('<<DateEntrySelected>>', on_date_save)
                    self.inline_edit_widget.bind('<FocusOut>', on_date_focus_out)
                    
                    # 修复DateEntry日历弹窗的焦点管理问题
                    # 内联编辑不在模态对话框中，传None只修复_on_focus_out_cal
                    fix_date_entry_for_modal(self.inline_edit_widget, None)
                    
                    # 确保日期选择器始终保持焦点
                    self.inline_edit_widget.focus_force()
                    # 初始化验证失败标记
                    self._inline_edit_validation_failed = False
                    
                except Exception as e:
                    # 如果日期选择器创建失败，使用普通的Entry控件替代
                    self.inline_edit_widget = ttk.Entry(tree, width=width // 10 - 2)
                    self.inline_edit_widget.insert(0, current_value)
                    self.inline_edit_widget.place(x=x, y=y, width=width, height=height)
                    
                    # 绑定事件
                    self.inline_edit_widget.bind('<Return>', on_widget_save)
                    self.inline_edit_widget.bind('<Escape>', on_widget_cancel)
                    self.inline_edit_widget.bind('<FocusOut>', on_widget_focus_out)
                    self.inline_edit_widget.bind('<Key>', on_widget_input)
                    
                    # 设置焦点
                    self.inline_edit_widget.focus_set()
                    self.inline_edit_widget.select_range(0, tk.END)
                    # 初始化验证失败标记
                    self._inline_edit_validation_failed = False
            else:
                # 如果DateEntry不可用，使用普通的Entry控件替代
                self.inline_edit_widget = ttk.Entry(tree, width=width // 10 - 2)
                self.inline_edit_widget.insert(0, current_value)
                self.inline_edit_widget.place(x=x, y=y, width=width, height=height)
                
                # 绑定事件
                self.inline_edit_widget.bind('<Return>', on_widget_save)
                self.inline_edit_widget.bind('<Escape>', on_widget_cancel)
                self.inline_edit_widget.bind('<FocusOut>', on_widget_focus_out)
                self.inline_edit_widget.bind('<Key>', on_widget_input)
                
                # 设置焦点
                self.inline_edit_widget.focus_set()
                self.inline_edit_widget.select_range(0, tk.END)
                # 初始化验证失败标记
                self._inline_edit_validation_failed = False
        else:
            # 默认使用Entry
            self.inline_edit_widget = ttk.Entry(tree, width=width // 10 - 2)
            self.inline_edit_widget.insert(0, current_value)
            self.inline_edit_widget.place(x=x, y=y, width=width, height=height)
            
            # 实时验证函数
            def validate_edit(text):
                # 获取处理器
                handlers = self._inline_edit_handlers[tree_name]
                validate = handlers.get('validate')
                if validate:
                    # 获取行数据用于验证
                    get_row_data = handlers.get('get_row_data')
                    if not get_row_data:
                        def default_get_row_data(item):
                            return {
                                'item': item,
                                'values': tree.item(item, 'values')
                            }
                        get_row_data = default_get_row_data
                    row_data = get_row_data(item)
                    row_data['tree_name'] = tree_name
                    
                    # 验证新值 - 支持二元组或三元组返回值
                    validate_result = validate(text, column_name, row_data)
                    if len(validate_result) == 3:
                        is_valid, _, _ = validate_result
                    else:
                        is_valid, _ = validate_result
                    # 设置文本颜色
                    self.inline_edit_widget.config(foreground='black' if is_valid else 'red')
                return "1"
            
            # 绑定实时验证
            self.inline_edit_widget.config(validate="all", validatecommand=(self.root.register(validate_edit), "%P"))
            
            # 绑定事件
            self.inline_edit_widget.bind('<Return>', on_widget_save)
            self.inline_edit_widget.bind('<Escape>', on_widget_cancel)
            self.inline_edit_widget.bind('<FocusOut>', on_widget_focus_out)
            self.inline_edit_widget.bind('<Key>', on_widget_input)
            
            # 绑定根窗口点击事件，用于关闭编辑状态
            def on_root_click(event):
                # 主动验证当前输入，如果无效则阻止点击其他位置
                is_valid, error_msg = self._validate_current_inline_edit()
                if not is_valid:
                    self._inline_edit_validation_failed = True
                    self.show_error("", error_msg)
                    self.root.after(1, self._refocus_inline_edit)
                    return
                
                # 检查点击的是否是编辑控件
                target = event.widget
                is_edit_widget = False
                is_treeview_click = False
                clicked_item = None
                
                # 遍历控件树，检查是否是编辑控件
                while target:
                    if hasattr(self, 'inline_edit_widget') and target == self.inline_edit_widget:
                        is_edit_widget = True
                        break
                    # 检查是否是Treeview控件
                    if str(target).endswith('.treeview'):
                        is_treeview_click = True
                        # 获取点击的行
                        try:
                            clicked_item = target.identify_row(event.y)
                        except Exception:
                            pass
                        break
                    target = target.master
                
                # 如果点击的不是编辑控件，关闭编辑状态
                if not is_edit_widget:
                    # 保存当前编辑
                    self.on_generic_inline_edit_save(None)
                    
                    # 如果点击的是Treeview控件且点击了有效的行
                    if is_treeview_click and clicked_item:
                        # 选中点击的行
                        target.selection_set(clicked_item)
            
            self.root.bind('<Button-1>', on_root_click, add='+')
            
            # 初始化验证失败标记
            self._inline_edit_validation_failed = False
        
        # 获取焦点
        self.inline_edit_widget.focus_set()
        
        # 如果是Entry，全选文本
        if column_type == 'entry':
            self.inline_edit_widget.select_range(0, tk.END)
        
        # 设置冷却期，防止三击被误判为两次双击
        self._in_double_click_cooldown = True
        if hasattr(self, '_double_click_cooldown_timer'):
            self.root.after_cancel(self._double_click_cooldown_timer)
        self._double_click_cooldown_timer = self.root.after(self.double_click_interval, 
            lambda: setattr(self, '_in_double_click_cooldown', False))
        
        # 阻止Treeview的默认双击行为（展开/收缩节点）
        return 'break'
    
    def on_generic_inline_edit_save(self, event):
        """通用的保存内联编辑内容方法
        """
        if not hasattr(self, 'inline_edit_widget') or not hasattr(self, 'inline_edit_data'):
            return
        
        # 检查DateEntry控件是否还存在
        if DateEntry is not None and isinstance(self.inline_edit_widget, DateEntry):
            try:
                # 尝试获取控件值，如果控件已被销毁会抛出异常
                exists = self.inline_edit_widget.winfo_exists()
            except tk.TclError:
                # 控件已被销毁，清理状态并返回
                self._destroy_inline_edit_widgets()
                return
        
        # 清理定时器，避免重复执行
        if hasattr(self, '_inline_edit_save_timer') and self._inline_edit_save_timer:
            self.root.after_cancel(self._inline_edit_save_timer)
            delattr(self, '_inline_edit_save_timer')
        
        # 当下拉列表展开时，不保存编辑内容
        if hasattr(self, '_combobox_is_posting') and self._combobox_is_posting:
            return
        
        # 检查事件类型，判断是否是从焦点丢失调用
        from_focus_out = event is None or (hasattr(event, 'type') and str(event.type) == 'FocusOut')
        
        # 获取新值
        if DateEntry is not None and isinstance(self.inline_edit_widget, DateEntry):
            # 对于DateEntry控件，直接获取格式化的日期字符串
            new_value = self.inline_edit_widget.get().strip()
        else:
            # 普通控件
            new_value = self.inline_edit_widget.get().strip()
        
        # 获取编辑数据
        edit_data = self.inline_edit_data
        item = edit_data['item']
        column_index = edit_data['column_index']
        column_name = edit_data['column_name']
        original_value = edit_data['original_value']
        tree = edit_data['tree']
        tree_name = edit_data['tree_name']
        
        # 检查是否有注册的处理器
        if tree_name not in self._inline_edit_handlers:
            self._destroy_inline_edit_widgets()
            return
        
        handlers = self._inline_edit_handlers[tree_name]
        
        # 如果值没有变化，直接清理
        if new_value == original_value:
            self._destroy_inline_edit_widgets()
            return
        
        # 获取行数据
        get_row_data = handlers.get('get_row_data')
        if not get_row_data:
            # 默认获取行数据的方法
            def default_get_row_data(item):
                return {
                    'item': item,
                    'values': tree.item(item, 'values')
                }
            get_row_data = default_get_row_data
        
        row_data = get_row_data(item)
        
        # 验证数据 - 空值检查（在try块外进行，验证失败时不销毁编辑控件）
        if not new_value:
            is_valid, error_msg = IPAMValidator.validate_inline_edit(column_name, new_value)
            if not is_valid:
                # 在信息栏显示错误提示，保持编辑状态，阻止焦点离开
                self.show_error("", _("input_cannot_be_empty"))
                # 标记验证失败，用于后续焦点处理
                self._inline_edit_validation_failed = True
                # 延迟重新获取焦点，确保在所有事件处理完成后执行
                self.root.after(1, self._refocus_inline_edit)
                return
        
        # 验证数据 - 自定义验证（在try块外进行，验证失败时不销毁编辑控件）
        validate = handlers.get('validate')
        if validate:
            validate_result = validate(new_value, column_name, row_data)
            # 支持返回二元组或三元组
            if len(validate_result) == 3:
                is_valid, validation_error, formatted_value = validate_result
            else:
                is_valid, validation_error = validate_result
                formatted_value = None
            
            if not is_valid:
                # 在信息栏显示错误提示，保持编辑状态，阻止焦点离开
                self.show_error("", validation_error)
                # 标记验证失败，用于后续焦点处理
                self._inline_edit_validation_failed = True
                # 延迟重新获取焦点，确保在所有事件处理完成后执行
                self.root.after(1, self._refocus_inline_edit)
                return
            
            # 使用格式化后的值（如果有）
            if formatted_value is not None:
                new_value = formatted_value
        
        # 验证通过，清除失败标记
        self._inline_edit_validation_failed = False
        
        # 验证通过，执行保存操作
        try:
            # 保存数据
            save = handlers.get('save')
            if save:
                success, message = save(new_value, column_name, row_data, item)
                if not success:
                    self.show_error("", message)
                    # 标记验证失败
                    self._inline_edit_validation_failed = True
                    # 延迟重新获取焦点
                    self.root.after(1, self._refocus_inline_edit)
                    return
                
                # 更新Treeview（如果保存方法没有自行刷新）
                try:
                    values = list(tree.item(item, 'values'))
                    values[column_index] = new_value
                    tree.item(item, values=values)
                    
                    # 更新斑马条纹
                    self.update_table_zebra_stripes(tree)
                except tk.TclError:
                    # item可能已被保存方法刷新，跳过Treeview更新
                    pass
        except Exception as e:
            self.show_error("", f"更新失败: {str(e)}")
            # 标记验证失败
            self._inline_edit_validation_failed = True
            # 延迟重新获取焦点
            self.root.after(1, self._refocus_inline_edit)
            return
        
        # 保存成功，清理编辑控件
        item_to_select = edit_data['item']
        tree_to_select = edit_data['tree']
        self._destroy_inline_edit_widgets()
        # 确保编辑的行仍然保持选中状态（可能因Treeview刷新而失效）
        try:
            tree_to_select.selection_set(item_to_select)
        except tk.TclError:
            pass
    
    def _refocus_inline_edit(self):
        """验证失败后延迟重新获取焦点，确保在所有事件处理完成后执行"""
        if hasattr(self, '_inline_edit_validation_failed') and self._inline_edit_validation_failed:
            if hasattr(self, 'inline_edit_widget') and self.inline_edit_widget.winfo_exists():
                self.inline_edit_widget.focus_force()
    
    def _validate_current_inline_edit(self):
        """主动验证当前内联编辑的输入值
        
        在用户尝试点击其他控件时提前验证，不依赖 _inline_edit_validation_failed 标记。
        该标记只在 on_generic_inline_edit_save 中设置，但用户可能还没尝试保存
        就直接点击了其他行，此时标记仍为 False。
        
        Returns:
            tuple: (is_valid, error_message) - 验证是否通过及错误信息
        """
        if not hasattr(self, 'inline_edit_data') or not self.inline_edit_data:
            return True, None
        if not hasattr(self, 'inline_edit_widget') or not self.inline_edit_widget.winfo_exists():
            return True, None
        
        try:
            current_value = self.inline_edit_widget.get()
        except Exception:
            return True, None
        
        column_name = self.inline_edit_data.get('column_name', '')
        tree_name = self.inline_edit_data.get('tree_name', '')
        tree = self.inline_edit_data.get('tree')
        item = self.inline_edit_data.get('item')
        
        # 空值检查
        is_valid, error_msg = IPAMValidator.validate_inline_edit(column_name, current_value)
        if not is_valid:
            return False, _("input_cannot_be_empty")
        
        # 自定义验证
        if tree_name in self._inline_edit_handlers:
            handlers = self._inline_edit_handlers[tree_name]
            validate = handlers.get('validate')
            if validate:
                get_row_data = handlers.get('get_row_data')
                if get_row_data and item:
                    row_data = get_row_data(item)
                elif tree and item:
                    row_data = {'item': item, 'values': tree.item(item, 'values')}
                else:
                    row_data = {}
                row_data['tree_name'] = tree_name
                
                validate_result = validate(current_value, column_name, row_data)
                if len(validate_result) >= 2:
                    is_valid = validate_result[0]
                    error_msg = validate_result[1] if not is_valid else None
                    return is_valid, error_msg
        
        return True, None

    def _handle_inline_edit_validation_before_click(self):
        """处理点击前的内联编辑验证逻辑
        
        在点击其他行或控件前，验证当前编辑内容并尝试保存。
        如果验证失败或保存失败，返回 'break' 阻止后续处理。
        
        Returns:
            str or None: 'break' 表示需要阻止后续处理，None 表示允许继续
        """
        if hasattr(self, 'inline_edit_data') and self.inline_edit_data:
            is_valid, error_msg = self._validate_current_inline_edit()
            if not is_valid:
                self._inline_edit_validation_failed = True
                self.show_error("", error_msg)
                self.root.after(1, self._refocus_inline_edit)
                return 'break'
            self.on_generic_inline_edit_save(None)
            # 检查编辑控件是否还存在（控件已被销毁时调用 winfo_exists() 会抛出异常）
            try:
                if hasattr(self, 'inline_edit_widget') and self.inline_edit_widget.winfo_exists():
                    self.root.after(1, self._refocus_inline_edit)
                    return 'break'
            except tk.TclError:
                # 控件已被销毁，说明保存成功
                pass
        return None

    def on_generic_inline_edit_cancel(self, event):
        """通用的取消内联编辑方法
        """
        # 清除验证失败标记
        self._inline_edit_validation_failed = False
        
        # 安全清理编辑控件
        if hasattr(self, 'inline_edit_data'):
            item_to_select = self.inline_edit_data['item']
            tree_to_select = self.inline_edit_data['tree']
        else:
            item_to_select = None
            tree_to_select = None
        
        self._destroy_inline_edit_widgets()
        
        # 确保编辑的行仍然保持选中状态
        if item_to_select and tree_to_select:
            try:
                tree_to_select.selection_set(item_to_select)
            except tk.TclError:
                pass
    
    def on_ip_tree_double_click_inline(self, event):
        """IP地址表格双击内联编辑处理"""
        # 此方法已被通用的on_generic_tree_double_click替代
        pass
    
    def _get_ip_row_data(self, item):
        """获取IP地址表格行数据
        
        Args:
            item: Treeview行ID
            
        Returns:
            dict: 行数据字典
        """
        try:
            values = self.ipam_ip_tree.item(item, 'values')
            return {
                'item': item,
                'values': values,
                'ip_address': values[0],
                'status': values[1],
                'hostname': values[2],
                'mac_address': values[3],
                'description': values[4],
                'allocated_at': values[5],
                'expiry_date': values[6]
            }
        except (tk.TclError, IndexError):
            # 树项不存在或数据格式错误
            return {
                'item': item,
                'values': [],
                'ip_address': '',
                'status': '',
                'hostname': '',
                'description': '',
                'allocated_at': '',
                'expiry_date': ''
            }
    
    def is_valid_ipv4(self, ip_address):
        """验证IPv4地址格式
        
        Args:
            ip_address: 要验证的IPv4地址
            
        Returns:
            bool: 是否为有效的IPv4地址
        """
        return self.validation_service.is_valid_ipv4(ip_address)
    
    def is_ip_in_network(self, ip_address, network):
        """检查IP地址是否在指定的网络范围内
        
        Args:
            ip_address: 要检查的IP地址
            network: 网络地址（CIDR格式）
            
        Returns:
            bool: IP地址是否在网络范围内
        """
        return self.validation_service.is_ip_in_network(ip_address, network)
    
    def _format_expiry_date(self, date_str: str) -> str:
        """格式化过期日期，确保包含时间部分
        
        Args:
            date_str: 日期字符串（YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS）
            
        Returns:
            str: 格式化后的日期字符串（YYYY-MM-DD HH:MM:SS）
        """
        if not date_str:
            return None
        
        date_str = date_str.strip()
        
        # 如果只有日期部分（长度为10），添加时间部分为23:59:59
        if len(date_str) == 10:
            return f"{date_str} 23:59:59"
        
        return date_str
    
    def _format_mac_address(self, mac):
        """格式化MAC地址为统一格式 XXXX-XXXX-XXXX，字母转为大写
        
        Args:
            mac: 原始MAC地址
            
        Returns:
            str: 格式化后的MAC地址
        """
        if not mac:
            return ''
        
        # 移除非十六进制字符（保留字母和数字）
        import re
        clean_mac = re.sub(r'[^0-9A-Fa-f]', '', mac).upper()
        
        # 根据长度格式化
        if len(clean_mac) == 12:
            # 格式化为 XXXX-XXXX-XXXX
            return f"{clean_mac[:4]}-{clean_mac[4:8]}-{clean_mac[8:]}"
        elif len(clean_mac) == 6:
            # 如果只有6个字符，可能是短格式，补齐前导零
            clean_mac = clean_mac.zfill(12)
            return f"{clean_mac[:4]}-{clean_mac[4:8]}-{clean_mac[8:]}"
        else:
            # 返回原始值（应该已经被验证过滤）
            return mac.upper()
    
    def _validate_ip_edit(self, new_value, column_name, row_data):
        """验证IP地址表的编辑值
        
        Args:
            new_value: 新编辑的值
            column_name: 列名
            row_data: 行数据字典
            
        Returns:
            tuple: (是否有效, 错误消息或None, 格式化后的值或None)
        """
        # 空值检查
        if not new_value:
            is_valid, error_msg = IPAMValidator.validate_inline_edit(column_name, new_value)
            if is_valid:
                return True, None, None
            else:
                return False, _('input_cannot_be_empty'), None
        
        # 获取选中的网络
        network_items = self.ipam_network_tree.selection()
        if not network_items:
            return False, _('please_select_network_for_allocation'), None
        network = self.ipam_network_tree.item(network_items[0], 'values')[0]
        
        if column_name == 'ip_address':
            # 验证IP地址格式
            if not self.is_valid_ipv4(new_value):
                return False, _('invalid_ip_address'), None
            
            # 验证IP地址是否在选定的网络范围内
            if not self.is_ip_in_network(new_value, network):
                return False, _('ip_not_in_network'), None
        elif column_name == 'status':
            # 验证状态合法性
            valid_statuses = [_('released'), _('allocated'), _('reserved')]
            if new_value not in valid_statuses:
                return False, _('invalid_status'), None
        elif column_name == 'hostname':
            # 主机名长度限制
            if len(new_value) > 255:
                return False, _('hostname_too_long'), None
        elif column_name == 'mac_address':
            # 验证MAC地址格式
            import re
            # 支持的MAC地址格式：
            # XX:XX:XX:XX:XX:XX 或 XX-XX-XX-XX-XX-XX 或 XXXXXXXXXXXX 或 XXXX-XXXX-XXXX
            mac_pattern = re.compile(r'^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$|^[0-9A-Fa-f]{12}$|^([0-9A-Fa-f]{4}-){2}[0-9A-Fa-f]{4}$')
            if not mac_pattern.match(new_value):
                return False, _('invalid_mac_address'), None
            # 格式化MAC地址
            formatted_mac = self._format_mac_address(new_value)
            return True, None, formatted_mac
        elif column_name == 'description':
            # 描述长度限制
            if len(new_value) > 1000:
                return False, _('description_too_long'), None
        elif column_name == 'expiry_date':
            # 验证日期格式（支持YYYY-MM-DD和YYYY-MM-DD HH:MM:SS）
            try:
                if len(new_value) == 10:
                    # 格式：YYYY-MM-DD
                    datetime.datetime.strptime(new_value, '%Y-%m-%d')
                elif len(new_value) == 19:
                    # 格式：YYYY-MM-DD HH:MM:SS
                    datetime.datetime.strptime(new_value, '%Y-%m-%d %H:%M:%S')
                else:
                    return False, _('invalid_date_format'), None
            except ValueError:
                return False, _('invalid_date_format'), None
        
        return True, None, None
    
    def _save_ip_edit(self, new_value, column_name, row_data, item):
        """保存IP地址表的编辑值
        
        Args:
            new_value: 新编辑的值
            column_name: 列名
            row_data: 行数据字典
            item: Treeview行ID
            
        Returns:
            tuple: (是否成功, 消息)
        """
        # 获取选中的网络
        network_items = self.ipam_network_tree.selection()
        if not network_items:
            return False, _('please_select_network')
        network = self.ipam_network_tree.item(network_items[0], 'values')[0]
        
        try:
            # 根据列名执行不同的更新操作
            if column_name == 'status':
                # 状态转换映射
                status_map = {
                    _('released'): 'released',
                    _('allocated'): 'allocated',
                    _('reserved'): 'reserved'
                }
                actual_status = status_map.get(new_value, new_value)
                
                # 根据状态执行不同的操作
                if actual_status == 'released':
                    # 释放IP
                    success, message = self.ipam.release_ip(row_data['ip_address'])
                elif actual_status == 'allocated':
                    # 分配IP
                    success, message = self.ipam.allocate_ip(
                        network, 
                        row_data['ip_address'], 
                        row_data['hostname'], 
                        row_data['description'],
                        row_data.get('expiry_date')
                    )
                elif actual_status == 'reserved':
                    # 保留IP
                    record_id = self._get_db_record_id(item)
                    success, message = self.ipam.reserve_ip(
                        network, 
                        row_data['ip_address'], 
                        row_data['hostname'],
                        row_data['description'],
                        row_data.get('expiry_date'),
                        record_id
                    )
                else:
                    success = False
                    message = f"未知状态: {actual_status}"
            elif column_name == 'ip_address':
                # 更新IP地址
                # 1. 先释放旧IP
                self.ipam.release_ip(row_data['ip_address'])
                # 2. 再分配新IP
                record_id = self._get_db_record_id(item)
                success, message = self.ipam.allocate_ip(
                    network, 
                    new_value, 
                    row_data['hostname'], 
                    row_data.get('description', ''),
                    row_data.get('expiry_date'),
                    record_id
                )
            elif column_name == 'hostname' or column_name == 'description' or column_name == 'mac_address':
                # 更新主机名、描述或MAC地址
                record_id_int = self._get_db_record_id(item)
                # 从row_data中获取当前值，确保使用最新数据
                hostname = new_value if column_name == 'hostname' else row_data.get('hostname', '')
                mac_address = new_value if column_name == 'mac_address' else row_data.get('mac_address', '')
                description = new_value if column_name == 'description' else row_data.get('description', '')
                # 获取过期日期
                expiry_date = row_data.get('expiry_date', None)
                if record_id_int:
                    # 使用update_ip_record方法更新特定记录
                    success, message = self.ipam.update_ip_record(record_id_int, hostname, mac_address, description, expiry_date)
                else:
                    # 如果没有数据库记录ID，使用默认方法
                    success, message = self.ipam.update_ip_info(
                        row_data['ip_address'], 
                        hostname=hostname,
                        description=description,
                        mac_address=mac_address
                    )
            elif column_name == 'expiry_date':
                # 更新过期日期
                formatted_date = self._format_expiry_date(new_value) if new_value else None
                record_id_int = self._get_db_record_id(item)
                if record_id_int:
                    # 使用记录ID更新过期日期
                    success, message = self.ipam.update_ip_expiry(row_data['ip_address'], formatted_date, record_id_int)
                else:
                    # 如果没有数据库记录ID，使用默认方法
                    success, message = self.ipam.update_ip_expiry(row_data['ip_address'], formatted_date)
            else:
                success = False
                message = f"不支持编辑的列: {column_name}"
            
            return success, message
        except Exception as e:
            return False, f"更新失败: {str(e)}"
    
    def _get_requirements_row_data(self, item, tree_name):
        """获取子网需求表或需求池表的行数据
        
        Args:
            item: Treeview行ID
            tree_name: 表格名称标识
            
        Returns:
            dict: 行数据字典
        """
        tree = self.requirements_tree if tree_name == 'requirements' else self.pool_tree
        values = tree.item(item, 'values')
        return {
            'item': item,
            'values': values,
            'index': values[0],
            'name': values[1],
            'hosts': values[2],
            'tree_name': tree_name
        }
    
    def _validate_requirements_edit(self, new_value, column_name, row_data):
        """验证子网需求表或需求池表的编辑值
        
        Args:
            new_value: 新编辑的值
            column_name: 列名
            row_data: 行数据字典
            
        Returns:
            tuple: (是否有效, 错误消息或None)
        """
        if not new_value:
            return False, _('input_cannot_be_empty')
        
        if column_name == 'hosts':
            # 验证主机数量是否为正整数
            try:
                hosts = int(new_value)
                if hosts <= 0:
                    return False, _('host_count_must_be_greater_than_0')
            except ValueError:
                return False, _('invalid_number')
        elif column_name == 'name':
            # 验证子网名称是否已存在
            tree_name = row_data['tree_name']
            for check_tree_name in ['requirements', 'pool']:
                tree = self.requirements_tree if check_tree_name == 'requirements' else self.pool_tree
                for check_item in tree.get_children():
                    # 排除当前编辑的行
                    if check_tree_name == tree_name and check_item == row_data['item']:
                        continue
                    check_values = tree.item(check_item, 'values')
                    if check_values[1] == new_value:
                        return False, _('subnet_already_exists', name=new_value)
        
        return True, None
    
    def _save_requirements_edit(self, new_value, column_name, row_data, item, tree_name):
        """保存子网需求表或需求池表的编辑值
        
        Args:
            new_value: 新编辑的值
            column_name: 列名
            row_data: 行数据字典
            item: Treeview行ID
            tree_name: 表格名称标识
            
        Returns:
            tuple: (是否成功, 消息)
        """
        try:
            tree = self.requirements_tree if tree_name == 'requirements' else self.pool_tree
            
            # 更新Treeview
            values = list(tree.item(item, 'values'))
            column_index = 1 if column_name == 'name' else 2
            values[column_index] = new_value
            tree.item(item, values=values)
            
            # 更新斑马条纹
            self.update_table_zebra_stripes(tree)
            
            # 保存当前状态
            self.save_current_state(f"编辑{tree_name}表格记录")
            
            return True, "更新成功"
        except Exception as e:
            return False, f"更新失败: {str(e)}"
    
    def _get_network_row_data(self, item):
        """获取网络管理表格行数据
        
        Args:
            item: Treeview行ID
            
        Returns:
            dict: 行数据字典
        """
        values = self.ipam_network_tree.item(item, 'values')
        # 确保 values 数组长度足够，处理旧数据的情况
        while len(values) < 5:
            values.append('')
        return {
            'item': item,
            'values': values,
            'network': values[0],
            'description': values[1],
            'vlan': values[2],
            'created_at': values[3],
            'ip_count': values[4]
        }
    
    def _validate_network_edit(self, new_value, column_name, row_data):
        """验证网络管理表的编辑值
        
        Args:
            new_value: 新编辑的值
            column_name: 列名
            row_data: 行数据字典
            
        Returns:
            tuple: (是否有效, 错误消息或None)
        """
        if column_name == 'network':
            # 验证网段地址格式
            if not new_value:
                return False, _('input_cannot_be_empty')
            try:
                ipaddress.ip_network(new_value, strict=False)
            except ValueError:
                return False, "无效的网段地址格式，请使用CIDR格式（如192.168.1.0/24）"
        elif column_name == 'description':
            # 描述长度限制
            if len(new_value) > 255:
                return False, _('description_too_long')
        elif column_name == 'vlan':
            # VLAN 验证
            if new_value:
                if not new_value.isdigit():
                    return False, _('vlan_invalid_format')
                vlan_num = int(new_value)
                if vlan_num < 1 or vlan_num > 4094:
                    return False, _('vlan_out_of_range')
        
        return True, None
    
    def _save_network_edit(self, new_value, column_name, row_data, item):
        """保存网络管理表的编辑值
        
        Args:
            new_value: 新编辑的值
            column_name: 列名
            row_data: 行数据字典
            item: Treeview行ID
            
        Returns:
            tuple: (是否成功, 消息)
        """
        try:
            if column_name == 'network':
                # 更新网段地址
                old_network = row_data['network']
                success, message = self.ipam.update_network(old_network, new_value)
                if success:
                    # 刷新IPAM数据并恢复选中状态
                    self.refresh_ipam_with_selection()
                return success, message
            elif column_name == 'description':
                # 更新网络描述
                success, message = self.ipam.update_network_description(row_data['network'], new_value)
                return success, message
            elif column_name == 'vlan':
                # 更新网络VLAN
                success, message = self.ipam.update_network_vlan(row_data['network'], new_value)
                return success, message
            else:
                return False, f"不支持编辑的列: {column_name}"
        except Exception as e:
            return False, f"更新失败: {str(e)}"
    
    def _validate_status(self, status_value):
        """验证状态值的合法性
        
        Args:
            status_value: 要验证的状态值
            
        Returns:
            tuple: (is_valid, actual_status) - 是否有效和转换后的实际状态
        """
        # 使用类常量获取所有有效状态
        valid_statuses = self._get_all_valid_statuses()
        if status_value not in valid_statuses:
            return False, None
        
        # 使用类方法获取状态转换映射
        status_map = self._get_status_map()
        actual_status = status_map.get(status_value, status_value)
        
        # 第一次检查已确保状态有效，无需再次验证
        return True, actual_status
    
    def on_inline_edit_save(self, event):
        """保存内联编辑的内容"""
        if hasattr(self, 'inline_edit_widget') and hasattr(self, 'inline_edit_data'):
            # 获取新值
            new_value = self.inline_edit_widget.get().strip()
            
            # 获取编辑数据
            edit_data = self.inline_edit_data
            item = edit_data['item']
            column_index = edit_data['column_index']
            column_name = edit_data['column_name']
            ip_address = edit_data['ip_address']
            
            # 获取选中的网络
            network_items = self.ipam_network_tree.selection()
            if not network_items:
                self.show_error(_('error'), _('please_select_network'))
                self.on_inline_edit_cancel(None)
                return
            network = self.ipam_network_tree.item(network_items[0], 'values')[0]
            
            # 输入验证
            validation_error = None
            actual_status = None
            if column_name == 'status':
                # 验证状态合法性
                is_valid, actual_status = self._validate_status(new_value)
                if not is_valid:
                    validation_error = "无效的状态值"
            elif column_name in ['hostname', 'description']:
                # 先过滤特殊字符（防止注入），确保验证使用正确的值
                if new_value:
                    # 移除控制字符（0-31和127），但保留所有可打印字符（包括中文、俄文等Unicode字符）
                    new_value = ''.join(char for char in new_value if not (0 <= ord(char) < 32 or ord(char) == 127))
                    
                    if column_name == 'hostname':
                        # 主机名：允许Unicode字符（包括中文、俄文等），同时只排除危险字符
                        # 允许的字符：Unicode单词字符、连字符、点、下划线
                        new_value = re.sub(r'[^\w\-\.]', '', new_value) if new_value else ''
                    else:
                        # 描述：只过滤真正危险的字符（SQL注入和XSS），保留常见标点符号
                        # 保留的字符：中文、俄文等Unicode字符，以及常见标点符号（逗号、句号、感叹号等）
                        new_value = re.sub(r'[\'\"<>;\\]', '', new_value) if new_value else ''
                
                # 验证长度（使用过滤后的值）
                if column_name == 'hostname' and len(new_value) > 255:
                    validation_error = "主机名长度不能超过255个字符"
                elif column_name == 'description' and len(new_value) > 1000:
                    validation_error = "描述长度不能超过1000个字符"
            
            if validation_error:
                self.show_error(_('error'), validation_error)
                return
            
            try:
                # 更新数据库
                if column_name == 'status':
                    # 获取当前值并转换（original_value是已知合法值，无需完整验证）
                    current_value = edit_data['original_value']
                    status_map = self._get_status_map()
                    current_actual_status = status_map.get(current_value, current_value)
                    
                    # 只有当状态真正改变时才执行操作
                    if actual_status != current_actual_status:
                        if actual_status == 'released':
                            # 释放IP
                            success, message = self.ipam.release_ip(ip_address)
                        elif actual_status == 'allocated':
                            # 分配IP
                            hostname = self.ipam_ip_tree.item(item, 'values')[2]
                            description = self.ipam_ip_tree.item(item, 'values')[3]
                            success, message = self.ipam.allocate_ip(network, ip_address, hostname, description)
                        elif actual_status == 'reserved':
                            # 保留IP
                            description = self.ipam_ip_tree.item(item, 'values')[3]
                            success, message = self.ipam.reserve_ip(network, ip_address, '', description)
                        else:
                            success = False
                            message = f"未知状态: {actual_status}"
                    else:
                        # 状态未改变，直接成功
                        success = True
                        message = "状态未改变"
                else:
                    # 更新主机名或描述
                    success, message = self.ipam.update_ip_info(ip_address, 
                                                              hostname=new_value if column_name == 'hostname' else None,
                                                              description=new_value if column_name == 'description' else None)
                
                if success:
                    # 更新Treeview
                    values = list(self.ipam_ip_tree.item(item, 'values'))
                    values[column_index] = new_value
                    self.ipam_ip_tree.item(item, values=values)
                else:
                    self.show_error(_('error'), message)
            except Exception as e:
                self.show_error(_('error'), f"更新失败: {str(e)}")
            finally:
                # 安全清理编辑控件
                self._destroy_inline_edit_widgets()
    
    def on_inline_edit_cancel(self, event):
        """取消内联编辑"""
        # 安全清理编辑控件
        self._destroy_inline_edit_widgets()
    
    def _get_db_record_id(self, tree_item):
        """从树项的tags中获取数据库记录ID
        
        Args:
            tree_item: 树项ID
        
        Returns:
            int or None: 数据库记录ID，如果获取失败返回None
        """
        tags = self.ipam_ip_tree.item(tree_item, 'tags')
        for tag in tags:
            if tag.startswith('dbid_'):
                try:
                    return int(tag[5:])
                except ValueError:
                    pass
        return None
    
    def on_ip_menu_action(self, action):
        """处理IP地址表格右键菜单的不同操作，支持多选"""
        selected_items = self.ipam_ip_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 获取选中的网络
        network_items = self.ipam_network_tree.selection()
        if not network_items:
            self.show_error(_('error'), _('please_select_network'))
            return
        network = self.ipam_network_tree.item(network_items[0], 'values')[0]
        
        if action == 'restore':
            # 恢复已释放的IP地址 - 支持多选
            success_count = 0
            error_count = 0
            
            error_details = []
            for ip_item in selected_items:
                ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(ip_item)
                
                try:
                    if db_record_id:
                        # 根据记录ID获取IP的历史信息
                        ip_info = self.ipam.get_ip_record_by_id(db_record_id)
                        if ip_info:
                            hostname = ip_info.get('hostname', '') or ''
                            description = ip_info.get('description', '') or ''
                            expiry_date = ip_info.get('expiry_date')
                            success, message = self.ipam.allocate_ip(network, ip_address, hostname, description, expiry_date, db_record_id)
                            if success:
                                success_count += 1
                            else:
                                error_count += 1
                                error_details.append(f"{ip_address}: {message}")
                        else:
                            error_count += 1
                            error_details.append(f"{ip_address}: 找不到记录信息")
                    else:
                        # 如果没有数据库记录ID，使用默认方法
                        ip_info = self.ipam.get_ip_info(ip_address)
                        if ip_info:
                            hostname = ip_info.get('hostname', '')
                            description = ip_info.get('description', '')
                            success, message = self.ipam.allocate_ip(network, ip_address, hostname, description)
                            if success:
                                success_count += 1
                            else:
                                error_count += 1
                                error_details.append(f"{ip_address}: {message}")
                        else:
                            error_count += 1
                            error_details.append(f"{ip_address}: 找不到IP信息")
                except Exception as e:
                    error_count += 1
                    error_details.append(f"{ip_address}: {str(e)}")
            
            # 显示结果
            if success_count > 0:
                self.show_info(_('success'), f"成功恢复 {success_count} 个IP地址")
                # 刷新IPAM数据并恢复选中状态
                self.refresh_ipam_with_selection()
            if error_count > 0:
                error_msg = f"{_('failed_to_restore_ips', count=error_count)}"
                if error_details:
                    error_msg += "\n" + "\n".join(error_details[:5])
                    if len(error_details) > 5:
                        error_msg += f"\n...还有 {len(error_details) - 5} 个错误"
                self.show_error(_('error'), error_msg)
                
        elif action == 'allocate':
            # 分配IP地址 - 只处理第一个选中的IP地址
            ip_item = selected_items[0]
            ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
            # 从tags中获取数据库记录ID
            db_record_id = self._get_db_record_id(ip_item)
            
            dialog_result = self.show_ip_address_dialog(_('allocate_address'), 'allocate', ip_address, record_id=db_record_id)
            if dialog_result:
                hostname = dialog_result['hostname']
                description = dialog_result['description']
                expiry_date = dialog_result['expiry_date']
                if hostname or description:
                    # 使用数据库记录ID来分配特定记录
                    success, message = self.ipam.allocate_ip(network, ip_address, hostname, description, expiry_date, db_record_id)
                    
                    if success:
                        self.show_info(_('success'), message)
                        self.refresh_ipam_with_selection()
                    else:
                        self.show_error(_('error'), message)
                        
        elif action == 'reserve':
            # 保留IP地址 - 只处理第一个选中的IP地址
            ip_item = selected_items[0]
            ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
            # 从tags中获取数据库记录ID
            db_record_id = self._get_db_record_id(ip_item)
            
            # 从树项中获取原始的主机名和描述
            original_hostname = self.ipam_ip_tree.item(ip_item, 'values')[2]
            original_description = self.ipam_ip_tree.item(ip_item, 'values')[4]
            
            # 显示对话框，传递原始的主机名和描述
            dialog_result = self.show_ip_address_dialog(_('reserve_address'), 'reserve', ip_address, record_id=db_record_id, original_hostname=original_hostname, original_description=original_description)
            if dialog_result:
                # 获取对话框中的值，如果为空则使用原始值
                hostname = dialog_result['hostname'] or original_hostname
                description = dialog_result['description'] or original_description
                expiry_date = dialog_result['expiry_date']
                
                # 使用数据库记录ID来保留特定记录
                success, message = self.ipam.reserve_ip(network, ip_address, hostname, description, expiry_date, db_record_id)
                
                if success:
                    self.show_info(_('success'), message)
                    self.refresh_ipam_with_selection()
                else:
                    self.show_error(_('error'), message)
                    
        elif action == 'release':
            # 释放IP地址 - 支持多选
            success_count = 0
            error_count = 0
            
            for ip_item in selected_items:
                ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(ip_item)
                
                if db_record_id:
                    # 使用specific策略释放特定记录
                    success, message = self.ipam.release_ip(ip_address, release_strategy="specific", record_id=db_record_id)
                else:
                    # 如果没有数据库记录ID，使用默认策略
                    success, message = self.ipam.release_ip(ip_address)
                
                if success:
                    success_count += 1
                else:
                    error_count += 1
            
            # 显示结果
            if success_count > 0 and error_count > 0:
                # 既有成功又有失败
                self.show_info(_('info'), f"成功释放 {success_count} 个IP地址，释放失败 {error_count} 个IP地址")
                # 刷新IPAM数据并恢复选中状态
                self.refresh_ipam_with_selection()
            elif success_count > 0:
                # 全部成功
                self.show_info(_('success'), f"{_('successfully_released_ips', count=success_count)}")
                # 刷新IPAM数据并恢复选中状态
                self.refresh_ipam_with_selection()
            elif error_count > 0:
                # 全部失败
                self.show_error(_('error'), f"释放失败 {error_count} 个IP地址")
        elif action == 'quick_allocate':
            # 快速分配IP地址 - 直接修改状态为已分配，不显示对话框，支持多选
            success_count = 0
            error_count = 0
            error_details = []
            
            for ip_item in selected_items:
                ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(ip_item)
                
                try:
                    if db_record_id:
                        # 根据记录ID获取当前记录信息
                        ip_info = self.ipam.get_ip_record_by_id(db_record_id)
                        if ip_info:
                            hostname = ip_info.get('hostname', 'Unnamed') or 'Unnamed'
                            description = ip_info.get('description', '快速分配') or '快速分配'
                            expiry_date = ip_info.get('expiry_date')
                        else:
                            hostname = 'Unnamed'
                            description = '快速分配'
                            expiry_date = None
                        # 快速分配，使用数据库记录ID
                        success, message = self.ipam.allocate_ip(network, ip_address, hostname, description, expiry_date, db_record_id)
                    else:
                        # 如果没有数据库记录ID，使用默认方法
                        ip_info = self.ipam.get_ip_info(ip_address)
                        if ip_info:
                            hostname = ip_info.get('hostname', 'Unnamed') or 'Unnamed'
                            description = ip_info.get('description', '快速分配') or '快速分配'
                            expiry_date = ip_info.get('expiry_date')
                        else:
                            hostname = 'Unnamed'
                            description = '快速分配'
                            expiry_date = None
                        success, message = self.ipam.allocate_ip(network, ip_address, hostname, description, expiry_date)
                    
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                        error_details.append(f"{ip_address}: {message}")
                except Exception as e:
                    print(f"快速分配失败: {e}")
                    error_count += 1
                    error_details.append(f"{ip_address}: {str(e)}")
            
            # 显示结果
            if success_count > 0:
                self.show_info(_('success'), f"{_('successfully_restored_ips_allocated', count=success_count)}")
                # 刷新IPAM数据并恢复选中状态
                self.refresh_ipam_with_selection()
            if error_count > 0:
                error_msg = f"{_('failed_to_restore_ips', count=error_count)}"
                if error_details:
                    error_msg += "\n" + "\n".join(error_details[:5])
                    if len(error_details) > 5:
                        error_msg += f"\n...还有 {len(error_details) - 5} 个错误"
                self.show_error(_('error'), error_msg)
        elif action == 'quick_reserve':
            # 快速保留IP地址 - 直接修改状态为保留，不显示对话框，支持多选
            success_count = 0
            error_count = 0
            error_details = []
            
            for ip_item in selected_items:
                ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(ip_item)
                
                try:
                    if db_record_id:
                        # 根据记录ID获取当前记录信息
                        ip_info = self.ipam.get_ip_record_by_id(db_record_id)
                        if ip_info:
                            description = ip_info.get('description', '快速保留') or '快速保留'
                        else:
                            description = '快速保留'
                        # 快速保留，使用数据库记录ID
                        success, message = self.ipam.reserve_ip(network, ip_address, '', description, None, db_record_id)
                    else:
                        # 如果没有数据库记录ID，使用默认方法
                        ip_info = self.ipam.get_ip_info(ip_address)
                        if ip_info:
                            description = ip_info.get('description', '快速保留') or '快速保留'
                        else:
                            description = '快速保留'
                        success, message = self.ipam.reserve_ip(network, ip_address, '', description)
                    
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                        error_details.append(f"{ip_address}: {message}")
                except Exception as e:
                    print(f"快速保留失败: {e}")
                    error_count += 1
                    error_details.append(f"{ip_address}: {str(e)}")
            
            # 显示结果
            if success_count > 0:
                self.show_info(_('success'), f"{_('successfully_restored_ips_reserved', count=success_count)}")
                # 刷新IPAM数据并恢复选中状态
                self.refresh_ipam_with_selection()
            if error_count > 0:
                error_msg = f"{_('failed_to_restore_ips', count=error_count)}"
                if error_details:
                    error_msg += "\n" + "\n".join(error_details[:5])
                    if len(error_details) > 5:
                        error_msg += f"\n...还有 {len(error_details) - 5} 个错误"
                self.show_error(_('error'), error_msg)
        elif action == 'cleanup':
            # 清理已释放的IP地址 - 支持多选
            success_count = 0
            error_count = 0
            error_details = []
            
            for ip_item in selected_items:
                ip_address = self.ipam_ip_tree.item(ip_item, 'values')[0]
                status = self.ipam_ip_tree.item(ip_item, 'values')[1]
                
                # 只清理状态为已释放的IP地址
                if status != _('released'):
                    continue
                
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(ip_item)
                
                if db_record_id:
                    try:
                        # 根据记录ID清理已释放的IP地址（数据库层会再次验证状态）
                        success, message = self.ipam.cleanup_released_ip_by_id(db_record_id)
                        if success:
                            success_count += 1
                        else:
                            error_count += 1
                            error_details.append(f"{ip_address}: {message}")
                    except Exception as e:
                        error_count += 1
                        error_details.append(f"{ip_address}: {str(e)}")
            
            # 显示结果
            if success_count > 0:
                self.show_info(_('success'), f"成功清理 {success_count} 个IP地址")
                # 刷新IPAM数据
                self.refresh_ipam_with_selection()
            if error_count > 0:
                error_msg = f"清理失败 {error_count} 个IP地址"
                if error_details:
                    error_msg += "\n" + "\n".join(error_details[:5])
                    if len(error_details) > 5:
                        error_msg += f"\n...还有 {len(error_details) - 5} 个错误"
                self.show_error(_('error'), error_msg)
    
    def sort_ip_table(self, column):
        """排序IP地址表格"""
        # 切换排序顺序
        if column == self.sort_column:
            self.sort_order = 'desc' if self.sort_order == 'asc' else 'asc'
        else:
            self.sort_column = column
            self.sort_order = 'asc'
        
        # 获取当前表格中的所有数据
        items = []
        for item in self.ipam_ip_tree.get_children():
            values = self.ipam_ip_tree.item(item, 'values')
            tags = self.ipam_ip_tree.item(item, 'tags')
            items.append((item, values, tags))
        
        # 根据排序列和排序顺序对数据进行排序
        def get_sort_value(item):
            values = item[1]
            if column == 'ip_address':
                # IP地址排序，按版本分组后按数值排序
                try:
                    addr = ipaddress.ip_address(values[0])
                    return (addr.version, int(addr))
                except (ValueError, AttributeError):
                    return (0, values[0])
            elif column == 'allocated_at':
                # 时间排序
                return values[5] if values[5] else ''
            elif column == 'expiry_date':
                # 过期日期排序
                return values[6] if values[6] else ''
            else:
                # 其他列直接排序
                index = {'status': 1, 'hostname': 2, 'mac_address': 3, 'description': 4}[column]
                return values[index]
        
        # 排序
        items.sort(key=get_sort_value, reverse=(self.sort_order == 'desc'))
        
        # 清空表格并重新插入排序后的数据
        for item in self.ipam_ip_tree.get_children():
            self.ipam_ip_tree.delete(item)
        
        for item_id, values, tags in items:
            self.ipam_ip_tree.insert('', tk.END, iid=item_id, tags=tags, values=values)
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(self.ipam_ip_tree)
    
    def on_search_input(self, event):
        """搜索输入事件处理"""
        # 获取搜索关键词
        search_text = self.ipam_search_entry.get().strip()
        
        # 执行过滤
        self.apply_filter()
        
    
    def on_filter_change(self, event):
        """过滤条件变化事件处理"""
        # 应用过滤
        self.apply_filter()
    
    def allocate_reserve_ip(self):
        """分配/保留IP地址"""
        selected_items = self.ipam_network_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_network_for_allocation'))
            return
        
        network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        
        # 显示IP地址分配/保留对话框，传递network参数
        dialog_result = self.show_ip_address_dialog(_('allocate_reserve_address'), 'allocate_reserve', network=network)
        if not dialog_result:
            return
        
        ip_address = dialog_result['ip']
        hostname = dialog_result['hostname']
        mac_address = dialog_result.get('mac_address', '')
        description = dialog_result['description']
        expiry_date = dialog_result.get('expiry_date', '')
        action = dialog_result.get('action', 'allocate')
        
        if not ip_address:
            self.show_error(_('error'), _('please_enter_ip_address'))
            return
        
        # 根据用户选择的操作类型执行相应的操作
        if action == 'allocate':
            is_valid, error_msg = IPAMValidator.validate_allocation_params(hostname, description)
            if not is_valid:
                self.show_error(_('error'), _('please_enter_description'))
                return
            # 调用IPAM模块分配IP地址
            success, message = self.ipam.allocate_ip(network, ip_address, hostname, description, expiry_date)
        else:  # reserve
            # 调用IPAM模块保留IP地址
            success, message = self.ipam.reserve_ip(network, ip_address, hostname, description, expiry_date)
        
        if success:
            # 显示成功消息
            self.show_info(_('success'), message)
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        else:
            # 显示错误消息
            self.show_error(_('error'), message)
    
    def allocate_ip(self):
        """分配IP地址"""
        selected_items = self.ipam_network_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_network'))
            return
        
        network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        
        # 显示IP地址分配对话框，传递network参数
        dialog_result = self.show_ip_address_dialog(_('allocate_address'), 'allocate', network=network)
        if not dialog_result:
            return
        
        ip_address = dialog_result['ip']
        hostname = dialog_result['hostname']
        description = dialog_result['description']
        expiry_date = dialog_result.get('expiry_date', '')
        
        if not ip_address:
            self.show_error(_('error'), _('please_enter_ip_address'))
            return
        
        is_valid, error_msg = IPAMValidator.validate_allocation_params(hostname, description)
        if not is_valid:
            self.show_error(_('error'), _('please_enter_description'))
            return
        
        success, message = self.ipam.allocate_ip(network, ip_address, hostname, description, expiry_date)
        if success:
            self.show_info(_('success'), message)
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        else:
            self.show_error(_('error'), message)
    
    def release_ip_address(self):
        """释放IP地址（支持单个或批量释放）"""
        # 检查是否选择了IP地址
        selected_ip_items = self.ipam_ip_tree.selection()
        if not selected_ip_items:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 保存当前选中的网络
        selected_network = None
        selected_network_items = self.ipam_network_tree.selection()
        if selected_network_items:
            selected_network = self.ipam_network_tree.item(selected_network_items[0], 'values')[0]
        
        # 批量释放IP地址
        success_count = 0
        error_count = 0
        for item in selected_ip_items:
            ip_address = self.ipam_ip_tree.item(item, 'values')[0]
            status = self.ipam_ip_tree.item(item, 'values')[1]
            # 只有状态为"已分配"或"已保留"的IP地址才能被释放
            if status in [_('allocated'), _('reserved')]:
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(item)
                
                if db_record_id:
                    # 使用specific策略释放特定记录
                    success, message = self.ipam.release_ip(ip_address, release_strategy="specific", record_id=db_record_id)
                else:
                    # 如果没有数据库记录ID，使用默认策略
                    success, message = self.ipam.release_ip(ip_address)
                
                if success:
                    success_count += 1
                else:
                    error_count += 1
                    # 记录释放失败的原因
                    print(f"释放IP {ip_address} 失败: {message}")
            else:
                # 跳过已经是已释放状态的IP地址
                print(f"跳过IP {ip_address}，状态为: {status}")
        
        # 显示结果
        if success_count > 0 and error_count > 0:
            # 既有成功又有失败
            self.show_info(_('info'), f"成功释放 {success_count} 个IP地址，释放失败 {error_count} 个IP地址")
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        elif success_count > 0:
            # 全部成功
            self.show_info(_('success'), f"{_('successfully_released_ips', count=success_count)}")
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        elif error_count > 0:
            # 全部失败
            self.show_error(_('error'), f"释放失败 {error_count} 个IP地址")
        else:
            # 检查是否所有选中的IP地址都是可用状态
            all_available = True
            for item in selected_ip_items:
                status = self.ipam_ip_tree.item(item, 'values')[1]
                if status not in [_('released')]:
                    all_available = False
                    break
            
            if all_available:
                self.show_info(_('hint'), _('ip_already_available_no_need_to_release'))
            else:
                self.show_error(_('error'), _('failed_to_release_ip'))

    def cleanup_available_ips(self):
        """清理所有可用状态的IP地址"""
        # 获取可用IP地址
        available_ips = self.ipam.get_available_ips()
        
        if available_ips:
            # 显示可用IP地址列表对话框
            self.show_available_ips_dialog(available_ips)
        else:
            self.show_info(_('info'), _('no_available_ips_found'))
    
    def show_available_ips_dialog(self, available_ips):
        """显示可用IP地址列表对话框
        
        Args:
            available_ips: 可用IP地址列表
        """
        # 创建对话框
        dialog = ComplexDialog(self.root, _('available_ips_detected'), 800, 600, resizable=True, modal=True)
        
        # 设置字体
        font_family, font_size = get_current_font_settings()
        
        # 创建框架
        main_frame = ttk.Frame(dialog.content_frame, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 添加统计信息
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 计算统计信息
        total_ips = len(available_ips)
        network_count = len(set(self.ipam.get_most_specific_network(ip['ip_address'])['network_address'] for ip in available_ips if self.ipam.get_most_specific_network(ip['ip_address'])))
        
        # 显示统计信息
        stats_label = ttk.Label(stats_frame, text=_('found_available_ips_count', count=total_ips, network_count=network_count), font=(font_family, font_size))
        stats_label.pack(anchor=tk.W)
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(main_frame)
        
        # 创建树状视图 - 不显示网络地址列
        tree = ttk.Treeview(main_frame, columns=('ip_address', 'status', 'hostname', 'mac_address', 'description', 'expiry_date'), show="headings")
        
        # 创建滚动条回调函数，实现自动隐藏功能
        def scrollbar_callback(*args):
            scrollbar.set(*args)
            if float(args[1]) - float(args[0]) >= 1.0 - 1e-9:
                scrollbar.pack_forget()
            else:
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 配置滚动条和Treeview
        tree.configure(yscrollcommand=scrollbar_callback)
        scrollbar.config(command=tree.yview)
        
        # 使用pack布局
        tree.pack(fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 初始调用回调函数
        scrollbar_callback(0.0, 1.0)
        
        # 配置列
        tree.heading('ip_address', text=_('ip_address'))
        tree.heading('status', text=_('status'))
        tree.heading('hostname', text=_('hostname'))
        tree.heading('mac_address', text=_('mac_address'))
        tree.heading('description', text=_('description'))
        tree.heading('expiry_date', text=_('expiry_date'))
        tree.column('ip_address', width=90, stretch=False)
        tree.column('status', width=60, stretch=False)
        tree.column('hostname', width=100, stretch=False)
        tree.column('mac_address', width=110, stretch=False)
        tree.column('description', width=100, stretch=True)
        tree.column('expiry_date', width=110, stretch=False)
        
        # 添加数据
        # 按IP地址排序
        sorted_ips = self._sort_ip_list(available_ips)
        for ip in sorted_ips:
            # 通过IP地址获取归属网络
            network_info = self.ipam.get_most_specific_network(ip['ip_address'])
            network = network_info['network_address'] if network_info else ''
            
            # 翻译状态值
            status = ip['status']
            if status == 'released':
                translated_status = _('released')
            elif status == 'allocated':
                translated_status = _('allocated')
            elif status == 'reserved':
                translated_status = _('reserved')
            else:
                translated_status = status
            
            # 处理过期日期，只显示日期部分
            expiry_date = ip['expiry_date'] or ''
            if expiry_date:
                # 提取日期部分（处理 YYYY-MM-DD HH:MM 和 YYYY-MM-DDTHH:MM 格式）
                if ' ' in expiry_date:
                    expiry_date = expiry_date.split(' ')[0]
                elif 'T' in expiry_date:
                    expiry_date = expiry_date.split('T')[0]
            
            record_id = ip.get('id', None)
            if record_id:
                tree.insert('', tk.END, iid=str(record_id),
                            values=(ip['ip_address'], translated_status,
                                    ip['hostname'] or '', ip.get('mac_address', '') or '',
                                    ip['description'] or '', expiry_date))
            else:
                tree.insert('', tk.END,
                            values=(ip['ip_address'], translated_status,
                                    ip['hostname'] or '', ip.get('mac_address', '') or '',
                                    ip['description'] or '', expiry_date))
        
        # 配置斑马纹样式并应用
        self.configure_treeview_styles(tree)
        self.update_table_zebra_stripes(tree)
        
        # 配置滚动条
        scrollbar.config(command=tree.yview)
        
        # 使用dialog.button_frame放置按钮
        # 左侧按钮
        cleanup_selected_button = ttk.Button(dialog.button_frame, text=_('cleanup_selected'), 
                                           command=lambda: self.cleanup_selected_available_ips(tree, dialog, available_ips))
        cleanup_selected_button.pack(side=tk.LEFT, padx=5)
        
        # 创建全部清理按钮
        cleanup_all_button = ttk.Button(dialog.button_frame, text=_('cleanup_all'), 
                                      command=lambda: self.cleanup_all_available_ips(dialog, available_ips))
        cleanup_all_button.pack(side=tk.LEFT, padx=5)
        
        # 右侧关闭按钮
        cancel_button = ttk.Button(dialog.button_frame, text=_('close'), command=dialog.destroy)
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        # 显示对话框
        dialog.show()
    
    def cleanup_selected_available_ips(self, tree, dialog, available_ips):
        """清理选中的可用IP地址
        
        Args:
            tree: 树状视图控件
            dialog: 对话框
            available_ips: 可用IP地址列表
        """
        selected_items = tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 构建以id为键的字典，用于精确匹配
        ip_by_id = {}
        for ip in available_ips:
            if 'id' in ip:
                ip_by_id[str(ip['id'])] = ip
        
        # 通过iid（数据库记录ID）获取选中的IP对象
        selected_ips = []
        for item_id in selected_items:
            if item_id in ip_by_id:
                selected_ips.append(ip_by_id[item_id])
        
        if not selected_ips:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 清理选中的IP地址
        cleaned_count = 0
        try:
            conn = sqlite3.connect(self.ipam.db_file)
            cursor = conn.cursor()
            
            # 记录清理历史
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for ip in selected_ips:
                # 记录清理历史
                cursor.execute('''
                INSERT INTO allocation_history (ip_address, action, performed_by, performed_at)
                VALUES (?, ?, ?, ?)
                ''', (ip['ip_address'], 'cleanup', 'admin', now))
                
                # 删除可用状态的IP地址
                cursor.execute('DELETE FROM ip_addresses WHERE id = ? AND status = ?', (ip['id'], 'released'))
                
                # 删除对应的隐藏信息记录
                cursor.execute('DELETE FROM ip_hidden_info WHERE ip_record_id = ?', (ip['id'],))
                
                cleaned_count += 1
            
            conn.commit()
            conn.close()
        except Exception as e:
            self.show_error(_('error'), f"清理IP地址失败: {str(e)}")
            return
        
        if cleaned_count > 0:
            self.show_info(_('success'), f"{_('successfully_cleaned_ips', count=cleaned_count)}")
            # 刷新网络列表和IP列表
            self.refresh_ipam_networks()
            selected_network_items = self.ipam_network_tree.selection()
            if selected_network_items:
                network = self.ipam_network_tree.item(selected_network_items[0], 'values')[0]
                self.refresh_ipam_ips(network)
            # 刷新对话框中的表格
            self.refresh_available_ips_table(tree)
        
        # 不关闭对话框，保持窗口打开
    
    def refresh_available_ips_table(self, tree):
        """刷新可用IP地址表格
        
        Args:
            tree: 树状视图控件
        """
        # 清空表格
        for item in tree.get_children():
            tree.delete(item)
        
        # 获取最新的可用IP地址列表
        available_ips = self.ipam.get_available_ips()
        
        # 添加新数据
        sorted_ips = self._sort_ip_list(available_ips)
        for ip in sorted_ips:
            network_info = self.ipam.get_most_specific_network(ip['ip_address'])
            network = network_info['network_address'] if network_info else ''
            
            status = ip['status']
            if status == 'released':
                translated_status = _('released')
            elif status == 'allocated':
                translated_status = _('allocated')
            elif status == 'reserved':
                translated_status = _('reserved')
            else:
                translated_status = status
            
            # 处理过期日期，只显示日期部分
            expiry_date = ip.get('expiry_date', '') or ''
            if expiry_date:
                # 提取日期部分（处理 YYYY-MM-DD HH:MM 和 YYYY-MM-DDTHH:MM 格式）
                if ' ' in expiry_date:
                    expiry_date = expiry_date.split(' ')[0]
                elif 'T' in expiry_date:
                    expiry_date = expiry_date.split('T')[0]
            
            record_id = ip.get('id', None)
            if record_id:
                tree.insert('', tk.END, iid=str(record_id), values=(ip['ip_address'], translated_status,
                                                                    ip['hostname'] or '', ip.get('mac_address', '') or '',
                                                                    ip['description'] or '', expiry_date))
            else:
                tree.insert('', tk.END, values=(ip['ip_address'], translated_status,
                                                ip['hostname'] or '', ip.get('mac_address', '') or '',
                                                ip['description'] or '', expiry_date))
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(tree)
    
    def cleanup_all_available_ips(self, dialog, available_ips):
        """清理所有可用IP地址
        
        Args:
            dialog: 对话框
            available_ips: 可用IP地址列表
        """
        # 清理所有可用IP
        success, message = self.ipam.cleanup_available_ips()
        if success:
            self.show_info(_('success'), message)
            # 刷新网络列表和IP列表
            self.refresh_ipam_networks()
            selected_network_items = self.ipam_network_tree.selection()
            if selected_network_items:
                network = self.ipam_network_tree.item(selected_network_items[0], 'values')[0]
                self.refresh_ipam_ips(network)
        else:
            self.show_error(_('error'), message)
        
        # 关闭对话框
        dialog.destroy()
    
    def check_expired_ips(self):
        """检查过期IP地址并处理"""
        # 获取过期IP地址
        expired_ips = self.ipam.get_expired_ips()
        
        if expired_ips:
            # 显示过期IP地址列表对话框
            self.show_expired_ips_dialog(expired_ips)
        else:
            self.show_info(_('info'), _('no_expired_ips_found'))
    
    def show_expired_ips_dialog(self, expired_ips):
        """显示过期IP地址列表对话框
        
        Args:
            expired_ips: 过期IP地址列表
        """
        # 创建对话框
        dialog = ComplexDialog(self.root, _('expired_ips_detected'), 800, 600, resizable=True, modal=True)
        
        # 设置字体
        font_family, font_size = get_current_font_settings()
        
        # 创建框架
        main_frame = ttk.Frame(dialog.content_frame, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 添加统计信息
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 计算统计信息
        total_ips = len(expired_ips)
        network_count = len(set(self.ipam.get_most_specific_network(ip['ip_address'])['network_address'] for ip in expired_ips if self.ipam.get_most_specific_network(ip['ip_address'])))
        
        # 显示统计信息
        stats_label = ttk.Label(stats_frame, text=_('found_expired_ips_count', count=total_ips, network_count=network_count), font=(font_family, font_size))
        stats_label.pack(anchor=tk.W)
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(main_frame)
        
        # 创建树状视图 - 不显示网络地址列
        tree = ttk.Treeview(main_frame, columns=('ip_address', 'status', 'hostname', 'mac_address', 'description', 'expiry_date'), show="headings")
        
        # 创建滚动条回调函数，实现自动隐藏功能
        def scrollbar_callback(*args):
            scrollbar.set(*args)
            if float(args[1]) - float(args[0]) >= 1.0 - 1e-9:
                scrollbar.pack_forget()
            else:
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 配置滚动条和Treeview
        tree.configure(yscrollcommand=scrollbar_callback)
        scrollbar.config(command=tree.yview)
        
        # 使用pack布局
        tree.pack(fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 初始调用回调函数
        scrollbar_callback(0.0, 1.0)
        
        # 配置列
        tree.heading('ip_address', text=_('ip_address'))
        tree.heading('status', text=_('status'))
        tree.heading('hostname', text=_('hostname'))
        tree.heading('mac_address', text=_('mac_address'))
        tree.heading('description', text=_('description'))
        tree.heading('expiry_date', text=_('expiry_date'))
        tree.column('ip_address', width=90, stretch=False)
        tree.column('status', width=60, stretch=False)
        tree.column('hostname', width=100, stretch=False)
        tree.column('mac_address', width=110, stretch=False)
        tree.column('description', width=100, stretch=True)
        tree.column('expiry_date', width=110, stretch=False)
        
        # 添加数据
        # 按IP地址排序
        sorted_ips = self._sort_ip_list(expired_ips)
        for ip in sorted_ips:
            # 翻译状态值
            status = ip['status']
            if status == 'released':
                translated_status = _('released')
            elif status == 'allocated':
                translated_status = _('allocated')
            elif status == 'reserved':
                translated_status = _('reserved')
            else:
                translated_status = status
            
            # 处理过期日期，只显示日期部分
            expiry_date = ip['expiry_date'] or ''
            if expiry_date:
                # 提取日期部分（处理 YYYY-MM-DD HH:MM 和 YYYY-MM-DDTHH:MM 格式）
                if ' ' in expiry_date:
                    expiry_date = expiry_date.split(' ')[0]
                elif 'T' in expiry_date:
                    expiry_date = expiry_date.split('T')[0]
            
            # 使用记录ID作为树项的ID
            record_id = ip.get('id', None)
            if record_id:
                tree.insert('', tk.END, iid=str(record_id), values=(ip['ip_address'], translated_status,
                                                                    ip['hostname'] or '', ip.get('mac_address', '') or '',
                                                                    ip['description'] or '', expiry_date))
            else:
                tree.insert('', tk.END, values=(ip['ip_address'], translated_status,
                                                ip['hostname'] or '', ip.get('mac_address', '') or '',
                                                ip['description'] or '', expiry_date))
        
        # 配置斑马纹样式并应用
        self.configure_treeview_styles(tree)
        self.update_table_zebra_stripes(tree)
        
        # 配置滚动条
        scrollbar.config(command=tree.yview)
        
        # 使用dialog.button_frame放置按钮
        # 左侧按钮
        release_button = ttk.Button(dialog.button_frame, text=_('release_selected'), 
                                   command=lambda: self.release_selected_expired_ips(tree, dialog, expired_ips))
        release_button.pack(side=tk.LEFT, padx=5)
        
        # 创建延期按钮
        extend_button = ttk.Button(dialog.button_frame, text=_('extend_expiry'), 
                                  command=lambda: self.extend_selected_expired_ips(tree, dialog, expired_ips))
        extend_button.pack(side=tk.LEFT, padx=5)
        
        # 创建全部释放按钮
        release_all_button = ttk.Button(dialog.button_frame, text=_('release_all'), 
                                      command=lambda: self.release_all_expired_ips(dialog, expired_ips))
        release_all_button.pack(side=tk.LEFT, padx=5)
        
        # 右侧关闭按钮
        cancel_button = ttk.Button(dialog.button_frame, text=_('close'), command=dialog.destroy)
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        # 显示对话框
        dialog.show()
    
    def release_selected_expired_ips(self, tree, dialog, expired_ips):
        """释放选中的过期IP地址
        
        Args:
            tree: 树状视图控件
            dialog: 对话框
            expired_ips: 过期IP地址列表
        """
        selected_items = tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 构建以id为键的字典，用于精确匹配
        ip_by_id = {}
        for ip in expired_ips:
            if 'id' in ip:
                ip_by_id[str(ip['id'])] = ip
        
        # 通过iid（数据库记录ID）获取选中的IP对象
        selected_ips = []
        for item_id in selected_items:
            if item_id in ip_by_id:
                selected_ips.append(ip_by_id[item_id])
        
        if not selected_ips:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 释放选中的IP地址
        released_count = 0
        for ip in selected_ips:
            try:
                db_record_id = ip.get('id')
                ip_address = ip['ip_address']
                if db_record_id:
                    success, message = self.ipam.release_ip(ip_address, release_strategy="specific", record_id=db_record_id)
                else:
                    success, message = self.ipam.release_ip(ip_address)
                if success:
                    released_count += 1
            except Exception as e:
                print(f"释放过期IP失败: {e}")
        
        if released_count > 0:
            self.show_info(_('success'), f"{_('successfully_released_ips', count=released_count)}")
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
            # 刷新对话框中的表格
            self.refresh_expired_ips_table(tree)
        
        # 不关闭对话框，保持窗口打开
    
    def refresh_expired_ips_table(self, tree):
        """刷新过期IP地址表格
        
        Args:
            tree: 树状视图控件
        """
        # 清空表格
        for item in tree.get_children():
            tree.delete(item)
        
        # 获取最新的过期IP地址列表
        expired_ips = self.ipam.get_expired_ips()
        
        # 添加新数据
        sorted_ips = self._sort_ip_list(expired_ips)
        for ip in sorted_ips:
            status = ip['status']
            if status == 'released':
                translated_status = _('released')
            elif status == 'allocated':
                translated_status = _('allocated')
            elif status == 'reserved':
                translated_status = _('reserved')
            else:
                translated_status = status
            
            # 处理过期日期，只显示日期部分
            expiry_date = ip.get('expiry_date', '') or ''
            if expiry_date:
                if ' ' in expiry_date:
                    expiry_date = expiry_date.split(' ')[0]
                elif 'T' in expiry_date:
                    expiry_date = expiry_date.split('T')[0]
            
            record_id = ip.get('id', None)
            if record_id:
                tree.insert('', tk.END, iid=str(record_id), values=(ip['ip_address'], translated_status,
                                                                    ip['hostname'] or '', ip.get('mac_address', '') or '',
                                                                    ip['description'] or '', expiry_date))
            else:
                tree.insert('', tk.END, values=(ip['ip_address'], translated_status,
                                                ip['hostname'] or '', ip.get('mac_address', '') or '',
                                                ip['description'] or '', expiry_date))
        
        # 更新斑马纹样式
        self.update_table_zebra_stripes(tree)
    
    def release_all_expired_ips(self, dialog, expired_ips):
        """释放所有过期IP地址
        
        Args:
            dialog: 对话框
            expired_ips: 过期IP地址列表
        """
        # 自动释放过期IP
        success, msg, released_count = self.ipam.auto_release_expired_ips()
        if success:
            self.show_info(_('success'), msg)
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        else:
            self.show_error(_('error'), msg)
        
        # 关闭对话框
        dialog.destroy()
    
    def extend_selected_expired_ips(self, tree, dialog, expired_ips):
        """延期选中的过期IP地址
        
        Args:
            tree: 树状视图控件
            dialog: 对话框
            expired_ips: 过期IP地址列表
        """
        selected_items = tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 构建以id为键的字典，用于精确匹配
        ip_by_id = {}
        for ip in expired_ips:
            if 'id' in ip:
                ip_by_id[str(ip['id'])] = ip
        
        # 通过iid（数据库记录ID）获取选中的IP对象
        selected_ips = []
        for item_id in selected_items:
            if item_id in ip_by_id:
                selected_ips.append(ip_by_id[item_id])
        
        if not selected_ips:
            self.show_error(_('error'), _('please_select_ip_address'))
            return
        
        # 显示延期对话框
        self.show_extend_expiry_dialog(selected_ips, tree, dialog)
    
    def show_extend_expiry_dialog(self, selected_ips, tree, parent_dialog):
        """显示延期对话框
        
        Args:
            selected_ips: 选中的IP地址列表
            tree: 树状视图控件
            parent_dialog: 父对话框
        """
        # 创建对话框
        dialog = ComplexDialog(parent_dialog.dialog, _('extend_expiry'), 400, 200, resizable=False, modal=True)
        
        # 设置字体
        font_family, font_size = get_current_font_settings()
        
        # 创建主框架
        main_frame = ttk.Frame(dialog.content_frame, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建标题
        title_label = ttk.Label(
            main_frame, 
            text=_('select_extend_duration'), 
            font=(font_family, font_size + 2, 'bold'),
            foreground='#333333'
        )
        title_label.pack(pady=(0, 20))
        
        # 延期时间选项
        extend_options = [
            (_('1_month'), 30),
            (_('3_months'), 90),
            (_('6_months'), 180),
            (_('1_year'), 365),
            (_('3_years'), 1095),
            (_('5_years'), 1825)
        ]
        
        # 变量
        selected_option = tk.StringVar()
        selected_option.set(extend_options[0][0])
        
        # 创建选项菜单
        option_frame = ttk.Frame(main_frame)
        option_frame.pack(fill=tk.X, pady=10)
        
        # 创建三列布局
        left_frame = ttk.Frame(option_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, expand=True, padx=10)
        
        middle_frame = ttk.Frame(option_frame)
        middle_frame.pack(side=tk.LEFT, fill=tk.Y, expand=True, padx=10)
        
        right_frame = ttk.Frame(option_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.Y, expand=True, padx=10)
        
        # 分配选项到三列
        for i, (text, days) in enumerate(extend_options):
            if i % 3 == 0:
                frame = left_frame
            elif i % 3 == 1:
                frame = middle_frame
            else:
                frame = right_frame
            
            radiobutton = ttk.Radiobutton(
                frame, 
                text=text, 
                variable=selected_option, 
                value=text
            )
            radiobutton.pack(anchor=tk.W, pady=5)
        
        # 创建确定按钮
        def on_confirm():
            # 获取选中的延期天数
            selected_text = selected_option.get()
            days = None
            for text, d in extend_options:
                if text == selected_text:
                    days = d
                    break
            
            if days is not None:
                # 执行延期操作
                extended_count = 0
                for ip in selected_ips:
                    current_expiry = ip.get('expiry_date', None)
                    if current_expiry:
                        try:
                            # 计算新的过期日期，设置为当天的最后一秒 (23:59:59)
                            new_expiry = datetime.datetime.now() + datetime.timedelta(days=days)
                            new_expiry = new_expiry.replace(hour=23, minute=59, second=59)
                            new_expiry_str = new_expiry.strftime('%Y-%m-%d %H:%M:%S')
                            
                            # 更新过期日期
                            record_id = ip.get('id', None)
                            success, message = self.ipam.update_ip_expiry(ip['ip_address'], new_expiry_str, record_id)
                            if success:
                                extended_count += 1
                        except Exception as e:
                            self.show_error(_('error'), f"更新过期日期失败: {str(e)}")
                
                if extended_count > 0:
                    self.show_info(_('success'), f"{_('successfully_extended_ips', count=extended_count)}")
                    # 刷新IPAM数据并恢复选中状态
                    self.refresh_ipam_with_selection()
                    # 刷新对话框中的表格
                    self.refresh_expired_ips_table(tree)
                
                # 关闭对话框
                dialog.destroy()
        
        # 添加确定按钮
        dialog.add_button(_('confirm'), on_confirm, column=2)
        
        # 添加取消按钮
        dialog.add_button(_('cancel'), dialog.destroy, column=1)
        
        # 显示对话框
        dialog.show()
    
    def refresh_ipam_with_selection(self):
        """刷新IPAM数据并恢复选中状态
        
        保存当前选中的网络，刷新网络列表和IP列表，然后重新选中之前的网络
        """
        # 保存当前选中的所有网络
        selected_networks = []
        selected_network_items = self.ipam_network_tree.selection()
        for item in selected_network_items:
            network = self.ipam_network_tree.item(item, 'values')[0]
            selected_networks.append(network)
        
        # 刷新网络列表
        self.refresh_ipam_networks()
        
        # 重新选中之前的所有网络
        if selected_networks:
            # 查找并选中之前的所有网络
            for item in self.ipam_network_tree.get_children():
                network = self.ipam_network_tree.item(item, 'values')[0]
                if network in selected_networks:
                    self.ipam_network_tree.selection_add(item)
            # 刷新第一个选中网络的IP列表
            self.refresh_ipam_ips(selected_networks[0])
        else:
            # 如果没有选中的网络，刷新第一个网络的IP列表
            network_items = self.ipam_network_tree.get_children()
            if network_items:
                network = self.ipam_network_tree.item(network_items[0], 'values')[0]
                self.ipam_network_tree.selection_set(network_items[0])
                self.refresh_ipam_ips(network)

    def auto_release_expired_ips(self):
        """手动触发自动释放过期IP地址"""
        success, msg, released_count = self.ipam.auto_release_expired_ips()
        if success:
            self.show_info(_('success'), msg)
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
        else:
            self.show_error(_('error'), msg)
    
    def batch_migrate_ip_addresses(self):
        """批量迁移IP地址到其他网络"""
        # 检查是否选择了IP地址
        selected_items = self.ipam_ip_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_ips_to_migrate'))
            return
        
        # 获取选中的记录ID和IP地址
        ip_records = []
        for item in selected_items:
            try:
                # 从tags中获取数据库记录ID
                db_record_id = self._get_db_record_id(item)
                values = self.ipam_ip_tree.item(item, 'values')
                if values and len(values) > 0:
                    ip_address = values[0]
                    ip_records.append({'id': db_record_id, 'ip_address': ip_address})
            except (ValueError, IndexError):
                # 如果树项ID不是整数，直接使用IP地址
                try:
                    values = self.ipam_ip_tree.item(item, 'values')
                    if values and len(values) > 0:
                        ip_address = values[0]
                        # 尝试通过IP地址获取记录ID
                        # 这里简化处理，直接使用IP地址作为标识
                        ip_records.append({'id': None, 'ip_address': ip_address})
                except (ValueError, IndexError):
                    pass
        
        if not ip_records:
            self.show_error(_('error'), _('please_select_ips_to_migrate'))
            return
        
        # 获取所有可用的网络列表
        all_networks = self.ipam.get_all_networks()
        network_list = [net['network'] for net in all_networks]
        
        # 对网络列表进行排序（按IP版本分组，再按地址数值排序）
        try:
            network_list.sort(key=lambda x: (ipaddress.ip_network(x, strict=False).version, int(ipaddress.ip_network(x, strict=False).network_address)))
        except Exception:
            # 如果排序失败，保持原顺序
            pass
        
        if not network_list:
            self.show_info(_('hint'), _('no_networks_available'))
            return
        
        # 创建对话框
        dialog = ComplexDialog(self.root, _('batch_migrate'), 400, 200, resizable=False, modal=True)
        
        # 创建主容器框架，增加外边距
        main_container = ttk.Frame(dialog.content_frame, padding=(15, 15, 15, 10))
        main_container.pack(fill='both', expand=True)
        
        # 显示选中的IP数量（靠左上放置）
        ttk.Label(main_container, text=_('selected_ips_count').format(count=len(ip_records))).pack(side='top', anchor='w', pady=(0, 10))
        
        # 目标网络选择 - 将标签和下拉框放在同一行
        network_frame = ttk.Frame(main_container)
        network_frame.pack(pady=5, fill='x')
        
        ttk.Label(network_frame, text=_('target_network') + ':').pack(side='left', padx=(0, 10))
        
        # 创建网络下拉列表
        network_var = tk.StringVar()
        network_combobox = ttk.Combobox(network_frame, textvariable=network_var, values=network_list, width=30)
        network_combobox.pack(side='left', fill='x', expand=True)
        network_combobox.focus_set()
        
        # 确认按钮
        def confirm():
            target_network = network_var.get().strip()
            if not target_network:
                self.show_info(_('hint'), _('please_select_target_network'))
                return
            
            # 执行批量迁移
            success, msg, migrated_count = self.ipam.batch_migrate_ips(ip_records, target_network)
            if success:
                self.show_info(_('success'), msg)
                # 刷新IPAM数据并切换到目标网络
                self.refresh_ipam_networks()
                # 选中目标网络并刷新其IP列表
                for item in self.ipam_network_tree.get_children():
                    network = self.ipam_network_tree.item(item, 'values')[0]
                    if network == target_network:
                        self.ipam_network_tree.selection_set(item)
                        self.refresh_ipam_ips(target_network)
                        break
            else:
                self.show_error(_('error'), msg)
            
            dialog.destroy()
        
        # 取消按钮
        def cancel():
            dialog.destroy()
        
        # 添加按钮
        dialog.add_button(_('cancel'), cancel, column=1)
        dialog.add_button(_('confirm'), confirm, column=2)
        
        # 显示对话框
        dialog.show()
    
    def batch_set_expiry_date(self):
        """批量设置IP地址的过期日期"""
        # 检查是否选择了IP地址
        selected_items = self.ipam_ip_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_ips_to_set_expiry'))
            return
        
        # 获取选中的记录ID和IP地址
        record_ids = []
        ip_addresses = []
        for item in selected_items:
            # 从tags中获取数据库记录ID
            db_record_id = self._get_db_record_id(item)
            ip_address = self.ipam_ip_tree.item(item, 'values')[0]
            ip_addresses.append(ip_address)
            if db_record_id:
                record_ids.append(db_record_id)
        
        # 创建对话框
        dialog = ComplexDialog(self.root, _('batch_set_expiry_date'), 400, 200, resizable=False, modal=True)
        
        # 创建日期选择器 - 使用grid布局将标签和下拉控件放在同一行
        dialog.content_frame.grid_columnconfigure(0, weight=0)
        dialog.content_frame.grid_columnconfigure(1, weight=1)
        
        # 显示选中的IP数量（靠左上放置）
        ttk.Label(dialog.content_frame, text=_('selected_ips_count').format(count=len(ip_addresses))).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10), padx=(30, 0))
        
        ttk.Label(dialog.content_frame, text=_('expiry_date') + ':').grid(row=1, column=0, sticky="e", pady=5, padx=(30, 5))
        
        # 创建日期输入框
        from datetime import datetime
        default_expiry = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
        
        date_entry = None
        
        if DateEntry:
            # 直接创建DateEntry日期选择器
            date_entry = DateEntry(dialog.content_frame, date_pattern='yyyy-MM-dd')
            date_entry.grid(row=1, column=1, sticky="ew", pady=5, padx=(5, 30))
            date_entry.set_date(default_expiry)
            
            # 修复模态对话框中DateEntry日历弹窗问题
            fix_date_entry_for_modal(date_entry, dialog.dialog)
        else:
            # 使用普通Entry输入日期
            date_var = tk.StringVar()
            date_var.set(default_expiry)
            date_entry = ttk.Entry(dialog.content_frame, textvariable=date_var)
            date_entry.grid(row=1, column=1, sticky="ew", pady=10, padx=(0, 10))
            ttk.Label(dialog.content_frame, text=_('date_format_yyyy_mm_dd')).grid(row=2, column=1, sticky="w", pady=5)
        
        # 获取日期值的辅助函数
        def get_expiry_date():
            if date_entry:
                return date_entry.get()
            return None
        
        # 确认按钮
        def confirm():
            expiry_date = get_expiry_date()
            if not expiry_date:
                self.show_info(_('hint'), _('please_enter_expiry_date'))
                return
            
            # 批量更新过期日期
            success, msg, updated_count = self.ipam.batch_update_ip_expiry(ip_addresses, expiry_date, record_ids)
            if success:
                self.show_info(_('success'), msg)
                # 刷新IP列表
                selected_network_items = self.ipam_network_tree.selection()
                if selected_network_items:
                    network = self.ipam_network_tree.item(selected_network_items[0], 'values')[0]
                    self.refresh_ipam_ips(network)
            else:
                self.show_error(_('error'), msg)
            
            dialog.destroy()
        
        # 清除过期日期按钮
        def clear_expiry():
            # 批量清除过期日期
            success, msg, updated_count = self.ipam.batch_update_ip_expiry(ip_addresses, None, record_ids)
            if success:
                self.show_info(_('success'), msg)
                # 刷新IP列表
                selected_network_items = self.ipam_network_tree.selection()
                if selected_network_items:
                    network = self.ipam_network_tree.item(selected_network_items[0], 'values')[0]
                    self.refresh_ipam_ips(network)
            else:
                self.show_error(_('error'), msg)
            
            dialog.destroy()
        
        # 添加按钮
        dialog.add_button(_('confirm'), confirm, column=1)
        dialog.add_button(_('clear_expiry_date'), clear_expiry, column=2)
        
        # 显示对话框
        dialog.show()

    def show_ip_address_dialog(self, title, action_type, ip_address=None, network=None, record_id=None, original_hostname='', original_description=''):
        """显示IP地址分配/保留对话框
        
        Args:
            title: 对话框标题
            action_type: 操作类型，'allocate', 'reserve', 'auto_allocate'
            ip_address: 可选，已选定的IP地址
            network: 可选，当前选定的网络网段
            record_id: 可选，记录ID，用于获取特定记录的信息
            original_hostname: 可选，原始主机名
            original_description: 可选，原始描述
        """
        # 根据操作类型调整对话框大小
        if action_type == 'auto_allocate':
            height = 260
        elif action_type == 'allocate_reserve':
            height = 290  # 减小分配/保留对话框高度，消除空白间隙
        else:
            height = 260
        
        # 设置默认过期日期为1年后
        from datetime import datetime
        default_expiry = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
        
        # 使用统一的create_dialog方法创建居中对话框
        dialog = self.create_dialog(title, 420, height)
        
        # 创建主框架（增加内边距）
        main_frame = ttk.Frame(dialog, padding="30 20 20 0")
        main_frame.pack(fill=tk.X, expand=False, side=tk.TOP)
        
        # 配置网格布局
        main_frame.grid_columnconfigure(0, weight=0)
        main_frame.grid_columnconfigure(1, weight=1)
        
        # 只有在非自动分配时才显示IP地址输入框
        show_ip_input = action_type != 'auto_allocate'
        ip_entry = None
        auto_allocate_var = tk.BooleanVar(value=True)  # 默认选中自动分配
        
        # IP地址输入
        if show_ip_input:
            # IP地址输入框
            ttk.Label(main_frame, text=_('ip_address') + ':').grid(row=0, column=0, sticky="e", pady=5, padx=(0, 10))
            ip_border, ip_entry = create_bordered_entry(main_frame)
            ip_border.grid(row=0, column=1, sticky="ew", pady=5, padx=(0, 10))
            
            # IP地址实时验证（必须不带前缀）
            def validate_ip_input(event=None):
                self.validate_cidr(ip_entry.get(), ip_entry, require_prefix=False)
            
            ip_entry.bind('<FocusOut>', validate_ip_input)
            ip_entry.bind('<KeyRelease>', validate_ip_input)
            
            # 自动分配选项（只在分配/保留对话框中显示）
            if action_type == 'allocate_reserve':
                auto_allocate_frame = ttk.Frame(main_frame)
                auto_allocate_frame.grid(row=1, column=1, sticky="w", pady=5, padx=(0, 0))
                
                def toggle_auto_allocate():
                    """切换自动分配状态"""
                    if auto_allocate_var.get():
                        ip_entry.config(state='readonly')
                        # 自动分配IP地址
                        if network:
                            try:
                                
                                # 获取当前选中网段的所有子网段
                                all_networks = self.ipam.get_all_networks()
                                selected_network_obj = ipaddress.ip_network(network, strict=False)
                                
                                # 筛选出属于当前选中网段的所有子网段
                                child_networks = []
                                for net in all_networks:
                                    try:
                                        net_obj = ipaddress.ip_network(net['network'], strict=False)
                                        if net_obj.subnet_of(selected_network_obj) and net_obj != selected_network_obj:
                                            child_networks.append(net['network'])
                                    except Exception:
                                        pass
                                
                                # 找到前缀大于等于24的子网段
                                prefix_24_plus_networks = []
                                for net in child_networks:
                                    try:
                                        net_obj = ipaddress.ip_network(net)
                                        if net_obj.prefixlen >= 24:
                                            prefix_24_plus_networks.append(net)
                                    except Exception:
                                        pass
                                
                                # 选择目标网络
                                target_network = network
                                if prefix_24_plus_networks:
                                    # 随机选择一个前缀大于等于24的子网段
                                    import random
                                    target_network = random.choice(prefix_24_plus_networks)
                                elif child_networks:
                                    # 如果没有前缀大于等于24的子网段，选择最底层的子网段
                                    child_networks.sort(key=lambda x: ipaddress.ip_network(x).prefixlen, reverse=True)
                                    target_network = child_networks[0]
                                
                                # 获取目标网络中的所有已分配IP地址
                                ips = self.ipam.get_network_ips(target_network)
                                allocated_ips = {ip['ip_address'] for ip in ips}
                                
                                target_network_obj = ipaddress.ip_network(target_network, strict=False)
                                
                                # 对于大型网络，使用更高效的方法
                                if target_network_obj.num_addresses > 1000:
                                    # 生成一个随机IP地址并检查是否可用
                                    import random
                                    max_attempts = 100
                                    for attempt in range(max_attempts):
                                        # 生成网络内的随机IP
                                        if target_network_obj.version == 4:
                                            # IPv4
                                            network_int = int(target_network_obj.network_address)
                                            host_bits = 32 - target_network_obj.prefixlen
                                            random_offset = random.randint(1, 2**host_bits - 2)  # 排除网络地址和广播地址
                                            ip_int = network_int + random_offset
                                            ip_str = str(ipaddress.IPv4Address(ip_int))
                                        else:
                                            # IPv6
                                            # 简化处理，生成网络前缀 + 随机后缀
                                            network_parts = list(target_network_obj.network_address.exploded.split(':'))
                                            # 生成随机后缀
                                            for i in range(len(network_parts)):
                                                if network_parts[i] == '0' * len(network_parts[i]):
                                                    network_parts[i] = format(random.randint(0, 65535), 'x')
                                            ip_str = ':'.join(network_parts)
                                            ip_str = str(ipaddress.IPv6Address(ip_str))
                                        
                                        if ip_str not in allocated_ips:
                                            # 检查IP是否在网络内
                                            if ipaddress.ip_address(ip_str) in target_network_obj:
                                                ip_entry.config(state='normal')
                                                ip_entry.delete(0, tk.END)
                                                ip_entry.insert(0, ip_str)
                                                ip_entry.config(state='readonly')
                                                break
                                else:
                                    # 对于小型网络，逐个检查
                                    for ip in target_network_obj.hosts():
                                        ip_str = str(ip)
                                        if ip_str not in allocated_ips:
                                            ip_entry.config(state='normal')
                                            ip_entry.delete(0, tk.END)
                                            ip_entry.insert(0, ip_str)
                                            ip_entry.config(state='readonly')
                                            break
                            except Exception:
                                pass
                    else:
                        ip_entry.config(state='normal')
                        # 重新填充前缀
                        if network and not ip_address:
                            try:
                                ip_network = ipaddress.ip_network(network, strict=False)
                                prefix = ""
                                
                                if ip_network.version == 4:
                                    # IPv4: 根据前缀长度动态提取网段
                                    user_network = network.split('/')[0]  # 提取IP部分，如192.168.1.0
                                    octets = user_network.split(".")
                                    prefix_len = ip_network.prefixlen
                                    
                                    # 根据前缀长度计算固定的字节数（整数部分）
                                    fixed_octets = prefix_len // 8
                                    
                                    # 特殊处理：如果固定字节数为0，至少提取1字节
                                    fixed_octets = max(1, fixed_octets)
                                    
                                    # 只提取固定的字节数
                                    if fixed_octets == 4:
                                        # /32 子网直接使用完整地址
                                        prefix = user_network
                                    else:
                                        # 提取固定数量的网段并添加点号
                                        prefix = ".".join(octets[:fixed_octets]) + "."
                                else:
                                    # IPv6: 提取前缀，例如 2001:db8::
                                    net_addr_str = str(ip_network.network_address)
                                    if "::" in net_addr_str:
                                        # 简化的IPv6地址
                                        prefix = net_addr_str
                                    else:
                                        # 完整的IPv6地址，提取前半部分
                                        parts = net_addr_str.split(":")
                                        prefix = ":".join(parts[:4]) + ":"
                                
                                # 预填充前缀到输入框
                                ip_entry.delete(0, tk.END)
                                ip_entry.insert(0, prefix)
                                # 设置光标位置到前缀末尾
                                ip_entry.icursor(len(prefix))
                            except Exception:
                                # 如果网段解析失败，不填充前缀
                                pass
            
            # 如果提供了IP地址，填充对话框
            if ip_address:
                ip_entry.insert(0, ip_address)
                ip_entry.config(state='readonly')  # 设置为只读
                auto_allocate_var.set(True)  # 有IP地址时默认启用自动分配
    
            # 延迟执行耗时操作，提升对话框打开速度
            def delayed_initialization():
                """延迟初始化：自动分配IP和加载历史信息"""
                # 初始化自动分配状态
                if action_type == 'allocate_reserve' and show_ip_input:
                    toggle_auto_allocate()
                
                # 延迟加载IP历史信息（仅当没有传入原始值时）
                if ip_address and not (original_hostname or original_description):
                    try:
                        ip_info = None
                        if record_id:
                            ip_info = self.ipam.get_ip_record_by_id(record_id)
                        
                        if not ip_info:
                            ip_info = self.ipam.get_ip_info(ip_address)
                        
                        if ip_info:
                            hostname_entry.insert(0, ip_info.get('hostname', ''))
                            mac_entry.insert(0, ip_info.get('mac_address', ''))
                            description_entry.insert(0, ip_info.get('description', ''))
                            expiry_date = ip_info.get('expiry_date', '')
                            if expiry_date:
                                try:
                                    from datetime import datetime
                                    expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d")
                                    new_expiry = (expiry_date_obj + timedelta(days=365)).strftime("%Y-%m-%d")
                                    if DateEntry:
                                        expiry_entry.set_date(new_expiry)
                                except Exception:
                                    pass
                    except Exception:
                        pass
        
        # 验证MAC地址格式函数
        def validate_mac_address(mac):
            if not mac:
                return True
            import re
            mac_pattern = re.compile(r'^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$|^[0-9A-Fa-f]{12}$|^([0-9A-Fa-f]{4}-){2}[0-9A-Fa-f]{4}$')
            return bool(mac_pattern.match(mac))
        
        # 主机名输入
        hostname_row = 0 if not show_ip_input else (2 if action_type == 'allocate_reserve' else 1)
        ttk.Label(main_frame, text=_('hostname') + ':').grid(row=hostname_row, column=0, sticky="e", pady=5, padx=(0, 10))
        hostname_border, hostname_entry = create_bordered_entry(main_frame)
        hostname_border.grid(row=hostname_row, column=1, sticky="ew", pady=5, padx=(0, 10))
        
        # MAC地址输入
        mac_row = 1 if not show_ip_input else (3 if action_type == 'allocate_reserve' else 2)
        ttk.Label(main_frame, text=_('mac_address') + ':').grid(row=mac_row, column=0, sticky="e", pady=5, padx=(0, 10))
        mac_border, mac_entry = create_bordered_entry(main_frame)
        mac_border.grid(row=mac_row, column=1, sticky="ew", pady=5, padx=(0, 10))
        
        # MAC地址实时验证
        def on_mac_validate(P):
            """MAC地址实时验证"""
            if not P:  # 空值允许
                mac_entry.config(foreground='black')
                return True
            # 使用统一的验证函数
            if validate_mac_address(P):
                mac_entry.config(foreground='black')
                return True
            else:
                mac_entry.config(foreground='red')
                return True  # 允许输入，但显示红色文字提示
        
        # 创建验证注册
        vcmd = (dialog.register(on_mac_validate), '%P')
        mac_entry.config(validate='key', validatecommand=vcmd)
        
        # 描述输入
        desc_row = 2 if not show_ip_input else (4 if action_type == 'allocate_reserve' else 3)
        ttk.Label(main_frame, text=_('description') + ':').grid(row=desc_row, column=0, sticky="e", pady=5, padx=(0, 10))
        desc_border, description_entry = create_bordered_entry(main_frame)
        desc_border.grid(row=desc_row, column=1, sticky="ew", pady=5, padx=(0, 10))
        
        # 过期日期输入
        expiry_row = 3 if not show_ip_input else (5 if action_type == 'allocate_reserve' else 4)
        ttk.Label(main_frame, text=_('expiry_date') + ':').grid(row=expiry_row, column=0, sticky="e", pady=5, padx=(0, 10))
        
        if DateEntry:
            # 使用DateEntry选择日期
            expiry_entry = DateEntry(main_frame, date_pattern='yyyy-MM-dd')
            expiry_entry.grid(row=expiry_row, column=1, sticky="ew", pady=0, padx=(0, 10))
            # DateEntry初始化时会覆盖textvariable的值，所以需要在创建后设置日期
            expiry_entry.set_date(default_expiry)
            
            # 延迟执行DateEntry修复函数，提升对话框打开速度
            dialog.after(30, lambda: fix_date_entry_for_modal(expiry_entry, dialog))
        else:
            # 使用普通Entry输入日期
            expiry_var = tk.StringVar()
            expiry_var.set(default_expiry)
            expiry_entry = ttk.Entry(main_frame, textvariable=expiry_var)
            expiry_entry.grid(row=expiry_row, column=1, sticky="ew", pady=0, padx=(0, 10))
            ttk.Label(main_frame, text="格式: YYYY-MM-DD").grid(row=expiry_row + 1, column=1, sticky="w", pady=0, padx=(0, 10))
        
        # 如果提供了IP地址，填充原始值（必须在控件创建之后执行）
        if ip_address and (original_hostname or original_description):
            hostname_entry.insert(0, original_hostname)
            description_entry.insert(0, original_description)
        
        # 添加批量和自动复选框（只在分配/保留对话框中显示）
        if action_type == 'allocate_reserve' and show_ip_input:
            # 批量复选框变量
            batch_var = tk.BooleanVar(value=False)  # 默认不选中批量
            
            # 批量模式控件
            batch_mode_frame = ttk.Frame(auto_allocate_frame)
            
            # 数量输入（调整字号）
            ttk.Label(batch_mode_frame, text=_("quantity"), font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=(0, 5))
            quantity_border, quantity_entry = create_bordered_entry(batch_mode_frame, width=5)
            quantity_border.pack(side=tk.LEFT, padx=(0, 10))
            quantity_entry.insert(0, "10")  # 默认值
            
            # 生成模式选择
            generate_mode_var = tk.StringVar(value='sequential')  # 默认连续生成
            ttk.Radiobutton(batch_mode_frame, text=_("sequential"), variable=generate_mode_var, value='sequential').pack(side=tk.LEFT, padx=(0, 10))
            ttk.Radiobutton(batch_mode_frame, text=_("random"), variable=generate_mode_var, value='random').pack(side=tk.LEFT)
            # 添加批量复选框（调整位置，减少缩进）
            batch_check = ttk.Checkbutton(
                auto_allocate_frame, 
                text=_("batch"), 
                variable=batch_var, 
                command=lambda: toggle_batch_mode()
            )
            batch_check.pack(side=tk.LEFT, padx=1)
            
            # 批量模式控件（默认隐藏）
            batch_mode_frame.pack(side=tk.LEFT, padx=10)
            batch_mode_frame.pack_forget()
            
            # 添加自动复选框
            auto_allocate_check = ttk.Checkbutton(
                auto_allocate_frame, 
                text=_("auto"), 
                variable=auto_allocate_var, 
                command=toggle_auto_allocate
            )
            auto_allocate_check.pack(side=tk.LEFT, padx=10)
            # 复选框会根据 auto_allocate_var 的值自动设置为选中状态，无需调用 invoke()
        
        # 按钮框架（放在dialog底部，固定位置）
        button_frame = ttk.Frame(dialog)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=(0, 15))
        
        button_container = ttk.Frame(button_frame)
        button_container.pack(side=tk.RIGHT)
        
        # 启动延迟初始化，提升对话框打开速度
        if show_ip_input and action_type == 'allocate_reserve':
            dialog.after(50, delayed_initialization)
        
        # 批量生成相关函数
        def get_target_network_for_batch():
            """获取批量生成的目标网络"""
            all_networks = self.ipam.get_all_networks()
            selected_network_obj = ipaddress.ip_network(network, strict=False)
            
            child_networks = []
            for net in all_networks:
                try:
                    net_obj = ipaddress.ip_network(net['network'], strict=False)
                    if net_obj.subnet_of(selected_network_obj) and net_obj != selected_network_obj:
                        child_networks.append(net['network'])
                except Exception:
                    pass
            
            prefix_24_plus_networks = []
            for net in child_networks:
                try:
                    net_obj = ipaddress.ip_network(net)
                    if net_obj.prefixlen >= 24:
                        prefix_24_plus_networks.append(net)
                except Exception:
                    pass
            
            target_network = network
            if prefix_24_plus_networks:
                import random
                target_network = random.choice(prefix_24_plus_networks)
            elif child_networks:
                child_networks.sort(key=lambda x: ipaddress.ip_network(x).prefixlen, reverse=True)
                target_network = child_networks[0]
            
            return target_network
        
        def get_ip_placeholder(target_network):
            """生成IP地址占位符"""
            try:
                net_obj = ipaddress.ip_network(target_network)
                ip_parts = str(net_obj.network_address).split('.')
                # 对于IPv4，显示前两位，后两位用*表示
                if len(ip_parts) == 4:
                    return f"{ip_parts[0]}.{ip_parts[1]}.*.*"
                return target_network
            except Exception:
                return target_network
        
        def on_batch_generate(action):
            """批量生成IP地址"""
            try:
                quantity = int(quantity_entry.get().strip())
                if quantity <= 0:
                    self.show_error(_('error'), _('invalid_quantity'))
                    return
                
                # 获取用户输入的主机名和描述
                user_hostname = hostname_entry.get().strip()
                user_description = description_entry.get().strip()
                
                mode = generate_mode_var.get()
                
                target_network = get_target_network_for_batch()
                target_network_obj = ipaddress.ip_network(target_network, strict=False)
                
                ips = self.ipam.get_network_ips(target_network)
                allocated_ips = {ip['ip_address'] for ip in ips}
                
                available_ips = []
                for ip in target_network_obj.hosts():
                    ip_str = str(ip)
                    if ip_str not in allocated_ips:
                        available_ips.append(ip_str)
                
                if len(available_ips) < quantity:
                    self.show_error(_('error'), _('insufficient_available_ips', available=len(available_ips)))
                    return
                
                selected_ips = []
                if mode == 'sequential':
                    selected_ips = available_ips[:quantity]
                else:
                    import random
                    selected_ips = random.sample(available_ips, quantity)
                
                success_count = 0
                for i, ip_str in enumerate(selected_ips, 1):
                    num_str = str(i).zfill(len(str(quantity)))
                    
                    # 根据用户输入生成主机名和描述
                    hostname = ""
                    description = ""
                    
                    if user_hostname:
                        hostname = f"{user_hostname}-{num_str}"
                    
                    if user_description:
                        description = f"{user_description}-{num_str}"
                    elif not user_hostname:
                        # 如果主机名和描述都不填，使用默认的描述
                        description = f"{_('batch_generate')}-{num_str}"
                    
                    if action == 'allocate':
                        if self.ipam.allocate_ip(target_network, ip_str, hostname, description):
                            success_count += 1
                    else:
                        if self.ipam.reserve_ip(target_network, ip_str, hostname, description):
                            success_count += 1
                
                if success_count > 0:
                    self.show_info(_('success'), _('batch_generate_success', count=success_count))
                    self.refresh_ipam_with_selection()
                    dialog.result = False
                    dialog.destroy()
                else:
                    self.show_error(_('error'), _('batch_generate_failed'))
            
            except ValueError:
                self.show_error(_('error'), _('invalid_quantity'))
            except Exception as e:
                self.show_error(_('error'), f"{_('batch_generate_failed')}: {str(e)}")
        
        # 通用的验证和保存函数
        def validate_and_save(action):
            # 获取IP地址
            ip = ip_entry.get().strip() if ip_entry else ''
            hostname = hostname_entry.get().strip()
            mac_address = mac_entry.get().strip()
            description = description_entry.get().strip()
            expiry_date = expiry_entry.get().strip()
            
            # 验证IP地址格式（必须不带前缀）
            if not ip:
                self.show_error(_('error'), _('please_enter_ip_address'))
                return False
            if not self.validate_cidr(ip, require_prefix=False):
                self.show_error(_('error'), _('invalid_ip_address_format'))
                return False
            
            # 使用统一验证规则：主机名和描述不能同时为空
            is_valid, error_msg = IPAMValidator.validate_allocation_params(hostname, description)
            if not is_valid:
                self.show_error(_('error'), error_msg or _('please_enter_description'))
                return False
            
            # 验证MAC地址格式
            if mac_address and not validate_mac_address(mac_address):
                self.show_error(_('error'), _('invalid_mac_address'))
                return False
            
            # 格式化MAC地址
            if mac_address:
                mac_address = self._format_mac_address(mac_address)
            
            # 验证IP地址是否在所选网络范围内
            if ip and network:
                try:
                    # 解析网络和IP地址
                    ip_network = ipaddress.ip_network(network, strict=False)
                    ip_address_obj = ipaddress.ip_address(ip)
                    
                    # 检查IP地址是否在网络范围内
                    if ip_address_obj not in ip_network:
                        # 显示错误提示
                        self.show_error(_('error'), _('ip_not_in_network'))
                        return False
                except Exception as e:
                    # IP地址格式错误或网络解析失败
                    self.show_error(_('error'), _('invalid_ip_address_format'))
                    return False
            
            # 保存输入值
            dialog.ip = ip
            dialog.hostname = hostname
            dialog.mac_address = mac_address
            dialog.description = description
            dialog.expiry_date = expiry_date
            dialog.action = action
            dialog.result = True
            dialog.destroy()
            return True
        
        # 取消按钮
        def on_cancel():
            dialog.result = False
            dialog.destroy()
        
        # 绑定Escape键关闭对话框
        dialog.bind('<Escape>', lambda e: on_cancel())
        
        # 按钮变量
        allocate_btn = None
        reserve_btn = None
        
        # 根据操作类型显示不同的按钮
        if action_type == 'allocate_reserve':
            # 分配/保留对话框：显示"分配地址"和"保留地址"按钮
            reserve_btn = ttk.Button(button_container, text=_('reserve_address'), command=lambda: validate_and_save('reserve'))
            reserve_btn.pack(side=tk.LEFT, padx=(0, 10))
            allocate_btn = ttk.Button(button_container, text=_('allocate_address'), command=lambda: validate_and_save('allocate'))
            allocate_btn.pack(side=tk.LEFT, padx=0)
        else:
            # 其他对话框：显示"确定"和"取消"按钮
            def on_ok():
                validate_and_save('allocate' if action_type == 'allocate' else 'reserve')
            
            ttk.Button(button_container, text=_('ok'), command=on_ok).pack(side=tk.LEFT, padx=(0, 10))
            ttk.Button(button_container, text=_('cancel'), command=on_cancel).pack(side=tk.LEFT, padx=0)
        
        # 批量模式切换函数
        def toggle_batch_mode():
            """切换批量模式"""
            if action_type != 'allocate_reserve' or not show_ip_input:
                return
            
            is_batch = batch_var.get()
            
            if is_batch:
                # 批量模式
                # 隐藏自动分配选项
                auto_allocate_check.pack_forget()
                # 显示批量模式控件
                batch_mode_frame.pack(side=tk.LEFT, padx=10)
                # 禁用IP和MAC地址输入控件，但主机名和描述允许编辑（不修改现有内容）
                ip_entry.config(state='readonly')
                mac_entry.config(state='readonly')
                hostname_entry.config(state='normal')  # 允许编辑
                description_entry.config(state='normal')  # 允许编辑
                # 设置IP地址为占位符
                target_network = get_target_network_for_batch()
                placeholder = get_ip_placeholder(target_network)
                ip_entry.config(state='normal')
                ip_entry.delete(0, tk.END)
                ip_entry.insert(0, placeholder)
                ip_entry.config(state='readonly')
                # 更改按钮文本
                if allocate_btn:
                    allocate_btn.config(text=_('batch_allocate'))
                    allocate_btn.config(command=lambda: on_batch_generate('allocate'))
                if reserve_btn:
                    reserve_btn.config(text=_('batch_reserve'))
                    reserve_btn.config(command=lambda: on_batch_generate('reserve'))
            else:
                # 非批量模式
                # 隐藏批量模式控件
                batch_mode_frame.pack_forget()
                # 显示自动分配选项
                auto_allocate_check.pack(side=tk.LEFT, padx=10)
                # 启用输入控件
                if not auto_allocate_var.get():
                    ip_entry.config(state='normal')
                    # 重新填充前缀（和自动选项关闭时的逻辑一样）
                    if network and not ip_address:
                        try:
                            ip_network = ipaddress.ip_network(network, strict=False)
                            prefix = ""
                            
                            if ip_network.version == 4:
                                # IPv4: 根据前缀长度动态提取网段
                                user_network = network.split('/')[0]  # 提取IP部分，如192.168.1.0
                                octets = user_network.split(".")
                                prefix_len = ip_network.prefixlen
                                
                                # 根据前缀长度计算固定的字节数（整数部分）
                                fixed_octets = prefix_len // 8
                                
                                # 特殊处理：如果固定字节数为0，至少提取1字节
                                fixed_octets = max(1, fixed_octets)
                                
                                # 只提取固定的字节数
                                if fixed_octets == 4:
                                    # /32 子网直接使用完整地址
                                    prefix = user_network
                                else:
                                    # 提取固定数量的网段并添加点号
                                    prefix = ".".join(octets[:fixed_octets]) + "."
                            else:
                                # IPv6: 提取前缀，例如 2001:db8::
                                net_addr_str = str(ip_network.network_address)
                                if "::" in net_addr_str:
                                    # 简化的IPv6地址
                                    prefix = net_addr_str
                                else:
                                    # 完整的IPv6地址，提取前半部分
                                    parts = net_addr_str.split(":")
                                    prefix = ":".join(parts[:4]) + ":"
                            
                            # 预填充前缀到输入框
                            ip_entry.delete(0, tk.END)
                            ip_entry.insert(0, prefix)
                            # 设置光标位置到前缀末尾
                            ip_entry.icursor(len(prefix))
                        except Exception:
                            # 如果网段解析失败，不填充前缀
                            pass
                hostname_entry.config(state='normal')
                mac_entry.config(state='normal')
                description_entry.config(state='normal')
                # 恢复按钮文本
                if allocate_btn:
                    allocate_btn.config(text=_('allocate_address'))
                    allocate_btn.config(command=lambda: validate_and_save('allocate'))
                if reserve_btn:
                    reserve_btn.config(text=_('reserve_address'))
                    reserve_btn.config(command=lambda: validate_and_save('reserve'))
                # 恢复IP地址自动分配
                if auto_allocate_var.get():
                    toggle_auto_allocate()
        
        # 等待对话框关闭
        dialog.wait_window()
        
        # 返回结果
        if hasattr(dialog, 'result') and dialog.result:
            return {
                'ip': getattr(dialog, 'ip', ''),
                'hostname': getattr(dialog, 'hostname', ''),
                'mac_address': getattr(dialog, 'mac_address', ''),
                'description': getattr(dialog, 'description', ''),
                'expiry_date': getattr(dialog, 'expiry_date', ''),
                'action': getattr(dialog, 'action', 'allocate')
            }
        else:
            return None

    def refresh_visualization(self, selected_network=None):
        """刷新可视化内容
        
        Args:
            selected_network: 要显示的网络CIDR，如果为None则使用当前选中的网段
        """
        # 如果没有传入网络参数，尝试从网段表的选中项获取
        if not selected_network:
            selected_items = self.ipam_network_tree.selection()
            if selected_items:
                selected_network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        
        if not selected_network:
            return
        
        # 刷新网络拓扑图
        # 从IPAM中获取实际的子网数据
        network_tree = {}
        try:
            parent_network = ipaddress.ip_network(selected_network)
            
            # 从IPAM获取所有网络
            all_networks = self.ipam.get_all_networks()
            
            # 筛选出属于当前父网络的子网，并按前缀长度排序（长在前，短在后）
            relevant_networks = []
            for network in all_networks:
                try:
                    subnet = ipaddress.ip_network(network['network'])
                    # 检查是否是父网络的子网
                    if subnet.subnet_of(parent_network):
                        relevant_networks.append({
                            "name": network.get('description', f"子网{len(relevant_networks) + 1}"),
                            "cidr": network['network'],
                            "network_obj": subnet,
                            # 保留完整原始数据用于智能颜色分配
                            "raw_data": network
                        })
                except Exception:
                    pass
            
            # 按前缀长度排序，长的在前（子网在前）
            relevant_networks.sort(key=lambda x: x['network_obj'].prefixlen, reverse=True)
            
            # 构建网络树结构
            network_tree = {}
            for net in relevant_networks:
                cidr = net['cidr']
                network_tree[cidr] = {
                    "name": net['name'],
                    "cidr": cidr,
                    "children": [],
                    # 保留原始数据用于类型推断
                    "raw_data": net.get('raw_data', {})
                }
            
            # 建立父子关系
            for net in relevant_networks:
                current_cidr = net['cidr']
                current_subnet = net['network_obj']
                
                # 找到直接父网络
                parent_cidr = None
                max_prefix_len = -1
                
                for other_cidr, other_net in network_tree.items():
                    if other_cidr == current_cidr:
                        continue
                    try:
                        other_subnet = ipaddress.ip_network(other_cidr)
                        if current_subnet.subnet_of(other_subnet):
                            if other_subnet.prefixlen > max_prefix_len:
                                max_prefix_len = other_subnet.prefixlen
                                parent_cidr = other_cidr
                    except Exception:
                        pass
                
                if parent_cidr:
                    network_tree[parent_cidr]['children'].append(current_cidr)
            
        except Exception as e:
            print(_("subnet_error", e=str(e)))
        
        # 生成网络数据
        network_data = []
        node_id_counter = 1
        
        # 递归生成网络数据
        def generate_network_data(cidr, level=0, parent_id=None):
            nonlocal node_id_counter
            
            net_info = network_tree.get(cidr)
            if not net_info:
                return None
            
            # 获取IP统计信息
            ip_info = {"total": 0, "allocated": 0, "reserved": 0, "available": 0, "registered": 0, "network_total": 0}
            try:
                subnet_network = ipaddress.ip_network(cidr)
                network_total = 2 ** (32 - subnet_network.prefixlen) if subnet_network.version == 4 else 2 ** (128 - subnet_network.prefixlen)
                
                # 获取子网及其所有子网络的IP地址列表
                subnet_ips = self.ipam.get_network_ips(cidr)
                # 只计算已分配和已保留的IP地址数量
                allocated = sum(1 for ip in subnet_ips if ip.get('status') == 'allocated')
                reserved = sum(1 for ip in subnet_ips if ip.get('status') == 'reserved')
                # 计算总注册IP数量（已分配 + 已保留）
                registered = allocated + reserved
                # 计算可用IP数量
                available = sum(1 for ip in subnet_ips if ip.get('status') == 'released')
                
                ip_info = {
                    "total": network_total,
                    "allocated": allocated,
                    "reserved": reserved,
                    "available": available,
                    "registered": registered,
                    "network_total": network_total
                }
            except Exception as e:
                print(_("ip_statistics_error", e=str(e)))
            
            node_id = f"node_{node_id_counter}"
            node_id_counter += 1
            
            # 生成子节点
            children = []
            for child_cidr in net_info['children']:
                child_node = generate_network_data(child_cidr, level + 1, node_id)
                if child_node:
                    children.append(child_node['id'])
                    network_data.append(child_node)
            
            # 根据网络名称和层级设置不同的设备类型
            device_type = "router" if level == 0 else "switch"
            
            # 智能推断 subnet_type（决定颜色）
            subnet_type = "default"
            name_lower = net_info['name'].lower()
            cidr_lower = cidr.lower()
            
            # 根据网络名称关键词判断类型（多语言支持）
            from i18n import get_language
            current_lang = get_language()
            
            # 设备类型和子网类型映射
            type_mapping = {
                "server": ("server", "server"),          # 暖橙色
                "management": ("client", "management"),    # 柔和紫色
                "wireless": ("wireless", "wireless"),      # 深蓝色
                "client": ("switch", "client"),          # 青绿色
                "office": ("office", "office"),          # 绿色
                "production": ("production", "production"),  # 橙色
                "test": ("test", "test"),              # 紫色
                "dmz": ("dmz", "dmz"),              # 玫红色
                "storage": ("storage", "storage"),        # 橙红色
                "backup": ("backup", "backup"),         # 浅绿色
                "iot": ("iot", "iot"),               # 浅蓝色
                "voip": ("voip", "voip"),             # 深蓝色
                "network": ("switch", "network")        # 柔和橙色
            }
            
            # 从翻译文件获取关键词
            from i18n import translator
            keywords_config = translator.translations.get('network_type_keywords', {})
            lang_keywords = keywords_config.get(current_lang, {})
            
            # 遍历关键词进行匹配
            matched_type = None
            for type_name, type_keywords in lang_keywords.items():
                for keyword in type_keywords:
                    if keyword.lower() in name_lower:
                        matched_type = type_name
                        break
                if matched_type:
                    break
            
            # 如果匹配到类型，设置对应的设备类型和子网类型
            if matched_type and matched_type in type_mapping:
                device_type, subnet_type = type_mapping[matched_type]
            # 根据网络层级设置设备类型和子网类型（如果没有关键词匹配）
            elif level > 0:
                if level == 1:
                    device_type = "switch"  # 一级子节点使用椭圆
                    subnet_type = "network"  # 柔和橙色
                elif level == 2:
                    device_type = "switch2"  # 二级子节点使用圆角矩形
                    subnet_type = "management"  # 柔和紫色
                elif level == 3:
                    device_type = "switch3"  # 三级子节点使用矩形
                    subnet_type = "office"  # 绿色
                else:
                    device_type = "switch3"  # 四级及以上子节点使用矩形
                    subnet_type = "production"  # 橙色
            # 根据 CIDR 前缀长度进一步区分（更细的层级）
            else:
                try:
                    prefix_len = ipaddress.ip_network(cidr).prefixlen
                    if prefix_len <= 8:  # 超大网段（如 10.0.0.0/8）
                        subnet_type = "extra_large"  # 红色
                    elif prefix_len <= 16:  # 大网段（如 172.16.0.0/12, 192.168.0.0/16）
                        subnet_type = "large"  # 亮青色
                    elif prefix_len <= 20:  # 中等网段（/17-20）
                        subnet_type = "medium"  # 深蓝色
                    elif prefix_len <= 24:  # 中等网段（/21-24）
                        subnet_type = "default"  # 主蓝色
                    else:  # 小网段（/25+）
                        subnet_type = "small"  # 紫色
                except Exception:
                    pass
            
            return {
                "id": node_id,
                "name": net_info['name'],
                "cidr": cidr,
                "level": level,
                "parent_id": parent_id,  # 添加父节点 ID
                "type": "network" if level == 0 else "client",
                "device_type": device_type,
                "subnet_type": subnet_type,
                "ip_info": ip_info,
                "children": children
            }
        
        # 生成根节点
        root_node = generate_network_data(selected_network, 0)
        if root_node:
            network_data.insert(0, root_node)
        
        self.topology_visualizer.draw_topology(network_data)
    
    def reserve_ip(self):
        """保留IP地址"""
        selected_items = self.ipam_network_tree.selection()
        if not selected_items:
            self.show_error(_('error'), _('please_select_network'))
            return
        
        network = self.ipam_network_tree.item(selected_items[0], 'values')[0]
        
        # 显示IP地址保留对话框，传递network参数以自动填充前缀
        dialog_result = self.show_ip_address_dialog(_('reserve_address'), 'reserve', network=network)
        if not dialog_result:
            return
        
        ip_address = dialog_result['ip']
        hostname = dialog_result['hostname']
        description = dialog_result['description']
        
        if not ip_address:
            self.show_info(_('hint'), _('please_enter_ip_address'))
            return
        
        success, message = self.ipam.reserve_ip(network, ip_address, hostname, description)
        if success:
            self.show_info(_('success'), message)
            # 刷新IPAM数据并恢复选中状态
            self.refresh_ipam_with_selection()
            # 清空输入框
            self.ipam_ip_entry.delete(0, tk.END)
            self.ipam_ip_description_entry.delete(0, tk.END)
        else:
            self.show_error(_('error'), message)


def load_application(splash, root):
    """加载应用程序模块
    
    Args:
        splash: 启动画面实例
        root: 主窗口实例
    """
    import time
    from window_utils import setup_window_settings
    from config_manager import get_config
    
    # 定义加载模块列表
    modules = [
        (_("config_file"), 0.8),
        (_("i18n_resources"), 0.6),
        (_("database_init"), 1.0),
        (_("ui_components"), 0.8),
        (_("network_services"), 0.8),
        (_("plugin_system"), 0.6),
        (_("style_system"), 0.6),
        (_("history"), 0.6),
        (_("user_settings"), 0.6),
        (_("app_instance"), 1.0),
        (_("loading_complete"), 1.2)
    ]
    
    print(_("start_loading"))
    
    # 加载配置
    try:
        config = get_config()
        print(_("config_loaded"))
    except Exception as e:
        print(f"Error loading config: {str(e)}")
    
    # 模拟加载过程
    for i, (module, delay) in enumerate(modules):
        print(_("loading_module", module=module))
        
        # 更新启动画面状态
        splash.update_progress(module)
        
        # 等待一段时间，模拟加载过程
        start_time = time.time()
        while time.time() - start_time < delay:
            # 保持主循环运行，确保启动画面能够更新
            try:
                root.update()
                splash.splash.update()
            except Exception as e:
                print(f"Error updating window: {str(e)}")
            time.sleep(0.1)
    
    print(_("loading_process_complete"))


if __name__ == "__main__":
    # 导入窗口工具模块
    from window_utils import setup_window_settings
    from config_manager import get_config
    
    # 创建主窗口
    root = tk.Tk()
    
    # 立即隐藏窗口，避免闪现
    root.withdraw()
    
    from version import get_version
    version = get_version()
    root.title(f"子网规划师 v{version}")
    
    # 设置窗口图标
    icon_path = os.path.join(os.path.dirname(__file__), "Subnet_Planner.ico")
    if os.path.exists(icon_path):
        try:
            root.iconbitmap(icon_path)
        except Exception as e:
            print(f"Error setting icon: {str(e)}")
    
    # 显示启动画面
    splash = SplashScreen()
    
    # 执行加载
    load_application(splash, root)
    
    # 设置双击间隔
    try:
        double_click_time = root.tk.call('tk', 'getDoubleClickTime')
    except Exception:
        double_click_time = 500
    
    try:
        root.tk.call('tk', 'setDoubleClickTime', double_click_time)
    except Exception:
        pass
    
    # 设置窗口初始大小
    BASE_WIDTH = 1050
    BASE_HEIGHT = 950
    WINDOW_WIDTH = BASE_WIDTH
    WINDOW_HEIGHT = BASE_HEIGHT
    
    # 调用窗口设置函数
    setup_window_settings(
        root, 
        WINDOW_WIDTH, 
        WINDOW_HEIGHT, 
        lock_width=False, 
        min_width=BASE_WIDTH, 
        min_height=BASE_HEIGHT, 
        max_width=10000, 
        max_height=10000,
        position=None
    )
    
    # 创建应用实例
    from windows_app import SubnetPlannerApp
    app = SubnetPlannerApp(root)
    
    # 关闭启动画面
    splash.close()
    
    # 显示主窗口
    root.deiconify()
    
    # 设置主窗口关闭事件处理程序
    def on_main_window_close():
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_main_window_close)
    
    # 运行应用
    root.mainloop()
