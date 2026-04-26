from exporters.base import DataExporter
from exporters.data_preparer import DataPreparer
from i18n import _ as translate


class ExcelExporter(DataExporter):
    def export(self, file_path, data_source, main_data, main_headers, remaining_data, remaining_headers):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        remaining_data_list = DataPreparer.convert_remaining_data(remaining_data, remaining_headers)
        mock_tree = DataPreparer.create_mock_tree(remaining_headers, remaining_data_list)

        wb = Workbook()

        parent_cidr = self._get_parent_cidr(data_source)

        main_sheet = wb.active
        main_title = translate("split_segment_info") if data_source["main_name"] == translate("split_segment_info") else translate("subnet_requirements")
        main_sheet.title = str(main_title) if main_title else "Sheet"

        if parent_cidr:
            main_sheet.cell(row=1, column=1, value=translate("parent_network"))
            main_sheet.cell(row=1, column=2, value=parent_cidr)
            data_start_row = 3
        else:
            data_start_row = 1

        for col_idx, header in enumerate(main_headers, 1):
            cell = main_sheet.cell(row=data_start_row, column=col_idx, value=header)
            if cell:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        for row_idx, values in enumerate(main_data, data_start_row + 1):
            for col_idx, value in enumerate(values, 1):
                main_sheet.cell(row=row_idx, column=col_idx, value=value)

        remaining_sheet = wb.create_sheet(title=str(translate("remaining_subnets")) if translate("remaining_subnets") else "Remaining")

        if parent_cidr:
            remaining_sheet.cell(row=1, column=1, value=translate("parent_network"))
            remaining_sheet.cell(row=1, column=2, value=parent_cidr)
            remaining_data_start_row = 3
        else:
            remaining_data_start_row = 1

        for col_idx, header in enumerate(remaining_headers, 1):
            cell = remaining_sheet.cell(row=remaining_data_start_row, column=col_idx, value=header)
            if cell:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(mock_tree.get_children(), remaining_data_start_row + 1):
            values = mock_tree.item(item, "values")
            for col_idx, value in enumerate(values, 1):
                remaining_sheet.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)

    def _get_parent_cidr(self, data_source):
        chart_data = data_source.get("chart_data")
        if chart_data and "parent" in chart_data:
            return chart_data["parent"].get("name", "")
        return ""

    def get_file_extension(self) -> str:
        return ".xlsx"
