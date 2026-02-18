#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据导出工具模块

提供多种格式的数据导出功能，包括:
1. PDF格式导出（支持中文）
2. JSON格式导出
3. TXT格式导出
4. CSV格式导出
5. Excel格式导出

所有导出功能均正确处理中文编码问题。
"""

import os
import json
import csv
import time
import traceback
import math
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
from version import get_version
from i18n import _ as translate, get_language  # type: ignore

__version__ = get_version()
from reportlab.lib.pagesizes import A4, landscape  # type: ignore
from reportlab.lib.units import cm  # type: ignore
from reportlab.lib import colors  # type: ignore
from reportlab.platypus import (  # type: ignore
    BaseDocTemplate,
    PageTemplate,
    Frame,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage,
    PageBreak,
    NextPageTemplate,
    KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
from reportlab.lib.enums import TA_CENTER, TA_LEFT  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, Alignment  # type: ignore
from font_config import FontConfig  # type: ignore

MAIN_TABLE_COLORS = {
    "header_bg": "#3498db",
    "header_text": "white",
    "box": "#3498db",
    "row_even": "#f0f4f8",
}

REMAINING_TABLE_COLORS = {
    "header_bg": "#27ae60",
    "header_text": "white",
    "box": "#27ae60",
    "row_even": "#f0f4f8",
}

TABLE_COMMON_STYLE = [
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("GRID", (0, 0), (-1, -1), 1, "#bdc3c7"),
    ("BACKGROUND", (0, 1), (-1, -1), "#f8f9fa"),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), ["white", "#f0f4f8"]),
    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ("LEFTPADDING", (0, 0), (-1, -1), 2),
    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ("TOPPADDING", (0, 0), (-1, -1), 2),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
]

HEADER_STYLE = [
    ("FONTSIZE", (0, 0), (-1, 0), 11),
    ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
    ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
    ("TOPPADDING", (0, 0), (-1, 0), 4),
]

K2V_HEADERS = [translate("item"), translate("value")]


class ExportUtils:
    """数据导出工具类"""

    # 字体缓存（类变量）
    _font_cache: dict[str, str] = {}
    _font_path_cache: str | None = None
    _font_path_lang: str | None = None
    
    def format_large_number(self, num, use_scientific=True):
        """格式化大数值，可选择使用科学计数法或千位分隔符
        
        Args:
            num: 要格式化的数值
            use_scientific: 是否使用科学计数法，默认为True
            
        Returns:
            str: 格式化后的字符串
        """
        # 导入ip_subnet_calculator中的format_large_number函数，统一使用一个实现
        from ip_subnet_calculator import format_large_number
        return format_large_number(num, use_scientific)

    @classmethod
    def clear_font_cache(cls) -> None:
        """清除字体缓存，在语言切换时调用"""
        cls._font_cache.clear()
        cls._font_path_cache = None
        cls._font_path_lang = None

        # 同时清除 ReportLab 的 PDF 字体注册
        try:
            from reportlab.pdfbase import pdfmetrics
            if "ChineseFont" in pdfmetrics.getRegisteredFontNames():
                if hasattr(pdfmetrics, '_fonts') and "ChineseFont" in pdfmetrics._fonts:  # type: ignore[attr-defined]
                    del pdfmetrics._fonts["ChineseFont"]  # type: ignore[attr-defined]
                    print("🧹 已清除 ReportLab PDF 字体注册")
        except Exception as e:
            print(f"⚠️ 清除 PDF 字体注册时出现警告: {e}")

        print("🧹 已清除字体缓存")

    def __init__(self) -> None:
        """初始化导出工具"""
        self.has_asian_font: bool = False
        self._register_asian_fonts()

    def _get_prioritized_font_candidates(self, current_lang: str) -> list[tuple[str, str]]:
        """根据当前语言获取优先级排序的字体候选列表

        Args:
            current_lang: 当前语言代码

        Returns:
            list: 优先级排序的字体候选列表
        """
        return FontConfig.get_font_candidates(current_lang)

    def _calculate_text_width(self, text: str, ascii_width: int = 8, non_ascii_width: int = 12, padding: int = 15) -> int:
        """计算文本宽度

        Args:
            text: 要计算的文本
            ascii_width: ASCII字符宽度
            non_ascii_width: 非ASCII字符宽度
            padding: 额外边距

        Returns:
            int: 文本宽度
        """
        text_width = 0
        for char in text:
            if ord(char) > 127:
                text_width += non_ascii_width
            else:
                text_width += ascii_width
        text_width += padding
        return text_width

    def _get_cell_text(self, cell: object) -> str:
        """获取单元格的文本内容

        Args:
            cell: 单元格对象

        Returns:
            str: 文本内容
        """
        if hasattr(cell, 'getPlainText'):
            return cell.getPlainText()  # type: ignore[attr-defined]
        elif hasattr(cell, 'text'):
            return cell.text  # type: ignore[attr-defined]
        else:
            return str(cell)

    def _adjust_table_font_size(self, table_data: list[list[object]], col_widths: list[float], table_text_style: ParagraphStyle, min_font_size: int = 8, style_prefix: str = "CustomStyle") -> list[list[object]]:
        """调整表格字体大小以适应列宽

        Args:
            table_data: 表格数据
            col_widths: 列宽列表
            table_text_style: 基础文本样式
            min_font_size: 最小字体大小
            style_prefix: 样式名称前缀

        Returns:
            list: 调整后的表格数据
        """
        adjusted_data = []
        for row_idx, row in enumerate(table_data):
            adjusted_row = []
            # 只处理与col_widths长度一致的列，避免IndexError
            for col_idx, cell in enumerate(row[:len(col_widths)]):
                text = self._get_cell_text(cell)
                text_width = self._calculate_text_width(text, ascii_width=5, non_ascii_width=10, padding=8)

                font_size = 11 if row_idx == 0 else 10
                if text_width > col_widths[col_idx] * 0.9:
                    scale_factor = (col_widths[col_idx] * 0.95) / text_width
                    font_size = max(font_size * scale_factor, min_font_size)

                custom_style = ParagraphStyle(
                    f"{style_prefix}_{row_idx}_{col_idx}",
                    parent=table_text_style,
                    fontSize=font_size,
                    wordWrap="None"
                )
                adjusted_cell = Paragraph(text, custom_style)
                adjusted_row.append(adjusted_cell)
            adjusted_data.append(adjusted_row)
        return adjusted_data

    def _register_asian_fonts(self):
        """注册亚洲字体（中文、韩语、日语）供PDF导出使用，使用缓存避免重复注册"""
        # 注意：字体注册已移至 _export_to_pdf 方法中，以支持语言切换
        # 这里只初始化 has_asian_font 标志
        from i18n import get_language
        current_lang = get_language()
        
        # 检查是否已经注册过相同语言的字体
        if ExportUtils._font_path_lang == current_lang and ExportUtils._font_path_cache:
            print(f"🔍 [init] 使用已注册的字体缓存 (语言: {current_lang})")
            self.has_asian_font = True
        else:
            # 语言不同或没有缓存，字体将在 _export_to_pdf 中注册
            print(f"🔍 [init] 字体将在导出PDF时注册 (语言: {current_lang})")
            self.has_asian_font = False

    def _calculate_auto_col_widths(self, table_data, table_width):
        """根据内容计算自适应列宽

        Args:
            table_data: 表格数据，包含Paragraph对象的列表
            table_width: 表格可用宽度

        Returns:
            list: 每列的自适应宽度
        """
        table_cols = len(table_data[0]) if table_data else 0
        if table_cols == 0:
            return []

        max_col_widths = [0] * table_cols
        min_col_width = 60  # 调整最小列宽

        # 计算每列的最大内容宽度
        for row in table_data:
            # 只处理与表头列数一致的数据列，避免IndexError
            for col_idx, cell in enumerate(row[:table_cols]):
                text = self._get_cell_text(cell)
                text_width = self._calculate_text_width(text, ascii_width=8, non_ascii_width=12, padding=15)
                if text_width > max_col_widths[col_idx]:
                    max_col_widths[col_idx] = text_width

        # 确保最小列宽
        for i, width in enumerate(max_col_widths):
            if width < min_col_width:
                max_col_widths[i] = min_col_width

        # 计算总宽度和比例
        total_width = sum(max_col_widths)
        final_widths = []

        # 收集所有数据行的内容宽度
        data_col_widths = [0] * table_cols
        for row_idx, row in enumerate(table_data):
            if row_idx == 0:  # 跳过表头，只考虑数据行
                continue
            # 只处理与表头列数一致的数据列，避免IndexError
            for col_idx, cell in enumerate(row[:table_cols]):
                text = self._get_cell_text(cell)
                text_width = self._calculate_text_width(text, ascii_width=8, non_ascii_width=12, padding=15)
                if text_width > data_col_widths[col_idx]:
                    data_col_widths[col_idx] = text_width

        if total_width > 0:
            # 结合表头和数据行的宽度
            combined_widths = [max(max_col_widths[i], data_col_widths[i]) for i in range(table_cols)]

            # 先分配最小宽度
            remaining_width = table_width
            for i in range(table_cols):
                final_widths.append(min_col_width)
                remaining_width -= min_col_width

            # 按比例分配额外宽度
            if remaining_width > 0:
                # 计算各列超出最小宽度的部分
                extra_widths = [max(0, combined_widths[i] - min_col_width) for i in range(table_cols)]
                total_extra = sum(extra_widths)

                if total_extra > 0:
                    # 按比例分配额外宽度
                    for i in range(table_cols):
                        if extra_widths[i] > 0:
                            extra = (extra_widths[i] / total_extra) * remaining_width
                            final_widths[i] += extra
                        # 对于数据行内容很少的列，限制最大宽度
                        if data_col_widths[i] < min_col_width * 1.5:
                            final_widths[i] = min(final_widths[i], min_col_width * 1.5)
        else:
            # 平均分配宽度
            final_widths = [table_width / table_cols] * table_cols

        return final_widths

    def _is_k2v_headers(self, headers):
        # 更灵活的判断：如果只有两列，就认为是键值对格式
        return len(headers) == 2

    def _get_col_widths(self, table_data, table_width, col_widths, num_cols):
        """获取表格列宽，确保自适应页面宽度"""
        try:
            # 始终使用自适应宽度计算
            auto_col_widths = self._calculate_auto_col_widths(table_data, table_width)
            col_widths = auto_col_widths
        except (ValueError, TypeError, AttributeError):
            traceback.print_exc()
            # 回退到平均分配宽度
            col_widths = [table_width / num_cols] * num_cols

        # 确保列宽有效
        valid = []
        for width in col_widths:
            if width is None or not isinstance(width, (int, float)) or width <= 0:
                valid.append(100)
            else:
                valid.append(width)

        return valid

    def _get_table_style(self, table_colors, has_asian_font):
        header_font = "ChineseFont" if has_asian_font else "Helvetica-Bold"
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), table_colors["header_bg"]),
            ("TEXTCOLOR", (0, 0), (-1, 0), table_colors["header_text"]),
            ("FONTNAME", (0, 0), (-1, 0), header_font),
            # 使用GRID代替BOX，避免重复线条
            ("GRID", (0, 0), (-1, -1), 1, "#bdc3c7"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]
        # 只添加必要的样式，避免与TABLE_COMMON_STYLE冲突
        return TableStyle(style + HEADER_STYLE + [("WORDWRAP", (0, 0), (-1, -1), False)])

    def _prepare_export_data(self, data_source):
        """准备导出数据

        Args:
            data_source: 字典，包含导出数据的源信息

        Returns:
            tuple: (main_data, main_headers, remaining_data, remaining_headers)
        """
        main_data = []
        main_tree = data_source["main_tree"]
        main_filter = data_source.get("main_filter", None)
        main_headers = data_source.get("main_headers")

        # 检测IP版本：通过检查列宽度或IP版本变量
        is_ipv6 = False
        
        # 方法1：检查是否有IP版本变量
        ip_version = data_source.get("ip_version")
        if ip_version:
            is_ipv6 = ip_version == "IPv6"
        else:
            # 方法2：检查是否有end_address列且宽度大于0
            for col in main_tree["columns"]:
                col_width = main_tree.column(col, "width")
                if col == "end_address" and col_width > 0:
                    is_ipv6 = True
                    break
            
            # 方法3：如果主表格没有end_address列，检查剩余表格
            if not is_ipv6 and "remaining_tree" in data_source:
                remaining_tree = data_source["remaining_tree"]
                for col in remaining_tree["columns"]:
                    col_width = remaining_tree.column(col, "width")
                    if col == "end_address" and col_width > 0:
                        is_ipv6 = True
                        break

        # 获取主表格的所有列名
        main_columns = main_tree["columns"]
        
        # 根据IP版本过滤主表格字段
        if is_ipv6:
            # IPv6：去掉子网掩码、通配符掩码、广播地址
            main_field_filter = lambda h: h not in [translate("subnet_mask"), translate("wildcard_mask"), translate("broadcast_address")]
        else:
            # IPv4：去掉网段结束地址
            main_field_filter = lambda h: h != translate("network_end_address")

        # 过滤主表格的标题
        if main_headers is None:
            all_main_headers = [main_tree.heading(col, "text") or "" for col in main_columns]
            # 根据IP版本过滤标题
            main_headers = [h for h in all_main_headers if main_field_filter(h)]
        else:
            # 如果提供了main_headers，也进行过滤
            main_headers = [h for h in main_headers if main_field_filter(h)]

        # 过滤主表格的列
        filtered_main_columns = []
        for i, col in enumerate(main_columns):
            header = main_tree.heading(col, "text") or ""
            if main_field_filter(header):
                filtered_main_columns.append(i)

        added_items = set()
        for item in main_tree.get_children():
            values = main_tree.item(item, "values")
            if main_filter:
                if main_filter(values):
                    if len(values) >= 2 and values[0] != "":
                        item_key = values[0]
                        if item_key not in added_items:
                            added_items.add(item_key)
                            # 过滤值列表
                            filtered_values = [values[i] for i in filtered_main_columns]
                            main_data.append(filtered_values)
                    else:
                        # 过滤值列表
                        filtered_values = [values[i] for i in filtered_main_columns]
                        main_data.append(filtered_values)
            elif values:
                if len(values) >= 2 and values[0] != "":
                    item_key = values[0]
                    if item_key not in added_items:
                        added_items.add(item_key)
                        # 过滤值列表
                        filtered_values = [values[i] for i in filtered_main_columns]
                        main_data.append(filtered_values)
                else:
                    # 过滤值列表
                    filtered_values = [values[i] for i in filtered_main_columns]
                    main_data.append(filtered_values)

        unique_main_data = []
        seen_rows = set()
        for row in main_data:
            row_tuple = tuple(row)
            if row_tuple not in seen_rows:
                seen_rows.add(row_tuple)
                unique_main_data.append(row)
        main_data = unique_main_data

        remaining_tree = data_source["remaining_tree"]
        remaining_columns = remaining_tree["columns"]
        remaining_headers = []
        
        # 根据IP版本过滤剩余表格列
        filtered_columns = []
        for col in remaining_columns:
            col_width = remaining_tree.column(col, "width")
            # 只保留显示的列（宽度大于0）
            if col_width > 0:
                header = remaining_tree.heading(col, "text") or ""
                if main_field_filter(header):
                    filtered_columns.append(col)
                    remaining_headers.append(header)
        
        remaining_data = []
        for item in remaining_tree.get_children():
            values = remaining_tree.item(item, "values")
            if values:
                # 创建字典，只包含过滤后的列
                filtered_values = []
                for i, col in enumerate(remaining_columns):
                    if col in filtered_columns:
                        filtered_values.append(values[i])
                # 创建字典，使用过滤后的列名和值
                filtered_dict = dict(zip(remaining_headers, filtered_values))
                remaining_data.append(filtered_dict)

        return main_data, main_headers, remaining_data, remaining_headers

    def _export_to_json(self, file_path, data_source, main_data, main_headers, remaining_data):
        """导出数据为JSON格式（UTF-8编码，不转义中文）"""
        if data_source["main_name"] == translate("split_segment_info"):
            export_data = {
                translate("split_segment_info"): dict(main_data), 
                translate("remaining_subnets"): remaining_data
            }
        else:
            export_data = {
                f"{data_source['main_name']}": [dict(zip(main_headers, item)) for item in main_data],
                
                translate("remaining_subnets"): remaining_data,
            }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)

    def _export_to_txt(self, file_path, data_source, main_data, main_headers, remaining_data=None, remaining_headers=None):
        """导出数据为TXT格式（UTF-8编码）"""
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(f"{data_source['main_name']}\n")
            f.write("=" * 80 + "\n")

            if self._is_k2v_headers(main_headers):
                # 计算项目名+冒号的最大宽度
                max_key_colon_width = 0
                for values in main_data:
                    if len(values) > 0:
                        key_colon_width = len(f"{str(values[0])}:")
                        if key_colon_width > max_key_colon_width:
                            max_key_colon_width = key_colon_width
                
                # 计算需要的Tab数量，确保所有值对齐在同一列（每个Tab对应8个字符宽度）
                tab_width = 8
                # 计算最大宽度需要的Tab数
                max_tabs = (max_key_colon_width + tab_width - 1) // tab_width
                
                # 特殊字段处理：需要少一个Tab的字段列表
                reduced_tab_fields = ["通配符掩码", "可用地址数"]
                
                for values in main_data:
                    key = str(values[0])
                    value = str(values[1])
                    key_colon = f"{key}:"
                    # 计算当前项目名+冒号需要的Tab数
                    current_tabs = (len(key_colon) + tab_width - 1) // tab_width
                    # 需要的额外Tab数 = 最大Tab数 - 当前Tab数 + 1（+1确保有足够的间距）
                    extra_tabs = max_tabs - current_tabs + 1
                    
                    # 特殊处理：通配符掩码和可用地址数少一个Tab
                    if key in reduced_tab_fields:
                        extra_tabs = max(1, extra_tabs - 1)  # 确保至少有一个Tab
                    
                    f.write(f"{key_colon}{'\t' * extra_tabs}{value}\n")
            else:
                # 计算每个列的最大宽度，使用自适应宽度
                max_widths = []
                # 先处理表头
                for i, header in enumerate(main_headers):
                    max_widths.append(len(str(header)) + 2)  # 表头宽度+2边距
                
                # 再处理数据行
                for values in main_data:
                    for i, value in enumerate(values):
                        if i < len(max_widths):
                            current_width = len(str(value)) + 2  # 数据宽度+2边距
                            if current_width > max_widths[i]:
                                max_widths[i] = current_width
                
                # 确保最小宽度为12
                max_widths = [max(w, 12) for w in max_widths]
                
                # 写入表头
                for i, header in enumerate(main_headers):
                    f.write(f"{header:<{max_widths[i]}}")
                f.write("\n")
                
                # 写入分隔线
                total_width = sum(max_widths)
                f.write("-" * total_width + "\n")
                
                # 写入数据行
                for values in main_data:
                    for i, value in enumerate(values):
                        if i < len(max_widths):
                            f.write(f"{str(value):<{max_widths[i]}}")
                    f.write("\n")

            f.write(f"\n\n{data_source['remaining_name']}\n")
            f.write("=" * 80 + "\n")

            if remaining_headers and remaining_data:
                # 使用传入的剩余数据表头和数据
                # 收集所有剩余数据行
                all_remaining_rows = []
                for item in remaining_data:
                    if isinstance(item, dict):
                        # 如果是字典，按照剩余表头的顺序提取值
                        values = [item.get(header, '') for header in remaining_headers]
                    else:
                        # 如果已经是列表，直接使用
                        values = item
                    all_remaining_rows.append(values)
                
                # 计算每个列的最大宽度，使用自适应宽度
                max_widths = []
                # 先处理表头
                for i, header in enumerate(remaining_headers):
                    max_widths.append(len(str(header)) + 2)  # 表头宽度+2边距
                
                # 再处理数据行
                for values in all_remaining_rows:
                    for i, value in enumerate(values):
                        if i < len(max_widths):
                            current_width = len(str(value)) + 2  # 数据宽度+2边距
                            if current_width > max_widths[i]:
                                max_widths[i] = current_width
                
                # 确保最小宽度为12
                max_widths = [max(w, 12) for w in max_widths]
                
                # 写入表头
                for i, header in enumerate(remaining_headers):
                    f.write(f"{header:<{max_widths[i]}}")
                f.write("\n")
                
                # 写入分隔线
                total_width = sum(max_widths)
                f.write("-" * total_width + "\n")
                
                # 写入数据行
                for values in all_remaining_rows:
                    for i, value in enumerate(values):
                        if i < len(max_widths):
                            f.write(f"{str(value):<{max_widths[i]}}")
                    f.write("\n")
            else:
                # 动态获取剩余数据的表头（备用方案）
                remaining_headers = [data_source["remaining_tree"].heading(col, "text") or ""
                                   for col in data_source["remaining_tree"]["columns"]]
                
                # 收集所有剩余数据行
                all_remaining_rows = []
                for item in data_source["remaining_tree"].get_children():
                    values = data_source["remaining_tree"].item(item, "values")
                    all_remaining_rows.append(values)
                
                # 计算每个列的最大宽度，使用自适应宽度
                max_widths = []
                # 先处理表头
                for i, header in enumerate(remaining_headers):
                    max_widths.append(len(str(header)) + 2)  # 表头宽度+2边距
                
                # 再处理数据行
                for values in all_remaining_rows:
                    for i, value in enumerate(values):
                        if i < len(max_widths):
                            current_width = len(str(value)) + 2  # 数据宽度+2边距
                            if current_width > max_widths[i]:
                                max_widths[i] = current_width
                
                # 确保最小宽度为12
                max_widths = [max(w, 12) for w in max_widths]
                
                # 写入表头
                for i, header in enumerate(remaining_headers):
                    f.write(f"{header:<{max_widths[i]}}")
                f.write("\n")
                
                # 写入分隔线
                total_width = sum(max_widths)
                f.write("-" * total_width + "\n")
                
                # 写入数据行
                for values in all_remaining_rows:
                    for i, value in enumerate(values):
                        if i < len(max_widths):
                            f.write(f"{str(value):<{max_widths[i]}}")
                    f.write("\n")

    def _export_to_csv(self, file_path, main_data, main_headers, remaining_tree, remaining_headers=None):
        """导出数据为CSV格式（UTF-8-BOM编码，Excel友好）
        
        Args:
            file_path: 文件路径
            main_data: 主数据
            main_headers: 主数据表头
            remaining_tree: 剩余数据的Treeview对象或MockTree对象
            remaining_headers: 剩余数据表头，如果为None则从tree中获取
        """
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)

            writer.writerow(main_headers)
            for values in main_data:
                writer.writerow(values)

            writer.writerow([])

            # 使用传递进来的headers，或者从tree中获取
            if remaining_headers:
                writer.writerow(remaining_headers)
            else:
                remaining_headers = [remaining_tree.heading(col, "text") or "" for col in remaining_tree["columns"]]
                writer.writerow(remaining_headers)
            
            # 写入剩余数据
            for item in remaining_tree.get_children():
                values = remaining_tree.item(item, "values")
                writer.writerow(values)

    def _export_to_excel(self, file_path, data_source, main_data, main_headers, remaining_tree, remaining_headers):
        """导出数据为Excel格式（原生支持中文）"""
        wb = Workbook()

        main_sheet = wb.active
        # 根据数据类型设置不同的sheet名称
        main_title = translate("split_segment_info") if data_source["main_name"] == translate("split_segment_info") else translate("subnet_requirements")
        main_sheet.title = str(main_title) if main_title else "Sheet"

        for col_idx, header in enumerate(main_headers, 1):
            cell = main_sheet.cell(row=1, column=col_idx, value=header)
            if cell:
                cell.font = Font(bold=True)  # type: ignore
                cell.alignment = Alignment(horizontal="center")  # type: ignore

        for row_idx, values in enumerate(main_data, 2):
            for col_idx, value in enumerate(values, 1):
                main_sheet.cell(row=row_idx, column=col_idx, value=value)

        remaining_sheet = wb.create_sheet(title=str(translate("remaining_subnets")) if translate("remaining_subnets") else "Remaining")

        for col_idx, header in enumerate(remaining_headers, 1):
            cell = remaining_sheet.cell(row=1, column=col_idx, value=header)
            if cell:
                cell.font = Font(bold=True)  # type: ignore
                cell.alignment = Alignment(horizontal="center")  # type: ignore

        for row_idx, item in enumerate(remaining_tree.get_children(), 2):
            values = remaining_tree.item(item, "values")
            for col_idx, value in enumerate(values, 1):
                remaining_sheet.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)

    def _export_to_pdf(self, file_path, data_source, main_data, main_headers, _remaining_data, remaining_headers):
        """导出数据为PDF格式（支持中文、韩语、日语）"""
        # 直接获取当前语言，不依赖于之前的注册
        current_lang = get_language()
        print(f"🔍 导出PDF时当前语言: {current_lang}")

        # 检查是否已经注册过相同语言的字体
        if ExportUtils._font_path_lang == current_lang and ExportUtils._font_path_cache:
            print(f"🔍 使用已注册的PDF字体缓存 (语言: {current_lang})")
            self.has_asian_font = True
        else:
            # 语言不同或没有缓存，需要重新注册字体
            print(f"🔍 语言已切换或无缓存，需要重新注册字体 (旧语言: {ExportUtils._font_path_lang}, 新语言: {current_lang})")
            
            # 使用系统环境变量获取字体目录
            font_dir = os.path.join(os.environ.get('WINDIR', r'C:\Windows'), 'Fonts')

            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # 从 font_config 获取字体文件名列表
            font_candidates = FontConfig.get_font_filenames(current_lang)
            
            font_path = None
            
            # 查找可用字体
            for font_file in font_candidates:
                potential_path = os.path.join(font_dir, font_file)
                if os.path.exists(potential_path):
                    font_path = potential_path
                    print(f"🔍 找到可用字体: {font_file} 在 {font_path}")
                    break
                else:
                    print(f"⚠️ 字体文件不存在: {potential_path}")
            
            # 如果找到字体，直接注册
            if font_path:
                try:
                    print(f"🔍 尝试注册字体: {font_path}")
                    print(f"🔍 当前已注册字体: {pdfmetrics.getRegisteredFontNames()}")

                    # 尝试注册字体,如果已存在则先清理
                    font_name = "ChineseFont"
                    if font_name in pdfmetrics.getRegisteredFontNames():
                        # 字体已注册,尝试删除并重新注册
                        print(f"🔍 字体 '{font_name}' 已注册,先删除再重新注册")
                        try:
                            # ReportLab没有直接的删除方法,但我们可以从内部字典中删除
                            if hasattr(pdfmetrics, '_fonts') and font_name in pdfmetrics._fonts:  # type: ignore
                                del pdfmetrics._fonts[font_name]  # type: ignore
                                print("🔍 已删除旧的字体注册")
                        except Exception as del_error:
                            print(f"⚠️ 删除旧字体时出现警告: {del_error}")

                    # 重新注册字体
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    self.has_asian_font = True
                    
                    # 缓存字体路径和语言
                    ExportUtils._font_path_cache = font_path
                    ExportUtils._font_path_lang = current_lang
                    
                    print(f"✅ 成功注册字体: {os.path.basename(font_path)} 作为 {font_name}")
                    print(f"🔍 注册后已注册字体: {pdfmetrics.getRegisteredFontNames()}")
                    
                    # 测试字体是否支持当前语言字符
                    test_text = FontConfig.get_font_test_text(current_lang)
                    try:
                        from reportlab.pdfbase.pdfmetrics import stringWidth
                        width = stringWidth(test_text, font_name, 10)
                        print(f"✅ 字体测试通过: '{test_text}' 宽度={width}")
                    except Exception as test_error:
                        print(f"⚠️ 字体测试失败: {test_error}")

                except Exception as e:
                    print(f"❌ 注册字体失败: {e}")
                    print(f"❌ 异常类型: {type(e).__name__}")
                    traceback.print_exc()

                    # 如果注册失败,检查是否有已注册的字体可以回退
                    if "ChineseFont" in pdfmetrics.getRegisteredFontNames():
                        print("🔍 使用之前注册的 ChineseFont 字体(可能不支持当前语言)")
                        self.has_asian_font = True
                    else:
                        print("🔍 无可用字体,将使用默认字体(Helvetica)")
                        self.has_asian_font = False
            else:
                print("🔍 未找到合适字体，将使用默认字体")
                self.has_asian_font = False

        print(f"🔍 使用的主要字体: ChineseFont, has_asian_font={self.has_asian_font}")
        
        margins = (2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm)

        doc = BaseDocTemplate(
            file_path,
            pagesize=landscape(A4),
            leftMargin=margins[0],
            rightMargin=margins[1],
            topMargin=margins[2],
            bottomMargin=margins[3],
            showBoundary=False,
        )

        landscape_width, landscape_height = landscape(A4)
        portrait_width, portrait_height = A4
        
        # 根据当前语言获取日期格式
        def get_date_format():
            lang = get_language()
            if lang == "ko":
                # 韩语日期格式：2025년01월11일 14:30:00
                return "%Y년%m월%d일 %H:%M:%S"
            elif lang in ["zh", "zh_tw"]:
                # 中文日期格式：2025年01月11日 14:30:00
                return "%Y年%m月%d日 %H:%M:%S"
            elif lang == "ja":
                # 日语日期格式：2025年01月11日 14:30:00
                return "%Y年%m月%d日 %H:%M:%S"
            else:  # 英文
                return "%Y-%m-%d %H:%M:%S"
        
        # 生成导出时间，用于页眉显示
        export_time = time.strftime(get_date_format())
        
        # 创建页眉回调函数
        def on_page(canvas, _event):
            """页面回调函数，用于绘制页眉"""
            canvas.saveState()
            # 获取当前页面尺寸
            current_width, current_height = canvas._pagesize
            # 设置字体
            font_name = "ChineseFont" if self.has_asian_font else "Helvetica"
            canvas.setFont(font_name, 10)
            canvas.setFillColor(colors.HexColor("#666666"))
            # 在页眉右侧绘制导出时间
            canvas.drawRightString(
                current_width - margins[1],  # x坐标：页面宽度 - 右边距
                current_height - 40,  # y坐标：页面顶部 - 20像素，留有适当边距
                f"{translate('export_time')}: {export_time}"
            )
            canvas.restoreState()
        
        landscape_frame = Frame(
            margins[0],
            margins[3],
            landscape_width - margins[0] - margins[1],
            landscape_height - margins[2] - margins[3],
            id='landscape_frame',
        )
        landscape_template = PageTemplate(id='landscape', frames=[landscape_frame], onPage=on_page)

        portrait_frame = Frame(
            margins[0],
            margins[3],
            portrait_width - margins[0] - margins[1],
            portrait_height - margins[2] - margins[3],
            id='portrait_frame',
        )
        portrait_template = PageTemplate(id='portrait', frames=[portrait_frame], pagesize=A4, onPage=on_page)

        doc.addPageTemplates([landscape_template, portrait_template])

        page_width = landscape_width
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ChineseTitle",
            parent=styles["Title"],
            fontName="ChineseFont" if self.has_asian_font else "Helvetica-Bold",
            fontSize=20,
            textColor=colors.HexColor("#2c3e50"),
            alignment=TA_CENTER,
            spaceAfter=20,
        )

        heading2_style = ParagraphStyle(
            "ChineseHeading2",
            parent=styles["Heading2"],
            fontName="ChineseFont" if self.has_asian_font else "Helvetica-Bold",
            fontSize=16,
            textColor=colors.HexColor("#34495e"),
            alignment=TA_LEFT,
            spaceBefore=20,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "ChineseNormal",
            parent=styles["Normal"],
            fontName="ChineseFont" if self.has_asian_font else "Helvetica",
            fontSize=11,
            textColor=colors.HexColor("#34495e"),
            spaceAfter=5,
        )

        table_text_style = ParagraphStyle(
            "ChineseTableText",
            parent=styles["Normal"],
            fontName="ChineseFont" if self.has_asian_font else "Helvetica",
            fontSize=10,
            alignment=TA_CENTER,
            wordWrap="None",  # 禁用自动换行
            leading=12,  # 设置行间距为12，确保垂直居中
            spaceBefore=0,
            spaceAfter=0,
        )

        elements.append(Paragraph(data_source["pdf_title"], title_style))
        elements.append(Spacer(1, 15))

        # 在切分段信息/已分配子网信息前添加父网段信息
        if data_source["main_name"] in [translate("split_segment_info"), translate("allocated_subnets")]:
            # 显示父网段信息
            elements.append(Paragraph(str(translate("parent_network_info")), heading2_style))
            
            # 从chart_data获取父网段信息
            chart_data = data_source.get("chart_data")
            if chart_data and "parent" in chart_data:
                parent_info = chart_data["parent"]
                
                # 获取完整的父网段信息
                from ip_subnet_calculator import get_subnet_info
                
                parent_cidr = parent_info.get("name", "")
                full_parent_info = get_subnet_info(parent_cidr)
                
                # 转置表格：将信息水平排列
                parent_table_data = []
                
                # 第一行：标题行
                # 根据IP版本确定显示哪些字段
                parent_ip_version = full_parent_info.get('version', 4)
                is_ipv6 = parent_ip_version == 6
                
                if is_ipv6:
                    # IPv6只显示必要字段
                    parent_table_data.append([
                        Paragraph(str(translate("parent_cidr")), table_text_style),
                        Paragraph(str(translate("network_address")), table_text_style),
                        Paragraph(str(translate("prefix_length")), table_text_style),
                        Paragraph(str(translate("available_addresses")), table_text_style),
                        Paragraph(str(translate("host_address_range")), table_text_style)
                    ])
                    
                    # 第二行：数据行
                    parent_table_data.append([
                        Paragraph(str(full_parent_info.get("cidr", parent_cidr)), table_text_style),
                        Paragraph(str(full_parent_info.get("network", "")), table_text_style),
                        Paragraph(str(full_parent_info.get("prefixlen", "")), table_text_style),
                        Paragraph(self.format_large_number(full_parent_info.get('usable_addresses', 0)), table_text_style),
                        Paragraph(f"{full_parent_info.get('host_range_start', '')} - {full_parent_info.get('host_range_end', '')}", table_text_style)
                    ])
                else:
                    # IPv4显示所有字段
                    parent_table_data.append([
                        Paragraph(str(translate("parent_cidr")), table_text_style),
                        Paragraph(str(translate("network_address")), table_text_style),
                        Paragraph(str(translate("subnet_mask")), table_text_style),
                        Paragraph(str(translate("wildcard_mask")), table_text_style),
                        Paragraph(str(translate("broadcast_address")), table_text_style),
                        Paragraph(str(translate("prefix_length")), table_text_style),
                        Paragraph(str(translate("available_addresses")), table_text_style),
                        Paragraph(str(translate("host_address_range")), table_text_style)
                    ])
                    
                    # 第二行：数据行
                    parent_table_data.append([
                        Paragraph(str(full_parent_info.get("cidr", parent_cidr)), table_text_style),
                        Paragraph(str(full_parent_info.get("network", "")), table_text_style),
                        Paragraph(str(full_parent_info.get("netmask", "")), table_text_style),
                        Paragraph(str(full_parent_info.get("wildcard", "")), table_text_style),
                        Paragraph(str(full_parent_info.get("broadcast", "")), table_text_style),
                        Paragraph(str(full_parent_info.get("prefixlen", "")), table_text_style),
                        Paragraph(self.format_large_number(full_parent_info.get('usable_addresses', 0)), table_text_style),
                        Paragraph(f"{full_parent_info.get('host_range_start', '')} - {full_parent_info.get('host_range_end', '')}", table_text_style)
                    ])
                
                # 创建父网段信息表格
                table_width = page_width - margins[0] - margins[1]
                
                # 使用自适应列宽
                num_cols = len(parent_table_data[0])
                # 获取自适应列宽
                col_widths = self._get_col_widths(parent_table_data, table_width, None, num_cols)
                
                # 确保所有列宽都是有效的数字
                valid_col_widths = []
                for width in col_widths:
                    if width is None or not isinstance(width, (int, float)) or width <= 0:
                        valid_col_widths.append(table_width / num_cols)
                    else:
                        valid_col_widths.append(width)
                
                # 确保表格数据有效
                valid_table_data = []
                for row in parent_table_data:
                    valid_row = []
                    for cell in row:
                        if cell is None:
                            valid_row.append(Paragraph("", table_text_style))
                        else:
                            valid_row.append(cell)
                    valid_table_data.append(valid_row)
                
                # 创建表格时直接传递列宽
                parent_table = Table(valid_table_data, colWidths=valid_col_widths, repeatRows=1)
                # 应用样式
                parent_table.setStyle(self._get_table_style(MAIN_TABLE_COLORS, self.has_asian_font))
                elements.append(parent_table)
                elements.append(Spacer(1, 20))
        
        # 显示切分段信息或已分配子网信息
        main_heading = Paragraph(data_source["main_name"], heading2_style)
        # 准备KeepTogether的内容列表
        keep_together_main = [main_heading]

        # 检测是否为IPv6模式：检查数据中是否包含IPv6地址
        is_ipv6 = False
        for values in main_data:
            for v in values:
                if isinstance(v, str) and ':' in v and len(v) > 10:  # 简单检测IPv6地址
                    is_ipv6 = True
                    break
            if is_ipv6:
                break

        # 特殊处理：子网切分PDF的切分段信息表格
        if data_source["main_name"] == translate("split_segment_info"):
            # 转置表格并移除指定列
            # 定义要移除的列名
            columns_to_remove = [translate("parent_network"), translate("split_line"), translate("prefix_length"), translate("cidr"), translate("separator"), translate("network_address")]
            
            # 对于IPv4版本的切分段信息表格，移除广播地址字段
            if not is_ipv6:
                columns_to_remove.append(translate("broadcast_address"))
            
            # 键值对格式处理
            filtered_data = {}
            for values in main_data:
                key = str(values[0]) if values[0] is not None else ""
                value = str(values[1]) if values[1] is not None else ""
                # 只保留非空键和不需要移除的键，同时移除虚线列
                if key and key not in columns_to_remove and not all(c == '-' for c in key):
                    filtered_data[key] = value
            
            # 转置：将键作为表头，值作为一行
            if filtered_data:
                # 创建表头行
                headers = list(filtered_data.keys())
                main_table_data = [[Paragraph(h, table_text_style) for h in headers]]
                # 创建数据行
                values = [filtered_data[h] for h in headers]
                main_table_data.append([Paragraph(str(v), table_text_style) for v in values])
            else:
                main_table_data = []
        elif self._is_k2v_headers(main_headers):
            main_table_data = [[Paragraph("项目", table_text_style), Paragraph("值", table_text_style)]]
            for values in main_data:
                main_table_data.append(
                    [
                        Paragraph(str(values[0]) if values[0] is not None else "", table_text_style),
                        Paragraph(str(values[1]) if values[1] is not None else "", table_text_style),
                    ]
                )
        else:
            # 根据IP版本过滤已分配子网表格的字段
            if is_ipv6 and data_source["main_name"] == translate("allocated_subnets"):
                # IPv6模式：过滤已分配子网表格的列
                main_headers = [h for h in main_headers if h not in [translate("subnet_mask"), translate("wildcard_mask"), translate("broadcast_address")]]
                main_table_data = [[Paragraph(h, table_text_style) for h in main_headers]]
                
                for values in main_data:
                    # 创建过滤后的值列表，移除不适用的字段
                    filtered_values = []
                    for i, h in enumerate(main_data[0]):
                        if i < len(values):
                            header = main_headers[i] if i < len(main_headers) else ""
                            if header not in [translate("subnet_mask"), translate("wildcard_mask"), translate("broadcast_address")]:
                                filtered_values.append(values[i])
                    main_table_data.append(
                        [Paragraph(str(v) if v is not None else "", table_text_style) for v in filtered_values]
                    )
            else:
                # IPv4模式：过滤已分配子网表格的列，移除"网段结束地址"字段
                filtered_headers = [h for h in main_headers if h != translate("network_end_address")]
                main_table_data = [[Paragraph(h, table_text_style) for h in filtered_headers]]
                
                for values in main_data:
                    # 创建过滤后的值列表，移除"网段结束地址"字段
                    filtered_values = []
                    for i, header in enumerate(main_headers):
                        if i < len(values):
                            if header != translate("network_end_address"):
                                filtered_values.append(values[i])
                    main_table_data.append(
                        [Paragraph(str(v) if v is not None else "", table_text_style) for v in filtered_values]
                    )

        if len(main_table_data) > 1:
            table_width = page_width - margins[0] - margins[1]
            table_cols = len(main_table_data[0])

            col_widths = data_source.get("main_table_cols")
            if isinstance(col_widths, str):
                try:
                    col_ratios = [float(w) for w in col_widths.split(":")]
                    if all(ratio < 10 for ratio in col_ratios):
                        total_ratio = sum(col_ratios)
                        col_widths = [table_width * (ratio / total_ratio) for ratio in col_ratios] if total_ratio > 0 else None
                    else:
                        col_widths = col_ratios
                except (ValueError, TypeError):
                    col_widths = None

            valid_col_widths = self._get_col_widths(main_table_data, table_width, col_widths, table_cols)

            # 动态调整表格单元格的字体大小
            adjusted_table_data = self._adjust_table_font_size(main_table_data, valid_col_widths, table_text_style, style_prefix="CustomStyle")

            main_table = Table(adjusted_table_data, colWidths=valid_col_widths, repeatRows=1)
            main_table.setStyle(self._get_table_style(MAIN_TABLE_COLORS, self.has_asian_font))
            keep_together_main.append(main_table)
        else:
            main_name_text = str(translate('no')) if translate('no') else "无"
            keep_together_main.append(Paragraph(f"{main_name_text}{str(data_source['main_name'])}", normal_style))
        
        # 将标题和表格包装在KeepTogether中
        elements.append(KeepTogether(keep_together_main))
        elements.append(Spacer(1, 20))

        # 显示剩余网段信息
        remaining_heading = Paragraph(str(data_source["remaining_name"]), heading2_style)
        # 准备KeepTogether的内容列表
        keep_together_remaining = [remaining_heading]
        
        # 使用传入的remaining_data和remaining_headers，而不是直接从Treeview获取
        remaining_table_data = [[Paragraph(str(h), table_text_style) for h in remaining_headers]]
        
        # 将remaining_data转换为表格数据
        for item in _remaining_data:
            if isinstance(item, dict):
                # 如果是字典，按照remaining_headers的顺序提取值
                values = [item.get(header, '') for header in remaining_headers]
            else:
                # 如果已经是列表，直接使用
                values = item
            remaining_table_data.append(
                [Paragraph(str(v) if v is not None else "", table_text_style) for v in values]
            )

        if len(remaining_table_data) > 1:
            table_width = page_width - margins[0] - margins[1]
            table_cols = len(remaining_table_data[0])

            col_widths = data_source.get("remaining_table_cols")
            if isinstance(col_widths, str):
                try:
                    col_ratios = [float(w) for w in col_widths.split(":")]
                    if all(ratio < 10 for ratio in col_ratios):
                        total_ratio = sum(col_ratios)
                        col_widths = [table_width * (ratio / total_ratio) for ratio in col_ratios] if total_ratio > 0 else None
                    else:
                        col_widths = col_ratios
                except (ValueError, TypeError):
                    col_widths = None

            valid_col_widths = self._get_col_widths(remaining_table_data, table_width, col_widths, table_cols)

            # 动态调整剩余网段表格单元格的字体大小
            adjusted_remaining_data = self._adjust_table_font_size(remaining_table_data, valid_col_widths, table_text_style, style_prefix="CustomRemainingStyle")

            remaining_table = Table(adjusted_remaining_data, colWidths=valid_col_widths, repeatRows=1)
            remaining_table.setStyle(self._get_table_style(REMAINING_TABLE_COLORS, self.has_asian_font))
            keep_together_remaining.append(remaining_table)
        else:
            no_text = str(translate('no')) if translate('no') else "无"
            remaining_name_text = str(data_source['remaining_name'])
            keep_together_remaining.append(Paragraph(f"{no_text}{remaining_name_text}", normal_style))
        
        # 将标题和表格包装在KeepTogether中
        elements.append(KeepTogether(keep_together_remaining))

        # 添加网段分布图（对子网切分和子网规划功能，放在剩余网段表之后）
        chart_data = data_source.get("chart_data")
        # 检查chart_data是否存在，而不依赖于硬编码的中文文本
        if chart_data:
            try:
                print("开始添加网段分布图到PDF...")
                self._add_chart_to_pdf(elements, chart_data, margins, portrait_width, portrait_height)
            except (IOError, ValueError, TypeError, AttributeError) as e:
                print(f"添加网段分布图失败: {e}")
                traceback.print_exc()

        doc.build(elements)

    def _load_system_font(self, font_size=36, bold_offset=4, verbose=False):
        """加载系统字体（支持中文、韩语、日语），使用缓存避免重复加载

        Args:
            font_size: 字体大小
            bold_offset: 粗体字体大小偏移量
            verbose: 是否打印详细信息

        Returns:
            tuple: (font, bold_font, font_loaded)
        """
        # 生成缓存键
        cache_key = (font_size, bold_offset)
        
        # 检查缓存
        if cache_key in ExportUtils._font_cache:
            if verbose:
                print(f"使用缓存的字体 (size={font_size}, bold_offset={bold_offset})")
            return ExportUtils._font_cache[cache_key]
        
        font = None
        bold_font = None
        font_loaded = False
        try:
            from i18n import get_language
            current_lang = get_language()
            
            system_font_dir = os.path.join(os.environ.get('WINDIR', r'C:\Windows'), 'Fonts')
            
            # 从 font_config 获取字体候选列表
            font_candidates_tuples = FontConfig.get_font_candidates(current_lang)
            font_candidates = [(font_file, font_size, font_name) for font_file, font_name in font_candidates_tuples]


            for font_file, size, font_name in font_candidates:
                font_path = os.path.join(system_font_dir, font_file)
                if os.path.exists(font_path):
                    try:
                        font = ImageFont.truetype(font_path, size)
                        bold_font = ImageFont.truetype(font_path, size + bold_offset)
                        font_loaded = True
                        if verbose:
                            print(f"成功加载{font_name}字体: {font_path}")
                        break
                    except (FileNotFoundError, IOError, OSError, ValueError, TypeError) as e:
                        if verbose:
                            print(f"尝试加载{font_name}失败: {e}")
                        continue

            if not font_loaded:
                font = ImageFont.load_default()
                bold_font = ImageFont.load_default()
                if verbose:
                    print("使用默认字体")
        except (IOError, OSError, ValueError, TypeError) as e:
            if verbose:
                print(f"加载系统字体失败: {e}")
            font = ImageFont.load_default()
            bold_font = ImageFont.load_default()
        
        # 缓存结果
        result = (font, bold_font, font_loaded)
        ExportUtils._font_cache[cache_key] = result
        
        return result

    def _calculate_chart_dimensions(self, networks):
        """计算图表所需的尺寸

        Args:
            networks: 网段列表

        Returns:
            tuple: (segment_height, required_height)
        """
        split_networks = [net for net in networks if net.get("type") == "split"]
        remaining_networks = [net for net in networks if net.get("type") != "split"]
        total_networks = len(split_networks) + len(remaining_networks)

        segment_height = 100 + 34
        # 动态计算基础高度和所需总高度
        base_height = 280 + 100 + 100 + 150 + 300
        required_height = base_height + total_networks * segment_height
        return segment_height, required_height

    def _draw_title(self, draw, high_res_width, title_font, title=None):
        """绘制图表标题

        Args:
            draw: ImageDraw对象
            high_res_width: 图像宽度
            title_font: 标题字体对象
            title: 标题文字，默认使用国际化的"distribution_chart"

        Returns:
            tuple: (title_font, title_x, title_y)
        """
        # 使用传入的字体对象，避免重复加载
        # 使用国际化标题，如果没有提供则使用默认翻译
        if not title:
            title = translate("distribution_chart")

        title_bbox = draw.textbbox((0, 0), title, font=title_font)
        title_x = (high_res_width - (title_bbox[2] - title_bbox[0])) // 2
        title_y = 100
        return title_font, title_x, title_y

    def _add_chart_to_pdf(self, elements, chart_data, margins, portrait_width, portrait_height):
        """添加网段分布图到PDF元素列表

        Args:
            elements: PDF元素列表
            chart_data: 网段分布图数据
            margins: 页面边距
            portrait_width: 纵向页面宽度
            portrait_height: 纵向页面高度
            main_name: 主数据名称，用于区分子网切分和子网规划
        """
        # 导入翻译函数，避免与局部变量冲突
        from i18n import _ as translate  # _ 是翻译函数，这里重命名为 translate 以避免冲突
        
        if not chart_data or 'networks' not in chart_data or len(chart_data['networks']) == 0:
            print("没有有效的网段分布图数据，跳过")
            return

        print(f"检测到有效网段分布图数据，包含 {len(chart_data['networks'])} 个网段")

        # 切换到纵向页面模板
        elements.append(NextPageTemplate('portrait'))
        elements.append(PageBreak())

        # 准备图表数据
        parent_info = chart_data.get("parent", {})
        parent_cidr = parent_info.get("name", translate("parent_network"))
        parent_range = parent_info.get("range", 1)
        networks = chart_data.get("networks", [])

        # 直接从chart_data中获取chart_type，如果没有则根据数据结构自动判断
        chart_type = chart_data.get("type", "split")
        
        # 计算切分段和剩余网段
        split_networks = [net for net in networks if net.get("type") == "split"]
        remaining_networks = [net for net in networks if net.get("type") != "split"]
        
        # 精确计算图表所需的总高度
        # 基本元素高度
        title_height = 280  # 标题部分高度
        parent_network_height = 150  # 父网段部分高度
        legend_height = 150  # 图例部分高度（减小，因为图例实际不需要那么高）
        
        # 网段列表高度
        segment_height = 134  # 每个网段行的高度（100高度 + 34间距）
        
        # 需求网段/切分段标题高度
        demand_title_height = 180 if chart_type == "plan" else 0  # 需求网段标题高度
        
        # 分隔线高度
        separator_height = 100  # 分隔线和间距高度
        
        # 剩余网段标题高度
        remaining_title_height = 180  # 剩余网段标题高度
        
        # 精确计算所需总高度
        required_height = (title_height
                          + parent_network_height
                          + separator_height
                          + demand_title_height
                          + len(split_networks) * segment_height
                          + separator_height
                          + remaining_title_height
                          + len(remaining_networks) * segment_height
                          + legend_height
                          + 50)  # 减小额外的安全间距到50

        # 创建高分辨率图像
        high_res_width = 2480
        high_res_height = max(3508, required_height)

        pil_image = Image.new('RGB', (high_res_width, high_res_height), color='#333333')
        draw = ImageDraw.Draw(pil_image)

        # 一次性加载所有需要的字体，避免重复加载
        # 标题字体
        title_font_size = 76
        _, title_font, _ = self._load_system_font(title_font_size, verbose=False)
        # 普通文本字体
        font, bold_font, _ = self._load_system_font(font_size=36, verbose=False)
        # 大号文本字体
        text_font_size = 50
        text_font, bold_text_font, _ = self._load_system_font(font_size=text_font_size, bold_offset=6, verbose=False)

        # 设置图表参数
        margin_left = 180
        margin_right = 150
        margin_top = 280
        chart_width = high_res_width - margin_left - margin_right
        chart_x = margin_left
        chart_right = chart_x + chart_width  # 计算图表右边缘位置
        ADDRESS_OFFSET = 50  # 地址文本距离图表右边缘的偏移量（左缩进）

        # 定义绘制带描边文字的辅助函数
        def draw_text_with_stroke(draw_obj, position, text, font, fill, stroke_color="#666666", stroke_width=4):
            """绘制带描边的文字

            Args:
                draw_obj: ImageDraw对象
                position: (x, y) 坐标
                text: 要绘制的文字
                font: 字体对象
                fill: 文字颜色
                stroke_color: 描边颜色
                stroke_width: 描边宽度
            """
            # 使用PIL的描边参数绘制文字
            try:
                draw_obj.text(
                    position,
                    text,
                    font=font,
                    fill=fill,
                    stroke_width=stroke_width,
                    stroke_fill=stroke_color
                )
            except (TypeError, AttributeError):
                # 如果PIL版本不支持描边参数，使用方向性格式
                x, y = position
                directions = [(-1, -1), (-1, 1), (1, -1), (1, 1)]
                for dx, dy in directions:
                    draw_obj.text((x + dx, y + dy), text, font=font, fill=stroke_color)
                draw_obj.text((x, y), text, font=font, fill=fill)

        # 使用对数比例尺
        log_max = math.log10(parent_range)
        log_min = 3

        min_bar_width = 120
        padding = 34
        bar_height = 100
        
        # 为所有网段分配颜色
        subnet_colors = [
            "#5e9c6a", "#db6679", "#f0ab55", "#8b6cb8",
            "#5b8fd9", "#3c70d8", "#e68838", "#a04132",
            "#6a9da8", "#87c569", "#6d8de8", "#c16fa0",
            "#a99bc6", "#a44d69", "#b9d0f8", "#5d4ea5",
            "#f5ad8c", "#5b8fd9", "#db6679", "#a6c589",
        ]

        # 绘制标题
        title_font, title_x, title_y = self._draw_title(draw, high_res_width, title_font)
        draw_text_with_stroke(draw, (title_x, title_y), translate("distribution_chart"), title_font, "#ffffff")

        y = margin_top

        # 绘制父网段
        log_value = max(log_min, math.log10(parent_range))
        bar_width = max(min_bar_width, ((log_value - log_min) / (log_max - log_min)) * chart_width)
        parent_color = "#636e72"
        draw.rectangle([chart_x, y, chart_x
            + bar_width, y
            + bar_height], fill=parent_color, outline=None, width=0)

        usable_addresses = parent_range - 2 if parent_range > 2 else parent_range
        segment_text = f"{translate('parent_network')}: {parent_cidr}"
        address_text = f"{translate('usable_addresses')}: {self.format_large_number(usable_addresses)}"

        def get_centered_y(box_y, box_height, _, _font):
            text_y = box_y + box_height // 2 - 38
            return text_y

        segment_bbox = draw.textbbox((0, 0), segment_text, font=bold_text_font)
        segment_text_y = get_centered_y(y, bar_height, segment_bbox, bold_text_font)
        address_bbox = draw.textbbox((0, 0), address_text, font=bold_text_font)
        address_text_y = get_centered_y(y, bar_height, address_bbox, bold_text_font)
        
        # 计算地址文本的右对齐位置：图表右边缘 - 左偏移 - 文本宽度
        address_width = address_bbox[2] - address_bbox[0]
        address_x = chart_right - ADDRESS_OFFSET - address_width

        draw_text_with_stroke(draw, (chart_x
                + 30, segment_text_y), segment_text, bold_text_font, "#ffffff")
        draw_text_with_stroke(draw, (address_x, address_text_y), address_text, bold_text_font, "#ffffff")

        y += bar_height + padding

        # 在父网段和需求/切分网段之间添加分割线（仅子网规划显示）
        if chart_type == "plan":
            draw.line([chart_x, y + 20, chart_x + chart_width, y + 20], fill="#cccccc", width=4)
            y += 60  # 增大间距，让需求网段与上部分割线的距离和剩余网段保持一致
        else:
            y += 0  # 子网切分保持较小间距
        
        # 绘制需求网段标题
        if chart_type == "plan":
            demand_count = len(split_networks)
            title_text = f"{translate('allocated_subnets')} ({demand_count} {translate('pieces')}):"
            title_bbox = draw.textbbox((0, 0), title_text, font=bold_text_font)
            title_text_y = get_centered_y(y, bar_height, title_bbox, bold_text_font)
            draw_text_with_stroke(draw, (chart_x, title_text_y), title_text, bold_text_font, "#ffffff")
            y += 100
        
        # 绘制切分/需求网段
        for i, network in enumerate(split_networks):
            network_range = network.get("range", 1)
            log_value = max(log_min, math.log10(network_range))
            bar_width = max(min_bar_width, ((log_value - log_min) / (log_max - log_min)) * chart_width)
            
            if chart_type == "split":
                # 子网切分 - 使用加粗字体
                split_color = "#4a7eb4"
                name = network.get("name", "")
                segment_text = f"{translate('split_segment')}: {name}"
                
                usable_addresses = network_range - 2 if network_range > 2 else network_range
                address_text = f"{translate("usable_addresses")}: {self.format_large_number(usable_addresses)}"
                
                segment_bbox = draw.textbbox((0, 0), segment_text, font=bold_text_font)
                segment_text_y = get_centered_y(y, bar_height, segment_bbox, bold_text_font)
                address_bbox = draw.textbbox((0, 0), address_text, font=bold_text_font)
                address_text_y = get_centered_y(y, bar_height, address_bbox, bold_text_font)
                
                # 计算地址文本的右对齐位置：图表右边缘 - 左偏移 - 文本宽度
                address_width = address_bbox[2] - address_bbox[0]
                address_x = chart_right - ADDRESS_OFFSET - address_width
                
                draw_text_with_stroke(draw, (chart_x
                    + 30, segment_text_y), segment_text, bold_text_font, "#ffffff")
                draw_text_with_stroke(draw, (address_x, address_text_y), address_text, bold_text_font, "#ffffff")
            else:
                # 子网规划 - 已分配子网，使用普通字体
                color_index = i % len(subnet_colors)
                split_color = subnet_colors[color_index]
                name = network.get("name", "")
                cidr = network.get("cidr", "")
                segment_text = f"{translate('segment')} {i + 1}: {name}    {cidr}"
                
                draw.rectangle([chart_x, y, chart_x
                    + bar_width, y
                    + bar_height], fill=split_color, outline=None, width=0)

                usable_addresses = network_range - 2 if network_range > 2 else network_range
                address_text = f"{translate("usable_addresses")}: {self.format_large_number(usable_addresses)}"

                segment_bbox = draw.textbbox((0, 0), segment_text, font=text_font)
                segment_text_y = get_centered_y(y, bar_height, segment_bbox, text_font)
                address_bbox = draw.textbbox((0, 0), address_text, font=text_font)
                address_text_y = get_centered_y(y, bar_height, address_bbox, text_font)
                
                # 计算地址文本的右对齐位置：图表右边缘 - 左偏移 - 文本宽度
                address_width = address_bbox[2] - address_bbox[0]
                address_x = chart_right - ADDRESS_OFFSET - address_width

                draw_text_with_stroke(draw, (chart_x
                    + 30, segment_text_y), segment_text, text_font, "#ffffff")
                draw_text_with_stroke(draw, (address_x, address_text_y), address_text, text_font, "#ffffff")

            y += bar_height + padding

        # 在需求/切分网段和剩余网段之间添加分割线
        draw.line([chart_x, y + 20, chart_x + chart_width, y + 20], fill="#cccccc", width=4)

        # 绘制剩余网段标题
        y += 80
        remaining_count = len(remaining_networks)
        title_text = f"{translate('remaining_subnets')} ({remaining_count} {translate('pieces')}):"

        title_bbox = draw.textbbox((0, 0), title_text, font=bold_text_font)
        title_text_y = get_centered_y(y, bar_height, title_bbox, bold_text_font)
        draw_text_with_stroke(draw, (chart_x, title_text_y), title_text, bold_text_font, "#ffffff")
        y += 100

        # 绘制剩余网段
        for i, network in enumerate(remaining_networks):
            network_range = network.get("range", 1)
            log_value = max(log_min, math.log10(network_range))
            bar_width = max(min_bar_width, ((log_value - log_min) / (log_max - log_min)) * chart_width)
            color_index = i % len(subnet_colors)
            color = subnet_colors[color_index]
            draw.rectangle([chart_x, y, chart_x
                + bar_width, y
                + bar_height], fill=color, outline=None, width=0)

            name = network.get("name", "")
            usable_addresses = network_range - 2 if network_range > 2 else network_range
            
            if chart_type == "split":
                segment_text = f"{translate('segment')} {i + 1}: {name}"
            else:
                segment_text = f"{translate('segment')} {i + 1}: {name}"
                
            address_text = f"{translate("usable_addresses")}: {self.format_large_number(usable_addresses)}"

            segment_bbox = draw.textbbox((0, 0), segment_text, font=text_font)
            segment_text_y = get_centered_y(y, bar_height, segment_bbox, text_font)
            address_bbox = draw.textbbox((0, 0), address_text, font=text_font)
            address_text_y = get_centered_y(y, bar_height, address_bbox, text_font)
            
            # 计算地址文本的右对齐位置：图表右边缘 - 左偏移 - 文本宽度
            address_width = address_bbox[2] - address_bbox[0]
            address_x = chart_right - ADDRESS_OFFSET - address_width

            draw_text_with_stroke(draw, (chart_x
                + 30, segment_text_y), segment_text, text_font, "#ffffff")
            draw_text_with_stroke(draw, (address_x, address_text_y), address_text, text_font, "#ffffff")

            y += bar_height + padding

        # 绘制图例
        y += 80
        legend_title = f"{translate('legend')}:"
        legend_title_bbox = draw.textbbox((0, 0), legend_title, font=bold_text_font)
        legend_title_y = y + (bar_height - (legend_title_bbox[3] - legend_title_bbox[1])) // 2
        draw_text_with_stroke(draw, (chart_x, legend_title_y), legend_title, bold_text_font, "#ffffff")
        y += 100

        legend_y = y
        legend_item_height = 60
        legend_container_y = legend_y
        legend_container_height = legend_item_height

        def get_centered_text_y(container_y, container_height, text_bbox):
            text_height = text_bbox[3] - text_bbox[1]
            container_center = container_y + container_height // 2
            text_y = container_center - text_height // 2 - int(text_height * 0.30)
            return text_y

        # 父网段图例
        parent_x = chart_x
        parent_color = "#636e72"
        parent_label = translate("parent_network")
        parent_block_size = 40
        parent_text_font = text_font
        parent_label_bbox = draw.textbbox((0, 0), parent_label, font=parent_text_font)

        parent_block_y = legend_container_y + (legend_container_height - parent_block_size) // 2
        parent_label_y = get_centered_text_y(legend_container_y, legend_container_height, parent_label_bbox)

        draw.rectangle([parent_x, parent_block_y, parent_x + parent_block_size, parent_block_y + parent_block_size],
                      fill=parent_color, outline=None, width=0)
        draw_text_with_stroke(draw, (parent_x
            + parent_block_size
            + 25, parent_label_y), parent_label, parent_text_font, "#ffffff")

        # 切分/需求网段图例 - 动态计算位置，增加间距
        # 先计算父网段图例的总宽度
        parent_label_width = parent_label_bbox[2] - parent_label_bbox[0]
        split_x = parent_x + parent_block_size + 25 + parent_label_width + 80  # 增加间距到80
        split_text_font = text_font
        
        if chart_type == "split":
            # 子网切分 - 单一颜色
            split_color = "#4a7eb4"
            split_label = translate("split_segment")
            split_block_size = 40
            
            split_block_y = legend_container_y + (legend_container_height - split_block_size) // 2
            split_label_bbox = draw.textbbox((0, 0), split_label, font=split_text_font)
            split_label_y = get_centered_text_y(legend_container_y, legend_container_height, split_label_bbox)
            
            draw.rectangle([split_x, split_block_y, split_x + split_block_size, split_block_y + split_block_size],
                          fill=split_color, outline=None, width=0)
            draw_text_with_stroke(draw, (split_x + split_block_size + 15, split_label_y), split_label, split_text_font, "#ffffff")
        else:
            # 子网规划 - 多彩样式
            split_label = f"{translate('allocated_subnets')}"
            legend_colors = ["#5e9c6a", "#db6679", "#f0ab55", "#8b6cb8"]
            split_block_size = 30
            split_block_gap = 15  # 减小颜色块之间的间距
            
            split_block_y = legend_container_y + (legend_container_height - split_block_size) // 2
            split_label_bbox = draw.textbbox((0, 0), split_label, font=split_text_font)
            split_label_y = get_centered_text_y(legend_container_y, legend_container_height, split_label_bbox)
            
            # 绘制多彩图例块
            for j, color in enumerate(legend_colors):
                draw.rectangle(
                    [split_x + j * (split_block_size + split_block_gap), split_block_y,
                     split_x + j * (split_block_size + split_block_gap) + split_block_size,
                     split_block_y + split_block_size],
                    fill=color, outline=None, width=0
                )
            
            draw_text_with_stroke(draw,
                (split_x + len(legend_colors) * (split_block_size + split_block_gap) + 15, split_label_y),
                split_label, text_font, "#ffffff"
            )

        # 剩余网段图例 - 动态计算位置，增加间距
        # 计算切分/需求网段图例的总宽度
        if chart_type == "split":
            split_label_text = translate("split_segment")
            split_label_width = draw.textbbox((0, 0), split_label_text, font=split_text_font)[2] - draw.textbbox((0, 0), split_label_text, font=split_text_font)[0]
            remaining_x = split_x + split_block_size + 15 + split_label_width + 80  # 增加间距到80
        else:
            split_label_text = f"{translate('allocated_subnets')}"
            split_label_width = draw.textbbox((0, 0), split_label_text, font=split_text_font)[2] - draw.textbbox((0, 0), split_label_text, font=split_text_font)[0]
            remaining_x = split_x + len(legend_colors) * (split_block_size + split_block_gap) + 15 + split_label_width + 80  # 增加间距到80
        remaining_label = f"{translate('remaining_subnets')}"
        legend_colors = ["#5e9c6a", "#db6679", "#f0ab55", "#8b6cb8"]
        remaining_block_size = 30
        remaining_block_gap = 15  # 减小颜色块之间的间距
        remaining_text_font = text_font
        remaining_label_bbox = draw.textbbox((0, 0), remaining_label, font=remaining_text_font)

        remaining_block_y = legend_container_y + (legend_container_height - remaining_block_size) // 2
        remaining_label_y = get_centered_text_y(legend_container_y, legend_container_height, remaining_label_bbox)

        for j, color in enumerate(legend_colors):
            draw.rectangle(
                [remaining_x + j * (remaining_block_size + remaining_block_gap), remaining_block_y,
                 remaining_x + j * (remaining_block_size + remaining_block_gap) + remaining_block_size,
                 remaining_block_y + remaining_block_size],
                fill=color, outline=None, width=0
            )

        draw_text_with_stroke(draw,
            (remaining_x + len(legend_colors) * (remaining_block_size + remaining_block_gap) + 15, remaining_label_y),
            remaining_label, text_font, "#ffffff"
        )

        print("成功创建网段分布图")

        # 保存图像为高DPI PNG
        img_byte_arr = BytesIO()
        pil_image.save(img_byte_arr, format='PNG', dpi=(300, 300))
        img_byte_arr.seek(0)
        print(f"成功保存高DPI PNG图像，尺寸: {pil_image.size}, DPI: 300")

        # 计算图像在PDF中的尺寸
        actual_image_height = high_res_height
        image_ratio = high_res_width / actual_image_height

        available_width = portrait_width - margins[0] - margins[1] - 20
        
        # 不固定高度，而是根据图像比例和可用宽度计算合适的高度
        # 先根据可用宽度计算理论高度
        final_pdf_width = available_width
        final_pdf_height = final_pdf_width / image_ratio
        
        # 确保高度不会超过页面的可用高度
        max_available_height = portrait_height - margins[2] - margins[3] - 20
        if final_pdf_height > max_available_height:
            final_pdf_height = max_available_height
            final_pdf_width = final_pdf_height * image_ratio

        # 添加图像到PDF
        elements.append(RLImage(img_byte_arr, width=final_pdf_width, height=final_pdf_height))

        # 切换回横向页面模板
        elements.append(NextPageTemplate('landscape'))
        elements.append(PageBreak())

        print("网段分布图成功添加到PDF")

    def export_data(self, data_source, _title, _success_msg, failure_msg):
        """通用数据导出函数

        Args:
            data_source: 字典，包含导出数据的源信息
            title: 文件对话框标题
            success_msg: 成功消息格式字符串
            failure_msg: 失败消息格式字符串

        Returns:
            tuple: (success: bool, message: str, file_path: str or None)
        """
        try:
            main_data, main_headers, remaining_data, remaining_headers = self._prepare_export_data(data_source)

            is_valid = True
            error_msg = ""

            # 使用翻译函数进行比较，确保在任何语言环境下都能正确匹配
            split_segment_info = translate("split_segment_info")
            allocated_subnets = translate("allocated_subnets")
            
            if data_source["main_name"] == split_segment_info:
                if not main_data:
                    is_valid = False
                    error_msg = translate("no_split_data_found")
            elif data_source["main_name"] == allocated_subnets:
                if not main_data:
                    is_valid = False
                    error_msg = translate("no_planning_data_found")

            if not is_valid:
                return False, error_msg, None

            return True, "data_prepared", (main_data, main_headers, remaining_data, remaining_headers)

        except (ValueError, TypeError, KeyError) as e:
            traceback.print_exc()
            return False, f"{failure_msg}: {str(e)}", None

    def _create_mock_tree(self, headers, data):
        """创建模拟Treeview对象

        Args:
            headers: 表头列表
            data: 数据列表

        Returns:
            MockTree对象
        """
        class MockTree:
            """模拟Treeview对象"""
            def __init__(self, headers, data):
                self.headers = headers
                self.data = data
                self.columns = list(range(len(headers)))

            def heading(self, col, _event):
                """获取列标题

                Args:
                    col: 列索引或列名
                    option: 选项名，通常是"text"

                Returns:
                    列标题字符串
                """
                if isinstance(col, str) and col.isdigit():
                    col = int(col)
                return self.headers[col] if isinstance(col, int) and 0 <= col < len(self.headers) else ""

            def get_children(self):
                """获取所有子项的ID列表"""
                return range(len(self.data))

            def item(self, item, option=None):
                """获取项目的属性

                Args:
                    item: 项目ID
                    option: 选项名，如果为None则返回所有选项的字典

                Returns:
                    如果option为None则返回字典，否则返回对应选项的值
                """
                values = self.data[item] if item < len(self.data) else []
                if option is None:
                    return {"values": values}
                elif option == "values":
                    return values
                else:
                    return None

            def __getitem__(self, key):
                """支持字典访问方式"""
                if key == "columns":
                    return self.columns
                raise KeyError(key)

        return MockTree(headers, data)

    def _convert_remaining_data(self, remaining_data, remaining_headers):
        """将字典列表转换为列表列表

        Args:
            remaining_data: 剩余数据（字典列表或列表列表）
            remaining_headers: 剩余数据表头

        Returns:
            list: 转换后的数据列表
        """
        remaining_data_list = []
        for item in remaining_data:
            if isinstance(item, dict):
                # 按照剩余表头的顺序提取值
                row = [item.get(header, '') for header in remaining_headers]
                remaining_data_list.append(row)
            else:
                # 如果已经是列表，直接添加
                remaining_data_list.append(item)
        return remaining_data_list

    def export_to_file(self, file_path, data_source, main_data, main_headers, remaining_data, remaining_headers):
        """将数据导出到指定文件

        Args:
            file_path: 文件路径
            data_source: 数据源字典
            main_data: 主数据
            main_headers: 主数据表头
            remaining_data: 剩余数据
            remaining_headers: 剩余数据表头

        Returns:
            tuple: (success: bool, message: str)
        """
        try:
            # 根据文件扩展名选择相应的导出方法
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext == ".json":
                self._export_to_json(file_path, data_source, main_data, main_headers, remaining_data)
            elif file_ext == ".txt":
                self._export_to_txt(file_path, data_source, main_data, main_headers, remaining_data, remaining_headers)
            elif file_ext == ".csv":
                # 转换剩余数据并创建模拟tree对象
                remaining_data_list = self._convert_remaining_data(remaining_data, remaining_headers)
                mock_tree = self._create_mock_tree(remaining_headers, remaining_data_list)
                self._export_to_csv(file_path, main_data, main_headers, mock_tree, remaining_headers)
            elif file_ext == ".xlsx":
                # 转换剩余数据并创建模拟tree对象
                remaining_data_list = self._convert_remaining_data(remaining_data, remaining_headers)
                mock_tree = self._create_mock_tree(remaining_headers, remaining_data_list)
                self._export_to_excel(file_path, data_source, main_data, main_headers, mock_tree, remaining_headers)
            elif file_ext == ".pdf":
                self._export_to_pdf(file_path, data_source, main_data, main_headers, remaining_data, remaining_headers)
            else:
                return False, f"不支持的文件格式: {file_ext}"

            return True, f"成功导出到: {file_path}"
        except PermissionError as e:
            return False, f"导出失败: 没有写入权限，请检查文件是否被占用或目录权限是否正确: {e}"
        except FileNotFoundError as e:
            return False, f"导出失败: 文件路径不存在: {e}"
        except (IOError, OSError) as e:
            return False, f"导出失败: IO错误: {e}"
        except ValueError as e:
            return False, f"导出失败: 数据格式错误: {e}"
        except TypeError as e:
            return False, f"导出失败: 类型错误: {e}"
        except Exception as e:
            print(f"导出过程中发生意外错误: {e}")
            traceback.print_exc()
            return False, f"导出失败: 意外错误: {type(e).__name__}: {e}"
